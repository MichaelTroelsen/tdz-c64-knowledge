"""File-type extraction, document add/update/remove, scraping, tagging and summarisation.

Extracted from server.py's KnowledgeBase by R12 step 2. These are mixins,
not standalone classes: every method still expects the attributes
KnowledgeBase.__init__ sets up, and nothing here imports server, so the
class can be assembled in server.py without an import cycle.
"""
from concurrent.futures import ThreadPoolExecutor
from concurrent.futures import as_completed
from datetime import datetime
from features import OCR_SUPPORT
from features import PDFPLUMBER_SUPPORT
from features import PDF_SUPPORT
from features import PdfReader
from features import convert_from_path
from features import pdfplumber
from features import pytesseract
from models import DocumentChunk
from models import DocumentMeta
from models import DocumentNotFoundError
from models import KnowledgeBaseError
from models import ProgressCallback
from models import ProgressUpdate
from models import SecurityError
from models import UnsupportedFileTypeError
from pathlib import Path
from text_utils import _expand_brace_pattern
from typing import Optional
from util import USER_AGENT
from util import _retry_on_db_locked
from util import http_get_polite
from util import http_headers
from util import robots_allows
import json
import os
import queue
import re
import sqlite3


class IngestMixin:

    def _chunk_text(self, text: str, chunk_size: int = 1500, overlap: int = 200) -> list[str]:
        """
        Split text into overlapping chunks with optimized algorithm.

        Uses single-pass processing to minimize string operations.
        For large documents (>10K words), this is ~15% faster than naive split/join.

        Args:
            text: Text to chunk
            chunk_size: Number of words per chunk (default: 1500)
            overlap: Number of words to overlap between chunks (default: 200)

        Returns:
            List of text chunks
        """
        words = text.split()

        if len(words) <= chunk_size:
            return [text]

        chunks = []
        start = 0

        while start < len(words):
            end = start + chunk_size
            chunk_words = words[start:end]
            chunks.append(' '.join(chunk_words))
            start = end - overlap

        return chunks

    def _extract_pdf_with_ocr(self, filepath: str) -> tuple[str, int]:
        """Extract text from scanned PDF using OCR, returns (text, page_count)."""
        if not self.use_ocr:
            raise RuntimeError("OCR not enabled. Set USE_OCR=1 and install Tesseract.")

        if not self.poppler_available:
            raise RuntimeError(
                "Poppler not found! OCR requires poppler-utils.\n"
                "Install instructions:\n"
                "  Windows: Download from https://github.com/oschwartz10612/poppler-windows/releases/\n"
                "  Set POPPLER_PATH environment variable to the bin directory\n"
                "  Example: POPPLER_PATH=C:\\path\\to\\poppler-24.08.0\\Library\\bin"
            )

        try:
            self.logger.info(f"Using OCR to extract text from scanned PDF: {filepath}")
            # Convert PDF pages to images
            if self.poppler_path:
                images = convert_from_path(filepath, poppler_path=self.poppler_path)
            else:
                images = convert_from_path(filepath)

            pages = []
            for i, image in enumerate(images):
                try:
                    # Extract text using Tesseract OCR
                    text = pytesseract.image_to_string(image)
                    pages.append(text)
                    self.logger.debug(f"OCR processed page {i + 1}/{len(images)}")
                except Exception as e:
                    self.logger.error(f"OCR failed for page {i + 1}: {e}")
                    pages.append("")  # Empty text for failed page

            full_text = "\n\n--- PAGE BREAK ---\n\n".join(pages)
            self.logger.info(f"OCR extraction complete: {len(full_text)} characters from {len(images)} pages")
            return full_text, len(images)

        except Exception as e:
            self.logger.error(f"OCR extraction failed: {e}")
            raise RuntimeError(f"OCR extraction failed: {e}")

    def _check_poppler_available(self) -> bool:
        """Check if poppler is available for pdf2image.

        Returns True if poppler can be found, False otherwise.
        """
        if not OCR_SUPPORT:
            return False

        try:
            # Try to import pdf2image

            # Create a minimal test - just try to access pdfinfo
            import subprocess

            # Determine the command to check based on poppler_path
            if self.poppler_path:
                # Check if pdfinfo exists in the specified path
                pdfinfo_cmd = os.path.join(self.poppler_path, 'pdfinfo')
                if os.name == 'nt':  # Windows
                    pdfinfo_cmd += '.exe'

                if not os.path.exists(pdfinfo_cmd):
                    return False

                # Try to run pdfinfo -v (poppler uses -v not --version)
                result = subprocess.run(
                    [pdfinfo_cmd, '-v'],
                    capture_output=True,
                    timeout=5
                )
                return result.returncode == 0
            else:
                # Check if pdfinfo is in PATH
                if os.name == 'nt':  # Windows
                    result = subprocess.run(
                        ['where', 'pdfinfo'],
                        capture_output=True,
                        timeout=5
                    )
                else:  # Unix-like
                    result = subprocess.run(
                        ['which', 'pdfinfo'],
                        capture_output=True,
                        timeout=5
                    )
                return result.returncode == 0

        except Exception as e:
            self.logger.debug(f"Poppler check failed: {e}")
            return False

    def _extract_pdf_text(self, filepath: str) -> tuple[str, int, dict]:
        """Extract text from PDF with automatic OCR fallback for scanned PDFs.

        Returns (text, page_count, metadata).
        """
        if not PDF_SUPPORT:
            raise RuntimeError("PDF support not available. Install pypdf: pip install pypdf")

        reader = PdfReader(filepath)
        pages = []
        total_text_length = 0

        for page in reader.pages:
            text = page.extract_text() or ""
            pages.append(text)
            total_text_length += len(text.strip())

        # Check if PDF appears to be scanned (very little or no text extracted)
        # Threshold: if we get less than 100 characters total, likely scanned
        is_scanned = total_text_length < 100 and len(reader.pages) > 0

        if is_scanned and self.use_ocr:
            self.logger.info(f"PDF appears to be scanned ({total_text_length} chars extracted), falling back to OCR")
            try:
                ocr_text, page_count = self._extract_pdf_with_ocr(filepath)
                # Use OCR text instead of extracted pages
                pages = [ocr_text]
                # Still extract metadata from PDF
            except Exception as e:
                self.logger.warning(f"OCR fallback failed: {e}, using extracted text anyway")

        # Extract metadata
        metadata = {}
        if reader.metadata:
            # Convert metadata values to strings to handle IndirectObject references
            author = reader.metadata.get('/Author')
            metadata['author'] = str(author) if author else None

            subject = reader.metadata.get('/Subject')
            metadata['subject'] = str(subject) if subject else None

            creator = reader.metadata.get('/Creator')
            metadata['creator'] = str(creator) if creator else None

            creation_date = reader.metadata.get('/CreationDate')
            if creation_date:
                # Try to parse PDF date format (D:YYYYMMDDHHmmSS)
                try:
                    creation_date_str = str(creation_date)
                    if creation_date_str.startswith('D:'):
                        date_str = creation_date_str[2:16]  # Extract YYYYMMDDHHmmSS
                        metadata['creation_date'] = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
                    else:
                        metadata['creation_date'] = creation_date_str
                except (ValueError, TypeError, AttributeError, IndexError):
                    metadata['creation_date'] = str(creation_date)

        return "\n\n--- PAGE BREAK ---\n\n".join(pages), len(reader.pages), metadata

    def _extract_text_file(self, filepath: str) -> str:
        """Extract text from a text file."""
        encodings = ['utf-8', 'latin-1', 'cp1252']
        for enc in encodings:
            try:
                with open(filepath, 'r', encoding=enc) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
        raise RuntimeError(f"Could not decode {filepath}")

    def _extract_excel_file(self, filepath: str) -> tuple[str, int]:
        """
        Extract text from Excel file (.xlsx, .xls).

        Returns:
            Tuple of (text_content, sheet_count)
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("openpyxl not installed. Install with: pip install openpyxl")

        try:
            workbook = load_workbook(filepath, data_only=True)
            sheets_text = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Add sheet header
                sheets_text.append(f"\n{'='*60}\nSheet: {sheet_name}\n{'='*60}\n")

                # Extract all cell values
                rows_text = []
                for row in sheet.iter_rows(values_only=True):
                    # Filter out None values and convert to strings
                    row_values = [str(cell) if cell is not None else '' for cell in row]
                    # Skip completely empty rows
                    if any(val.strip() for val in row_values):
                        rows_text.append('\t'.join(row_values))

                sheets_text.append('\n'.join(rows_text))

            text_content = '\n\n'.join(sheets_text)
            sheet_count = len(workbook.sheetnames)

            self.logger.info(f"Extracted {sheet_count} sheets from Excel file")
            return text_content, sheet_count

        except Exception as e:
            raise RuntimeError(f"Error reading Excel file: {str(e)}")

    def _extract_html_file(self, filepath: str) -> str:
        """
        Extract text from HTML file (.html, .htm).

        Returns:
            Extracted text content
        """
        try:
            from bs4 import BeautifulSoup
        except ImportError:
            raise RuntimeError("beautifulsoup4 not installed. Install with: pip install beautifulsoup4")

        try:
            # Read the HTML file with encoding detection
            with open(filepath, 'rb') as f:
                raw_data = f.read()

            # Detect encoding
            import chardet
            detected = chardet.detect(raw_data)
            encoding = detected.get('encoding', 'utf-8')

            # Decode with detected encoding
            html_content = raw_data.decode(encoding, errors='replace')

            # Parse with BeautifulSoup
            soup = BeautifulSoup(html_content, 'html.parser')

            # Remove script and style elements
            for script in soup(['script', 'style', 'nav', 'footer', 'header']):
                script.decompose()

            # Extract text content
            text_parts = []

            # Get title if available
            if soup.title and soup.title.string:
                text_parts.append(f"Title: {soup.title.string.strip()}\n")

            # Process body or entire document
            body = soup.body if soup.body else soup

            # Handle code blocks specially to preserve formatting
            for pre in body.find_all(['pre', 'code']):
                # Mark code blocks so they're preserved
                pre_text = pre.get_text()
                if pre_text.strip():
                    text_parts.append(f"\n--- CODE BLOCK ---\n{pre_text}\n--- END CODE BLOCK ---\n")
                pre.decompose()  # Remove so we don't process again

            # Get remaining text
            main_text = body.get_text(separator='\n', strip=True)

            # Clean up excessive whitespace
            lines = [line.strip() for line in main_text.split('\n')]
            lines = [line for line in lines if line]  # Remove empty lines
            text_parts.append('\n'.join(lines))

            text_content = '\n\n'.join(text_parts)

            self.logger.info(f"Extracted HTML file ({len(text_content)} characters)")
            return text_content

        except Exception as e:
            raise RuntimeError(f"Error reading HTML file: {str(e)}")

    def _extract_tables(self, filepath: str) -> list[dict]:
        """Extract tables from PDF using pdfplumber.

        Returns a list of table dictionaries with structure:
        {
            'table_id': int,
            'page': int,
            'markdown': str,
            'searchable_text': str,
            'row_count': int,
            'col_count': int
        }
        """
        if not PDFPLUMBER_SUPPORT:
            self.logger.debug("pdfplumber not available, skipping table extraction")
            return []

        tables = []
        table_id = 0

        try:
            with pdfplumber.open(filepath) as pdf:
                for page_num, page in enumerate(pdf.pages, start=1):
                    # Extract tables from this page
                    page_tables = page.extract_tables()

                    if page_tables:
                        for table_data in page_tables:
                            if not table_data or len(table_data) == 0:
                                continue

                            # Convert table to markdown
                            markdown = self._table_to_markdown(table_data)

                            # Create searchable text (all cells joined with spaces)
                            searchable_text = " ".join(
                                str(cell).strip()
                                for row in table_data
                                for cell in row
                                if cell and str(cell).strip()
                            )

                            tables.append({
                                'table_id': table_id,
                                'page': page_num,
                                'markdown': markdown,
                                'searchable_text': searchable_text,
                                'row_count': len(table_data),
                                'col_count': len(table_data[0]) if table_data else 0
                            })
                            table_id += 1

        except Exception as e:
            self.logger.warning(f"Error extracting tables from {filepath}: {e}")
            return []

        self.logger.info(f"Extracted {len(tables)} tables from PDF")
        return tables

    def _table_to_markdown(self, table_data: list[list]) -> str:
        """Convert a table (list of lists) to markdown format."""
        if not table_data or len(table_data) == 0:
            return ""

        lines = []

        # Header row
        header = table_data[0]
        lines.append("| " + " | ".join(str(cell or "").strip() for cell in header) + " |")

        # Separator row
        lines.append("| " + " | ".join("---" for _ in header) + " |")

        # Data rows
        for row in table_data[1:]:
            lines.append("| " + " | ".join(str(cell or "").strip() for cell in row) + " |")

        return "\n".join(lines)

    def _detect_code_blocks(self, text: str) -> list[dict]:
        """Detect code blocks in text (BASIC, Assembly, Hex dumps).

        Returns a list of code block dictionaries with structure:
        {
            'block_id': int,
            'page': None,  # Page detection happens in add_document
            'block_type': str,  # 'basic', 'assembly', or 'hex'
            'code': str,
            'searchable_text': str,
            'line_count': int
        }
        """
        code_blocks = []
        block_id = 0

        # Pattern 1: BASIC code (lines starting with line numbers)
        # Example: "10 PRINT "HELLO"", "20 GOTO 10"
        basic_pattern = r'(?:^|\n)((?:\d+\s+[A-Z]+[^\n]*\n?){3,})'

        for match in re.finditer(basic_pattern, text, re.MULTILINE):
            code = match.group(1).strip()
            lines = code.split('\n')

            code_blocks.append({
                'block_id': block_id,
                'page': None,
                'block_type': 'basic',
                'code': code,
                'searchable_text': code,
                'line_count': len(lines)
            })
            block_id += 1

        # Pattern 2: Assembly code (mnemonics: LDA, STA, JMP, etc.)
        # Example: "    LDA #$00", "    STA $D020"
        assembly_pattern = r'(?:^|\n)((?:\s*(?:LDA|STA|LDX|STX|LDY|STY|JMP|JSR|RTS|BEQ|BNE|BCC|BCS|ADC|SBC|AND|ORA|EOR|INC|DEC|CMP|CPX|CPY|ASL|LSR|ROL|ROR|BIT|NOP|CLC|SEC|CLI|SEI|CLD|SED|CLV|PHA|PLA|PHP|PLP|TAX|TAY|TXA|TYA|TSX|TXS|INX|INY|DEX|DEY|BMI|BPL|BVC|BVS)[^\n]*\n?){3,})'

        for match in re.finditer(assembly_pattern, text, re.MULTILINE | re.IGNORECASE):
            code = match.group(1).strip()
            lines = code.split('\n')

            # Avoid duplicates (check if this code is already captured)
            if not any(block['code'] == code for block in code_blocks):
                code_blocks.append({
                    'block_id': block_id,
                    'page': None,
                    'block_type': 'assembly',
                    'code': code,
                    'searchable_text': code,
                    'line_count': len(lines)
                })
                block_id += 1

        # Pattern 3: Hex dumps (lines with hex values)
        # Example: "D000: 00 01 02 03 04 05 06 07"
        hex_pattern = r'(?:^|\n)((?:[0-9A-F]{4}:\s*(?:[0-9A-F]{2}\s*){8,}\n?){3,})'

        for match in re.finditer(hex_pattern, text, re.MULTILINE | re.IGNORECASE):
            code = match.group(1).strip()
            lines = code.split('\n')

            code_blocks.append({
                'block_id': block_id,
                'page': None,
                'block_type': 'hex',
                'code': code,
                'searchable_text': code,
                'line_count': len(lines)
            })
            block_id += 1

        basic_count = sum(1 for b in code_blocks if b['block_type'] == 'basic')
        assembly_count = sum(1 for b in code_blocks if b['block_type'] == 'assembly')
        hex_count = sum(1 for b in code_blocks if b['block_type'] == 'hex')
        self.logger.info(f"Detected {len(code_blocks)} code blocks ({basic_count} BASIC, {assembly_count} Assembly, {hex_count} Hex)")
        return code_blocks

    def _extract_facets(self, text: str) -> dict[str, set[str]]:
        """Extract categorizable terms for faceted search.

        Returns a dictionary of facet types to sets of values:
        {
            'hardware': {'SID', 'VIC-II', 'CIA'},
            'instruction': {'LDA', 'STA', 'JMP'},
            'register': {'$D000', '$D400'}
        }
        """
        facets = {
            'hardware': set(),
            'instruction': set(),
            'register': set()
        }

        # Extract hardware components
        facets['hardware'] = self._extract_hardware_refs(text)

        # Extract 6502 instructions
        facets['instruction'] = self._extract_instructions(text)

        # Extract register addresses
        facets['register'] = self._extract_registers(text)

        return facets

    def _extract_hardware_refs(self, text: str) -> set[str]:
        """Extract hardware component mentions from text."""
        hardware = set()

        # Hardware patterns (case-insensitive)
        patterns = {
            'SID': r'\b(?:SID|6581|8580|Sound\s+Interface\s+Device)\b',
            'VIC-II': r'\b(?:VIC-?II|VIC\s*2|6569|6567|Video\s+Interface\s+Chip)\b',
            'CIA': r'\b(?:CIA|6526|Complex\s+Interface\s+Adapter)\b',
            '6502': r'\b6502\b',
            'PLA': r'\b(?:PLA|82S100|Programmable\s+Logic\s+Array)\b',
            'Datasette': r'\b(?:Datasette|1530|C2N)\b',
            'Disk Drive': r'\b(?:1541|1571|1581|Disk\s+Drive)\b',
        }

        for component, pattern in patterns.items():
            if re.search(pattern, text, re.IGNORECASE):
                hardware.add(component)

        return hardware

    def _extract_instructions(self, text: str) -> set[str]:
        """Extract 6502 assembly instructions from text."""
        instructions = set()

        # Common 6502 mnemonics
        mnemonics = [
            'LDA', 'STA', 'LDX', 'STX', 'LDY', 'STY',
            'JMP', 'JSR', 'RTS', 'RTI',
            'BEQ', 'BNE', 'BCC', 'BCS', 'BMI', 'BPL', 'BVC', 'BVS',
            'ADC', 'SBC', 'AND', 'ORA', 'EOR',
            'INC', 'DEC', 'INX', 'INY', 'DEX', 'DEY',
            'CMP', 'CPX', 'CPY',
            'ASL', 'LSR', 'ROL', 'ROR',
            'BIT', 'NOP',
            'CLC', 'SEC', 'CLI', 'SEI', 'CLD', 'SED', 'CLV',
            'PHA', 'PLA', 'PHP', 'PLP',
            'TAX', 'TAY', 'TXA', 'TYA', 'TSX', 'TXS'
        ]

        for mnemonic in mnemonics:
            # Look for mnemonic as whole word (not part of another word)
            pattern = r'\b' + mnemonic + r'\b'
            if re.search(pattern, text, re.IGNORECASE):
                instructions.add(mnemonic)

        return instructions

    def _extract_registers(self, text: str) -> set[str]:
        """Extract memory register addresses from text."""
        registers = set()

        # Find all 4-digit hex addresses with $ prefix
        # Common C64 ranges: $D000-$DFFF (I/O), $A000-$BFFF (BASIC ROM), $E000-$FFFF (KERNAL ROM)
        register_pattern = r'\$[0-9A-Fa-f]{4}'
        matches = re.findall(register_pattern, text)

        # Normalize to uppercase and add to set
        for match in matches:
            registers.add(match.upper())

        return registers

    def _extract_cross_references(self, chunks: list[DocumentChunk], doc_id: str) -> list[dict]:
        """
        Extract cross-references from document chunks.

        Returns list of cross-reference dictionaries with keys:
        - doc_id, chunk_id, ref_type, ref_value, context
        """
        cross_refs = []

        for chunk in chunks:
            text = chunk.content
            chunk_id = chunk.chunk_id

            # Extract memory addresses ($D000-$FFFF)
            addresses = self._extract_memory_addresses(text)
            for addr in addresses:
                # Get context (sentence containing the address)
                context = self._get_reference_context(text, addr)
                cross_refs.append({
                    'doc_id': doc_id,
                    'chunk_id': chunk_id,
                    'ref_type': 'memory_address',
                    'ref_value': addr,
                    'context': context
                })

            # Extract register offsets (VIC+0, SID+4, etc.)
            offsets = self._extract_register_offsets(text)
            for offset in offsets:
                context = self._get_reference_context(text, offset)
                cross_refs.append({
                    'doc_id': doc_id,
                    'chunk_id': chunk_id,
                    'ref_type': 'register_offset',
                    'ref_value': offset,
                    'context': context
                })

            # Extract page references ("see page 156")
            page_refs = self._extract_page_references(text)
            for page_ref in page_refs:
                context = self._get_reference_context(text, str(page_ref))
                cross_refs.append({
                    'doc_id': doc_id,
                    'chunk_id': chunk_id,
                    'ref_type': 'page_reference',
                    'ref_value': str(page_ref),
                    'context': context
                })

        return cross_refs

    def _extract_memory_addresses(self, text: str) -> set[str]:
        """Extract memory addresses like $D000, $D020, etc."""
        addresses = set()
        # Match $xxxx format (4-digit hex)
        pattern = r'\$[0-9A-Fa-f]{4}\b'
        matches = re.findall(pattern, text)
        for match in matches:
            addresses.add(match.upper())
        return addresses

    def _extract_register_offsets(self, text: str) -> set[str]:
        """Extract register offset references like VIC+0, SID+4, CIA1+0."""
        offsets = set()
        # Match patterns like: VIC+0, SID+4, CIA1+12, etc.
        pattern = r'\b(VIC|SID|CIA[12]?|PLA)\s*\+\s*(\d+)\b'
        matches = re.findall(pattern, text, re.IGNORECASE)
        for chip, offset in matches:
            offsets.add(f"{chip.upper()}+{offset}")
        return offsets

    def _extract_page_references(self, text: str) -> set[int]:
        """Extract page number references like 'see page 156', 'page 42'."""
        page_nums = set()
        # Match patterns like: "page 123", "see page 456", "on page 789"
        pattern = r'\b(?:see\s+)?page\s+(\d+)\b'
        matches = re.findall(pattern, text, re.IGNORECASE)
        for page_num in matches:
            page_nums.add(int(page_num))
        return page_nums

    def _get_reference_context(self, text: str, reference: str, context_chars: int = 100) -> str:
        """Get surrounding context for a reference."""
        # Find the reference in text
        pos = text.find(reference)
        if pos == -1:
            # Try case-insensitive
            pos = text.lower().find(reference.lower())

        if pos == -1:
            return ""

        # Get surrounding context
        start = max(0, pos - context_chars)
        end = min(len(text), pos + len(reference) + context_chars)

        context = text[start:end].strip()

        # Add ellipsis if truncated
        if start > 0:
            context = "..." + context
        if end < len(text):
            context = context + "..."

        return context

    def _find_mdscrape_executable(self) -> Optional[str]:
        """Find mdscrape executable in common locations.

        Returns:
            Path to mdscrape executable, or None if not found
        """
        import shutil

        # Check if mdscrape is in PATH
        mdscrape = shutil.which('mdscrape')
        if mdscrape:
            self.logger.info(f"Found mdscrape in PATH: {mdscrape}")
            return mdscrape

        # Check common Windows/Linux paths
        common_paths = [
            Path(r'C:\Users\mit\claude\mdscrape\mdscrape.exe'),  # User-specified location
            Path(r'C:\Users\mit\claude\mdscrape\mdscrape'),
            Path.home() / 'claude' / 'mdscrape' / 'mdscrape.exe',
            Path.home() / 'claude' / 'mdscrape' / 'mdscrape',
            Path(__file__).parent.parent / 'mdscrape' / 'mdscrape.exe',
            Path(__file__).parent.parent / 'mdscrape' / 'mdscrape',
        ]

        for path in common_paths:
            if path.exists():
                self.logger.info(f"Found mdscrape at: {path}")
                return str(path)

        # Check MDSCRAPE_PATH environment variable
        env_path = os.environ.get('MDSCRAPE_PATH')
        if env_path:
            path = Path(env_path)
            if path.exists():
                self.logger.info(f"Found mdscrape via MDSCRAPE_PATH: {path}")
                return str(path)

        self.logger.warning("mdscrape executable not found. Install from: https://github.com/MichaelTroelsen/mdscrape")
        return None

    def _extract_source_url_from_md(self, md_file: Path) -> Optional[str]:
        """Extract source URL from YAML frontmatter in markdown file.

        Args:
            md_file: Path to markdown file

        Returns:
            Source URL if found, None otherwise
        """
        try:
            with open(md_file, 'r', encoding='utf-8') as f:
                content = f.read()

            # Parse YAML frontmatter (between --- delimiters)
            if content.startswith('---'):
                parts = content.split('---', 2)
                if len(parts) >= 3:
                    frontmatter = parts[1]
                    # Simple YAML parsing for 'source:' or 'url:' field
                    for line in frontmatter.split('\n'):
                        line = line.strip()
                        if line.startswith('source:') or line.startswith('url:'):
                            # Extract URL after colon
                            url = line.split(':', 1)[1].strip().strip('"\'')
                            if url:
                                return url
        except Exception as e:
            self.logger.warning(f"Failed to extract URL from {md_file}: {e}")

        return None

    def _add_scraped_document(self, filepath: str, source_url: str, title: Optional[str],
                              tags: Optional[list[str]], scrape_config: str,
                              scrape_date: str) -> DocumentMeta:
        """Add a scraped markdown document with URL metadata.

        Args:
            filepath: Path to scraped markdown file
            source_url: Original URL that was scraped
            title: Optional title for document
            tags: Optional list of tags
            scrape_config: JSON string with scraping configuration
            scrape_date: ISO timestamp of scrape

        Returns:
            DocumentMeta object for added document
        """
        # First, add document using normal flow
        doc = self.add_document(filepath, title, tags)

        # Compute content hash for change detection
        url_content_hash = self._compute_file_hash(filepath)

        # Update database with URL metadata
        with self._lock:
            cursor = self.db_conn.cursor()
            cursor.execute("""
                UPDATE documents
                SET source_url = ?,
                    scrape_date = ?,
                    scrape_config = ?,
                    scrape_status = 'success',
                    url_content_hash = ?
                WHERE doc_id = ?
            """, (source_url, scrape_date, scrape_config, url_content_hash, doc.doc_id))

            self.db_conn.commit()

        # Update in-memory object
        doc.source_url = source_url
        doc.scrape_date = scrape_date
        doc.scrape_config = scrape_config
        doc.scrape_status = 'success'
        doc.url_content_hash = url_content_hash

        # Update in documents dict
        self.documents[doc.doc_id] = doc

        self.logger.info(f"Added scraped document: {doc.title} (from {source_url})")
        return doc

    def _is_path_allowed(self, filepath: str) -> bool:
        """
        Check if a file path is within allowed directories.

        Args:
            filepath: Path to check

        Returns:
            True if path is allowed (or no restrictions configured), False otherwise
        """
        # No restrictions if allowed_dirs not configured
        if not self.allowed_dirs:
            return True

        # Resolve to absolute path to prevent path traversal
        try:
            resolved_path = Path(filepath).resolve()
        except (OSError, ValueError):
            # Invalid path
            return False

        # Check if path is within any allowed directory
        return any(
            resolved_path.is_relative_to(allowed_dir)
            for allowed_dir in self.allowed_dirs
        )

    def _extract_text_for_file(self, filepath: str) -> tuple[str, str, Optional[int], dict]:
        """Extract raw text and file-type metadata for a document file.

        Shared by add_document (full ingest) and update_document (which peeks
        a new file's declared card id before deciding what to supersede).

        Returns:
            (text, file_type, total_pages, pdf_metadata)
        """
        filename = os.path.basename(filepath)
        file_ext = os.path.splitext(filename)[1].lower()

        total_pages = None
        pdf_metadata = {}
        try:
            if file_ext == '.pdf':
                text, total_pages, pdf_metadata = self._extract_pdf_text(filepath)
                file_type = 'pdf'
                self.logger.info(f"Extracted {total_pages} pages from PDF")
            elif file_ext in ['.xlsx', '.xls']:
                text, sheet_count = self._extract_excel_file(filepath)
                file_type = 'excel'
                total_pages = sheet_count  # Treat sheets as "pages"
                self.logger.info(f"Extracted Excel file with {sheet_count} sheets ({len(text)} characters)")
            elif file_ext in ['.html', '.htm']:
                text = self._extract_html_file(filepath)
                file_type = 'html'
                self.logger.info(f"Extracted HTML file ({len(text)} characters)")
            elif file_ext in ['.txt', '.md', '.asm', '.bas', '.inc', '.s']:
                text = self._extract_text_file(filepath)
                file_type = 'text'
                self.logger.info(f"Extracted text file ({len(text)} characters)")
            else:
                raise UnsupportedFileTypeError(f"Unsupported file type: {file_ext}")
        except (UnsupportedFileTypeError, DocumentNotFoundError):
            raise
        except Exception as e:
            self.logger.error(f"Error extracting {filepath}: {e}")
            raise KnowledgeBaseError(f"Error extracting document: {e}")

        return text, file_type, total_pages, pdf_metadata

    def _extract_card_id(self, text: str) -> Optional[str]:
        """Parse the logical `id` out of a knowledge card's fenced ```json block.

        Cards are markdown documents whose identity is the `id` field of their
        first fenced json block, not their (content-hash) doc_id. Anything
        without a well-formed json block with a string `id` is not a card -
        PDFs, scrapes, ezines, etc. return None here and are left untouched
        by the upsert/refuse logic in add_document/update_document.
        """
        if not text:
            return None
        match = re.search(r'```json\s*\n(.*?)```', text, re.DOTALL)
        if not match:
            return None
        try:
            data = json.loads(match.group(1))
        except (json.JSONDecodeError, ValueError):
            return None
        if isinstance(data, dict):
            card_id = data.get('id')
            if isinstance(card_id, str) and card_id.strip():
                return card_id.strip()
        return None

    def get_document_by_card_id(self, card_id: str, include_superseded: bool = False) -> Optional[DocumentMeta]:
        """Resolve a card's logical id to its document.

        Returns the live (non-superseded) card by default - by construction
        there is at most one, since add_document/update_document refuse to
        create a second live document for the same card_id. With
        include_superseded=True, falls back to the most recently indexed
        superseded version if no live one exists.
        """
        matches = [d for d in self.documents.values() if d.card_id == card_id]
        live = [d for d in matches if not d.superseded_by]
        if live:
            return live[0]
        if include_superseded and matches:
            return sorted(matches, key=lambda d: d.indexed_at, reverse=True)[0]
        return None

    def _rebuild_entity_relationships(self) -> None:
        """Recompute entity_relationships from scratch across all live documents.

        entity_relationships has no per-document attribution - it's a running
        aggregate keyed only on (entity1_text, entity2_text, relationship_type)
        with no doc_id column - so a superseded document's contribution can't
        be surgically subtracted. The only correct fix is to wipe the table
        and rebuild it from whatever document_entities currently holds for
        live documents.
        """
        cursor = self.db_conn.cursor()
        cursor.execute("DELETE FROM entity_relationships")
        self.db_conn.commit()

        cursor.execute("SELECT DISTINCT doc_id FROM document_entities")
        doc_ids = [row[0] for row in cursor.fetchall()]
        for doc_id in doc_ids:
            doc = self.documents.get(doc_id)
            if doc is None or doc.superseded_by:
                continue
            try:
                self.extract_entity_relationships(doc_id, force_regenerate=True)
            except Exception as e:
                self.logger.warning(f"Relationship rebuild failed for {doc_id}: {e}")

    def _mark_superseded(self, old_doc_id: str, new_doc_id: str) -> None:
        """Mark old_doc_id as superseded by new_doc_id and purge its contribution
        from derived artifacts that would otherwise keep answering with retracted
        content (entities, entity relationships, cached graph artifacts).

        Chunks/embeddings for old_doc_id are deliberately left in place - the
        card's prior content stays retrievable by doc_id for history/audit -
        but old_doc_id is excluded from default search results and from
        get_document_by_card_id once superseded_by is set.
        """
        if old_doc_id not in self.documents or old_doc_id == new_doc_id:
            return

        cursor = self.db_conn.cursor()
        cursor.execute(
            "UPDATE documents SET superseded_by = ? WHERE doc_id = ?",
            (new_doc_id, old_doc_id)
        )
        self.db_conn.commit()
        self.documents[old_doc_id].superseded_by = new_doc_id

        # Purge stale entities contributed by the retracted content, then
        # rebuild the globally-aggregated relationship table from what's left
        # so stale co-occurrence edges can't survive it.
        cursor.execute("SELECT COUNT(*) FROM document_entities WHERE doc_id = ?", (old_doc_id,))
        had_entities = cursor.fetchone()[0] > 0
        if had_entities:
            cursor.execute("DELETE FROM document_entities WHERE doc_id = ?", (old_doc_id,))
            self.db_conn.commit()
            if self._entity_cache is not None:
                self._entity_cache.clear()
            try:
                self.extract_entities(new_doc_id, confidence_threshold=0.6, force_regenerate=True)
            except Exception as e:
                self.logger.warning(f"Entity re-extraction for {new_doc_id} failed after supersede: {e}")
            self._rebuild_entity_relationships()

        # Drop cached graph artifacts so they can't keep answering with
        # retracted claims - build_knowledge_graph rebuilds from
        # document_entities/entity_relationships on every call, but other
        # code paths may read these caches directly.
        for table in ("graph_cache", "graph_metrics", "graph_paths"):
            try:
                cursor.execute(f"DELETE FROM {table}")
            except sqlite3.OperationalError:
                pass
        self.db_conn.commit()

        # Invalidate search caches - the old doc is now excluded from default results.
        self._invalidate_caches()

        self.logger.info(f"Marked {old_doc_id} as superseded by {new_doc_id}")

    def add_document(self, filepath: str, title: Optional[str] = None, tags: Optional[list[str]] = None,
                     progress_callback: ProgressCallback = None, replace: bool = False) -> DocumentMeta:
        """Add a document to the knowledge base.

        Args:
            filepath: Path to the document file
            title: Optional title for the document
            tags: Optional list of tags
            progress_callback: Optional callback for progress updates
            replace: If the file is a knowledge card (has a ```json id``` block)
                and a live card with the same id already exists, add_document
                refuses by default (raises KnowledgeBaseError naming the
                existing doc id). Pass replace=True to supersede it instead.
                Non-card documents (no json id block) are never affected by
                this check and are always created, matching prior behavior.
        """
        # Report progress: Start
        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="add_document",
                current=0,
                total=4,
                message="Starting document ingestion",
                item=filepath
            ))

        # Resolve to absolute path to prevent path traversal
        resolved_path = Path(filepath).resolve()

        # Security: Validate path is within allowed directories
        if not self._is_path_allowed(filepath):
            self.logger.error(f"Security violation: Path outside allowed directories: {resolved_path}")
            raise SecurityError(
                f"Path outside allowed directories. File must be within: {self.allowed_dirs}"
            )

        filepath = str(resolved_path)
        self.logger.info(f"Adding document: {filepath}")

        if not os.path.exists(filepath):
            self.logger.error(f"File not found: {filepath}")
            raise DocumentNotFoundError(f"File not found: {filepath}")

        filename = os.path.basename(filepath)

        # Extract text based on file type
        text, file_type, total_pages, pdf_metadata = self._extract_text_for_file(filepath)

        # Report progress: Text extraction complete
        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="add_document",
                current=1,
                total=4,
                message=f"Text extraction complete ({len(text)} characters)",
                item=filename
            ))

        # Extract tables from PDFs
        tables = []
        if file_type == 'pdf':
            tables = self._extract_tables(filepath)
            if tables:
                self.logger.info(f"Extracted {len(tables)} tables from PDF")

        # Detect code blocks in text
        code_blocks = self._detect_code_blocks(text)
        if code_blocks:
            self.logger.info(f"Detected {len(code_blocks)} code blocks")

        # Extract facets for faceted search
        facets = self._extract_facets(text)
        facet_count = sum(len(values) for values in facets.values())
        if facet_count > 0:
            self.logger.info(f"Extracted {facet_count} facets ({len(facets['hardware'])} hardware, {len(facets['instruction'])} instructions, {len(facets['register'])} registers)")

        # Generate content-based doc_id for deduplication
        doc_id = self._generate_doc_id(filepath, text)

        # Parse the card's logical identity, if this is a knowledge card
        card_id = self._extract_card_id(text)

        # Thread-safe duplicate check
        superseded_doc_id = None
        with self._lock:
            # Check for duplicate content
            if doc_id in self.documents:
                existing_doc = self.documents[doc_id]
                self.logger.warning(f"Duplicate content detected: {filepath}")
                self.logger.warning(f"  Matches existing document: {existing_doc.filepath}")
                self.logger.info(f"Skipping duplicate - returning existing document {doc_id}")
                return existing_doc

            # Card-identity guard: refuse to silently fork a card that already
            # exists under the same logical id. This is the fix for the
            # "two live documents both claim id: X" failure mode - callers
            # must explicitly opt into replacing via replace=True or
            # update_document().
            if card_id:
                existing_card = self.get_document_by_card_id(card_id, include_superseded=False)
                if existing_card and existing_card.doc_id != doc_id:
                    if not replace:
                        raise KnowledgeBaseError(
                            f"Card '{card_id}' already exists as document {existing_card.doc_id} "
                            f"('{existing_card.title}'). Use update_document() to replace it, "
                            f"or pass replace=true to add_document()."
                        )
                    superseded_doc_id = existing_card.doc_id

        # Create chunks
        text_chunks = self._chunk_text(text)
        chunks = []
        for i, chunk_text in enumerate(text_chunks):
            # Estimate page number for PDFs based on PAGE BREAK markers
            page_num = None
            if file_type == 'pdf' and '--- PAGE BREAK ---' in text:
                # Count PAGE BREAK markers before this chunk
                chunk_start_pos = text.find(chunk_text[:100])  # Find chunk in full text
                if chunk_start_pos >= 0:
                    page_breaks_before = text[:chunk_start_pos].count('--- PAGE BREAK ---')
                    page_num = page_breaks_before + 1  # Pages are 1-indexed

            chunk = DocumentChunk(
                doc_id=doc_id,
                filename=filename,
                title=title or filename,
                chunk_id=i,
                page=page_num,
                content=chunk_text,
                word_count=len(chunk_text.split())
            )
            chunks.append(chunk)

        # Report progress: Chunking complete
        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="add_document",
                current=2,
                total=4,
                message=f"Created {len(chunks)} chunks",
                item=filename
            ))

        # Compute file modification time and content hash for update detection
        file_mtime = os.path.getmtime(resolved_path)
        file_hash = self._compute_file_hash(resolved_path)

        # Extract cross-references for content linking
        cross_refs = self._extract_cross_references(chunks, doc_id)
        if cross_refs:
            self.logger.info(f"Extracted {len(cross_refs)} cross-references")

        # Create metadata
        doc_meta = DocumentMeta(
            doc_id=doc_id,
            filename=filename,
            title=title or filename,
            filepath=filepath,
            file_type=file_type,
            total_pages=total_pages,
            total_chunks=len(chunks),
            indexed_at=datetime.now().isoformat(),
            tags=tags or [],
            author=pdf_metadata.get('author'),
            subject=pdf_metadata.get('subject'),
            creator=pdf_metadata.get('creator'),
            creation_date=pdf_metadata.get('creation_date'),
            file_mtime=file_mtime,
            file_hash=file_hash,
            card_id=card_id
        )

        # Thread-safe database insertion and cache invalidation
        with self._lock:
            # Add to database (with tables, code blocks, facets, and cross-references)
            _retry_on_db_locked(
                self._add_document_db, doc_meta, chunks,
                tables=tables, code_blocks=code_blocks, facets=facets, cross_refs=cross_refs
            )
            self.documents[doc_id] = doc_meta

            # Report progress: Database insertion complete
            if progress_callback:
                progress_callback(ProgressUpdate(
                    operation="add_document",
                    current=3,
                    total=4,
                    message="Stored in database",
                    item=filename
                ))

            # Invalidate BM25 index (will be rebuilt on next search)
            self.bm25 = None

            # Incrementally add chunks to embeddings (faster than full rebuild).
            # Must load the model first: _add_chunks_to_embeddings silently
            # no-ops when embeddings_model is None, which it is for every
            # process that hasn't yet run a semantic search - i.e. every
            # ingest-only session. Without this, newly-added documents are
            # never embedded and there is no error to notice.
            if self.use_semantic:
                self._ensure_embeddings_loaded()
                if self.embeddings_model is not None:
                    self._add_chunks_to_embeddings(chunks)
                else:
                    self.logger.warning(
                        f"Skipping embeddings for {doc_id}: embeddings model unavailable"
                    )

            # Update query suggestions with new terms
            self._update_suggestions_for_chunks(chunks)

            # Invalidate search caches
            self._invalidate_caches()

        # If this replaces an existing card, retire the old one and refresh
        # everything derived from its (now-retracted) content.
        if superseded_doc_id:
            self._mark_superseded(superseded_doc_id, doc_id)

        self.logger.info(f"Successfully indexed document {doc_id}: {filename} ({len(chunks)} chunks)")

        # Report progress: Complete
        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="add_document",
                current=4,
                total=4,
                message="Document indexed successfully",
                item=filename
            ))

        # Auto-queue entity extraction (configurable via environment variable)
        auto_extract = os.getenv('AUTO_EXTRACT_ENTITIES', '1') == '1'
        if auto_extract:
            try:
                result = self.queue_entity_extraction(
                    doc_id=doc_meta.doc_id,
                    confidence_threshold=0.6,
                    skip_if_exists=True
                )
                if result.get('queued'):
                    self.logger.info(f"Auto-queued entity extraction job {result['job_id']} for document {doc_meta.doc_id}")
                else:
                    self.logger.debug(f"Entity extraction not queued: {result.get('reason', 'unknown')}")
            except Exception as e:
                # Don't fail document ingestion if extraction queueing fails
                self.logger.warning(f"Failed to auto-queue entity extraction: {e}")

        return doc_meta

    def _detect_and_extract_frames(self, url: str) -> list[str]:
        """Detect HTML frames/iframes and extract their source URLs.

        Args:
            url: The URL to check for frames

        Returns:
            List of frame source URLs (relative URLs converted to absolute)
        """
        import re
        import requests  # for the exception types in the handlers below

        try:
            # Fetch the HTML content with longer timeout
            response = http_get_polite(url, timeout=30)
            response.raise_for_status()
            html_content = response.text

            # Check if this is a frameset page
            if '<FRAMESET' not in html_content.upper() and '<FRAME' not in html_content.upper() and '<IFRAME' not in html_content.upper():
                return []

            self.logger.info(f"Detected frameset page at {url}")

            # Extract frame/iframe src attributes using regex
            # Match <frame src="..."> and <iframe src="...">
            frame_pattern = r'<(?:frame|iframe)[^>]+src=["\']([\w\-\.\/]+)["\']'
            matches = re.findall(frame_pattern, html_content, re.IGNORECASE)

            if not matches:
                return []

            # Convert relative URLs to absolute
            from urllib.parse import urljoin
            frame_urls = []
            for src in matches:
                # Skip external URLs and special targets
                if src.startswith('http://') or src.startswith('https://') or src.startswith('javascript:'):
                    continue

                absolute_url = urljoin(url, src)
                frame_urls.append(absolute_url)
                self.logger.info(f"Found frame source: {absolute_url}")

            return frame_urls

        except requests.exceptions.Timeout:
            self.logger.debug(f"Timeout detecting frames at {url} (will use mdscrape instead)")
            return []
        except Exception as e:
            self.logger.debug(f"Could not detect frames at {url}: {e} (will use mdscrape instead)")
            return []

    def scrape_url(self, url: str, title: Optional[str] = None, tags: Optional[list[str]] = None,
                   follow_links: bool = True, same_domain_only: bool = True,
                   max_pages: int = 50, depth: int = 3, limit: Optional[str] = None,
                   threads: int = 3, delay: int = 500, selector: Optional[str] = None,
                   progress_callback: ProgressCallback = None) -> dict:
        """Scrape a URL using mdscrape and add resulting documents to knowledge base.

        Supports recursive scraping of entire websites by following links.

        Args:
            url: Starting URL to scrape (e.g., http://www.sidmusic.org/sid/)
            title: Optional base title for scraped documents
            tags: Optional list of tags (domain name auto-added)
            follow_links: Follow links to scrape sub-pages (default: True)
            same_domain_only: Only follow links on the same domain (default: True)
            max_pages: Maximum number of pages to scrape (default: 50)
            depth: Maximum crawl depth - how many link levels to follow (default: 3)
            limit: Advanced: Limit scraping to URLs with this prefix (overrides same_domain_only)
            threads: Number of concurrent threads (default: 3 - these sources are
                small volunteer-run sites; a high thread count invites a ban)
            delay: Delay between requests in ms (default: 500)
            selector: CSS selector for main content (optional)
            progress_callback: Optional callback for progress updates

        Examples:
            # Scrape single page only
            kb.scrape_url("http://example.com/page.html", follow_links=False)

            # Scrape entire site (stay on same domain, max 3 levels deep)
            kb.scrape_url("http://www.sidmusic.org/sid/", follow_links=True, same_domain_only=True, depth=3)

            # Scrape specific section (limit to /sid/ prefix)
            kb.scrape_url("http://www.sidmusic.org/sid/", limit="http://www.sidmusic.org/sid/")

        Returns:
            Dictionary with scraping results:
            {
                'status': 'success' | 'partial' | 'failed',
                'url': original_url,
                'output_dir': path_to_scraped_files,
                'files_scraped': count,
                'docs_added': count,
                'docs_updated': count,
                'docs_failed': count,
                'pages_scraped': list_of_urls,
                'error': error_message (if failed),
                'doc_ids': [list of added doc_ids]
            }
        """
        import subprocess
        from urllib.parse import urlparse
        from datetime import datetime

        # 1. Validate URL
        try:
            parsed = urlparse(url)
            if not parsed.scheme or not parsed.netloc:
                raise ValueError(f"Invalid URL: {url}")
            if parsed.scheme not in ['http', 'https']:
                raise ValueError(f"Only HTTP/HTTPS URLs supported: {url}")
        except Exception as e:
            return {
                'status': 'failed',
                'url': url,
                'error': f"Invalid URL: {str(e)}"
            }

        # 2. Extract domain for auto-tagging
        domain = parsed.netloc.replace('www.', '')
        if tags is None:
            tags = []
        tags = list(tags) + [domain, 'scraped']

        # 3. Setup output directory in scraped_docs
        scraped_base = self.data_dir / "scraped_docs"
        scraped_base.mkdir(exist_ok=True)

        # Use domain + timestamp for unique output dir
        safe_domain = domain.replace('.', '_').replace(':', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = scraped_base / f"{safe_domain}_{timestamp}"

        # 3a. Handle follow_links and same_domain_only parameters
        if not follow_links:
            # Don't follow any links - just scrape the single page
            depth = 1
            self.logger.info("follow_links=False: Scraping single page only (depth=1)")

        # 3a-bis. Respect robots.txt before touching the site at all.
        if not robots_allows(url):
            self.logger.warning(f"robots.txt disallows scraping {url}")
            return {
                'status': 'failed',
                'url': url,
                'files_scraped': 0,
                'docs_added': 0,
                'docs_updated': 0,
                'docs_failed': 0,
                'error': (
                    "robots.txt on this host disallows fetching this URL. "
                    "Set TDZ_RESPECT_ROBOTS=0 to override if you have permission "
                    "(e.g. your own mirror)."
                ),
            }

        # 3b. Detect and handle HTML frames
        frame_urls = self._detect_and_extract_frames(url)
        if frame_urls:
            self.logger.info(f"Detected {len(frame_urls)} frame(s), will scrape each individually")

            # Scrape each frame source recursively
            all_doc_ids = []
            all_files_scraped = 0
            all_docs_added = 0

            for frame_url in frame_urls:
                self.logger.info(f"Scraping frame: {frame_url}")

                # For frames, use the parent directory as limit if same_domain_only
                # This allows following links from the frame
                frame_limit = limit
                if same_domain_only and limit is None:
                    # Use the parent directory of the original URL
                    frame_limit = f"{parsed.scheme}://{parsed.netloc}{parsed.path.rsplit('/', 1)[0]}"
                    if not frame_limit.endswith('/'):
                        frame_limit += '/'

                frame_result = self.scrape_url(
                    url=frame_url,
                    title=title,
                    tags=tags,
                    follow_links=follow_links,
                    same_domain_only=False,  # Disable auto-limit for frames
                    max_pages=max_pages,
                    depth=depth,
                    limit=frame_limit,  # Use parent directory as limit
                    threads=threads,
                    delay=delay,
                    selector=selector,
                    progress_callback=progress_callback
                )

                if frame_result['status'] == 'success':
                    all_doc_ids.extend(frame_result.get('doc_ids', []))
                    all_files_scraped += frame_result.get('files_scraped', 0)
                    all_docs_added += frame_result.get('docs_added', 0)

            # Return combined results from all frames
            return {
                'status': 'success',
                'url': url,
                'frames_detected': len(frame_urls),
                'files_scraped': all_files_scraped,
                'docs_added': all_docs_added,
                'docs_updated': 0,
                'docs_failed': 0,
                'doc_ids': all_doc_ids,
                'message': f'Scraped {len(frame_urls)} frames with {all_docs_added} total documents'
            }

        if same_domain_only and limit is None:
            # Automatically set limit to base domain URL to stay on same domain
            # Extract base URL (scheme + netloc + path up to last /)
            base_url = f"{parsed.scheme}://{parsed.netloc}"

            # If URL has a path, use it as the limit prefix
            if parsed.path and parsed.path != '/':
                # Get the directory part of the path (not the file)
                path_parts = parsed.path.rstrip('/').split('/')
                if path_parts:
                    # Use the full path as prefix to stay within that section
                    base_path = '/'.join(path_parts)
                    limit = f"{base_url}{base_path}"
                else:
                    limit = base_url
            else:
                limit = base_url

            self.logger.info(f"same_domain_only=True: Limiting to URLs starting with '{limit}'")

        # 4. Build mdscrape command
        mdscrape_path = self._find_mdscrape_executable()
        if not mdscrape_path:
            return {
                'status': 'failed',
                'url': url,
                'error': 'mdscrape executable not found. Set MDSCRAPE_PATH or install from: https://github.com/MichaelTroelsen/mdscrape'
            }

        cmd = [
            mdscrape_path,
            url,
            '--output', str(output_dir),
            '--depth', str(depth),
            '--threads', str(threads),
            '--delay', str(delay),
            # mdscrape otherwise announces itself as the generic "mdscrape/1.0",
            # which tells a site operator nothing about who is crawling them.
            '--user-agent', USER_AGENT,
        ]

        if limit:
            cmd.extend(['--limit', limit])
        if selector:
            cmd.extend(['--selector', selector])
        # Note: max_pages is a UI parameter only - mdscrape doesn't support it yet
        # Use depth to control crawl scope instead

        # 5. Store scrape config
        scrape_config = {
            'url': url,
            'follow_links': follow_links,
            'same_domain_only': same_domain_only,
            'max_pages': max_pages,
            'depth': depth,
            'limit': limit,
            'threads': threads,
            'delay': delay,
            'selector': selector,
            'timestamp': datetime.now().isoformat()
        }
        scrape_config_json = json.dumps(scrape_config)

        # 6. Execute mdscrape
        self.logger.info(f"Scraping URL: {url}")
        self.logger.info(f"Command: {' '.join(cmd)}")

        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="scrape_url",
                current=0,
                total=100,
                message="Starting web scraping",
                item=url
            ))

        try:
            # Execute with real-time output streaming
            import time
            import re
            from threading import Thread, Event
            from queue import Queue, Empty

            process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                bufsize=1  # Line buffered
            )

            # Queues for capturing output
            stdout_queue = Queue()
            stderr_queue = Queue()

            # Helper to read output in background thread
            def enqueue_output(stream, queue, stop_event):
                try:
                    for line in iter(stream.readline, ''):
                        if stop_event.is_set():
                            break
                        queue.put(line)
                except Exception:
                    pass
                finally:
                    stream.close()

            stop_event = Event()
            stdout_thread = Thread(target=enqueue_output, args=(process.stdout, stdout_queue, stop_event))
            stderr_thread = Thread(target=enqueue_output, args=(process.stderr, stderr_queue, stop_event))
            stdout_thread.daemon = True
            stderr_thread.daemon = True
            stdout_thread.start()
            stderr_thread.start()

            # Track progress
            pages_scraped = 0
            current_url = None
            last_update_time = time.time()
            timeout_warned = False
            stdout_lines = []
            stderr_lines = []

            # This loop used to have no wall-clock deadline at all: it only
            # LOGGED a warning after 60s of no progress and never actually
            # terminated the process, so a stalled or hung mdscrape process
            # blocked this call (and, per issue #12's investigation, the
            # whole asyncio event loop) indefinitely. scrape_start_time and
            # stop_reason below give it a real, enforced deadline.
            scrape_start_time = time.time()
            scrape_timeout_s = float(os.getenv('SCRAPE_TIMEOUT_S', '3600'))
            stop_reason = None  # 'max_pages' | 'timeout' | None (ran to completion)

            # Process output in real-time
            while process.poll() is None:
                # Check stdout
                try:
                    line = stdout_queue.get(timeout=0.1)
                    stdout_lines.append(line)

                    # Parse mdscrape output for current URL
                    # mdscrape typically outputs: "Scraping: https://example.com/page"
                    url_match = re.search(r'(?:Scraping|Processing|Fetching)[:\s]+(\S+)', line, re.IGNORECASE)
                    if url_match:
                        current_url = url_match.group(1)
                        pages_scraped += 1
                        last_update_time = time.time()
                        timeout_warned = False

                        # Update progress
                        self.logger.info(f"[{pages_scraped}/{max_pages}] Scraping: {current_url}")

                        if progress_callback:
                            progress_callback(ProgressUpdate(
                                operation="scrape_url",
                                current=min(pages_scraped, max_pages),
                                total=max_pages,
                                message=f"Scraping page {pages_scraped}/{max_pages}",
                                item=current_url
                            ))

                        # mdscrape has no concept of max_pages itself (see the
                        # comment above where the command is built) - this is
                        # the actual enforcement. Reaching the requested cap
                        # is success, not an error, so it's tracked separately
                        # from the timeout case below.
                        if pages_scraped >= max_pages:
                            stop_reason = 'max_pages'
                            self.logger.info(f"Reached max_pages={max_pages}, stopping crawl")
                            process.terminate()
                            break

                except Empty:
                    pass

                # Check stderr
                try:
                    line = stderr_queue.get_nowait()
                    stderr_lines.append(line)
                    # Log errors but don't stop
                    if 'error' in line.lower() and 'image' not in line.lower():
                        self.logger.warning(f"Scrape warning: {line.strip()}")
                except Empty:
                    pass

                # No-progress warning stays informational (mdscrape can go
                # quiet on a slow page and still recover), but the overall
                # wall-clock deadline below is a hard stop - previously
                # nothing in this loop ever terminated the process, so a
                # truly hung mdscrape run blocked this call forever.
                time_since_update = time.time() - last_update_time
                if time_since_update > 60 and not timeout_warned:
                    timeout_warned = True
                    warning_msg = f"⚠️ No progress for {int(time_since_update)} seconds"
                    if current_url:
                        warning_msg += f" (current: {current_url})"
                    self.logger.warning(warning_msg)

                    if progress_callback:
                        progress_callback(ProgressUpdate(
                            operation="scrape_url",
                            current=pages_scraped,
                            total=max_pages,
                            message=f"⚠️ Page taking longer than 60s...",
                            item=current_url or "unknown"
                        ))

                if time.time() - scrape_start_time > scrape_timeout_s:
                    stop_reason = 'timeout'
                    self.logger.error(
                        f"Scraping exceeded {scrape_timeout_s:.0f}s wall-clock limit "
                        f"({pages_scraped} pages scraped so far), terminating"
                    )
                    process.terminate()
                    break

            # A deliberate stop (page cap or timeout) needs its own
            # terminate-then-kill sequence; a process that already exited on
            # its own just needs reaping, which is what the original
            # unconditional wait(timeout=60) here provided.
            if stop_reason is not None:
                try:
                    process.wait(timeout=10)
                except subprocess.TimeoutExpired:
                    self.logger.warning("Process did not exit after terminate(), killing it")
                    process.kill()
                    process.wait(timeout=10)
            else:
                process.wait(timeout=60)

            # Stop background threads
            stop_event.set()
            stdout_thread.join(timeout=1)
            stderr_thread.join(timeout=1)

            # Collect remaining output
            while not stdout_queue.empty():
                try:
                    stdout_lines.append(stdout_queue.get_nowait())
                except Empty:
                    break

            while not stderr_queue.empty():
                try:
                    stderr_lines.append(stderr_queue.get_nowait())
                except Empty:
                    break

            # Check for errors but don't fail if files were scraped
            stdout_output = ''.join(stdout_lines)
            stderr_output = ''.join(stderr_lines)

            if stop_reason is not None:
                # We deliberately terminated the process ourselves, so a
                # nonzero returncode here is expected (the classic
                # image-error/generic-error classification below is for
                # mdscrape's OWN exit status and would just be noise).
                self.logger.info(
                    f"Scraping stopped by {stop_reason} after {pages_scraped} pages "
                    f"(not a failure - proceeding with what was scraped)"
                )
            elif process.returncode != 0:
                error_msg = stderr_output or stdout_output or "Unknown error"

                # Count image-related errors (not critical failures)
                image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.svg', '.webp', '.ico']
                error_lines = error_msg.split('\n')
                image_errors = sum(1 for line in error_lines
                                 if any(ext in line.lower() for ext in image_extensions))
                total_errors = len([line for line in error_lines if 'Error:' in line or 'error' in line.lower()])

                # If all errors are image-related, treat as warning not failure
                if image_errors > 0 and total_errors > 0 and image_errors == total_errors:
                    self.logger.warning(f"Scraping completed with {image_errors} image-related errors (expected)")
                else:
                    # Log full error but continue - we'll check if any files were scraped
                    self.logger.warning(f"Scraping completed with errors: {error_msg[:500]}...")
            else:
                self.logger.info(f"[OK] Scraping completed successfully ({pages_scraped} pages)")

        except subprocess.TimeoutExpired:
            # The loop above now enforces scrape_timeout_s itself and
            # terminates/kills the process well before this can fire from
            # the crawl running long - reaching here means the process
            # didn't die even after kill(), which is the genuinely
            # exceptional case worth surfacing distinctly.
            self.logger.error("Scrape process did not exit even after being killed")
            return {
                'status': 'failed',
                'url': url,
                'error': 'Scrape process did not exit even after being killed'
            }
        except Exception as e:
            self.logger.error(f"Scraping error: {e}")
            return {
                'status': 'failed',
                'url': url,
                'error': f"Scraping error: {str(e)}"
            }

        # 7. Find all generated markdown files
        if not output_dir.exists():
            return {
                'status': 'failed',
                'url': url,
                'error': f"Output directory not created: {output_dir}",
                'stop_reason': stop_reason,
            }

        md_files = list(output_dir.rglob('*.md'))

        if not md_files:
            return {
                'status': 'failed',
                'url': url,
                'error': f"No markdown files generated in {output_dir}",
                'stop_reason': stop_reason,
            }

        self.logger.info(f"Found {len(md_files)} markdown files to process")

        # 8. Add each file to knowledge base
        added_docs = []
        failed_docs = []
        scrape_date = datetime.now().isoformat()

        for i, md_file in enumerate(md_files):
            if progress_callback:
                progress_callback(ProgressUpdate(
                    operation="scrape_url",
                    current=i,
                    total=len(md_files),
                    message="Adding scraped document",
                    item=md_file.name
                ))

            try:
                # Extract source URL from frontmatter
                source_url_for_file = self._extract_source_url_from_md(md_file)
                if not source_url_for_file:
                    source_url_for_file = url  # Fallback to base URL

                # Generate title from domain + page path
                if title:
                    # Use provided base title
                    doc_title = title
                else:
                    # Extract page name from URL path
                    parsed_source = urlparse(source_url_for_file)
                    page_path = parsed_source.path.strip('/')

                    # If it's the index/root, use domain name
                    if not page_path or page_path.lower() in ['index', 'index.html', 'index.htm']:
                        page_name = "Home"
                    else:
                        # Use the last part of the path as page name
                        page_name = page_path.split('/')[-1]
                        # Remove file extensions
                        page_name = page_name.replace('.html', '').replace('.htm', '').replace('.php', '')
                        # Clean up formatting
                        page_name = page_name.replace('_', ' ').replace('-', ' ').title()

                    # Combine domain + page name
                    doc_title = f"{domain} - {page_name}"

                # Add document with URL metadata
                doc = self._add_scraped_document(
                    filepath=str(md_file),
                    source_url=source_url_for_file,
                    title=doc_title,
                    tags=tags,
                    scrape_config=scrape_config_json,
                    scrape_date=scrape_date
                )
                added_docs.append(doc.doc_id)
                self.logger.info(f"Added: {doc.title} ({doc.doc_id})")

            except Exception as e:
                self.logger.error(f"Failed to add {md_file}: {e}")
                failed_docs.append(str(md_file))

        # 9. Return results
        status = 'success' if not failed_docs else ('partial' if added_docs else 'failed')
        if stop_reason == 'timeout':
            # A deliberate max_pages cap is a normal, successful outcome, but
            # hitting the wall-clock timeout means the crawl was cut short
            # before it necessarily finished what was asked - the caller
            # should be able to tell "got everything requested" apart from
            # "ran out of time partway through".
            status = 'partial' if added_docs else 'failed'

        result_dict = {
            'status': status,
            'url': url,
            'output_dir': str(output_dir),
            'files_scraped': len(md_files),
            'docs_added': len(added_docs),
            'docs_updated': 0,
            'docs_failed': len(failed_docs),
            'doc_ids': added_docs,
            'stop_reason': stop_reason,
        }

        if failed_docs:
            result_dict['error'] = f"{len(failed_docs)} files failed to add"
        if stop_reason == 'timeout':
            result_dict['error'] = (
                result_dict.get('error', '') +
                (' ' if result_dict.get('error') else '') +
                f"Crawl exceeded the {scrape_timeout_s:.0f}s time limit and was stopped early "
                f"after {pages_scraped} pages - results may be incomplete."
            ).strip()

        self.logger.info(f"Scraping complete: {status} - Added {len(added_docs)}/{len(md_files)} documents")

        return result_dict

    def rescrape_document(self, doc_id: str, progress_callback: ProgressCallback = None) -> dict:
        """Re-scrape an existing URL-sourced document.

        Args:
            doc_id: Document ID to re-scrape
            progress_callback: Optional progress callback

        Returns:
            Dictionary with re-scrape results (same format as scrape_url)
        """
        # Get document metadata
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        doc = self.documents[doc_id]

        # Check if document has source URL
        if not doc.source_url:
            raise ValueError(f"Document is not URL-sourced: {doc_id}")

        self.logger.info(f"Re-scraping document: {doc.title} (from {doc.source_url})")

        # Parse original scrape config
        scrape_config = {}
        if doc.scrape_config:
            try:
                scrape_config = json.loads(doc.scrape_config)
            except Exception as e:
                self.logger.warning(f"Failed to parse scrape config: {e}")

        # Scrape BEFORE touching the existing document. This used to remove
        # the old document first, unconditionally, with no rollback - a
        # dead/renamed page or a hung/failed scrape then permanently
        # destroyed the only copy of the content. Scraping first means a
        # failure just leaves the original document exactly as it was.
        # depth also no longer silently falls back to 50 (a much deeper
        # crawl than scrape_url's own default of 3) when the stored config
        # predates that field being recorded.
        result = self.scrape_url(
            url=doc.source_url,
            title=doc.title,
            tags=doc.tags,
            depth=scrape_config.get('depth', 3),
            limit=scrape_config.get('limit'),
            threads=scrape_config.get('threads', 10),
            delay=scrape_config.get('delay', 100),
            selector=scrape_config.get('selector'),
            progress_callback=progress_callback
        )

        # Add rescrape metadata to result
        result['rescrape'] = True
        result['old_doc_id'] = doc_id

        if result.get('status') == 'failed' or not result.get('doc_ids'):
            self.logger.warning(
                f"Re-scrape of {doc_id} produced no documents (status={result.get('status')}); "
                "keeping the original document unchanged"
            )
            result['old_doc_kept'] = True
            self.logger.info(f"Re-scrape complete: {result['status']}")
            return result

        if doc_id in result['doc_ids']:
            # Content-hash dedup (see add_document) matched the existing
            # document byte-for-byte - there is nothing new to swap in.
            self.logger.info(f"Re-scrape of {doc_id} found identical content; nothing to replace")
            result['old_doc_kept'] = True
            self.logger.info(f"Re-scrape complete: {result['status']}")
            return result

        # Only now, with a confirmed successful scrape of different content
        # in hand, is it safe to retire the old version.
        self.logger.info(f"Removing superseded document version: {doc_id}")
        self.remove_document(doc_id)
        result['old_doc_kept'] = False

        self.logger.info(f"Re-scrape complete: {result['status']}")
        return result

    def _discover_urls(self, base_url: str, config: Optional[dict] = None, max_pages: int = 100) -> set:
        """Discover all URLs at a website by crawling.

        Args:
            base_url: Starting URL to crawl
            config: Optional scrape configuration with depth, limit, etc.
            max_pages: Maximum number of pages to crawl (default: 100)

        Returns:
            Set of discovered URLs
        """
        from urllib.parse import urljoin, urlparse
        from bs4 import BeautifulSoup
        import requests

        discovered = set()
        to_visit = {base_url}
        visited = set()

        # Extract config parameters
        if config:
            depth = min(config.get('depth', 3), 5)  # Cap at 5 to prevent excessive crawling
            limit = config.get('limit')
            same_domain_only = config.get('same_domain_only', True)
            follow_links = config.get('follow_links', True)
        else:
            depth = 3
            limit = None
            same_domain_only = True
            follow_links = True

        # If follow_links is False, just check the single URL
        if not follow_links:
            depth = 1
            self.logger.info("follow_links=False, checking single URL only")

        base_parsed = urlparse(base_url)
        base_domain = base_parsed.netloc

        self.logger.info(f"URL Discovery: depth={depth}, same_domain={same_domain_only}, limit={limit}, max_pages={max_pages}")

        # Crawl with depth tracking
        current_depth = 0
        total_fetched = 0

        while to_visit and current_depth < depth and total_fetched < max_pages:
            current_level = to_visit.copy()
            to_visit.clear()

            self.logger.info(f"Crawl depth {current_depth + 1}/{depth}: {len(current_level)} URLs to check")

            for url in current_level:
                if url in visited:
                    continue

                if total_fetched >= max_pages:
                    self.logger.info(f"Reached max_pages limit ({max_pages})")
                    break

                visited.add(url)

                if not robots_allows(url):
                    self.logger.info(f"robots.txt disallows, skipping: {url}")
                    continue

                total_fetched += 1

                try:
                    # Fetch page with timeout
                    self.logger.debug(f"Fetching: {url}")
                    response = http_get_polite(url, timeout=15)

                    if response.status_code != 200:
                        self.logger.debug(f"Non-200 status ({response.status_code}): {url}")
                        continue

                    # Add to discovered (successful fetch)
                    discovered.add(url)
                    self.logger.debug(f"Discovered: {url}")

                    # Parse HTML to find links (only if following links and not at max depth)
                    if follow_links and current_depth < depth - 1:
                        if 'text/html' in response.headers.get('Content-Type', ''):
                            soup = BeautifulSoup(response.content, 'html.parser')

                            # Find all links
                            links_found = 0
                            for link in soup.find_all('a', href=True):
                                href = link['href']

                                # Convert relative URLs to absolute
                                absolute_url = urljoin(url, href)

                                # Remove fragments
                                absolute_url = absolute_url.split('#')[0]

                                # Skip if already visited or queued
                                if absolute_url in visited or absolute_url in to_visit:
                                    continue

                                # Parse the discovered URL
                                link_parsed = urlparse(absolute_url)

                                # Apply same_domain_only filter
                                if same_domain_only and link_parsed.netloc != base_domain:
                                    continue

                                # Apply limit filter
                                if limit and not absolute_url.startswith(limit):
                                    continue

                                # Skip non-HTTP(S) URLs
                                if link_parsed.scheme not in ['http', 'https']:
                                    continue

                                # Add to next level
                                to_visit.add(absolute_url)
                                links_found += 1

                            if links_found > 0:
                                self.logger.debug(f"Found {links_found} links on {url}")

                except requests.Timeout:
                    self.logger.warning(f"Timeout fetching {url}")
                    continue
                except Exception as e:
                    self.logger.debug(f"Error discovering {url}: {e}")
                    continue

            current_depth += 1

        self.logger.info(f"Discovery complete: {len(discovered)} URLs found (visited {total_fetched} pages at depth {current_depth})")
        return discovered

    def check_url_updates(self, auto_rescrape: bool = False, check_structure: bool = True) -> dict:
        """Check all URL-sourced documents for updates.

        Args:
            auto_rescrape: If True, automatically re-scrape changed URLs
            check_structure: If True, check for new/missing sub-pages (slower but comprehensive)

        Returns:
            Dictionary with update information:
            {
                'unchanged': [list of docs with no changes],
                'changed': [list of docs with updates available],
                'failed': [list of docs where check failed],
                'rescraped': [list of doc_ids that were re-scraped],
                'new_pages': [list of newly discovered URLs not in database],
                'missing_pages': [list of docs in database but no longer accessible],
                'scrape_sessions': [list of detected scrape sessions checked]
            }
        """
        from datetime import datetime, timezone
        import json

        results = {
            'unchanged': [],
            'changed': [],
            'failed': [],
            'rescraped': [],
            'new_pages': [],
            'missing_pages': [],
            'scrape_sessions': []
        }

        # Find all URL-sourced documents
        url_docs = [doc for doc in self.documents.values() if doc.source_url]

        if not url_docs:
            self.logger.info("No URL-sourced documents to check")
            return results

        self.logger.info(f"Checking {len(url_docs)} URL-sourced documents for updates")

        # Group documents by scrape session (same base URL)
        scrape_sessions = {}
        for doc in url_docs:
            # Parse scrape config to get base URL
            if doc.scrape_config:
                try:
                    config = json.loads(doc.scrape_config)
                    base_url = config.get('url', doc.source_url)
                except (json.JSONDecodeError, KeyError, AttributeError, TypeError):
                    base_url = doc.source_url
            else:
                base_url = doc.source_url

            # Group by base URL
            if base_url not in scrape_sessions:
                scrape_sessions[base_url] = {
                    'base_url': base_url,
                    'config': config if doc.scrape_config else None,
                    'docs': []
                }
            scrape_sessions[base_url]['docs'].append(doc)

        self.logger.info(f"Found {len(scrape_sessions)} scrape sessions to check")

        # Check each scrape session
        for session_key, session in scrape_sessions.items():
            session_result = {
                'base_url': session['base_url'],
                'docs_count': len(session['docs']),
                'changed': 0,
                'unchanged': 0,
                'new': 0,
                'missing': 0
            }
            results['scrape_sessions'].append(session_result)

            # Get stored URLs for this session
            stored_urls = {doc.source_url: doc for doc in session['docs']}

            # If check_structure is enabled, discover current URLs
            discovered_urls = set()
            if check_structure:
                try:
                    self.logger.info(f"Discovering URLs for: {session['base_url']}")
                    discovered_urls = self._discover_urls(
                        session['base_url'],
                        session['config']
                    )
                    self.logger.info(f"Discovered {len(discovered_urls)} URLs")
                except Exception as e:
                    self.logger.error(f"Failed to discover URLs for {session['base_url']}: {e}")

            # Check each stored document
            for doc in session['docs']:
                try:
                    import requests

                    # Try HEAD request first (faster)
                    response = requests.head(
                        doc.source_url, timeout=10, allow_redirects=True,
                        headers=http_headers(),
                    )

                    # Update last_checked timestamp
                    with self._lock:
                        cursor = self.db_conn.cursor()
                        cursor.execute("""
                            UPDATE documents
                            SET url_last_checked = ?
                            WHERE doc_id = ?
                        """, (datetime.now().isoformat(), doc.doc_id))
                        self.db_conn.commit()

                    # Check if page still exists
                    if response.status_code == 404:
                        self.logger.warning(f"Page no longer exists: {doc.source_url}")
                        results['missing_pages'].append({
                            'doc_id': doc.doc_id,
                            'title': doc.title,
                            'url': doc.source_url
                        })
                        session_result['missing'] += 1
                        continue

                    # Check Last-Modified header if available
                    page_changed = False
                    if 'Last-Modified' in response.headers:
                        from email.utils import parsedate_to_datetime
                        last_modified = parsedate_to_datetime(response.headers['Last-Modified'])

                        if doc.scrape_date:
                            scrape_dt = datetime.fromisoformat(doc.scrape_date)
                            # Ensure both datetimes are timezone-aware for comparison
                            if scrape_dt.tzinfo is None:
                                scrape_dt = scrape_dt.replace(tzinfo=timezone.utc)
                            if last_modified > scrape_dt:
                                page_changed = True
                                self.logger.info(f"Update available: {doc.title} ({doc.source_url})")
                                results['changed'].append({
                                    'doc_id': doc.doc_id,
                                    'title': doc.title,
                                    'url': doc.source_url,
                                    'last_modified': last_modified.isoformat(),
                                    'scraped_date': doc.scrape_date,
                                    'reason': 'content_modified'
                                })
                                session_result['changed'] += 1

                                # Auto-rescrape if requested
                                if auto_rescrape:
                                    self.logger.info(f"Auto-rescaping: {doc.title}")
                                    try:
                                        rescrape_result = self.rescrape_document(doc.doc_id)
                                        if rescrape_result['status'] == 'success':
                                            results['rescraped'].append(doc.doc_id)
                                    except Exception as e:
                                        self.logger.error(f"Auto-rescrape failed: {e}")

                    if not page_changed:
                        # No change detected
                        results['unchanged'].append({
                            'doc_id': doc.doc_id,
                            'title': doc.title,
                            'url': doc.source_url
                        })
                        session_result['unchanged'] += 1

                except Exception as e:
                    self.logger.error(f"Failed to check {doc.source_url}: {e}")
                    results['failed'].append({
                        'doc_id': doc.doc_id,
                        'title': doc.title,
                        'url': doc.source_url,
                        'error': str(e)
                    })

            # Check for new pages
            if check_structure and discovered_urls:
                new_urls = discovered_urls - set(stored_urls.keys())
                if new_urls:
                    self.logger.info(f"Found {len(new_urls)} new pages for {session['base_url']}")
                    for new_url in new_urls:
                        results['new_pages'].append({
                            'url': new_url,
                            'base_url': session['base_url'],
                            'scrape_config': session['config']
                        })
                        session_result['new'] += 1

                # Check for missing pages (in database but not discovered)
                if discovered_urls:  # Only if discovery was successful
                    missing_urls = set(stored_urls.keys()) - discovered_urls
                    # Filter out already detected 404s
                    existing_missing = {p['url'] for p in results['missing_pages']}
                    for missing_url in missing_urls:
                        if missing_url not in existing_missing:
                            doc = stored_urls[missing_url]
                            self.logger.warning(f"Page not discovered during crawl: {missing_url}")
                            results['missing_pages'].append({
                                'doc_id': doc.doc_id,
                                'title': doc.title,
                                'url': missing_url,
                                'reason': 'not_discovered'
                            })
                            session_result['missing'] += 1

        self.logger.info(
            f"Update check complete: {len(results['unchanged'])} unchanged, "
            f"{len(results['changed'])} changed, {len(results['new_pages'])} new pages, "
            f"{len(results['missing_pages'])} missing pages, {len(results['failed'])} failed"
        )

        return results

    def remove_document(self, doc_id: str) -> bool:
        """Remove a document from the knowledge base."""
        self.logger.info(f"Removing document: {doc_id}")

        if doc_id not in self.documents:
            self.logger.warning(f"Document not found for removal: {doc_id}")
            return False

        filename = self.documents[doc_id].filename

        # Remove from database (chunks cascade automatically)
        success = _retry_on_db_locked(self._remove_document_db, doc_id)

        if success:
            # Remove from in-memory index
            del self.documents[doc_id]

            # Prune this doc's chunks from the in-memory chunk cache. Without
            # this, self.chunks keeps serving the deleted content forever in
            # a long-running process: _build_bm25_index() only reloads from
            # the database when self.chunks is empty, so invalidating
            # self.bm25 alone doesn't pick up the DB-level cascade delete.
            self.chunks = [c for c in self.chunks if c.doc_id != doc_id]

            # Invalidate BM25 index (will be rebuilt on next search)
            self.bm25 = None

            # Remove this document's vectors from the shared embeddings index
            # in place. Nulling the in-memory index here (the old behaviour)
            # left the on-disk .faiss/.json files untouched but out of sync
            # with self.embeddings_index; the next add_document would then
            # see a "no index" state, build a fresh index from only its own
            # new chunks, and overwrite the full-corpus file with it -
            # silently destroying every other document's embeddings.
            if self.use_semantic:
                self._remove_doc_embeddings(doc_id)

            # Invalidate search caches
            self._invalidate_caches()

            self.logger.info(f"Successfully removed document {doc_id}: {filename}")

        return success

    def needs_reindex(self, filepath: str, doc_id: str) -> bool:
        """
        Check if a document needs re-indexing based on file modification time and content hash.

        Args:
            filepath: Path to the document file
            doc_id: Document ID to check

        Returns:
            True if the document needs re-indexing, False otherwise
        """
        doc = self.documents.get(doc_id)
        if not doc:
            return True  # Document doesn't exist, needs indexing

        # If no mtime/hash stored, can't check - assume needs reindex
        if doc.file_mtime is None or doc.file_hash is None:
            self.logger.info(f"Document {doc_id} has no update detection data, assuming needs reindex")
            return True

        # Quick check: modification time
        try:
            current_mtime = os.path.getmtime(filepath)
            if current_mtime <= doc.file_mtime:
                # File hasn't been modified since last index
                return False
        except OSError:
            # File doesn't exist or can't be accessed
            self.logger.warning(f"Cannot access file: {filepath}")
            return False

        # File was modified - do deep check with content hash
        try:
            current_hash = self._compute_file_hash(filepath)
            if current_hash == doc.file_hash:
                # Content is same despite mtime change (e.g., touched)
                self.logger.info(f"File mtime changed but content unchanged: {filepath}")
                return False
            else:
                # Content has actually changed
                self.logger.info(f"File content changed: {filepath}")
                return True
        except Exception as e:
            self.logger.error(f"Error computing hash for {filepath}: {e}")
            return False

    def _reindex_document_if_changed(self, filepath: str, title: Optional[str] = None, tags: Optional[list[str]] = None) -> DocumentMeta:
        """
        Re-index a document (matched by filepath) if its file content has changed
        since it was indexed, or add it if it doesn't exist yet. Used by the bulk
        directory-scan path (add_documents_bulk / check_for_updates) to catch
        files that were edited on disk. Not card-aware - it does a full
        remove-then-add (fresh content-hash doc_id) rather than an in-place
        update, so it does not preserve doc_id or history. For card documents,
        prefer update_document(card_id_or_doc_id, filepath) instead.

        Args:
            filepath: Path to the document file
            title: Optional title (if not provided, uses filename)
            tags: Optional list of tags

        Returns:
            DocumentMeta for the document (existing or newly indexed)
        """
        # Find existing doc by filepath
        existing_doc = None
        for doc in self.documents.values():
            if doc.filepath == filepath:
                existing_doc = doc
                break

        if not existing_doc:
            # Document doesn't exist, add it
            self.logger.info(f"Document not found, adding: {filepath}")
            return self.add_document(filepath, title, tags)

        if not self.needs_reindex(filepath, existing_doc.doc_id):
            # Document unchanged
            self.logger.info(f"Document unchanged, skipping reindex: {filepath}")
            return existing_doc

        # Document has changed, re-index it
        self.logger.info(f"Document changed, re-indexing: {filepath}")
        self.remove_document(existing_doc.doc_id)
        return self.add_document(filepath, title, tags)

    def update_document(self, card_id_or_doc_id: str, filepath: str,
                        title: Optional[str] = None, tags: Optional[list[str]] = None) -> DocumentMeta:
        """
        Replace an existing card's content (and all derived artifacts) at a
        stable logical identity.

        Resolves card_id_or_doc_id to an existing LIVE document (first as an
        exact doc_id, then as a card's logical id via get_document_by_card_id),
        ingests filepath as the new content, and retires the old document
        (marks it superseded_by the new doc, purges its stale entities, and
        rebuilds entity_relationships) so exactly one live document answers
        for that card afterwards.

        This is a whole-file replace, not a merge: the new file's content
        entirely replaces the old card's content. If the new file declares a
        different (or no) card id than the document being updated, the update
        is refused - update_document does not change a card's identity.

        Re-running the same file through update_document is idempotent: since
        doc_id is content-derived, ingesting identical content resolves to the
        same doc_id as the card already live, so no second supersede happens.

        Args:
            card_id_or_doc_id: The card's logical id (from its json block) or
                an exact doc_id of an existing, live document.
            filepath: Path to the new content to replace it with.
            title: Optional new title (defaults to the existing document's title).
            tags: Optional new tags (defaults to the existing document's tags).

        Returns:
            DocumentMeta for the (new or unchanged) live document.

        Raises:
            DocumentNotFoundError: If card_id_or_doc_id does not resolve to a
                live document. Use add_document to create a new card.
            KnowledgeBaseError: If the new file's declared card id conflicts
                with the document being updated.
        """
        old_doc = self.documents.get(card_id_or_doc_id)
        if old_doc is None or old_doc.superseded_by:
            old_doc = self.get_document_by_card_id(card_id_or_doc_id, include_superseded=False)

        if old_doc is None:
            raise DocumentNotFoundError(
                f"No live document or card found for '{card_id_or_doc_id}'. "
                f"Use add_document() to create a new card."
            )

        # Peek the new file's declared identity BEFORE ingesting anything.
        # add_document(replace=True) will supersede whatever live document
        # currently owns the incoming card_id - so if we didn't validate
        # first, a mismatched file could silently supersede an unrelated
        # card instead of (or in addition to) old_doc.
        resolved_path = Path(filepath).resolve()
        if not self._is_path_allowed(filepath):
            raise SecurityError(
                f"Path outside allowed directories. File must be within: {self.allowed_dirs}"
            )
        if not os.path.exists(str(resolved_path)):
            raise DocumentNotFoundError(f"File not found: {filepath}")
        peek_text, _, _, _ = self._extract_text_for_file(str(resolved_path))
        new_card_id = self._extract_card_id(peek_text)

        if old_doc.card_id and new_card_id != old_doc.card_id:
            raise KnowledgeBaseError(
                f"Refusing update: {filepath} declares card id {new_card_id!r}, "
                f"but you're updating card {old_doc.card_id!r} (doc {old_doc.doc_id}). "
                f"update_document() does not change a card's identity - fix the "
                f"file's id or use add_document() to create a separate card."
            )

        if new_card_id:
            colliding = self.get_document_by_card_id(new_card_id, include_superseded=False)
            if colliding and colliding.doc_id != old_doc.doc_id:
                raise KnowledgeBaseError(
                    f"Refusing update: {filepath} declares card id {new_card_id!r}, "
                    f"which already belongs to a different live document "
                    f"{colliding.doc_id} ('{colliding.title}'). update_document() "
                    f"will not supersede a document other than the one you asked "
                    f"to update ({old_doc.doc_id})."
                )

        resolved_title = title if title is not None else old_doc.title
        resolved_tags = tags if tags is not None else old_doc.tags

        new_doc = self.add_document(filepath, resolved_title, resolved_tags, replace=True)

        if new_doc.doc_id == old_doc.doc_id:
            # Identical content - add_document's own dedupe short-circuited.
            return new_doc

        # add_document(replace=True) already superseded old_doc for us when
        # old_doc.card_id was set (its own card-identity lookup resolves to
        # exactly old_doc, guaranteed by the checks above). Cover the case
        # where old_doc has no card_id (generic content replace) by
        # superseding explicitly here.
        if not old_doc.card_id and not old_doc.superseded_by:
            self._mark_superseded(old_doc.doc_id, new_doc.doc_id)

        return new_doc

    def update_document_title(self, doc_id: str, title: str) -> None:
        """
        Update the title of a document.

        Args:
            doc_id: Document ID
            title: New title for the document

        Raises:
            ValueError: If document not found
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        doc = self.documents[doc_id]
        old_title = doc.title

        # DB first, memory second: mutating in-memory before the write meant a
        # failed UPDATE left this process serving a title that was never saved.
        try:
            cursor = self.db_conn.cursor()
            cursor.execute(
                "UPDATE documents SET title = ? WHERE doc_id = ?",
                (title, doc_id)
            )
            self.db_conn.commit()
        except Exception:
            self.db_conn.rollback()
            doc.title = old_title
            self.logger.exception(f"Failed to update title for {doc_id}; rolled back")
            raise

        doc.title = title

        self.logger.info(f"Updated title for document {doc_id[:12]}: {title}")

    def update_document_tags(self, doc_id: str, tags: list[str]) -> None:
        """
        Update the tags for a document.

        Args:
            doc_id: Document ID
            tags: New list of tags (replaces existing tags)

        Raises:
            ValueError: If document not found
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        doc = self.documents[doc_id]
        old_tags = doc.tags

        # Write the DB first, then mutate memory only once it committed -
        # otherwise a failed write leaves this process reporting tags that were
        # never persisted.
        #
        # The column holds a JSON array: every reader parses it with
        # json.loads (see _load_documents / _reload_documents). This method
        # used to write ','.join(tags), which is not valid JSON, so a single
        # call poisoned the row and the next document reload - i.e. every new
        # session - died with JSONDecodeError and loaded no documents at all.
        try:
            cursor = self.db_conn.cursor()
            cursor.execute(
                "UPDATE documents SET tags = ? WHERE doc_id = ?",
                (json.dumps(tags or []), doc_id)
            )
            self.db_conn.commit()
        except Exception:
            self.db_conn.rollback()
            doc.tags = old_tags
            self.logger.exception(f"Failed to update tags for {doc_id}; rolled back")
            raise

        doc.tags = tags

        self.logger.info(f"Updated tags for document {doc_id[:12]}: {tags}")

    def suggest_tags(self, doc_id: str, confidence_threshold: float = 0.6) -> list[dict]:
        """
        Suggest tags for a document based on content analysis.

        Uses heuristic-based tag suggestion (no LLM required):
        - Detects hardware components (SID, VIC-II, CIA, 6502)
        - Identifies programming topics (assembly, BASIC, machine code)
        - Recognizes document types (reference, tutorial, guide)
        - Extracts memory addresses and registers

        Args:
            doc_id: Document ID to analyze
            confidence_threshold: Minimum confidence for suggestions (0.0-1.0)

        Returns:
            List of suggested tags with confidence scores:
            [
                {'tag': 'sid-programming', 'confidence': 0.95, 'category': 'hardware'},
                {'tag': 'assembly', 'confidence': 0.85, 'category': 'programming'},
                ...
            ]

        Raises:
            ValueError: If document not found
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        suggested_tags = []

        # Get sample of document content (first 3 chunks to avoid overhead)
        try:
            chunks = self._get_chunks_db(doc_id)
            sample_text = '\n'.join([c['content'] for c in chunks[:3]]) if chunks else ''
        except Exception as e:
            self.logger.warning(f"Failed to get chunks for {doc_id}: {e}")
            sample_text = ''

        if not sample_text:
            return suggested_tags

        text_lower = sample_text.lower()

        # Hardware detection
        hardware_patterns = {
            'sid-chip': (r'\bsid\b|\b6581\b', 0.9),
            'vic-ii': (r'\bvic-?ii\b|\bvic\s*2\b|\b6569\b|\b6567\b', 0.9),
            'cia': (r'\bcia\b|\b6526\b', 0.85),
            '6502-processor': (r'\b6502\b|\b6510\b', 0.9),
            'joystick': (r'\bjoystick\b|\bcontroller\b', 0.7),
            'disk-drive': (r'\bdisk\s*drive\b|\b1541\b|\b1571\b', 0.8),
        }

        for tag, (pattern, confidence) in hardware_patterns.items():
            if re.search(pattern, text_lower, re.IGNORECASE):
                if confidence >= confidence_threshold:
                    suggested_tags.append({
                        'tag': tag,
                        'confidence': confidence,
                        'category': 'hardware'
                    })

        # Programming topic detection
        programming_patterns = {
            'assembly': (r'\bassembly\b|\bmachine\s*code\b|\basync\b', 0.85),
            'basic': (r'\bbasic\b|\bprogram\s*line\b|\bline\s*numbers\b', 0.8),
            'graphics': (r'\bgraphics\b|\bsprite\b|\bbitmap\b|\bcharacter\s*set\b', 0.85),
            'sound-music': (r'\bsound\b|\bmusic\b|\baudio\b|\benvelop\b|\bsynthesis\b', 0.8),
            'interrupts': (r'\binterrupt\b|\birq\b|\bnmi\b', 0.9),
            'memory-management': (r'\bmemory\s*map\b|\bmemory\s*address\b|\b\$[0-9a-f]{4}\b', 0.75),
        }

        for tag, (pattern, confidence) in programming_patterns.items():
            if re.search(pattern, text_lower, re.IGNORECASE):
                if confidence >= confidence_threshold:
                    suggested_tags.append({
                        'tag': tag,
                        'confidence': confidence,
                        'category': 'programming'
                    })

        # Document type detection
        doc_type_patterns = {
            'reference': (r'\breference\b|\bopcode\s*table\b|\binstruction\s*set\b', 0.85),
            'tutorial': (r'\btutorial\b|\bhow\s*to\b|\bguide\b|\blearn\b', 0.75),
            'specification': (r'\bspecification\b|\bspec\b|\bdatasheet\b|\bmanual\b', 0.9),
            'code-example': (r'\bexample\b|\bcode\s*sample\b|\broutine\b', 0.7),
        }

        for tag, (pattern, confidence) in doc_type_patterns.items():
            if re.search(pattern, text_lower, re.IGNORECASE):
                if confidence >= confidence_threshold:
                    suggested_tags.append({
                        'tag': tag,
                        'confidence': confidence,
                        'category': 'document-type'
                    })

        # Difficulty level detection
        difficulty_patterns = {
            'beginner': (r'\bbeginning\b|\bstarter\b|\bintroduction\b|\bfundamentals?\b', 0.75),
            'intermediate': (r'\bintermediate\b|\badvanced-beginner\b', 0.7),
            'advanced': (r'\badvanced\b|\bexpert\b|\bdeep-dive\b', 0.75),
        }

        for tag, (pattern, confidence) in difficulty_patterns.items():
            if re.search(pattern, text_lower, re.IGNORECASE):
                if confidence >= confidence_threshold:
                    suggested_tags.append({
                        'tag': tag,
                        'confidence': confidence,
                        'category': 'difficulty'
                    })

        # Sort by confidence (descending)
        suggested_tags.sort(key=lambda x: x['confidence'], reverse=True)

        self.logger.info(f"Suggested {len(suggested_tags)} tags for document {doc_id[:12]}")

        return suggested_tags

    def add_tags_to_document(self, doc_id: str, new_tags: list[str],
                            merge: bool = True) -> list[str]:
        """
        Add tags to a document, optionally merging with existing tags.

        Args:
            doc_id: Document ID
            new_tags: Tags to add
            merge: If True, merge with existing tags. If False, replace tags.

        Returns:
            Updated list of all tags for the document

        Raises:
            ValueError: If document not found
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        # Normalize and deduplicate tags
        new_tags = list(set([t.lower().replace(' ', '-') for t in new_tags if t]))

        if merge:
            # Merge with existing tags
            existing_tags = set(self.documents[doc_id].tags or [])
            all_tags = list(existing_tags | set(new_tags))
        else:
            all_tags = new_tags

        # Update document
        self.update_document_tags(doc_id, all_tags)

        return all_tags

    def get_tags_by_category(self) -> dict[str, list[dict]]:
        """
        Get all tags organized by category for easier browsing.

        Returns:
            Dictionary with categories as keys and tag lists as values:
            {
                'hardware': [
                    {'tag': 'sid-chip', 'count': 15, 'documents': ['doc1', 'doc2', ...]},
                    ...
                ],
                'programming': [...],
                ...
            }
        """
        # Categorize known tags
        tag_categories = {
            'hardware': ['sid-chip', 'vic-ii', 'cia', '6502-processor', 'joystick', 'disk-drive'],
            'programming': ['assembly', 'basic', 'graphics', 'sound-music', 'interrupts', 'memory-management'],
            'document-type': ['reference', 'tutorial', 'specification', 'code-example'],
            'difficulty': ['beginner', 'intermediate', 'advanced'],
        }

        result = {}

        # For each category, count tag usage
        for category, known_tags in tag_categories.items():
            result[category] = []

            for tag in known_tags:
                # Count documents with this tag
                doc_ids = [doc_id for doc_id, doc in self.documents.items()
                          if tag in (doc.tags or [])]

                if doc_ids:  # Only include tags that are actually used
                    result[category].append({
                        'tag': tag,
                        'count': len(doc_ids),
                        'documents': doc_ids[:10]  # Show first 10 docs
                    })

            # Sort by count (descending)
            result[category].sort(key=lambda x: x['count'], reverse=True)

        # Add custom/user-defined tags that don't fit in categories
        all_known_tags = set()
        for tags_list in tag_categories.values():
            all_known_tags.update(tags_list)

        custom_tags = {}
        for doc in self.documents.values():
            for tag in (doc.tags or []):
                if tag not in all_known_tags:
                    if tag not in custom_tags:
                        custom_tags[tag] = []
                    custom_tags[tag].append(doc.doc_id)

        if custom_tags:
            result['custom'] = [
                {
                    'tag': tag,
                    'count': len(doc_ids),
                    'documents': doc_ids[:10]
                }
                for tag, doc_ids in sorted(custom_tags.items(),
                                         key=lambda x: len(x[1]),
                                         reverse=True)
            ]

        return result

    def _call_llm(self, prompt: str, max_tokens: int = 1024, temperature: float = 0.3) -> str:
        """
        Call LLM with a prompt (helper method for LLM operations).

        Args:
            prompt: Text prompt
            max_tokens: Maximum tokens to generate
            temperature: Sampling temperature (0.0-1.0)

        Returns:
            LLM response text

        Raises:
            ValueError: If LLM not available or call fails
        """
        # Check if LLM client is available
        if not hasattr(self, 'llm_client') or self.llm_client is None:
            # Try to initialize it
            try:
                from llm_integration import LLMClient
                self.llm_client = LLMClient()
            except Exception as e:
                raise ValueError(f"LLM client not available: {e}")

        try:
            response = self.llm_client.call(prompt, max_tokens=max_tokens, temperature=temperature)
            return response.strip()
        except Exception as e:
            self.logger.error(f"LLM call failed: {e}")
            raise ValueError(f"LLM call failed: {e}")

    def summarize_document(self, doc_id: str,
                          max_length: int = 500,
                          style: str = "technical") -> str:
        """
        Generate an AI summary of a document.

        Args:
            doc_id: Document ID to summarize
            max_length: Maximum summary length in words
            style: Summary style (technical, simple, or detailed)

        Returns:
            Summary text

        Raises:
            ValueError: If document not found or LLM not available
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        # Check if LLM client is available
        if not hasattr(self, 'llm_client') or self.llm_client is None:
            # Try to initialize it
            try:
                from llm_integration import LLMClient
                self.llm_client = LLMClient()
            except Exception as e:
                raise ValueError(f"LLM client not available: {e}")

        doc = self.documents[doc_id]

        # Get document content (first 10 chunks to keep context reasonable)
        chunks = self._get_chunks_db(doc_id)
        content_chunks = chunks[:10] if len(chunks) > 10 else chunks
        content = '\n\n'.join([chunk.content for chunk in content_chunks])

        # Truncate content if too long (max 20000 chars)
        if len(content) > 20000:
            content = content[:20000] + "..."

        # Build prompt based on style
        style_prompts = {
            "technical": "Provide a concise technical summary focusing on key concepts, technologies, and implementation details.",
            "simple": "Provide a simple, easy-to-understand summary suitable for beginners.",
            "detailed": "Provide a comprehensive detailed summary covering all major topics and subtopics."
        }

        style_instruction = style_prompts.get(style, style_prompts["technical"])

        prompt = f"""Summarize the following document in approximately {max_length} words.

Document Title: {doc.title}

{style_instruction}

Document Content:
{content}

Summary:"""

        try:
            summary = self.llm_client.call(prompt, max_tokens=max_length * 2, temperature=0.3)
            return summary.strip()
        except Exception as e:
            self.logger.error(f"Failed to generate summary: {e}")
            raise ValueError(f"Failed to generate summary: {e}")

    def check_all_updates(self, auto_update: bool = False) -> dict:
        """
        Check all indexed documents for updates.

        Args:
            auto_update: If True, automatically re-index changed documents

        Returns:
            Dictionary with lists of unchanged, changed, and missing documents
        """
        results = {
            'unchanged': [],
            'changed': [],
            'missing': [],
            'updated': []  # Only populated if auto_update=True
        }

        for doc_id, doc in list(self.documents.items()):
            filepath = doc.filepath

            # Check if file still exists
            if not os.path.exists(filepath):
                results['missing'].append({
                    'doc_id': doc_id,
                    'filepath': filepath,
                    'title': doc.title
                })
                continue

            # Check if needs reindex
            if self.needs_reindex(filepath, doc_id):
                results['changed'].append({
                    'doc_id': doc_id,
                    'filepath': filepath,
                    'title': doc.title
                })

                if auto_update:
                    try:
                        updated_doc = self._reindex_document_if_changed(filepath, doc.title, doc.tags)
                        results['updated'].append({
                            'doc_id': updated_doc.doc_id,
                            'filepath': filepath,
                            'title': updated_doc.title
                        })
                    except Exception as e:
                        self.logger.error(f"Failed to update {filepath}: {e}")
            else:
                results['unchanged'].append({
                    'doc_id': doc_id,
                    'filepath': filepath,
                    'title': doc.title
                })

        return results

    def add_documents_bulk(self, directory: str, pattern: str = "**/*.{pdf,txt,md,html,htm,xlsx,xls}",
                           tags: Optional[list[str]] = None, recursive: bool = True,
                           skip_duplicates: bool = True, progress_callback: ProgressCallback = None) -> dict:
        """
        Add multiple documents from a directory matching a glob pattern.

        Args:
            directory: Directory to search for documents
            pattern: Glob pattern (default: **/*.{pdf,txt})
            tags: Tags to apply to all documents
            recursive: Search subdirectories (default: True)
            skip_duplicates: Skip files with duplicate content (default: True)
            progress_callback: Optional callback for progress updates

        Returns:
            Dictionary with lists of added, skipped, and failed documents
        """
        from pathlib import Path

        dir_path = Path(directory).resolve()
        if not dir_path.exists():
            raise ValueError(f"Directory does not exist: {directory}")

        # Find matching files. Brace alternation must be expanded by hand -
        # pathlib.glob treats "{pdf,txt}" as a literal extension and matches
        # nothing, which made the default pattern a silent no-op.
        search_pattern = pattern if recursive else pattern.replace('**/', '')
        files = []
        seen_paths = set()
        for expanded in _expand_brace_pattern(search_pattern):
            for match in dir_path.glob(expanded):
                # Dedupe: overlapping alternatives can match the same file.
                if match not in seen_paths:
                    seen_paths.add(match)
                    files.append(match)

        results = {
            'added': [],
            'skipped': [],
            'failed': []
        }

        self.logger.info(f"Bulk add: found {len(files)} files matching pattern '{pattern}' in {directory}")

        # Report progress: Start
        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="add_documents_bulk",
                current=0,
                total=len(files),
                message=f"Starting bulk add of {len(files)} files",
                item=directory
            ))

        # Get worker count (configurable via environment variable, default to CPU count)
        max_workers = int(os.getenv('PARALLEL_WORKERS', str(os.cpu_count() or 4)))
        self.logger.info(f"Using {max_workers} workers for parallel processing")

        # Process files in parallel using ThreadPoolExecutor
        def process_file(file_path):
            """Process a single file and return result."""
            if not file_path.is_file():
                return None

            try:
                # Generate title from filename
                title = file_path.stem
                doc = self.add_document(str(file_path), title=title, tags=tags)

                return {
                    'status': 'added',
                    'doc_id': doc.doc_id,
                    'filepath': str(file_path),
                    'title': title,
                    'chunks': doc.total_chunks
                }

            except Exception as e:
                return {
                    'status': 'failed',
                    'filepath': str(file_path),
                    'error': str(e)
                }

        # Use ThreadPoolExecutor for parallel processing
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit all files for processing
            future_to_file = {executor.submit(process_file, fp): fp for fp in files}

            # Process completed tasks as they finish
            completed = 0
            seen_doc_ids = set()

            for future in as_completed(future_to_file):
                completed += 1
                file_path = future_to_file[future]

                # Report progress: Processing file
                if progress_callback:
                    progress_callback(ProgressUpdate(
                        operation="add_documents_bulk",
                        current=completed,
                        total=len(files),
                        message=f"Processing file {completed}/{len(files)}",
                        item=str(file_path.name)
                    ))

                try:
                    result = future.result()
                    if result is None:
                        continue

                    if result['status'] == 'added':
                        # Check for duplicates
                        if skip_duplicates and result['doc_id'] in seen_doc_ids:
                            results['skipped'].append({
                                'filepath': result['filepath'],
                                'reason': 'duplicate content',
                                'doc_id': result['doc_id']
                            })
                        else:
                            seen_doc_ids.add(result['doc_id'])
                            results['added'].append({
                                'doc_id': result['doc_id'],
                                'filepath': result['filepath'],
                                'title': result['title'],
                                'chunks': result['chunks']
                            })
                    elif result['status'] == 'failed':
                        results['failed'].append({
                            'filepath': result['filepath'],
                            'error': result['error']
                        })
                        self.logger.error(f"Failed to add {result['filepath']}: {result['error']}")

                except Exception as e:
                    results['failed'].append({
                        'filepath': str(file_path),
                        'error': str(e)
                    })
                    self.logger.error(f"Failed to process {file_path}: {e}")

        self.logger.info(f"Bulk add complete: {len(results['added'])} added, "
                        f"{len(results['skipped'])} skipped, {len(results['failed'])} failed")

        # Report progress: Complete
        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="add_documents_bulk",
                current=len(files),
                total=len(files),
                message=f"Bulk add complete: {len(results['added'])} added, "
                        f"{len(results['skipped'])} skipped, {len(results['failed'])} failed"
            ))

        return results

    def remove_documents_bulk(self, doc_ids: Optional[list[str]] = None,
                              tags: Optional[list[str]] = None) -> dict:
        """
        Remove multiple documents by doc IDs or tags.

        Args:
            doc_ids: List of document IDs to remove
            tags: Remove all documents with any of these tags

        Returns:
            Dictionary with lists of removed and failed document IDs
        """
        if not doc_ids and not tags:
            raise ValueError("Must provide either doc_ids or tags")

        results = {
            'removed': [],
            'failed': []
        }

        # Collect doc_ids to remove
        ids_to_remove = set()

        if doc_ids:
            ids_to_remove.update(doc_ids)

        if tags:
            # Find all documents with any of the specified tags
            for doc_id, doc in self.documents.items():
                if any(tag in doc.tags for tag in tags):
                    ids_to_remove.add(doc_id)

        self.logger.info(f"Bulk remove: removing {len(ids_to_remove)} documents")

        for doc_id in ids_to_remove:
            try:
                if self.remove_document(doc_id):
                    results['removed'].append(doc_id)
                else:
                    results['failed'].append({
                        'doc_id': doc_id,
                        'error': 'Document not found'
                    })
            except Exception as e:
                results['failed'].append({
                    'doc_id': doc_id,
                    'error': str(e)
                })
                self.logger.error(f"Failed to remove {doc_id}: {e}")

        self.logger.info(f"Bulk remove complete: {len(results['removed'])} removed, "
                        f"{len(results['failed'])} failed")

        return results

    def update_tags_bulk(self, doc_ids: Optional[list[str]] = None,
                         existing_tags: Optional[list[str]] = None,
                         add_tags: Optional[list[str]] = None,
                         remove_tags: Optional[list[str]] = None,
                         replace_tags: Optional[list[str]] = None) -> dict:
        """
        Update tags for multiple documents in bulk.

        Args:
            doc_ids: List of document IDs to update (if None, uses existing_tags to find docs)
            existing_tags: Find documents with any of these tags (alternative to doc_ids)
            add_tags: Tags to add to the documents
            remove_tags: Tags to remove from the documents
            replace_tags: Replace all tags with these tags

        Returns:
            Dictionary with lists of updated and failed document IDs

        Examples:
            # Add 'assembly' tag to specific documents
            kb.update_tags_bulk(doc_ids=['doc1', 'doc2'], add_tags=['assembly'])

            # Remove 'draft' tag from all documents that have it
            kb.update_tags_bulk(existing_tags=['draft'], remove_tags=['draft'])

            # Replace all tags with 'archive' for specific documents
            kb.update_tags_bulk(doc_ids=['doc1', 'doc2'], replace_tags=['archive'])

            # Add 'reviewed' and remove 'draft' for documents with 'pending' tag
            kb.update_tags_bulk(existing_tags=['pending'], add_tags=['reviewed'], remove_tags=['draft'])
        """
        if not doc_ids and not existing_tags:
            raise ValueError("Must provide either doc_ids or existing_tags")

        if not add_tags and not remove_tags and not replace_tags:
            raise ValueError("Must provide at least one of: add_tags, remove_tags, replace_tags")

        results = {
            'updated': [],
            'failed': []
        }

        # Collect doc_ids to update
        ids_to_update = set()

        if doc_ids:
            ids_to_update.update(doc_ids)

        if existing_tags:
            # Find all documents with any of the specified tags
            for doc_id, doc in self.documents.items():
                if any(tag in doc.tags for tag in existing_tags):
                    ids_to_update.add(doc_id)

        self.logger.info(f"Bulk tag update: updating {len(ids_to_update)} documents")

        for doc_id in ids_to_update:
            # Tracked outside the try so the handler can restore it: doc.tags is
            # mutated in memory before the UPDATE lands.
            old_tags = None
            try:
                if doc_id not in self.documents:
                    results['failed'].append({
                        'doc_id': doc_id,
                        'error': 'Document not found'
                    })
                    continue

                doc = self.documents[doc_id]
                old_tags = doc.tags.copy()

                # Apply tag operations
                if replace_tags is not None:
                    doc.tags = replace_tags.copy()
                else:
                    if add_tags:
                        # Add tags (avoiding duplicates)
                        for tag in add_tags:
                            if tag not in doc.tags:
                                doc.tags.append(tag)

                    if remove_tags:
                        # Remove tags
                        doc.tags = [tag for tag in doc.tags if tag not in remove_tags]

                # Update in database
                cursor = self.db_conn.cursor()
                cursor.execute("""
                    UPDATE documents
                    SET tags = ?
                    WHERE doc_id = ?
                """, (json.dumps(doc.tags), doc_id))
                self.db_conn.commit()

                results['updated'].append({
                    'doc_id': doc_id,
                    'old_tags': old_tags,
                    'new_tags': doc.tags
                })

                self.logger.debug(f"Updated tags for {doc_id}: {old_tags} -> {doc.tags}")

            except Exception as e:
                # Restore the pre-mutation tags. Without this, a failed write
                # left this process reporting tags that were never persisted -
                # they silently reverted on the next restart.
                doc = self.documents.get(doc_id)
                if doc is not None and old_tags is not None:
                    try:
                        self.db_conn.rollback()
                    except Exception:
                        pass
                    doc.tags = old_tags
                results['failed'].append({
                    'doc_id': doc_id,
                    'error': str(e)
                })
                self.logger.error(f"Failed to update tags for {doc_id}: {e}")

        self.logger.info(f"Bulk tag update complete: {len(results['updated'])} updated, "
                        f"{len(results['failed'])} failed")

        return results

    def auto_tag_document(self, doc_id: str, confidence_threshold: float = 0.7,
                         max_tags: int = 10, append: bool = True) -> dict:
        """
        Generate tags automatically using LLM analysis.

        Args:
            doc_id: Document to tag
            confidence_threshold: Minimum confidence to accept tag (0.0-1.0)
            max_tags: Maximum number of tags to suggest
            append: If True, append to existing tags; if False, replace

        Returns:
            {
                'doc_id': str,
                'suggested_tags': [{'tag': str, 'confidence': float}, ...],
                'applied_tags': [str],
                'skipped_tags': [str],  # Below confidence threshold
                'existing_tags': [str],
                'new_tags': [str]  # Final tag list
            }

        Example:
            result = kb.auto_tag_document('doc123', confidence_threshold=0.7)
            # {
            #     'suggested_tags': [
            #         {'tag': 'sid-programming', 'confidence': 0.95},
            #         {'tag': 'assembly', 'confidence': 0.88},
            #         {'tag': 'beginner', 'confidence': 0.65}  # Below threshold
            #     ],
            #     'applied_tags': ['sid-programming', 'assembly'],
            #     'skipped_tags': ['beginner'],
            #     ...
            # }
        """
        # Import LLM client
        try:
            from llm_integration import get_llm_client
        except ImportError:
            raise ImportError("llm_integration module not found. Auto-tagging requires LLM integration.")

        # Get LLM client
        llm_client = get_llm_client()
        if not llm_client:
            raise ValueError("LLM not configured. Set LLM_PROVIDER and appropriate API key.")

        # Get document
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        doc = self.documents[doc_id]

        # Get sample text (first 3 chunks for analysis)
        chunks = self._get_chunks_db(doc_id)
        sample_chunks = chunks[:3] if len(chunks) > 3 else chunks
        sample_text = "\n\n".join([c.content for c in sample_chunks])

        # Limit text size (first 3000 chars)
        if len(sample_text) > 3000:
            sample_text = sample_text[:3000] + "..."

        # Build prompt
        prompt = f"""Analyze this Commodore 64 technical documentation and suggest relevant tags.

Consider these categories:
1. Hardware components (sid, vic-ii, cia, 6502, memory, cartridge, disk-drive, etc.)
2. Programming topics (assembly, basic, machine-code, graphics, sound, sprites, etc.)
3. Document type (tutorial, reference, manual, guide, example, etc.)
4. Difficulty level (beginner, intermediate, advanced, expert)
5. Content area (programming, hardware, history, repair, modification, etc.)

Document title: {doc.title}
Document filename: {doc.filename}

Sample text:
{sample_text}

Return a JSON object with this structure:
{{
    "tags": [
        {{"tag": "sid-programming", "confidence": 0.95, "reason": "Document extensively discusses SID chip programming"}},
        {{"tag": "assembly", "confidence": 0.88, "reason": "Contains assembly code examples"}}
    ]
}}

Important:
- Use lowercase with hyphens (e.g., "sid-programming" not "SID Programming")
- Provide {max_tags} or fewer tags
- Include confidence score (0.0-1.0) for each tag
- Brief reason for each tag suggestion
- Return ONLY the JSON, no other text"""

        # Call LLM
        self.logger.info(f"Auto-tagging document {doc_id} ({doc.title})")

        try:
            response = llm_client.call_json(prompt, max_tokens=1024, temperature=0.3)
        except Exception as e:
            self.logger.error(f"LLM call failed: {e}")
            raise ValueError(f"Failed to generate tags: {e}")

        # Parse response
        suggested_tags = response.get('tags', [])

        # Filter by confidence
        high_confidence_tags = [
            t for t in suggested_tags
            if t['confidence'] >= confidence_threshold
        ]

        low_confidence_tags = [
            t for t in suggested_tags
            if t['confidence'] < confidence_threshold
        ]

        # Extract tag names
        applied_tag_names = [t['tag'] for t in high_confidence_tags]
        skipped_tag_names = [t['tag'] for t in low_confidence_tags]

        # Get existing tags
        existing_tags = doc.tags.copy()

        # Apply tags
        if append:
            # Add new tags to existing (avoid duplicates)
            new_tags = existing_tags.copy()
            for tag in applied_tag_names:
                if tag not in new_tags:
                    new_tags.append(tag)
        else:
            # Replace all tags
            new_tags = applied_tag_names

        # Update document
        doc.tags = new_tags

        # Update in database
        cursor = self.db_conn.cursor()
        cursor.execute("""
            UPDATE documents
            SET tags = ?
            WHERE doc_id = ?
        """, (json.dumps(new_tags), doc_id))
        self.db_conn.commit()

        result = {
            'doc_id': doc_id,
            'doc_title': doc.title,
            'suggested_tags': suggested_tags,
            'applied_tags': applied_tag_names,
            'skipped_tags': skipped_tag_names,
            'existing_tags': existing_tags,
            'new_tags': new_tags,
            'confidence_threshold': confidence_threshold
        }

        self.logger.info(f"Auto-tagged {doc_id}: applied {len(applied_tag_names)} tags, "
                        f"skipped {len(skipped_tag_names)} low-confidence tags")

        return result

    def auto_tag_all_documents(self, confidence_threshold: float = 0.7,
                               max_tags: int = 10, append: bool = True,
                               skip_tagged: bool = True, max_docs: Optional[int] = None) -> dict:
        """
        Bulk auto-tag all documents using LLM.

        Args:
            confidence_threshold: Minimum confidence to accept tag (0.0-1.0)
            max_tags: Maximum tags per document
            append: If True, append to existing tags; if False, replace
            skip_tagged: If True, skip documents that already have tags
            max_docs: Maximum number of documents to process (None = all)

        Returns:
            {
                'processed': int,
                'skipped': int,
                'failed': int,
                'total_tags_added': int,
                'results': [list of individual results]
            }

        Example:
            results = kb.auto_tag_all_documents(
                confidence_threshold=0.7,
                skip_tagged=True,
                max_docs=10
            )
        """
        results = {
            'processed': 0,
            'skipped': 0,
            'failed': 0,
            'total_tags_added': 0,
            'results': []
        }

        # Get documents to process
        docs_to_process = []

        for doc_id, doc in self.documents.items():
            # Skip if already has tags (optional)
            if skip_tagged and doc.tags:
                results['skipped'] += 1
                continue

            docs_to_process.append(doc_id)

            # Limit number of documents
            if max_docs and len(docs_to_process) >= max_docs:
                break

        self.logger.info(f"Auto-tagging {len(docs_to_process)} documents "
                        f"(skipped {results['skipped']} already tagged)")

        # Process each document
        for i, doc_id in enumerate(docs_to_process, 1):
            try:
                self.logger.info(f"Auto-tagging {i}/{len(docs_to_process)}: {doc_id}")

                result = self.auto_tag_document(
                    doc_id,
                    confidence_threshold=confidence_threshold,
                    max_tags=max_tags,
                    append=append
                )

                # Count new tags added
                tags_added = len(set(result['new_tags']) - set(result['existing_tags']))
                results['total_tags_added'] += tags_added

                results['processed'] += 1
                results['results'].append(result)

            except Exception as e:
                results['failed'] += 1
                self.logger.error(f"Failed to auto-tag {doc_id}: {e}")
                results['results'].append({
                    'doc_id': doc_id,
                    'error': str(e)
                })

        self.logger.info(f"Auto-tagging complete: processed={results['processed']}, "
                        f"failed={results['failed']}, tags_added={results['total_tags_added']}")

        return results

    def generate_summary(self, doc_id: str, summary_type: str = 'brief',
                        force_regenerate: bool = False) -> str:
        """
        Generate an AI-powered summary of a document.

        Args:
            doc_id: Document ID to summarize
            summary_type: Type of summary ('brief', 'detailed', 'bullet')
                - 'brief': 1-2 paragraph overview (200-300 words)
                - 'detailed': Comprehensive summary with key points (500-800 words)
                - 'bullet': Bullet-point summary of main topics
            force_regenerate: If True, regenerate even if cached summary exists

        Returns:
            Summary text as a string

        Raises:
            ValueError: If document not found or LLM not configured
            DocumentNotFoundError: If document doesn't exist

        Examples:
            # Generate brief summary
            summary = kb.generate_summary('doc123', 'brief')

            # Get detailed summary
            summary = kb.generate_summary('doc456', 'detailed')

            # Force regeneration (bypass cache)
            summary = kb.generate_summary('doc789', 'brief', force_regenerate=True)
        """
        # Validate document exists
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        doc = self.documents[doc_id]

        # Check for cached summary
        if not force_regenerate:
            cursor = self.db_conn.cursor()
            cursor.execute("""
                SELECT summary_text FROM document_summaries
                WHERE doc_id = ? AND summary_type = ?
            """, (doc_id, summary_type))
            result = cursor.fetchone()
            if result:
                self.logger.debug(f"Using cached summary for {doc_id} ({summary_type})")
                return result[0]

        # Import LLM client
        try:
            from llm_integration import get_llm_client
        except ImportError:
            raise ImportError("llm_integration module not found. Summarization requires LLM integration.")

        # Get LLM client
        llm_client = get_llm_client()
        if not llm_client:
            raise ValueError("LLM not configured. Set LLM_PROVIDER and appropriate API key.")

        # Get document content
        chunks = self._get_chunks_db(doc_id)
        if not chunks:
            raise ValueError(f"No content found for document: {doc_id}")

        # For brief summaries, use first 5 chunks; for detailed, use more
        if summary_type == 'brief':
            sample_chunks = chunks[:5]
            word_limit = 300
            length_guidance = "1-2 paragraphs, approximately 200-300 words"
        elif summary_type == 'detailed':
            sample_chunks = chunks[:15] if len(chunks) > 15 else chunks
            word_limit = 800
            length_guidance = "3-5 paragraphs with detailed explanations, approximately 500-800 words"
        elif summary_type == 'bullet':
            sample_chunks = chunks[:10]
            word_limit = 400
            length_guidance = "8-12 bullet points covering main topics"
        else:
            raise ValueError(f"Invalid summary type: {summary_type}. Must be 'brief', 'detailed', or 'bullet'.")

        # Join content
        content = "\n\n".join([c.content for c in sample_chunks])

        # Limit content size to first 10k chars to control API costs
        if len(content) > 10000:
            content = content[:10000] + "..."

        # Build prompt based on summary type
        if summary_type == 'bullet':
            prompt = f"""Create a bullet-point summary of this Commodore 64 technical documentation.

Document Title: {doc.title}
Document Type: {doc.file_type}

Content:
{content}

Create a concise bullet-point summary with 8-12 main topics. Each bullet should be clear and informative.
Return ONLY the bullet points, one per line, starting with "- ". No introduction or explanation needed."""

        else:
            prompt = f"""Create a {summary_type} summary of this Commodore 64 technical documentation.

Document Title: {doc.title}
Document Type: {doc.file_type}

Content:
{content}

Write a {summary_type} summary that is {length_guidance}.
Focus on:
- Key concepts and main topics
- Technical details relevant to programmers
- Important procedures or examples
- Practical applications

Return ONLY the summary text, no preamble."""

        # Call LLM
        self.logger.info(f"Generating {summary_type} summary for {doc_id} ({doc.title})")

        try:
            summary_text = llm_client.call(prompt, max_tokens=word_limit + 200, temperature=0.4)
        except Exception as e:
            self.logger.error(f"LLM call failed: {e}")
            raise ValueError(f"Failed to generate summary: {e}")

        # Clean up summary text
        if not summary_text or not summary_text.strip():
            raise ValueError("LLM returned empty summary")

        summary_text = summary_text.strip()

        # Store summary in database
        cursor = self.db_conn.cursor()
        try:
            # Get model name from LLM client
            model = os.getenv('LLM_MODEL', 'unknown')

            cursor.execute("""
                INSERT OR REPLACE INTO document_summaries
                (doc_id, summary_type, summary_text, generated_at, model, token_count)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (doc_id, summary_type, summary_text, datetime.now().isoformat(),
                  model, len(summary_text.split())))

            self.db_conn.commit()
            self.logger.info(f"Saved {summary_type} summary for {doc_id}")

        except Exception as e:
            self.logger.exception("Failed to save summary to database")
            # Return summary even if save failed
            pass

        return summary_text

    def generate_summary_all(self, summary_types: Optional[list[str]] = None,
                            force_regenerate: bool = False,
                            max_docs: Optional[int] = None) -> dict:
        """
        Bulk generate summaries for all documents.

        Args:
            summary_types: List of summary types to generate (['brief'], ['brief', 'detailed'], etc.)
                          Default: ['brief']
            force_regenerate: If True, regenerate all summaries
            max_docs: Maximum number of documents to process (None = all)

        Returns:
            {
                'processed': int,
                'failed': int,
                'total_summaries': int,
                'by_type': {'brief': int, 'detailed': int, 'bullet': int},
                'results': [list of individual results]
            }

        Example:
            results = kb.generate_summary_all(
                summary_types=['brief', 'detailed'],
                max_docs=50
            )
        """
        if summary_types is None:
            summary_types = ['brief']

        results = {
            'processed': 0,
            'failed': 0,
            'total_summaries': 0,
            'by_type': {st: 0 for st in summary_types},
            'results': []
        }

        # Get documents to process
        docs_to_process = list(self.documents.keys())

        if max_docs:
            docs_to_process = docs_to_process[:max_docs]

        self.logger.info(f"Generating summaries for {len(docs_to_process)} documents "
                        f"(types: {', '.join(summary_types)})")

        # Process each document
        for i, doc_id in enumerate(docs_to_process, 1):
            doc_results = {
                'doc_id': doc_id,
                'title': self.documents[doc_id].title,
                'summaries': {}
            }

            for summary_type in summary_types:
                try:
                    self.logger.info(f"[{i}/{len(docs_to_process)}] {doc_id} ({summary_type})")

                    summary = self.generate_summary(
                        doc_id,
                        summary_type=summary_type,
                        force_regenerate=force_regenerate
                    )

                    doc_results['summaries'][summary_type] = {
                        'success': True,
                        'length': len(summary),
                        'word_count': len(summary.split())
                    }

                    results['total_summaries'] += 1
                    results['by_type'][summary_type] += 1

                except Exception as e:
                    results['failed'] += 1
                    self.logger.error(f"Failed to summarize {doc_id} ({summary_type}): {e}")
                    doc_results['summaries'][summary_type] = {
                        'success': False,
                        'error': str(e)
                    }

            results['processed'] += 1
            results['results'].append(doc_results)

        self.logger.info(f"Summary generation complete: processed={results['processed']}, "
                        f"failed={results['failed']}, total_summaries={results['total_summaries']}")

        return results

    def get_summary(self, doc_id: str, summary_type: str = 'brief') -> Optional[str]:
        """
        Retrieve a cached summary without regenerating.

        Args:
            doc_id: Document ID
            summary_type: Type of summary ('brief', 'detailed', 'bullet')

        Returns:
            Summary text if it exists, None otherwise
        """
        cursor = self.db_conn.cursor()
        cursor.execute("""
            SELECT summary_text FROM document_summaries
            WHERE doc_id = ? AND summary_type = ?
        """, (doc_id, summary_type))
        result = cursor.fetchone()
        return result[0] if result else None

    def export_documents_bulk(self, doc_ids: Optional[list[str]] = None,
                              tags: Optional[list[str]] = None,
                              format: str = 'json') -> str:
        """
        Export metadata for multiple documents.

        Args:
            doc_ids: List of document IDs to export (if None, uses tags or exports all)
            tags: Export documents with any of these tags
            format: Export format ('json', 'csv', or 'markdown')

        Returns:
            Exported data as a string

        Examples:
            # Export all documents as JSON
            data = kb.export_documents_bulk(format='json')

            # Export documents with 'reference' tag as CSV
            data = kb.export_documents_bulk(tags=['reference'], format='csv')

            # Export specific documents as Markdown
            data = kb.export_documents_bulk(doc_ids=['doc1', 'doc2'], format='markdown')
        """
        # Collect docs to export
        docs_to_export = []

        if doc_ids:
            # Export specific documents
            for doc_id in doc_ids:
                if doc_id in self.documents:
                    docs_to_export.append(self.documents[doc_id])
        elif tags:
            # Export documents with specified tags
            for doc in self.documents.values():
                if any(tag in doc.tags for tag in tags):
                    docs_to_export.append(doc)
        else:
            # Export all documents
            docs_to_export = list(self.documents.values())

        self.logger.info(f"Bulk export: exporting {len(docs_to_export)} documents as {format}")

        # Format the output
        if format == 'json':
            export_data = []
            for doc in docs_to_export:
                export_data.append({
                    'doc_id': doc.doc_id,
                    'filename': doc.filename,
                    'title': doc.title,
                    'filepath': doc.filepath,
                    'file_type': doc.file_type,
                    'total_pages': doc.total_pages,
                    'total_chunks': doc.total_chunks,
                    'indexed_at': doc.indexed_at,
                    'tags': doc.tags,
                    'author': doc.author,
                    'subject': doc.subject,
                    'creator': doc.creator,
                    'creation_date': doc.creation_date
                })
            return json.dumps(export_data, indent=2)

        elif format == 'csv':
            import csv
            from io import StringIO

            output = StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow(['doc_id', 'filename', 'title', 'filepath', 'file_type',
                           'total_pages', 'total_chunks', 'indexed_at', 'tags',
                           'author', 'subject', 'creator', 'creation_date'])

            # Write data
            for doc in docs_to_export:
                writer.writerow([
                    doc.doc_id,
                    doc.filename,
                    doc.title,
                    doc.filepath,
                    doc.file_type,
                    doc.total_pages,
                    doc.total_chunks,
                    doc.indexed_at,
                    ', '.join(doc.tags),
                    doc.author or '',
                    doc.subject or '',
                    doc.creator or '',
                    doc.creation_date or ''
                ])

            return output.getvalue()

        elif format == 'markdown':
            lines = []
            lines.append("# Document Export")
            lines.append(f"\n**Total Documents:** {len(docs_to_export)}")
            lines.append(f"**Exported:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            lines.append("---\n")

            for i, doc in enumerate(docs_to_export, 1):
                lines.append(f"## {i}. {doc.title}")
                lines.append(f"- **ID:** `{doc.doc_id}`")
                lines.append(f"- **Filename:** {doc.filename}")
                lines.append(f"- **Type:** {doc.file_type}")
                lines.append(f"- **Pages:** {doc.total_pages}")
                lines.append(f"- **Chunks:** {doc.total_chunks}")
                lines.append(f"- **Tags:** {', '.join(doc.tags) if doc.tags else 'None'}")
                if doc.author:
                    lines.append(f"- **Author:** {doc.author}")
                if doc.subject:
                    lines.append(f"- **Subject:** {doc.subject}")
                lines.append(f"- **Indexed:** {doc.indexed_at}")
                lines.append(f"- **Path:** `{doc.filepath}`")
                lines.append("")

            return '\n'.join(lines)

        else:
            raise ValueError(f"Unsupported format: {format}. Use 'json', 'csv', or 'markdown'")
