#!/usr/bin/env python3
"""
TDZ C64 Knowledge - MCP Server
A Model Context Protocol server for searching C64 documentation.
"""

import os
import sys
import json
import re
import hashlib
import logging
import time
import sqlite3
import threading
import queue
from pathlib import Path
from typing import Optional, Callable
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()  # Load .env file from current directory

# Import version information
from version import __build_date__, get_full_version_string

# Anomaly detection support
try:
    from anomaly_detector import AnomalyDetector, CheckResult, AnomalyScore, Baseline
    ANOMALY_SUPPORT = True
except ImportError:
    ANOMALY_SUPPORT = False
    AnomalyDetector = None
    print("Warning: anomaly_detector not found. Anomaly detection disabled.", file=sys.stderr)

# Caching support
try:
    from cachetools import TTLCache
    CACHE_SUPPORT = True
except ImportError:
    CACHE_SUPPORT = False
    print("Warning: cachetools not installed. Search caching disabled.", file=sys.stderr)

from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import (
    Tool,
    TextContent,
    Resource,
)

# Optional imports for PDF support
try:
    from pypdf import PdfReader
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    print("Warning: pypdf not installed. PDF support disabled.", file=sys.stderr)

# PDF table extraction support
try:
    import pdfplumber
    PDFPLUMBER_SUPPORT = True
except ImportError:
    PDFPLUMBER_SUPPORT = False
    print("Warning: pdfplumber not installed. Table extraction disabled.", file=sys.stderr)

# BM25 search support
try:
    from rank_bm25 import BM25Okapi
    BM25_SUPPORT = True
except ImportError:
    BM25_SUPPORT = False
    print("Warning: rank-bm25 not installed. Using simple search.", file=sys.stderr)

# NLTK for query preprocessing
try:
    import nltk
    from nltk.corpus import stopwords
    from nltk.stem import PorterStemmer
    from nltk.tokenize import word_tokenize
    NLTK_SUPPORT = True

    # Ensure NLTK data is available
    try:
        stopwords.words('english')
    except LookupError:
        nltk.download('stopwords', quiet=True)
        nltk.download('punkt', quiet=True)
        nltk.download('punkt_tab', quiet=True)
except ImportError:
    NLTK_SUPPORT = False
    print("Warning: nltk not installed. Query preprocessing disabled.", file=sys.stderr)

# Semantic search support
try:
    from sentence_transformers import SentenceTransformer
    import faiss
    import numpy as np
    SEMANTIC_SUPPORT = True
except ImportError:
    SEMANTIC_SUPPORT = False
    print("Warning: sentence-transformers or faiss-cpu not installed. Semantic search disabled.", file=sys.stderr)

# Fuzzy search support
try:
    from rapidfuzz import fuzz
    FUZZY_SUPPORT = True
except ImportError:
    FUZZY_SUPPORT = False
    print("Warning: rapidfuzz not installed. Fuzzy search disabled.", file=sys.stderr)

# OCR support
try:
    import pytesseract
    from pdf2image import convert_from_path
    OCR_SUPPORT = True
except ImportError:
    OCR_SUPPORT = False
    print("Warning: pytesseract/pdf2image/Pillow not installed. OCR disabled.", file=sys.stderr)


# Custom Exceptions
class KnowledgeBaseError(Exception):
    """Base exception for knowledge base errors."""
    pass


class DocumentNotFoundError(KnowledgeBaseError):
    """Raised when a document is not found."""
    pass


class ChunkNotFoundError(KnowledgeBaseError):
    """Raised when a chunk is not found."""
    pass


class UnsupportedFileTypeError(KnowledgeBaseError):
    """Raised when an unsupported file type is provided."""
    pass


class IndexCorruptedError(KnowledgeBaseError):
    """Raised when the index is corrupted."""
    pass


class SecurityError(KnowledgeBaseError):
    """Raised when a security violation is detected."""
    pass


@dataclass
class DocumentChunk:
    """A searchable chunk of a document."""
    doc_id: str
    filename: str
    title: str
    chunk_id: int
    page: Optional[int]
    content: str
    word_count: int


@dataclass
class DocumentMeta:
    """Metadata about an indexed document."""
    doc_id: str
    filename: str
    title: str
    filepath: str
    file_type: str
    total_pages: Optional[int]
    total_chunks: int
    indexed_at: str
    tags: list[str]
    # PDF metadata (optional)
    author: Optional[str] = None
    subject: Optional[str] = None
    creator: Optional[str] = None
    creation_date: Optional[str] = None
    # Update detection fields
    file_mtime: Optional[float] = None  # File modification time
    file_hash: Optional[str] = None  # MD5 hash of file content
    # URL scraping fields (optional)
    source_url: Optional[str] = None  # Original URL if document was scraped
    scrape_date: Optional[str] = None  # ISO timestamp of last scrape
    scrape_config: Optional[str] = None  # JSON string with scraping config
    scrape_status: Optional[str] = None  # 'success', 'partial', 'failed'
    scrape_error: Optional[str] = None  # Error message if scrape failed
    url_last_checked: Optional[str] = None  # ISO timestamp of last update check
    url_content_hash: Optional[str] = None  # Hash of scraped content for change detection


@dataclass
class ProgressUpdate:
    """Progress update for long-running operations."""
    operation: str  # Operation name (e.g., "add_document", "add_documents_bulk")
    current: int  # Current progress (items processed)
    total: int  # Total items to process
    message: str  # Status message
    item: Optional[str] = None  # Current item being processed (e.g., filename)
    percentage: float = 0.0  # Percentage complete (0-100)

    def __post_init__(self):
        """Calculate percentage after initialization."""
        if self.total > 0:
            self.percentage = (self.current / self.total) * 100.0


# Type alias for progress callback function
ProgressCallback = Optional[Callable[[ProgressUpdate], None]]


class KnowledgeBase:
    """Manages the document index and search."""

    def __init__(self, data_dir: str):
        self.data_dir = Path(data_dir)
        self.data_dir.mkdir(parents=True, exist_ok=True)

        # Setup logging
        log_file = self.data_dir / "server.log"
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler(sys.stderr)
            ]
        )
        self.logger = logging.getLogger(__name__)
        self.logger.info("=" * 60)
        self.logger.info(f"{get_full_version_string()}")
        self.logger.info(f"Build Date: {__build_date__}")
        self.logger.info("=" * 60)
        self.logger.info(f"Initializing KnowledgeBase at {data_dir}")

        # Thread safety for parallel document processing
        self._lock = threading.Lock()

        # Database setup
        self.db_file = self.data_dir / "knowledge_base.db"
        self.db_conn = None
        self._init_database()

        # Legacy file paths (for migration)
        self.index_file = self.data_dir / "index.json"
        self.chunks_dir = self.data_dir / "chunks"

        self.documents: dict[str, DocumentMeta] = {}
        self.chunks: list[DocumentChunk] = []  # Only loaded on demand for BM25
        self.bm25 = None  # BM25 index, built on demand

        # Initialize caching layer
        if CACHE_SUPPORT:
            cache_size = int(os.getenv('SEARCH_CACHE_SIZE', '100'))
            cache_ttl = int(os.getenv('SEARCH_CACHE_TTL', '300'))  # 5 minutes
            self._search_cache = TTLCache(maxsize=cache_size, ttl=cache_ttl)
            self._similar_cache = TTLCache(maxsize=cache_size, ttl=cache_ttl)
            self._semantic_cache = TTLCache(maxsize=cache_size, ttl=cache_ttl)
            self._hybrid_cache = TTLCache(maxsize=cache_size, ttl=cache_ttl)
            self._faceted_cache = TTLCache(maxsize=cache_size, ttl=cache_ttl)

            # Query embedding cache for semantic search (1 hour TTL for embeddings)
            embedding_cache_ttl = int(os.getenv('EMBEDDING_CACHE_TTL', '3600'))  # 1 hour
            self._embedding_cache = TTLCache(maxsize=cache_size, ttl=embedding_cache_ttl)

            # Entity extraction result cache (24 hour TTL for expensive LLM operations)
            entity_cache_ttl = int(os.getenv('ENTITY_CACHE_TTL', '86400'))  # 24 hours
            self._entity_cache = TTLCache(maxsize=50, ttl=entity_cache_ttl)

            # Health check cache (5 minute TTL for system metrics)
            health_cache_ttl = int(os.getenv('HEALTH_CACHE_TTL', '300'))  # 5 minutes
            self._health_cache = TTLCache(maxsize=1, ttl=health_cache_ttl)

            # Stats cache (1 minute TTL for statistics)
            stats_cache_ttl = int(os.getenv('STATS_CACHE_TTL', '60'))  # 1 minute
            self._stats_cache = TTLCache(maxsize=1, ttl=stats_cache_ttl)

            self.logger.info(f"Search result caching enabled (size={cache_size}, ttl={cache_ttl}s)")
            self.logger.info(f"Embedding caching enabled (size={cache_size}, ttl={embedding_cache_ttl}s)")
            self.logger.info(f"Entity caching enabled (size=50, ttl={entity_cache_ttl}s)")
            self.logger.info(f"Health/stats caching enabled (health={health_cache_ttl}s, stats={stats_cache_ttl}s)")
        else:
            self._search_cache = None
            self._similar_cache = None
            self._semantic_cache = None
            self._hybrid_cache = None
            self._faceted_cache = None
            self._embedding_cache = None
            self._entity_cache = None
            self._health_cache = None
            self._stats_cache = None

        # Initialize query preprocessing
        self.use_preprocessing = NLTK_SUPPORT and os.getenv('USE_QUERY_PREPROCESSING', '1') == '1'
        if self.use_preprocessing:
            self.stemmer = PorterStemmer()
            self.stop_words = set(stopwords.words('english'))
            self.logger.info("Query preprocessing enabled (stemming + stopwords)")
        else:
            self.stemmer = None
            self.stop_words = set()
            if NLTK_SUPPORT:
                self.logger.info("Query preprocessing disabled via USE_QUERY_PREPROCESSING=0")

        # Security: Allowed directories for document ingestion (optional)
        # Set via ALLOWED_DOCS_DIRS environment variable (comma-separated paths)
        # Always include: scraped_docs directory and current working directory
        allowed_dirs_env = os.getenv('ALLOWED_DOCS_DIRS', '')

        # Start with scraped_docs, downloads, temp, and current working directory
        default_allowed_dirs = [
            self.data_dir / "scraped_docs",  # Always allow scraped documents
            self.data_dir / "downloads",     # Always allow downloaded files (Archive Search)
            self.data_dir / "temp",          # Always allow temp files (Quick Add)
            Path.cwd()  # Always allow current working directory
        ]

        if allowed_dirs_env:
            # Add user-specified directories
            user_dirs = [Path(d.strip()).resolve() for d in allowed_dirs_env.split(',') if d.strip()]
            # Combine and remove duplicates while preserving order
            all_dirs = default_allowed_dirs + user_dirs
            seen = set()
            self.allowed_dirs = []
            for d in all_dirs:
                if d not in seen:
                    seen.add(d)
                    self.allowed_dirs.append(d)
        else:
            self.allowed_dirs = default_allowed_dirs

        self.logger.info(f"Path traversal protection enabled for: {self.allowed_dirs}")

        # Semantic search initialization (lazy loading for faster startup)
        self.use_semantic = SEMANTIC_SUPPORT and os.getenv('USE_SEMANTIC_SEARCH', '0') == '1'
        self.embeddings_model = None
        self.embeddings_index = None
        self.embeddings_doc_map = []  # Maps FAISS index positions to (doc_id, chunk_id)
        self._embeddings_loaded = False  # Track if model has been loaded

        if self.use_semantic:
            # Don't load model yet - will load on first use for faster startup
            self.embeddings_file = self.data_dir / "embeddings.faiss"
            self.embeddings_map_file = self.data_dir / "embeddings_map.json"
            self.logger.info("Semantic search enabled (lazy loading - model will load on first use)")
        else:
            if SEMANTIC_SUPPORT:
                self.logger.info("Semantic search disabled via USE_SEMANTIC_SEARCH=0")

        # Fuzzy search initialization
        self.use_fuzzy = FUZZY_SUPPORT and os.getenv('USE_FUZZY_SEARCH', '1') == '1'
        self.fuzzy_threshold = int(os.getenv('FUZZY_THRESHOLD', '80'))  # 0-100, default 80%

        if self.use_fuzzy:
            self.logger.info(f"Fuzzy search enabled (threshold={self.fuzzy_threshold}%)")
        else:
            if FUZZY_SUPPORT:
                self.logger.info("Fuzzy search disabled via USE_FUZZY_SEARCH=0")

        # OCR initialization
        self.use_ocr = OCR_SUPPORT and os.getenv('USE_OCR', '1') == '1'
        self.poppler_path = os.getenv('POPPLER_PATH', None)  # Optional Poppler path for pdf2image
        self.poppler_available = False  # Track if poppler is actually available

        if self.use_ocr:
            # Check if Tesseract is installed
            try:
                pytesseract.get_tesseract_version()
                self.logger.info("OCR enabled (Tesseract found)")

                # Check if poppler is available
                self.poppler_available = self._check_poppler_available()

                if self.poppler_available:
                    if self.poppler_path:
                        self.logger.info(f"Poppler found at: {self.poppler_path}")
                    else:
                        self.logger.info("Poppler found in system PATH")
                else:
                    self.logger.warning("[WARNING] Poppler not found! OCR will not work for scanned PDFs.")
                    self.logger.warning("Install poppler-utils:")
                    self.logger.warning("  Windows: Download from https://github.com/oschwartz10612/poppler-windows/releases/")
                    self.logger.warning("  Set POPPLER_PATH environment variable to the bin directory")
                    self.logger.warning("  Example: POPPLER_PATH=C:\\path\\to\\poppler-24.08.0\\Library\\bin")

            except Exception as e:
                self.logger.warning(f"OCR libraries installed but Tesseract not found: {e}")
                self.logger.warning("Install Tesseract from: https://github.com/UB-Mannheim/tesseract/wiki")
                self.use_ocr = False
        else:
            if OCR_SUPPORT:
                self.logger.info("OCR disabled via USE_OCR=0")

        # Load documents (with automatic migration if needed)
        self._load_documents()
        self.logger.info(f"Loaded {len(self.documents)} documents")

        # Initialize background entity extraction queue
        self._extraction_queue = queue.Queue()
        self._extraction_shutdown = threading.Event()
        self._extraction_worker = threading.Thread(
            target=self._extraction_worker_loop,
            daemon=True,
            name="EntityExtractionWorker"
        )
        self._extraction_worker.start()
        self.logger.info("Background entity extraction worker started")

        # Initialize anomaly detection
        self.anomaly_detector = None
        if ANOMALY_SUPPORT and os.getenv('USE_ANOMALY_DETECTION', '1') == '1':
            try:
                self.anomaly_detector = AnomalyDetector(self)
                self.logger.info("Anomaly detection enabled")
            except Exception as e:
                self.logger.warning(f"Anomaly detection initialization failed: {e}")
                self.anomaly_detector = None
        else:
            if ANOMALY_SUPPORT:
                self.logger.info("Anomaly detection disabled via USE_ANOMALY_DETECTION=0")

    def _init_database(self):
        """Initialize SQLite database and create schema if needed."""
        db_exists = self.db_file.exists()

        # Connect to database with enable foreign keys
        self.db_conn = sqlite3.connect(str(self.db_file), check_same_thread=False)
        self.db_conn.execute("PRAGMA foreign_keys = ON")

        if not db_exists:
            self.logger.info("Creating new database schema")
            cursor = self.db_conn.cursor()

            # Create documents table
            cursor.execute("""
                CREATE TABLE documents (
                    doc_id TEXT PRIMARY KEY,
                    filename TEXT NOT NULL,
                    title TEXT NOT NULL,
                    filepath TEXT NOT NULL UNIQUE,
                    file_type TEXT NOT NULL,
                    total_pages INTEGER,
                    total_chunks INTEGER NOT NULL,
                    indexed_at TEXT NOT NULL,
                    tags TEXT NOT NULL,
                    author TEXT,
                    subject TEXT,
                    creator TEXT,
                    creation_date TEXT,
                    file_mtime REAL,
                    file_hash TEXT,
                    source_url TEXT,
                    scrape_date TEXT,
                    scrape_config TEXT,
                    scrape_status TEXT,
                    scrape_error TEXT,
                    url_last_checked TEXT,
                    url_content_hash TEXT
                )
            """)

            # Create indexes on documents table
            cursor.execute("CREATE INDEX idx_documents_filepath ON documents(filepath)")
            cursor.execute("CREATE INDEX idx_documents_file_type ON documents(file_type)")
            cursor.execute("CREATE INDEX idx_documents_source_url ON documents(source_url)")
            cursor.execute("CREATE INDEX idx_documents_scrape_status ON documents(scrape_status)")

            # Create chunks table
            cursor.execute("""
                CREATE TABLE chunks (
                    doc_id TEXT NOT NULL,
                    chunk_id INTEGER NOT NULL,
                    page INTEGER,
                    content TEXT NOT NULL,
                    word_count INTEGER NOT NULL,
                    PRIMARY KEY (doc_id, chunk_id),
                    FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                )
            """)

            # Create index on chunks table
            cursor.execute("CREATE INDEX idx_chunks_doc_id ON chunks(doc_id)")

            # Create FTS5 virtual table for full-text search
            cursor.execute("""
                CREATE VIRTUAL TABLE IF NOT EXISTS chunks_fts5 USING fts5(
                    doc_id UNINDEXED,
                    chunk_id UNINDEXED,
                    content,
                    tokenize='porter unicode61'
                )
            """)

            # Trigger: Keep FTS5 in sync on INSERT
            cursor.execute("""
                CREATE TRIGGER IF NOT EXISTS chunks_fts5_insert AFTER INSERT ON chunks BEGIN
                    INSERT INTO chunks_fts5(rowid, doc_id, chunk_id, content)
                    VALUES (new.rowid, new.doc_id, new.chunk_id, new.content);
                END
            """)

            # Trigger: Keep FTS5 in sync on DELETE
            cursor.execute("""
                CREATE TRIGGER IF NOT EXISTS chunks_fts5_delete AFTER DELETE ON chunks BEGIN
                    DELETE FROM chunks_fts5 WHERE rowid = old.rowid;
                END
            """)

            # Trigger: Keep FTS5 in sync on UPDATE
            cursor.execute("""
                CREATE TRIGGER IF NOT EXISTS chunks_fts5_update AFTER UPDATE ON chunks BEGIN
                    DELETE FROM chunks_fts5 WHERE rowid = old.rowid;
                    INSERT INTO chunks_fts5(rowid, doc_id, chunk_id, content)
                    VALUES (new.rowid, new.doc_id, new.chunk_id, new.content);
                END
            """)

            # Create document_tables table
            cursor.execute("""
                CREATE TABLE document_tables (
                    doc_id TEXT NOT NULL,
                    table_id INTEGER NOT NULL,
                    page INTEGER,
                    markdown TEXT NOT NULL,
                    searchable_text TEXT NOT NULL,
                    row_count INTEGER,
                    col_count INTEGER,
                    PRIMARY KEY (doc_id, table_id),
                    FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                )
            """)

            # Create FTS5 index for table search
            cursor.execute("""
                CREATE VIRTUAL TABLE tables_fts USING fts5(
                    doc_id UNINDEXED,
                    table_id UNINDEXED,
                    searchable_text,
                    tokenize='porter unicode61'
                )
            """)

            # Triggers for tables_fts
            cursor.execute("""
                CREATE TRIGGER tables_fts_insert AFTER INSERT ON document_tables BEGIN
                    INSERT INTO tables_fts(rowid, doc_id, table_id, searchable_text)
                    VALUES ((SELECT COALESCE(MAX(rowid), 0) + 1 FROM tables_fts),
                            new.doc_id, new.table_id, new.searchable_text);
                END
            """)

            cursor.execute("""
                CREATE TRIGGER tables_fts_delete AFTER DELETE ON document_tables BEGIN
                    DELETE FROM tables_fts WHERE doc_id = old.doc_id AND table_id = old.table_id;
                END
            """)

            cursor.execute("""
                CREATE TRIGGER tables_fts_update AFTER UPDATE ON document_tables BEGIN
                    DELETE FROM tables_fts WHERE doc_id = old.doc_id AND table_id = old.table_id;
                    INSERT INTO tables_fts(rowid, doc_id, table_id, searchable_text)
                    VALUES ((SELECT COALESCE(MAX(rowid), 0) + 1 FROM tables_fts),
                            new.doc_id, new.table_id, new.searchable_text);
                END
            """)

            # Create document_code_blocks table
            cursor.execute("""
                CREATE TABLE document_code_blocks (
                    doc_id TEXT NOT NULL,
                    block_id INTEGER NOT NULL,
                    page INTEGER,
                    block_type TEXT NOT NULL,
                    code TEXT NOT NULL,
                    searchable_text TEXT NOT NULL,
                    line_count INTEGER,
                    PRIMARY KEY (doc_id, block_id),
                    FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                )
            """)

            # Create FTS5 index for code search
            cursor.execute("""
                CREATE VIRTUAL TABLE code_fts USING fts5(
                    doc_id UNINDEXED,
                    block_id UNINDEXED,
                    block_type UNINDEXED,
                    searchable_text,
                    tokenize='porter unicode61'
                )
            """)

            # Triggers for code_fts
            cursor.execute("""
                CREATE TRIGGER code_fts_insert AFTER INSERT ON document_code_blocks BEGIN
                    INSERT INTO code_fts(rowid, doc_id, block_id, block_type, searchable_text)
                    VALUES ((SELECT COALESCE(MAX(rowid), 0) + 1 FROM code_fts),
                            new.doc_id, new.block_id, new.block_type, new.searchable_text);
                END
            """)

            cursor.execute("""
                CREATE TRIGGER code_fts_delete AFTER DELETE ON document_code_blocks BEGIN
                    DELETE FROM code_fts WHERE doc_id = old.doc_id AND block_id = old.block_id;
                END
            """)

            cursor.execute("""
                CREATE TRIGGER code_fts_update AFTER UPDATE ON document_code_blocks BEGIN
                    DELETE FROM code_fts WHERE doc_id = old.doc_id AND block_id = old.block_id;
                    INSERT INTO code_fts(rowid, doc_id, block_id, block_type, searchable_text)
                    VALUES ((SELECT COALESCE(MAX(rowid), 0) + 1 FROM code_fts),
                            new.doc_id, new.block_id, new.block_type, new.searchable_text);
                END
            """)

            # Create document_facets table for faceted search
            cursor.execute("""
                CREATE TABLE document_facets (
                    doc_id TEXT NOT NULL,
                    facet_type TEXT NOT NULL,
                    facet_value TEXT NOT NULL,
                    PRIMARY KEY (doc_id, facet_type, facet_value),
                    FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                )
            """)

            # Create indexes for faceted search
            cursor.execute("CREATE INDEX idx_facets_type_value ON document_facets(facet_type, facet_value)")
            cursor.execute("CREATE INDEX idx_facets_doc_id ON document_facets(doc_id)")

            # Create search_log table for analytics
            cursor.execute("""
                CREATE TABLE search_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT DEFAULT CURRENT_TIMESTAMP,
                    query TEXT NOT NULL,
                    search_mode TEXT NOT NULL,
                    results_count INTEGER NOT NULL,
                    clicked_doc_id TEXT,
                    execution_time_ms REAL,
                    tags TEXT
                )
            """)

            # Create indexes for search analytics
            cursor.execute("CREATE INDEX idx_search_log_query ON search_log(query)")
            cursor.execute("CREATE INDEX idx_search_log_timestamp ON search_log(timestamp)")
            cursor.execute("CREATE INDEX idx_search_log_mode ON search_log(search_mode)")

            # Create cross_references table for content linking
            cursor.execute("""
                CREATE TABLE cross_references (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    doc_id TEXT NOT NULL,
                    chunk_id INTEGER NOT NULL,
                    ref_type TEXT NOT NULL,
                    ref_value TEXT NOT NULL,
                    context TEXT,
                    FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                )
            """)

            # Create indexes for cross-reference lookup
            cursor.execute("CREATE INDEX idx_xref_type_value ON cross_references(ref_type, ref_value)")
            cursor.execute("CREATE INDEX idx_xref_doc_id ON cross_references(doc_id)")

            # Create query_suggestions table for autocomplete
            cursor.execute("""
                CREATE VIRTUAL TABLE query_suggestions USING fts5(
                    term,
                    frequency UNINDEXED,
                    category UNINDEXED
                )
            """)

            # Create document_summaries table for AI-generated summaries
            cursor.execute("""
                CREATE TABLE document_summaries (
                    doc_id TEXT NOT NULL,
                    summary_type TEXT NOT NULL,
                    summary_text TEXT NOT NULL,
                    generated_at TEXT NOT NULL,
                    model TEXT,
                    token_count INTEGER,
                    PRIMARY KEY (doc_id, summary_type),
                    FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                )
            """)

            # Create indexes for summary queries
            cursor.execute("CREATE INDEX idx_summaries_doc_id ON document_summaries(doc_id)")
            cursor.execute("CREATE INDEX idx_summaries_type ON document_summaries(summary_type)")

            # Create document_entities table for named entity extraction
            cursor.execute("""
                CREATE TABLE document_entities (
                    doc_id TEXT NOT NULL,
                    entity_id INTEGER NOT NULL,
                    entity_text TEXT NOT NULL,
                    entity_type TEXT NOT NULL,
                    confidence REAL NOT NULL,
                    context TEXT,
                    first_chunk_id INTEGER,
                    occurrence_count INTEGER DEFAULT 1,
                    generated_at TEXT NOT NULL,
                    model TEXT,
                    PRIMARY KEY (doc_id, entity_id),
                    FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                )
            """)

            # Create FTS5 index for entity search
            cursor.execute("""
                CREATE VIRTUAL TABLE entities_fts USING fts5(
                    doc_id UNINDEXED,
                    entity_id UNINDEXED,
                    entity_text,
                    entity_type UNINDEXED,
                    context,
                    tokenize='porter unicode61'
                )
            """)

            # Triggers to keep entities_fts in sync with document_entities
            cursor.execute("""
                CREATE TRIGGER entities_fts_insert AFTER INSERT ON document_entities BEGIN
                    INSERT INTO entities_fts(rowid, doc_id, entity_id, entity_text, entity_type, context)
                    VALUES (new.rowid, new.doc_id, new.entity_id, new.entity_text, new.entity_type, new.context);
                END
            """)

            cursor.execute("""
                CREATE TRIGGER entities_fts_delete AFTER DELETE ON document_entities BEGIN
                    DELETE FROM entities_fts WHERE rowid = old.rowid;
                END
            """)

            cursor.execute("""
                CREATE TRIGGER entities_fts_update AFTER UPDATE ON document_entities BEGIN
                    DELETE FROM entities_fts WHERE rowid = old.rowid;
                    INSERT INTO entities_fts(rowid, doc_id, entity_id, entity_text, entity_type, context)
                    VALUES (new.rowid, new.doc_id, new.entity_id, new.entity_text, new.entity_type, new.context);
                END
            """)

            # Create indexes for entity queries
            cursor.execute("CREATE INDEX idx_entities_doc_id ON document_entities(doc_id)")
            cursor.execute("CREATE INDEX idx_entities_type ON document_entities(entity_type)")
            cursor.execute("CREATE INDEX idx_entities_text ON document_entities(entity_text)")

            # Create entity relationships table
            cursor.execute("""
                CREATE TABLE entity_relationships (
                    entity1_text TEXT NOT NULL,
                    entity1_type TEXT NOT NULL,
                    entity2_text TEXT NOT NULL,
                    entity2_type TEXT NOT NULL,
                    relationship_type TEXT NOT NULL,
                    strength REAL NOT NULL,
                    doc_count INTEGER NOT NULL DEFAULT 1,
                    first_seen_doc TEXT,
                    context_sample TEXT,
                    last_updated TEXT NOT NULL,
                    PRIMARY KEY (entity1_text, entity2_text, relationship_type)
                )
            """)

            # Create indexes for relationship queries
            cursor.execute("CREATE INDEX idx_relationships_entity1 ON entity_relationships(entity1_text)")
            cursor.execute("CREATE INDEX idx_relationships_entity2 ON entity_relationships(entity2_text)")
            cursor.execute("CREATE INDEX idx_relationships_type ON entity_relationships(relationship_type)")
            cursor.execute("CREATE INDEX idx_relationships_strength ON entity_relationships(strength)")

            # Create extraction_jobs table for background entity extraction
            cursor.execute("""
                CREATE TABLE extraction_jobs (
                    job_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    doc_id TEXT NOT NULL,
                    status TEXT NOT NULL,
                    confidence_threshold REAL NOT NULL,
                    queued_at TEXT NOT NULL,
                    started_at TEXT,
                    completed_at TEXT,
                    error_message TEXT,
                    entities_extracted INTEGER,
                    FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                )
            """)

            # Create indexes for extraction jobs queries
            cursor.execute("CREATE INDEX idx_extraction_jobs_doc_id ON extraction_jobs(doc_id)")
            cursor.execute("CREATE INDEX idx_extraction_jobs_status ON extraction_jobs(status)")
            cursor.execute("CREATE INDEX idx_extraction_jobs_queued_at ON extraction_jobs(queued_at)")

            # Create graph_cache table for knowledge graph caching (v2.24.0)
            cursor.execute("""
                CREATE TABLE graph_cache (
                    cache_id TEXT PRIMARY KEY,
                    graph_version INTEGER NOT NULL,
                    graph_data BLOB NOT NULL,
                    node_count INTEGER NOT NULL,
                    edge_count INTEGER NOT NULL,
                    created_date TEXT NOT NULL,
                    last_accessed TEXT
                )
            """)

            cursor.execute("CREATE INDEX idx_graph_cache_created ON graph_cache(created_date)")
            cursor.execute("CREATE INDEX idx_graph_cache_accessed ON graph_cache(last_accessed)")

            # Create graph_metrics table for graph analysis metrics (v2.24.0)
            cursor.execute("""
                CREATE TABLE graph_metrics (
                    metric_id TEXT PRIMARY KEY,
                    entity_text TEXT NOT NULL,
                    entity_type TEXT NOT NULL,
                    pagerank REAL,
                    betweenness_centrality REAL,
                    closeness_centrality REAL,
                    degree_centrality REAL,
                    community_id INTEGER,
                    computed_date TEXT NOT NULL
                )
            """)

            cursor.execute("CREATE INDEX idx_graph_metrics_entity ON graph_metrics(entity_text)")
            cursor.execute("CREATE INDEX idx_graph_metrics_pagerank ON graph_metrics(pagerank DESC)")
            cursor.execute("CREATE INDEX idx_graph_metrics_community ON graph_metrics(community_id)")
            cursor.execute("CREATE INDEX idx_graph_metrics_type ON graph_metrics(entity_type)")

            # Create graph_paths table for path finding cache (v2.24.0)
            cursor.execute("""
                CREATE TABLE graph_paths (
                    path_id TEXT PRIMARY KEY,
                    entity1 TEXT NOT NULL,
                    entity2 TEXT NOT NULL,
                    path_length INTEGER NOT NULL,
                    path_nodes TEXT NOT NULL,
                    path_weight REAL,
                    computed_date TEXT NOT NULL
                )
            """)

            cursor.execute("CREATE INDEX idx_graph_paths_entity1 ON graph_paths(entity1)")
            cursor.execute("CREATE INDEX idx_graph_paths_entity2 ON graph_paths(entity2)")
            cursor.execute("CREATE INDEX idx_graph_paths_entities ON graph_paths(entity1, entity2)")

            self.db_conn.commit()
            self.logger.info("Database schema created successfully (with FTS5, tables, code blocks, facets, analytics, suggestions, summaries, entities, relationships, extraction jobs, and knowledge graph)")
        else:
            self.logger.info("Using existing database")

            # Migrate database schema: Add file_mtime and file_hash columns if missing
            cursor = self.db_conn.cursor()
            cursor.execute("PRAGMA table_info(documents)")
            columns = [row[1] for row in cursor.fetchall()]

            if 'file_mtime' not in columns:
                self.logger.info("Migrating database: adding file_mtime column")
                cursor.execute("ALTER TABLE documents ADD COLUMN file_mtime REAL")
                self.db_conn.commit()

            if 'file_hash' not in columns:
                self.logger.info("Migrating database: adding file_hash column")
                cursor.execute("ALTER TABLE documents ADD COLUMN file_hash TEXT")
                self.db_conn.commit()

            # Migrate: Add URL scraping columns if missing
            if 'source_url' not in columns:
                self.logger.info("Migrating database: adding URL scraping columns")
                cursor.execute("ALTER TABLE documents ADD COLUMN source_url TEXT")
                cursor.execute("ALTER TABLE documents ADD COLUMN scrape_date TEXT")
                cursor.execute("ALTER TABLE documents ADD COLUMN scrape_config TEXT")
                cursor.execute("ALTER TABLE documents ADD COLUMN scrape_status TEXT")
                cursor.execute("ALTER TABLE documents ADD COLUMN scrape_error TEXT")
                cursor.execute("ALTER TABLE documents ADD COLUMN url_last_checked TEXT")
                cursor.execute("ALTER TABLE documents ADD COLUMN url_content_hash TEXT")

                # Create indexes for URL columns
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_documents_source_url ON documents(source_url)")
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_documents_scrape_status ON documents(scrape_status)")

                self.db_conn.commit()
                self.logger.info("URL scraping columns and indexes added successfully")

            # Check if FTS5 table exists and populate if needed
            try:
                cursor = self.db_conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM chunks_fts5")
                fts5_count = cursor.fetchone()[0]

                # If FTS5 table is empty but chunks exist, populate it
                if fts5_count == 0:
                    cursor.execute("SELECT COUNT(*) FROM chunks")
                    chunks_count = cursor.fetchone()[0]

                    if chunks_count > 0:
                        self.logger.info(f"Populating FTS5 index with {chunks_count} existing chunks")
                        cursor.execute("""
                            INSERT INTO chunks_fts5(rowid, doc_id, chunk_id, content)
                            SELECT rowid, doc_id, chunk_id, content FROM chunks
                        """)
                        self.db_conn.commit()
                        self.logger.info("FTS5 index populated successfully")
            except Exception as e:
                # FTS5 table doesn't exist, create it
                self.logger.info(f"Creating FTS5 table for existing database: {e}")
                cursor = self.db_conn.cursor()

                cursor.execute("""
                    CREATE VIRTUAL TABLE IF NOT EXISTS chunks_fts5 USING fts5(
                        doc_id UNINDEXED,
                        chunk_id UNINDEXED,
                        content,
                        tokenize='porter unicode61'
                    )
                """)

                cursor.execute("""
                    CREATE TRIGGER IF NOT EXISTS chunks_fts5_insert AFTER INSERT ON chunks BEGIN
                        INSERT INTO chunks_fts5(rowid, doc_id, chunk_id, content)
                        VALUES (new.rowid, new.doc_id, new.chunk_id, new.content);
                    END
                """)

                cursor.execute("""
                    CREATE TRIGGER IF NOT EXISTS chunks_fts5_delete AFTER DELETE ON chunks BEGIN
                        DELETE FROM chunks_fts5 WHERE rowid = old.rowid;
                    END
                """)

                cursor.execute("""
                    CREATE TRIGGER IF NOT EXISTS chunks_fts5_update AFTER UPDATE ON chunks BEGIN
                        DELETE FROM chunks_fts5 WHERE rowid = old.rowid;
                        INSERT INTO chunks_fts5(rowid, doc_id, chunk_id, content)
                        VALUES (new.rowid, new.doc_id, new.chunk_id, new.content);
                    END
                """)

                # Populate FTS5 from existing chunks
                cursor.execute("""
                    INSERT INTO chunks_fts5(rowid, doc_id, chunk_id, content)
                    SELECT rowid, doc_id, chunk_id, content FROM chunks
                """)

                self.db_conn.commit()
                self.logger.info("FTS5 table created and populated for existing database")

            # Migrate: Add document_tables table if not exists
            cursor = self.db_conn.cursor()
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='document_tables'
            """)
            if not cursor.fetchone():
                self.logger.info("Creating document_tables table")
                cursor.execute("""
                    CREATE TABLE document_tables (
                        doc_id TEXT NOT NULL,
                        table_id INTEGER NOT NULL,
                        page INTEGER,
                        markdown TEXT NOT NULL,
                        searchable_text TEXT NOT NULL,
                        row_count INTEGER,
                        col_count INTEGER,
                        PRIMARY KEY (doc_id, table_id),
                        FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                    )
                """)

                # Create FTS5 index for table search
                cursor.execute("""
                    CREATE VIRTUAL TABLE IF NOT EXISTS tables_fts USING fts5(
                        doc_id UNINDEXED,
                        table_id UNINDEXED,
                        searchable_text,
                        tokenize='porter unicode61'
                    )
                """)

                # Triggers for tables_fts
                cursor.execute("""
                    CREATE TRIGGER IF NOT EXISTS tables_fts_insert AFTER INSERT ON document_tables BEGIN
                        INSERT INTO tables_fts(rowid, doc_id, table_id, searchable_text)
                        VALUES (new.rowid, new.doc_id, new.table_id, new.searchable_text);
                    END
                """)

                cursor.execute("""
                    CREATE TRIGGER IF NOT EXISTS tables_fts_delete AFTER DELETE ON document_tables BEGIN
                        DELETE FROM tables_fts WHERE rowid = old.rowid;
                    END
                """)

                cursor.execute("""
                    CREATE TRIGGER IF NOT EXISTS tables_fts_update AFTER UPDATE ON document_tables BEGIN
                        DELETE FROM tables_fts WHERE rowid = old.rowid;
                        INSERT INTO tables_fts(rowid, doc_id, table_id, searchable_text)
                        VALUES (new.rowid, new.doc_id, new.table_id, new.searchable_text);
                    END
                """)

                self.db_conn.commit()
                self.logger.info("document_tables and tables_fts created")

            # Migrate: Add document_code_blocks table if not exists
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='document_code_blocks'
            """)
            if not cursor.fetchone():
                self.logger.info("Creating document_code_blocks table")
                cursor.execute("""
                    CREATE TABLE document_code_blocks (
                        doc_id TEXT NOT NULL,
                        block_id INTEGER NOT NULL,
                        page INTEGER,
                        block_type TEXT NOT NULL,
                        code TEXT NOT NULL,
                        searchable_text TEXT NOT NULL,
                        line_count INTEGER,
                        PRIMARY KEY (doc_id, block_id),
                        FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                    )
                """)

                # Create FTS5 index for code search
                cursor.execute("""
                    CREATE VIRTUAL TABLE IF NOT EXISTS code_fts USING fts5(
                        doc_id UNINDEXED,
                        block_id UNINDEXED,
                        block_type UNINDEXED,
                        searchable_text,
                        tokenize='porter unicode61'
                    )
                """)

                # Triggers for code_fts
                cursor.execute("""
                    CREATE TRIGGER IF NOT EXISTS code_fts_insert AFTER INSERT ON document_code_blocks BEGIN
                        INSERT INTO code_fts(rowid, doc_id, block_id, block_type, searchable_text)
                        VALUES (new.rowid, new.doc_id, new.block_id, new.block_type, new.searchable_text);
                    END
                """)

                cursor.execute("""
                    CREATE TRIGGER IF NOT EXISTS code_fts_delete AFTER DELETE ON document_code_blocks BEGIN
                        DELETE FROM code_fts WHERE rowid = old.rowid;
                    END
                """)

                cursor.execute("""
                    CREATE TRIGGER IF NOT EXISTS code_fts_update AFTER UPDATE ON document_code_blocks BEGIN
                        DELETE FROM code_fts WHERE rowid = old.rowid;
                        INSERT INTO code_fts(rowid, doc_id, block_id, block_type, searchable_text)
                        VALUES (new.rowid, new.doc_id, new.block_id, new.block_type, new.searchable_text);
                    END
                """)

                self.db_conn.commit()
                self.logger.info("document_code_blocks and code_fts created")

            # Migrate: Add document_facets table if not exists
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='document_facets'
            """)
            if not cursor.fetchone():
                self.logger.info("Creating document_facets table for faceted search")
                cursor.execute("""
                    CREATE TABLE document_facets (
                        doc_id TEXT NOT NULL,
                        facet_type TEXT NOT NULL,
                        facet_value TEXT NOT NULL,
                        PRIMARY KEY (doc_id, facet_type, facet_value),
                        FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                    )
                """)

                cursor.execute("CREATE INDEX idx_facets_type_value ON document_facets(facet_type, facet_value)")
                cursor.execute("CREATE INDEX idx_facets_doc_id ON document_facets(doc_id)")

                self.db_conn.commit()
                self.logger.info("document_facets table created")

            # Migrate: Add search_log table if not exists
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='search_log'
            """)
            if not cursor.fetchone():
                self.logger.info("Creating search_log table for analytics")
                cursor.execute("""
                    CREATE TABLE search_log (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        timestamp TEXT DEFAULT CURRENT_TIMESTAMP,
                        query TEXT NOT NULL,
                        search_mode TEXT NOT NULL,
                        results_count INTEGER NOT NULL,
                        clicked_doc_id TEXT,
                        execution_time_ms REAL,
                        tags TEXT
                    )
                """)

                cursor.execute("CREATE INDEX idx_search_log_query ON search_log(query)")
                cursor.execute("CREATE INDEX idx_search_log_timestamp ON search_log(timestamp)")
                cursor.execute("CREATE INDEX idx_search_log_mode ON search_log(search_mode)")

                self.db_conn.commit()
                self.logger.info("search_log table created")

            # Migrate: Add cross_references table if not exists
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='cross_references'
            """)
            if not cursor.fetchone():
                self.logger.info("Creating cross_references table for content linking")
                cursor.execute("""
                    CREATE TABLE cross_references (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        doc_id TEXT NOT NULL,
                        chunk_id INTEGER NOT NULL,
                        ref_type TEXT NOT NULL,
                        ref_value TEXT NOT NULL,
                        context TEXT,
                        FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                    )
                """)

                cursor.execute("CREATE INDEX idx_xref_type_value ON cross_references(ref_type, ref_value)")
                cursor.execute("CREATE INDEX idx_xref_doc_id ON cross_references(doc_id)")

                self.db_conn.commit()
                self.logger.info("cross_references table created")

            # Migrate: Add query_suggestions table if not exists
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='query_suggestions'
            """)
            if not cursor.fetchone():
                self.logger.info("Creating query_suggestions table for autocomplete")
                cursor.execute("""
                    CREATE VIRTUAL TABLE query_suggestions USING fts5(
                        term,
                        frequency UNINDEXED,
                        category UNINDEXED
                    )
                """)

                self.db_conn.commit()
                self.logger.info("query_suggestions table created")

            # Migrate: Add document_summaries table if not exists
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='document_summaries'
            """)
            if not cursor.fetchone():
                self.logger.info("Creating document_summaries table for AI-generated summaries")
                cursor.execute("""
                    CREATE TABLE document_summaries (
                        doc_id TEXT NOT NULL,
                        summary_type TEXT NOT NULL,
                        summary_text TEXT NOT NULL,
                        generated_at TEXT NOT NULL,
                        model TEXT,
                        token_count INTEGER,
                        PRIMARY KEY (doc_id, summary_type),
                        FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                    )
                """)

                # Create indexes for summary queries
                cursor.execute("CREATE INDEX idx_summaries_doc_id ON document_summaries(doc_id)")
                cursor.execute("CREATE INDEX idx_summaries_type ON document_summaries(summary_type)")

                self.db_conn.commit()
                self.logger.info("document_summaries table created")

            # Migrate: Add document_entities table if not exists
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='document_entities'
            """)
            if not cursor.fetchone():
                self.logger.info("Creating document_entities table for named entity extraction")
                cursor.execute("""
                    CREATE TABLE document_entities (
                        doc_id TEXT NOT NULL,
                        entity_id INTEGER NOT NULL,
                        entity_text TEXT NOT NULL,
                        entity_type TEXT NOT NULL,
                        confidence REAL NOT NULL,
                        context TEXT,
                        first_chunk_id INTEGER,
                        occurrence_count INTEGER DEFAULT 1,
                        generated_at TEXT NOT NULL,
                        model TEXT,
                        PRIMARY KEY (doc_id, entity_id),
                        FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                    )
                """)

                # Create FTS5 index for entity search
                cursor.execute("""
                    CREATE VIRTUAL TABLE entities_fts USING fts5(
                        doc_id UNINDEXED,
                        entity_id UNINDEXED,
                        entity_text,
                        entity_type UNINDEXED,
                        context,
                        tokenize='porter unicode61'
                    )
                """)

                # Triggers to keep entities_fts in sync
                cursor.execute("""
                    CREATE TRIGGER entities_fts_insert AFTER INSERT ON document_entities BEGIN
                        INSERT INTO entities_fts(rowid, doc_id, entity_id, entity_text, entity_type, context)
                        VALUES (new.rowid, new.doc_id, new.entity_id, new.entity_text, new.entity_type, new.context);
                    END
                """)

                cursor.execute("""
                    CREATE TRIGGER entities_fts_delete AFTER DELETE ON document_entities BEGIN
                        DELETE FROM entities_fts WHERE rowid = old.rowid;
                    END
                """)

                cursor.execute("""
                    CREATE TRIGGER entities_fts_update AFTER UPDATE ON document_entities BEGIN
                        DELETE FROM entities_fts WHERE rowid = old.rowid;
                        INSERT INTO entities_fts(rowid, doc_id, entity_id, entity_text, entity_type, context)
                        VALUES (new.rowid, new.doc_id, new.entity_id, new.entity_text, new.entity_type, new.context);
                    END
                """)

                # Create indexes for entity queries
                cursor.execute("CREATE INDEX idx_entities_doc_id ON document_entities(doc_id)")
                cursor.execute("CREATE INDEX idx_entities_type ON document_entities(entity_type)")
                cursor.execute("CREATE INDEX idx_entities_text ON document_entities(entity_text)")

                self.db_conn.commit()
                self.logger.info("document_entities table created")

            # Migrate: Add entity_relationships table if not exists
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='entity_relationships'
            """)
            if not cursor.fetchone():
                self.logger.info("Creating entity_relationships table for entity co-occurrence tracking")
                cursor.execute("""
                    CREATE TABLE entity_relationships (
                        entity1_text TEXT NOT NULL,
                        entity1_type TEXT NOT NULL,
                        entity2_text TEXT NOT NULL,
                        entity2_type TEXT NOT NULL,
                        relationship_type TEXT NOT NULL,
                        strength REAL NOT NULL,
                        doc_count INTEGER NOT NULL DEFAULT 1,
                        first_seen_doc TEXT,
                        context_sample TEXT,
                        last_updated TEXT NOT NULL,
                        PRIMARY KEY (entity1_text, entity2_text, relationship_type)
                    )
                """)

                # Create indexes for relationship queries
                cursor.execute("CREATE INDEX idx_relationships_entity1 ON entity_relationships(entity1_text)")
                cursor.execute("CREATE INDEX idx_relationships_entity2 ON entity_relationships(entity2_text)")
                cursor.execute("CREATE INDEX idx_relationships_type ON entity_relationships(relationship_type)")
                cursor.execute("CREATE INDEX idx_relationships_strength ON entity_relationships(strength)")

                self.db_conn.commit()
                self.logger.info("entity_relationships table created")

            # Migrate: Add extraction_jobs table if not exists
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='extraction_jobs'
            """)
            if not cursor.fetchone():
                self.logger.info("Creating extraction_jobs table for background entity extraction")
                cursor.execute("""
                    CREATE TABLE extraction_jobs (
                        job_id INTEGER PRIMARY KEY AUTOINCREMENT,
                        doc_id TEXT NOT NULL,
                        status TEXT NOT NULL,
                        confidence_threshold REAL NOT NULL,
                        queued_at TEXT NOT NULL,
                        started_at TEXT,
                        completed_at TEXT,
                        error_message TEXT,
                        entities_extracted INTEGER,
                        FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE
                    )
                """)

                # Create indexes for extraction jobs queries
                cursor.execute("CREATE INDEX idx_extraction_jobs_doc_id ON extraction_jobs(doc_id)")
                cursor.execute("CREATE INDEX idx_extraction_jobs_status ON extraction_jobs(status)")
                cursor.execute("CREATE INDEX idx_extraction_jobs_queued_at ON extraction_jobs(queued_at)")

                self.db_conn.commit()
                self.logger.info("extraction_jobs table created")

            # Check for graph_cache table (v2.24.0 - Knowledge Graph Analysis)
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='graph_cache'")
            if not cursor.fetchone():
                self.logger.info("Creating graph_cache table for knowledge graph caching")
                cursor.execute("""
                    CREATE TABLE graph_cache (
                        cache_id TEXT PRIMARY KEY,
                        graph_version INTEGER NOT NULL,
                        graph_data BLOB NOT NULL,
                        node_count INTEGER NOT NULL,
                        edge_count INTEGER NOT NULL,
                        created_date TEXT NOT NULL,
                        last_accessed TEXT
                    )
                """)

                # Create index for cache management
                cursor.execute("CREATE INDEX idx_graph_cache_created ON graph_cache(created_date)")
                cursor.execute("CREATE INDEX idx_graph_cache_accessed ON graph_cache(last_accessed)")

                self.db_conn.commit()
                self.logger.info("graph_cache table created")

            # Check for graph_metrics table (v2.24.0 - Knowledge Graph Analysis)
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='graph_metrics'")
            if not cursor.fetchone():
                self.logger.info("Creating graph_metrics table for graph analysis metrics")
                cursor.execute("""
                    CREATE TABLE graph_metrics (
                        metric_id TEXT PRIMARY KEY,
                        entity_text TEXT NOT NULL,
                        entity_type TEXT NOT NULL,
                        pagerank REAL,
                        betweenness_centrality REAL,
                        closeness_centrality REAL,
                        degree_centrality REAL,
                        community_id INTEGER,
                        computed_date TEXT NOT NULL
                    )
                """)

                # Create indexes for metric queries
                cursor.execute("CREATE INDEX idx_graph_metrics_entity ON graph_metrics(entity_text)")
                cursor.execute("CREATE INDEX idx_graph_metrics_pagerank ON graph_metrics(pagerank DESC)")
                cursor.execute("CREATE INDEX idx_graph_metrics_community ON graph_metrics(community_id)")
                cursor.execute("CREATE INDEX idx_graph_metrics_type ON graph_metrics(entity_type)")

                self.db_conn.commit()
                self.logger.info("graph_metrics table created")

            # Check for graph_paths table (v2.24.0 - Knowledge Graph Analysis)
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='graph_paths'")
            if not cursor.fetchone():
                self.logger.info("Creating graph_paths table for path finding cache")
                cursor.execute("""
                    CREATE TABLE graph_paths (
                        path_id TEXT PRIMARY KEY,
                        entity1 TEXT NOT NULL,
                        entity2 TEXT NOT NULL,
                        path_length INTEGER NOT NULL,
                        path_nodes TEXT NOT NULL,
                        path_weight REAL,
                        computed_date TEXT NOT NULL
                    )
                """)

                # Create indexes for path queries
                cursor.execute("CREATE INDEX idx_graph_paths_entity1 ON graph_paths(entity1)")
                cursor.execute("CREATE INDEX idx_graph_paths_entity2 ON graph_paths(entity2)")
                cursor.execute("CREATE INDEX idx_graph_paths_entities ON graph_paths(entity1, entity2)")

                self.db_conn.commit()
                self.logger.info("graph_paths table created")

            # ============================================================
            # Phase 2 Topic & Cluster Tables (v2.24.0 - Discovery)
            # ============================================================

            # Check for topics table (v2.24.0 - Topic Modeling)
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='topics'")
            if not cursor.fetchone():
                self.logger.info("Creating topics table for topic modeling")
                cursor.execute("""
                    CREATE TABLE topics (
                        topic_id TEXT PRIMARY KEY,
                        model_type TEXT NOT NULL,
                        topic_number INTEGER NOT NULL,
                        top_words TEXT NOT NULL,
                        word_weights TEXT NOT NULL,
                        num_documents INTEGER DEFAULT 0,
                        coherence_score REAL,
                        created_date TEXT NOT NULL,
                        UNIQUE(model_type, topic_number)
                    )
                """)

                # Create indexes for topic queries
                cursor.execute("CREATE INDEX idx_topics_model ON topics(model_type)")
                cursor.execute("CREATE INDEX idx_topics_number ON topics(model_type, topic_number)")

                self.db_conn.commit()
                self.logger.info("topics table created")

            # Check for document_topics table (v2.24.0 - Topic Modeling)
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='document_topics'")
            if not cursor.fetchone():
                self.logger.info("Creating document_topics table for topic assignments")
                cursor.execute("""
                    CREATE TABLE document_topics (
                        assignment_id TEXT PRIMARY KEY,
                        doc_id TEXT NOT NULL,
                        topic_id TEXT NOT NULL,
                        probability REAL NOT NULL,
                        model_type TEXT NOT NULL,
                        assigned_date TEXT NOT NULL,
                        FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE,
                        FOREIGN KEY (topic_id) REFERENCES topics(topic_id) ON DELETE CASCADE
                    )
                """)

                # Create indexes for topic assignment queries
                cursor.execute("CREATE INDEX idx_doc_topics_doc ON document_topics(doc_id)")
                cursor.execute("CREATE INDEX idx_doc_topics_topic ON document_topics(topic_id)")
                cursor.execute("CREATE INDEX idx_doc_topics_model ON document_topics(model_type)")
                cursor.execute("CREATE INDEX idx_doc_topics_probability ON document_topics(probability DESC)")

                self.db_conn.commit()
                self.logger.info("document_topics table created")

            # Check for clusters table (v2.24.0 - Document Clustering)
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='clusters'")
            if not cursor.fetchone():
                self.logger.info("Creating clusters table for clustering analysis")
                cursor.execute("""
                    CREATE TABLE clusters (
                        cluster_id TEXT PRIMARY KEY,
                        algorithm TEXT NOT NULL,
                        cluster_number INTEGER NOT NULL,
                        centroid_vector BLOB,
                        num_documents INTEGER DEFAULT 0,
                        representative_docs TEXT,
                        top_terms TEXT,
                        silhouette_score REAL,
                        created_date TEXT NOT NULL,
                        UNIQUE(algorithm, cluster_number)
                    )
                """)

                # Create indexes for cluster queries
                cursor.execute("CREATE INDEX idx_clusters_algorithm ON clusters(algorithm)")
                cursor.execute("CREATE INDEX idx_clusters_number ON clusters(algorithm, cluster_number)")

                self.db_conn.commit()
                self.logger.info("clusters table created")

            # Check for document_clusters table (v2.24.0 - Document Clustering)
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='document_clusters'")
            if not cursor.fetchone():
                self.logger.info("Creating document_clusters table for cluster assignments")
                cursor.execute("""
                    CREATE TABLE document_clusters (
                        assignment_id TEXT PRIMARY KEY,
                        doc_id TEXT NOT NULL,
                        cluster_id TEXT NOT NULL,
                        distance REAL,
                        algorithm TEXT NOT NULL,
                        assigned_date TEXT NOT NULL,
                        FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE,
                        FOREIGN KEY (cluster_id) REFERENCES clusters(cluster_id) ON DELETE CASCADE
                    )
                """)

                # ========================================
                # PHASE 3: TEMPORAL ANALYSIS TABLES
                # ========================================

                # Events table - stores detected temporal events
                cursor.execute("""
                    CREATE TABLE events (
                        event_id TEXT PRIMARY KEY,
                        event_type TEXT NOT NULL,
                        title TEXT NOT NULL,
                        description TEXT,
                        date_extracted TEXT,
                        date_normalized TEXT,
                        year INTEGER,
                        month INTEGER,
                        day INTEGER,
                        confidence REAL DEFAULT 0.5,
                        entities TEXT,
                        metadata TEXT,
                        created_date TEXT NOT NULL
                    )
                """)

                # Document-events mapping table
                cursor.execute("""
                    CREATE TABLE document_events (
                        mapping_id TEXT PRIMARY KEY,
                        doc_id TEXT NOT NULL,
                        event_id TEXT NOT NULL,
                        context TEXT,
                        position INTEGER,
                        created_date TEXT NOT NULL,
                        FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE,
                        FOREIGN KEY (event_id) REFERENCES events(event_id) ON DELETE CASCADE
                    )
                """)

                # Timeline entries table
                cursor.execute("""
                    CREATE TABLE timeline_entries (
                        entry_id TEXT PRIMARY KEY,
                        event_id TEXT NOT NULL,
                        display_date TEXT NOT NULL,
                        sort_order INTEGER NOT NULL,
                        category TEXT,
                        importance INTEGER DEFAULT 3,
                        created_date TEXT NOT NULL,
                        FOREIGN KEY (event_id) REFERENCES events(event_id) ON DELETE CASCADE
                    )
                """)

                # Create indexes for Phase 3 tables
                self.logger.info("Creating indexes for Phase 3 tables")
                cursor.execute("CREATE INDEX idx_events_type ON events(event_type)")
                cursor.execute("CREATE INDEX idx_events_year ON events(year)")
                cursor.execute("CREATE INDEX idx_events_date ON events(date_normalized)")
                cursor.execute("CREATE INDEX idx_document_events_doc ON document_events(doc_id)")
                cursor.execute("CREATE INDEX idx_document_events_event ON document_events(event_id)")
                cursor.execute("CREATE INDEX idx_timeline_entries_event ON timeline_entries(event_id)")
                cursor.execute("CREATE INDEX idx_timeline_entries_sort ON timeline_entries(sort_order)")
                cursor.execute("CREATE INDEX idx_timeline_entries_category ON timeline_entries(category)")

                # Create indexes for cluster assignment queries
                cursor.execute("CREATE INDEX idx_doc_clusters_doc ON document_clusters(doc_id)")
                cursor.execute("CREATE INDEX idx_doc_clusters_cluster ON document_clusters(cluster_id)")
                cursor.execute("CREATE INDEX idx_doc_clusters_algorithm ON document_clusters(algorithm)")
                cursor.execute("CREATE INDEX idx_doc_clusters_distance ON document_clusters(distance)")

                self.db_conn.commit()
                self.logger.info("document_clusters table created")

        # Always run migrations for schema updates (regardless of db_exists)
        self._migrate_phase3_schema()

    def _migrate_phase3_schema(self):
        """Migrate existing databases to include Phase 3 temporal analysis tables."""
        cursor = self.db_conn.cursor()

        # Check if Phase 3 tables already exist
        result = cursor.execute("""
            SELECT name FROM sqlite_master
            WHERE type='table' AND name='events'
        """).fetchone()

        if result:
            # Phase 3 tables already exist, skip migration
            return

        self.logger.info("Creating Phase 3 temporal analysis tables")

        try:
            # Create events table
            cursor.execute("""
                CREATE TABLE events (
                    event_id TEXT PRIMARY KEY,
                    event_type TEXT NOT NULL,
                    title TEXT NOT NULL,
                    description TEXT,
                    date_extracted TEXT,
                    date_normalized TEXT,
                    year INTEGER,
                    month INTEGER,
                    day INTEGER,
                    confidence REAL DEFAULT 0.5,
                    entities TEXT,
                    metadata TEXT,
                    created_date TEXT NOT NULL
                )
            """)

            # Create document-events mapping table
            cursor.execute("""
                CREATE TABLE document_events (
                    mapping_id TEXT PRIMARY KEY,
                    doc_id TEXT NOT NULL,
                    event_id TEXT NOT NULL,
                    context TEXT,
                    position INTEGER,
                    created_date TEXT NOT NULL,
                    FOREIGN KEY (doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE,
                    FOREIGN KEY (event_id) REFERENCES events(event_id) ON DELETE CASCADE
                )
            """)

            # Create timeline entries table
            cursor.execute("""
                CREATE TABLE timeline_entries (
                    entry_id TEXT PRIMARY KEY,
                    event_id TEXT NOT NULL,
                    display_date TEXT NOT NULL,
                    sort_order INTEGER NOT NULL,
                    category TEXT,
                    importance INTEGER DEFAULT 3,
                    created_date TEXT NOT NULL,
                    FOREIGN KEY (event_id) REFERENCES events(event_id) ON DELETE CASCADE
                )
            """)

            # Create indexes for Phase 3 tables
            cursor.execute("CREATE INDEX idx_events_type ON events(event_type)")
            cursor.execute("CREATE INDEX idx_events_year ON events(year)")
            cursor.execute("CREATE INDEX idx_events_date ON events(date_normalized)")
            cursor.execute("CREATE INDEX idx_document_events_doc ON document_events(doc_id)")
            cursor.execute("CREATE INDEX idx_document_events_event ON document_events(event_id)")
            cursor.execute("CREATE INDEX idx_timeline_entries_event ON timeline_entries(event_id)")
            cursor.execute("CREATE INDEX idx_timeline_entries_sort ON timeline_entries(sort_order)")
            cursor.execute("CREATE INDEX idx_timeline_entries_category ON timeline_entries(category)")

            self.db_conn.commit()
            self.logger.info("Phase 3 temporal analysis tables created successfully")

        except Exception as e:
            self.logger.error(f"Failed to create Phase 3 tables: {e}")
            self.db_conn.rollback()
            raise

    def _fts5_available(self) -> bool:
        """Check if FTS5 is available and table exists."""
        try:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM chunks_fts5 LIMIT 1")
            return True
        except Exception:
            return False

    def _load_documents(self):
        """Load documents from database, with automatic migration from JSON if needed."""
        cursor = self.db_conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM documents")
        doc_count = cursor.fetchone()[0]

        # Check if database is empty but JSON files exist (migration needed)
        if doc_count == 0 and self.index_file.exists():
            self.logger.info("Found legacy JSON index, performing automatic migration to SQLite")
            self._migrate_from_json()
        else:
            # Load documents from database
            self.logger.info(f"Loading {doc_count} documents from database")
            cursor.execute("SELECT * FROM documents")
            rows = cursor.fetchall()

            for row in rows:
                doc = DocumentMeta(
                    doc_id=row[0],
                    filename=row[1],
                    title=row[2],
                    filepath=row[3],
                    file_type=row[4],
                    total_pages=row[5],
                    total_chunks=row[6],
                    indexed_at=row[7],
                    tags=json.loads(row[8]),
                    author=row[9],
                    subject=row[10],
                    creator=row[11],
                    creation_date=row[12],
                    file_mtime=row[13] if len(row) > 13 else None,
                    file_hash=row[14] if len(row) > 14 else None,
                    source_url=row[15] if len(row) > 15 else None,
                    scrape_date=row[16] if len(row) > 16 else None,
                    scrape_config=row[17] if len(row) > 17 else None,
                    scrape_status=row[18] if len(row) > 18 else None,
                    scrape_error=row[19] if len(row) > 19 else None,
                    url_last_checked=row[20] if len(row) > 20 else None,
                    url_content_hash=row[21] if len(row) > 21 else None
                )
                self.documents[doc.doc_id] = doc

    def _migrate_from_json(self):
        """Migrate existing JSON index and chunks to SQLite database."""
        self.logger.info("Starting migration from JSON to SQLite")

        try:
            with open(self.index_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            migrated_count = 0
            for doc_data in data.get('documents', []):
                doc_id = doc_data['doc_id']

                # Load chunks from chunks/{doc_id}.json
                chunk_file = self.chunks_dir / f"{doc_id}.json"
                if not chunk_file.exists():
                    self.logger.warning(f"Missing chunk file for {doc_id}, skipping")
                    continue

                with open(chunk_file, 'r', encoding='utf-8') as f:
                    chunk_list = json.load(f)

                # Insert into database
                doc_meta = DocumentMeta(**doc_data)
                chunks = [DocumentChunk(**c) for c in chunk_list]
                self._add_document_db(doc_meta, chunks)

                self.documents[doc_id] = doc_meta
                migrated_count += 1

            self.logger.info(f"Successfully migrated {migrated_count} documents to SQLite")
            self.logger.info("JSON files preserved as backup (can be manually deleted)")

        except Exception as e:
            self.logger.error(f"Migration failed: {e}")
            raise KnowledgeBaseError(f"Failed to migrate from JSON: {e}")

    def _add_document_db(self, doc_meta: DocumentMeta, chunks: list[DocumentChunk],
                         tables: Optional[list[dict]] = None, code_blocks: Optional[list[dict]] = None,
                         facets: Optional[dict[str, set[str]]] = None, cross_refs: Optional[list[dict]] = None):
        """Add a document, chunks, tables, code blocks, facets, and cross-references to the database using a transaction."""
        cursor = self.db_conn.cursor()

        try:
            # Start transaction
            cursor.execute("BEGIN TRANSACTION")

            # Insert document
            cursor.execute("""
                INSERT OR REPLACE INTO documents
                (doc_id, filename, title, filepath, file_type, total_pages, total_chunks,
                 indexed_at, tags, author, subject, creator, creation_date, file_mtime, file_hash,
                 source_url, scrape_date, scrape_config, scrape_status, scrape_error,
                 url_last_checked, url_content_hash)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                doc_meta.doc_id,
                doc_meta.filename,
                doc_meta.title,
                doc_meta.filepath,
                doc_meta.file_type,
                doc_meta.total_pages,
                doc_meta.total_chunks,
                doc_meta.indexed_at,
                json.dumps(doc_meta.tags),
                doc_meta.author,
                doc_meta.subject,
                doc_meta.creator,
                doc_meta.creation_date,
                doc_meta.file_mtime,
                doc_meta.file_hash,
                doc_meta.source_url,
                doc_meta.scrape_date,
                doc_meta.scrape_config,
                doc_meta.scrape_status,
                doc_meta.scrape_error,
                doc_meta.url_last_checked,
                doc_meta.url_content_hash
            ))

            # Delete old chunks if re-indexing
            cursor.execute("DELETE FROM chunks WHERE doc_id = ?", (doc_meta.doc_id,))

            # Insert chunks
            for chunk in chunks:
                cursor.execute("""
                    INSERT INTO chunks
                    (doc_id, chunk_id, page, content, word_count)
                    VALUES (?, ?, ?, ?, ?)
                """, (
                    chunk.doc_id,
                    chunk.chunk_id,
                    chunk.page,
                    chunk.content,
                    chunk.word_count
                ))

            # Delete old tables if re-indexing
            cursor.execute("DELETE FROM document_tables WHERE doc_id = ?", (doc_meta.doc_id,))

            # Insert tables
            if tables:
                for table in tables:
                    cursor.execute("""
                        INSERT INTO document_tables
                        (doc_id, table_id, page, markdown, searchable_text, row_count, col_count)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (
                        doc_meta.doc_id,
                        table['table_id'],
                        table['page'],
                        table['markdown'],
                        table['searchable_text'],
                        table['row_count'],
                        table['col_count']
                    ))

            # Delete old code blocks if re-indexing
            cursor.execute("DELETE FROM document_code_blocks WHERE doc_id = ?", (doc_meta.doc_id,))

            # Insert code blocks
            if code_blocks:
                for block in code_blocks:
                    cursor.execute("""
                        INSERT INTO document_code_blocks
                        (doc_id, block_id, page, block_type, code, searchable_text, line_count)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (
                        doc_meta.doc_id,
                        block['block_id'],
                        block['page'],
                        block['block_type'],
                        block['code'],
                        block['searchable_text'],
                        block['line_count']
                    ))

            # Delete old facets if re-indexing
            cursor.execute("DELETE FROM document_facets WHERE doc_id = ?", (doc_meta.doc_id,))

            # Insert facets
            if facets:
                for facet_type, facet_values in facets.items():
                    for facet_value in facet_values:
                        cursor.execute("""
                            INSERT INTO document_facets (doc_id, facet_type, facet_value)
                            VALUES (?, ?, ?)
                        """, (doc_meta.doc_id, facet_type, facet_value))

            # Delete old cross-references if re-indexing
            cursor.execute("DELETE FROM cross_references WHERE doc_id = ?", (doc_meta.doc_id,))

            # Insert cross-references
            if cross_refs:
                for ref in cross_refs:
                    cursor.execute("""
                        INSERT INTO cross_references (doc_id, chunk_id, ref_type, ref_value, context)
                        VALUES (?, ?, ?, ?, ?)
                    """, (ref['doc_id'], ref['chunk_id'], ref['ref_type'], ref['ref_value'], ref['context']))

            # Commit transaction
            try:
                self.db_conn.commit()
            except (SystemError, Exception) as commit_error:
                # Python 3.14 SQLite bug workaround - commit may fail with various errors
                # Check if "not an error" message (Python 3.14 bug) or SystemError
                error_msg = str(commit_error).lower()
                if isinstance(commit_error, SystemError) or "not an error" in error_msg:
                    # Verify document was actually added by checking it exists
                    check_cursor = self.db_conn.cursor()
                    result = check_cursor.execute(
                        "SELECT 1 FROM documents WHERE doc_id = ?", (doc_meta.doc_id,)
                    ).fetchone()
                    if result is not None:
                        # Document was added, commit succeeded despite error
                        self.logger.warning(
                            f"Commit error (Python 3.14 bug), but document {doc_meta.doc_id} was added successfully: {commit_error}"
                        )
                        return  # Success despite error
                # Re-raise if not the Python 3.14 bug or if verification failed
                raise

        except Exception as e:
            # Rollback on error (but skip if it's the Python 3.14 "not an error" bug that we already handled)
            error_msg = str(e).lower()
            if not (isinstance(e, SystemError) or "not an error" in error_msg):
                try:
                    self.db_conn.rollback()
                except SystemError:
                    # Rollback may also fail with same bug, ignore
                    pass
                self.logger.error(f"Error adding document to database: {e}")
                raise KnowledgeBaseError(f"Failed to add document to database: {e}")
            # If it's the "not an error" bug, we already verified and returned above, so this shouldn't be reached
            # But if it is, don't wrap it again
            raise

    def _remove_document_db(self, doc_id: str) -> bool:
        """Remove a document from the database (chunks cascade automatically)."""
        cursor = self.db_conn.cursor()

        try:
            cursor.execute("DELETE FROM documents WHERE doc_id = ?", (doc_id,))
            rowcount = cursor.rowcount
            try:
                self.db_conn.commit()
            except SystemError as se:
                # Python 3.14 SQLite bug workaround - commit may return NULL
                # Check if deletion actually happened by verifying doc doesn't exist
                check_cursor = self.db_conn.cursor()
                result = check_cursor.execute("SELECT 1 FROM documents WHERE doc_id = ?", (doc_id,)).fetchone()
                if result is None:
                    # Document was deleted, commit succeeded despite SystemError
                    self.logger.warning(f"SystemError during commit (Python 3.14 bug), but deletion succeeded: {se}")
                    return rowcount > 0
                else:
                    # Deletion failed, re-raise
                    raise
            return rowcount > 0
        except Exception as e:
            try:
                self.db_conn.rollback()
            except SystemError:
                # Rollback may also fail with same bug, ignore
                pass
            self.logger.error(f"Error removing document from database: {e}")
            raise KnowledgeBaseError(f"Failed to remove document from database: {e}")

    def _get_chunks_db(self, doc_id: Optional[str] = None) -> list[DocumentChunk]:
        """Load chunks from database. If doc_id is None, load all chunks."""
        cursor = self.db_conn.cursor()

        if doc_id:
            cursor.execute("""
                SELECT c.doc_id, d.filename, d.title, c.chunk_id, c.page, c.content, c.word_count
                FROM chunks c
                JOIN documents d ON c.doc_id = d.doc_id
                WHERE c.doc_id = ?
                ORDER BY c.chunk_id
            """, (doc_id,))
        else:
            cursor.execute("""
                SELECT c.doc_id, d.filename, d.title, c.chunk_id, c.page, c.content, c.word_count
                FROM chunks c
                JOIN documents d ON c.doc_id = d.doc_id
                ORDER BY c.doc_id, c.chunk_id
            """)

        chunks = []
        for row in cursor.fetchall():
            chunk = DocumentChunk(
                doc_id=row[0],
                filename=row[1],
                title=row[2],
                chunk_id=row[3],
                page=row[4],
                content=row[5],
                word_count=row[6]
            )
            chunks.append(chunk)

        return chunks

    def _load_index(self):
        """Load existing index from disk."""
        if self.index_file.exists():
            with open(self.index_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for doc_data in data.get('documents', []):
                    doc = DocumentMeta(**doc_data)
                    self.documents[doc.doc_id] = doc
                    
            # Load chunks
            for chunk_file in self.chunks_dir.glob("*.json"):
                with open(chunk_file, 'r', encoding='utf-8') as f:
                    chunk_data = json.load(f)
                    for c in chunk_data:
                        self.chunks.append(DocumentChunk(**c))

    def _build_bm25_index(self):
        """Build BM25 index from chunks for fast searching (lazy loading from database)."""
        import time
        from concurrent.futures import ThreadPoolExecutor, as_completed

        start_time = time.time()

        if not BM25_SUPPORT:
            self.logger.info("BM25 index not built (no support)")
            return

        # Lazy load all chunks from database if not already in memory
        if not self.chunks:
            self.logger.info("Loading all chunks from database for BM25 index")
            load_start = time.time()
            self.chunks = self._get_chunks_db()
            load_time = time.time() - load_start
            self.logger.info(f"Loaded {len(self.chunks)} chunks in {load_time:.2f}s")

        if not self.chunks:
            self.logger.info("BM25 index not built (no chunks)")
            return

        # Tokenize all chunk content with preprocessing if enabled
        # Use parallel processing for faster tokenization
        self.logger.info(f"Tokenizing {len(self.chunks)} chunks...")
        tokenize_start = time.time()

        if self.use_preprocessing and len(self.chunks) > 100:
            # Parallel tokenization for large datasets
            tokenized_corpus = []
            with ThreadPoolExecutor(max_workers=4) as executor:
                # Submit all tokenization tasks
                future_to_idx = {
                    executor.submit(self._preprocess_text, chunk.content): idx
                    for idx, chunk in enumerate(self.chunks)
                }

                # Collect results in order
                results = [None] * len(self.chunks)
                for future in as_completed(future_to_idx):
                    idx = future_to_idx[future]
                    results[idx] = future.result()

                tokenized_corpus = results
        else:
            # Sequential for small datasets or no preprocessing
            tokenized_corpus = [self._preprocess_text(chunk.content) for chunk in self.chunks]

        tokenize_time = time.time() - tokenize_start
        self.logger.info(f"Tokenization completed in {tokenize_time:.2f}s")

        # Build BM25 index
        index_start = time.time()
        self.bm25 = BM25Okapi(tokenized_corpus)
        index_time = time.time() - index_start

        total_time = time.time() - start_time
        preprocessing_status = "with preprocessing" if self.use_preprocessing else "without preprocessing"
        self.logger.info(f"Built BM25 index with {len(self.chunks)} chunks ({preprocessing_status}) - Total: {total_time:.2f}s (load: {load_time:.2f}s, tokenize: {tokenize_time:.2f}s, index: {index_time:.2f}s)")

    def _fuzzy_match_terms(self, query_terms: list[str], content: str) -> tuple[bool, float]:
        """
        Check if query terms fuzzy match content using rapidfuzz.

        Args:
            query_terms: List of query terms to match
            content: Content to search in

        Returns:
            Tuple of (match_found, average_similarity_score)
        """
        if not self.use_fuzzy:
            # Fuzzy search disabled, do exact matching
            content_lower = content.lower()
            matches = sum(1 for term in query_terms if term.lower() in content_lower)
            return (matches > 0, matches / len(query_terms) if query_terms else 0.0)

        # Split content into words for fuzzy matching
        content_words = content.lower().split()

        match_scores = []
        for query_term in query_terms:
            query_term_lower = query_term.lower()

            # Check for exact match first (fastest)
            if query_term_lower in content.lower():
                match_scores.append(100.0)
                continue

            # Try fuzzy matching against content words
            best_score = 0.0
            for content_word in content_words:
                score = fuzz.ratio(query_term_lower, content_word)
                if score > best_score:
                    best_score = score
                if score >= self.fuzzy_threshold:
                    break  # Found a good enough match

            match_scores.append(best_score)

        # Consider it a match if at least one term meets the threshold
        match_found = any(score >= self.fuzzy_threshold for score in match_scores)
        avg_score = sum(match_scores) / len(match_scores) if match_scores else 0.0

        return (match_found, avg_score)

    def _preprocess_text(self, text: str) -> list[str]:
        """Preprocess text for searching: tokenize, lowercase, remove stopwords, stem.

        Args:
            text: The text to preprocess

        Returns:
            List of processed tokens
        """
        if not self.use_preprocessing:
            # No preprocessing - just lowercase and split
            return text.lower().split()

        # Tokenize and lowercase
        try:
            tokens = word_tokenize(text.lower())
        except Exception:
            # Fallback if tokenization fails
            tokens = text.lower().split()

        # Remove stopwords and apply stemming
        processed_tokens = []
        for token in tokens:
            # Keep alphanumeric tokens and hyphenated words (like VIC-II, 6502)
            # Remove pure punctuation tokens
            if token.isalnum() or ('-' in token and any(c.isalnum() for c in token)):
                # Remove stopwords (but keep technical terms with hyphens)
                if token not in self.stop_words:
                    # Apply stemming only to pure alphanumeric tokens
                    # Don't stem technical terms with hyphens/numbers
                    if self.stemmer and token.isalpha():
                        stemmed = self.stemmer.stem(token)
                        processed_tokens.append(stemmed)
                    else:
                        processed_tokens.append(token)

        return processed_tokens

    def _save_index(self):
        """Save index to disk."""
        data = {
            'documents': [asdict(doc) for doc in self.documents.values()],
            'last_updated': datetime.now().isoformat()
        }
        with open(self.index_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
    
    def _save_chunks(self, doc_id: str, chunks: list[DocumentChunk]):
        """Save chunks for a document."""
        chunk_file = self.chunks_dir / f"{doc_id}.json"
        with open(chunk_file, 'w', encoding='utf-8') as f:
            json.dump([asdict(c) for c in chunks], f)
    
    def _generate_doc_id(self, filepath: str, text_content: str = None) -> str:
        """
        Generate a unique document ID based on content hash.

        If text_content is provided, generates ID from content hash (deduplication).
        If not provided, falls back to filepath hash (legacy behavior).

        Args:
            filepath: Path to the document
            text_content: Extracted text content (optional)

        Returns:
            12-character hex string document ID
        """
        if text_content:
            # Content-based ID for deduplication
            # Normalize text: lowercase, strip whitespace
            normalized = text_content.lower().strip()
            # Hash first 10k words to handle large documents efficiently
            words = normalized.split()[:10000]
            content_sample = ' '.join(words)
            return hashlib.md5(content_sample.encode('utf-8')).hexdigest()[:12]
        else:
            # Filepath-based ID (legacy)
            return hashlib.md5(filepath.encode()).hexdigest()[:12]
    
    def _chunk_text(self, text: str, chunk_size: int = 1500, overlap: int = 200) -> list[str]:
        """
        Split text into overlapping chunks with optimized algorithm.

        Uses single-pass processing to minimize string operations.
        For large documents (>10K words), this is ~15% faster than naive split/join.

        Args:
            text: Text to chunk
            chunk_size: Number of words per chunk (default: 1500)
            overlap: Number of words to overlap between chunks (default: 200)

        Returns:
            List of text chunks
        """
        words = text.split()

        if len(words) <= chunk_size:
            return [text]

        chunks = []
        start = 0

        while start < len(words):
            end = start + chunk_size
            chunk_words = words[start:end]
            chunks.append(' '.join(chunk_words))
            start = end - overlap

        return chunks
    
    def _extract_pdf_with_ocr(self, filepath: str) -> tuple[str, int]:
        """Extract text from scanned PDF using OCR, returns (text, page_count)."""
        if not self.use_ocr:
            raise RuntimeError("OCR not enabled. Set USE_OCR=1 and install Tesseract.")

        if not self.poppler_available:
            raise RuntimeError(
                "Poppler not found! OCR requires poppler-utils.\n"
                "Install instructions:\n"
                "  Windows: Download from https://github.com/oschwartz10612/poppler-windows/releases/\n"
                "  Set POPPLER_PATH environment variable to the bin directory\n"
                "  Example: POPPLER_PATH=C:\\path\\to\\poppler-24.08.0\\Library\\bin"
            )

        try:
            self.logger.info(f"Using OCR to extract text from scanned PDF: {filepath}")
            # Convert PDF pages to images
            if self.poppler_path:
                images = convert_from_path(filepath, poppler_path=self.poppler_path)
            else:
                images = convert_from_path(filepath)

            pages = []
            for i, image in enumerate(images):
                try:
                    # Extract text using Tesseract OCR
                    text = pytesseract.image_to_string(image)
                    pages.append(text)
                    self.logger.debug(f"OCR processed page {i + 1}/{len(images)}")
                except Exception as e:
                    self.logger.error(f"OCR failed for page {i + 1}: {e}")
                    pages.append("")  # Empty text for failed page

            full_text = "\n\n--- PAGE BREAK ---\n\n".join(pages)
            self.logger.info(f"OCR extraction complete: {len(full_text)} characters from {len(images)} pages")
            return full_text, len(images)

        except Exception as e:
            self.logger.error(f"OCR extraction failed: {e}")
            raise RuntimeError(f"OCR extraction failed: {e}")

    def _check_poppler_available(self) -> bool:
        """Check if poppler is available for pdf2image.

        Returns True if poppler can be found, False otherwise.
        """
        if not OCR_SUPPORT:
            return False

        try:
            # Try to import pdf2image

            # Create a minimal test - just try to access pdfinfo
            import subprocess

            # Determine the command to check based on poppler_path
            if self.poppler_path:
                # Check if pdfinfo exists in the specified path
                pdfinfo_cmd = os.path.join(self.poppler_path, 'pdfinfo')
                if os.name == 'nt':  # Windows
                    pdfinfo_cmd += '.exe'

                if not os.path.exists(pdfinfo_cmd):
                    return False

                # Try to run pdfinfo -v (poppler uses -v not --version)
                result = subprocess.run(
                    [pdfinfo_cmd, '-v'],
                    capture_output=True,
                    timeout=5
                )
                return result.returncode == 0
            else:
                # Check if pdfinfo is in PATH
                if os.name == 'nt':  # Windows
                    result = subprocess.run(
                        ['where', 'pdfinfo'],
                        capture_output=True,
                        timeout=5
                    )
                else:  # Unix-like
                    result = subprocess.run(
                        ['which', 'pdfinfo'],
                        capture_output=True,
                        timeout=5
                    )
                return result.returncode == 0

        except Exception as e:
            self.logger.debug(f"Poppler check failed: {e}")
            return False

    def _extract_pdf_text(self, filepath: str) -> tuple[str, int, dict]:
        """Extract text from PDF with automatic OCR fallback for scanned PDFs.

        Returns (text, page_count, metadata).
        """
        if not PDF_SUPPORT:
            raise RuntimeError("PDF support not available. Install pypdf: pip install pypdf")

        reader = PdfReader(filepath)
        pages = []
        total_text_length = 0

        for page in reader.pages:
            text = page.extract_text() or ""
            pages.append(text)
            total_text_length += len(text.strip())

        # Check if PDF appears to be scanned (very little or no text extracted)
        # Threshold: if we get less than 100 characters total, likely scanned
        is_scanned = total_text_length < 100 and len(reader.pages) > 0

        if is_scanned and self.use_ocr:
            self.logger.info(f"PDF appears to be scanned ({total_text_length} chars extracted), falling back to OCR")
            try:
                ocr_text, page_count = self._extract_pdf_with_ocr(filepath)
                # Use OCR text instead of extracted pages
                pages = [ocr_text]
                # Still extract metadata from PDF
            except Exception as e:
                self.logger.warning(f"OCR fallback failed: {e}, using extracted text anyway")

        # Extract metadata
        metadata = {}
        if reader.metadata:
            # Convert metadata values to strings to handle IndirectObject references
            author = reader.metadata.get('/Author')
            metadata['author'] = str(author) if author else None

            subject = reader.metadata.get('/Subject')
            metadata['subject'] = str(subject) if subject else None

            creator = reader.metadata.get('/Creator')
            metadata['creator'] = str(creator) if creator else None

            creation_date = reader.metadata.get('/CreationDate')
            if creation_date:
                # Try to parse PDF date format (D:YYYYMMDDHHmmSS)
                try:
                    creation_date_str = str(creation_date)
                    if creation_date_str.startswith('D:'):
                        date_str = creation_date_str[2:16]  # Extract YYYYMMDDHHmmSS
                        metadata['creation_date'] = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
                    else:
                        metadata['creation_date'] = creation_date_str
                except (ValueError, TypeError, AttributeError, IndexError):
                    metadata['creation_date'] = str(creation_date)

        return "\n\n--- PAGE BREAK ---\n\n".join(pages), len(reader.pages), metadata
    
    def _extract_text_file(self, filepath: str) -> str:
        """Extract text from a text file."""
        encodings = ['utf-8', 'latin-1', 'cp1252']
        for enc in encodings:
            try:
                with open(filepath, 'r', encoding=enc) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
        raise RuntimeError(f"Could not decode {filepath}")

    def _extract_excel_file(self, filepath: str) -> tuple[str, int]:
        """
        Extract text from Excel file (.xlsx, .xls).

        Returns:
            Tuple of (text_content, sheet_count)
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("openpyxl not installed. Install with: pip install openpyxl")

        try:
            workbook = load_workbook(filepath, data_only=True)
            sheets_text = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Add sheet header
                sheets_text.append(f"\n{'='*60}\nSheet: {sheet_name}\n{'='*60}\n")

                # Extract all cell values
                rows_text = []
                for row in sheet.iter_rows(values_only=True):
                    # Filter out None values and convert to strings
                    row_values = [str(cell) if cell is not None else '' for cell in row]
                    # Skip completely empty rows
                    if any(val.strip() for val in row_values):
                        rows_text.append('\t'.join(row_values))

                sheets_text.append('\n'.join(rows_text))

            text_content = '\n\n'.join(sheets_text)
            sheet_count = len(workbook.sheetnames)

            self.logger.info(f"Extracted {sheet_count} sheets from Excel file")
            return text_content, sheet_count

        except Exception as e:
            raise RuntimeError(f"Error reading Excel file: {str(e)}")

    def _extract_html_file(self, filepath: str) -> str:
        """
        Extract text from HTML file (.html, .htm).

        Returns:
            Extracted text content
        """
        try:
            from bs4 import BeautifulSoup
        except ImportError:
            raise RuntimeError("beautifulsoup4 not installed. Install with: pip install beautifulsoup4")

        try:
            # Read the HTML file with encoding detection
            with open(filepath, 'rb') as f:
                raw_data = f.read()

            # Detect encoding
            import chardet
            detected = chardet.detect(raw_data)
            encoding = detected.get('encoding', 'utf-8')

            # Decode with detected encoding
            html_content = raw_data.decode(encoding, errors='replace')

            # Parse with BeautifulSoup
            soup = BeautifulSoup(html_content, 'html.parser')

            # Remove script and style elements
            for script in soup(['script', 'style', 'nav', 'footer', 'header']):
                script.decompose()

            # Extract text content
            text_parts = []

            # Get title if available
            if soup.title and soup.title.string:
                text_parts.append(f"Title: {soup.title.string.strip()}\n")

            # Process body or entire document
            body = soup.body if soup.body else soup

            # Handle code blocks specially to preserve formatting
            for pre in body.find_all(['pre', 'code']):
                # Mark code blocks so they're preserved
                pre_text = pre.get_text()
                if pre_text.strip():
                    text_parts.append(f"\n--- CODE BLOCK ---\n{pre_text}\n--- END CODE BLOCK ---\n")
                pre.decompose()  # Remove so we don't process again

            # Get remaining text
            main_text = body.get_text(separator='\n', strip=True)

            # Clean up excessive whitespace
            lines = [line.strip() for line in main_text.split('\n')]
            lines = [line for line in lines if line]  # Remove empty lines
            text_parts.append('\n'.join(lines))

            text_content = '\n\n'.join(text_parts)

            self.logger.info(f"Extracted HTML file ({len(text_content)} characters)")
            return text_content

        except Exception as e:
            raise RuntimeError(f"Error reading HTML file: {str(e)}")

    def _extract_tables(self, filepath: str) -> list[dict]:
        """Extract tables from PDF using pdfplumber.

        Returns a list of table dictionaries with structure:
        {
            'table_id': int,
            'page': int,
            'markdown': str,
            'searchable_text': str,
            'row_count': int,
            'col_count': int
        }
        """
        if not PDFPLUMBER_SUPPORT:
            self.logger.debug("pdfplumber not available, skipping table extraction")
            return []

        tables = []
        table_id = 0

        try:
            with pdfplumber.open(filepath) as pdf:
                for page_num, page in enumerate(pdf.pages, start=1):
                    # Extract tables from this page
                    page_tables = page.extract_tables()

                    if page_tables:
                        for table_data in page_tables:
                            if not table_data or len(table_data) == 0:
                                continue

                            # Convert table to markdown
                            markdown = self._table_to_markdown(table_data)

                            # Create searchable text (all cells joined with spaces)
                            searchable_text = " ".join(
                                str(cell).strip()
                                for row in table_data
                                for cell in row
                                if cell and str(cell).strip()
                            )

                            tables.append({
                                'table_id': table_id,
                                'page': page_num,
                                'markdown': markdown,
                                'searchable_text': searchable_text,
                                'row_count': len(table_data),
                                'col_count': len(table_data[0]) if table_data else 0
                            })
                            table_id += 1

        except Exception as e:
            self.logger.warning(f"Error extracting tables from {filepath}: {e}")
            return []

        self.logger.info(f"Extracted {len(tables)} tables from PDF")
        return tables

    def _table_to_markdown(self, table_data: list[list]) -> str:
        """Convert a table (list of lists) to markdown format."""
        if not table_data or len(table_data) == 0:
            return ""

        lines = []

        # Header row
        header = table_data[0]
        lines.append("| " + " | ".join(str(cell or "").strip() for cell in header) + " |")

        # Separator row
        lines.append("| " + " | ".join("---" for _ in header) + " |")

        # Data rows
        for row in table_data[1:]:
            lines.append("| " + " | ".join(str(cell or "").strip() for cell in row) + " |")

        return "\n".join(lines)

    def _detect_code_blocks(self, text: str) -> list[dict]:
        """Detect code blocks in text (BASIC, Assembly, Hex dumps).

        Returns a list of code block dictionaries with structure:
        {
            'block_id': int,
            'page': None,  # Page detection happens in add_document
            'block_type': str,  # 'basic', 'assembly', or 'hex'
            'code': str,
            'searchable_text': str,
            'line_count': int
        }
        """
        code_blocks = []
        block_id = 0

        # Pattern 1: BASIC code (lines starting with line numbers)
        # Example: "10 PRINT "HELLO"", "20 GOTO 10"
        basic_pattern = r'(?:^|\n)((?:\d+\s+[A-Z]+[^\n]*\n?){3,})'

        for match in re.finditer(basic_pattern, text, re.MULTILINE):
            code = match.group(1).strip()
            lines = code.split('\n')

            code_blocks.append({
                'block_id': block_id,
                'page': None,
                'block_type': 'basic',
                'code': code,
                'searchable_text': code,
                'line_count': len(lines)
            })
            block_id += 1

        # Pattern 2: Assembly code (mnemonics: LDA, STA, JMP, etc.)
        # Example: "    LDA #$00", "    STA $D020"
        assembly_pattern = r'(?:^|\n)((?:\s*(?:LDA|STA|LDX|STX|LDY|STY|JMP|JSR|RTS|BEQ|BNE|BCC|BCS|ADC|SBC|AND|ORA|EOR|INC|DEC|CMP|CPX|CPY|ASL|LSR|ROL|ROR|BIT|NOP|CLC|SEC|CLI|SEI|CLD|SED|CLV|PHA|PLA|PHP|PLP|TAX|TAY|TXA|TYA|TSX|TXS|INX|INY|DEX|DEY|BMI|BPL|BVC|BVS)[^\n]*\n?){3,})'

        for match in re.finditer(assembly_pattern, text, re.MULTILINE | re.IGNORECASE):
            code = match.group(1).strip()
            lines = code.split('\n')

            # Avoid duplicates (check if this code is already captured)
            if not any(block['code'] == code for block in code_blocks):
                code_blocks.append({
                    'block_id': block_id,
                    'page': None,
                    'block_type': 'assembly',
                    'code': code,
                    'searchable_text': code,
                    'line_count': len(lines)
                })
                block_id += 1

        # Pattern 3: Hex dumps (lines with hex values)
        # Example: "D000: 00 01 02 03 04 05 06 07"
        hex_pattern = r'(?:^|\n)((?:[0-9A-F]{4}:\s*(?:[0-9A-F]{2}\s*){8,}\n?){3,})'

        for match in re.finditer(hex_pattern, text, re.MULTILINE | re.IGNORECASE):
            code = match.group(1).strip()
            lines = code.split('\n')

            code_blocks.append({
                'block_id': block_id,
                'page': None,
                'block_type': 'hex',
                'code': code,
                'searchable_text': code,
                'line_count': len(lines)
            })
            block_id += 1

        basic_count = sum(1 for b in code_blocks if b['block_type'] == 'basic')
        assembly_count = sum(1 for b in code_blocks if b['block_type'] == 'assembly')
        hex_count = sum(1 for b in code_blocks if b['block_type'] == 'hex')
        self.logger.info(f"Detected {len(code_blocks)} code blocks ({basic_count} BASIC, {assembly_count} Assembly, {hex_count} Hex)")
        return code_blocks

    def _extract_facets(self, text: str) -> dict[str, set[str]]:
        """Extract categorizable terms for faceted search.

        Returns a dictionary of facet types to sets of values:
        {
            'hardware': {'SID', 'VIC-II', 'CIA'},
            'instruction': {'LDA', 'STA', 'JMP'},
            'register': {'$D000', '$D400'}
        }
        """
        facets = {
            'hardware': set(),
            'instruction': set(),
            'register': set()
        }

        # Extract hardware components
        facets['hardware'] = self._extract_hardware_refs(text)

        # Extract 6502 instructions
        facets['instruction'] = self._extract_instructions(text)

        # Extract register addresses
        facets['register'] = self._extract_registers(text)

        return facets

    def _extract_hardware_refs(self, text: str) -> set[str]:
        """Extract hardware component mentions from text."""
        hardware = set()

        # Hardware patterns (case-insensitive)
        patterns = {
            'SID': r'\b(?:SID|6581|8580|Sound\s+Interface\s+Device)\b',
            'VIC-II': r'\b(?:VIC-?II|VIC\s*2|6569|6567|Video\s+Interface\s+Chip)\b',
            'CIA': r'\b(?:CIA|6526|Complex\s+Interface\s+Adapter)\b',
            '6502': r'\b6502\b',
            'PLA': r'\b(?:PLA|82S100|Programmable\s+Logic\s+Array)\b',
            'Datasette': r'\b(?:Datasette|1530|C2N)\b',
            'Disk Drive': r'\b(?:1541|1571|1581|Disk\s+Drive)\b',
        }

        for component, pattern in patterns.items():
            if re.search(pattern, text, re.IGNORECASE):
                hardware.add(component)

        return hardware

    def _extract_instructions(self, text: str) -> set[str]:
        """Extract 6502 assembly instructions from text."""
        instructions = set()

        # Common 6502 mnemonics
        mnemonics = [
            'LDA', 'STA', 'LDX', 'STX', 'LDY', 'STY',
            'JMP', 'JSR', 'RTS', 'RTI',
            'BEQ', 'BNE', 'BCC', 'BCS', 'BMI', 'BPL', 'BVC', 'BVS',
            'ADC', 'SBC', 'AND', 'ORA', 'EOR',
            'INC', 'DEC', 'INX', 'INY', 'DEX', 'DEY',
            'CMP', 'CPX', 'CPY',
            'ASL', 'LSR', 'ROL', 'ROR',
            'BIT', 'NOP',
            'CLC', 'SEC', 'CLI', 'SEI', 'CLD', 'SED', 'CLV',
            'PHA', 'PLA', 'PHP', 'PLP',
            'TAX', 'TAY', 'TXA', 'TYA', 'TSX', 'TXS'
        ]

        for mnemonic in mnemonics:
            # Look for mnemonic as whole word (not part of another word)
            pattern = r'\b' + mnemonic + r'\b'
            if re.search(pattern, text, re.IGNORECASE):
                instructions.add(mnemonic)

        return instructions

    def _extract_registers(self, text: str) -> set[str]:
        """Extract memory register addresses from text."""
        registers = set()

        # Find all 4-digit hex addresses with $ prefix
        # Common C64 ranges: $D000-$DFFF (I/O), $A000-$BFFF (BASIC ROM), $E000-$FFFF (KERNAL ROM)
        register_pattern = r'\$[0-9A-Fa-f]{4}'
        matches = re.findall(register_pattern, text)

        # Normalize to uppercase and add to set
        for match in matches:
            registers.add(match.upper())

        return registers

    def _extract_cross_references(self, chunks: list[DocumentChunk], doc_id: str) -> list[dict]:
        """
        Extract cross-references from document chunks.

        Returns list of cross-reference dictionaries with keys:
        - doc_id, chunk_id, ref_type, ref_value, context
        """
        cross_refs = []

        for chunk in chunks:
            text = chunk.content
            chunk_id = chunk.chunk_id

            # Extract memory addresses ($D000-$FFFF)
            addresses = self._extract_memory_addresses(text)
            for addr in addresses:
                # Get context (sentence containing the address)
                context = self._get_reference_context(text, addr)
                cross_refs.append({
                    'doc_id': doc_id,
                    'chunk_id': chunk_id,
                    'ref_type': 'memory_address',
                    'ref_value': addr,
                    'context': context
                })

            # Extract register offsets (VIC+0, SID+4, etc.)
            offsets = self._extract_register_offsets(text)
            for offset in offsets:
                context = self._get_reference_context(text, offset)
                cross_refs.append({
                    'doc_id': doc_id,
                    'chunk_id': chunk_id,
                    'ref_type': 'register_offset',
                    'ref_value': offset,
                    'context': context
                })

            # Extract page references ("see page 156")
            page_refs = self._extract_page_references(text)
            for page_ref in page_refs:
                context = self._get_reference_context(text, str(page_ref))
                cross_refs.append({
                    'doc_id': doc_id,
                    'chunk_id': chunk_id,
                    'ref_type': 'page_reference',
                    'ref_value': str(page_ref),
                    'context': context
                })

        return cross_refs

    def _extract_memory_addresses(self, text: str) -> set[str]:
        """Extract memory addresses like $D000, $D020, etc."""
        addresses = set()
        # Match $xxxx format (4-digit hex)
        pattern = r'\$[0-9A-Fa-f]{4}\b'
        matches = re.findall(pattern, text)
        for match in matches:
            addresses.add(match.upper())
        return addresses

    def _extract_register_offsets(self, text: str) -> set[str]:
        """Extract register offset references like VIC+0, SID+4, CIA1+0."""
        offsets = set()
        # Match patterns like: VIC+0, SID+4, CIA1+12, etc.
        pattern = r'\b(VIC|SID|CIA[12]?|PLA)\s*\+\s*(\d+)\b'
        matches = re.findall(pattern, text, re.IGNORECASE)
        for chip, offset in matches:
            offsets.add(f"{chip.upper()}+{offset}")
        return offsets

    def _extract_page_references(self, text: str) -> set[int]:
        """Extract page number references like 'see page 156', 'page 42'."""
        page_nums = set()
        # Match patterns like: "page 123", "see page 456", "on page 789"
        pattern = r'\b(?:see\s+)?page\s+(\d+)\b'
        matches = re.findall(pattern, text, re.IGNORECASE)
        for page_num in matches:
            page_nums.add(int(page_num))
        return page_nums

    def _get_reference_context(self, text: str, reference: str, context_chars: int = 100) -> str:
        """Get surrounding context for a reference."""
        # Find the reference in text
        pos = text.find(reference)
        if pos == -1:
            # Try case-insensitive
            pos = text.lower().find(reference.lower())

        if pos == -1:
            return ""

        # Get surrounding context
        start = max(0, pos - context_chars)
        end = min(len(text), pos + len(reference) + context_chars)

        context = text[start:end].strip()

        # Add ellipsis if truncated
        if start > 0:
            context = "..." + context
        if end < len(text):
            context = context + "..."

        return context

    def build_suggestion_dictionary(self, rebuild: bool = False):
        """
        Build autocomplete suggestion dictionary from all documents.

        Args:
            rebuild: If True, clear existing suggestions and rebuild from scratch
        """
        self.logger.info("Building query suggestion dictionary...")
        start_time = time.time()

        cursor = self.db_conn.cursor()

        # Clear existing if rebuilding
        if rebuild:
            cursor.execute("DELETE FROM query_suggestions")
            self.db_conn.commit()

        # Extract terms from all chunks
        from collections import defaultdict
        terms = defaultdict(int)

        chunks = self._get_chunks_db()
        for chunk in chunks:
            text = chunk.content

            # Extract technical terms (ALL CAPS, 2+ chars)
            tech_terms = re.findall(r'\b[A-Z]{2,}(?:-[A-Z]+)?\b', text)  # VIC-II, SID, CIA
            for term in tech_terms:
                terms[(term, 'hardware')] += 1

            # Extract memory addresses
            addresses = re.findall(r'\$[0-9A-Fa-f]{4}', text)
            for addr in addresses:
                terms[(addr.upper(), 'register')] += 1

            # Extract 6502 instructions
            instructions = re.findall(
                r'\b(?:LDA|STA|LDX|STX|LDY|STY|TAX|TAY|TXA|TYA|TSX|TXS|'
                r'ADC|SBC|AND|ORA|EOR|INC|DEC|INX|INY|DEX|DEY|'
                r'CMP|CPX|CPY|ASL|LSR|ROL|ROR|BIT|NOP|'
                r'JMP|JSR|RTS|RTI|BEQ|BNE|BCC|BCS|BMI|BPL|BVC|BVS|'
                r'CLC|SEC|CLI|SEI|CLD|SED|CLV|PHA|PLA|PHP|PLP)\b',
                text, re.IGNORECASE
            )
            for instr in instructions:
                terms[(instr.upper(), 'instruction')] += 1

            # Extract common technical phrases (2-3 words)
            # Look for capitalized phrases like "Sprite Multiplexing", "Sound Interface Device"
            phrases = re.findall(r'\b[A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,2}\b', text)
            for phrase in phrases:
                if len(phrase) > 5:  # Avoid very short phrases
                    terms[(phrase, 'concept')] += 1

        # Store top N terms (limit to avoid bloat)
        top_terms = sorted(terms.items(), key=lambda x: x[1], reverse=True)[:2000]

        for (term, category), freq in top_terms:
            cursor.execute("""
                INSERT INTO query_suggestions (term, frequency, category)
                VALUES (?, ?, ?)
            """, (term, freq, category))

        self.db_conn.commit()

        elapsed = time.time() - start_time
        self.logger.info(f"Built suggestion dictionary with {len(top_terms)} terms in {elapsed:.2f}s")

    def get_query_suggestions(self, partial: str, max_suggestions: int = 5,
                             category: Optional[str] = None) -> list[dict]:
        """
        Get autocomplete suggestions for partial query.

        Args:
            partial: Partial query string (e.g., "VIC")
            max_suggestions: Maximum number of suggestions to return
            category: Optional category filter ('hardware', 'register', 'instruction', 'concept')

        Returns:
            List of suggestion dicts with 'term', 'frequency', and 'category'
        """
        if not partial or len(partial) < 2:
            return []

        cursor = self.db_conn.cursor()

        # Escape special FTS5 characters by quoting the query
        # FTS5 special chars: $ * " and others need to be quoted
        escaped_partial = f'"{partial}"*'

        # Use FTS5 prefix matching
        if category:
            cursor.execute("""
                SELECT term, frequency, category
                FROM query_suggestions
                WHERE term MATCH ? AND category = ?
                ORDER BY rank, frequency DESC
                LIMIT ?
            """, (escaped_partial, category, max_suggestions))
        else:
            cursor.execute("""
                SELECT term, frequency, category
                FROM query_suggestions
                WHERE term MATCH ?
                ORDER BY rank, frequency DESC
                LIMIT ?
            """, (escaped_partial, max_suggestions))

        results = []
        for row in cursor.fetchall():
            results.append({
                'term': row[0],
                'frequency': row[1],
                'category': row[2]
            })

        return results

    def _update_suggestions_for_chunks(self, chunks: list[DocumentChunk]):
        """
        Incrementally update query suggestions with terms from new chunks.

        Args:
            chunks: List of newly added chunks
        """
        from collections import defaultdict
        terms = defaultdict(int)

        # Extract terms from new chunks
        for chunk in chunks:
            text = chunk.content

            # Extract technical terms (ALL CAPS, 2+ chars)
            tech_terms = re.findall(r'\b[A-Z]{2,}(?:-[A-Z]+)?\b', text)
            for term in tech_terms:
                terms[(term, 'hardware')] += 1

            # Extract memory addresses
            addresses = re.findall(r'\$[0-9A-Fa-f]{4}', text)
            for addr in addresses:
                terms[(addr.upper(), 'register')] += 1

            # Extract 6502 instructions
            instructions = re.findall(
                r'\b(?:LDA|STA|LDX|STX|LDY|STY|TAX|TAY|TXA|TYA|TSX|TXS|'
                r'ADC|SBC|AND|ORA|EOR|INC|DEC|INX|INY|DEX|DEY|'
                r'CMP|CPX|CPY|ASL|LSR|ROL|ROR|BIT|NOP|'
                r'JMP|JSR|RTS|RTI|BEQ|BNE|BCC|BCS|BMI|BPL|BVC|BVS|'
                r'CLC|SEC|CLI|SEI|CLD|SED|CLV|PHA|PLA|PHP|PLP)\b',
                text, re.IGNORECASE
            )
            for instr in instructions:
                terms[(instr.upper(), 'instruction')] += 1

            # Extract common technical phrases (2-3 words)
            phrases = re.findall(r'\b[A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,2}\b', text)
            for phrase in phrases:
                if len(phrase) > 5:
                    terms[(phrase, 'concept')] += 1

        if not terms:
            return

        cursor = self.db_conn.cursor()

        # Update or insert terms (upsert logic)
        for (term, category), freq in terms.items():
            # Check if term already exists
            cursor.execute("""
                SELECT frequency FROM query_suggestions
                WHERE term = ? AND category = ?
            """, (term, category))

            existing = cursor.fetchone()
            if existing:
                # Update frequency
                new_freq = existing[0] + freq
                cursor.execute("""
                    DELETE FROM query_suggestions WHERE term = ? AND category = ?
                """, (term, category))
                cursor.execute("""
                    INSERT INTO query_suggestions (term, frequency, category)
                    VALUES (?, ?, ?)
                """, (term, new_freq, category))
            else:
                # Insert new term
                cursor.execute("""
                    INSERT INTO query_suggestions (term, frequency, category)
                    VALUES (?, ?, ?)
                """, (term, freq, category))

        self.db_conn.commit()
        self.logger.debug(f"Updated suggestion dictionary with {len(terms)} terms")

    def export_search_results(self, results: list[dict], format: str = 'markdown',
                             query: Optional[str] = None) -> str:
        """
        Export search results to various formats.

        Args:
            results: List of search result dicts
            format: Output format ('markdown', 'json', 'html')
            query: Optional query string to include in export

        Returns:
            Formatted string in requested format
        """
        if format == 'markdown':
            return self._export_markdown(results, query)
        elif format == 'json':
            return self._export_json(results, query)
        elif format == 'html':
            return self._export_html(results, query)
        else:
            raise ValueError(f"Unsupported export format: {format}. Use 'markdown', 'json', or 'html'.")

    def _export_markdown(self, results: list[dict], query: Optional[str] = None) -> str:
        """Export results as Markdown."""
        output = "# Search Results\n\n"

        if query:
            output += f"**Query:** {query}\n"
        output += f"**Results:** {len(results)}\n"
        output += f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        output += "---\n\n"

        for i, result in enumerate(results, 1):
            output += f"## {i}. {result.get('title', 'Untitled')}\n\n"

            # Add score if present
            if 'score' in result:
                output += f"**Score:** {result['score']:.3f}\n"
            elif 'similarity' in result:
                output += f"**Similarity:** {result['similarity']:.3f}\n"

            # Add metadata
            if 'filename' in result:
                output += f"**File:** {result['filename']}\n"
            if 'page' in result and result['page']:
                output += f"**Page:** {result['page']}\n"
            if 'doc_id' in result:
                output += f"**Doc ID:** {result['doc_id']}\n"
            if 'chunk_id' in result:
                output += f"**Chunk:** {result['chunk_id']}\n"

            output += "\n"

            # Add snippet/content
            if 'snippet' in result:
                output += f"### Excerpt\n\n{result['snippet']}\n\n"
            elif 'context' in result:
                output += f"### Context\n\n{result['context']}\n\n"

            # Add tags if present
            if 'tags' in result and result['tags']:
                tags = result['tags'] if isinstance(result['tags'], list) else []
                if tags:
                    output += f"**Tags:** {', '.join(tags)}\n\n"

            output += "---\n\n"

        return output

    def _export_json(self, results: list[dict], query: Optional[str] = None) -> str:
        """Export results as JSON."""
        export_data = {
            'query': query,
            'result_count': len(results),
            'generated_at': datetime.now().isoformat(),
            'results': results
        }
        return json.dumps(export_data, indent=2)

    def _export_html(self, results: list[dict], query: Optional[str] = None) -> str:
        """Export results as HTML."""
        html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Search Results</title>
    <style>
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
            max-width: 900px;
            margin: 0 auto;
            padding: 20px;
            line-height: 1.6;
            color: #333;
        }
        h1 { color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }
        .meta { color: #7f8c8d; margin-bottom: 20px; }
        .result {
            background: #f8f9fa;
            border-left: 4px solid #3498db;
            padding: 15px;
            margin: 20px 0;
            border-radius: 4px;
        }
        .result h2 { color: #2c3e50; margin-top: 0; }
        .result-meta { color: #7f8c8d; font-size: 0.9em; margin: 10px 0; }
        .snippet {
            background: white;
            padding: 10px;
            border-radius: 4px;
            margin: 10px 0;
            font-family: 'Courier New', monospace;
            white-space: pre-wrap;
        }
        .tags { margin-top: 10px; }
        .tag {
            display: inline-block;
            background: #3498db;
            color: white;
            padding: 3px 8px;
            border-radius: 3px;
            margin: 2px;
            font-size: 0.85em;
        }
    </style>
</head>
<body>
    <h1>🔍 Search Results</h1>
"""

        if query:
            html += f"    <div class='meta'><strong>Query:</strong> {query}</div>\n"
        html += f"    <div class='meta'><strong>Results:</strong> {len(results)}</div>\n"
        html += f"    <div class='meta'><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>\n\n"

        for i, result in enumerate(results, 1):
            html += "    <div class='result'>\n"
            html += f"        <h2>{i}. {result.get('title', 'Untitled')}</h2>\n"

            html += "        <div class='result-meta'>\n"
            if 'score' in result:
                html += f"            <strong>Score:</strong> {result['score']:.3f}<br>\n"
            elif 'similarity' in result:
                html += f"            <strong>Similarity:</strong> {result['similarity']:.3f}<br>\n"
            if 'filename' in result:
                html += f"            <strong>File:</strong> {result['filename']}<br>\n"
            if 'page' in result and result['page']:
                html += f"            <strong>Page:</strong> {result['page']}<br>\n"
            html += "        </div>\n"

            if 'snippet' in result:
                html += f"        <div class='snippet'>{result['snippet']}</div>\n"
            elif 'context' in result:
                html += f"        <div class='snippet'>{result['context']}</div>\n"

            if 'tags' in result and result['tags']:
                tags = result['tags'] if isinstance(result['tags'], list) else []
                if tags:
                    html += "        <div class='tags'>\n"
                    for tag in tags:
                        html += f"            <span class='tag'>{tag}</span>\n"
                    html += "        </div>\n"

            html += "    </div>\n\n"

        html += """</body>
</html>"""

        return html

    def _cache_key(self, method: str, **kwargs) -> str:
        """Generate a cache key from method name and arguments."""
        # Sort kwargs for consistent hashing
        sorted_items = sorted(kwargs.items())
        key_str = f"{method}:{sorted_items}"
        return hashlib.md5(key_str.encode()).hexdigest()

    def _invalidate_caches(self):
        """Clear all caches when data changes."""
        if self._search_cache is not None:
            self._search_cache.clear()
        if self._similar_cache is not None:
            self._similar_cache.clear()
        if self._semantic_cache is not None:
            self._semantic_cache.clear()
        if self._hybrid_cache is not None:
            self._hybrid_cache.clear()
        if self._faceted_cache is not None:
            self._faceted_cache.clear()
        self.logger.info("All search result caches invalidated")

    def _compute_file_hash(self, filepath: str) -> str:
        """Compute MD5 hash of file content."""
        md5_hash = hashlib.md5()
        with open(filepath, 'rb') as f:
            # Read file in chunks for memory efficiency
            for chunk in iter(lambda: f.read(8192), b''):
                md5_hash.update(chunk)
        return md5_hash.hexdigest()

    def _find_mdscrape_executable(self) -> Optional[str]:
        """Find mdscrape executable in common locations.

        Returns:
            Path to mdscrape executable, or None if not found
        """
        import shutil

        # Check if mdscrape is in PATH
        mdscrape = shutil.which('mdscrape')
        if mdscrape:
            self.logger.info(f"Found mdscrape in PATH: {mdscrape}")
            return mdscrape

        # Check common Windows/Linux paths
        common_paths = [
            Path(r'C:\Users\mit\claude\mdscrape\mdscrape.exe'),  # User-specified location
            Path(r'C:\Users\mit\claude\mdscrape\mdscrape'),
            Path.home() / 'claude' / 'mdscrape' / 'mdscrape.exe',
            Path.home() / 'claude' / 'mdscrape' / 'mdscrape',
            Path(__file__).parent.parent / 'mdscrape' / 'mdscrape.exe',
            Path(__file__).parent.parent / 'mdscrape' / 'mdscrape',
        ]

        for path in common_paths:
            if path.exists():
                self.logger.info(f"Found mdscrape at: {path}")
                return str(path)

        # Check MDSCRAPE_PATH environment variable
        env_path = os.environ.get('MDSCRAPE_PATH')
        if env_path:
            path = Path(env_path)
            if path.exists():
                self.logger.info(f"Found mdscrape via MDSCRAPE_PATH: {path}")
                return str(path)

        self.logger.warning("mdscrape executable not found. Install from: https://github.com/MichaelTroelsen/mdscrape")
        return None

    def _extract_source_url_from_md(self, md_file: Path) -> Optional[str]:
        """Extract source URL from YAML frontmatter in markdown file.

        Args:
            md_file: Path to markdown file

        Returns:
            Source URL if found, None otherwise
        """
        try:
            with open(md_file, 'r', encoding='utf-8') as f:
                content = f.read()

            # Parse YAML frontmatter (between --- delimiters)
            if content.startswith('---'):
                parts = content.split('---', 2)
                if len(parts) >= 3:
                    frontmatter = parts[1]
                    # Simple YAML parsing for 'source:' or 'url:' field
                    for line in frontmatter.split('\n'):
                        line = line.strip()
                        if line.startswith('source:') or line.startswith('url:'):
                            # Extract URL after colon
                            url = line.split(':', 1)[1].strip().strip('"\'')
                            if url:
                                return url
        except Exception as e:
            self.logger.warning(f"Failed to extract URL from {md_file}: {e}")

        return None

    def _add_scraped_document(self, filepath: str, source_url: str, title: Optional[str],
                              tags: Optional[list[str]], scrape_config: str,
                              scrape_date: str) -> DocumentMeta:
        """Add a scraped markdown document with URL metadata.

        Args:
            filepath: Path to scraped markdown file
            source_url: Original URL that was scraped
            title: Optional title for document
            tags: Optional list of tags
            scrape_config: JSON string with scraping configuration
            scrape_date: ISO timestamp of scrape

        Returns:
            DocumentMeta object for added document
        """
        # First, add document using normal flow
        doc = self.add_document(filepath, title, tags)

        # Compute content hash for change detection
        url_content_hash = self._compute_file_hash(filepath)

        # Update database with URL metadata
        with self._lock:
            cursor = self.db_conn.cursor()
            cursor.execute("""
                UPDATE documents
                SET source_url = ?,
                    scrape_date = ?,
                    scrape_config = ?,
                    scrape_status = 'success',
                    url_content_hash = ?
                WHERE doc_id = ?
            """, (source_url, scrape_date, scrape_config, url_content_hash, doc.doc_id))

            self.db_conn.commit()

        # Update in-memory object
        doc.source_url = source_url
        doc.scrape_date = scrape_date
        doc.scrape_config = scrape_config
        doc.scrape_status = 'success'
        doc.url_content_hash = url_content_hash

        # Update in documents dict
        self.documents[doc.doc_id] = doc

        self.logger.info(f"Added scraped document: {doc.title} (from {source_url})")
        return doc

    def _is_path_allowed(self, filepath: str) -> bool:
        """
        Check if a file path is within allowed directories.

        Args:
            filepath: Path to check

        Returns:
            True if path is allowed (or no restrictions configured), False otherwise
        """
        # No restrictions if allowed_dirs not configured
        if not self.allowed_dirs:
            return True

        # Resolve to absolute path to prevent path traversal
        try:
            resolved_path = Path(filepath).resolve()
        except (OSError, ValueError):
            # Invalid path
            return False

        # Check if path is within any allowed directory
        return any(
            resolved_path.is_relative_to(allowed_dir)
            for allowed_dir in self.allowed_dirs
        )

    def add_document(self, filepath: str, title: Optional[str] = None, tags: Optional[list[str]] = None,
                     progress_callback: ProgressCallback = None) -> DocumentMeta:
        """Add a document to the knowledge base.

        Args:
            filepath: Path to the document file
            title: Optional title for the document
            tags: Optional list of tags
            progress_callback: Optional callback for progress updates
        """
        # Report progress: Start
        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="add_document",
                current=0,
                total=4,
                message="Starting document ingestion",
                item=filepath
            ))

        # Resolve to absolute path to prevent path traversal
        resolved_path = Path(filepath).resolve()

        # Security: Validate path is within allowed directories
        if not self._is_path_allowed(filepath):
            self.logger.error(f"Security violation: Path outside allowed directories: {resolved_path}")
            raise SecurityError(
                f"Path outside allowed directories. File must be within: {self.allowed_dirs}"
            )

        filepath = str(resolved_path)
        self.logger.info(f"Adding document: {filepath}")

        if not os.path.exists(filepath):
            self.logger.error(f"File not found: {filepath}")
            raise DocumentNotFoundError(f"File not found: {filepath}")

        filename = os.path.basename(filepath)
        file_ext = os.path.splitext(filename)[1].lower()

        # Extract text based on file type
        total_pages = None
        pdf_metadata = {}
        try:
            if file_ext == '.pdf':
                text, total_pages, pdf_metadata = self._extract_pdf_text(filepath)
                file_type = 'pdf'
                self.logger.info(f"Extracted {total_pages} pages from PDF")
            elif file_ext in ['.xlsx', '.xls']:
                text, sheet_count = self._extract_excel_file(filepath)
                file_type = 'excel'
                total_pages = sheet_count  # Treat sheets as "pages"
                self.logger.info(f"Extracted Excel file with {sheet_count} sheets ({len(text)} characters)")
            elif file_ext in ['.html', '.htm']:
                text = self._extract_html_file(filepath)
                file_type = 'html'
                self.logger.info(f"Extracted HTML file ({len(text)} characters)")
            elif file_ext in ['.txt', '.md', '.asm', '.bas', '.inc', '.s']:
                text = self._extract_text_file(filepath)
                file_type = 'text'
                self.logger.info(f"Extracted text file ({len(text)} characters)")
            else:
                raise UnsupportedFileTypeError(f"Unsupported file type: {file_ext}")
        except (UnsupportedFileTypeError, DocumentNotFoundError):
            raise
        except Exception as e:
            self.logger.error(f"Error extracting {filepath}: {e}")
            raise KnowledgeBaseError(f"Error extracting document: {e}")

        # Report progress: Text extraction complete
        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="add_document",
                current=1,
                total=4,
                message=f"Text extraction complete ({len(text)} characters)",
                item=filename
            ))

        # Extract tables from PDFs
        tables = []
        if file_type == 'pdf':
            tables = self._extract_tables(filepath)
            if tables:
                self.logger.info(f"Extracted {len(tables)} tables from PDF")

        # Detect code blocks in text
        code_blocks = self._detect_code_blocks(text)
        if code_blocks:
            self.logger.info(f"Detected {len(code_blocks)} code blocks")

        # Extract facets for faceted search
        facets = self._extract_facets(text)
        facet_count = sum(len(values) for values in facets.values())
        if facet_count > 0:
            self.logger.info(f"Extracted {facet_count} facets ({len(facets['hardware'])} hardware, {len(facets['instruction'])} instructions, {len(facets['register'])} registers)")

        # Generate content-based doc_id for deduplication
        doc_id = self._generate_doc_id(filepath, text)

        # Thread-safe duplicate check
        with self._lock:
            # Check for duplicate content
            if doc_id in self.documents:
                existing_doc = self.documents[doc_id]
                self.logger.warning(f"Duplicate content detected: {filepath}")
                self.logger.warning(f"  Matches existing document: {existing_doc.filepath}")
                self.logger.info(f"Skipping duplicate - returning existing document {doc_id}")
                return existing_doc
        
        # Create chunks
        text_chunks = self._chunk_text(text)
        chunks = []
        for i, chunk_text in enumerate(text_chunks):
            # Estimate page number for PDFs based on PAGE BREAK markers
            page_num = None
            if file_type == 'pdf' and '--- PAGE BREAK ---' in text:
                # Count PAGE BREAK markers before this chunk
                chunk_start_pos = text.find(chunk_text[:100])  # Find chunk in full text
                if chunk_start_pos >= 0:
                    page_breaks_before = text[:chunk_start_pos].count('--- PAGE BREAK ---')
                    page_num = page_breaks_before + 1  # Pages are 1-indexed

            chunk = DocumentChunk(
                doc_id=doc_id,
                filename=filename,
                title=title or filename,
                chunk_id=i,
                page=page_num,
                content=chunk_text,
                word_count=len(chunk_text.split())
            )
            chunks.append(chunk)

        # Report progress: Chunking complete
        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="add_document",
                current=2,
                total=4,
                message=f"Created {len(chunks)} chunks",
                item=filename
            ))

        # Compute file modification time and content hash for update detection
        file_mtime = os.path.getmtime(resolved_path)
        file_hash = self._compute_file_hash(resolved_path)

        # Extract cross-references for content linking
        cross_refs = self._extract_cross_references(chunks, doc_id)
        if cross_refs:
            self.logger.info(f"Extracted {len(cross_refs)} cross-references")

        # Create metadata
        doc_meta = DocumentMeta(
            doc_id=doc_id,
            filename=filename,
            title=title or filename,
            filepath=filepath,
            file_type=file_type,
            total_pages=total_pages,
            total_chunks=len(chunks),
            indexed_at=datetime.now().isoformat(),
            tags=tags or [],
            author=pdf_metadata.get('author'),
            subject=pdf_metadata.get('subject'),
            creator=pdf_metadata.get('creator'),
            creation_date=pdf_metadata.get('creation_date'),
            file_mtime=file_mtime,
            file_hash=file_hash
        )

        # Thread-safe database insertion and cache invalidation
        with self._lock:
            # Add to database (with tables, code blocks, facets, and cross-references)
            self._add_document_db(doc_meta, chunks, tables=tables, code_blocks=code_blocks, facets=facets, cross_refs=cross_refs)
            self.documents[doc_id] = doc_meta

            # Report progress: Database insertion complete
            if progress_callback:
                progress_callback(ProgressUpdate(
                    operation="add_document",
                    current=3,
                    total=4,
                    message="Stored in database",
                    item=filename
                ))

            # Invalidate BM25 index (will be rebuilt on next search)
            self.bm25 = None

            # Incrementally add chunks to embeddings (faster than full rebuild)
            if self.use_semantic:
                self._add_chunks_to_embeddings(chunks)

            # Update query suggestions with new terms
            self._update_suggestions_for_chunks(chunks)

            # Invalidate search caches
            self._invalidate_caches()

        self.logger.info(f"Successfully indexed document {doc_id}: {filename} ({len(chunks)} chunks)")

        # Report progress: Complete
        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="add_document",
                current=4,
                total=4,
                message="Document indexed successfully",
                item=filename
            ))

        # Auto-queue entity extraction (configurable via environment variable)
        auto_extract = os.getenv('AUTO_EXTRACT_ENTITIES', '1') == '1'
        if auto_extract:
            try:
                result = self.queue_entity_extraction(
                    doc_id=doc_meta.doc_id,
                    confidence_threshold=0.6,
                    skip_if_exists=True
                )
                if result.get('queued'):
                    self.logger.info(f"Auto-queued entity extraction job {result['job_id']} for document {doc_meta.doc_id}")
                else:
                    self.logger.debug(f"Entity extraction not queued: {result.get('reason', 'unknown')}")
            except Exception as e:
                # Don't fail document ingestion if extraction queueing fails
                self.logger.warning(f"Failed to auto-queue entity extraction: {e}")

        return doc_meta

    def _detect_and_extract_frames(self, url: str) -> list[str]:
        """Detect HTML frames/iframes and extract their source URLs.

        Args:
            url: The URL to check for frames

        Returns:
            List of frame source URLs (relative URLs converted to absolute)
        """
        import re
        import requests

        try:
            # Fetch the HTML content with longer timeout
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            html_content = response.text

            # Check if this is a frameset page
            if '<FRAMESET' not in html_content.upper() and '<FRAME' not in html_content.upper() and '<IFRAME' not in html_content.upper():
                return []

            self.logger.info(f"Detected frameset page at {url}")

            # Extract frame/iframe src attributes using regex
            # Match <frame src="..."> and <iframe src="...">
            frame_pattern = r'<(?:frame|iframe)[^>]+src=["\']([\w\-\.\/]+)["\']'
            matches = re.findall(frame_pattern, html_content, re.IGNORECASE)

            if not matches:
                return []

            # Convert relative URLs to absolute
            from urllib.parse import urljoin
            frame_urls = []
            for src in matches:
                # Skip external URLs and special targets
                if src.startswith('http://') or src.startswith('https://') or src.startswith('javascript:'):
                    continue

                absolute_url = urljoin(url, src)
                frame_urls.append(absolute_url)
                self.logger.info(f"Found frame source: {absolute_url}")

            return frame_urls

        except requests.exceptions.Timeout:
            self.logger.debug(f"Timeout detecting frames at {url} (will use mdscrape instead)")
            return []
        except Exception as e:
            self.logger.debug(f"Could not detect frames at {url}: {e} (will use mdscrape instead)")
            return []

    def scrape_url(self, url: str, title: Optional[str] = None, tags: Optional[list[str]] = None,
                   follow_links: bool = True, same_domain_only: bool = True,
                   max_pages: int = 50, depth: int = 3, limit: Optional[str] = None,
                   threads: int = 10, delay: int = 100, selector: Optional[str] = None,
                   progress_callback: ProgressCallback = None) -> dict:
        """Scrape a URL using mdscrape and add resulting documents to knowledge base.

        Supports recursive scraping of entire websites by following links.

        Args:
            url: Starting URL to scrape (e.g., http://www.sidmusic.org/sid/)
            title: Optional base title for scraped documents
            tags: Optional list of tags (domain name auto-added)
            follow_links: Follow links to scrape sub-pages (default: True)
            same_domain_only: Only follow links on the same domain (default: True)
            max_pages: Maximum number of pages to scrape (default: 50)
            depth: Maximum crawl depth - how many link levels to follow (default: 3)
            limit: Advanced: Limit scraping to URLs with this prefix (overrides same_domain_only)
            threads: Number of concurrent threads (default: 10)
            delay: Delay between requests in ms (default: 100)
            selector: CSS selector for main content (optional)
            progress_callback: Optional callback for progress updates

        Examples:
            # Scrape single page only
            kb.scrape_url("http://example.com/page.html", follow_links=False)

            # Scrape entire site (stay on same domain, max 3 levels deep)
            kb.scrape_url("http://www.sidmusic.org/sid/", follow_links=True, same_domain_only=True, depth=3)

            # Scrape specific section (limit to /sid/ prefix)
            kb.scrape_url("http://www.sidmusic.org/sid/", limit="http://www.sidmusic.org/sid/")

        Returns:
            Dictionary with scraping results:
            {
                'status': 'success' | 'partial' | 'failed',
                'url': original_url,
                'output_dir': path_to_scraped_files,
                'files_scraped': count,
                'docs_added': count,
                'docs_updated': count,
                'docs_failed': count,
                'pages_scraped': list_of_urls,
                'error': error_message (if failed),
                'doc_ids': [list of added doc_ids]
            }
        """
        import subprocess
        from urllib.parse import urlparse
        from datetime import datetime

        # 1. Validate URL
        try:
            parsed = urlparse(url)
            if not parsed.scheme or not parsed.netloc:
                raise ValueError(f"Invalid URL: {url}")
            if parsed.scheme not in ['http', 'https']:
                raise ValueError(f"Only HTTP/HTTPS URLs supported: {url}")
        except Exception as e:
            return {
                'status': 'failed',
                'url': url,
                'error': f"Invalid URL: {str(e)}"
            }

        # 2. Extract domain for auto-tagging
        domain = parsed.netloc.replace('www.', '')
        if tags is None:
            tags = []
        tags = list(tags) + [domain, 'scraped']

        # 3. Setup output directory in scraped_docs
        scraped_base = self.data_dir / "scraped_docs"
        scraped_base.mkdir(exist_ok=True)

        # Use domain + timestamp for unique output dir
        safe_domain = domain.replace('.', '_').replace(':', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = scraped_base / f"{safe_domain}_{timestamp}"

        # 3a. Handle follow_links and same_domain_only parameters
        if not follow_links:
            # Don't follow any links - just scrape the single page
            depth = 1
            self.logger.info("follow_links=False: Scraping single page only (depth=1)")

        # 3b. Detect and handle HTML frames
        frame_urls = self._detect_and_extract_frames(url)
        if frame_urls:
            self.logger.info(f"Detected {len(frame_urls)} frame(s), will scrape each individually")

            # Scrape each frame source recursively
            all_doc_ids = []
            all_files_scraped = 0
            all_docs_added = 0

            for frame_url in frame_urls:
                self.logger.info(f"Scraping frame: {frame_url}")

                # For frames, use the parent directory as limit if same_domain_only
                # This allows following links from the frame
                frame_limit = limit
                if same_domain_only and limit is None:
                    # Use the parent directory of the original URL
                    frame_limit = f"{parsed.scheme}://{parsed.netloc}{parsed.path.rsplit('/', 1)[0]}"
                    if not frame_limit.endswith('/'):
                        frame_limit += '/'

                frame_result = self.scrape_url(
                    url=frame_url,
                    title=title,
                    tags=tags,
                    follow_links=follow_links,
                    same_domain_only=False,  # Disable auto-limit for frames
                    max_pages=max_pages,
                    depth=depth,
                    limit=frame_limit,  # Use parent directory as limit
                    threads=threads,
                    delay=delay,
                    selector=selector,
                    progress_callback=progress_callback
                )

                if frame_result['status'] == 'success':
                    all_doc_ids.extend(frame_result.get('doc_ids', []))
                    all_files_scraped += frame_result.get('files_scraped', 0)
                    all_docs_added += frame_result.get('docs_added', 0)

            # Return combined results from all frames
            return {
                'status': 'success',
                'url': url,
                'frames_detected': len(frame_urls),
                'files_scraped': all_files_scraped,
                'docs_added': all_docs_added,
                'docs_updated': 0,
                'docs_failed': 0,
                'doc_ids': all_doc_ids,
                'message': f'Scraped {len(frame_urls)} frames with {all_docs_added} total documents'
            }

        if same_domain_only and limit is None:
            # Automatically set limit to base domain URL to stay on same domain
            # Extract base URL (scheme + netloc + path up to last /)
            base_url = f"{parsed.scheme}://{parsed.netloc}"

            # If URL has a path, use it as the limit prefix
            if parsed.path and parsed.path != '/':
                # Get the directory part of the path (not the file)
                path_parts = parsed.path.rstrip('/').split('/')
                if path_parts:
                    # Use the full path as prefix to stay within that section
                    base_path = '/'.join(path_parts)
                    limit = f"{base_url}{base_path}"
                else:
                    limit = base_url
            else:
                limit = base_url

            self.logger.info(f"same_domain_only=True: Limiting to URLs starting with '{limit}'")

        # 4. Build mdscrape command
        mdscrape_path = self._find_mdscrape_executable()
        if not mdscrape_path:
            return {
                'status': 'failed',
                'url': url,
                'error': 'mdscrape executable not found. Set MDSCRAPE_PATH or install from: https://github.com/MichaelTroelsen/mdscrape'
            }

        cmd = [
            mdscrape_path,
            url,
            '--output', str(output_dir),
            '--depth', str(depth),
            '--threads', str(threads),
            '--delay', str(delay)
        ]

        if limit:
            cmd.extend(['--limit', limit])
        if selector:
            cmd.extend(['--selector', selector])
        # Note: max_pages is a UI parameter only - mdscrape doesn't support it yet
        # Use depth to control crawl scope instead

        # 5. Store scrape config
        scrape_config = {
            'url': url,
            'follow_links': follow_links,
            'same_domain_only': same_domain_only,
            'max_pages': max_pages,
            'depth': depth,
            'limit': limit,
            'threads': threads,
            'delay': delay,
            'selector': selector,
            'timestamp': datetime.now().isoformat()
        }
        scrape_config_json = json.dumps(scrape_config)

        # 6. Execute mdscrape
        self.logger.info(f"Scraping URL: {url}")
        self.logger.info(f"Command: {' '.join(cmd)}")

        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="scrape_url",
                current=0,
                total=100,
                message="Starting web scraping",
                item=url
            ))

        try:
            # Execute with real-time output streaming
            import time
            import re
            from threading import Thread, Event
            from queue import Queue, Empty

            process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                bufsize=1  # Line buffered
            )

            # Queues for capturing output
            stdout_queue = Queue()
            stderr_queue = Queue()

            # Helper to read output in background thread
            def enqueue_output(stream, queue, stop_event):
                try:
                    for line in iter(stream.readline, ''):
                        if stop_event.is_set():
                            break
                        queue.put(line)
                except Exception:
                    pass
                finally:
                    stream.close()

            stop_event = Event()
            stdout_thread = Thread(target=enqueue_output, args=(process.stdout, stdout_queue, stop_event))
            stderr_thread = Thread(target=enqueue_output, args=(process.stderr, stderr_queue, stop_event))
            stdout_thread.daemon = True
            stderr_thread.daemon = True
            stdout_thread.start()
            stderr_thread.start()

            # Track progress
            pages_scraped = 0
            current_url = None
            last_update_time = time.time()
            timeout_warned = False
            stdout_lines = []
            stderr_lines = []

            # Process output in real-time
            while process.poll() is None:
                # Check stdout
                try:
                    line = stdout_queue.get(timeout=0.1)
                    stdout_lines.append(line)

                    # Parse mdscrape output for current URL
                    # mdscrape typically outputs: "Scraping: https://example.com/page"
                    url_match = re.search(r'(?:Scraping|Processing|Fetching)[:\s]+(\S+)', line, re.IGNORECASE)
                    if url_match:
                        current_url = url_match.group(1)
                        pages_scraped += 1
                        last_update_time = time.time()
                        timeout_warned = False

                        # Update progress
                        self.logger.info(f"[{pages_scraped}/{max_pages}] Scraping: {current_url}")

                        if progress_callback:
                            progress_callback(ProgressUpdate(
                                operation="scrape_url",
                                current=min(pages_scraped, max_pages),
                                total=max_pages,
                                message=f"Scraping page {pages_scraped}/{max_pages}",
                                item=current_url
                            ))

                except Empty:
                    pass

                # Check stderr
                try:
                    line = stderr_queue.get_nowait()
                    stderr_lines.append(line)
                    # Log errors but don't stop
                    if 'error' in line.lower() and 'image' not in line.lower():
                        self.logger.warning(f"Scrape warning: {line.strip()}")
                except Empty:
                    pass

                # Check for timeout on individual page (60 seconds)
                time_since_update = time.time() - last_update_time
                if time_since_update > 60 and not timeout_warned:
                    timeout_warned = True
                    warning_msg = f"⚠️ No progress for {int(time_since_update)} seconds"
                    if current_url:
                        warning_msg += f" (current: {current_url})"
                    self.logger.warning(warning_msg)

                    if progress_callback:
                        progress_callback(ProgressUpdate(
                            operation="scrape_url",
                            current=pages_scraped,
                            total=max_pages,
                            message=f"⚠️ Page taking longer than 60s...",
                            item=current_url or "unknown"
                        ))

            # Wait for process to complete
            process.wait(timeout=60)

            # Stop background threads
            stop_event.set()
            stdout_thread.join(timeout=1)
            stderr_thread.join(timeout=1)

            # Collect remaining output
            while not stdout_queue.empty():
                try:
                    stdout_lines.append(stdout_queue.get_nowait())
                except Empty:
                    break

            while not stderr_queue.empty():
                try:
                    stderr_lines.append(stderr_queue.get_nowait())
                except Empty:
                    break

            # Check for errors but don't fail if files were scraped
            stdout_output = ''.join(stdout_lines)
            stderr_output = ''.join(stderr_lines)

            if process.returncode != 0:
                error_msg = stderr_output or stdout_output or "Unknown error"

                # Count image-related errors (not critical failures)
                image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.svg', '.webp', '.ico']
                error_lines = error_msg.split('\n')
                image_errors = sum(1 for line in error_lines
                                 if any(ext in line.lower() for ext in image_extensions))
                total_errors = len([line for line in error_lines if 'Error:' in line or 'error' in line.lower()])

                # If all errors are image-related, treat as warning not failure
                if image_errors > 0 and total_errors > 0 and image_errors == total_errors:
                    self.logger.warning(f"Scraping completed with {image_errors} image-related errors (expected)")
                else:
                    # Log full error but continue - we'll check if any files were scraped
                    self.logger.warning(f"Scraping completed with errors: {error_msg[:500]}...")
            else:
                self.logger.info(f"[OK] Scraping completed successfully ({pages_scraped} pages)")

        except subprocess.TimeoutExpired:
            self.logger.error("Scraping timeout (>1 hour)")
            return {
                'status': 'failed',
                'url': url,
                'error': 'Scraping timeout (>1 hour)'
            }
        except Exception as e:
            self.logger.error(f"Scraping error: {e}")
            return {
                'status': 'failed',
                'url': url,
                'error': f"Scraping error: {str(e)}"
            }

        # 7. Find all generated markdown files
        if not output_dir.exists():
            return {
                'status': 'failed',
                'url': url,
                'error': f"Output directory not created: {output_dir}"
            }

        md_files = list(output_dir.rglob('*.md'))

        if not md_files:
            return {
                'status': 'failed',
                'url': url,
                'error': f"No markdown files generated in {output_dir}"
            }

        self.logger.info(f"Found {len(md_files)} markdown files to process")

        # 8. Add each file to knowledge base
        added_docs = []
        failed_docs = []
        scrape_date = datetime.now().isoformat()

        for i, md_file in enumerate(md_files):
            if progress_callback:
                progress_callback(ProgressUpdate(
                    operation="scrape_url",
                    current=i,
                    total=len(md_files),
                    message="Adding scraped document",
                    item=md_file.name
                ))

            try:
                # Extract source URL from frontmatter
                source_url_for_file = self._extract_source_url_from_md(md_file)
                if not source_url_for_file:
                    source_url_for_file = url  # Fallback to base URL

                # Generate title from domain + page path
                if title:
                    # Use provided base title
                    doc_title = title
                else:
                    # Extract page name from URL path
                    parsed_source = urlparse(source_url_for_file)
                    page_path = parsed_source.path.strip('/')

                    # If it's the index/root, use domain name
                    if not page_path or page_path.lower() in ['index', 'index.html', 'index.htm']:
                        page_name = "Home"
                    else:
                        # Use the last part of the path as page name
                        page_name = page_path.split('/')[-1]
                        # Remove file extensions
                        page_name = page_name.replace('.html', '').replace('.htm', '').replace('.php', '')
                        # Clean up formatting
                        page_name = page_name.replace('_', ' ').replace('-', ' ').title()

                    # Combine domain + page name
                    doc_title = f"{domain} - {page_name}"

                # Add document with URL metadata
                doc = self._add_scraped_document(
                    filepath=str(md_file),
                    source_url=source_url_for_file,
                    title=doc_title,
                    tags=tags,
                    scrape_config=scrape_config_json,
                    scrape_date=scrape_date
                )
                added_docs.append(doc.doc_id)
                self.logger.info(f"Added: {doc.title} ({doc.doc_id})")

            except Exception as e:
                self.logger.error(f"Failed to add {md_file}: {e}")
                failed_docs.append(str(md_file))

        # 9. Return results
        status = 'success' if not failed_docs else ('partial' if added_docs else 'failed')

        result_dict = {
            'status': status,
            'url': url,
            'output_dir': str(output_dir),
            'files_scraped': len(md_files),
            'docs_added': len(added_docs),
            'docs_updated': 0,
            'docs_failed': len(failed_docs),
            'doc_ids': added_docs
        }

        if failed_docs:
            result_dict['error'] = f"{len(failed_docs)} files failed to add"

        self.logger.info(f"Scraping complete: {status} - Added {len(added_docs)}/{len(md_files)} documents")

        return result_dict

    def rescrape_document(self, doc_id: str, progress_callback: ProgressCallback = None) -> dict:
        """Re-scrape an existing URL-sourced document.

        Args:
            doc_id: Document ID to re-scrape
            progress_callback: Optional progress callback

        Returns:
            Dictionary with re-scrape results (same format as scrape_url)
        """
        # Get document metadata
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        doc = self.documents[doc_id]

        # Check if document has source URL
        if not doc.source_url:
            raise ValueError(f"Document is not URL-sourced: {doc_id}")

        self.logger.info(f"Re-scraping document: {doc.title} (from {doc.source_url})")

        # Parse original scrape config
        scrape_config = {}
        if doc.scrape_config:
            try:
                scrape_config = json.loads(doc.scrape_config)
            except Exception as e:
                self.logger.warning(f"Failed to parse scrape config: {e}")

        # Remove old document
        self.logger.info(f"Removing old document version: {doc_id}")
        self.remove_document(doc_id)

        # Re-scrape with original config
        result = self.scrape_url(
            url=doc.source_url,
            title=doc.title,
            tags=doc.tags,
            depth=scrape_config.get('depth', 50),
            limit=scrape_config.get('limit'),
            threads=scrape_config.get('threads', 10),
            delay=scrape_config.get('delay', 100),
            selector=scrape_config.get('selector'),
            progress_callback=progress_callback
        )

        # Add rescrape metadata to result
        result['rescrape'] = True
        result['old_doc_id'] = doc_id

        self.logger.info(f"Re-scrape complete: {result['status']}")
        return result

    def _discover_urls(self, base_url: str, config: Optional[dict] = None, max_pages: int = 100) -> set:
        """Discover all URLs at a website by crawling.

        Args:
            base_url: Starting URL to crawl
            config: Optional scrape configuration with depth, limit, etc.
            max_pages: Maximum number of pages to crawl (default: 100)

        Returns:
            Set of discovered URLs
        """
        from urllib.parse import urljoin, urlparse
        from bs4 import BeautifulSoup
        import requests

        discovered = set()
        to_visit = {base_url}
        visited = set()

        # Extract config parameters
        if config:
            depth = min(config.get('depth', 3), 5)  # Cap at 5 to prevent excessive crawling
            limit = config.get('limit')
            same_domain_only = config.get('same_domain_only', True)
            follow_links = config.get('follow_links', True)
        else:
            depth = 3
            limit = None
            same_domain_only = True
            follow_links = True

        # If follow_links is False, just check the single URL
        if not follow_links:
            depth = 1
            self.logger.info("follow_links=False, checking single URL only")

        base_parsed = urlparse(base_url)
        base_domain = base_parsed.netloc

        self.logger.info(f"URL Discovery: depth={depth}, same_domain={same_domain_only}, limit={limit}, max_pages={max_pages}")

        # Crawl with depth tracking
        current_depth = 0
        total_fetched = 0

        while to_visit and current_depth < depth and total_fetched < max_pages:
            current_level = to_visit.copy()
            to_visit.clear()

            self.logger.info(f"Crawl depth {current_depth + 1}/{depth}: {len(current_level)} URLs to check")

            for url in current_level:
                if url in visited:
                    continue

                if total_fetched >= max_pages:
                    self.logger.info(f"Reached max_pages limit ({max_pages})")
                    break

                visited.add(url)
                total_fetched += 1

                try:
                    # Fetch page with timeout
                    self.logger.debug(f"Fetching: {url}")
                    response = requests.get(url, timeout=15, allow_redirects=True)

                    if response.status_code != 200:
                        self.logger.debug(f"Non-200 status ({response.status_code}): {url}")
                        continue

                    # Add to discovered (successful fetch)
                    discovered.add(url)
                    self.logger.debug(f"Discovered: {url}")

                    # Parse HTML to find links (only if following links and not at max depth)
                    if follow_links and current_depth < depth - 1:
                        if 'text/html' in response.headers.get('Content-Type', ''):
                            soup = BeautifulSoup(response.content, 'html.parser')

                            # Find all links
                            links_found = 0
                            for link in soup.find_all('a', href=True):
                                href = link['href']

                                # Convert relative URLs to absolute
                                absolute_url = urljoin(url, href)

                                # Remove fragments
                                absolute_url = absolute_url.split('#')[0]

                                # Skip if already visited or queued
                                if absolute_url in visited or absolute_url in to_visit:
                                    continue

                                # Parse the discovered URL
                                link_parsed = urlparse(absolute_url)

                                # Apply same_domain_only filter
                                if same_domain_only and link_parsed.netloc != base_domain:
                                    continue

                                # Apply limit filter
                                if limit and not absolute_url.startswith(limit):
                                    continue

                                # Skip non-HTTP(S) URLs
                                if link_parsed.scheme not in ['http', 'https']:
                                    continue

                                # Add to next level
                                to_visit.add(absolute_url)
                                links_found += 1

                            if links_found > 0:
                                self.logger.debug(f"Found {links_found} links on {url}")

                except requests.Timeout:
                    self.logger.warning(f"Timeout fetching {url}")
                    continue
                except Exception as e:
                    self.logger.debug(f"Error discovering {url}: {e}")
                    continue

            current_depth += 1

        self.logger.info(f"Discovery complete: {len(discovered)} URLs found (visited {total_fetched} pages at depth {current_depth})")
        return discovered

    def check_url_updates(self, auto_rescrape: bool = False, check_structure: bool = True) -> dict:
        """Check all URL-sourced documents for updates.

        Args:
            auto_rescrape: If True, automatically re-scrape changed URLs
            check_structure: If True, check for new/missing sub-pages (slower but comprehensive)

        Returns:
            Dictionary with update information:
            {
                'unchanged': [list of docs with no changes],
                'changed': [list of docs with updates available],
                'failed': [list of docs where check failed],
                'rescraped': [list of doc_ids that were re-scraped],
                'new_pages': [list of newly discovered URLs not in database],
                'missing_pages': [list of docs in database but no longer accessible],
                'scrape_sessions': [list of detected scrape sessions checked]
            }
        """
        from datetime import datetime, timezone
        import json

        results = {
            'unchanged': [],
            'changed': [],
            'failed': [],
            'rescraped': [],
            'new_pages': [],
            'missing_pages': [],
            'scrape_sessions': []
        }

        # Find all URL-sourced documents
        url_docs = [doc for doc in self.documents.values() if doc.source_url]

        if not url_docs:
            self.logger.info("No URL-sourced documents to check")
            return results

        self.logger.info(f"Checking {len(url_docs)} URL-sourced documents for updates")

        # Group documents by scrape session (same base URL)
        scrape_sessions = {}
        for doc in url_docs:
            # Parse scrape config to get base URL
            if doc.scrape_config:
                try:
                    config = json.loads(doc.scrape_config)
                    base_url = config.get('url', doc.source_url)
                except (json.JSONDecodeError, KeyError, AttributeError, TypeError):
                    base_url = doc.source_url
            else:
                base_url = doc.source_url

            # Group by base URL
            if base_url not in scrape_sessions:
                scrape_sessions[base_url] = {
                    'base_url': base_url,
                    'config': config if doc.scrape_config else None,
                    'docs': []
                }
            scrape_sessions[base_url]['docs'].append(doc)

        self.logger.info(f"Found {len(scrape_sessions)} scrape sessions to check")

        # Check each scrape session
        for session_key, session in scrape_sessions.items():
            session_result = {
                'base_url': session['base_url'],
                'docs_count': len(session['docs']),
                'changed': 0,
                'unchanged': 0,
                'new': 0,
                'missing': 0
            }
            results['scrape_sessions'].append(session_result)

            # Get stored URLs for this session
            stored_urls = {doc.source_url: doc for doc in session['docs']}

            # If check_structure is enabled, discover current URLs
            discovered_urls = set()
            if check_structure:
                try:
                    self.logger.info(f"Discovering URLs for: {session['base_url']}")
                    discovered_urls = self._discover_urls(
                        session['base_url'],
                        session['config']
                    )
                    self.logger.info(f"Discovered {len(discovered_urls)} URLs")
                except Exception as e:
                    self.logger.error(f"Failed to discover URLs for {session['base_url']}: {e}")

            # Check each stored document
            for doc in session['docs']:
                try:
                    import requests

                    # Try HEAD request first (faster)
                    response = requests.head(doc.source_url, timeout=10, allow_redirects=True)

                    # Update last_checked timestamp
                    with self._lock:
                        cursor = self.db_conn.cursor()
                        cursor.execute("""
                            UPDATE documents
                            SET url_last_checked = ?
                            WHERE doc_id = ?
                        """, (datetime.now().isoformat(), doc.doc_id))
                        self.db_conn.commit()

                    # Check if page still exists
                    if response.status_code == 404:
                        self.logger.warning(f"Page no longer exists: {doc.source_url}")
                        results['missing_pages'].append({
                            'doc_id': doc.doc_id,
                            'title': doc.title,
                            'url': doc.source_url
                        })
                        session_result['missing'] += 1
                        continue

                    # Check Last-Modified header if available
                    page_changed = False
                    if 'Last-Modified' in response.headers:
                        from email.utils import parsedate_to_datetime
                        last_modified = parsedate_to_datetime(response.headers['Last-Modified'])

                        if doc.scrape_date:
                            scrape_dt = datetime.fromisoformat(doc.scrape_date)
                            # Ensure both datetimes are timezone-aware for comparison
                            if scrape_dt.tzinfo is None:
                                scrape_dt = scrape_dt.replace(tzinfo=timezone.utc)
                            if last_modified > scrape_dt:
                                page_changed = True
                                self.logger.info(f"Update available: {doc.title} ({doc.source_url})")
                                results['changed'].append({
                                    'doc_id': doc.doc_id,
                                    'title': doc.title,
                                    'url': doc.source_url,
                                    'last_modified': last_modified.isoformat(),
                                    'scraped_date': doc.scrape_date,
                                    'reason': 'content_modified'
                                })
                                session_result['changed'] += 1

                                # Auto-rescrape if requested
                                if auto_rescrape:
                                    self.logger.info(f"Auto-rescaping: {doc.title}")
                                    try:
                                        rescrape_result = self.rescrape_document(doc.doc_id)
                                        if rescrape_result['status'] == 'success':
                                            results['rescraped'].append(doc.doc_id)
                                    except Exception as e:
                                        self.logger.error(f"Auto-rescrape failed: {e}")

                    if not page_changed:
                        # No change detected
                        results['unchanged'].append({
                            'doc_id': doc.doc_id,
                            'title': doc.title,
                            'url': doc.source_url
                        })
                        session_result['unchanged'] += 1

                except Exception as e:
                    self.logger.error(f"Failed to check {doc.source_url}: {e}")
                    results['failed'].append({
                        'doc_id': doc.doc_id,
                        'title': doc.title,
                        'url': doc.source_url,
                        'error': str(e)
                    })

            # Check for new pages
            if check_structure and discovered_urls:
                new_urls = discovered_urls - set(stored_urls.keys())
                if new_urls:
                    self.logger.info(f"Found {len(new_urls)} new pages for {session['base_url']}")
                    for new_url in new_urls:
                        results['new_pages'].append({
                            'url': new_url,
                            'base_url': session['base_url'],
                            'scrape_config': session['config']
                        })
                        session_result['new'] += 1

                # Check for missing pages (in database but not discovered)
                if discovered_urls:  # Only if discovery was successful
                    missing_urls = set(stored_urls.keys()) - discovered_urls
                    # Filter out already detected 404s
                    existing_missing = {p['url'] for p in results['missing_pages']}
                    for missing_url in missing_urls:
                        if missing_url not in existing_missing:
                            doc = stored_urls[missing_url]
                            self.logger.warning(f"Page not discovered during crawl: {missing_url}")
                            results['missing_pages'].append({
                                'doc_id': doc.doc_id,
                                'title': doc.title,
                                'url': missing_url,
                                'reason': 'not_discovered'
                            })
                            session_result['missing'] += 1

        self.logger.info(
            f"Update check complete: {len(results['unchanged'])} unchanged, "
            f"{len(results['changed'])} changed, {len(results['new_pages'])} new pages, "
            f"{len(results['missing_pages'])} missing pages, {len(results['failed'])} failed"
        )

        return results

    def remove_document(self, doc_id: str) -> bool:
        """Remove a document from the knowledge base."""
        self.logger.info(f"Removing document: {doc_id}")

        if doc_id not in self.documents:
            self.logger.warning(f"Document not found for removal: {doc_id}")
            return False

        filename = self.documents[doc_id].filename

        # Remove from database (chunks cascade automatically)
        success = self._remove_document_db(doc_id)

        if success:
            # Remove from in-memory index
            del self.documents[doc_id]

            # Invalidate BM25 index (will be rebuilt on next search)
            self.bm25 = None

            # Invalidate embeddings (will be rebuilt on next semantic search)
            if self.use_semantic:
                self.embeddings_index = None
                self.embeddings_doc_map = []

            # Invalidate search caches
            self._invalidate_caches()

            self.logger.info(f"Successfully removed document {doc_id}: {filename}")

        return success

    def needs_reindex(self, filepath: str, doc_id: str) -> bool:
        """
        Check if a document needs re-indexing based on file modification time and content hash.

        Args:
            filepath: Path to the document file
            doc_id: Document ID to check

        Returns:
            True if the document needs re-indexing, False otherwise
        """
        doc = self.documents.get(doc_id)
        if not doc:
            return True  # Document doesn't exist, needs indexing

        # If no mtime/hash stored, can't check - assume needs reindex
        if doc.file_mtime is None or doc.file_hash is None:
            self.logger.info(f"Document {doc_id} has no update detection data, assuming needs reindex")
            return True

        # Quick check: modification time
        try:
            current_mtime = os.path.getmtime(filepath)
            if current_mtime <= doc.file_mtime:
                # File hasn't been modified since last index
                return False
        except OSError:
            # File doesn't exist or can't be accessed
            self.logger.warning(f"Cannot access file: {filepath}")
            return False

        # File was modified - do deep check with content hash
        try:
            current_hash = self._compute_file_hash(filepath)
            if current_hash == doc.file_hash:
                # Content is same despite mtime change (e.g., touched)
                self.logger.info(f"File mtime changed but content unchanged: {filepath}")
                return False
            else:
                # Content has actually changed
                self.logger.info(f"File content changed: {filepath}")
                return True
        except Exception as e:
            self.logger.error(f"Error computing hash for {filepath}: {e}")
            return False

    def update_document(self, filepath: str, title: Optional[str] = None, tags: Optional[list[str]] = None) -> DocumentMeta:
        """
        Update an existing document if it has changed, or add it if it doesn't exist.

        Args:
            filepath: Path to the document file
            title: Optional title (if not provided, uses filename)
            tags: Optional list of tags

        Returns:
            DocumentMeta for the document (existing or newly indexed)
        """
        # Find existing doc by filepath
        existing_doc = None
        for doc in self.documents.values():
            if doc.filepath == filepath:
                existing_doc = doc
                break

        if not existing_doc:
            # Document doesn't exist, add it
            self.logger.info(f"Document not found, adding: {filepath}")
            return self.add_document(filepath, title, tags)

        if not self.needs_reindex(filepath, existing_doc.doc_id):
            # Document unchanged
            self.logger.info(f"Document unchanged, skipping reindex: {filepath}")
            return existing_doc

        # Document has changed, re-index it
        self.logger.info(f"Document changed, re-indexing: {filepath}")
        self.remove_document(existing_doc.doc_id)
        return self.add_document(filepath, title, tags)

    def update_document_title(self, doc_id: str, title: str) -> None:
        """
        Update the title of a document.

        Args:
            doc_id: Document ID
            title: New title for the document

        Raises:
            ValueError: If document not found
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        # Update in-memory
        self.documents[doc_id].title = title

        # Update in database
        cursor = self.db_conn.cursor()
        cursor.execute(
            "UPDATE documents SET title = ? WHERE doc_id = ?",
            (title, doc_id)
        )
        self.db_conn.commit()

        self.logger.info(f"Updated title for document {doc_id[:12]}: {title}")

    def update_document_tags(self, doc_id: str, tags: list[str]) -> None:
        """
        Update the tags for a document.

        Args:
            doc_id: Document ID
            tags: New list of tags (replaces existing tags)

        Raises:
            ValueError: If document not found
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        # Update in-memory
        self.documents[doc_id].tags = tags

        # Update in database
        cursor = self.db_conn.cursor()
        tags_str = ','.join(tags) if tags else ''
        cursor.execute(
            "UPDATE documents SET tags = ? WHERE doc_id = ?",
            (tags_str, doc_id)
        )
        self.db_conn.commit()

        self.logger.info(f"Updated tags for document {doc_id[:12]}: {tags}")

    def suggest_tags(self, doc_id: str, confidence_threshold: float = 0.6) -> list[dict]:
        """
        Suggest tags for a document based on content analysis.

        Uses heuristic-based tag suggestion (no LLM required):
        - Detects hardware components (SID, VIC-II, CIA, 6502)
        - Identifies programming topics (assembly, BASIC, machine code)
        - Recognizes document types (reference, tutorial, guide)
        - Extracts memory addresses and registers

        Args:
            doc_id: Document ID to analyze
            confidence_threshold: Minimum confidence for suggestions (0.0-1.0)

        Returns:
            List of suggested tags with confidence scores:
            [
                {'tag': 'sid-programming', 'confidence': 0.95, 'category': 'hardware'},
                {'tag': 'assembly', 'confidence': 0.85, 'category': 'programming'},
                ...
            ]

        Raises:
            ValueError: If document not found
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        suggested_tags = []

        # Get sample of document content (first 3 chunks to avoid overhead)
        try:
            chunks = self._get_chunks_db(doc_id)
            sample_text = '\n'.join([c['content'] for c in chunks[:3]]) if chunks else ''
        except Exception as e:
            self.logger.warning(f"Failed to get chunks for {doc_id}: {e}")
            sample_text = ''

        if not sample_text:
            return suggested_tags

        text_lower = sample_text.lower()

        # Hardware detection
        hardware_patterns = {
            'sid-chip': (r'\bsid\b|\b6581\b', 0.9),
            'vic-ii': (r'\bvic-?ii\b|\bvic\s*2\b|\b6569\b|\b6567\b', 0.9),
            'cia': (r'\bcia\b|\b6526\b', 0.85),
            '6502-processor': (r'\b6502\b|\b6510\b', 0.9),
            'joystick': (r'\bjoystick\b|\bcontroller\b', 0.7),
            'disk-drive': (r'\bdisk\s*drive\b|\b1541\b|\b1571\b', 0.8),
        }

        for tag, (pattern, confidence) in hardware_patterns.items():
            if re.search(pattern, text_lower, re.IGNORECASE):
                if confidence >= confidence_threshold:
                    suggested_tags.append({
                        'tag': tag,
                        'confidence': confidence,
                        'category': 'hardware'
                    })

        # Programming topic detection
        programming_patterns = {
            'assembly': (r'\bassembly\b|\bmachine\s*code\b|\basync\b', 0.85),
            'basic': (r'\bbasic\b|\bprogram\s*line\b|\bline\s*numbers\b', 0.8),
            'graphics': (r'\bgraphics\b|\bsprite\b|\bbitmap\b|\bcharacter\s*set\b', 0.85),
            'sound-music': (r'\bsound\b|\bmusic\b|\baudio\b|\benvelop\b|\bsynthesis\b', 0.8),
            'interrupts': (r'\binterrupt\b|\birq\b|\bnmi\b', 0.9),
            'memory-management': (r'\bmemory\s*map\b|\bmemory\s*address\b|\b\$[0-9a-f]{4}\b', 0.75),
        }

        for tag, (pattern, confidence) in programming_patterns.items():
            if re.search(pattern, text_lower, re.IGNORECASE):
                if confidence >= confidence_threshold:
                    suggested_tags.append({
                        'tag': tag,
                        'confidence': confidence,
                        'category': 'programming'
                    })

        # Document type detection
        doc_type_patterns = {
            'reference': (r'\breference\b|\bopcode\s*table\b|\binstruction\s*set\b', 0.85),
            'tutorial': (r'\btutorial\b|\bhow\s*to\b|\bguide\b|\blearn\b', 0.75),
            'specification': (r'\bspecification\b|\bspec\b|\bdatasheet\b|\bmanual\b', 0.9),
            'code-example': (r'\bexample\b|\bcode\s*sample\b|\broutine\b', 0.7),
        }

        for tag, (pattern, confidence) in doc_type_patterns.items():
            if re.search(pattern, text_lower, re.IGNORECASE):
                if confidence >= confidence_threshold:
                    suggested_tags.append({
                        'tag': tag,
                        'confidence': confidence,
                        'category': 'document-type'
                    })

        # Difficulty level detection
        difficulty_patterns = {
            'beginner': (r'\bbeginning\b|\bstarter\b|\bintroduction\b|\bfundamentals?\b', 0.75),
            'intermediate': (r'\bintermediate\b|\badvanced-beginner\b', 0.7),
            'advanced': (r'\badvanced\b|\bexpert\b|\bdeep-dive\b', 0.75),
        }

        for tag, (pattern, confidence) in difficulty_patterns.items():
            if re.search(pattern, text_lower, re.IGNORECASE):
                if confidence >= confidence_threshold:
                    suggested_tags.append({
                        'tag': tag,
                        'confidence': confidence,
                        'category': 'difficulty'
                    })

        # Sort by confidence (descending)
        suggested_tags.sort(key=lambda x: x['confidence'], reverse=True)

        self.logger.info(f"Suggested {len(suggested_tags)} tags for document {doc_id[:12]}")

        return suggested_tags

    def add_tags_to_document(self, doc_id: str, new_tags: list[str],
                            merge: bool = True) -> list[str]:
        """
        Add tags to a document, optionally merging with existing tags.

        Args:
            doc_id: Document ID
            new_tags: Tags to add
            merge: If True, merge with existing tags. If False, replace tags.

        Returns:
            Updated list of all tags for the document

        Raises:
            ValueError: If document not found
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        # Normalize and deduplicate tags
        new_tags = list(set([t.lower().replace(' ', '-') for t in new_tags if t]))

        if merge:
            # Merge with existing tags
            existing_tags = set(self.documents[doc_id].tags or [])
            all_tags = list(existing_tags | set(new_tags))
        else:
            all_tags = new_tags

        # Update document
        self.update_document_tags(doc_id, all_tags)

        return all_tags

    def get_tags_by_category(self) -> dict[str, list[dict]]:
        """
        Get all tags organized by category for easier browsing.

        Returns:
            Dictionary with categories as keys and tag lists as values:
            {
                'hardware': [
                    {'tag': 'sid-chip', 'count': 15, 'documents': ['doc1', 'doc2', ...]},
                    ...
                ],
                'programming': [...],
                ...
            }
        """
        # Categorize known tags
        tag_categories = {
            'hardware': ['sid-chip', 'vic-ii', 'cia', '6502-processor', 'joystick', 'disk-drive'],
            'programming': ['assembly', 'basic', 'graphics', 'sound-music', 'interrupts', 'memory-management'],
            'document-type': ['reference', 'tutorial', 'specification', 'code-example'],
            'difficulty': ['beginner', 'intermediate', 'advanced'],
        }

        result = {}

        # For each category, count tag usage
        for category, known_tags in tag_categories.items():
            result[category] = []

            for tag in known_tags:
                # Count documents with this tag
                doc_ids = [doc_id for doc_id, doc in self.documents.items()
                          if tag in (doc.tags or [])]

                if doc_ids:  # Only include tags that are actually used
                    result[category].append({
                        'tag': tag,
                        'count': len(doc_ids),
                        'documents': doc_ids[:10]  # Show first 10 docs
                    })

            # Sort by count (descending)
            result[category].sort(key=lambda x: x['count'], reverse=True)

        # Add custom/user-defined tags that don't fit in categories
        all_known_tags = set()
        for tags_list in tag_categories.values():
            all_known_tags.update(tags_list)

        custom_tags = {}
        for doc in self.documents.values():
            for tag in (doc.tags or []):
                if tag not in all_known_tags:
                    if tag not in custom_tags:
                        custom_tags[tag] = []
                    custom_tags[tag].append(doc.doc_id)

        if custom_tags:
            result['custom'] = [
                {
                    'tag': tag,
                    'count': len(doc_ids),
                    'documents': doc_ids[:10]
                }
                for tag, doc_ids in sorted(custom_tags.items(),
                                         key=lambda x: len(x[1]),
                                         reverse=True)
            ]

        return result

    def summarize_document(self, doc_id: str,
                          max_length: int = 500,
                          style: str = "technical") -> str:
        """
        Generate an AI summary of a document.

        Args:
            doc_id: Document ID to summarize
            max_length: Maximum summary length in words
            style: Summary style (technical, simple, or detailed)

        Returns:
            Summary text

        Raises:
            ValueError: If document not found or LLM not available
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        # Check if LLM client is available
        if not hasattr(self, 'llm_client') or self.llm_client is None:
            # Try to initialize it
            try:
                from llm_integration import LLMClient
                self.llm_client = LLMClient()
            except Exception as e:
                raise ValueError(f"LLM client not available: {e}")

        doc = self.documents[doc_id]

        # Get document content (first 10 chunks to keep context reasonable)
        chunks = self._get_chunks_db(doc_id)
        content_chunks = chunks[:10] if len(chunks) > 10 else chunks
        content = '\n\n'.join([chunk.content for chunk in content_chunks])

        # Truncate content if too long (max 20000 chars)
        if len(content) > 20000:
            content = content[:20000] + "..."

        # Build prompt based on style
        style_prompts = {
            "technical": "Provide a concise technical summary focusing on key concepts, technologies, and implementation details.",
            "simple": "Provide a simple, easy-to-understand summary suitable for beginners.",
            "detailed": "Provide a comprehensive detailed summary covering all major topics and subtopics."
        }

        style_instruction = style_prompts.get(style, style_prompts["technical"])

        prompt = f"""Summarize the following document in approximately {max_length} words.

Document Title: {doc.title}

{style_instruction}

Document Content:
{content}

Summary:"""

        try:
            summary = self.llm_client.call(prompt, max_tokens=max_length * 2, temperature=0.3)
            return summary.strip()
        except Exception as e:
            self.logger.error(f"Failed to generate summary: {e}")
            raise ValueError(f"Failed to generate summary: {e}")

    def check_all_updates(self, auto_update: bool = False) -> dict:
        """
        Check all indexed documents for updates.

        Args:
            auto_update: If True, automatically re-index changed documents

        Returns:
            Dictionary with lists of unchanged, changed, and missing documents
        """
        results = {
            'unchanged': [],
            'changed': [],
            'missing': [],
            'updated': []  # Only populated if auto_update=True
        }

        for doc_id, doc in list(self.documents.items()):
            filepath = doc.filepath

            # Check if file still exists
            if not os.path.exists(filepath):
                results['missing'].append({
                    'doc_id': doc_id,
                    'filepath': filepath,
                    'title': doc.title
                })
                continue

            # Check if needs reindex
            if self.needs_reindex(filepath, doc_id):
                results['changed'].append({
                    'doc_id': doc_id,
                    'filepath': filepath,
                    'title': doc.title
                })

                if auto_update:
                    try:
                        updated_doc = self.update_document(filepath, doc.title, doc.tags)
                        results['updated'].append({
                            'doc_id': updated_doc.doc_id,
                            'filepath': filepath,
                            'title': updated_doc.title
                        })
                    except Exception as e:
                        self.logger.error(f"Failed to update {filepath}: {e}")
            else:
                results['unchanged'].append({
                    'doc_id': doc_id,
                    'filepath': filepath,
                    'title': doc.title
                })

        return results

    def add_documents_bulk(self, directory: str, pattern: str = "**/*.{pdf,txt,md,html,htm,xlsx,xls}",
                           tags: Optional[list[str]] = None, recursive: bool = True,
                           skip_duplicates: bool = True, progress_callback: ProgressCallback = None) -> dict:
        """
        Add multiple documents from a directory matching a glob pattern.

        Args:
            directory: Directory to search for documents
            pattern: Glob pattern (default: **/*.{pdf,txt})
            tags: Tags to apply to all documents
            recursive: Search subdirectories (default: True)
            skip_duplicates: Skip files with duplicate content (default: True)
            progress_callback: Optional callback for progress updates

        Returns:
            Dictionary with lists of added, skipped, and failed documents
        """
        from pathlib import Path

        dir_path = Path(directory).resolve()
        if not dir_path.exists():
            raise ValueError(f"Directory does not exist: {directory}")

        # Find matching files
        if recursive:
            files = list(dir_path.glob(pattern))
        else:
            # Non-recursive: remove ** from pattern
            non_recursive_pattern = pattern.replace('**/', '')
            files = list(dir_path.glob(non_recursive_pattern))

        results = {
            'added': [],
            'skipped': [],
            'failed': []
        }

        self.logger.info(f"Bulk add: found {len(files)} files matching pattern '{pattern}' in {directory}")

        # Report progress: Start
        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="add_documents_bulk",
                current=0,
                total=len(files),
                message=f"Starting bulk add of {len(files)} files",
                item=directory
            ))

        # Get worker count (configurable via environment variable, default to CPU count)
        max_workers = int(os.getenv('PARALLEL_WORKERS', str(os.cpu_count() or 4)))
        self.logger.info(f"Using {max_workers} workers for parallel processing")

        # Process files in parallel using ThreadPoolExecutor
        def process_file(file_path):
            """Process a single file and return result."""
            if not file_path.is_file():
                return None

            try:
                # Generate title from filename
                title = file_path.stem
                doc = self.add_document(str(file_path), title=title, tags=tags)

                return {
                    'status': 'added',
                    'doc_id': doc.doc_id,
                    'filepath': str(file_path),
                    'title': title,
                    'chunks': doc.total_chunks
                }

            except Exception as e:
                return {
                    'status': 'failed',
                    'filepath': str(file_path),
                    'error': str(e)
                }

        # Use ThreadPoolExecutor for parallel processing
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit all files for processing
            future_to_file = {executor.submit(process_file, fp): fp for fp in files}

            # Process completed tasks as they finish
            completed = 0
            seen_doc_ids = set()

            for future in as_completed(future_to_file):
                completed += 1
                file_path = future_to_file[future]

                # Report progress: Processing file
                if progress_callback:
                    progress_callback(ProgressUpdate(
                        operation="add_documents_bulk",
                        current=completed,
                        total=len(files),
                        message=f"Processing file {completed}/{len(files)}",
                        item=str(file_path.name)
                    ))

                try:
                    result = future.result()
                    if result is None:
                        continue

                    if result['status'] == 'added':
                        # Check for duplicates
                        if skip_duplicates and result['doc_id'] in seen_doc_ids:
                            results['skipped'].append({
                                'filepath': result['filepath'],
                                'reason': 'duplicate content',
                                'doc_id': result['doc_id']
                            })
                        else:
                            seen_doc_ids.add(result['doc_id'])
                            results['added'].append({
                                'doc_id': result['doc_id'],
                                'filepath': result['filepath'],
                                'title': result['title'],
                                'chunks': result['chunks']
                            })
                    elif result['status'] == 'failed':
                        results['failed'].append({
                            'filepath': result['filepath'],
                            'error': result['error']
                        })
                        self.logger.error(f"Failed to add {result['filepath']}: {result['error']}")

                except Exception as e:
                    results['failed'].append({
                        'filepath': str(file_path),
                        'error': str(e)
                    })
                    self.logger.error(f"Failed to process {file_path}: {e}")

        self.logger.info(f"Bulk add complete: {len(results['added'])} added, "
                        f"{len(results['skipped'])} skipped, {len(results['failed'])} failed")

        # Report progress: Complete
        if progress_callback:
            progress_callback(ProgressUpdate(
                operation="add_documents_bulk",
                current=len(files),
                total=len(files),
                message=f"Bulk add complete: {len(results['added'])} added, "
                        f"{len(results['skipped'])} skipped, {len(results['failed'])} failed"
            ))

        return results

    def remove_documents_bulk(self, doc_ids: Optional[list[str]] = None,
                              tags: Optional[list[str]] = None) -> dict:
        """
        Remove multiple documents by doc IDs or tags.

        Args:
            doc_ids: List of document IDs to remove
            tags: Remove all documents with any of these tags

        Returns:
            Dictionary with lists of removed and failed document IDs
        """
        if not doc_ids and not tags:
            raise ValueError("Must provide either doc_ids or tags")

        results = {
            'removed': [],
            'failed': []
        }

        # Collect doc_ids to remove
        ids_to_remove = set()

        if doc_ids:
            ids_to_remove.update(doc_ids)

        if tags:
            # Find all documents with any of the specified tags
            for doc_id, doc in self.documents.items():
                if any(tag in doc.tags for tag in tags):
                    ids_to_remove.add(doc_id)

        self.logger.info(f"Bulk remove: removing {len(ids_to_remove)} documents")

        for doc_id in ids_to_remove:
            try:
                if self.remove_document(doc_id):
                    results['removed'].append(doc_id)
                else:
                    results['failed'].append({
                        'doc_id': doc_id,
                        'error': 'Document not found'
                    })
            except Exception as e:
                results['failed'].append({
                    'doc_id': doc_id,
                    'error': str(e)
                })
                self.logger.error(f"Failed to remove {doc_id}: {e}")

        self.logger.info(f"Bulk remove complete: {len(results['removed'])} removed, "
                        f"{len(results['failed'])} failed")

        return results

    def update_tags_bulk(self, doc_ids: Optional[list[str]] = None,
                         existing_tags: Optional[list[str]] = None,
                         add_tags: Optional[list[str]] = None,
                         remove_tags: Optional[list[str]] = None,
                         replace_tags: Optional[list[str]] = None) -> dict:
        """
        Update tags for multiple documents in bulk.

        Args:
            doc_ids: List of document IDs to update (if None, uses existing_tags to find docs)
            existing_tags: Find documents with any of these tags (alternative to doc_ids)
            add_tags: Tags to add to the documents
            remove_tags: Tags to remove from the documents
            replace_tags: Replace all tags with these tags

        Returns:
            Dictionary with lists of updated and failed document IDs

        Examples:
            # Add 'assembly' tag to specific documents
            kb.update_tags_bulk(doc_ids=['doc1', 'doc2'], add_tags=['assembly'])

            # Remove 'draft' tag from all documents that have it
            kb.update_tags_bulk(existing_tags=['draft'], remove_tags=['draft'])

            # Replace all tags with 'archive' for specific documents
            kb.update_tags_bulk(doc_ids=['doc1', 'doc2'], replace_tags=['archive'])

            # Add 'reviewed' and remove 'draft' for documents with 'pending' tag
            kb.update_tags_bulk(existing_tags=['pending'], add_tags=['reviewed'], remove_tags=['draft'])
        """
        if not doc_ids and not existing_tags:
            raise ValueError("Must provide either doc_ids or existing_tags")

        if not add_tags and not remove_tags and not replace_tags:
            raise ValueError("Must provide at least one of: add_tags, remove_tags, replace_tags")

        results = {
            'updated': [],
            'failed': []
        }

        # Collect doc_ids to update
        ids_to_update = set()

        if doc_ids:
            ids_to_update.update(doc_ids)

        if existing_tags:
            # Find all documents with any of the specified tags
            for doc_id, doc in self.documents.items():
                if any(tag in doc.tags for tag in existing_tags):
                    ids_to_update.add(doc_id)

        self.logger.info(f"Bulk tag update: updating {len(ids_to_update)} documents")

        for doc_id in ids_to_update:
            try:
                if doc_id not in self.documents:
                    results['failed'].append({
                        'doc_id': doc_id,
                        'error': 'Document not found'
                    })
                    continue

                doc = self.documents[doc_id]
                old_tags = doc.tags.copy()

                # Apply tag operations
                if replace_tags is not None:
                    doc.tags = replace_tags.copy()
                else:
                    if add_tags:
                        # Add tags (avoiding duplicates)
                        for tag in add_tags:
                            if tag not in doc.tags:
                                doc.tags.append(tag)

                    if remove_tags:
                        # Remove tags
                        doc.tags = [tag for tag in doc.tags if tag not in remove_tags]

                # Update in database
                cursor = self.db_conn.cursor()
                cursor.execute("""
                    UPDATE documents
                    SET tags = ?
                    WHERE doc_id = ?
                """, (json.dumps(doc.tags), doc_id))
                self.db_conn.commit()

                results['updated'].append({
                    'doc_id': doc_id,
                    'old_tags': old_tags,
                    'new_tags': doc.tags
                })

                self.logger.debug(f"Updated tags for {doc_id}: {old_tags} -> {doc.tags}")

            except Exception as e:
                results['failed'].append({
                    'doc_id': doc_id,
                    'error': str(e)
                })
                self.logger.error(f"Failed to update tags for {doc_id}: {e}")

        self.logger.info(f"Bulk tag update complete: {len(results['updated'])} updated, "
                        f"{len(results['failed'])} failed")

        return results

    def auto_tag_document(self, doc_id: str, confidence_threshold: float = 0.7,
                         max_tags: int = 10, append: bool = True) -> dict:
        """
        Generate tags automatically using LLM analysis.

        Args:
            doc_id: Document to tag
            confidence_threshold: Minimum confidence to accept tag (0.0-1.0)
            max_tags: Maximum number of tags to suggest
            append: If True, append to existing tags; if False, replace

        Returns:
            {
                'doc_id': str,
                'suggested_tags': [{'tag': str, 'confidence': float}, ...],
                'applied_tags': [str],
                'skipped_tags': [str],  # Below confidence threshold
                'existing_tags': [str],
                'new_tags': [str]  # Final tag list
            }

        Example:
            result = kb.auto_tag_document('doc123', confidence_threshold=0.7)
            # {
            #     'suggested_tags': [
            #         {'tag': 'sid-programming', 'confidence': 0.95},
            #         {'tag': 'assembly', 'confidence': 0.88},
            #         {'tag': 'beginner', 'confidence': 0.65}  # Below threshold
            #     ],
            #     'applied_tags': ['sid-programming', 'assembly'],
            #     'skipped_tags': ['beginner'],
            #     ...
            # }
        """
        # Import LLM client
        try:
            from llm_integration import get_llm_client
        except ImportError:
            raise ImportError("llm_integration module not found. Auto-tagging requires LLM integration.")

        # Get LLM client
        llm_client = get_llm_client()
        if not llm_client:
            raise ValueError("LLM not configured. Set LLM_PROVIDER and appropriate API key.")

        # Get document
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        doc = self.documents[doc_id]

        # Get sample text (first 3 chunks for analysis)
        chunks = self._get_chunks_db(doc_id)
        sample_chunks = chunks[:3] if len(chunks) > 3 else chunks
        sample_text = "\n\n".join([c.content for c in sample_chunks])

        # Limit text size (first 3000 chars)
        if len(sample_text) > 3000:
            sample_text = sample_text[:3000] + "..."

        # Build prompt
        prompt = f"""Analyze this Commodore 64 technical documentation and suggest relevant tags.

Consider these categories:
1. Hardware components (sid, vic-ii, cia, 6502, memory, cartridge, disk-drive, etc.)
2. Programming topics (assembly, basic, machine-code, graphics, sound, sprites, etc.)
3. Document type (tutorial, reference, manual, guide, example, etc.)
4. Difficulty level (beginner, intermediate, advanced, expert)
5. Content area (programming, hardware, history, repair, modification, etc.)

Document title: {doc.title}
Document filename: {doc.filename}

Sample text:
{sample_text}

Return a JSON object with this structure:
{{
    "tags": [
        {{"tag": "sid-programming", "confidence": 0.95, "reason": "Document extensively discusses SID chip programming"}},
        {{"tag": "assembly", "confidence": 0.88, "reason": "Contains assembly code examples"}}
    ]
}}

Important:
- Use lowercase with hyphens (e.g., "sid-programming" not "SID Programming")
- Provide {max_tags} or fewer tags
- Include confidence score (0.0-1.0) for each tag
- Brief reason for each tag suggestion
- Return ONLY the JSON, no other text"""

        # Call LLM
        self.logger.info(f"Auto-tagging document {doc_id} ({doc.title})")

        try:
            response = llm_client.call_json(prompt, max_tokens=1024, temperature=0.3)
        except Exception as e:
            self.logger.error(f"LLM call failed: {e}")
            raise ValueError(f"Failed to generate tags: {e}")

        # Parse response
        suggested_tags = response.get('tags', [])

        # Filter by confidence
        high_confidence_tags = [
            t for t in suggested_tags
            if t['confidence'] >= confidence_threshold
        ]

        low_confidence_tags = [
            t for t in suggested_tags
            if t['confidence'] < confidence_threshold
        ]

        # Extract tag names
        applied_tag_names = [t['tag'] for t in high_confidence_tags]
        skipped_tag_names = [t['tag'] for t in low_confidence_tags]

        # Get existing tags
        existing_tags = doc.tags.copy()

        # Apply tags
        if append:
            # Add new tags to existing (avoid duplicates)
            new_tags = existing_tags.copy()
            for tag in applied_tag_names:
                if tag not in new_tags:
                    new_tags.append(tag)
        else:
            # Replace all tags
            new_tags = applied_tag_names

        # Update document
        doc.tags = new_tags

        # Update in database
        cursor = self.db_conn.cursor()
        cursor.execute("""
            UPDATE documents
            SET tags = ?
            WHERE doc_id = ?
        """, (json.dumps(new_tags), doc_id))
        self.db_conn.commit()

        result = {
            'doc_id': doc_id,
            'doc_title': doc.title,
            'suggested_tags': suggested_tags,
            'applied_tags': applied_tag_names,
            'skipped_tags': skipped_tag_names,
            'existing_tags': existing_tags,
            'new_tags': new_tags,
            'confidence_threshold': confidence_threshold
        }

        self.logger.info(f"Auto-tagged {doc_id}: applied {len(applied_tag_names)} tags, "
                        f"skipped {len(skipped_tag_names)} low-confidence tags")

        return result

    def auto_tag_all_documents(self, confidence_threshold: float = 0.7,
                               max_tags: int = 10, append: bool = True,
                               skip_tagged: bool = True, max_docs: Optional[int] = None) -> dict:
        """
        Bulk auto-tag all documents using LLM.

        Args:
            confidence_threshold: Minimum confidence to accept tag (0.0-1.0)
            max_tags: Maximum tags per document
            append: If True, append to existing tags; if False, replace
            skip_tagged: If True, skip documents that already have tags
            max_docs: Maximum number of documents to process (None = all)

        Returns:
            {
                'processed': int,
                'skipped': int,
                'failed': int,
                'total_tags_added': int,
                'results': [list of individual results]
            }

        Example:
            results = kb.auto_tag_all_documents(
                confidence_threshold=0.7,
                skip_tagged=True,
                max_docs=10
            )
        """
        results = {
            'processed': 0,
            'skipped': 0,
            'failed': 0,
            'total_tags_added': 0,
            'results': []
        }

        # Get documents to process
        docs_to_process = []

        for doc_id, doc in self.documents.items():
            # Skip if already has tags (optional)
            if skip_tagged and doc.tags:
                results['skipped'] += 1
                continue

            docs_to_process.append(doc_id)

            # Limit number of documents
            if max_docs and len(docs_to_process) >= max_docs:
                break

        self.logger.info(f"Auto-tagging {len(docs_to_process)} documents "
                        f"(skipped {results['skipped']} already tagged)")

        # Process each document
        for i, doc_id in enumerate(docs_to_process, 1):
            try:
                self.logger.info(f"Auto-tagging {i}/{len(docs_to_process)}: {doc_id}")

                result = self.auto_tag_document(
                    doc_id,
                    confidence_threshold=confidence_threshold,
                    max_tags=max_tags,
                    append=append
                )

                # Count new tags added
                tags_added = len(set(result['new_tags']) - set(result['existing_tags']))
                results['total_tags_added'] += tags_added

                results['processed'] += 1
                results['results'].append(result)

            except Exception as e:
                results['failed'] += 1
                self.logger.error(f"Failed to auto-tag {doc_id}: {e}")
                results['results'].append({
                    'doc_id': doc_id,
                    'error': str(e)
                })

        self.logger.info(f"Auto-tagging complete: processed={results['processed']}, "
                        f"failed={results['failed']}, tags_added={results['total_tags_added']}")

        return results

    def generate_summary(self, doc_id: str, summary_type: str = 'brief',
                        force_regenerate: bool = False) -> str:
        """
        Generate an AI-powered summary of a document.

        Args:
            doc_id: Document ID to summarize
            summary_type: Type of summary ('brief', 'detailed', 'bullet')
                - 'brief': 1-2 paragraph overview (200-300 words)
                - 'detailed': Comprehensive summary with key points (500-800 words)
                - 'bullet': Bullet-point summary of main topics
            force_regenerate: If True, regenerate even if cached summary exists

        Returns:
            Summary text as a string

        Raises:
            ValueError: If document not found or LLM not configured
            DocumentNotFoundError: If document doesn't exist

        Examples:
            # Generate brief summary
            summary = kb.generate_summary('doc123', 'brief')

            # Get detailed summary
            summary = kb.generate_summary('doc456', 'detailed')

            # Force regeneration (bypass cache)
            summary = kb.generate_summary('doc789', 'brief', force_regenerate=True)
        """
        # Validate document exists
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        doc = self.documents[doc_id]

        # Check for cached summary
        if not force_regenerate:
            cursor = self.db_conn.cursor()
            cursor.execute("""
                SELECT summary_text FROM document_summaries
                WHERE doc_id = ? AND summary_type = ?
            """, (doc_id, summary_type))
            result = cursor.fetchone()
            if result:
                self.logger.debug(f"Using cached summary for {doc_id} ({summary_type})")
                return result[0]

        # Import LLM client
        try:
            from llm_integration import get_llm_client
        except ImportError:
            raise ImportError("llm_integration module not found. Summarization requires LLM integration.")

        # Get LLM client
        llm_client = get_llm_client()
        if not llm_client:
            raise ValueError("LLM not configured. Set LLM_PROVIDER and appropriate API key.")

        # Get document content
        chunks = self._get_chunks_db(doc_id)
        if not chunks:
            raise ValueError(f"No content found for document: {doc_id}")

        # For brief summaries, use first 5 chunks; for detailed, use more
        if summary_type == 'brief':
            sample_chunks = chunks[:5]
            word_limit = 300
            length_guidance = "1-2 paragraphs, approximately 200-300 words"
        elif summary_type == 'detailed':
            sample_chunks = chunks[:15] if len(chunks) > 15 else chunks
            word_limit = 800
            length_guidance = "3-5 paragraphs with detailed explanations, approximately 500-800 words"
        elif summary_type == 'bullet':
            sample_chunks = chunks[:10]
            word_limit = 400
            length_guidance = "8-12 bullet points covering main topics"
        else:
            raise ValueError(f"Invalid summary type: {summary_type}. Must be 'brief', 'detailed', or 'bullet'.")

        # Join content
        content = "\n\n".join([c.content for c in sample_chunks])

        # Limit content size to first 10k chars to control API costs
        if len(content) > 10000:
            content = content[:10000] + "..."

        # Build prompt based on summary type
        if summary_type == 'bullet':
            prompt = f"""Create a bullet-point summary of this Commodore 64 technical documentation.

Document Title: {doc.title}
Document Type: {doc.file_type}

Content:
{content}

Create a concise bullet-point summary with 8-12 main topics. Each bullet should be clear and informative.
Return ONLY the bullet points, one per line, starting with "- ". No introduction or explanation needed."""

        else:
            prompt = f"""Create a {summary_type} summary of this Commodore 64 technical documentation.

Document Title: {doc.title}
Document Type: {doc.file_type}

Content:
{content}

Write a {summary_type} summary that is {length_guidance}.
Focus on:
- Key concepts and main topics
- Technical details relevant to programmers
- Important procedures or examples
- Practical applications

Return ONLY the summary text, no preamble."""

        # Call LLM
        self.logger.info(f"Generating {summary_type} summary for {doc_id} ({doc.title})")

        try:
            summary_text = llm_client.call(prompt, max_tokens=word_limit + 200, temperature=0.4)
        except Exception as e:
            self.logger.error(f"LLM call failed: {e}")
            raise ValueError(f"Failed to generate summary: {e}")

        # Clean up summary text
        if not summary_text or not summary_text.strip():
            raise ValueError("LLM returned empty summary")

        summary_text = summary_text.strip()

        # Store summary in database
        cursor = self.db_conn.cursor()
        try:
            # Get model name from LLM client
            model = os.getenv('LLM_MODEL', 'unknown')

            cursor.execute("""
                INSERT OR REPLACE INTO document_summaries
                (doc_id, summary_type, summary_text, generated_at, model, token_count)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (doc_id, summary_type, summary_text, datetime.now().isoformat(),
                  model, len(summary_text.split())))

            self.db_conn.commit()
            self.logger.info(f"Saved {summary_type} summary for {doc_id}")

        except Exception as e:
            self.logger.error(f"Failed to save summary to database: {e}")
            # Return summary even if save failed
            pass

        return summary_text

    def generate_summary_all(self, summary_types: Optional[list[str]] = None,
                            force_regenerate: bool = False,
                            max_docs: Optional[int] = None) -> dict:
        """
        Bulk generate summaries for all documents.

        Args:
            summary_types: List of summary types to generate (['brief'], ['brief', 'detailed'], etc.)
                          Default: ['brief']
            force_regenerate: If True, regenerate all summaries
            max_docs: Maximum number of documents to process (None = all)

        Returns:
            {
                'processed': int,
                'failed': int,
                'total_summaries': int,
                'by_type': {'brief': int, 'detailed': int, 'bullet': int},
                'results': [list of individual results]
            }

        Example:
            results = kb.generate_summary_all(
                summary_types=['brief', 'detailed'],
                max_docs=50
            )
        """
        if summary_types is None:
            summary_types = ['brief']

        results = {
            'processed': 0,
            'failed': 0,
            'total_summaries': 0,
            'by_type': {st: 0 for st in summary_types},
            'results': []
        }

        # Get documents to process
        docs_to_process = list(self.documents.keys())

        if max_docs:
            docs_to_process = docs_to_process[:max_docs]

        self.logger.info(f"Generating summaries for {len(docs_to_process)} documents "
                        f"(types: {', '.join(summary_types)})")

        # Process each document
        for i, doc_id in enumerate(docs_to_process, 1):
            doc_results = {
                'doc_id': doc_id,
                'title': self.documents[doc_id].title,
                'summaries': {}
            }

            for summary_type in summary_types:
                try:
                    self.logger.info(f"[{i}/{len(docs_to_process)}] {doc_id} ({summary_type})")

                    summary = self.generate_summary(
                        doc_id,
                        summary_type=summary_type,
                        force_regenerate=force_regenerate
                    )

                    doc_results['summaries'][summary_type] = {
                        'success': True,
                        'length': len(summary),
                        'word_count': len(summary.split())
                    }

                    results['total_summaries'] += 1
                    results['by_type'][summary_type] += 1

                except Exception as e:
                    results['failed'] += 1
                    self.logger.error(f"Failed to summarize {doc_id} ({summary_type}): {e}")
                    doc_results['summaries'][summary_type] = {
                        'success': False,
                        'error': str(e)
                    }

            results['processed'] += 1
            results['results'].append(doc_results)

        self.logger.info(f"Summary generation complete: processed={results['processed']}, "
                        f"failed={results['failed']}, total_summaries={results['total_summaries']}")

        return results

    def get_summary(self, doc_id: str, summary_type: str = 'brief') -> Optional[str]:
        """
        Retrieve a cached summary without regenerating.

        Args:
            doc_id: Document ID
            summary_type: Type of summary ('brief', 'detailed', 'bullet')

        Returns:
            Summary text if it exists, None otherwise
        """
        cursor = self.db_conn.cursor()
        cursor.execute("""
            SELECT summary_text FROM document_summaries
            WHERE doc_id = ? AND summary_type = ?
        """, (doc_id, summary_type))
        result = cursor.fetchone()
        return result[0] if result else None

    def _normalize_entity_text(self, text: str, entity_type: str) -> str:
        """
        Normalize entity text for consistent matching and deduplication.

        Args:
            text: Raw entity text
            entity_type: Entity type (hardware, memory_address, etc.)

        Returns:
            Normalized entity text
        """
        normalized = text.strip()

        if entity_type == 'hardware':
            # Normalize hardware names
            # VIC II, VIC 2, VIC-II → VIC-II
            normalized = re.sub(r'VIC\s*-?\s*II|VIC\s+2', 'VIC-II', normalized, flags=re.IGNORECASE)
            # CIA 1, CIA-1 → CIA1
            normalized = re.sub(r'CIA\s*-?\s*([12])', r'CIA\1', normalized, flags=re.IGNORECASE)
            # Preserve standard forms
            if normalized.lower() == 'sid':
                normalized = 'SID'
            if normalized.lower() == 'cia':
                normalized = 'CIA'
            if normalized.lower() == 'pla':
                normalized = 'PLA'
            if normalized.lower() == 'c64':
                normalized = 'C64'
            if normalized.lower() == 'c128':
                normalized = 'C128'
            if 'vic-ii' in normalized.lower():
                normalized = 'VIC-II'

        elif entity_type == 'memory_address':
            # Normalize memory addresses to uppercase $ format
            if normalized.startswith('$'):
                normalized = normalized.upper()
            elif normalized.startswith('0x') or normalized.startswith('&H'):
                # Convert 0xD000 → $D000
                hex_part = normalized[2:] if normalized.startswith('0x') else normalized[2:]
                normalized = f'${hex_part.upper()}'
            elif normalized.isdigit():
                # Keep decimal as-is for now
                pass

        elif entity_type == 'instruction':
            # Instructions are always uppercase
            normalized = normalized.upper().split()[0]  # Take just the mnemonic

        elif entity_type == 'concept':
            # Normalize concept capitalization
            concept_map = {
                'sprite': 'sprite',
                'sprites': 'sprite',  # Singular form
                'raster interrupt': 'raster interrupt',
                'raster interrupts': 'raster interrupt',
                'irq': 'IRQ',
                'nmi': 'NMI',
                'dma': 'DMA',
                'bitmap mode': 'bitmap mode',
                'character mode': 'character mode',
                'multicolor mode': 'multicolor mode',
                'hires mode': 'hires mode',
            }
            lower = normalized.lower()
            normalized = concept_map.get(lower, normalized)

        return normalized

    def _extract_entities_regex(self, text: str) -> list[dict]:
        """
        Extract C64-specific entities using regex patterns (fast, no LLM needed).

        This supplements LLM-based extraction with high-confidence pattern matching
        for well-known C64 entities.

        Returns:
            List of entities with text, type, and confidence
        """
        entities = []

        # Hardware patterns (high confidence)
        hardware_patterns = [
            (r'\bVIC-?II\b', 'VIC-II', 0.98),
            (r'\bVIC\s*2\b', 'VIC-II', 0.95),
            (r'\b6569\b', '6569', 0.98),
            (r'\b6567\b', '6567', 0.98),
            (r'\bSID\b', 'SID', 0.98),
            (r'\b6581\b', '6581', 0.98),
            (r'\b8580\b', '8580', 0.98),
            (r'\bCIA(?:\s*[12])?\b', None, 0.98),  # CIA, CIA1, CIA2
            (r'\b6526\b', '6526', 0.98),
            (r'\b6502\b', '6502', 0.98),
            (r'\b6510\b', '6510', 0.98),
            (r'\bPLA\b', 'PLA', 0.95),
            (r'\b(?:C-?)?64(?:\s*C)?\b', 'C64', 0.90),
            (r'\bC-?128\b', 'C128', 0.95),
            (r'\bVIC-?20\b', 'VIC-20', 0.95),
            (r'\b1541\b', '1541', 0.95),
            (r'\b1571\b', '1571', 0.95),
            (r'\b1581\b', '1581', 0.95),
        ]

        for pattern, entity_name, confidence in hardware_patterns:
            for match in re.finditer(pattern, text, re.IGNORECASE):
                matched_text = match.group(0)
                # Get context (50 chars before and after)
                start = max(0, match.start() - 50)
                end = min(len(text), match.end() + 50)
                context = text[start:end].replace('\n', ' ')

                entities.append({
                    'entity_text': entity_name if entity_name else matched_text,
                    'entity_type': 'hardware',
                    'confidence': confidence,
                    'context': context,
                    'source': 'regex'
                })

        # Memory addresses (very high confidence)
        # Matches $D000, $d020, 53280 (decimal), 0xD000
        addr_patterns = [
            (r'\$[0-9A-Fa-f]{4}\b', 0.99),  # $D000
            (r'\b(?:0x|&H)[0-9A-Fa-f]{4}\b', 0.98),  # 0xD000 or &HD000
            (r'\b(?:53|54|55|56|57|58|59)\d{3}\b', 0.85),  # Decimal (53280-59999 range)
        ]

        for pattern, confidence in addr_patterns:
            for match in re.finditer(pattern, text):
                matched_text = match.group(0)
                start = max(0, match.start() - 50)
                end = min(len(text), match.end() + 50)
                context = text[start:end].replace('\n', ' ')

                entities.append({
                    'entity_text': matched_text.upper() if matched_text.startswith('$') else matched_text,
                    'entity_type': 'memory_address',
                    'confidence': confidence,
                    'context': context,
                    'source': 'regex'
                })

        # 6502 instructions (high confidence)
        instructions = [
            'LDA', 'STA', 'LDX', 'STX', 'LDY', 'STY',
            'TAX', 'TAY', 'TXA', 'TYA', 'TSX', 'TXS',
            'PHA', 'PLA', 'PHP', 'PLP',
            'AND', 'ORA', 'EOR',
            'ADC', 'SBC', 'CMP', 'CPX', 'CPY',
            'INC', 'INX', 'INY', 'DEC', 'DEX', 'DEY',
            'ASL', 'LSR', 'ROL', 'ROR',
            'JMP', 'JSR', 'RTS', 'RTI',
            'BCC', 'BCS', 'BEQ', 'BNE', 'BMI', 'BPL', 'BVC', 'BVS',
            'CLC', 'CLD', 'CLI', 'CLV', 'SEC', 'SED', 'SEI',
            'BRK', 'NOP', 'BIT'
        ]

        for instruction in instructions:
            # Match instruction at word boundary or with addressing mode
            pattern = r'\b' + instruction + r'\b(?:\s+[#$%@]?[\w,()]+)?'
            for match in re.finditer(pattern, text, re.IGNORECASE):
                matched_text = match.group(0).strip()
                start = max(0, match.start() - 40)
                end = min(len(text), match.end() + 40)
                context = text[start:end].replace('\n', ' ')

                # Higher confidence if followed by addressing mode
                has_operand = len(matched_text) > len(instruction)
                conf = 0.95 if has_operand else 0.85

                entities.append({
                    'entity_text': instruction.upper(),
                    'entity_type': 'instruction',
                    'confidence': conf,
                    'context': context,
                    'source': 'regex'
                })

        # Common C64 concepts (medium-high confidence)
        concept_patterns = [
            (r'\bsprite[s]?\b', 'sprite', 0.90),
            (r'\braster\s+interrupt[s]?\b', 'raster interrupt', 0.95),
            (r'\b(?:IRQ|NMI)\b', None, 0.92),
            (r'\bDMA\b', 'DMA', 0.90),
            (r'\bbitmap\s+mode\b', 'bitmap mode', 0.92),
            (r'\bcharacter\s+mode\b', 'character mode', 0.92),
            (r'\bmulti-?color\s+mode\b', 'multicolor mode', 0.92),
            (r'\bhi-?res\s+mode\b', 'hires mode', 0.90),
            (r'\bborder\s+color\b', 'border color', 0.88),
            (r'\bbackground\s+color\b', 'background color', 0.88),
            (r'\bcolor\s+RAM\b', 'color RAM', 0.90),
            (r'\bscreen\s+memory\b', 'screen memory', 0.90),
            (r'\bcharacter\s+set\b', 'character set', 0.88),
            (r'\bkernel\s+ROM\b', 'kernel ROM', 0.92),
            (r'\bBASIC\s+ROM\b', 'BASIC ROM', 0.92),
        ]

        for pattern, entity_name, confidence in concept_patterns:
            for match in re.finditer(pattern, text, re.IGNORECASE):
                matched_text = match.group(0)
                start = max(0, match.start() - 50)
                end = min(len(text), match.end() + 50)
                context = text[start:end].replace('\n', ' ')

                entities.append({
                    'entity_text': entity_name if entity_name else matched_text,
                    'entity_type': 'concept',
                    'confidence': confidence,
                    'context': context,
                    'source': 'regex'
                })

        return entities

    def extract_entities(self, doc_id: str,
                        confidence_threshold: float = 0.6,
                        force_regenerate: bool = False) -> dict:
        """
        Extract named entities from a document using LLM analysis.

        Args:
            doc_id: Document ID to extract entities from
            confidence_threshold: Minimum confidence to include entity (0.0-1.0, default: 0.6)
            force_regenerate: If True, extract even if entities already exist

        Returns:
            {
                'doc_id': str,
                'doc_title': str,
                'entities': [
                    {
                        'entity_text': 'VIC-II',
                        'entity_type': 'hardware',
                        'confidence': 0.95,
                        'context': '...snippet...',
                        'occurrence_count': 5
                    },
                    ...
                ],
                'entity_count': 42,
                'types': {'hardware': 10, 'memory_address': 8, ...}
            }

        Example:
            result = kb.extract_entities('my-doc-id', confidence_threshold=0.7)
            for entity in result['entities']:
                print(f"{entity['entity_type']}: {entity['entity_text']} ({entity['confidence']})")
        """
        # Validate document exists
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        doc = self.documents[doc_id]
        self.logger.info(f"Extracting entities from document {doc_id} ({doc.title})")

        # Check in-memory cache first (fastest)
        cache_key = f"{doc_id}:{confidence_threshold}"
        if not force_regenerate and self._entity_cache is not None:
            cached_result = self._entity_cache.get(cache_key)
            if cached_result is not None:
                self.logger.debug(f"Entity cache HIT for document: {doc_id}")
                return cached_result

        # Check if entities already exist in database
        if not force_regenerate:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM document_entities WHERE doc_id = ?", (doc_id,))
            existing_count = cursor.fetchone()[0]
            if existing_count > 0:
                self.logger.info(f"Document {doc_id} already has {existing_count} entities (use force_regenerate=True to re-extract)")
                # Return existing entities from database
                result = self.get_entities(doc_id)

                # Cache in memory for faster future access
                if self._entity_cache is not None:
                    self._entity_cache[cache_key] = result
                    self.logger.debug(f"Cached entity result from database for document: {doc_id}")

                return result

        # Get LLM client
        try:
            from llm_integration import get_llm_client
        except ImportError:
            raise ImportError("llm_integration module not found. Install required dependencies.")

        llm_client = get_llm_client()
        if not llm_client:
            raise ValueError("LLM not configured. Set LLM_PROVIDER and appropriate API key (ANTHROPIC_API_KEY or OPENAI_API_KEY)")

        # Get document chunks for sampling
        chunks = self._get_chunks_db(doc_id)
        if not chunks:
            raise ValueError(f"No chunks found for document {doc_id}")

        # Sample first 5 chunks (balance between coverage and cost)
        sample_chunks = chunks[:5] if len(chunks) > 5 else chunks
        sample_text = "\n\n".join([c.content for c in sample_chunks])

        # Limit to 5000 characters for cost control
        if len(sample_text) > 5000:
            sample_text = sample_text[:5000] + "..."

        # Build prompt for entity extraction
        prompt = f"""Extract named entities from this Commodore 64 technical documentation.

Document Title: {doc.title}

Content:
{sample_text}

Extract entities in these categories:
1. hardware - Chip names and components (SID, VIC-II, CIA, 6502, 6526, 6581, etc.)
2. memory_address - Memory addresses in any format ($D000, $D020, 53280, 0xD020, etc.)
3. instruction - Assembly instructions (LDA, STA, JMP, JSR, RTS, BRK, etc.)
4. person - People mentioned (Bob Yannes, Jack Tramiel, etc.)
5. company - Organizations (Commodore, MOS Technology, etc.)
6. product - Hardware/software products (VIC-20, C128, 1541, etc.)
7. concept - Technical concepts (sprite, raster interrupt, IRQ, DMA, etc.)

For each entity found, provide:
- entity_text: The entity name as it appears in the document
- entity_type: One of the categories above (lowercase with underscores)
- confidence: How confident you are this is a valid entity (0.0-1.0)
- context: Brief surrounding text showing how the entity is used (max 100 chars)

Return ONLY valid JSON in this exact format:
{{
    "entities": [
        {{
            "entity_text": "VIC-II",
            "entity_type": "hardware",
            "confidence": 0.95,
            "context": "The VIC-II chip controls all graphics"
        }},
        {{
            "entity_text": "$D020",
            "entity_type": "memory_address",
            "confidence": 0.98,
            "context": "Border color is controlled by $D020"
        }}
    ]
}}

Important:
- Extract 20-50 entities maximum
- Include confidence scores (0.0-1.0)
- Provide brief context snippets
- Preserve original capitalization/format
- Return ONLY JSON, no other text
"""

        # Extract entities using regex patterns first (fast, high-confidence)
        self.logger.info("Extracting entities using regex patterns")
        regex_entities = self._extract_entities_regex(sample_text)
        self.logger.info(f"Regex extraction found {len(regex_entities)} entities")

        # Call LLM for additional entity extraction
        self.logger.info(f"Calling LLM for entity extraction ({len(sample_text)} chars)")
        llm_entities = []
        try:
            response = llm_client.call_json(prompt, max_tokens=2048, temperature=0.3)
            llm_entities = response.get('entities', [])
            self.logger.info(f"LLM returned {len(llm_entities)} entities")
        except Exception as e:
            self.logger.warning(f"LLM call failed, using regex-only extraction: {e}")
            # Continue with regex entities only

        # Combine regex and LLM entities
        all_entities = regex_entities + llm_entities

        # Filter by confidence threshold
        filtered_entities = [
            e for e in all_entities
            if e.get('confidence', 0) >= confidence_threshold
        ]

        # Enhanced deduplication with entity normalization
        entity_map = {}
        for entity in filtered_entities:
            # Normalize entity text for better matching
            normalized_text = self._normalize_entity_text(entity['entity_text'], entity['entity_type'])
            key = (normalized_text.lower(), entity['entity_type'])

            if key in entity_map:
                entity_map[key]['occurrence_count'] += 1

                # Combine confidences: prefer regex (higher quality), boost if both sources agree
                current_source = entity.get('source', 'llm')
                existing_source = entity_map[key].get('source', 'llm')

                if current_source == 'regex' and existing_source == 'llm':
                    # Regex trumps LLM (higher precision)
                    entity_map[key]['confidence'] = entity['confidence']
                    entity_map[key]['source'] = 'regex'
                elif current_source == 'llm' and existing_source == 'regex':
                    # Both sources agree - boost confidence slightly
                    entity_map[key]['confidence'] = min(0.99, entity_map[key]['confidence'] * 1.05)
                    entity_map[key]['source'] = 'both'
                else:
                    # Keep highest confidence
                    if entity['confidence'] > entity_map[key]['confidence']:
                        entity_map[key]['confidence'] = entity['confidence']
                        entity_map[key]['context'] = entity.get('context', '')
            else:
                entity_map[key] = {
                    'entity_text': normalized_text,  # Use normalized form
                    'entity_type': entity['entity_type'],
                    'confidence': entity['confidence'],
                    'context': entity.get('context', ''),
                    'occurrence_count': 1,
                    'source': entity.get('source', 'llm')
                }

        unique_entities = list(entity_map.values())

        # Store in database
        cursor = self.db_conn.cursor()
        try:
            # Check if document still exists (might have been removed during background processing)
            doc_exists = cursor.execute(
                "SELECT 1 FROM documents WHERE doc_id = ?", (doc_id,)
            ).fetchone()

            if not doc_exists:
                self.logger.warning(
                    f"Document {doc_id} no longer exists, skipping entity storage "
                    f"(likely removed during background extraction)"
                )
                # Return entities even though storage was skipped
                result = {
                    'doc_id': doc_id,
                    'doc_title': 'Document Removed',
                    'entity_count': len(unique_entities),
                    'entities_by_type': {},
                    'entities': unique_entities
                }
                return result

            # Delete existing entities for this document
            cursor.execute("DELETE FROM document_entities WHERE doc_id = ?", (doc_id,))

            # Insert new entities
            from datetime import datetime
            generated_at = datetime.now().isoformat()
            model = llm_client.model if hasattr(llm_client, 'model') else 'unknown'

            for i, entity in enumerate(unique_entities, 1):
                cursor.execute("""
                    INSERT INTO document_entities
                    (doc_id, entity_id, entity_text, entity_type, confidence, context,
                     occurrence_count, generated_at, model)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    doc_id,
                    i,
                    entity['entity_text'],
                    entity['entity_type'],
                    entity['confidence'],
                    entity.get('context', ''),
                    entity.get('occurrence_count', 1),
                    generated_at,
                    model
                ))

            self.db_conn.commit()
            self.logger.info(f"Stored {len(unique_entities)} entities for document {doc_id}")

        except Exception as e:
            self.db_conn.rollback()
            self.logger.error(f"Failed to store entities in database: {e}")
            # Return entities even if storage failed
            pass

        # Build result
        types = {}
        for entity in unique_entities:
            entity_type = entity['entity_type']
            types[entity_type] = types.get(entity_type, 0) + 1

        result = {
            'doc_id': doc_id,
            'doc_title': doc.title,
            'entities': unique_entities,
            'entity_count': len(unique_entities),
            'types': types
        }

        # Cache result for future access (expensive LLM operation)
        if self._entity_cache is not None:
            self._entity_cache[cache_key] = result
            self.logger.debug(f"Cached entity extraction result for document: {doc_id}")

        return result

    def _extraction_worker_loop(self):
        """Background worker that processes entity extraction jobs from the queue."""
        self.logger.info("Entity extraction worker started")

        while not self._extraction_shutdown.is_set():
            try:
                # Wait for a job with timeout to allow checking shutdown flag
                try:
                    job = self._extraction_queue.get(timeout=1.0)
                except queue.Empty:
                    continue

                doc_id = job['doc_id']
                job_id = job['job_id']
                confidence_threshold = job['confidence_threshold']

                self.logger.info(f"Processing extraction job {job_id} for document {doc_id}")

                # Update job status to 'running'
                cursor = self.db_conn.cursor()
                cursor.execute("""
                    UPDATE extraction_jobs
                    SET status = 'running', started_at = ?
                    WHERE job_id = ?
                """, (datetime.now(timezone.utc).isoformat(), job_id))
                self.db_conn.commit()

                try:
                    # Extract entities (this calls the existing method)
                    result = self.extract_entities(
                        doc_id=doc_id,
                        confidence_threshold=confidence_threshold,
                        force_regenerate=False
                    )

                    # Update job status to 'completed'
                    cursor.execute("""
                        UPDATE extraction_jobs
                        SET status = 'completed',
                            completed_at = ?,
                            entities_extracted = ?
                        WHERE job_id = ?
                    """, (
                        datetime.now(timezone.utc).isoformat(),
                        result.get('entity_count', 0),
                        job_id
                    ))
                    self.db_conn.commit()

                    self.logger.info(f"Completed extraction job {job_id}: {result.get('entity_count', 0)} entities extracted")

                except Exception as e:
                    # Update job status to 'failed'
                    error_msg = str(e)
                    self.logger.error(f"Extraction job {job_id} failed: {error_msg}")

                    cursor.execute("""
                        UPDATE extraction_jobs
                        SET status = 'failed',
                            completed_at = ?,
                            error_message = ?
                        WHERE job_id = ?
                    """, (datetime.now(timezone.utc).isoformat(), error_msg, job_id))
                    self.db_conn.commit()

                # Mark job as done in the queue
                self._extraction_queue.task_done()

            except Exception as e:
                self.logger.error(f"Error in extraction worker loop: {e}")

        self.logger.info("Entity extraction worker stopped")

    def queue_entity_extraction(self, doc_id: str,
                                confidence_threshold: float = 0.6,
                                skip_if_exists: bool = True) -> dict:
        """
        Queue a document for background entity extraction.

        Args:
            doc_id: Document ID to extract entities from
            confidence_threshold: Minimum confidence threshold (0.0-1.0, default: 0.6)
            skip_if_exists: If True, skip if entities already exist or job is queued

        Returns:
            {
                'queued': bool,          # True if job was queued
                'job_id': int,           # Job ID if queued
                'reason': str,           # Reason if not queued
                'existing_job_id': int   # Existing job ID if already queued
            }
        """
        # Validate document exists
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        # Check if entities already exist
        if skip_if_exists:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM document_entities WHERE doc_id = ?", (doc_id,))
            existing_count = cursor.fetchone()[0]
            if existing_count > 0:
                return {
                    'queued': False,
                    'reason': f"Document already has {existing_count} entities",
                    'existing_entities': existing_count
                }

            # Check if job already queued or running
            cursor.execute("""
                SELECT job_id, status FROM extraction_jobs
                WHERE doc_id = ? AND status IN ('queued', 'running')
                ORDER BY queued_at DESC
                LIMIT 1
            """, (doc_id,))
            existing_job = cursor.fetchone()
            if existing_job:
                return {
                    'queued': False,
                    'reason': f"Extraction already {existing_job[1]} for this document",
                    'existing_job_id': existing_job[0]
                }

        # Create extraction job record
        cursor = self.db_conn.cursor()
        cursor.execute("""
            INSERT INTO extraction_jobs (doc_id, status, confidence_threshold, queued_at)
            VALUES (?, 'queued', ?, ?)
        """, (doc_id, confidence_threshold, datetime.now(timezone.utc).isoformat()))
        self.db_conn.commit()

        job_id = cursor.lastrowid

        # Add job to queue
        self._extraction_queue.put({
            'job_id': job_id,
            'doc_id': doc_id,
            'confidence_threshold': confidence_threshold
        })

        self.logger.info(f"Queued entity extraction job {job_id} for document {doc_id}")

        return {
            'queued': True,
            'job_id': job_id
        }

    def get_extraction_status(self, doc_id: str) -> dict:
        """
        Get the entity extraction status for a document.

        Args:
            doc_id: Document ID

        Returns:
            {
                'doc_id': str,
                'has_entities': bool,
                'entity_count': int,
                'jobs': [list of job dicts with status, queued_at, etc.]
            }
        """
        # Check if entities exist
        cursor = self.db_conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM document_entities WHERE doc_id = ?", (doc_id,))
        entity_count = cursor.fetchone()[0]

        # Get extraction jobs for this document
        cursor.execute("""
            SELECT job_id, status, confidence_threshold, queued_at,
                   started_at, completed_at, error_message, entities_extracted
            FROM extraction_jobs
            WHERE doc_id = ?
            ORDER BY queued_at DESC
        """, (doc_id,))

        jobs = []
        for row in cursor.fetchall():
            jobs.append({
                'job_id': row[0],
                'status': row[1],
                'confidence_threshold': row[2],
                'queued_at': row[3],
                'started_at': row[4],
                'completed_at': row[5],
                'error_message': row[6],
                'entities_extracted': row[7]
            })

        return {
            'doc_id': doc_id,
            'has_entities': entity_count > 0,
            'entity_count': entity_count,
            'jobs': jobs
        }

    def get_all_extraction_jobs(self, status_filter: Optional[str] = None,
                                limit: int = 100) -> list[dict]:
        """
        Get all entity extraction jobs.

        Args:
            status_filter: Optional status filter ('queued', 'running', 'completed', 'failed')
            limit: Maximum number of jobs to return (default: 100)

        Returns:
            List of job dicts with doc_id, status, timestamps, etc.
        """
        cursor = self.db_conn.cursor()

        if status_filter:
            cursor.execute("""
                SELECT j.job_id, j.doc_id, d.title, j.status, j.confidence_threshold,
                       j.queued_at, j.started_at, j.completed_at, j.error_message,
                       j.entities_extracted
                FROM extraction_jobs j
                LEFT JOIN documents d ON j.doc_id = d.doc_id
                WHERE j.status = ?
                ORDER BY j.queued_at DESC
                LIMIT ?
            """, (status_filter, limit))
        else:
            cursor.execute("""
                SELECT j.job_id, j.doc_id, d.title, j.status, j.confidence_threshold,
                       j.queued_at, j.started_at, j.completed_at, j.error_message,
                       j.entities_extracted
                FROM extraction_jobs j
                LEFT JOIN documents d ON j.doc_id = d.doc_id
                ORDER BY j.queued_at DESC
                LIMIT ?
            """, (limit,))

        jobs = []
        for row in cursor.fetchall():
            jobs.append({
                'job_id': row[0],
                'doc_id': row[1],
                'doc_title': row[2],
                'status': row[3],
                'confidence_threshold': row[4],
                'queued_at': row[5],
                'started_at': row[6],
                'completed_at': row[7],
                'error_message': row[8],
                'entities_extracted': row[9]
            })

        return jobs

    def get_entities(self, doc_id: str,
                    entity_types: Optional[list[str]] = None,
                    min_confidence: float = 0.0) -> dict:
        """
        Get all entities for a specific document.

        Args:
            doc_id: Document ID
            entity_types: Optional list of entity types to filter by
                         (e.g., ['hardware', 'memory_address'])
            min_confidence: Minimum confidence threshold (default: 0.0)

        Returns:
            {
                'doc_id': str,
                'doc_title': str,
                'entities': [list of entity dicts],
                'entity_count': int,
                'types': {'hardware': count, ...}
            }

        Example:
            # Get all entities
            result = kb.get_entities('my-doc-id')

            # Get only hardware entities with high confidence
            result = kb.get_entities('my-doc-id',
                                    entity_types=['hardware'],
                                    min_confidence=0.8)
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        doc = self.documents[doc_id]
        cursor = self.db_conn.cursor()

        # Build query with optional filters
        query = """
            SELECT entity_text, entity_type, confidence, context, occurrence_count
            FROM document_entities
            WHERE doc_id = ? AND confidence >= ?
        """
        params = [doc_id, min_confidence]

        if entity_types:
            placeholders = ','.join(['?'] * len(entity_types))
            query += f" AND entity_type IN ({placeholders})"
            params.extend(entity_types)

        query += " ORDER BY entity_type, confidence DESC"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        entities = []
        types = {}
        for row in rows:
            entity = {
                'entity_text': row[0],
                'entity_type': row[1],
                'confidence': row[2],
                'context': row[3],
                'occurrence_count': row[4]
            }
            entities.append(entity)

            entity_type = row[1]
            types[entity_type] = types.get(entity_type, 0) + 1

        return {
            'doc_id': doc_id,
            'doc_title': doc.title,
            'entities': entities,
            'entity_count': len(entities),
            'types': types
        }

    def search_entities(self, query: str,
                       entity_types: Optional[list[str]] = None,
                       min_confidence: float = 0.0,
                       max_results: int = 20) -> dict:
        """
        Search for entities across all documents using full-text search.

        Args:
            query: Search query (e.g., "VIC-II", "sprite", "$D000")
            entity_types: Filter by entity types (e.g., ['hardware', 'memory_address'])
            min_confidence: Minimum confidence threshold (0.0-1.0)
            max_results: Maximum number of results to return

        Returns:
            Dictionary with search results grouped by document:
            {
                'query': str,
                'total_matches': int,
                'documents': [
                    {
                        'doc_id': str,
                        'doc_title': str,
                        'matches': [
                            {
                                'entity_text': str,
                                'entity_type': str,
                                'confidence': float,
                                'context': str,
                                'occurrence_count': int
                            },
                            ...
                        ],
                        'match_count': int
                    },
                    ...
                ]
            }

        Examples:
            # Search for VIC-II chip mentions
            results = kb.search_entities('VIC-II')

            # Search for memory addresses only
            results = kb.search_entities('$D0', entity_types=['memory_address'])

            # Search with confidence threshold
            results = kb.search_entities('sprite', min_confidence=0.7)
        """
        if not query or not query.strip():
            raise ValueError("Search query cannot be empty")

        # Build FTS5 query
        # Escape special FTS5 characters and wrap in quotes for literal search
        escaped_query = query.strip().replace('"', '""')
        fts_query = f'"{escaped_query}"'

        # Build WHERE clause for filtering
        where_clauses = []
        params = []

        if entity_types:
            placeholders = ','.join('?' * len(entity_types))
            where_clauses.append(f"e.entity_type IN ({placeholders})")
            params.extend(entity_types)

        if min_confidence > 0.0:
            where_clauses.append("e.confidence >= ?")
            params.append(min_confidence)

        where_clause = " AND " + " AND ".join(where_clauses) if where_clauses else ""

        # Search entities_fts and join with document_entities for full data
        query_sql = f"""
            SELECT e.doc_id, e.entity_text, e.entity_type, e.confidence,
                   e.context, e.occurrence_count
            FROM entities_fts fts
            JOIN document_entities e ON fts.rowid = e.rowid
            WHERE entities_fts MATCH ?{where_clause}
            ORDER BY rank
            LIMIT ?
        """

        # Execute search
        cursor = self.db_conn.cursor()
        cursor.execute(query_sql, [fts_query] + params + [max_results * 10])  # Get extra for grouping
        rows = cursor.fetchall()

        # Group results by document
        doc_matches = {}
        for row in rows:
            doc_id, entity_text, entity_type, confidence, context, occurrence_count = row

            if doc_id not in doc_matches:
                doc_matches[doc_id] = []

            doc_matches[doc_id].append({
                'entity_text': entity_text,
                'entity_type': entity_type,
                'confidence': confidence,
                'context': context or '',
                'occurrence_count': occurrence_count
            })

        # Build result list with document titles
        documents = []
        for doc_id, matches in list(doc_matches.items())[:max_results]:
            doc = self.documents.get(doc_id)
            if doc:
                documents.append({
                    'doc_id': doc_id,
                    'doc_title': doc.title,
                    'matches': matches,
                    'match_count': len(matches)
                })

        return {
            'query': query,
            'total_matches': sum(len(matches) for matches in doc_matches.values()),
            'documents': documents
        }

    def find_docs_by_entity(self, entity_text: str,
                           entity_type: Optional[str] = None,
                           min_confidence: float = 0.0,
                           max_results: int = 20) -> dict:
        """
        Find all documents that contain a specific entity.

        Args:
            entity_text: Exact entity text to search for (e.g., "VIC-II", "$D000")
            entity_type: Optional entity type filter (e.g., 'hardware', 'memory_address')
            min_confidence: Minimum confidence threshold (0.0-1.0)
            max_results: Maximum number of documents to return

        Returns:
            Dictionary with documents containing the entity:
            {
                'entity_text': str,
                'entity_type': str or None,
                'total_documents': int,
                'documents': [
                    {
                        'doc_id': str,
                        'doc_title': str,
                        'entity_type': str,
                        'confidence': float,
                        'context': str,
                        'occurrence_count': int
                    },
                    ...
                ]
            }

        Examples:
            # Find all documents mentioning VIC-II
            results = kb.find_docs_by_entity('VIC-II')

            # Find documents with $D000 memory address
            results = kb.find_docs_by_entity('$D000', entity_type='memory_address')

            # Find with confidence threshold
            results = kb.find_docs_by_entity('sprite', min_confidence=0.7)
        """
        if not entity_text or not entity_text.strip():
            raise ValueError("Entity text cannot be empty")

        # Build WHERE clause
        where_clauses = ["e.entity_text = ?"]
        params = [entity_text.strip()]

        if entity_type:
            where_clauses.append("e.entity_type = ?")
            params.append(entity_type)

        if min_confidence > 0.0:
            where_clauses.append("e.confidence >= ?")
            params.append(min_confidence)

        where_clause = " AND ".join(where_clauses)

        # Query database
        query_sql = f"""
            SELECT e.doc_id, e.entity_type, e.confidence, e.context,
                   e.occurrence_count
            FROM document_entities e
            WHERE {where_clause}
            ORDER BY e.confidence DESC, e.occurrence_count DESC
            LIMIT ?
        """

        cursor = self.db_conn.cursor()
        cursor.execute(query_sql, params + [max_results])
        rows = cursor.fetchall()

        # Build result list with document titles
        documents = []
        for row in rows:
            doc_id, ent_type, confidence, context, occurrence_count = row
            doc = self.documents.get(doc_id)
            if doc:
                documents.append({
                    'doc_id': doc_id,
                    'doc_title': doc.title,
                    'entity_type': ent_type,
                    'confidence': confidence,
                    'context': context or '',
                    'occurrence_count': occurrence_count
                })

        return {
            'entity_text': entity_text.strip(),
            'entity_type': entity_type,
            'total_documents': len(documents),
            'documents': documents
        }

    def get_entity_stats(self, entity_type: Optional[str] = None) -> dict:
        """
        Get statistics about extracted entities in the knowledge base.

        Args:
            entity_type: Optional filter by entity type (e.g., 'hardware', 'memory_address')

        Returns:
            Dictionary with entity statistics:
            {
                'total_entities': int,
                'total_documents_with_entities': int,
                'by_type': {
                    'hardware': int,
                    'memory_address': int,
                    ...
                },
                'top_entities': [
                    {
                        'entity_text': str,
                        'entity_type': str,
                        'document_count': int,
                        'total_occurrences': int,
                        'avg_confidence': float
                    },
                    ...
                ],
                'documents_with_most_entities': [
                    {
                        'doc_id': str,
                        'doc_title': str,
                        'entity_count': int
                    },
                    ...
                ]
            }

        Examples:
            # Get overall statistics
            stats = kb.get_entity_stats()

            # Get statistics for hardware entities only
            stats = kb.get_entity_stats(entity_type='hardware')
        """
        cursor = self.db_conn.cursor()

        # Total entities
        if entity_type:
            cursor.execute(
                "SELECT COUNT(*) FROM document_entities WHERE entity_type = ?",
                (entity_type,)
            )
        else:
            cursor.execute("SELECT COUNT(*) FROM document_entities")
        total_entities = cursor.fetchone()[0]

        # Total documents with entities
        if entity_type:
            cursor.execute(
                "SELECT COUNT(DISTINCT doc_id) FROM document_entities WHERE entity_type = ?",
                (entity_type,)
            )
        else:
            cursor.execute("SELECT COUNT(DISTINCT doc_id) FROM document_entities")
        total_docs = cursor.fetchone()[0]

        # Breakdown by type
        by_type = {}
        if entity_type:
            by_type[entity_type] = total_entities
        else:
            cursor.execute("""
                SELECT entity_type, COUNT(*)
                FROM document_entities
                GROUP BY entity_type
                ORDER BY COUNT(*) DESC
            """)
            by_type = {row[0]: row[1] for row in cursor.fetchall()}

        # Top entities by document count
        type_filter = "WHERE entity_type = ?" if entity_type else ""
        params = [entity_type] if entity_type else []

        cursor.execute(f"""
            SELECT entity_text, entity_type,
                   COUNT(DISTINCT doc_id) as doc_count,
                   SUM(occurrence_count) as total_occurrences,
                   AVG(confidence) as avg_confidence
            FROM document_entities
            {type_filter}
            GROUP BY entity_text, entity_type
            ORDER BY doc_count DESC, total_occurrences DESC
            LIMIT 20
        """, params)

        top_entities = [
            {
                'entity_text': row[0],
                'entity_type': row[1],
                'document_count': row[2],
                'total_occurrences': row[3],
                'avg_confidence': round(row[4], 3)
            }
            for row in cursor.fetchall()
        ]

        # Documents with most entities
        cursor.execute(f"""
            SELECT doc_id, COUNT(*) as entity_count
            FROM document_entities
            {type_filter}
            GROUP BY doc_id
            ORDER BY entity_count DESC
            LIMIT 10
        """, params)

        docs_with_most = []
        for row in cursor.fetchall():
            doc_id = row[0]
            entity_count = row[1]
            doc = self.documents.get(doc_id)
            if doc:
                docs_with_most.append({
                    'doc_id': doc_id,
                    'doc_title': doc.title,
                    'entity_count': entity_count
                })

        return {
            'total_entities': total_entities,
            'total_documents_with_entities': total_docs,
            'by_type': by_type,
            'top_entities': top_entities,
            'documents_with_most_entities': docs_with_most
        }

    def export_entities(self, format: str = 'csv',
                       entity_types: Optional[list] = None,
                       min_confidence: float = 0.0,
                       output_path: Optional[str] = None) -> str:
        """
        Export entities to CSV or JSON format.

        Args:
            format: 'csv' or 'json'
            entity_types: Filter by entity types (None = all types)
            min_confidence: Minimum confidence threshold (0.0-1.0)
            output_path: Optional file path to write to (if None, returns string)

        Returns:
            Exported data as string (or writes to file if output_path provided)

        Example CSV:
            entity_text,entity_type,confidence,doc_count,occurrence_count,first_seen_doc
            VIC-II,hardware,0.95,15,87,89d0943d6009
            SID,hardware,0.92,12,56,89d0943d6009

        Example JSON:
            [
                {
                    "entity_text": "VIC-II",
                    "entity_type": "hardware",
                    "confidence": 0.95,
                    "doc_count": 15,
                    "occurrence_count": 87,
                    "first_seen_doc": "89d0943d6009"
                },
                ...
            ]
        """
        import csv
        import json
        from io import StringIO

        cursor = self.db_conn.cursor()

        # Build query with filters
        query = """
            SELECT entity_text, entity_type,
                   AVG(confidence) as avg_confidence,
                   COUNT(DISTINCT doc_id) as doc_count,
                   SUM(occurrence_count) as total_occurrences,
                   MIN(doc_id) as first_seen_doc
            FROM document_entities
            WHERE confidence >= ?
        """
        params = [min_confidence]

        if entity_types:
            placeholders = ','.join(['?'] * len(entity_types))
            query += f" AND entity_type IN ({placeholders})"
            params.extend(entity_types)

        query += """
            GROUP BY entity_text, entity_type
            ORDER BY doc_count DESC, total_occurrences DESC
        """

        cursor.execute(query, params)
        rows = cursor.fetchall()

        if format.lower() == 'csv':
            output = StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow(['entity_text', 'entity_type', 'avg_confidence',
                           'doc_count', 'total_occurrences', 'first_seen_doc'])

            # Write data
            for row in rows:
                writer.writerow([
                    row[0],  # entity_text
                    row[1],  # entity_type
                    f"{row[2]:.3f}",  # avg_confidence
                    row[3],  # doc_count
                    row[4],  # total_occurrences
                    row[5][:12]  # first_seen_doc (truncated)
                ])

            result = output.getvalue()
            output.close()

        elif format.lower() == 'json':
            entities = []
            for row in rows:
                entities.append({
                    'entity_text': row[0],
                    'entity_type': row[1],
                    'avg_confidence': round(row[2], 3),
                    'doc_count': row[3],
                    'total_occurrences': row[4],
                    'first_seen_doc': row[5][:12]
                })

            result = json.dumps(entities, indent=2)

        else:
            raise ValueError(f"Unsupported format: {format}. Use 'csv' or 'json'.")

        # Write to file if path provided
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(result)
            self.logger.info(f"Exported {len(rows)} entities to {output_path}")

        return result

    def export_relationships(self, format: str = 'csv',
                            min_strength: float = 0.0,
                            entity_types: Optional[list] = None,
                            output_path: Optional[str] = None) -> str:
        """
        Export entity relationships to CSV or JSON format.

        Args:
            format: 'csv' or 'json'
            min_strength: Minimum relationship strength (0.0-1.0)
            entity_types: Filter by entity types (None = all types)
            output_path: Optional file path to write to (if None, returns string)

        Returns:
            Exported data as string (or writes to file if output_path provided)

        Example CSV:
            entity1,entity1_type,entity2,entity2_type,strength,doc_count
            VIC-II,hardware,sprite,concept,0.85,12
            SID,hardware,sound,concept,0.78,9

        Example JSON:
            [
                {
                    "entity1": "VIC-II",
                    "entity1_type": "hardware",
                    "entity2": "sprite",
                    "entity2_type": "concept",
                    "strength": 0.85,
                    "doc_count": 12
                },
                ...
            ]
        """
        import csv
        import json
        from io import StringIO

        cursor = self.db_conn.cursor()

        # Build query with filters
        query = """
            SELECT entity1_text, entity1_type,
                   entity2_text, entity2_type,
                   strength, doc_count
            FROM entity_relationships
            WHERE strength >= ?
        """
        params = [min_strength]

        if entity_types:
            placeholders = ','.join(['?'] * len(entity_types))
            query += f" AND (entity1_type IN ({placeholders}) OR entity2_type IN ({placeholders}))"
            params.extend(entity_types * 2)  # Add for both entity1_type and entity2_type

        query += """
            ORDER BY strength DESC, doc_count DESC
        """

        cursor.execute(query, params)
        rows = cursor.fetchall()

        if format.lower() == 'csv':
            output = StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow(['entity1', 'entity1_type', 'entity2', 'entity2_type',
                           'strength', 'doc_count'])

            # Write data
            for row in rows:
                writer.writerow([
                    row[0],  # entity1_text
                    row[1],  # entity1_type
                    row[2],  # entity2_text
                    row[3],  # entity2_type
                    f"{row[4]:.3f}",  # strength
                    row[5]   # doc_count
                ])

            result = output.getvalue()
            output.close()

        elif format.lower() == 'json':
            relationships = []
            for row in rows:
                relationships.append({
                    'entity1': row[0],
                    'entity1_type': row[1],
                    'entity2': row[2],
                    'entity2_type': row[3],
                    'strength': round(row[4], 3),
                    'doc_count': row[5]
                })

            result = json.dumps(relationships, indent=2)

        else:
            raise ValueError(f"Unsupported format: {format}. Use 'csv' or 'json'.")

        # Write to file if path provided
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(result)
            self.logger.info(f"Exported {len(rows)} relationships to {output_path}")

        return result

    def extract_entities_bulk(self, confidence_threshold: float = 0.6,
                             force_regenerate: bool = False,
                             max_docs: Optional[int] = None,
                             skip_existing: bool = True) -> dict:
        """
        Bulk extract entities for multiple documents.

        Args:
            confidence_threshold: Minimum confidence threshold (0.0-1.0, default: 0.6)
            force_regenerate: If True, re-extract even if entities already exist
            max_docs: Maximum number of documents to process (None = all)
            skip_existing: If True, skip documents that already have entities (unless force_regenerate)

        Returns:
            {
                'processed': int,
                'failed': int,
                'skipped': int,
                'total_entities': int,
                'by_type': {'hardware': int, 'memory_address': int, ...},
                'results': [list of individual results]
            }

        Examples:
            # Extract entities for all documents
            results = kb.extract_entities_bulk()

            # Extract for first 10 documents only
            results = kb.extract_entities_bulk(max_docs=10)

            # Force re-extraction for all documents
            results = kb.extract_entities_bulk(force_regenerate=True)
        """
        results = {
            'processed': 0,
            'failed': 0,
            'skipped': 0,
            'total_entities': 0,
            'by_type': {},
            'results': []
        }

        # Get documents to process
        docs_to_process = list(self.documents.keys())

        if max_docs:
            docs_to_process = docs_to_process[:max_docs]

        self.logger.info(f"Extracting entities for {len(docs_to_process)} documents "
                        f"(confidence threshold: {confidence_threshold})")

        # Process each document
        for i, doc_id in enumerate(docs_to_process, 1):
            doc_results = {
                'doc_id': doc_id,
                'title': self.documents[doc_id].title,
                'status': 'unknown',
                'entity_count': 0,
                'error': None
            }

            try:
                # Check if entities already exist (unless force_regenerate)
                if skip_existing and not force_regenerate:
                    cursor = self.db_conn.cursor()
                    cursor.execute(
                        "SELECT COUNT(*) FROM document_entities WHERE doc_id = ?",
                        (doc_id,)
                    )
                    existing_count = cursor.fetchone()[0]

                    if existing_count > 0:
                        self.logger.info(f"[{i}/{len(docs_to_process)}] Skipping {doc_id} (already has {existing_count} entities)")
                        doc_results['status'] = 'skipped'
                        doc_results['entity_count'] = existing_count
                        results['skipped'] += 1
                        results['results'].append(doc_results)
                        continue

                # Extract entities
                self.logger.info(f"[{i}/{len(docs_to_process)}] Extracting entities from {doc_id}")

                result = self.extract_entities(
                    doc_id,
                    confidence_threshold=confidence_threshold,
                    force_regenerate=force_regenerate
                )

                doc_results['status'] = 'success'
                doc_results['entity_count'] = result['entity_count']

                # Update counts
                results['processed'] += 1
                results['total_entities'] += result['entity_count']

                # Update by_type counts
                for entity_type, count in result['types'].items():
                    results['by_type'][entity_type] = results['by_type'].get(entity_type, 0) + count

                results['results'].append(doc_results)

            except Exception as e:
                self.logger.error(f"[{i}/{len(docs_to_process)}] Failed to extract entities from {doc_id}: {e}")
                doc_results['status'] = 'failed'
                doc_results['error'] = str(e)
                results['failed'] += 1
                results['results'].append(doc_results)

        self.logger.info(f"Bulk entity extraction complete: processed={results['processed']}, "
                        f"failed={results['failed']}, skipped={results['skipped']}, "
                        f"total_entities={results['total_entities']}")

        return results

    # ========== ENTITY RELATIONSHIP METHODS ==========

    def extract_entity_relationships(self, doc_id: str, min_confidence: float = 0.6,
                                   force_regenerate: bool = False) -> dict:
        """
        Extract entity co-occurrence relationships from a document.

        This method analyzes how entities appear together in document chunks
        to identify related concepts, hardware, instructions, etc.

        Args:
            doc_id: Document ID to extract relationships from
            min_confidence: Minimum confidence threshold for entities (default: 0.6)
            force_regenerate: If True, regenerate relationships even if they exist

        Returns:
            {
                'doc_id': str,
                'relationship_count': int,
                'relationships': [
                    {
                        'entity1': str,
                        'entity1_type': str,
                        'entity2': str,
                        'entity2_type': str,
                        'strength': float,
                        'context': str
                    },
                    ...
                ]
            }

        Examples:
            # Extract relationships from a document
            result = kb.extract_entity_relationships('doc-id-123')

            # Force regeneration with higher confidence
            result = kb.extract_entity_relationships('doc-id-123',
                min_confidence=0.7, force_regenerate=True)
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        self.logger.info(f"Extracting entity relationships from document {doc_id}")

        # Get all entities for this document
        entity_result = self.get_entities(doc_id, min_confidence=min_confidence)

        if not entity_result['entities']:
            self.logger.info(f"No entities found for document {doc_id}")
            return {
                'doc_id': doc_id,
                'relationship_count': 0,
                'relationships': []
            }

        # Get all chunks for this document
        chunks = self._get_chunks_db(doc_id)

        if not chunks:
            self.logger.info(f"No chunks found for document {doc_id}")
            return {
                'doc_id': doc_id,
                'relationship_count': 0,
                'relationships': []
            }

        # Build entity -> type mapping from the entities list
        entity_map = {e['entity_text']: e['entity_type'] for e in entity_result['entities']}

        # Find co-occurrences with enhanced distance-based strength calculation
        relationships = {}  # (entity1, entity2) -> strength
        relationship_contexts = {}  # (entity1, entity2) -> context

        for chunk in chunks:
            content = chunk.content
            # Find which entities appear in this chunk and their positions
            entity_positions = {}
            for entity in entity_map.keys():
                pos = content.find(entity)
                if pos != -1:
                    entity_positions[entity] = pos

            entities_in_chunk = list(entity_positions.keys())

            # Create pairs of co-occurring entities with distance-based scoring
            for i, e1 in enumerate(entities_in_chunk):
                for e2 in entities_in_chunk[i+1:]:
                    # Sort entities alphabetically to avoid duplicates (A,B) vs (B,A)
                    pair = tuple(sorted([e1, e2]))

                    # Calculate distance-based strength
                    pos1 = entity_positions[e1]
                    pos2 = entity_positions[e2]
                    distance = abs(pos2 - pos1)

                    # Distance-based weight: closer = stronger
                    # Uses exponential decay: weight = e^(-distance/decay_factor)
                    import math
                    decay_factor = 500  # Characters - tune this for sensitivity
                    distance_weight = math.exp(-distance / decay_factor)

                    # Base strength from co-occurrence + distance weight
                    # This gives 1.0 for same position, ~0.61 for 500 chars apart, ~0.37 for 1000 chars
                    strength_increment = 1.0 * distance_weight

                    # Add to cumulative strength
                    relationships[pair] = relationships.get(pair, 0) + strength_increment

                    # Store context (first occurrence)
                    if pair not in relationship_contexts:
                        # Extract context around both entities
                        start_idx = max(0, min(pos1, pos2) - 50)
                        end_idx = min(len(content), max(pos1, pos2) + max(len(e1), len(e2)) + 50)
                        context = content[start_idx:end_idx].strip()
                        relationship_contexts[pair] = context

        # Enhanced normalization with logarithmic scaling
        # This prevents a few very high counts from dominating
        if relationships:
            max_strength = max(relationships.values())
            # Use log scaling for better distribution
            normalized_relationships = {}
            for pair, strength in relationships.items():
                # Log-scale normalization: more even distribution of scores
                # Formula: log(1 + strength) / log(1 + max_strength)
                import math
                normalized_relationships[pair] = math.log(1 + strength) / math.log(1 + max_strength)
        else:
            normalized_relationships = {}

        # Store relationships in database
        cursor = self.db_conn.cursor()
        from datetime import datetime
        now = datetime.now().isoformat()

        relationship_list = []
        for (e1, e2), strength in normalized_relationships.items():
            e1_type = entity_map[e1]
            e2_type = entity_map[e2]
            context = relationship_contexts[(e1, e2)]

            # Check if relationship already exists
            cursor.execute("""
                SELECT strength, doc_count FROM entity_relationships
                WHERE entity1_text = ? AND entity2_text = ? AND relationship_type = 'co-occurrence'
            """, (e1, e2))

            existing = cursor.fetchone()

            if existing:
                # Update existing relationship
                old_strength, doc_count = existing
                new_strength = (old_strength * doc_count + strength) / (doc_count + 1)
                new_doc_count = doc_count + 1

                cursor.execute("""
                    UPDATE entity_relationships
                    SET strength = ?, doc_count = ?, last_updated = ?, context_sample = ?
                    WHERE entity1_text = ? AND entity2_text = ? AND relationship_type = 'co-occurrence'
                """, (new_strength, new_doc_count, now, context, e1, e2))
            else:
                # Insert new relationship
                cursor.execute("""
                    INSERT INTO entity_relationships
                    (entity1_text, entity1_type, entity2_text, entity2_type,
                     relationship_type, strength, doc_count, first_seen_doc,
                     context_sample, last_updated)
                    VALUES (?, ?, ?, ?, 'co-occurrence', ?, 1, ?, ?, ?)
                """, (e1, e1_type, e2, e2_type, strength, doc_id, context, now))

            relationship_list.append({
                'entity1': e1,
                'entity1_type': e1_type,
                'entity2': e2,
                'entity2_type': e2_type,
                'strength': strength,
                'context': context
            })

        self.db_conn.commit()

        self.logger.info(f"Extracted {len(relationship_list)} relationships from {doc_id}")

        return {
            'doc_id': doc_id,
            'relationship_count': len(relationship_list),
            'relationships': sorted(relationship_list, key=lambda x: x['strength'], reverse=True)
        }

    def get_entity_relationships(self, entity_text: str,
                                relationship_type: Optional[str] = None,
                                min_strength: float = 0.0,
                                max_results: int = 20) -> list:
        """
        Get all relationships for a given entity.

        Args:
            entity_text: The entity to find relationships for
            relationship_type: Filter by relationship type (default: all types)
            min_strength: Minimum relationship strength (0.0-1.0)
            max_results: Maximum number of results to return

        Returns:
            List of relationships:
            [
                {
                    'related_entity': str,
                    'related_type': str,
                    'relationship_type': str,
                    'strength': float,
                    'doc_count': int,
                    'context_sample': str
                },
                ...
            ]

        Examples:
            # Find entities related to VIC-II
            relationships = kb.get_entity_relationships('VIC-II')

            # Find strong co-occurrences only
            relationships = kb.get_entity_relationships('VIC-II',
                relationship_type='co-occurrence', min_strength=0.5)
        """
        cursor = self.db_conn.cursor()

        # Build query
        query = """
            SELECT entity1_text, entity1_type, entity2_text, entity2_type,
                   relationship_type, strength, doc_count, context_sample
            FROM entity_relationships
            WHERE (entity1_text = ? OR entity2_text = ?)
              AND strength >= ?
        """
        params = [entity_text, entity_text, min_strength]

        if relationship_type:
            query += " AND relationship_type = ?"
            params.append(relationship_type)

        query += " ORDER BY strength DESC, doc_count DESC LIMIT ?"
        params.append(max_results)

        cursor.execute(query, params)
        rows = cursor.fetchall()

        results = []
        for row in rows:
            e1_text, e1_type, e2_text, e2_type, rel_type, strength, doc_count, context = row

            # Determine which is the "other" entity
            if e1_text == entity_text:
                related_entity = e2_text
                related_type = e2_type
            else:
                related_entity = e1_text
                related_type = e1_type

            results.append({
                'related_entity': related_entity,
                'related_type': related_type,
                'relationship_type': rel_type,
                'strength': strength,
                'doc_count': doc_count,
                'context_sample': context
            })

        return results

    def find_related_entities(self, entity_text: str, max_results: int = 10) -> list:
        """
        Discover entities related to a given entity (convenience method).

        This is a simplified version of get_entity_relationships() optimized
        for entity discovery and exploration.

        Args:
            entity_text: The entity to find related entities for
            max_results: Maximum number of related entities to return

        Returns:
            List of related entities with strength scores

        Examples:
            # Discover entities related to SID chip
            related = kb.find_related_entities('SID')
        """
        return self.get_entity_relationships(
            entity_text=entity_text,
            relationship_type='co-occurrence',
            min_strength=0.3,
            max_results=max_results
        )

    def search_by_entity_pair(self, entity1: str, entity2: str,
                             max_results: int = 10) -> list:
        """
        Find documents that contain both entities.

        Args:
            entity1: First entity to search for
            entity2: Second entity to search for
            max_results: Maximum number of documents to return

        Returns:
            List of documents containing both entities:
            [
                {
                    'doc_id': str,
                    'title': str,
                    'entity1_count': int,
                    'entity2_count': int,
                    'contexts': [str, ...]  # Snippets showing both entities
                },
                ...
            ]

        Examples:
            # Find docs about VIC-II and raster interrupts
            docs = kb.search_by_entity_pair('VIC-II', 'raster interrupt')
        """
        cursor = self.db_conn.cursor()

        # Find documents containing both entities
        query = """
            SELECT e1.doc_id,
                   COUNT(DISTINCT e1.entity_id) as entity1_count,
                   COUNT(DISTINCT e2.entity_id) as entity2_count
            FROM document_entities e1
            JOIN document_entities e2 ON e1.doc_id = e2.doc_id
            WHERE e1.entity_text = ? AND e2.entity_text = ?
            GROUP BY e1.doc_id
            ORDER BY entity1_count + entity2_count DESC
            LIMIT ?
        """

        cursor.execute(query, (entity1, entity2, max_results))
        rows = cursor.fetchall()

        results = []
        for doc_id, e1_count, e2_count in rows:
            doc = self.documents.get(doc_id)
            if not doc:
                continue

            # Get context snippets showing both entities
            chunks = self._get_chunks_db(doc_id)
            contexts = []

            for chunk in chunks:
                if entity1 in chunk.content and entity2 in chunk.content:
                    # Extract snippet around both entities
                    idx1 = chunk.content.find(entity1)
                    idx2 = chunk.content.find(entity2)
                    start = max(0, min(idx1, idx2) - 50)
                    end = min(len(chunk.content), max(idx1 + len(entity1), idx2 + len(entity2)) + 50)
                    context = chunk.content[start:end].strip()
                    contexts.append(context)

            results.append({
                'doc_id': doc_id,
                'title': doc.title,
                'entity1_count': e1_count,
                'entity2_count': e2_count,
                'contexts': contexts[:3]  # Return top 3 contexts
            })

        return results

    # ============================================================
    # Knowledge Graph Methods (v2.24.0 - Phase 1)
    # ============================================================

    def build_knowledge_graph(self, entity_types: Optional[list[str]] = None,
                             min_occurrences: int = 2,
                             min_relationship_strength: float = 0.3,
                             use_cache: bool = True):
        """
        Build NetworkX graph from entities and relationships.

        Args:
            entity_types: Filter to specific entity types (None = all types)
            min_occurrences: Minimum entity occurrences to include in graph
            min_relationship_strength: Minimum relationship strength to include
            use_cache: Try to load from cache if available

        Returns:
            NetworkX Graph with weighted nodes and edges

        Examples:
            # Build graph from all entities
            G = kb.build_knowledge_graph()

            # Build graph for hardware entities only
            G = kb.build_knowledge_graph(entity_types=['hardware', 'register'])

            # Build graph with minimum 5 occurrences
            G = kb.build_knowledge_graph(min_occurrences=5)
        """
        import networkx as nx
        import hashlib
        import pickle

        # Try to load from cache first
        if use_cache:
            cache_key = f"{entity_types}_{min_occurrences}_{min_relationship_strength}"
            cache_id = hashlib.md5(cache_key.encode()).hexdigest()[:16]

            cached_graph = self._load_cached_graph(cache_id)
            if cached_graph:
                self.logger.info(f"Loaded knowledge graph from cache ({cached_graph.number_of_nodes()} nodes, {cached_graph.number_of_edges()} edges)")
                return cached_graph

        self.logger.info(f"Building knowledge graph (min_occurrences={min_occurrences}, min_strength={min_relationship_strength})")

        # Create empty graph
        G = nx.Graph()

        # Build query for entities
        query = """
            SELECT entity_text, entity_type, SUM(occurrence_count) as occurrences
            FROM document_entities
            GROUP BY entity_text, entity_type
            HAVING occurrences >= ?
        """
        params = [min_occurrences]

        # Add entity type filter if specified
        if entity_types:
            placeholders = ','.join('?' * len(entity_types))
            query += f" AND entity_type IN ({placeholders})"
            params.extend(entity_types)

        # Fetch entities
        cursor = self.db_conn.cursor()
        entities = cursor.execute(query, params).fetchall()

        if not entities:
            self.logger.warning("No entities found matching criteria")
            return G

        # Add nodes to graph
        for text, etype, count in entities:
            G.add_node(text,
                      type=etype,
                      occurrences=count,
                      weight=count)  # Node weight = occurrence count

        self.logger.info(f"Added {G.number_of_nodes()} nodes to graph")

        # Fetch relationships
        rel_query = """
            SELECT entity1_text, entity2_text, strength, doc_count
            FROM entity_relationships
            WHERE strength >= ?
        """

        relationships = cursor.execute(rel_query, [min_relationship_strength]).fetchall()

        # Add edges to graph (only if both nodes exist)
        edges_added = 0
        for e1, e2, strength, count in relationships:
            if G.has_node(e1) and G.has_node(e2):
                G.add_edge(e1, e2,
                          weight=strength,
                          doc_count=count)
                edges_added += 1

        self.logger.info(f"Added {edges_added} edges to graph")

        # Log graph statistics
        if G.number_of_nodes() > 0:
            density = nx.density(G)
            self.logger.info(f"Graph density: {density:.4f}")

            # Count connected components
            num_components = nx.number_connected_components(G)
            self.logger.info(f"Connected components: {num_components}")

        # Cache the graph
        if use_cache and G.number_of_nodes() > 0:
            cache_id = self._cache_graph(G, entity_types, min_occurrences, min_relationship_strength)
            self.logger.info(f"Cached graph with ID: {cache_id}")

        return G

    def _cache_graph(self, G, entity_types=None, min_occurrences=2, min_relationship_strength=0.3) -> str:
        """
        Cache NetworkX graph to database for quick reloading.

        Args:
            G: NetworkX graph to cache
            entity_types: Entity types filter used
            min_occurrences: Minimum occurrences filter used
            min_relationship_strength: Minimum strength filter used

        Returns:
            cache_id: ID of the cached graph
        """
        import pickle
        import hashlib
        from datetime import datetime

        # Generate cache ID from parameters
        cache_key = f"{entity_types}_{min_occurrences}_{min_relationship_strength}"
        cache_id = hashlib.md5(cache_key.encode()).hexdigest()[:16]

        # Serialize graph
        try:
            graph_data = pickle.dumps(G, protocol=pickle.HIGHEST_PROTOCOL)
        except Exception as e:
            self.logger.error(f"Failed to pickle graph: {e}")
            return ""

        # Store in database
        cursor = self.db_conn.cursor()

        # Delete old cache with same ID if exists
        cursor.execute("DELETE FROM graph_cache WHERE cache_id = ?", (cache_id,))

        # Insert new cache
        cursor.execute("""
            INSERT INTO graph_cache
            (cache_id, graph_version, graph_data, node_count, edge_count, created_date, last_accessed)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (cache_id, 1, graph_data, G.number_of_nodes(),
              G.number_of_edges(), datetime.now().isoformat(), datetime.now().isoformat()))

        self.db_conn.commit()

        self.logger.debug(f"Cached graph: {G.number_of_nodes()} nodes, {G.number_of_edges()} edges")

        return cache_id

    def _load_cached_graph(self, cache_id: str):
        """
        Load cached graph from database.

        Args:
            cache_id: ID of the cached graph

        Returns:
            NetworkX graph if found, None otherwise
        """
        import pickle
        from datetime import datetime

        cursor = self.db_conn.cursor()
        row = cursor.execute("""
            SELECT graph_data, node_count, edge_count, created_date
            FROM graph_cache
            WHERE cache_id = ?
        """, (cache_id,)).fetchone()

        if not row:
            return None

        try:
            # Deserialize graph
            graph_data, node_count, edge_count, created_date = row
            G = pickle.loads(graph_data)

            # Update last accessed time
            cursor.execute("""
                UPDATE graph_cache SET last_accessed = ? WHERE cache_id = ?
            """, (datetime.now().isoformat(), cache_id))
            self.db_conn.commit()

            self.logger.debug(f"Loaded cached graph from {created_date}: {node_count} nodes, {edge_count} edges")

            return G

        except Exception as e:
            self.logger.warning(f"Failed to load cached graph: {e}")
            return None

    def clear_graph_cache(self, older_than_days: Optional[int] = None) -> int:
        """
        Clear cached graphs from database.

        Args:
            older_than_days: Only clear caches older than this many days (None = clear all)

        Returns:
            Number of caches cleared
        """
        from datetime import datetime, timedelta

        cursor = self.db_conn.cursor()

        if older_than_days is None:
            # Clear all caches
            cursor.execute("DELETE FROM graph_cache")
        else:
            # Clear old caches
            cutoff_date = (datetime.now() - timedelta(days=older_than_days)).isoformat()
            cursor.execute("""
                DELETE FROM graph_cache WHERE created_date < ?
            """, (cutoff_date,))

        deleted = cursor.rowcount
        self.db_conn.commit()

        self.logger.info(f"Cleared {deleted} cached graphs")
        return deleted

    # ============================================================
    # Graph Analysis Algorithms (v2.24.0 - Phase 1, Task 1.3)
    # ============================================================

    def compute_graph_metrics(self, G=None, entity_types: Optional[list[str]] = None,
                              min_occurrences: int = 2,
                              min_relationship_strength: float = 0.3,
                              store_results: bool = True) -> dict:
        """
        Compute comprehensive graph analysis metrics.

        Computes:
        - PageRank centrality (importance based on connections)
        - Betweenness centrality (bridge nodes)
        - Degree centrality (connection count)
        - Community detection (Louvain method)

        Args:
            G: Pre-built NetworkX graph (None = build new graph)
            entity_types: Filter to specific entity types
            min_occurrences: Minimum entity occurrences for graph building
            min_relationship_strength: Minimum relationship strength
            store_results: Save metrics to database

        Returns:
            {
                'pagerank': {entity: score, ...},
                'betweenness': {entity: score, ...},
                'degree': {entity: score, ...},
                'communities': {entity: community_id, ...},
                'num_communities': int,
                'graph_stats': {...}
            }
        """
        import networkx as nx
        from datetime import datetime

        # Build or use provided graph
        if G is None:
            self.logger.info("Building knowledge graph for analysis...")
            G = self.build_knowledge_graph(
                entity_types=entity_types,
                min_occurrences=min_occurrences,
                min_relationship_strength=min_relationship_strength
            )

        if G.number_of_nodes() == 0:
            self.logger.warning("Empty graph - no metrics to compute")
            return {
                'pagerank': {},
                'betweenness': {},
                'degree': {},
                'communities': {},
                'num_communities': 0,
                'graph_stats': {}
            }

        self.logger.info(f"Computing graph metrics for {G.number_of_nodes()} nodes, {G.number_of_edges()} edges")

        results = {}

        # 1. PageRank - Measures entity importance based on connections
        try:
            self.logger.debug("Computing PageRank centrality...")
            pagerank = nx.pagerank(G, weight='weight', max_iter=100, tol=1e-6)
            results['pagerank'] = pagerank
            top_pr = sorted(pagerank.items(), key=lambda x: x[1], reverse=True)[:5]
            self.logger.info(f"Top 5 PageRank entities: {[f'{e}({s:.4f})' for e, s in top_pr]}")
        except Exception as e:
            self.logger.error(f"PageRank computation failed: {e}")
            results['pagerank'] = {}

        # 2. Betweenness Centrality - Measures entities that bridge communities
        try:
            self.logger.debug("Computing betweenness centrality...")
            betweenness = nx.betweenness_centrality(G, weight='weight', normalized=True)
            results['betweenness'] = betweenness
            top_bt = sorted(betweenness.items(), key=lambda x: x[1], reverse=True)[:5]
            self.logger.info(f"Top 5 betweenness entities: {[f'{e}({s:.4f})' for e, s in top_bt]}")
        except Exception as e:
            self.logger.error(f"Betweenness computation failed: {e}")
            results['betweenness'] = {}

        # 3. Degree Centrality - Measures connection count
        try:
            self.logger.debug("Computing degree centrality...")
            degree = nx.degree_centrality(G)
            results['degree'] = degree
            top_deg = sorted(degree.items(), key=lambda x: x[1], reverse=True)[:5]
            self.logger.info(f"Top 5 degree entities: {[f'{e}({s:.4f})' for e, s in top_deg]}")
        except Exception as e:
            self.logger.error(f"Degree computation failed: {e}")
            results['degree'] = {}

        # 4. Community Detection - Louvain method
        try:
            self.logger.debug("Detecting communities (Louvain method)...")
            # Import community detection
            try:
                import community.community_louvain as community_louvain
                communities = community_louvain.best_partition(G, weight='weight')
                results['communities'] = communities
                results['num_communities'] = len(set(communities.values()))
                self.logger.info(f"Detected {results['num_communities']} communities")

                # Show community sizes
                comm_sizes = {}
                for entity, comm_id in communities.items():
                    comm_sizes[comm_id] = comm_sizes.get(comm_id, 0) + 1
                top_comms = sorted(comm_sizes.items(), key=lambda x: x[1], reverse=True)[:5]
                self.logger.info(f"Top 5 community sizes: {top_comms}")

            except ImportError:
                self.logger.warning("python-louvain not installed - using greedy modularity communities instead")
                from networkx.algorithms import community as nx_community
                communities_list = nx_community.greedy_modularity_communities(G, weight='weight')
                # Convert to dict format
                communities = {}
                for idx, comm in enumerate(communities_list):
                    for entity in comm:
                        communities[entity] = idx
                results['communities'] = communities
                results['num_communities'] = len(communities_list)
                self.logger.info(f"Detected {results['num_communities']} communities (greedy modularity)")

        except Exception as e:
            self.logger.error(f"Community detection failed: {e}")
            results['communities'] = {}
            results['num_communities'] = 0

        # Graph statistics
        results['graph_stats'] = {
            'nodes': G.number_of_nodes(),
            'edges': G.number_of_edges(),
            'density': nx.density(G) if G.number_of_nodes() > 0 else 0,
            'connected_components': nx.number_connected_components(G),
            'computed_at': datetime.now().isoformat()
        }

        # Store results in database
        if store_results and results['pagerank']:
            try:
                self._store_graph_metrics(results)
                self.logger.info("Stored graph metrics to database")
            except Exception as e:
                self.logger.error(f"Failed to store graph metrics: {e}")

        return results

    def _store_graph_metrics(self, results: dict) -> None:
        """Store computed graph metrics to database."""
        from datetime import datetime
        import hashlib

        cursor = self.db_conn.cursor()
        timestamp = datetime.now().isoformat()

        # Clear old metrics first
        cursor.execute("DELETE FROM graph_metrics")

        # Get entity types from database for all entities
        entity_types = {}
        for entity in results.get('pagerank', {}).keys():
            # Get entity type from document_entities table
            row = cursor.execute("""
                SELECT entity_type FROM document_entities
                WHERE entity_text = ?
                LIMIT 1
            """, (entity,)).fetchone()
            entity_types[entity] = row[0] if row else 'unknown'

        # Insert new metrics (one row per entity with all metrics)
        for entity in results.get('pagerank', {}).keys():
            # Generate metric_id
            metric_id = hashlib.md5(f"{entity}_{timestamp}".encode()).hexdigest()[:16]

            cursor.execute("""
                INSERT INTO graph_metrics
                (metric_id, entity_text, entity_type, pagerank, betweenness_centrality,
                 degree_centrality, community_id, computed_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                metric_id,
                entity,
                entity_types.get(entity, 'unknown'),
                results['pagerank'].get(entity),
                results['betweenness'].get(entity),
                results['degree'].get(entity),
                results['communities'].get(entity),
                timestamp
            ))

        self.db_conn.commit()
        self.logger.debug(f"Stored metrics for {len(results.get('pagerank', {}))} entities")

    def find_shortest_path(self, entity1: str, entity2: str,
                          G=None, max_path_length: int = 6,
                          store_result: bool = True) -> Optional[dict]:
        """
        Find shortest path between two entities in the knowledge graph.

        Args:
            entity1: Source entity text
            entity2: Target entity text
            G: Pre-built graph (None = build default graph)
            max_path_length: Maximum path length to search
            store_result: Save path to database

        Returns:
            {
                'path': [entity1, intermediate1, ..., entity2],
                'length': int,
                'exists': bool,
                'relationships': [rel_data, ...]
            }
            or None if no path found
        """
        import networkx as nx

        # Build graph if not provided
        if G is None:
            self.logger.debug("Building graph for path finding...")
            G = self.build_knowledge_graph(min_occurrences=2)

        # Check if both entities exist in graph
        if not G.has_node(entity1):
            self.logger.warning(f"Entity '{entity1}' not found in graph")
            return None

        if not G.has_node(entity2):
            self.logger.warning(f"Entity '{entity2}' not found in graph")
            return None

        # Find shortest path
        try:
            path = nx.shortest_path(G, source=entity1, target=entity2, weight='weight')
            path_length = len(path) - 1  # Number of edges

            if path_length > max_path_length:
                self.logger.info(f"Path found but exceeds max length ({path_length} > {max_path_length})")
                return None

            self.logger.info(f"Found path of length {path_length}: {' -> '.join(path)}")

            # Get relationship data for each edge in path
            relationships = []
            for i in range(len(path) - 1):
                e1, e2 = path[i], path[i + 1]
                edge_data = G.get_edge_data(e1, e2)
                relationships.append({
                    'from': e1,
                    'to': e2,
                    'weight': edge_data.get('weight', 0),
                    'doc_count': edge_data.get('doc_count', 0)
                })

            result = {
                'path': path,
                'length': path_length,
                'exists': True,
                'relationships': relationships
            }

            # Store in database
            if store_result:
                try:
                    self._store_graph_path(entity1, entity2, path, path_length)
                except Exception as e:
                    self.logger.error(f"Failed to store path: {e}")

            return result

        except nx.NetworkXNoPath:
            self.logger.info(f"No path exists between '{entity1}' and '{entity2}'")
            return {
                'path': [],
                'length': -1,
                'exists': False,
                'relationships': []
            }
        except Exception as e:
            self.logger.error(f"Path finding failed: {e}")
            return None

    def _store_graph_path(self, entity1: str, entity2: str, path: list, length: int) -> None:
        """Store computed shortest path to database."""
        from datetime import datetime
        import json
        import hashlib

        cursor = self.db_conn.cursor()

        # Generate path_id
        path_id = hashlib.md5(f"{entity1}_{entity2}_{datetime.now().isoformat()}".encode()).hexdigest()[:16]

        # Check if path already exists
        cursor.execute("""
            DELETE FROM graph_paths
            WHERE entity1 = ? AND entity2 = ?
        """, (entity1, entity2))

        # Insert new path
        cursor.execute("""
            INSERT INTO graph_paths
            (path_id, entity1, entity2, path_length, path_nodes, path_weight, computed_date)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            path_id,
            entity1,
            entity2,
            length,
            json.dumps(path),
            None,  # path_weight can be computed later if needed
            datetime.now().isoformat()
        ))

        self.db_conn.commit()
        self.logger.debug(f"Stored path: {entity1} -> {entity2} (length {length})")

    def get_entity_metrics(self, entity_text: str, metric_types: Optional[list[str]] = None) -> dict:
        """
        Retrieve stored graph metrics for an entity.

        Args:
            entity_text: Entity to get metrics for
            metric_types: List of metric types to retrieve (None = all)
                         Options: 'pagerank', 'betweenness', 'degree', 'community'

        Returns:
            {
                'entity': str,
                'entity_type': str,
                'metrics': {
                    'pagerank': float,
                    'betweenness': float,
                    'degree': float,
                    'community': int
                },
                'computed_date': str,
                'found': bool
            }
        """
        cursor = self.db_conn.cursor()

        # Query metrics table
        row = cursor.execute("""
            SELECT entity_type, pagerank, betweenness_centrality,
                   degree_centrality, community_id, computed_date
            FROM graph_metrics
            WHERE entity_text = ?
        """, (entity_text,)).fetchone()

        if not row:
            return {
                'entity': entity_text,
                'entity_type': None,
                'metrics': {},
                'computed_date': None,
                'found': False
            }

        entity_type, pagerank, betweenness, degree, community, computed_date = row

        # Build metrics dict (filter by metric_types if provided)
        all_metrics = {
            'pagerank': pagerank,
            'betweenness': betweenness,
            'degree': degree,
            'community': community
        }

        if metric_types:
            metrics = {k: v for k, v in all_metrics.items() if k in metric_types}
        else:
            metrics = all_metrics

        return {
            'entity': entity_text,
            'entity_type': entity_type,
            'metrics': metrics,
            'computed_date': computed_date,
            'found': True
        }

    def visualize_knowledge_graph(self, G=None,
                                  output_path: str = "knowledge_graph.html",
                                  entity_types: Optional[list[str]] = None,
                                  min_occurrences: int = 2,
                                  min_relationship_strength: float = 0.3,
                                  color_by: str = "entity_type",
                                  size_by: str = "pagerank",
                                  highlight_communities: bool = False,
                                  physics_enabled: bool = True,
                                  height: str = "800px",
                                  width: str = "100%") -> str:
        """
        Generate interactive HTML visualization of the knowledge graph using PyVis.

        Args:
            G: Pre-built NetworkX graph (None = build new graph)
            output_path: Path to save HTML file
            entity_types: Filter to specific entity types
            min_occurrences: Minimum entity occurrences for graph building
            min_relationship_strength: Minimum relationship strength
            color_by: Node coloring scheme: 'entity_type', 'community', or 'uniform'
            size_by: Node sizing metric: 'pagerank', 'degree', 'betweenness', or 'uniform'
            highlight_communities: Add community borders/grouping
            physics_enabled: Enable physics simulation for layout
            height: Visualization height (CSS format)
            width: Visualization width (CSS format)

        Returns:
            Path to generated HTML file

        Examples:
            # Basic visualization
            path = kb.visualize_knowledge_graph()

            # Color by community, size by PageRank
            path = kb.visualize_knowledge_graph(
                color_by='community',
                size_by='pagerank',
                highlight_communities=True
            )

            # Hardware entities only
            path = kb.visualize_knowledge_graph(
                entity_types=['hardware'],
                min_occurrences=3
            )
        """
        from pyvis.network import Network
        import networkx as nx
        from pathlib import Path

        # Build graph if not provided
        if G is None:
            self.logger.info("Building knowledge graph for visualization...")
            G = self.build_knowledge_graph(
                entity_types=entity_types,
                min_occurrences=min_occurrences,
                min_relationship_strength=min_relationship_strength
            )

        if G.number_of_nodes() == 0:
            self.logger.warning("Empty graph - cannot visualize")
            return ""

        self.logger.info(f"Visualizing graph: {G.number_of_nodes()} nodes, {G.number_of_edges()} edges")

        # Create PyVis network
        net = Network(height=height, width=width, directed=False, notebook=False)

        # Configure physics
        if physics_enabled:
            net.barnes_hut(gravity=-8000, central_gravity=0.3, spring_length=200, spring_strength=0.05)
        else:
            net.toggle_physics(False)

        # Get metrics if using them for sizing or coloring
        metrics = None
        if size_by in ['pagerank', 'degree', 'betweenness'] or color_by == 'community':
            # Try to load from database first
            cursor = self.db_conn.cursor()
            stored_metrics = cursor.execute("SELECT COUNT(*) FROM graph_metrics").fetchone()[0]

            if stored_metrics == 0:
                self.logger.info("Computing graph metrics for visualization...")
                metrics = self.compute_graph_metrics(G=G, store_results=False)
            else:
                # Load from database
                self.logger.info("Loading stored metrics for visualization...")
                metrics = {
                    'pagerank': {},
                    'betweenness': {},
                    'degree': {},
                    'communities': {}
                }
                rows = cursor.execute("""
                    SELECT entity_text, pagerank, betweenness_centrality,
                           degree_centrality, community_id
                    FROM graph_metrics
                """).fetchall()

                for entity, pr, betw, deg, comm in rows:
                    if entity in G.nodes():
                        metrics['pagerank'][entity] = pr
                        metrics['betweenness'][entity] = betw or 0
                        metrics['degree'][entity] = deg
                        metrics['communities'][entity] = comm

        # Define color scheme for entity types
        entity_type_colors = {
            'hardware': '#FF6B6B',      # Red
            'memory_address': '#4ECDC4', # Teal
            'instruction': '#45B7D1',    # Blue
            'person': '#FFA07A',         # Light salmon
            'company': '#98D8C8',        # Mint
            'product': '#F7DC6F',        # Yellow
            'concept': '#BB8FCE',        # Purple
            'unknown': '#95A5A6'         # Gray
        }

        # Generate community colors if needed
        community_colors = {}
        if color_by == 'community' and metrics:
            import random
            random.seed(42)  # Consistent colors
            unique_communities = set(metrics['communities'].values()) if metrics['communities'] else set()
            colors = ['#%06x' % random.randint(0, 0xFFFFFF) for _ in range(len(unique_communities))]
            community_colors = dict(zip(sorted(unique_communities), colors))

        # Calculate node sizes if using metric
        node_sizes = {}
        if size_by != 'uniform' and metrics:
            metric_values = metrics.get(size_by, {})
            if metric_values:
                max_val = max(metric_values.values()) if metric_values else 1
                min_val = min(metric_values.values()) if metric_values else 0

                for node in G.nodes():
                    val = metric_values.get(node, 0)
                    # Scale to size range 10-50
                    if max_val > min_val:
                        normalized = (val - min_val) / (max_val - min_val)
                        node_sizes[node] = 10 + (normalized * 40)
                    else:
                        node_sizes[node] = 20

        # Add nodes to PyVis network
        for node in G.nodes(data=True):
            node_id = node[0]
            node_data = node[1]

            # Determine color
            if color_by == 'entity_type':
                entity_type = node_data.get('type', 'unknown')
                color = entity_type_colors.get(entity_type, '#95A5A6')
            elif color_by == 'community' and metrics:
                comm_id = metrics['communities'].get(node_id, 0)
                color = community_colors.get(comm_id, '#95A5A6')
            else:  # uniform
                color = '#4ECDC4'

            # Determine size
            size = node_sizes.get(node_id, 20)

            # Create hover title with entity info
            title = f"<b>{node_id}</b><br>"
            title += f"Type: {node_data.get('type', 'unknown')}<br>"
            title += f"Occurrences: {node_data.get('occurrences', 0)}<br>"

            if metrics:
                if node_id in metrics.get('pagerank', {}):
                    title += f"PageRank: {metrics['pagerank'][node_id]:.6f}<br>"
                if node_id in metrics.get('betweenness', {}):
                    title += f"Betweenness: {metrics['betweenness'][node_id]:.6f}<br>"
                if node_id in metrics.get('degree', {}):
                    title += f"Degree: {metrics['degree'][node_id]:.6f}<br>"
                if node_id in metrics.get('communities', {}):
                    title += f"Community: {metrics['communities'][node_id]}"

            net.add_node(node_id, label=node_id, color=color, size=size, title=title)

        # Add edges to PyVis network
        for edge in G.edges(data=True):
            source = edge[0]
            target = edge[1]
            edge_data = edge[2]

            weight = edge_data.get('weight', 0.5)
            doc_count = edge_data.get('doc_count', 0)

            # Edge width based on weight
            width = 1 + (weight * 3)

            # Edge title
            title = f"Strength: {weight:.3f}<br>Documents: {doc_count}"

            net.add_edge(source, target, value=width, title=title)

        # Generate HTML
        output_path_obj = Path(output_path)
        if not output_path_obj.is_absolute():
            # Save to data directory if relative path
            output_path_obj = Path(self.data_dir) / output_path

        net.save_graph(str(output_path_obj))

        self.logger.info(f"Visualization saved to: {output_path_obj}")

        return str(output_path_obj)

    # ============================================================
    # Topic Modeling Methods (v2.24.0 - Phase 2, Task 2.2)
    # ============================================================

    def _prepare_topic_model_corpus(self, min_df: int = 2,
                                    max_df: float = 0.8,
                                    max_features: int = 1000) -> tuple:
        """
        Prepare document corpus for topic modeling.

        Args:
            min_df: Minimum document frequency (ignore terms in fewer docs)
            max_df: Maximum document frequency (ignore terms in more than this fraction)
            max_features: Maximum number of features to extract

        Returns:
            (doc_ids, vectorizer, document_term_matrix, feature_names)

        Examples:
            doc_ids, vectorizer, dtm, features = kb._prepare_topic_model_corpus()
        """
        from sklearn.feature_extraction.text import CountVectorizer

        self.logger.info("Preparing corpus for topic modeling...")

        # Get all document texts
        docs = []
        doc_ids = []

        for doc_id, doc in self.documents.items():
            # Get all chunks for this document
            chunks = self._get_chunks_db(doc_id)
            if chunks:
                # Combine all chunks into full document text
                full_text = " ".join(chunk.content for chunk in chunks)
                docs.append(full_text)
                doc_ids.append(doc_id)

        if not docs:
            self.logger.warning("No documents found for topic modeling")
            return [], None, None, []

        self.logger.info(f"Preparing corpus from {len(docs)} documents...")

        # Create CountVectorizer for LDA (LDA works better with raw counts than TF-IDF)
        vectorizer = CountVectorizer(
            max_df=max_df,
            min_df=min_df,
            stop_words='english',
            max_features=max_features,
            ngram_range=(1, 2),  # Unigrams and bigrams
            lowercase=True,
            token_pattern=r'\b[a-zA-Z]{3,}\b'  # Words with 3+ letters
        )

        # Transform documents to document-term matrix
        doc_term_matrix = vectorizer.fit_transform(docs)
        feature_names = vectorizer.get_feature_names_out()

        self.logger.info(f"Corpus prepared: {doc_term_matrix.shape[0]} docs, {doc_term_matrix.shape[1]} terms")

        return doc_ids, vectorizer, doc_term_matrix, feature_names

    def train_lda_model(self, num_topics: int = 10,
                       max_iter: int = 100,
                       random_state: int = 42,
                       min_df: int = 2,
                       max_df: float = 0.8,
                       max_features: int = 1000,
                       store_results: bool = True) -> dict:
        """
        Train Latent Dirichlet Allocation (LDA) topic model.

        LDA discovers hidden topics in document collections by finding
        patterns of word co-occurrence. Each document is modeled as a
        mixture of topics, and each topic is modeled as a distribution
        over words.

        Args:
            num_topics: Number of topics to discover
            max_iter: Maximum iterations for LDA training
            random_state: Random seed for reproducibility
            min_df: Minimum document frequency for vocabulary
            max_df: Maximum document frequency for vocabulary
            max_features: Maximum vocabulary size
            store_results: Save topics and assignments to database

        Returns:
            {
                'model_type': 'lda',
                'num_topics': int,
                'topics': [{
                    'topic_id': str,
                    'topic_number': int,
                    'top_words': [str, ...],
                    'word_weights': {word: weight, ...},
                    'coherence_score': float
                }, ...],
                'document_topics': [{
                    'doc_id': str,
                    'topic_assignments': [(topic_num, probability), ...]
                }, ...],
                'perplexity': float,
                'training_time': float
            }

        Examples:
            # Train 10-topic model
            results = kb.train_lda_model(num_topics=10)

            # Train model without storing to database
            results = kb.train_lda_model(num_topics=15, store_results=False)
        """
        from sklearn.decomposition import LatentDirichletAllocation
        import time
        from datetime import datetime

        self.logger.info(f"Training LDA model with {num_topics} topics...")
        start_time = time.time()

        # Prepare corpus
        doc_ids, vectorizer, doc_term_matrix, feature_names = self._prepare_topic_model_corpus(
            min_df=min_df,
            max_df=max_df,
            max_features=max_features
        )

        if doc_term_matrix is None:
            return {
                'model_type': 'lda',
                'num_topics': 0,
                'topics': [],
                'document_topics': [],
                'error': 'No documents available for topic modeling'
            }

        # Train LDA model
        self.logger.info("Training LDA model...")
        lda = LatentDirichletAllocation(
            n_components=num_topics,
            max_iter=max_iter,
            learning_method='online',
            learning_offset=50.0,
            random_state=random_state,
            n_jobs=-1  # Use all CPUs
        )

        lda.fit(doc_term_matrix)

        training_time = time.time() - start_time
        self.logger.info(f"LDA training completed in {training_time:.2f} seconds")

        # Calculate perplexity (lower is better)
        perplexity = lda.perplexity(doc_term_matrix)
        self.logger.info(f"Model perplexity: {perplexity:.2f}")

        # Extract topics
        topics = []
        n_top_words = 10

        for topic_idx, topic in enumerate(lda.components_):
            # Get top words for this topic
            top_indices = topic.argsort()[-n_top_words:][::-1]
            top_words = [feature_names[i] for i in top_indices]
            top_weights = [float(topic[i]) for i in top_indices]

            # Create word weights dict
            word_weights = dict(zip(top_words, top_weights))

            topic_info = {
                'topic_id': f"lda_topic_{topic_idx}",
                'topic_number': topic_idx,
                'top_words': top_words,
                'word_weights': word_weights,
                'coherence_score': None  # Can be computed separately if needed
            }

            topics.append(topic_info)

            self.logger.debug(f"Topic {topic_idx}: {', '.join(top_words[:5])}")

        # Get document-topic distributions
        doc_topic_dist = lda.transform(doc_term_matrix)

        document_topics = []
        for doc_idx, doc_id in enumerate(doc_ids):
            # Get topic probabilities for this document
            topic_probs = doc_topic_dist[doc_idx]

            # Get all topics with their probabilities
            topic_assignments = [
                (topic_num, float(prob))
                for topic_num, prob in enumerate(topic_probs)
            ]

            # Sort by probability descending
            topic_assignments.sort(key=lambda x: x[1], reverse=True)

            document_topics.append({
                'doc_id': doc_id,
                'topic_assignments': topic_assignments
            })

        # Build results
        results = {
            'model_type': 'lda',
            'num_topics': num_topics,
            'topics': topics,
            'document_topics': document_topics,
            'perplexity': perplexity,
            'training_time': training_time,
            'vocabulary_size': len(feature_names),
            'num_documents': len(doc_ids)
        }

        # Store to database if requested
        if store_results:
            self.logger.info("Storing LDA results to database...")
            self._store_topics_to_db(topics, 'lda')
            self._store_document_topics(document_topics, 'lda')
            self.logger.info("LDA results stored to database")

        return results

    def _store_topics_to_db(self, topics: list, model_type: str) -> None:
        """Store topics to database."""
        import json
        from datetime import datetime
        import hashlib

        cursor = self.db_conn.cursor()
        timestamp = datetime.now().isoformat()

        # Clear existing topics for this model type
        cursor.execute("DELETE FROM topics WHERE model_type = ?", (model_type,))

        # Insert new topics
        for topic in topics:
            topic_id = topic['topic_id']

            cursor.execute("""
                INSERT INTO topics
                (topic_id, model_type, topic_number, top_words, word_weights,
                 num_documents, coherence_score, created_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                topic_id,
                model_type,
                topic['topic_number'],
                json.dumps(topic['top_words']),
                json.dumps(topic['word_weights']),
                0,  # Will be updated when document assignments are stored
                topic.get('coherence_score'),
                timestamp
            ))

        self.db_conn.commit()
        self.logger.debug(f"Stored {len(topics)} {model_type} topics to database")

    def _store_document_topics(self, document_topics: list, model_type: str) -> None:
        """Store document-topic assignments to database."""
        from datetime import datetime
        import hashlib

        cursor = self.db_conn.cursor()
        timestamp = datetime.now().isoformat()

        # Clear existing assignments for this model type
        cursor.execute("DELETE FROM document_topics WHERE model_type = ?", (model_type,))

        # Store new assignments
        for doc_topic in document_topics:
            doc_id = doc_topic['doc_id']

            # Store each topic assignment
            for topic_num, probability in doc_topic['topic_assignments']:
                # Generate assignment ID
                assignment_id = hashlib.md5(
                    f"{doc_id}_{model_type}_{topic_num}".encode()
                ).hexdigest()[:16]

                # Get topic_id
                topic_id = f"{model_type}_topic_{topic_num}"

                cursor.execute("""
                    INSERT INTO document_topics
                    (assignment_id, doc_id, topic_id, probability, model_type, assigned_date)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    assignment_id,
                    doc_id,
                    topic_id,
                    probability,
                    model_type,
                    timestamp
                ))

        self.db_conn.commit()

        # Update document counts in topics table
        cursor.execute("""
            UPDATE topics
            SET num_documents = (
                SELECT COUNT(DISTINCT doc_id)
                FROM document_topics
                WHERE document_topics.topic_id = topics.topic_id
            )
            WHERE model_type = ?
        """, (model_type,))

        self.db_conn.commit()
        self.logger.debug(f"Stored {len(document_topics)} document-topic assignments to database")

    def train_nmf_model(self, num_topics: int = 10,
                       max_iter: int = 200,
                       random_state: int = 42,
                       min_df: int = 2,
                       max_df: float = 0.8,
                       max_features: int = 1000,
                       store_results: bool = True) -> dict:
        """
        Train Non-negative Matrix Factorization (NMF) topic model.

        NMF often produces more coherent topics than LDA and is faster.
        Uses TF-IDF representation which works better for NMF.

        Args:
            num_topics: Number of topics to discover
            max_iter: Maximum iterations for NMF
            random_state: Random seed for reproducibility
            min_df: Minimum document frequency for terms
            max_df: Maximum document frequency for terms
            max_features: Maximum vocabulary size
            store_results: Whether to store results to database

        Returns:
            {
                'model_type': 'nmf',
                'num_topics': int,
                'topics': list of topic dicts,
                'document_topics': list of document-topic assignments,
                'reconstruction_error': float,
                'training_time': float,
                'vocabulary_size': int,
                'num_documents': int
            }

        Examples:
            >>> kb = KnowledgeBase()
            >>> results = kb.train_nmf_model(num_topics=5, max_iter=200)
            >>> print(f"Found {results['num_topics']} topics")
            >>> print(f"Reconstruction error: {results['reconstruction_error']:.2f}")
        """
        from sklearn.decomposition import NMF
        from sklearn.feature_extraction.text import TfidfVectorizer
        import time

        self.logger.info(f"Training NMF model with {num_topics} topics...")

        # Get all document texts
        self.logger.info("Preparing corpus from documents...")
        docs = []
        doc_ids = []
        for doc_id, doc in self.documents.items():
            chunks = self._get_chunks_db(doc_id)
            if chunks:
                full_text = " ".join(chunk.content for chunk in chunks)
                docs.append(full_text)
                doc_ids.append(doc_id)

        if not docs:
            raise ValueError("No documents available for topic modeling")

        # Create TF-IDF vectorizer (better for NMF than raw counts)
        vectorizer = TfidfVectorizer(
            max_df=max_df,
            min_df=min_df,
            stop_words='english',
            max_features=max_features,
            ngram_range=(1, 2),
            lowercase=True,
            token_pattern=r'\b[a-zA-Z]{3,}\b'
        )

        tfidf_matrix = vectorizer.fit_transform(docs)
        feature_names = vectorizer.get_feature_names_out()

        self.logger.info(f"Corpus prepared: {len(docs)} docs, {len(feature_names)} terms")

        # Train NMF model
        self.logger.info("Training NMF model...")
        start_time = time.time()

        nmf = NMF(
            n_components=num_topics,
            max_iter=max_iter,
            random_state=random_state,
            init='nndsvda',  # Better initialization for sparse matrices
            solver='mu',     # Multiplicative update solver
            beta_loss='frobenius',
            alpha_W=0.1,     # L1 regularization for topic-word distribution
            alpha_H=0.1,     # L1 regularization for document-topic distribution
            l1_ratio=0.5
        )

        doc_topic_dist = nmf.fit_transform(tfidf_matrix)
        training_time = time.time() - start_time

        self.logger.info(f"NMF training completed in {training_time:.2f} seconds")

        # Calculate reconstruction error
        reconstruction_error = nmf.reconstruction_err_
        self.logger.info(f"Reconstruction error: {reconstruction_error:.2f}")

        # Extract topics (top 10 words per topic)
        topics = []
        for topic_idx, topic in enumerate(nmf.components_):
            # Get top 10 words for this topic
            top_indices = topic.argsort()[-10:][::-1]
            top_words = [feature_names[i] for i in top_indices]
            top_weights = [float(topic[i]) for i in top_indices]
            word_weights = dict(zip(top_words, top_weights))

            topics.append({
                'topic_id': f"nmf_topic_{topic_idx}",
                'topic_number': topic_idx,
                'top_words': top_words,
                'word_weights': word_weights,
                'coherence_score': None  # Can be calculated separately
            })

        # Get document-topic distributions
        document_topics = []
        for doc_idx, doc_id in enumerate(doc_ids):
            topic_probs = doc_topic_dist[doc_idx]
            # Normalize to probabilities (NMF outputs can be > 1)
            topic_probs = topic_probs / topic_probs.sum() if topic_probs.sum() > 0 else topic_probs

            topic_assignments = [
                (topic_num, float(prob))
                for topic_num, prob in enumerate(topic_probs)
            ]
            # Sort by probability descending
            topic_assignments.sort(key=lambda x: x[1], reverse=True)

            document_topics.append({
                'doc_id': doc_id,
                'topic_assignments': topic_assignments
            })

        # Store results to database if requested
        if store_results:
            self.logger.info("Storing NMF results to database...")
            self._store_topics_to_db(topics, 'nmf')
            self._store_document_topics(document_topics, 'nmf')
            self.logger.info("NMF results stored to database")

        return {
            'model_type': 'nmf',
            'num_topics': num_topics,
            'topics': topics,
            'document_topics': document_topics,
            'reconstruction_error': reconstruction_error,
            'training_time': training_time,
            'vocabulary_size': len(feature_names),
            'num_documents': len(doc_ids)
        }

    def train_bertopic_model(self, num_topics: int = 10,
                            min_topic_size: int = 5,
                            store_results: bool = True) -> dict:
        """
        Train BERTopic model using embeddings + UMAP + HDBSCAN.

        BERTopic is a state-of-the-art topic modeling technique that uses:
        1. Document embeddings (from sentence transformers)
        2. UMAP dimensionality reduction
        3. HDBSCAN clustering
        4. c-TF-IDF for topic representation

        Args:
            num_topics: Target number of topics (actual may vary with HDBSCAN)
            min_topic_size: Minimum cluster size for HDBSCAN
            store_results: Whether to store results to database

        Returns:
            {
                'model_type': 'bertopic',
                'num_topics': int,
                'topics': list of topic dicts,
                'document_topics': list of document-topic assignments,
                'training_time': float,
                'num_documents': int,
                'outliers': int (documents in topic -1)
            }

        Examples:
            >>> kb = KnowledgeBase()
            >>> results = kb.train_bertopic_model(num_topics=5, min_topic_size=3)
            >>> print(f"Found {results['num_topics']} topics")
            >>> print(f"Outliers: {results['outliers']} documents")
        """
        from bertopic import BERTopic
        from umap import UMAP
        from hdbscan import HDBSCAN
        import numpy as np
        import time

        self.logger.info(f"Training BERTopic model (target: {num_topics} topics)...")

        # Get documents and generate embeddings
        self.logger.info("Collecting documents and generating embeddings...")
        doc_ids = []
        documents = []

        for doc_id, doc in self.documents.items():
            chunks = self._get_chunks_db(doc_id)
            if chunks:
                # Use full document text
                full_text = " ".join(chunk.content for chunk in chunks)
                documents.append(full_text)
                doc_ids.append(doc_id)

        if not documents:
            raise ValueError("No documents available for topic modeling")

        # Generate embeddings for all documents
        self.logger.info(f"Generating embeddings for {len(documents)} documents...")
        if not self._embeddings_loaded:
            self._ensure_embeddings_loaded()

        embeddings = self.embeddings_model.encode(documents, show_progress_bar=False)
        embeddings = np.array(embeddings)
        self.logger.info(f"Generated embeddings: {embeddings.shape}")

        # Configure UMAP for dimensionality reduction
        umap_model = UMAP(
            n_neighbors=15,
            n_components=5,
            min_dist=0.0,
            metric='cosine',
            random_state=42
        )

        # Configure HDBSCAN for clustering
        hdbscan_model = HDBSCAN(
            min_cluster_size=min_topic_size,
            metric='euclidean',
            cluster_selection_method='eom',
            prediction_data=True
        )

        # Create BERTopic model
        self.logger.info("Training BERTopic model...")
        start_time = time.time()

        topic_model = BERTopic(
            umap_model=umap_model,
            hdbscan_model=hdbscan_model,
            nr_topics=num_topics if num_topics > 0 else 'auto',
            calculate_probabilities=True,
            verbose=False
        )

        # Train model
        topics, probs = topic_model.fit_transform(documents, embeddings)
        training_time = time.time() - start_time

        self.logger.info(f"BERTopic training completed in {training_time:.2f} seconds")

        # Get topic information
        topic_info = topic_model.get_topic_info()
        num_actual_topics = len(topic_info) - 1  # Exclude outlier topic (-1)

        self.logger.info(f"Found {num_actual_topics} topics (excluding outliers)")

        # Count outliers (topic -1)
        outliers = sum(1 for t in topics if t == -1)
        self.logger.info(f"Outliers: {outliers} documents")

        # Extract topics (excluding outlier topic -1)
        extracted_topics = []
        for topic_num in topic_info['Topic'].tolist():
            if topic_num == -1:
                continue  # Skip outlier topic

            topic_words = topic_model.get_topic(topic_num)
            if topic_words:
                top_words = [word for word, _ in topic_words[:10]]
                word_weights = {word: float(score) for word, score in topic_words[:10]}

                extracted_topics.append({
                    'topic_id': f"bertopic_topic_{topic_num}",
                    'topic_number': topic_num,
                    'top_words': top_words,
                    'word_weights': word_weights,
                    'coherence_score': None
                })

        # Get document-topic assignments
        document_topics = []
        for doc_idx, doc_id in enumerate(doc_ids):
            topic_num = topics[doc_idx]

            # Get probability distribution if available
            if probs is not None and doc_idx < len(probs):
                topic_probs = probs[doc_idx]
                # Create assignments for all topics with their probabilities
                topic_assignments = []
                for i, prob in enumerate(topic_probs):
                    if i - 1 >= 0:  # Adjust for BERTopic indexing
                        topic_assignments.append((i - 1, float(prob)))
                topic_assignments.sort(key=lambda x: x[1], reverse=True)
            else:
                # Fallback: single topic assignment with probability 1.0
                if topic_num != -1:
                    topic_assignments = [(topic_num, 1.0)]
                else:
                    topic_assignments = [(-1, 1.0)]

            document_topics.append({
                'doc_id': doc_id,
                'topic_assignments': topic_assignments
            })

        # Store results to database if requested
        if store_results and extracted_topics:
            self.logger.info("Storing BERTopic results to database...")
            # Filter out outlier assignments (topic -1)
            valid_document_topics = []
            for dt in document_topics:
                valid_assignments = [(t, p) for t, p in dt['topic_assignments'] if t != -1]
                if valid_assignments:
                    valid_document_topics.append({
                        'doc_id': dt['doc_id'],
                        'topic_assignments': valid_assignments
                    })

            self._store_topics_to_db(extracted_topics, 'bertopic')
            self._store_document_topics(valid_document_topics, 'bertopic')
            self.logger.info("BERTopic results stored to database")

        return {
            'model_type': 'bertopic',
            'num_topics': num_actual_topics,
            'topics': extracted_topics,
            'document_topics': document_topics,
            'training_time': training_time,
            'num_documents': len(doc_ids),
            'outliers': outliers
        }

    def compare_topic_models(self) -> dict:
        """
        Compare all topic models (LDA, NMF, BERTopic) stored in the database.

        Provides a comprehensive comparison of different topic modeling approaches
        to help select the best model for the corpus.

        Returns:
            {
                'models': {
                    'lda': model_stats,
                    'nmf': model_stats,
                    'bertopic': model_stats
                },
                'comparison': {
                    'fastest': str,
                    'most_topics': str,
                    'best_coverage': str
                }
            }

        Examples:
            >>> kb = KnowledgeBase()
            >>> kb.train_lda_model(num_topics=5)
            >>> kb.train_nmf_model(num_topics=5)
            >>> kb.train_bertopic_model(num_topics=5)
            >>> comparison = kb.compare_topic_models()
            >>> print(f"Fastest model: {comparison['comparison']['fastest']}")
        """
        cursor = self.db_conn.cursor()

        models = {}
        model_types = ['lda', 'nmf', 'bertopic']

        for model_type in model_types:
            # Get topics for this model
            topics = cursor.execute("""
                SELECT COUNT(*), AVG(num_documents)
                FROM topics
                WHERE model_type = ?
            """, (model_type,)).fetchone()

            if topics and topics[0] > 0:
                num_topics = topics[0]
                avg_docs_per_topic = topics[1] or 0

                # Get document coverage
                doc_count = cursor.execute("""
                    SELECT COUNT(DISTINCT doc_id)
                    FROM document_topics
                    WHERE model_type = ?
                """, (model_type,)).fetchone()[0]

                # Get top topics by document count
                top_topics = cursor.execute("""
                    SELECT topic_number, num_documents, top_words
                    FROM topics
                    WHERE model_type = ?
                    ORDER BY num_documents DESC
                    LIMIT 5
                """, (model_type,)).fetchall()

                # Calculate topic diversity (unique words across all topics)
                all_topics_data = cursor.execute("""
                    SELECT top_words
                    FROM topics
                    WHERE model_type = ?
                """, (model_type,)).fetchall()

                unique_words = set()
                for (top_words_json,) in all_topics_data:
                    import json
                    words = json.loads(top_words_json)
                    unique_words.update(words)

                models[model_type] = {
                    'num_topics': num_topics,
                    'documents_covered': doc_count,
                    'avg_docs_per_topic': round(avg_docs_per_topic, 1),
                    'vocabulary_diversity': len(unique_words),
                    'top_topics': [
                        {
                            'topic_number': t[0],
                            'num_documents': t[1],
                            'top_words': json.loads(t[2])[:5]  # First 5 words
                        }
                        for t in top_topics
                    ]
                }
            else:
                models[model_type] = None

        # Calculate comparisons
        comparison = {}

        # Find model with most topics
        valid_models = {k: v for k, v in models.items() if v is not None}
        if valid_models:
            most_topics = max(valid_models.items(),
                            key=lambda x: x[1]['num_topics'])
            comparison['most_topics'] = {
                'model': most_topics[0],
                'count': most_topics[1]['num_topics']
            }

            # Find model with best coverage
            best_coverage = max(valid_models.items(),
                              key=lambda x: x[1]['documents_covered'])
            comparison['best_coverage'] = {
                'model': best_coverage[0],
                'documents': best_coverage[1]['documents_covered']
            }

            # Find model with highest vocabulary diversity
            most_diverse = max(valid_models.items(),
                             key=lambda x: x[1]['vocabulary_diversity'])
            comparison['most_diverse_vocabulary'] = {
                'model': most_diverse[0],
                'unique_words': most_diverse[1]['vocabulary_diversity']
            }

            # Summary recommendation
            if len(valid_models) >= 2:
                # Simple heuristic: prioritize coverage and diversity
                scores = {}
                for model, stats in valid_models.items():
                    score = (stats['documents_covered'] * 0.4 +
                            stats['vocabulary_diversity'] * 0.3 +
                            stats['num_topics'] * 0.3)
                    scores[model] = score

                recommended = max(scores.items(), key=lambda x: x[1])
                comparison['recommended'] = {
                    'model': recommended[0],
                    'score': round(recommended[1], 2),
                    'reason': f"Best balance of coverage, diversity, and topic count"
                }

        return {
            'models': models,
            'comparison': comparison,
            'total_models': len(valid_models)
        }

    def _store_clusters_to_db(self, clusters: list, algorithm: str) -> None:
        """
        Store clusters to database.

        Args:
            clusters: List of cluster dicts with cluster info
            algorithm: Clustering algorithm name (kmeans, dbscan, hdbscan)
        """
        import json
        import pickle

        cursor = self.db_conn.cursor()

        # Clear existing clusters for this algorithm
        cursor.execute("DELETE FROM clusters WHERE algorithm = ?", (algorithm,))

        for cluster in clusters:
            # Serialize centroid vector if present
            centroid_blob = None
            if cluster.get('centroid') is not None:
                centroid_blob = pickle.dumps(cluster['centroid'])

            cursor.execute("""
                INSERT INTO clusters
                (cluster_id, algorithm, cluster_number, centroid_vector,
                 num_documents, representative_docs, top_terms, silhouette_score, created_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                cluster['cluster_id'],
                algorithm,
                cluster['cluster_number'],
                centroid_blob,
                cluster.get('num_documents', 0),
                json.dumps(cluster.get('representative_docs', [])),
                json.dumps(cluster.get('top_terms', [])),
                cluster.get('silhouette_score'),
                datetime.now().isoformat()
            ))

        self.db_conn.commit()
        self.logger.debug(f"Stored {len(clusters)} clusters to database")

    def _store_document_clusters(self, document_clusters: list, algorithm: str) -> None:
        """
        Store document-cluster assignments to database.

        Args:
            document_clusters: List of {doc_id, cluster_assignments} dicts
            algorithm: Clustering algorithm name
        """
        cursor = self.db_conn.cursor()

        # Clear existing assignments for this algorithm
        cursor.execute("DELETE FROM document_clusters WHERE algorithm = ?", (algorithm,))

        for doc_cluster in document_clusters:
            doc_id = doc_cluster['doc_id']
            cluster_num = doc_cluster['cluster_number']
            distance = doc_cluster.get('distance')

            assignment_id = hashlib.md5(
                f"{doc_id}_{algorithm}_{cluster_num}".encode()
            ).hexdigest()[:16]
            cluster_id = f"{algorithm}_cluster_{cluster_num}"

            cursor.execute("""
                INSERT INTO document_clusters
                (assignment_id, doc_id, cluster_id, distance, algorithm, assigned_date)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (assignment_id, doc_id, cluster_id, distance,
                  algorithm, datetime.now().isoformat()))

        self.db_conn.commit()

        # Update document counts in clusters table
        cursor.execute("""
            UPDATE clusters
            SET num_documents = (
                SELECT COUNT(DISTINCT doc_id)
                FROM document_clusters
                WHERE document_clusters.cluster_id = clusters.cluster_id
            )
            WHERE algorithm = ?
        """, (algorithm,))
        self.db_conn.commit()
        self.logger.debug(f"Stored {len(document_clusters)} document-cluster assignments to database")

    def cluster_documents_kmeans(self, num_clusters: int = 10,
                                 store_results: bool = True) -> dict:
        """
        Cluster documents using K-Means on embeddings.

        K-Means is a centroid-based clustering algorithm that partitions documents
        into K clusters based on embedding similarity.

        Args:
            num_clusters: Number of clusters (K)
            store_results: Whether to store results to database

        Returns:
            {
                'algorithm': 'kmeans',
                'num_clusters': int,
                'clusters': list of cluster dicts,
                'document_clusters': list of assignments,
                'silhouette_score': float,
                'training_time': float,
                'num_documents': int
            }

        Examples:
            >>> kb = KnowledgeBase()
            >>> results = kb.cluster_documents_kmeans(num_clusters=5)
            >>> print(f"Silhouette score: {results['silhouette_score']:.3f}")
        """
        from sklearn.cluster import KMeans
        from sklearn.metrics import silhouette_score
        import numpy as np
        import time

        self.logger.info(f"Clustering documents with K-Means (K={num_clusters})...")

        # Get documents and generate embeddings
        self.logger.info("Collecting documents and generating embeddings...")
        doc_ids = []
        documents = []

        for doc_id, doc in self.documents.items():
            chunks = self._get_chunks_db(doc_id)
            if chunks:
                full_text = " ".join(chunk.content for chunk in chunks)
                documents.append(full_text)
                doc_ids.append(doc_id)

        if not documents:
            raise ValueError("No documents available for clustering")

        # Generate embeddings
        if not self._embeddings_loaded:
            self._ensure_embeddings_loaded()

        self.logger.info(f"Generating embeddings for {len(documents)} documents...")
        embeddings = self.embeddings_model.encode(documents, show_progress_bar=False)
        embeddings = np.array(embeddings)
        self.logger.info(f"Embeddings shape: {embeddings.shape}")

        # Train K-Means
        self.logger.info("Training K-Means model...")
        start_time = time.time()

        kmeans = KMeans(
            n_clusters=num_clusters,
            random_state=42,
            n_init=10,
            max_iter=300
        )
        labels = kmeans.fit_predict(embeddings)
        training_time = time.time() - start_time

        self.logger.info(f"K-Means training completed in {training_time:.2f} seconds")

        # Calculate silhouette score
        silhouette = silhouette_score(embeddings, labels)
        self.logger.info(f"Silhouette score: {silhouette:.3f}")

        # Extract clusters
        clusters = []
        document_clusters = []

        for cluster_num in range(num_clusters):
            # Get documents in this cluster
            cluster_doc_indices = [i for i, label in enumerate(labels) if label == cluster_num]
            cluster_doc_ids = [doc_ids[i] for i in cluster_doc_indices]

            # Get cluster centroid
            centroid = kmeans.cluster_centers_[cluster_num]

            # Get representative documents (closest to centroid)
            if cluster_doc_indices:
                cluster_embeddings = embeddings[cluster_doc_indices]
                distances = np.linalg.norm(cluster_embeddings - centroid, axis=1)
                top_indices = distances.argsort()[:5]  # Top 5 closest
                representative_docs = [cluster_doc_ids[i] for i in top_indices]
            else:
                representative_docs = []

            # Calculate top terms for this cluster
            cluster_texts = [documents[i] for i in cluster_doc_indices]
            top_terms = self._extract_top_terms_from_texts(cluster_texts, top_n=10)

            clusters.append({
                'cluster_id': f"kmeans_cluster_{cluster_num}",
                'cluster_number': cluster_num,
                'centroid': centroid,
                'num_documents': len(cluster_doc_ids),
                'representative_docs': representative_docs,
                'top_terms': top_terms,
                'silhouette_score': silhouette
            })

            # Create document-cluster assignments
            for doc_idx in cluster_doc_indices:
                doc_id = doc_ids[doc_idx]
                distance = float(np.linalg.norm(embeddings[doc_idx] - centroid))

                document_clusters.append({
                    'doc_id': doc_id,
                    'cluster_number': cluster_num,
                    'distance': distance
                })

        # Store results to database if requested
        if store_results:
            self.logger.info("Storing K-Means results to database...")
            self._store_clusters_to_db(clusters, 'kmeans')
            self._store_document_clusters(document_clusters, 'kmeans')
            self.logger.info("K-Means results stored to database")

        return {
            'algorithm': 'kmeans',
            'num_clusters': num_clusters,
            'clusters': clusters,
            'document_clusters': document_clusters,
            'silhouette_score': silhouette,
            'training_time': training_time,
            'num_documents': len(doc_ids)
        }

    def _extract_top_terms_from_texts(self, texts: list, top_n: int = 10) -> list:
        """
        Extract top N terms from a list of texts using TF-IDF.

        Args:
            texts: List of text strings
            top_n: Number of top terms to extract

        Returns:
            List of top terms
        """
        if not texts:
            return []

        from sklearn.feature_extraction.text import TfidfVectorizer

        try:
            vectorizer = TfidfVectorizer(
                max_features=top_n,
                stop_words='english',
                ngram_range=(1, 2),
                max_df=0.8,
                min_df=1
            )
            vectorizer.fit(texts)
            return list(vectorizer.get_feature_names_out())
        except Exception as e:
            self.logger.warning(f"Could not extract top terms: {e}")
            return []

    def cluster_documents_dbscan(self, eps: float = 0.5,
                                 min_samples: int = 3,
                                 store_results: bool = True) -> dict:
        """
        Cluster documents using DBSCAN (Density-Based Spatial Clustering).

        DBSCAN finds arbitrary-shaped clusters and identifies outliers.
        Does not require specifying the number of clusters.

        Args:
            eps: Maximum distance between two samples to be considered neighbors
            min_samples: Minimum samples in a neighborhood for a core point
            store_results: Whether to store results to database

        Returns:
            {
                'algorithm': 'dbscan',
                'num_clusters': int (excluding noise),
                'clusters': list of cluster dicts,
                'document_clusters': list of assignments,
                'num_outliers': int (noise points),
                'training_time': float,
                'num_documents': int
            }

        Examples:
            >>> kb = KnowledgeBase()
            >>> results = kb.cluster_documents_dbscan(eps=0.5, min_samples=3)
            >>> print(f"Found {results['num_clusters']} clusters")
            >>> print(f"Outliers: {results['num_outliers']}")
        """
        from sklearn.cluster import DBSCAN
        import numpy as np
        import time

        self.logger.info(f"Clustering documents with DBSCAN (eps={eps}, min_samples={min_samples})...")

        # Get documents and generate embeddings
        self.logger.info("Collecting documents and generating embeddings...")
        doc_ids = []
        documents = []

        for doc_id, doc in self.documents.items():
            chunks = self._get_chunks_db(doc_id)
            if chunks:
                full_text = " ".join(chunk.content for chunk in chunks)
                documents.append(full_text)
                doc_ids.append(doc_id)

        if not documents:
            raise ValueError("No documents available for clustering")

        # Generate embeddings
        if not self._embeddings_loaded:
            self._ensure_embeddings_loaded()

        self.logger.info(f"Generating embeddings for {len(documents)} documents...")
        embeddings = self.embeddings_model.encode(documents, show_progress_bar=False)
        embeddings = np.array(embeddings)
        self.logger.info(f"Embeddings shape: {embeddings.shape}")

        # Train DBSCAN
        self.logger.info("Training DBSCAN model...")
        start_time = time.time()

        dbscan = DBSCAN(
            eps=eps,
            min_samples=min_samples,
            metric='cosine',
            n_jobs=-1
        )
        labels = dbscan.fit_predict(embeddings)
        training_time = time.time() - start_time

        self.logger.info(f"DBSCAN training completed in {training_time:.2f} seconds")

        # Count clusters (excluding noise label -1)
        unique_labels = set(labels)
        num_clusters = len(unique_labels) - (1 if -1 in unique_labels else 0)
        num_outliers = sum(1 for label in labels if label == -1)

        self.logger.info(f"Found {num_clusters} clusters and {num_outliers} outliers")

        # Extract clusters
        clusters = []
        document_clusters = []

        for cluster_num in unique_labels:
            if cluster_num == -1:
                continue  # Skip noise/outliers

            # Get documents in this cluster
            cluster_doc_indices = [i for i, label in enumerate(labels) if label == cluster_num]
            cluster_doc_ids = [doc_ids[i] for i in cluster_doc_indices]

            # Calculate centroid (mean of cluster embeddings)
            cluster_embeddings = embeddings[cluster_doc_indices]
            centroid = np.mean(cluster_embeddings, axis=0)

            # Get representative documents (closest to centroid)
            if cluster_doc_indices:
                distances = np.linalg.norm(cluster_embeddings - centroid, axis=1)
                top_indices = distances.argsort()[:5]  # Top 5 closest
                representative_docs = [cluster_doc_ids[i] for i in top_indices]
            else:
                representative_docs = []

            # Calculate top terms for this cluster
            cluster_texts = [documents[i] for i in cluster_doc_indices]
            top_terms = self._extract_top_terms_from_texts(cluster_texts, top_n=10)

            clusters.append({
                'cluster_id': f"dbscan_cluster_{cluster_num}",
                'cluster_number': cluster_num,
                'centroid': centroid,
                'num_documents': len(cluster_doc_ids),
                'representative_docs': representative_docs,
                'top_terms': top_terms,
                'silhouette_score': None
            })

            # Create document-cluster assignments
            for doc_idx in cluster_doc_indices:
                doc_id = doc_ids[doc_idx]
                distance = float(np.linalg.norm(embeddings[doc_idx] - centroid))

                document_clusters.append({
                    'doc_id': doc_id,
                    'cluster_number': cluster_num,
                    'distance': distance
                })

        # Store results to database if requested
        if store_results and clusters:
            self.logger.info("Storing DBSCAN results to database...")
            self._store_clusters_to_db(clusters, 'dbscan')
            self._store_document_clusters(document_clusters, 'dbscan')
            self.logger.info("DBSCAN results stored to database")

        return {
            'algorithm': 'dbscan',
            'num_clusters': num_clusters,
            'clusters': clusters,
            'document_clusters': document_clusters,
            'num_outliers': num_outliers,
            'training_time': training_time,
            'num_documents': len(doc_ids)
        }

    def cluster_documents_hdbscan(self, min_cluster_size: int = 5,
                                  min_samples: int = 3,
                                  store_results: bool = True) -> dict:
        """
        Cluster documents using HDBSCAN (Hierarchical DBSCAN).

        HDBSCAN is an improved version of DBSCAN that automatically selects
        clusters and handles varying density clusters better.

        Args:
            min_cluster_size: Minimum cluster size
            min_samples: Minimum samples in a neighborhood
            store_results: Whether to store results to database

        Returns:
            {
                'algorithm': 'hdbscan',
                'num_clusters': int (excluding noise),
                'clusters': list of cluster dicts,
                'document_clusters': list of assignments,
                'num_outliers': int (noise points),
                'training_time': float,
                'num_documents': int
            }

        Examples:
            >>> kb = KnowledgeBase()
            >>> results = kb.cluster_documents_hdbscan(min_cluster_size=5)
            >>> print(f"Found {results['num_clusters']} clusters")
            >>> print(f"Outliers: {results['num_outliers']}")
        """
        from hdbscan import HDBSCAN
        import numpy as np
        import time

        self.logger.info(f"Clustering documents with HDBSCAN (min_cluster_size={min_cluster_size})...")

        # Get documents and generate embeddings
        self.logger.info("Collecting documents and generating embeddings...")
        doc_ids = []
        documents = []

        for doc_id, doc in self.documents.items():
            chunks = self._get_chunks_db(doc_id)
            if chunks:
                full_text = " ".join(chunk.content for chunk in chunks)
                documents.append(full_text)
                doc_ids.append(doc_id)

        if not documents:
            raise ValueError("No documents available for clustering")

        # Generate embeddings
        if not self._embeddings_loaded:
            self._ensure_embeddings_loaded()

        self.logger.info(f"Generating embeddings for {len(documents)} documents...")
        embeddings = self.embeddings_model.encode(documents, show_progress_bar=False)
        embeddings = np.array(embeddings)
        self.logger.info(f"Embeddings shape: {embeddings.shape}")

        # Train HDBSCAN
        self.logger.info("Training HDBSCAN model...")
        start_time = time.time()

        hdbscan = HDBSCAN(
            min_cluster_size=min_cluster_size,
            min_samples=min_samples,
            metric='euclidean',
            cluster_selection_method='eom',
            prediction_data=True
        )
        labels = hdbscan.fit_predict(embeddings)
        training_time = time.time() - start_time

        self.logger.info(f"HDBSCAN training completed in {training_time:.2f} seconds")

        # Count clusters (excluding noise label -1)
        unique_labels = set(labels)
        num_clusters = len(unique_labels) - (1 if -1 in unique_labels else 0)
        num_outliers = sum(1 for label in labels if label == -1)

        self.logger.info(f"Found {num_clusters} clusters and {num_outliers} outliers")

        # Extract clusters
        clusters = []
        document_clusters = []

        for cluster_num in unique_labels:
            if cluster_num == -1:
                continue  # Skip noise/outliers

            # Get documents in this cluster
            cluster_doc_indices = [i for i, label in enumerate(labels) if label == cluster_num]
            cluster_doc_ids = [doc_ids[i] for i in cluster_doc_indices]

            # Calculate centroid (mean of cluster embeddings)
            cluster_embeddings = embeddings[cluster_doc_indices]
            centroid = np.mean(cluster_embeddings, axis=0)

            # Get representative documents (closest to centroid)
            if cluster_doc_indices:
                distances = np.linalg.norm(cluster_embeddings - centroid, axis=1)
                top_indices = distances.argsort()[:5]  # Top 5 closest
                representative_docs = [cluster_doc_ids[i] for i in top_indices]
            else:
                representative_docs = []

            # Calculate top terms for this cluster
            cluster_texts = [documents[i] for i in cluster_doc_indices]
            top_terms = self._extract_top_terms_from_texts(cluster_texts, top_n=10)

            clusters.append({
                'cluster_id': f"hdbscan_cluster_{cluster_num}",
                'cluster_number': cluster_num,
                'centroid': centroid,
                'num_documents': len(cluster_doc_ids),
                'representative_docs': representative_docs,
                'top_terms': top_terms,
                'silhouette_score': None
            })

            # Create document-cluster assignments
            for doc_idx in cluster_doc_indices:
                doc_id = doc_ids[doc_idx]
                distance = float(np.linalg.norm(embeddings[doc_idx] - centroid))

                document_clusters.append({
                    'doc_id': doc_id,
                    'cluster_number': cluster_num,
                    'distance': distance
                })

        # Store results to database if requested
        if store_results and clusters:
            self.logger.info("Storing HDBSCAN results to database...")
            self._store_clusters_to_db(clusters, 'hdbscan')
            self._store_document_clusters(document_clusters, 'hdbscan')
            self.logger.info("HDBSCAN results stored to database")

        return {
            'algorithm': 'hdbscan',
            'num_clusters': num_clusters,
            'clusters': clusters,
            'document_clusters': document_clusters,
            'num_outliers': num_outliers,
            'training_time': training_time,
            'num_documents': len(doc_ids)
        }

    def evaluate_clustering(self, algorithm: str) -> dict:
        """
        Evaluate clustering quality with multiple metrics.

        Calculates Silhouette score, Davies-Bouldin index, and
        Calinski-Harabasz score for the specified clustering algorithm.

        Args:
            algorithm: Clustering algorithm name (kmeans, dbscan, hdbscan)

        Returns:
            {
                'algorithm': str,
                'silhouette_score': float (higher is better, -1 to 1),
                'davies_bouldin_score': float (lower is better, ≥0),
                'calinski_harabasz_score': float (higher is better, ≥0),
                'num_clusters': int,
                'num_documents': int
            }

        Examples:
            >>> kb = KnowledgeBase()
            >>> kb.cluster_documents_kmeans(num_clusters=5)
            >>> metrics = kb.evaluate_clustering('kmeans')
            >>> print(f"Silhouette: {metrics['silhouette_score']:.3f}")
        """
        from sklearn.metrics import silhouette_score, davies_bouldin_score, calinski_harabasz_score
        import numpy as np

        self.logger.info(f"Evaluating {algorithm} clustering quality...")

        # Get document cluster assignments
        cursor = self.db_conn.cursor()
        assignments = cursor.execute("""
            SELECT doc_id, cluster_id
            FROM document_clusters
            WHERE algorithm = ?
            ORDER BY doc_id
        """, (algorithm,)).fetchall()

        if not assignments:
            raise ValueError(f"No clustering results found for algorithm: {algorithm}")

        # Get documents and generate embeddings
        doc_ids = [doc_id for doc_id, _ in assignments]
        documents = []

        for doc_id in doc_ids:
            chunks = self._get_chunks_db(doc_id)
            if chunks:
                full_text = " ".join(chunk.content for chunk in chunks)
                documents.append(full_text)

        # Generate embeddings
        if not self._embeddings_loaded:
            self._ensure_embeddings_loaded()

        embeddings = self.embeddings_model.encode(documents, show_progress_bar=False)
        embeddings = np.array(embeddings)

        # Get cluster labels
        cluster_ids = [cluster_id for _, cluster_id in assignments]
        unique_cluster_ids = list(set(cluster_ids))
        cluster_to_label = {cid: i for i, cid in enumerate(unique_cluster_ids)}
        labels = np.array([cluster_to_label[cid] for cid in cluster_ids])

        # Calculate metrics
        silhouette = silhouette_score(embeddings, labels)
        davies_bouldin = davies_bouldin_score(embeddings, labels)
        calinski_harabasz = calinski_harabasz_score(embeddings, labels)

        self.logger.info(f"Silhouette score: {silhouette:.3f}")
        self.logger.info(f"Davies-Bouldin index: {davies_bouldin:.3f}")
        self.logger.info(f"Calinski-Harabasz score: {calinski_harabasz:.2f}")

        return {
            'algorithm': algorithm,
            'silhouette_score': float(silhouette),
            'davies_bouldin_score': float(davies_bouldin),
            'calinski_harabasz_score': float(calinski_harabasz),
            'num_clusters': len(unique_cluster_ids),
            'num_documents': len(doc_ids)
        }

    # ============================================================
    # VISUALIZATION METHODS (Phase 2, Task 2.6)
    # ============================================================

    def visualize_topics_wordcloud(self, model_type: str = 'lda',
                                   output_dir: str = "topic_wordclouds") -> List[str]:
        """
        Generate word cloud visualizations for each topic.

        Creates PNG images showing word importance in each topic using
        word clouds. Words that are more important in the topic appear
        larger in the visualization.

        Args:
            model_type: Topic model type (lda, nmf, bertopic)
            output_dir: Directory to save word cloud images

        Returns:
            List of output file paths

        Examples:
            >>> kb = KnowledgeBase()
            >>> kb.train_lda_model(num_topics=5)
            >>> files = kb.visualize_topics_wordcloud('lda')
            >>> print(f"Created {len(files)} word clouds")
        """
        import json
        import os
        from pathlib import Path

        # Import visualization libraries
        try:
            from wordcloud import WordCloud
            import matplotlib
            matplotlib.use('Agg')  # Non-interactive backend
            import matplotlib.pyplot as plt
        except ImportError:
            raise ImportError("wordcloud and matplotlib required for visualizations. Install with: pip install wordcloud matplotlib")

        self.logger.info(f"Generating word clouds for {model_type} topics...")

        # Create output directory
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        # Get topics from database
        cursor = self.db_conn.cursor()
        topics = cursor.execute("""
            SELECT topic_number, word_weights, top_words
            FROM topics
            WHERE model_type = ?
            ORDER BY topic_number
        """, (model_type,)).fetchall()

        if not topics:
            raise ValueError(f"No topics found for model type: {model_type}")

        output_files = []

        for topic_num, word_weights_json, top_words_json in topics:
            # Parse word weights
            word_weights = json.loads(word_weights_json)

            # Create word cloud
            wc = WordCloud(
                width=800,
                height=400,
                background_color='white',
                colormap='viridis',
                relative_scaling=0.5,
                min_font_size=10
            ).generate_from_frequencies(word_weights)

            # Create figure
            plt.figure(figsize=(10, 5))
            plt.imshow(wc, interpolation='bilinear')
            plt.axis('off')
            plt.title(f"{model_type.upper()} Topic {topic_num}", fontsize=16, fontweight='bold')

            # Save image
            output_file = output_path / f"topic_{model_type}_{topic_num}.png"
            plt.savefig(output_file, bbox_inches='tight', dpi=150)
            plt.close()

            output_files.append(str(output_file))
            self.logger.info(f"Created word cloud: {output_file}")

        self.logger.info(f"Generated {len(output_files)} word cloud visualizations")
        return output_files

    def visualize_topic_distribution(self, model_type: str = 'lda',
                                     output_path: str = "topic_distribution.html") -> str:
        """
        Create interactive bar chart showing topic distribution across corpus.

        Shows how many documents are primarily associated with each topic,
        with interactive hover information.

        Args:
            model_type: Topic model type (lda, nmf, bertopic)
            output_path: Output HTML file path

        Returns:
            Path to output HTML file

        Examples:
            >>> kb = KnowledgeBase()
            >>> kb.train_lda_model(num_topics=5)
            >>> file = kb.visualize_topic_distribution('lda')
            >>> print(f"Chart saved to {file}")
        """
        import json
        from pathlib import Path

        try:
            import plotly.graph_objects as go
        except ImportError:
            raise ImportError("plotly required for visualizations. Install with: pip install plotly")

        self.logger.info(f"Creating topic distribution chart for {model_type}...")

        # Get topics with document counts
        cursor = self.db_conn.cursor()
        topics = cursor.execute("""
            SELECT topic_number, num_documents, top_words
            FROM topics
            WHERE model_type = ?
            ORDER BY topic_number
        """, (model_type,)).fetchall()

        if not topics:
            raise ValueError(f"No topics found for model type: {model_type}")

        # Parse data
        topic_numbers = []
        doc_counts = []
        top_words_list = []

        for topic_num, num_docs, top_words_json in topics:
            topic_numbers.append(f"Topic {topic_num}")
            doc_counts.append(num_docs)
            top_words = json.loads(top_words_json)
            top_words_list.append(", ".join(top_words[:5]))

        # Create bar chart
        fig = go.Figure(data=[
            go.Bar(
                x=topic_numbers,
                y=doc_counts,
                text=doc_counts,
                textposition='auto',
                hovertemplate='<b>%{x}</b><br>Documents: %{y}<br>Top words: %{customdata}<extra></extra>',
                customdata=top_words_list,
                marker=dict(
                    color=doc_counts,
                    colorscale='Viridis',
                    showscale=True,
                    colorbar=dict(title="Documents")
                )
            )
        ])

        fig.update_layout(
            title=f"{model_type.upper()} Topic Distribution Across Corpus",
            xaxis_title="Topic",
            yaxis_title="Number of Documents",
            height=500,
            template="plotly_white",
            hovermode='closest'
        )

        # Save to HTML
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        fig.write_html(str(output_file))

        self.logger.info(f"Topic distribution chart saved to {output_file}")
        return str(output_file)

    def visualize_clusters_2d(self, algorithm: str = 'kmeans',
                              output_path: str = "clusters_2d.html") -> str:
        """
        Visualize document clusters in 2D using UMAP dimensionality reduction.

        Creates interactive scatter plot showing how documents cluster together
        in semantic space. Uses UMAP to reduce high-dimensional embeddings to 2D.

        Args:
            algorithm: Clustering algorithm (kmeans, dbscan, hdbscan)
            output_path: Output HTML file path

        Returns:
            Path to output HTML file

        Examples:
            >>> kb = KnowledgeBase()
            >>> kb.cluster_documents_kmeans(num_clusters=5)
            >>> file = kb.visualize_clusters_2d('kmeans')
            >>> print(f"Cluster visualization saved to {file}")
        """
        import numpy as np
        from pathlib import Path

        try:
            import plotly.express as px
            import pandas as pd
        except ImportError:
            raise ImportError("plotly and pandas required. Install with: pip install plotly pandas")

        try:
            from umap import UMAP
        except ImportError:
            raise ImportError("umap-learn required. Install with: pip install umap-learn")

        self.logger.info(f"Creating 2D cluster visualization for {algorithm}...")

        # Get cluster assignments
        cursor = self.db_conn.cursor()
        assignments = cursor.execute("""
            SELECT dc.doc_id, dc.cluster_id, c.cluster_number
            FROM document_clusters dc
            JOIN clusters c ON dc.cluster_id = c.cluster_id
            WHERE dc.algorithm = ?
            ORDER BY dc.doc_id
        """, (algorithm,)).fetchall()

        if not assignments:
            raise ValueError(f"No clustering results found for algorithm: {algorithm}")

        # Get documents and generate embeddings
        doc_ids = [doc_id for doc_id, _, _ in assignments]
        cluster_labels = [cluster_num for _, _, cluster_num in assignments]
        documents = []
        titles = []

        for doc_id in doc_ids:
            chunks = self._get_chunks_db(doc_id)
            if chunks:
                full_text = " ".join(chunk.content for chunk in chunks)
                documents.append(full_text)
                titles.append(self.documents[doc_id].title if doc_id in self.documents else doc_id[:12])

        # Generate embeddings
        if not self._embeddings_loaded:
            self._ensure_embeddings_loaded()

        self.logger.info(f"Generating embeddings for {len(documents)} documents...")
        embeddings = self.embeddings_model.encode(documents, show_progress_bar=False)
        embeddings = np.array(embeddings)

        # Reduce to 2D with UMAP
        self.logger.info("Reducing embeddings to 2D with UMAP...")
        reducer = UMAP(
            n_components=2,
            random_state=42,
            n_neighbors=15,
            min_dist=0.1,
            metric='cosine'
        )
        embedding_2d = reducer.fit_transform(embeddings)

        # Create DataFrame
        df = pd.DataFrame({
            'x': embedding_2d[:, 0],
            'y': embedding_2d[:, 1],
            'cluster': [f"Cluster {label}" for label in cluster_labels],
            'cluster_num': cluster_labels,
            'doc_id': doc_ids,
            'title': titles
        })

        # Create scatter plot
        fig = px.scatter(
            df,
            x='x',
            y='y',
            color='cluster',
            hover_data=['title', 'doc_id'],
            title=f"Document Clusters in 2D ({algorithm.upper()})",
            labels={'x': 'UMAP Dimension 1', 'y': 'UMAP Dimension 2'},
            color_discrete_sequence=px.colors.qualitative.Vivid
        )

        fig.update_traces(marker=dict(size=8, line=dict(width=0.5, color='DarkSlateGrey')))
        fig.update_layout(
            height=600,
            template="plotly_white",
            hovermode='closest',
            legend=dict(title="Cluster")
        )

        # Save to HTML
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        fig.write_html(str(output_file))

        self.logger.info(f"2D cluster visualization saved to {output_file}")
        return str(output_file)

    def visualize_cluster_dendrogram(self, algorithm: str = 'kmeans',
                                     output_path: str = "cluster_dendrogram.html") -> str:
        """
        Create dendrogram visualization showing hierarchical cluster relationships.

        Builds a hierarchical tree showing how clusters relate to each other
        based on their centroid distances. Useful for understanding cluster
        similarity and structure.

        Args:
            algorithm: Clustering algorithm (kmeans, dbscan, hdbscan)
            output_path: Output HTML file path

        Returns:
            Path to output HTML file

        Examples:
            >>> kb = KnowledgeBase()
            >>> kb.cluster_documents_kmeans(num_clusters=5)
            >>> file = kb.visualize_cluster_dendrogram('kmeans')
            >>> print(f"Dendrogram saved to {file}")
        """
        import numpy as np
        from pathlib import Path

        try:
            from scipy.cluster.hierarchy import dendrogram, linkage
            from scipy.spatial.distance import pdist
        except ImportError:
            raise ImportError("scipy required. Install with: pip install scipy")

        try:
            import plotly.figure_factory as ff
        except ImportError:
            raise ImportError("plotly required. Install with: pip install plotly")

        self.logger.info(f"Creating dendrogram for {algorithm} clusters...")

        # Get cluster centroids
        cursor = self.db_conn.cursor()
        clusters = cursor.execute("""
            SELECT cluster_number, centroid_vector
            FROM clusters
            WHERE algorithm = ?
            ORDER BY cluster_number
        """, (algorithm,)).fetchall()

        if not clusters:
            raise ValueError(f"No clustering results found for algorithm: {algorithm}")

        # Parse centroids
        cluster_numbers = []
        centroids = []

        for cluster_num, centroid_blob in clusters:
            cluster_numbers.append(cluster_num)
            centroid = np.frombuffer(centroid_blob, dtype=np.float32)
            centroids.append(centroid)

        centroids = np.array(centroids, dtype=np.float64)  # Convert to float64
        self.logger.info(f"Loaded {len(centroids)} cluster centroids with shape {centroids.shape}")

        # Check for and handle non-finite values
        if not np.all(np.isfinite(centroids)):
            self.logger.warning("Found non-finite values in centroids, replacing with zeros")
            centroids = np.nan_to_num(centroids, nan=0.0, posinf=0.0, neginf=0.0)

        # Normalize centroids to prevent numerical issues
        from sklearn.preprocessing import normalize
        centroids = normalize(centroids)

        # Perform hierarchical clustering on centroids
        self.logger.info("Computing hierarchical clustering...")
        linkage_matrix = linkage(centroids, method='ward')

        # Create labels
        labels = [f"Cluster {num}" for num in cluster_numbers]

        # Create dendrogram using plotly
        fig = ff.create_dendrogram(
            centroids,
            labels=labels,
            linkagefun=lambda x: linkage(x, method='ward')
        )

        fig.update_layout(
            title=f"Cluster Dendrogram ({algorithm.upper()})",
            xaxis_title="Cluster",
            yaxis_title="Distance",
            height=500,
            template="plotly_white"
        )

        # Save to HTML
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        fig.write_html(str(output_file))

        self.logger.info(f"Dendrogram saved to {output_file}")
        return str(output_file)

    # ============================================================
    # Phase 3: Temporal Analysis Methods
    # ============================================================

    def extract_dates_from_text(self, text: str) -> list[dict]:
        """
        Extract dates and temporal references from text using regex patterns.

        This method identifies various date formats commonly found in C64 documentation:
        - Years: "1982", "1983"
        - Month/Year: "January 1982", "Jan 1982"
        - Full dates: "January 15, 1982", "15-Jan-1982"
        - Decades: "1980s", "early 80s"
        - Date ranges: "1982-1985", "1982 to 1985"

        Args:
            text: Text to extract dates from

        Returns:
            List of dictionaries with:
            - 'text': Original matched text
            - 'type': Date type (year, month_year, full_date, decade, range)
            - 'year': Extracted year (int or None)
            - 'month': Extracted month (int or None)
            - 'day': Extracted day (int or None)
            - 'start_pos': Character position in text
            - 'end_pos': Character position in text

        Examples:
            >>> kb.extract_dates_from_text("The C64 was released in August 1982")
            [{'text': 'August 1982', 'type': 'month_year', 'year': 1982, 'month': 8, ...}]

            >>> kb.extract_dates_from_text("Popular throughout the 1980s")
            [{'text': '1980s', 'type': 'decade', 'year': 1980, ...}]
        """
        import re

        dates = []

        # Month names mapping
        months = {
            'january': 1, 'jan': 1,
            'february': 2, 'feb': 2,
            'march': 3, 'mar': 3,
            'april': 4, 'apr': 4,
            'may': 5,
            'june': 6, 'jun': 6,
            'july': 7, 'jul': 7,
            'august': 8, 'aug': 8,
            'september': 9, 'sep': 9, 'sept': 9,
            'october': 10, 'oct': 10,
            'november': 11, 'nov': 11,
            'december': 12, 'dec': 12
        }

        # Pattern 1a: Full dates - Month Day, Year (January 15, 1982)
        pattern1a = r'\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec|January|February|March|April|May|June|July|August|September|October|November|December)[\s]+(\d{1,2}),?[\s]+(\d{4})\b'
        for match in re.finditer(pattern1a, text, re.IGNORECASE):
            month_str = match.group(1).lower()
            day = int(match.group(2))
            year = int(match.group(3))
            month = months.get(month_str, None)

            dates.append({
                'text': match.group(0),
                'type': 'full_date',
                'year': year,
                'month': month,
                'day': day,
                'start_pos': match.start(),
                'end_pos': match.end()
            })

        # Pattern 1b: Full dates - Day Month Year (15 Jan 1982 | 15-Jan-1982)
        pattern1b = r'\b(\d{1,2})[\s\-](Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec|January|February|March|April|May|June|July|August|September|October|November|December)[\s,\-]+(\d{4})\b'
        for match in re.finditer(pattern1b, text, re.IGNORECASE):
            day = int(match.group(1))
            month_str = match.group(2).lower()
            year = int(match.group(3))
            month = months.get(month_str, None)

            # Skip if already matched
            if not any(d['start_pos'] <= match.start() < d['end_pos'] for d in dates):
                dates.append({
                    'text': match.group(0),
                    'type': 'full_date',
                    'year': year,
                    'month': month,
                    'day': day,
                    'start_pos': match.start(),
                    'end_pos': match.end()
                })

        # Pattern 2: Month Year (January 1982 | Jan 1982)
        pattern2 = r'\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec|January|February|March|April|May|June|July|August|September|October|November|December)[\s,]+(\d{4})\b'
        for match in re.finditer(pattern2, text, re.IGNORECASE):
            month_str = match.group(1).lower()
            year = int(match.group(2))
            month = months.get(month_str, None)

            # Skip if already matched as full date
            if not any(d['start_pos'] <= match.start() < d['end_pos'] for d in dates):
                dates.append({
                    'text': match.group(0),
                    'type': 'month_year',
                    'year': year,
                    'month': month,
                    'day': None,
                    'start_pos': match.start(),
                    'end_pos': match.end()
                })

        # Pattern 3: Year only (1982, 1983) - must be in C64 era (1975-2000)
        pattern3 = r'\b(19[7-9]\d|200\d)\b'
        for match in re.finditer(pattern3, text):
            year = int(match.group(0))

            # Skip if already matched in a more specific date
            if not any(d['start_pos'] <= match.start() < d['end_pos'] for d in dates):
                dates.append({
                    'text': match.group(0),
                    'type': 'year',
                    'year': year,
                    'month': None,
                    'day': None,
                    'start_pos': match.start(),
                    'end_pos': match.end()
                })

        # Pattern 4: Decades (1980s, early 80s, mid-80s)
        pattern4 = r'\b(early\s+|mid[\s\-]?|late\s+)?(19)?([7-9]0)s\b'
        for match in re.finditer(pattern4, text, re.IGNORECASE):
            decade = match.group(3)
            prefix = match.group(1) if match.group(1) else ''

            # Determine decade start year
            if match.group(2):  # Has "19" prefix
                year = int(f"19{decade}")
            else:
                year = int(f"19{decade}")  # Assume 1900s

            dates.append({
                'text': match.group(0),
                'type': 'decade',
                'year': year,
                'month': None,
                'day': None,
                'prefix': prefix.strip(),
                'start_pos': match.start(),
                'end_pos': match.end()
            })

        # Pattern 5: Date ranges (1982-1985, 1982 to 1985)
        pattern5 = r'\b(19[7-9]\d)[\s]*([\-to]+)[\s]*(19[7-9]\d)\b'
        for match in re.finditer(pattern5, text, re.IGNORECASE):
            start_year = int(match.group(1))
            end_year = int(match.group(3))

            dates.append({
                'text': match.group(0),
                'type': 'range',
                'year': start_year,
                'end_year': end_year,
                'month': None,
                'day': None,
                'start_pos': match.start(),
                'end_pos': match.end()
            })

        # Sort by position in text
        dates.sort(key=lambda x: x['start_pos'])

        return dates

    def normalize_date(self, date_dict: dict) -> tuple[str, int, int, int]:
        """
        Normalize a date dictionary to ISO format and component integers.

        Args:
            date_dict: Dictionary from extract_dates_from_text()

        Returns:
            Tuple of (iso_string, year, month, day)
            - iso_string: ISO 8601 date string (YYYY-MM-DD or YYYY-MM or YYYY)
            - year: Integer year
            - month: Integer month or 0 if not specified
            - day: Integer day or 0 if not specified

        Examples:
            >>> date_dict = {'type': 'month_year', 'year': 1982, 'month': 8, 'day': None}
            >>> kb.normalize_date(date_dict)
            ('1982-08', 1982, 8, 0)

            >>> date_dict = {'type': 'year', 'year': 1982, 'month': None, 'day': None}
            >>> kb.normalize_date(date_dict)
            ('1982', 1982, 0, 0)
        """
        year = date_dict.get('year', 0)
        month = date_dict.get('month', 0) if date_dict.get('month') else 0
        day = date_dict.get('day', 0) if date_dict.get('day') else 0

        # Build ISO string based on precision
        if day > 0 and month > 0:
            iso_string = f"{year:04d}-{month:02d}-{day:02d}"
        elif month > 0:
            iso_string = f"{year:04d}-{month:02d}"
        else:
            iso_string = f"{year:04d}"

        return iso_string, year, month, day

    def detect_events_in_text(self, text: str, doc_id: Optional[str] = None) -> list[dict]:
        """
        Detect significant events in text using pattern matching.

        This method identifies events commonly found in C64 documentation:
        - Product releases: "released", "launched", "introduced"
        - Company milestones: "founded", "acquired", "established"
        - Technical innovations: "first", "invented", "developed", "created"
        - Cultural events: "competition", "demo", "conference", "meeting"

        Args:
            text: Text to analyze for events
            doc_id: Optional document ID for context

        Returns:
            List of event dictionaries with:
            - 'type': Event type (release, milestone, innovation, cultural)
            - 'title': Brief event title
            - 'description': Event description with context
            - 'date_info': Associated date information
            - 'confidence': Confidence score (0.0-1.0)
            - 'entities': Related entities (if any)
            - 'position': Character position in text

        Examples:
            >>> kb.detect_events_in_text("The C64 was released in August 1982")
            [{'type': 'release', 'title': 'C64 released', 'date_info': {...}, ...}]
        """
        import re

        events = []

        # First, extract all dates from the text
        dates = self.extract_dates_from_text(text)

        # Event patterns with trigger words and types
        event_patterns = [
            # Product releases
            {
                'pattern': r'\b(released|launched|introduced|unveiled|announced|shipped|available)\b',
                'type': 'release',
                'confidence': 0.85
            },
            # Company milestones
            {
                'pattern': r'\b(founded|established|acquired|merged|created|formed|incorporated)\b',
                'type': 'milestone',
                'confidence': 0.80
            },
            # Technical innovations
            {
                'pattern': r'\b(first|invented|developed|created|designed|pioneered|innovated|breakthrough)\b',
                'type': 'innovation',
                'confidence': 0.75
            },
            # Cultural events
            {
                'pattern': r'\b(competition|contest|demo|demonstration|conference|convention|meeting|expo|exhibition|show)\b',
                'type': 'cultural',
                'confidence': 0.70
            },
            # Version/update events
            {
                'pattern': r'\b(updated|upgraded|revised|version|release|edition)\b',
                'type': 'update',
                'confidence': 0.65
            }
        ]

        # Find all event trigger words
        for event_pattern in event_patterns:
            for match in re.finditer(event_pattern['pattern'], text, re.IGNORECASE):
                trigger_word = match.group(0)
                trigger_pos = match.start()

                # Find the nearest date (within 200 characters)
                nearest_date = None
                min_distance = float('inf')

                for date in dates:
                    distance = abs(date['start_pos'] - trigger_pos)
                    if distance < min_distance and distance < 200:
                        min_distance = distance
                        nearest_date = date

                # Extract context around the trigger word (100 chars before and after)
                context_start = max(0, trigger_pos - 100)
                context_end = min(len(text), trigger_pos + 100)
                context = text[context_start:context_end].strip()

                # Try to extract a title from the surrounding sentence
                # Find sentence boundaries
                sentence_start = max(0, text.rfind('.', 0, trigger_pos) + 1)
                sentence_end = text.find('.', trigger_pos)
                if sentence_end == -1:
                    sentence_end = len(text)

                sentence = text[sentence_start:sentence_end].strip()

                # Create event title (first 100 chars of sentence or context)
                title = sentence[:100] if len(sentence) <= 100 else sentence[:97] + "..."

                # Adjust confidence based on date proximity
                confidence = event_pattern['confidence']
                if nearest_date:
                    # Higher confidence if date is very close
                    if min_distance < 50:
                        confidence = min(1.0, confidence + 0.1)
                else:
                    # Lower confidence if no date found
                    confidence = max(0.3, confidence - 0.2)

                # Try to extract entities from the context (basic noun extraction)
                entities = []
                # Look for capitalized words (potential proper nouns)
                entity_matches = re.findall(r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\b', sentence)
                entities = list(set(entity_matches))[:5]  # Max 5 entities

                event = {
                    'type': event_pattern['type'],
                    'title': title,
                    'description': context,
                    'trigger_word': trigger_word,
                    'date_info': nearest_date,
                    'confidence': confidence,
                    'entities': entities,
                    'position': trigger_pos,
                    'sentence': sentence
                }

                events.append(event)

        # Sort by position in text
        events.sort(key=lambda x: x['position'])

        return events

    def detect_anomalies(self, min_severity: str = 'moderate', days: int = 7) -> dict:
        """
        Detect anomalies in document monitoring history.

        This method analyzes URL monitoring history to identify unusual patterns:
        - Unexpected update frequencies (sites changing much more/less than normal)
        - Performance degradation (significantly slower response times)
        - Change magnitude anomalies (unusually large or small content changes)

        The system learns baseline patterns for each document over a configurable
        learning period (default: 30 days) and scores deviations across multiple
        dimensions.

        Args:
            min_severity: Minimum severity level to include ('normal', 'minor', 'moderate', 'critical')
            days: Number of days of history to analyze (default: 7)

        Returns:
            Dictionary containing:
            - 'anomalies': List of anomaly records with scores and details
            - 'total_anomalies': Total number of anomalies found
            - 'by_severity': Count breakdown by severity level
            - 'avg_score': Average anomaly score
            - 'time_range_days': Days of history analyzed

        Severity Levels:
            - normal: 0-30 (no alert needed)
            - minor: 31-60 (include in digest)
            - moderate: 61-85 (immediate notification)
            - critical: 86-100 (urgent alert)

        Examples:
            >>> # Get critical and moderate anomalies from last 7 days
            >>> results = kb.detect_anomalies(min_severity='moderate', days=7)
            >>> print(f"Found {results['total_anomalies']} anomalies")
            >>> for anomaly in results['anomalies']:
            ...     print(f"  {anomaly['doc_title']}: {anomaly['severity']} ({anomaly['score']}/100)")

            >>> # Get all anomalies (including minor) from last 30 days
            >>> results = kb.detect_anomalies(min_severity='minor', days=30)
            >>> print(f"Breakdown: {results['by_severity']}")

        Note:
            Requires anomaly detection to be enabled (USE_ANOMALY_DETECTION=1).
            Will return error if anomaly_detector is not initialized.
        """
        if not self.anomaly_detector:
            return {
                'error': 'Anomaly detection not available',
                'anomalies': [],
                'total_anomalies': 0,
                'by_severity': {},
                'avg_score': 0.0,
                'time_range_days': days
            }

        try:
            # Get anomalies from detector
            anomalies = self.anomaly_detector.get_anomalies(
                min_severity=min_severity,
                days=days
            )

            # Calculate statistics
            total_anomalies = len(anomalies)
            by_severity = {}
            total_score = 0.0

            for anomaly in anomalies:
                severity = anomaly.get('severity', 'unknown')
                by_severity[severity] = by_severity.get(severity, 0) + 1
                total_score += anomaly.get('score', 0.0)

            avg_score = total_score / total_anomalies if total_anomalies > 0 else 0.0

            return {
                'anomalies': anomalies,
                'total_anomalies': total_anomalies,
                'by_severity': by_severity,
                'avg_score': round(avg_score, 2),
                'time_range_days': days
            }

        except Exception as e:
            self.logger.error(f"Anomaly detection failed: {e}", exc_info=True)
            return {
                'error': f'Anomaly detection failed: {str(e)}',
                'anomalies': [],
                'total_anomalies': 0,
                'by_severity': {},
                'avg_score': 0.0,
                'time_range_days': days
            }

    def extract_document_events(self, doc_id: str, min_confidence: float = 0.5) -> dict:
        """
        Extract events from a specific document and store them in the database.

        Args:
            doc_id: Document ID to process
            min_confidence: Minimum confidence threshold (0.0-1.0)

        Returns:
            Dictionary with:
            - 'doc_id': Document ID
            - 'title': Document title
            - 'event_count': Number of events extracted
            - 'events': List of extracted events
            - 'stored_count': Number of events stored to database

        Examples:
            >>> result = kb.extract_document_events('doc123', min_confidence=0.6)
            >>> print(f"Found {result['event_count']} events")
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document {doc_id} not found")

        doc = self.documents[doc_id]

        # Load document chunks
        chunks = self._get_chunks_db(doc_id)

        # Combine chunks into full text (with chunk boundaries marked)
        full_text = ""
        chunk_positions = []  # Track which chunk each character belongs to

        for chunk in chunks:
            start_pos = len(full_text)
            full_text += chunk.content + "\n\n"
            end_pos = len(full_text)
            chunk_positions.append({
                'chunk_id': chunk.chunk_id,
                'start': start_pos,
                'end': end_pos
            })

        # Detect events in the full text
        events = self.detect_events_in_text(full_text, doc_id)

        # Filter by confidence
        filtered_events = [e for e in events if e['confidence'] >= min_confidence]

        # Store events to database
        stored_count = 0
        cursor = self.db_conn.cursor()

        try:
            cursor.execute("BEGIN TRANSACTION")

            for event in filtered_events:
                # Generate event ID
                import uuid
                from datetime import datetime

                event_id = str(uuid.uuid4())

                # Normalize date if available
                date_normalized = None
                year = None
                month = None
                day = None

                if event['date_info']:
                    iso_string, year, month, day = self.normalize_date(event['date_info'])
                    date_normalized = iso_string

                # Store event
                cursor.execute("""
                    INSERT INTO events
                    (event_id, event_type, title, description, date_extracted,
                     date_normalized, year, month, day, confidence, entities,
                     metadata, created_date)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    event_id,
                    event['type'],
                    event['title'],
                    event['description'],
                    event['date_info']['text'] if event['date_info'] else None,
                    date_normalized,
                    year if year else None,
                    month if month else None,
                    day if day else None,
                    event['confidence'],
                    json.dumps(event['entities']),
                    json.dumps({
                        'trigger_word': event['trigger_word'],
                        'sentence': event['sentence']
                    }),
                    datetime.utcnow().isoformat()
                ))

                # Find which chunk this event belongs to
                event_chunk_id = None
                for chunk_pos in chunk_positions:
                    if chunk_pos['start'] <= event['position'] < chunk_pos['end']:
                        event_chunk_id = chunk_pos['chunk_id']
                        break

                # Store document-event mapping
                mapping_id = str(uuid.uuid4())
                cursor.execute("""
                    INSERT INTO document_events
                    (mapping_id, doc_id, event_id, context, position, created_date)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    mapping_id,
                    doc_id,
                    event_id,
                    event['description'],
                    event['position'],
                    datetime.utcnow().isoformat()
                ))

                stored_count += 1

            self.db_conn.commit()
            self.logger.info(f"Stored {stored_count} events for document {doc_id}")

        except Exception as e:
            self.db_conn.rollback()
            self.logger.error(f"Failed to store events: {e}")
            raise

        return {
            'doc_id': doc_id,
            'title': doc.title,
            'event_count': len(events),
            'filtered_count': len(filtered_events),
            'stored_count': stored_count,
            'events': filtered_events
        }

    def build_timeline(self, min_confidence: float = 0.5,
                       categorize: bool = True) -> dict:
        """
        Build a chronological timeline from events in the database.

        This method creates timeline_entries records from events, organizing them
        chronologically and optionally categorizing them by event type and time period.

        Args:
            min_confidence: Minimum confidence threshold for events to include
            categorize: If True, categorize events by decade and type

        Returns:
            Dictionary with:
            - 'total_events': Total events processed
            - 'timeline_entries': Number of timeline entries created
            - 'date_range': Earliest and latest years
            - 'categories': Event counts by category
            - 'by_year': Event counts by year

        Examples:
            >>> result = kb.build_timeline(min_confidence=0.6)
            >>> print(f"Created timeline with {result['timeline_entries']} entries")
        """
        import uuid
        from datetime import datetime

        cursor = self.db_conn.cursor()

        # Get all events with dates above confidence threshold
        events = cursor.execute("""
            SELECT event_id, event_type, title, description,
                   date_normalized, year, month, day, confidence
            FROM events
            WHERE confidence >= ? AND year IS NOT NULL
            ORDER BY year, month, day
        """, (min_confidence,)).fetchall()

        if not events:
            self.logger.warning("No events found with dates to build timeline")
            return {
                'total_events': 0,
                'timeline_entries': 0,
                'date_range': None,
                'categories': {},
                'by_year': {}
            }

        # Clear existing timeline entries (rebuild)
        cursor.execute("DELETE FROM timeline_entries")

        # Track statistics
        categories = {}
        by_year = {}
        created_count = 0

        try:
            for event in events:
                event_id, event_type, title, description, date_normalized, year, month, day, confidence = event

                # Create display date
                if day and month:
                    display_date = f"{year}-{month:02d}-{day:02d}"
                elif month:
                    display_date = f"{year}-{month:02d}"
                else:
                    display_date = str(year)

                # Create sort order (YYYYMMDD as integer)
                sort_order = year * 10000
                if month:
                    sort_order += month * 100
                if day:
                    sort_order += day

                # Determine category
                if categorize:
                    decade = (year // 10) * 10
                    category = f"{decade}s-{event_type}"
                else:
                    category = event_type

                # Determine importance (1-5 scale, 5 is highest)
                # Higher confidence = higher importance
                importance = min(5, max(1, int(confidence * 5)))

                # Create timeline entry
                entry_id = str(uuid.uuid4())
                cursor.execute("""
                    INSERT INTO timeline_entries
                    (entry_id, event_id, display_date, sort_order,
                     category, importance, created_date)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    entry_id,
                    event_id,
                    display_date,
                    sort_order,
                    category,
                    importance,
                    datetime.utcnow().isoformat()
                ))

                created_count += 1

                # Update statistics
                categories[category] = categories.get(category, 0) + 1
                by_year[year] = by_year.get(year, 0) + 1

            self.db_conn.commit()
            self.logger.info(f"Built timeline with {created_count} entries from {len(events)} events")

        except Exception as e:
            self.db_conn.rollback()
            self.logger.error(f"Failed to build timeline: {e}")
            raise

        # Determine date range
        years = list(by_year.keys())
        date_range = (min(years), max(years)) if years else None

        return {
            'total_events': len(events),
            'timeline_entries': created_count,
            'date_range': date_range,
            'categories': categories,
            'by_year': by_year
        }

    def get_timeline(self, start_year: Optional[int] = None,
                     end_year: Optional[int] = None,
                     category: Optional[str] = None,
                     min_importance: int = 1,
                     limit: Optional[int] = None) -> list[dict]:
        """
        Query timeline entries with optional filtering.

        Args:
            start_year: Filter events from this year onwards (inclusive)
            end_year: Filter events up to this year (inclusive)
            category: Filter by category (e.g., "1980s-release")
            min_importance: Minimum importance level (1-5)
            limit: Maximum number of entries to return

        Returns:
            List of timeline entry dictionaries with:
            - 'entry_id': Timeline entry ID
            - 'event_id': Associated event ID
            - 'display_date': Formatted date string
            - 'sort_order': Integer for chronological sorting
            - 'category': Event category
            - 'importance': Importance level (1-5)
            - 'event_type': Event type
            - 'title': Event title
            - 'description': Event description
            - 'confidence': Event confidence

        Examples:
            >>> # Get all events from the 1980s
            >>> timeline = kb.get_timeline(start_year=1980, end_year=1989)

            >>> # Get important release events
            >>> releases = kb.get_timeline(category="1980s-release", min_importance=4)
        """
        cursor = self.db_conn.cursor()

        # Build query with filters
        query = """
            SELECT
                t.entry_id, t.event_id, t.display_date, t.sort_order,
                t.category, t.importance,
                e.event_type, e.title, e.description, e.confidence, e.year
            FROM timeline_entries t
            JOIN events e ON t.event_id = e.event_id
            WHERE t.importance >= ?
        """
        params = [min_importance]

        if start_year:
            query += " AND e.year >= ?"
            params.append(start_year)

        if end_year:
            query += " AND e.year <= ?"
            params.append(end_year)

        if category:
            query += " AND t.category = ?"
            params.append(category)

        query += " ORDER BY t.sort_order ASC"

        if limit:
            query += " LIMIT ?"
            params.append(limit)

        results = cursor.execute(query, params).fetchall()

        timeline = []
        for row in results:
            timeline.append({
                'entry_id': row[0],
                'event_id': row[1],
                'display_date': row[2],
                'sort_order': row[3],
                'category': row[4],
                'importance': row[5],
                'event_type': row[6],
                'title': row[7],
                'description': row[8],
                'confidence': row[9],
                'year': row[10]
            })

        return timeline

    def search_events_by_date(self, start_year: Optional[int] = None,
                              end_year: Optional[int] = None,
                              event_type: Optional[str] = None,
                              min_confidence: float = 0.5) -> list[dict]:
        """
        Search for events within a date range.

        Args:
            start_year: Start year (inclusive)
            end_year: End year (inclusive)
            event_type: Filter by event type (release, milestone, innovation, etc.)
            min_confidence: Minimum confidence threshold

        Returns:
            List of event dictionaries ordered chronologically

        Examples:
            >>> # Find all releases in the 1980s
            >>> events = kb.search_events_by_date(
            ...     start_year=1980,
            ...     end_year=1989,
            ...     event_type='release'
            ... )
        """
        cursor = self.db_conn.cursor()

        query = """
            SELECT event_id, event_type, title, description,
                   date_extracted, date_normalized, year, month, day,
                   confidence, entities, metadata
            FROM events
            WHERE confidence >= ?
        """
        params = [min_confidence]

        if start_year:
            query += " AND year >= ?"
            params.append(start_year)

        if end_year:
            query += " AND year <= ?"
            params.append(end_year)

        if event_type:
            query += " AND event_type = ?"
            params.append(event_type)

        query += " ORDER BY year, month, day"

        results = cursor.execute(query, params).fetchall()

        events = []
        for row in results:
            events.append({
                'event_id': row[0],
                'event_type': row[1],
                'title': row[2],
                'description': row[3],
                'date_extracted': row[4],
                'date_normalized': row[5],
                'year': row[6],
                'month': row[7],
                'day': row[8],
                'confidence': row[9],
                'entities': json.loads(row[10]) if row[10] else [],
                'metadata': json.loads(row[11]) if row[11] else {}
            })

        return events

    def get_historical_context(self, year: int, context_years: int = 2) -> dict:
        """
        Get historical context for a specific year.

        Returns events from the target year plus surrounding years to provide context.

        Args:
            year: Target year
            context_years: Number of years before/after to include (default: 2)

        Returns:
            Dictionary with:
            - 'target_year': The requested year
            - 'year_range': (start_year, end_year)
            - 'events': List of events in the range
            - 'events_by_year': Events grouped by year
            - 'total_events': Total event count

        Examples:
            >>> # Get context for 1982 (the C64 release year)
            >>> context = kb.get_historical_context(1982, context_years=2)
            >>> # Returns events from 1980-1984
        """
        start_year = year - context_years
        end_year = year + context_years

        events = self.search_events_by_date(start_year=start_year, end_year=end_year)

        # Group events by year
        events_by_year = {}
        for event in events:
            event_year = event['year']
            if event_year not in events_by_year:
                events_by_year[event_year] = []
            events_by_year[event_year].append(event)

        return {
            'target_year': year,
            'year_range': (start_year, end_year),
            'events': events,
            'events_by_year': events_by_year,
            'total_events': len(events)
        }

    def visualize_timeline(self, start_year: Optional[int] = None,
                          end_year: Optional[int] = None,
                          output_path: str = "timeline.html") -> str:
        """
        Create interactive timeline visualization using Plotly.

        Generates a horizontal timeline showing events chronologically with:
        - Color-coded event types
        - Hover information with event details
        - Importance-based marker sizes
        - Zoomable and interactive

        Args:
            start_year: Filter events from this year (optional)
            end_year: Filter events to this year (optional)
            output_path: Path to save HTML file

        Returns:
            Path to saved HTML file

        Examples:
            >>> kb.visualize_timeline(start_year=1980, end_year=1990)
            'timeline.html'
        """
        import plotly.graph_objects as go
        from pathlib import Path

        # Get timeline entries
        timeline = self.get_timeline(start_year=start_year, end_year=end_year)

        if not timeline:
            self.logger.warning("No timeline entries to visualize")
            return ""

        # Color map for event types
        color_map = {
            'release': '#FF6B6B',      # Red
            'milestone': '#4ECDC4',    # Teal
            'innovation': '#95E1D3',   # Mint
            'cultural': '#FFA07A',     # Salmon
            'update': '#9B59B6'        # Purple
        }

        # Prepare data for plotting
        dates = []
        titles = []
        types = []
        importances = []
        confidences = []
        descriptions = []
        colors = []

        for entry in timeline:
            # Parse date for plotting
            date_str = entry['display_date']
            if len(date_str) == 4:  # Year only
                date_str += '-01-01'
            elif len(date_str) == 7:  # Year-month
                date_str += '-01'

            dates.append(date_str)
            titles.append(entry['title'][:100])
            types.append(entry['event_type'])
            importances.append(entry['importance'] * 5)  # Scale for marker size
            confidences.append(entry['confidence'])
            descriptions.append(entry['description'][:200] if entry['description'] else 'No description')
            colors.append(color_map.get(entry['event_type'], '#95A5A6'))

        # Create scatter plot for timeline
        fig = go.Figure()

        # Add trace for each event type
        for event_type in set(types):
            type_indices = [i for i, t in enumerate(types) if t == event_type]

            fig.add_trace(go.Scatter(
                x=[dates[i] for i in type_indices],
                y=[1] * len(type_indices),  # All on same horizontal line
                mode='markers',
                name=event_type.capitalize(),
                marker=dict(
                    size=[importances[i] for i in type_indices],
                    color=color_map.get(event_type, '#95A5A6'),
                    line=dict(width=1, color='white'),
                    symbol='circle'
                ),
                text=[titles[i] for i in type_indices],
                customdata=[[
                    confidences[i],
                    importances[i] / 5,
                    descriptions[i]
                ] for i in type_indices],
                hovertemplate=(
                    '<b>%{text}</b><br>'
                    'Type: ' + event_type + '<br>'
                    'Date: %{x}<br>'
                    'Confidence: %{customdata[0]:.2f}<br>'
                    'Importance: %{customdata[1]}/5<br>'
                    '<extra></extra>'
                )
            ))

        # Update layout
        title_text = f"C64 Knowledge Base Timeline"
        if start_year and end_year:
            title_text += f" ({start_year}-{end_year})"

        fig.update_layout(
            title=title_text,
            xaxis_title="Date",
            yaxis_title="",
            yaxis=dict(
                showgrid=False,
                showticklabels=False,
                range=[0.5, 1.5]
            ),
            hovermode='closest',
            height=600,
            showlegend=True,
            legend=dict(
                yanchor="top",
                y=0.99,
                xanchor="left",
                x=0.01
            )
        )

        # Save to HTML
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        fig.write_html(str(output_file))

        self.logger.info(f"Timeline visualization saved to {output_file}")
        return str(output_file)

    def visualize_event_network(self, start_year: Optional[int] = None,
                                end_year: Optional[int] = None,
                                output_path: str = "event_network.html") -> str:
        """
        Create network visualization showing relationships between events.

        Events are connected based on:
        - Shared entities (people, companies, technologies)
        - Temporal proximity (events in same year)
        - Event type similarity

        Args:
            start_year: Filter events from this year (optional)
            end_year: Filter events to this year (optional)
            output_path: Path to save HTML file

        Returns:
            Path to saved HTML file

        Examples:
            >>> kb.visualize_event_network(start_year=1980, end_year=1990)
            'event_network.html'
        """
        import plotly.graph_objects as go
        from pathlib import Path
        import networkx as nx

        # Get events
        events = self.search_events_by_date(start_year=start_year, end_year=end_year)

        if not events or len(events) < 2:
            self.logger.warning("Need at least 2 events for network visualization")
            return ""

        # Build network graph
        G = nx.Graph()

        # Add nodes (events)
        for i, event in enumerate(events):
            G.add_node(i,
                      title=event['title'][:50],
                      year=event['year'],
                      type=event['event_type'],
                      confidence=event['confidence'],
                      entities=event['entities'])

        # Add edges based on relationships
        for i, event1 in enumerate(events):
            for j, event2 in enumerate(events[i+1:], i+1):
                weight = 0

                # Shared entities
                entities1 = set(event1['entities'])
                entities2 = set(event2['entities'])
                shared = entities1.intersection(entities2)
                if shared:
                    weight += len(shared) * 2

                # Same year
                if event1['year'] == event2['year']:
                    weight += 1

                # Same type
                if event1['event_type'] == event2['event_type']:
                    weight += 0.5

                # Add edge if there's a connection
                if weight > 0:
                    G.add_edge(i, j, weight=weight)

        # Calculate layout
        pos = nx.spring_layout(G, k=0.5, iterations=50)

        # Prepare edge trace
        edge_x = []
        edge_y = []
        for edge in G.edges():
            x0, y0 = pos[edge[0]]
            x1, y1 = pos[edge[1]]
            edge_x.extend([x0, x1, None])
            edge_y.extend([y0, y1, None])

        edge_trace = go.Scatter(
            x=edge_x, y=edge_y,
            line=dict(width=0.5, color='#888'),
            hoverinfo='none',
            mode='lines'
        )

        # Prepare node traces by type
        color_map = {
            'release': '#FF6B6B',
            'milestone': '#4ECDC4',
            'innovation': '#95E1D3',
            'cultural': '#FFA07A',
            'update': '#9B59B6'
        }

        node_traces = []
        for event_type in set(node['type'] for _, node in G.nodes(data=True)):
            node_x = []
            node_y = []
            node_text = []
            node_size = []

            for node_id, node_data in G.nodes(data=True):
                if node_data['type'] == event_type:
                    x, y = pos[node_id]
                    node_x.append(x)
                    node_y.append(y)
                    node_text.append(
                        f"{node_data['title']}<br>"
                        f"Year: {node_data['year']}<br>"
                        f"Type: {node_data['type']}<br>"
                        f"Confidence: {node_data['confidence']:.2f}<br>"
                        f"Connections: {G.degree(node_id)}"
                    )
                    # Node size based on number of connections
                    node_size.append(10 + G.degree(node_id) * 3)

            node_trace = go.Scatter(
                x=node_x, y=node_y,
                mode='markers',
                name=event_type.capitalize(),
                hoverinfo='text',
                text=node_text,
                marker=dict(
                    size=node_size,
                    color=color_map.get(event_type, '#95A5A6'),
                    line=dict(width=2, color='white')
                )
            )
            node_traces.append(node_trace)

        # Create figure
        fig = go.Figure(data=[edge_trace] + node_traces,
                       layout=go.Layout(
                           title=dict(text='Event Network - C64 Knowledge Base', font=dict(size=16)),
                           showlegend=True,
                           hovermode='closest',
                           margin=dict(b=0, l=0, r=0, t=40),
                           xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                           yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                           height=800
                       ))

        # Save to HTML
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        fig.write_html(str(output_file))

        self.logger.info(f"Event network visualization saved to {output_file}")
        return str(output_file)

    def visualize_event_trends(self, output_path: str = "event_trends.html") -> str:
        """
        Create trend charts showing event distribution over time.

        Generates:
        - Events per year bar chart
        - Events by type over time (stacked area chart)
        - Cumulative events over time

        Args:
            output_path: Path to save HTML file

        Returns:
            Path to saved HTML file

        Examples:
            >>> kb.visualize_event_trends()
            'event_trends.html'
        """
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
        from pathlib import Path

        cursor = self.db_conn.cursor()

        # Get events by year and type
        events_data = cursor.execute("""
            SELECT year, event_type, COUNT(*) as count
            FROM events
            WHERE year IS NOT NULL
            GROUP BY year, event_type
            ORDER BY year, event_type
        """).fetchall()

        if not events_data:
            self.logger.warning("No events with dates for trend visualization")
            return ""

        # Organize data
        years_set = set()
        type_year_counts = {}

        for year, event_type, count in events_data:
            years_set.add(year)
            if event_type not in type_year_counts:
                type_year_counts[event_type] = {}
            type_year_counts[event_type][year] = count

        years = sorted(years_set)

        # Create subplots
        fig = make_subplots(
            rows=3, cols=1,
            subplot_titles=(
                'Events Per Year',
                'Events by Type Over Time',
                'Cumulative Events'
            ),
            vertical_spacing=0.12,
            row_heights=[0.33, 0.33, 0.33]
        )

        # 1. Events per year (bar chart)
        year_totals = {}
        for year in years:
            year_totals[year] = sum(
                type_year_counts[et].get(year, 0)
                for et in type_year_counts
            )

        fig.add_trace(
            go.Bar(
                x=years,
                y=[year_totals[y] for y in years],
                name='Total Events',
                marker_color='#3498db',
                showlegend=False
            ),
            row=1, col=1
        )

        # 2. Events by type (stacked area chart)
        color_map = {
            'release': '#FF6B6B',
            'milestone': '#4ECDC4',
            'innovation': '#95E1D3',
            'cultural': '#FFA07A',
            'update': '#9B59B6'
        }

        for event_type in sorted(type_year_counts.keys()):
            type_counts = [type_year_counts[event_type].get(y, 0) for y in years]

            fig.add_trace(
                go.Scatter(
                    x=years,
                    y=type_counts,
                    name=event_type.capitalize(),
                    mode='lines',
                    stackgroup='one',
                    fillcolor=color_map.get(event_type, '#95A5A6'),
                    line=dict(width=0.5, color=color_map.get(event_type, '#95A5A6'))
                ),
                row=2, col=1
            )

        # 3. Cumulative events
        cumulative = []
        total = 0
        for year in years:
            total += year_totals[year]
            cumulative.append(total)

        fig.add_trace(
            go.Scatter(
                x=years,
                y=cumulative,
                name='Cumulative',
                mode='lines+markers',
                line=dict(color='#e74c3c', width=3),
                marker=dict(size=8),
                showlegend=False
            ),
            row=3, col=1
        )

        # Update layout
        fig.update_xaxes(title_text="Year", row=3, col=1)
        fig.update_yaxes(title_text="Count", row=1, col=1)
        fig.update_yaxes(title_text="Count", row=2, col=1)
        fig.update_yaxes(title_text="Total Events", row=3, col=1)

        fig.update_layout(
            title_text="Event Trends - C64 Knowledge Base",
            height=1200,
            showlegend=True,
            hovermode='x unified'
        )

        # Save to HTML
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        fig.write_html(str(output_file))

        self.logger.info(f"Event trends visualization saved to {output_file}")
        return str(output_file)

    def visualize_knowledge_graph_3d(self, max_entities: int = 50,
                                     min_confidence: float = 0.7,
                                     output_path: str = "knowledge_graph_3d.html") -> str:
        """
        Create 3D interactive knowledge graph showing entities, relationships, and documents.

        Args:
            max_entities: Maximum number of entities to include (top by frequency)
            min_confidence: Minimum confidence for entities
            output_path: Output HTML file path

        Returns:
            Path to generated HTML file
        """
        import plotly.graph_objects as go
        from pathlib import Path
        import networkx as nx
        import numpy as np

        cursor = self.db_conn.cursor()

        # Get top entities
        entities_data = cursor.execute("""
            SELECT entity_text, entity_type, COUNT(*) as frequency,
                   AVG(confidence) as avg_confidence
            FROM document_entities
            WHERE confidence >= ?
            GROUP BY entity_text, entity_type
            ORDER BY frequency DESC
            LIMIT ?
        """, (min_confidence, max_entities)).fetchall()

        if not entities_data:
            self.logger.warning("No entities found for 3D knowledge graph")
            return ""

        # Build graph
        G = nx.Graph()

        # Add entity nodes
        entity_map = {}
        for i, (name, entity_type, freq, conf) in enumerate(entities_data):
            G.add_node(f"entity_{i}", label=name, node_type='entity',
                      entity_type=entity_type, frequency=freq, confidence=conf)
            entity_map[name] = f"entity_{i}"

        # Get relationships between entities
        relationships = cursor.execute("""
            SELECT entity1_text, entity2_text, relationship_type, strength, doc_count
            FROM entity_relationships
            WHERE entity1_text IN ({}) AND entity2_text IN ({})
            AND strength >= 0.5
        """.format(','.join('?' * len(entity_map)), ','.join('?' * len(entity_map))),
                                     list(entity_map.keys()) * 2).fetchall()

        # Add edges
        for entity1, entity2, rel_type, strength, co_occur in relationships:
            if entity1 in entity_map and entity2 in entity_map:
                G.add_edge(entity_map[entity1], entity_map[entity2],
                          relationship=rel_type, strength=strength, co_occur=co_occur)

        # Calculate 3D spring layout
        pos_3d = nx.spring_layout(G, dim=3, k=0.5, iterations=50)

        # Extract positions
        x_nodes = [pos_3d[node][0] for node in G.nodes()]
        y_nodes = [pos_3d[node][1] for node in G.nodes()]
        z_nodes = [pos_3d[node][2] for node in G.nodes()]

        # Prepare node data
        node_labels = [G.nodes[node]['label'] for node in G.nodes()]
        node_types = [G.nodes[node]['entity_type'] for node in G.nodes()]
        node_frequencies = [G.nodes[node]['frequency'] for node in G.nodes()]

        # Color map for entity types
        type_color_map = {
            'PERSON': '#FF6B6B',
            'ORG': '#4ECDC4',
            'PRODUCT': '#95E1D3',
            'TECH': '#FFA07A',
            'LOCATION': '#9B59B6'
        }
        node_colors = [type_color_map.get(t, '#95A5A6') for t in node_types]

        # Create edge traces
        edge_x = []
        edge_y = []
        edge_z = []

        for edge in G.edges():
            x0, y0, z0 = pos_3d[edge[0]]
            x1, y1, z1 = pos_3d[edge[1]]
            edge_x.extend([x0, x1, None])
            edge_y.extend([y0, y1, None])
            edge_z.extend([z0, z1, None])

        # Create edge trace
        edge_trace = go.Scatter3d(
            x=edge_x, y=edge_y, z=edge_z,
            mode='lines',
            line=dict(color='rgba(125, 125, 125, 0.3)', width=1),
            hoverinfo='none',
            name='Relationships'
        )

        # Create node trace
        node_trace = go.Scatter3d(
            x=x_nodes, y=y_nodes, z=z_nodes,
            mode='markers+text',
            marker=dict(
                size=[min(30, freq * 2) for freq in node_frequencies],
                color=node_colors,
                line=dict(color='white', width=0.5),
                opacity=0.8
            ),
            text=node_labels,
            textposition='top center',
            textfont=dict(size=8),
            customdata=list(zip(node_types, node_frequencies)),
            hovertemplate=(
                '<b>%{text}</b><br>'
                'Type: %{customdata[0]}<br>'
                'Frequency: %{customdata[1]}<br>'
                '<extra></extra>'
            ),
            name='Entities'
        )

        # Create figure
        fig = go.Figure(data=[edge_trace, node_trace])

        fig.update_layout(
            title=dict(text='3D Knowledge Graph - C64 Knowledge Base', font=dict(size=16)),
            showlegend=True,
            scene=dict(
                xaxis=dict(showgrid=False, showticklabels=False, title=''),
                yaxis=dict(showgrid=False, showticklabels=False, title=''),
                zaxis=dict(showgrid=False, showticklabels=False, title=''),
                bgcolor='rgba(240, 240, 240, 0.9)'
            ),
            margin=dict(l=0, r=0, b=0, t=40),
            height=800
        )

        # Save to HTML
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        fig.write_html(str(output_file))

        self.logger.info(f"3D knowledge graph saved to {output_file} ({len(G.nodes())} nodes, {len(G.edges())} edges)")
        return str(output_file)

    def visualize_hierarchical_bundling(self, max_entities: int = 30,
                                       output_path: str = "hierarchical_bundling.html") -> str:
        """
        Create hierarchical edge bundling visualization showing entity relationships.

        Uses circular layout with entities grouped by type, showing relationships
        with curved bundled edges.

        Args:
            max_entities: Maximum number of entities to include
            output_path: Output HTML file path

        Returns:
            Path to generated HTML file
        """
        import plotly.graph_objects as go
        from pathlib import Path
        import numpy as np

        cursor = self.db_conn.cursor()

        # Get top entities grouped by type
        entities_data = cursor.execute("""
            SELECT entity_text, entity_type, COUNT(*) as frequency,
                   AVG(confidence) as avg_confidence
            FROM document_entities
            WHERE confidence >= 0.7
            GROUP BY entity_text, entity_type
            ORDER BY entity_type, frequency DESC
        """).fetchall()

        # Limit and group by type
        entities_by_type = {}
        for name, entity_type, freq, conf in entities_data:
            if entity_type not in entities_by_type:
                entities_by_type[entity_type] = []
            if len(entities_by_type[entity_type]) < max_entities // len(set(e[1] for e in entities_data[:max_entities])):
                entities_by_type[entity_type].append((name, freq, conf))

        # Calculate circular positions for each type
        entity_positions = {}
        angle = 0
        total_entities = sum(len(entities) for entities in entities_by_type.values())

        if total_entities == 0:
            self.logger.warning("No entities found for hierarchical bundling")
            return ""

        angle_step = 2 * np.pi / total_entities

        for entity_type, entities in entities_by_type.items():
            for name, freq, conf in entities:
                x = np.cos(angle)
                y = np.sin(angle)
                entity_positions[name] = (x, y, entity_type, freq, conf)
                angle += angle_step

        # Get relationships
        entity_names = list(entity_positions.keys())
        relationships = cursor.execute("""
            SELECT entity1_text, entity2_text, relationship_type, strength, doc_count
            FROM entity_relationships
            WHERE entity1_text IN ({}) AND entity2_text IN ({})
            AND strength >= 0.3
        """.format(','.join('?' * len(entity_names)), ','.join('?' * len(entity_names))),
                                     entity_names * 2).fetchall()

        # Create edge traces with bundling effect
        edge_traces = []

        for entity1, entity2, rel_type, strength, co_occur in relationships:
            if entity1 in entity_positions and entity2 in entity_positions:
                x1, y1, type1, freq1, conf1 = entity_positions[entity1]
                x2, y2, type2, freq2, conf2 = entity_positions[entity2]

                # Create curved path (quadratic Bezier curve through origin)
                t = np.linspace(0, 1, 20)
                # Control point at origin for bundling effect
                cx, cy = 0, 0
                x_curve = (1-t)**2 * x1 + 2*(1-t)*t * cx + t**2 * x2
                y_curve = (1-t)**2 * y1 + 2*(1-t)*t * cy + t**2 * y2

                # Color and width based on strength
                color = f'rgba(100, 100, 100, {strength * 0.5})'
                width = max(0.5, strength * 3)

                edge_trace = go.Scatter(
                    x=x_curve, y=y_curve,
                    mode='lines',
                    line=dict(color=color, width=width),
                    hoverinfo='skip',
                    showlegend=False
                )
                edge_traces.append(edge_trace)

        # Create node traces by type
        type_color_map = {
            'PERSON': '#FF6B6B',
            'ORG': '#4ECDC4',
            'PRODUCT': '#95E1D3',
            'TECH': '#FFA07A',
            'LOCATION': '#9B59B6'
        }

        node_traces = []
        for entity_type, color in type_color_map.items():
            type_entities = [(name, pos) for name, pos in entity_positions.items() if pos[2] == entity_type]

            if type_entities:
                x_nodes = [pos[0] for name, pos in type_entities]
                y_nodes = [pos[1] for name, pos in type_entities]
                labels = [name for name, pos in type_entities]
                frequencies = [pos[3] for name, pos in type_entities]

                node_trace = go.Scatter(
                    x=x_nodes, y=y_nodes,
                    mode='markers+text',
                    marker=dict(
                        size=[min(20, f * 2) for f in frequencies],
                        color=color,
                        line=dict(color='white', width=1)
                    ),
                    text=labels,
                    textposition='top center',
                    textfont=dict(size=8),
                    name=entity_type,
                    hovertemplate='<b>%{text}</b><br>Type: ' + entity_type + '<br><extra></extra>'
                )
                node_traces.append(node_trace)

        # Create figure
        fig = go.Figure(data=edge_traces + node_traces)

        fig.update_layout(
            title=dict(text='Hierarchical Edge Bundling - Entity Relationships', font=dict(size=16)),
            showlegend=True,
            xaxis=dict(showgrid=False, zeroline=False, showticklabels=False, range=[-1.5, 1.5]),
            yaxis=dict(showgrid=False, zeroline=False, showticklabels=False, range=[-1.5, 1.5]),
            plot_bgcolor='rgba(240, 240, 240, 0.5)',
            height=800,
            hovermode='closest'
        )

        # Save to HTML
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        fig.write_html(str(output_file))

        self.logger.info(f"Hierarchical bundling visualization saved to {output_file}")
        return str(output_file)

    def visualize_topic_flow_sankey(self, time_period: str = 'decade',
                                   output_path: str = "topic_flow.html") -> str:
        """
        Create Sankey diagram showing flow of topics/entities over time.

        Args:
            time_period: Time grouping ('year' or 'decade')
            output_path: Output HTML file path

        Returns:
            Path to generated HTML file
        """
        import plotly.graph_objects as go
        from pathlib import Path

        cursor = self.db_conn.cursor()

        # Get events with entities
        events_data = cursor.execute("""
            SELECT e.year, e.event_type, e.entities
            FROM events e
            WHERE e.year IS NOT NULL AND e.entities IS NOT NULL
            ORDER BY e.year
        """).fetchall()

        if not events_data:
            self.logger.warning("No events with entities found for Sankey diagram")
            return ""

        # Process data into time periods
        import json

        time_entity_map = {}
        entity_type_map = {}

        for year, event_type, entities_json in events_data:
            # Determine time period
            if time_period == 'decade':
                period = f"{(year // 10) * 10}s"
            else:
                period = str(year)

            # Parse entities
            try:
                entities = json.loads(entities_json)
                for entity in entities:
                    if period not in time_entity_map:
                        time_entity_map[period] = {}

                    if entity not in time_entity_map[period]:
                        time_entity_map[period][entity] = 0
                    time_entity_map[period][entity] += 1

                    # Track entity types
                    if entity not in entity_type_map:
                        entity_type_map[entity] = event_type
            except:
                continue

        if len(time_entity_map) < 2:
            self.logger.warning("Need at least 2 time periods for Sankey diagram")
            return ""

        # Build Sankey data
        nodes = []
        node_map = {}
        node_colors = []

        # Color scheme for time periods
        period_colors = ['#FF6B6B', '#4ECDC4', '#95E1D3', '#FFA07A', '#9B59B6', '#F7DC6F']

        # Create nodes for each time period and their entities
        color_idx = 0
        for period in sorted(time_entity_map.keys()):
            # Get top entities for this period
            top_entities = sorted(time_entity_map[period].items(),
                                key=lambda x: x[1], reverse=True)[:10]

            base_color = period_colors[color_idx % len(period_colors)]

            for entity, count in top_entities:
                node_label = f"{entity} ({period})"
                node_map[node_label] = len(nodes)
                nodes.append(node_label)
                node_colors.append(base_color)

            color_idx += 1

        # Create links between consecutive periods
        links_source = []
        links_target = []
        links_value = []
        links_color = []

        sorted_periods = sorted(time_entity_map.keys())
        for i in range(len(sorted_periods) - 1):
            period1 = sorted_periods[i]
            period2 = sorted_periods[i + 1]

            # Find entities that appear in both periods
            entities1 = set(time_entity_map[period1].keys())
            entities2 = set(time_entity_map[period2].keys())
            common_entities = entities1.intersection(entities2)

            for entity in common_entities:
                node1 = f"{entity} ({period1})"
                node2 = f"{entity} ({period2})"

                if node1 in node_map and node2 in node_map:
                    # Flow value is minimum of the two counts
                    value = min(time_entity_map[period1][entity],
                              time_entity_map[period2][entity])

                    links_source.append(node_map[node1])
                    links_target.append(node_map[node2])
                    links_value.append(value)
                    links_color.append('rgba(100, 100, 100, 0.3)')

        if not links_source:
            self.logger.warning("No entity flows found between periods")
            return ""

        # Create Sankey diagram
        fig = go.Figure(data=[go.Sankey(
            node=dict(
                pad=15,
                thickness=20,
                line=dict(color='white', width=0.5),
                label=nodes,
                color=node_colors
            ),
            link=dict(
                source=links_source,
                target=links_target,
                value=links_value,
                color=links_color
            )
        )])

        fig.update_layout(
            title=dict(text=f'Topic Flow Over Time ({time_period.capitalize()})',
                      font=dict(size=16)),
            font=dict(size=10),
            height=800
        )

        # Save to HTML
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        fig.write_html(str(output_file))

        self.logger.info(f"Sankey diagram saved to {output_file} ({len(nodes)} nodes, {len(links_source)} flows)")
        return str(output_file)

    # ========================================================================
    # Phase 1: Knowledge Graph Builder (v2.24.0)
    # ========================================================================

    def build_knowledge_graph(self, entity_types: Optional[List[str]] = None,
                             min_occurrences: int = 2,
                             min_relationship_strength: float = 0.3,
                             use_cache: bool = True) -> 'nx.Graph':
        """
        Build NetworkX knowledge graph from entities and relationships.

        Creates a weighted graph where:
        - Nodes = entities (with type, occurrences, weight attributes)
        - Edges = relationships (with weight=strength, co_occurrences attributes)

        Args:
            entity_types: Filter to specific entity types (None = all types)
            min_occurrences: Minimum entity occurrences to include (default: 2)
            min_relationship_strength: Minimum relationship strength (0.0-1.0, default: 0.3)
            use_cache: Try to load from cache first (default: True)

        Returns:
            NetworkX Graph with entities as nodes and relationships as edges

        Example:
            >>> kb = KnowledgeBase()
            >>> G = kb.build_knowledge_graph(entity_types=['person', 'org'], min_occurrences=5)
            >>> print(f"Graph: {G.number_of_nodes()} nodes, {G.number_of_edges()} edges")

        Raises:
            ImportError: If NetworkX is not installed
            ValueError: If parameters are invalid
        """
        import hashlib
        from datetime import datetime

        try:
            import networkx as nx
        except ImportError:
            self.logger.error("NetworkX not installed. Run: pip install networkx")
            raise ImportError("NetworkX required for knowledge graph. Install with: pip install networkx>=3.0")

        # Validate parameters
        if min_occurrences < 1:
            raise ValueError("min_occurrences must be >= 1")
        if not 0.0 <= min_relationship_strength <= 1.0:
            raise ValueError("min_relationship_strength must be between 0.0 and 1.0")

        self.logger.info(f"Building knowledge graph (min_occurrences={min_occurrences}, min_strength={min_relationship_strength})")

        cursor = self.db_conn.cursor()

        # Add nodes from entities
        query = """
            SELECT entity_text, entity_type, COUNT(*) as occurrences
            FROM document_entities
            WHERE confidence >= 0.5
            GROUP BY entity_text, entity_type
            HAVING occurrences >= ?
        """
        params = [min_occurrences]

        if entity_types:
            placeholders = ','.join('?' * len(entity_types))
            query += f" AND entity_type IN ({placeholders})"
            params.extend(entity_types)

        query += " ORDER BY occurrences DESC"

        entities = cursor.execute(query, params).fetchall()

        if not entities:
            self.logger.warning("No entities found matching criteria")
            return nx.Graph()

        # Create graph and add nodes
        G = nx.Graph()
        entity_set = set()

        for entity_text, entity_type, occurrences in entities:
            G.add_node(entity_text,
                      type=entity_type,
                      occurrences=occurrences,
                      weight=occurrences)  # Node weight = occurrence count
            entity_set.add(entity_text)

        self.logger.info(f"Added {G.number_of_nodes()} nodes to graph")

        # Add edges from relationships
        rel_query = """
            SELECT entity1_text, entity2_text, strength, doc_count
            FROM entity_relationships
            WHERE strength >= ?
            ORDER BY strength DESC
        """

        relationships = cursor.execute(rel_query, [min_relationship_strength]).fetchall()

        edge_count = 0
        for e1, e2, strength, doc_count in relationships:
            # Only add edge if both entities are in the graph
            if e1 in entity_set and e2 in entity_set:
                G.add_edge(e1, e2,
                          weight=strength,
                          co_occurrences=doc_count)
                edge_count += 1

        self.logger.info(f"Added {edge_count} edges to graph")

        # Validate graph
        if G.number_of_nodes() == 0:
            self.logger.warning("Graph has no nodes")
            return G

        # Log graph statistics
        density = nx.density(G) if G.number_of_nodes() > 1 else 0.0
        num_components = nx.number_connected_components(G)
        largest_cc = max(nx.connected_components(G), key=len) if num_components > 0 else set()

        self.logger.info(f"Graph statistics:")
        self.logger.info(f"  Nodes: {G.number_of_nodes()}")
        self.logger.info(f"  Edges: {G.number_of_edges()}")
        self.logger.info(f"  Density: {density:.4f}")
        self.logger.info(f"  Connected components: {num_components}")
        self.logger.info(f"  Largest component size: {len(largest_cc)}")

        # Cache graph
        if use_cache and G.number_of_nodes() > 0:
            cache_id = self._cache_graph(G)
            self.logger.info(f"Graph cached with ID: {cache_id}")

        return G

    def _cache_graph(self, G: 'nx.Graph') -> str:
        """
        Cache NetworkX graph to database for quick reloading.

        Args:
            G: NetworkX graph to cache

        Returns:
            cache_id: Unique identifier for cached graph

        Example:
            >>> cache_id = kb._cache_graph(G)
            >>> print(f"Cached as: {cache_id}")
        """
        import pickle
        import hashlib
        from datetime import datetime

        # Generate cache ID from graph properties and timestamp
        cache_str = f"{G.number_of_nodes()}_{G.number_of_edges()}_{datetime.now().isoformat()}"
        cache_id = hashlib.sha256(cache_str.encode()).hexdigest()[:16]

        # Serialize graph
        try:
            graph_data = pickle.dumps(G)
        except Exception as e:
            self.logger.error(f"Failed to pickle graph: {e}")
            raise

        # Store to database
        cursor = self.db_conn.cursor()
        cursor.execute("""
            INSERT INTO graph_cache
            (cache_id, graph_version, graph_data, node_count, edge_count, created_date)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (cache_id, 1, graph_data, G.number_of_nodes(),
              G.number_of_edges(), datetime.now().isoformat()))

        self.db_conn.commit()

        self.logger.debug(f"Cached graph {cache_id}: {G.number_of_nodes()} nodes, {G.number_of_edges()} edges")
        return cache_id

    def _load_cached_graph(self, cache_id: str) -> Optional['nx.Graph']:
        """
        Load cached NetworkX graph from database.

        Args:
            cache_id: Unique identifier of cached graph

        Returns:
            NetworkX graph if found, None otherwise

        Example:
            >>> G = kb._load_cached_graph("abc123def456")
            >>> if G:
            >>>     print(f"Loaded: {G.number_of_nodes()} nodes")
        """
        import pickle
        from datetime import datetime

        cursor = self.db_conn.cursor()
        row = cursor.execute("""
            SELECT graph_data FROM graph_cache WHERE cache_id = ?
        """, (cache_id,)).fetchone()

        if not row:
            self.logger.debug(f"Cache miss: {cache_id}")
            return None

        # Update last accessed timestamp
        cursor.execute("""
            UPDATE graph_cache SET last_accessed = ? WHERE cache_id = ?
        """, (datetime.now().isoformat(), cache_id))
        self.db_conn.commit()

        # Deserialize graph
        try:
            G = pickle.loads(row[0])
            self.logger.debug(f"Cache hit: {cache_id} ({G.number_of_nodes()} nodes, {G.number_of_edges()} edges)")
            return G
        except Exception as e:
            self.logger.error(f"Failed to unpickle graph {cache_id}: {e}")
            return None

    def analyze_pagerank(self, G: 'nx.Graph', alpha: float = 0.85,
                        max_iter: int = 100, store_to_db: bool = True) -> Dict[str, float]:
        """
        Calculate PageRank scores for all entities in knowledge graph.

        PageRank identifies the most "important" entities based on their
        connections to other entities in the graph.

        Args:
            G: NetworkX graph (from build_knowledge_graph)
            alpha: Damping parameter (0.0-1.0, default: 0.85)
            max_iter: Maximum iterations for convergence (default: 100)
            store_to_db: Store results to graph_metrics table (default: True)

        Returns:
            Dict mapping entity_text -> PageRank score (sorted by score descending)

        Example:
            >>> G = kb.build_knowledge_graph()
            >>> pagerank = kb.analyze_pagerank(G)
            >>> top_5 = list(pagerank.items())[:5]
            >>> for entity, score in top_5:
            >>>     print(f"{entity}: {score:.4f}")

        Raises:
            ImportError: If NetworkX is not installed
        """
        try:
            import networkx as nx
        except ImportError:
            raise ImportError("NetworkX required. Install with: pip install networkx>=3.0")

        if G.number_of_nodes() == 0:
            self.logger.warning("Cannot compute PageRank on empty graph")
            return {}

        self.logger.info(f"Computing PageRank (alpha={alpha}, max_iter={max_iter})")

        # Calculate PageRank with edge weights
        pagerank = nx.pagerank(G, alpha=alpha, max_iter=max_iter, weight='weight')

        # Sort by score descending
        sorted_pr = sorted(pagerank.items(), key=lambda x: x[1], reverse=True)

        self.logger.info(f"PageRank computed for {len(pagerank)} entities")
        self.logger.info(f"Top entity: {sorted_pr[0][0]} (score: {sorted_pr[0][1]:.6f})")

        # Store to database
        if store_to_db:
            stored = self._store_graph_metrics(pagerank, metric_type='pagerank', G=G)
            self.logger.info(f"Stored PageRank metrics for {stored} entities")

        return dict(sorted_pr)

    def detect_communities(self, G: 'nx.Graph', algorithm: str = 'louvain',
                          store_to_db: bool = True) -> Dict[str, int]:
        """
        Detect communities (clusters) in knowledge graph.

        Communities are groups of entities that are more densely connected
        to each other than to the rest of the graph.

        Args:
            G: NetworkX graph (from build_knowledge_graph)
            algorithm: Detection algorithm:
                - 'louvain': Louvain method (best for large graphs)
                - 'label_propagation': Fast, non-deterministic
                - 'greedy_modularity': Greedy optimization
            store_to_db: Store results to graph_metrics table (default: True)

        Returns:
            Dict mapping entity_text -> community_id

        Example:
            >>> G = kb.build_knowledge_graph()
            >>> communities = kb.detect_communities(G, algorithm='louvain')
            >>> print(f"Found {len(set(communities.values()))} communities")

        Raises:
            ImportError: If NetworkX is not installed
            ValueError: If algorithm is unknown
        """
        try:
            import networkx as nx
        except ImportError:
            raise ImportError("NetworkX required. Install with: pip install networkx>=3.0")

        if G.number_of_nodes() == 0:
            self.logger.warning("Cannot detect communities in empty graph")
            return {}

        self.logger.info(f"Detecting communities using {algorithm} algorithm")

        # Run community detection
        if algorithm == 'louvain':
            communities = nx.community.louvain_communities(G, weight='weight')
        elif algorithm == 'label_propagation':
            communities = nx.community.label_propagation_communities(G)
        elif algorithm == 'greedy_modularity':
            communities = nx.community.greedy_modularity_communities(G, weight='weight')
        else:
            raise ValueError(f"Unknown algorithm: {algorithm}. Use 'louvain', 'label_propagation', or 'greedy_modularity'")

        # Convert to dict: entity -> community_id
        entity_to_community = {}
        for idx, community in enumerate(communities):
            for entity in community:
                entity_to_community[entity] = idx

        num_communities = len(communities)
        avg_size = len(entity_to_community) / num_communities if num_communities > 0 else 0

        self.logger.info(f"Detected {num_communities} communities (avg size: {avg_size:.1f})")

        # Store to database
        if store_to_db:
            stored = self._store_graph_metrics(entity_to_community, metric_type='community', G=G)
            self.logger.info(f"Stored community assignments for {stored} entities")

        return entity_to_community

    def calculate_centrality(self, G: 'nx.Graph', store_to_db: bool = True) -> Dict[str, Dict[str, float]]:
        """
        Calculate multiple centrality measures for all entities.

        Centrality measures identify important/influential entities:
        - Betweenness: Entities that bridge different parts of the graph
        - Closeness: Entities that are close to all other entities
        - Degree: Entities with many direct connections

        Args:
            G: NetworkX graph (from build_knowledge_graph)
            store_to_db: Store results to graph_metrics table (default: True)

        Returns:
            Dict with keys 'betweenness', 'closeness', 'degree', each mapping entity -> score

        Example:
            >>> G = kb.build_knowledge_graph()
            >>> centrality = kb.calculate_centrality(G)
            >>> top_betweenness = sorted(centrality['betweenness'].items(),
            >>>                          key=lambda x: x[1], reverse=True)[:5]

        Raises:
            ImportError: If NetworkX is not installed
        """
        try:
            import networkx as nx
        except ImportError:
            raise ImportError("NetworkX required. Install with: pip install networkx>=3.0")

        if G.number_of_nodes() == 0:
            self.logger.warning("Cannot calculate centrality on empty graph")
            return {'betweenness': {}, 'closeness': {}, 'degree': {}}

        self.logger.info("Calculating centrality measures...")

        centrality = {}

        # Betweenness centrality (weighted)
        self.logger.debug("Computing betweenness centrality")
        centrality['betweenness'] = nx.betweenness_centrality(G, weight='weight')

        # Closeness centrality (weighted as distance)
        self.logger.debug("Computing closeness centrality")
        centrality['closeness'] = nx.closeness_centrality(G, distance='weight')

        # Degree centrality (unweighted)
        self.logger.debug("Computing degree centrality")
        centrality['degree'] = nx.degree_centrality(G)

        self.logger.info(f"Computed 3 centrality measures for {G.number_of_nodes()} entities")

        # Store all metrics to database
        if store_to_db:
            for metric_type, values in centrality.items():
                stored = self._store_graph_metrics(values, metric_type=metric_type, G=G)
                self.logger.debug(f"Stored {metric_type} for {stored} entities")

        return centrality

    def find_shortest_path(self, G: 'nx.Graph', entity1: str, entity2: str,
                          cache_result: bool = True) -> Optional[List[str]]:
        """
        Find shortest path between two entities in knowledge graph.

        Args:
            G: NetworkX graph (from build_knowledge_graph)
            entity1: Source entity name
            entity2: Target entity name
            cache_result: Cache path to database (default: True)

        Returns:
            List of entity names forming the path, or None if no path exists

        Example:
            >>> G = kb.build_knowledge_graph()
            >>> path = kb.find_shortest_path(G, "VIC-II", "sprites")
            >>> if path:
            >>>     print(" -> ".join(path))

        Raises:
            ImportError: If NetworkX is not installed
        """
        try:
            import networkx as nx
        except ImportError:
            raise ImportError("NetworkX required. Install with: pip install networkx>=3.0")

        if entity1 not in G.nodes():
            self.logger.warning(f"Entity not in graph: {entity1}")
            return None

        if entity2 not in G.nodes():
            self.logger.warning(f"Entity not in graph: {entity2}")
            return None

        try:
            # Find shortest path using edge weights
            path = nx.shortest_path(G, entity1, entity2, weight='weight')

            self.logger.info(f"Found path from '{entity1}' to '{entity2}': {len(path)} steps")

            # Cache path to database
            if cache_result:
                self._cache_path(entity1, entity2, path, G=G)

            return path

        except nx.NetworkXNoPath:
            self.logger.info(f"No path found between '{entity1}' and '{entity2}'")
            return None

    def _store_graph_metrics(self, metrics: Dict[str, float], metric_type: str,
                            G: Optional['nx.Graph'] = None) -> int:
        """
        Store graph analysis metrics to database.

        Args:
            metrics: Dict mapping entity_text -> metric_value
            metric_type: Type of metric ('pagerank', 'community', 'betweenness', 'closeness', 'degree')
            G: NetworkX graph (optional, for entity type lookup)

        Returns:
            Number of metrics stored
        """
        from datetime import datetime

        cursor = self.db_conn.cursor()
        stored = 0

        for entity_text, value in metrics.items():
            # Get entity type from graph or database
            entity_type = None
            if G and entity_text in G.nodes():
                entity_type = G.nodes[entity_text].get('type', 'unknown')
            else:
                # Lookup from database
                row = cursor.execute("""
                    SELECT entity_type FROM document_entities
                    WHERE entity_text = ?
                    LIMIT 1
                """, (entity_text,)).fetchone()
                if row:
                    entity_type = row[0]

            if not entity_type:
                entity_type = 'unknown'

            # Generate metric ID
            import hashlib
            metric_id = hashlib.sha256(f"{entity_text}_{metric_type}_{datetime.now().isoformat()}".encode()).hexdigest()[:16]

            # Determine which column to update based on metric type
            if metric_type == 'pagerank':
                cursor.execute("""
                    INSERT OR REPLACE INTO graph_metrics
                    (metric_id, entity_text, entity_type, pagerank, computed_date)
                    VALUES (?, ?, ?, ?, ?)
                """, (metric_id, entity_text, entity_type, value, datetime.now().isoformat()))

            elif metric_type == 'community':
                cursor.execute("""
                    INSERT OR REPLACE INTO graph_metrics
                    (metric_id, entity_text, entity_type, community_id, computed_date)
                    VALUES (?, ?, ?, ?, ?)
                """, (metric_id, entity_text, entity_type, int(value), datetime.now().isoformat()))

            elif metric_type == 'betweenness':
                cursor.execute("""
                    INSERT OR REPLACE INTO graph_metrics
                    (metric_id, entity_text, entity_type, betweenness_centrality, computed_date)
                    VALUES (?, ?, ?, ?, ?)
                """, (metric_id, entity_text, entity_type, value, datetime.now().isoformat()))

            elif metric_type == 'closeness':
                cursor.execute("""
                    INSERT OR REPLACE INTO graph_metrics
                    (metric_id, entity_text, entity_type, closeness_centrality, computed_date)
                    VALUES (?, ?, ?, ?, ?)
                """, (metric_id, entity_text, entity_type, value, datetime.now().isoformat()))

            elif metric_type == 'degree':
                cursor.execute("""
                    INSERT OR REPLACE INTO graph_metrics
                    (metric_id, entity_text, entity_type, degree_centrality, computed_date)
                    VALUES (?, ?, ?, ?, ?)
                """, (metric_id, entity_text, entity_type, value, datetime.now().isoformat()))

            stored += 1

        self.db_conn.commit()
        return stored

    def _cache_path(self, entity1: str, entity2: str, path: List[str],
                   G: Optional['nx.Graph'] = None) -> str:
        """
        Cache shortest path to database.

        Args:
            entity1: Source entity
            entity2: Target entity
            path: List of entities forming the path
            G: NetworkX graph (optional, for weight calculation)

        Returns:
            path_id: Unique identifier for cached path
        """
        import hashlib
        import json
        from datetime import datetime

        # Generate path ID
        path_str = f"{entity1}_{entity2}_{len(path)}_{datetime.now().isoformat()}"
        path_id = hashlib.sha256(path_str.encode()).hexdigest()[:16]

        # Calculate path weight if graph provided
        path_weight = None
        if G:
            path_weight = 0.0
            for i in range(len(path) - 1):
                if G.has_edge(path[i], path[i + 1]):
                    path_weight += G.edges[path[i], path[i + 1]].get('weight', 1.0)

        # Store to database
        cursor = self.db_conn.cursor()
        cursor.execute("""
            INSERT INTO graph_paths
            (path_id, entity1, entity2, path_length, path_nodes, path_weight, computed_date)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (path_id, entity1, entity2, len(path), json.dumps(path),
              path_weight, datetime.now().isoformat()))

        self.db_conn.commit()

        return path_id

    def visualize_knowledge_graph_pyvis(self, G: 'nx.Graph',
                                       output_path: str = "knowledge_graph.html",
                                       color_by: str = "type",
                                       size_by: str = "pagerank",
                                       show_labels: bool = True,
                                       max_nodes: int = 100) -> str:
        """
        Generate interactive HTML visualization of knowledge graph using PyVis.

        Creates an interactive, physics-based graph visualization that can be
        panned, zoomed, and explored in a web browser.

        Args:
            G: NetworkX graph (from build_knowledge_graph)
            output_path: Output HTML file path (default: "knowledge_graph.html")
            color_by: Node coloring attribute:
                - 'type': Color by entity type
                - 'community': Color by community ID
            size_by: Node sizing attribute:
                - 'pagerank': Size by PageRank score
                - 'occurrences': Size by occurrence count
                - 'degree': Size by node degree
            show_labels: Show entity names on nodes (default: True)
            max_nodes: Maximum nodes to display (default: 100)

        Returns:
            Path to generated HTML file

        Example:
            >>> G = kb.build_knowledge_graph()
            >>> pagerank = kb.analyze_pagerank(G)
            >>> communities = kb.detect_communities(G)
            >>> kb.visualize_knowledge_graph_pyvis(G, color_by='community', size_by='pagerank')

        Raises:
            ImportError: If PyVis is not installed
        """
        try:
            from pyvis.network import Network
        except ImportError:
            self.logger.error("PyVis not installed. Run: pip install pyvis")
            raise ImportError("PyVis required for visualization. Install with: pip install pyvis>=0.3.0")

        try:
            import networkx as nx
        except ImportError:
            raise ImportError("NetworkX required. Install with: pip install networkx>=3.0")

        from pathlib import Path

        if G.number_of_nodes() == 0:
            self.logger.warning("Cannot visualize empty graph")
            return ""

        self.logger.info(f"Creating PyVis visualization (color_by={color_by}, size_by={size_by})")

        # Calculate metrics if needed for sizing
        if size_by == "pagerank":
            # Check if PageRank already computed
            sample_node = list(G.nodes())[0]
            if 'pagerank' not in G.nodes[sample_node]:
                self.logger.info("Computing PageRank for node sizing")
                pagerank = self.analyze_pagerank(G, store_to_db=False)
                nx.set_node_attributes(G, pagerank, 'pagerank')

        elif size_by == "degree":
            # Calculate degree centrality
            degree_cent = nx.degree_centrality(G)
            nx.set_node_attributes(G, degree_cent, 'degree')

        # Calculate communities if needed for coloring
        if color_by == "community":
            sample_node = list(G.nodes())[0]
            if 'community' not in G.nodes[sample_node]:
                self.logger.info("Computing communities for node coloring")
                communities = self.detect_communities(G, store_to_db=False)
                nx.set_node_attributes(G, communities, 'community')

        # Filter to top nodes if graph is too large
        if G.number_of_nodes() > max_nodes:
            self.logger.info(f"Filtering graph to top {max_nodes} nodes by degree centrality")
            centrality = nx.degree_centrality(G)
            top_nodes = sorted(centrality.items(), key=lambda x: x[1], reverse=True)[:max_nodes]
            G = G.subgraph([n for n, _ in top_nodes]).copy()

        # Create PyVis network
        net = Network(
            height="750px",
            width="100%",
            bgcolor="#222222",
            font_color="white",
            notebook=False
        )

        # Configure physics for better layout
        net.barnes_hut(
            gravity=-80000,
            central_gravity=0.3,
            spring_length=250,
            spring_strength=0.001,
            damping=0.09
        )

        # Get color mapping
        color_map = self._get_color_map_for_graph(G, color_by)

        # Add nodes with styling
        for node, attrs in G.nodes(data=True):
            # Determine size
            if size_by == "pagerank":
                size = attrs.get('pagerank', 0.001) * 1000  # Scale PageRank
            elif size_by == "occurrences":
                size = attrs.get('occurrences', 1) * 2
            elif size_by == "degree":
                size = attrs.get('degree', 0.01) * 100
            else:
                size = 10

            # Clamp size to reasonable range
            size = max(10, min(50, size))

            # Determine color
            if color_by == "type":
                color_key = attrs.get('type', 'default')
            elif color_by == "community":
                color_key = attrs.get('community', 0)
            else:
                color_key = 'default'

            color = color_map.get(color_key, "#97c2fc")

            # Create tooltip
            tooltip = self._get_node_tooltip(node, attrs)

            # Add node
            net.add_node(
                node,
                label=node if show_labels else "",
                size=size,
                color=color,
                title=tooltip
            )

        # Add edges
        for e1, e2, attrs in G.edges(data=True):
            weight = attrs.get('weight', 0.5)
            co_occurrences = attrs.get('co_occurrences', 0)

            net.add_edge(
                e1, e2,
                value=weight * 5,  # Scale edge width
                title=f"Strength: {weight:.2f}, Co-occurrences: {co_occurrences}"
            )

        # Save to HTML
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        net.save_graph(str(output_file))

        self.logger.info(f"PyVis visualization saved to {output_file} ({G.number_of_nodes()} nodes, {G.number_of_edges()} edges)")
        return str(output_file)

    def _get_color_map_for_graph(self, G: 'nx.Graph', color_by: str) -> Dict[str, str]:
        """
        Get color mapping for graph nodes.

        Args:
            G: NetworkX graph
            color_by: Attribute to color by ('type' or 'community')

        Returns:
            Dict mapping attribute value -> hex color
        """
        if color_by == "type":
            return {
                'hardware': '#e74c3c',       # Red
                'instruction': '#3498db',    # Blue
                'register': '#2ecc71',       # Green
                'memory_address': '#f39c12', # Orange
                'person': '#9b59b6',         # Purple
                'company': '#1abc9c',        # Teal
                'product': '#e67e22',        # Dark orange
                'org': '#16a085',            # Dark teal
                'tech': '#d35400',           # Dark red-orange
                'location': '#8e44ad',       # Dark purple
                'default': '#95a5a6'         # Gray
            }

        elif color_by == "community":
            # Generate distinct colors for communities
            import colorsys
            try:
                import networkx as nx
            except ImportError:
                return {}

            communities = nx.get_node_attributes(G, 'community')
            if not communities:
                return {}

            num_communities = len(set(communities.values()))
            color_map = {}

            for i in range(num_communities):
                hue = i / num_communities
                rgb = colorsys.hsv_to_rgb(hue, 0.8, 0.9)
                hex_color = '#%02x%02x%02x' % tuple(int(c * 255) for c in rgb)
                color_map[i] = hex_color

            return color_map

        return {'default': '#97c2fc'}

    def _get_node_tooltip(self, node: str, attrs: Dict) -> str:
        """
        Generate HTML tooltip for graph node.

        Args:
            node: Node name/text
            attrs: Node attributes dict

        Returns:
            HTML string for tooltip
        """
        lines = [
            f"<b>{node}</b>",
            f"Type: {attrs.get('type', 'unknown')}",
            f"Occurrences: {attrs.get('occurrences', 0)}"
        ]

        if 'pagerank' in attrs:
            lines.append(f"PageRank: {attrs['pagerank']:.6f}")

        if 'community' in attrs:
            lines.append(f"Community: {attrs['community']}")

        if 'betweenness_centrality' in attrs:
            lines.append(f"Betweenness: {attrs['betweenness_centrality']:.4f}")

        if 'degree' in attrs:
            lines.append(f"Degree Centrality: {attrs['degree']:.4f}")

        return "<br>".join(lines)

    def export_graph(self, G: 'nx.Graph', output_path: str,
                    format: str = 'graphml') -> str:
        """
        Export knowledge graph to various formats.

        Args:
            G: NetworkX graph
            output_path: Output file path
            format: Export format:
                - 'graphml': GraphML XML format
                - 'gexf': GEXF XML format (Gephi)
                - 'json': JSON graph format
                - 'gml': GML format

        Returns:
            Path to exported file

        Example:
            >>> G = kb.build_knowledge_graph()
            >>> kb.export_graph(G, "graph.graphml", format="graphml")

        Raises:
            ImportError: If NetworkX is not installed
            ValueError: If format is unknown
        """
        try:
            import networkx as nx
        except ImportError:
            raise ImportError("NetworkX required. Install with: pip install networkx>=3.0")

        from pathlib import Path

        if format not in ['graphml', 'gexf', 'json', 'gml']:
            raise ValueError(f"Unknown format: {format}. Use 'graphml', 'gexf', 'json', or 'gml'")

        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        self.logger.info(f"Exporting graph to {format} format: {output_file}")

        if format == 'graphml':
            nx.write_graphml(G, str(output_file))
        elif format == 'gexf':
            nx.write_gexf(G, str(output_file))
        elif format == 'json':
            from networkx.readwrite import json_graph
            import json
            data = json_graph.node_link_data(G)
            with open(output_file, 'w') as f:
                json.dump(data, f, indent=2)
        elif format == 'gml':
            nx.write_gml(G, str(output_file))

        self.logger.info(f"Graph exported to {output_file} ({G.number_of_nodes()} nodes, {G.number_of_edges()} edges)")
        return str(output_file)

    # ========================================================================
    # Phase 2: Topic Modeling & Clustering (v2.24.0)
    # ========================================================================

    def _prepare_topic_model_corpus(self, min_df: int = 2,
                                   max_df: float = 0.8,
                                   max_features: int = 1000) -> Tuple[List[str], Any, Any]:
        """
        Prepare document corpus for topic modeling.

        Creates TF-IDF matrix from all documents in the knowledge base.

        Args:
            min_df: Minimum document frequency (default: 2)
            max_df: Maximum document frequency as fraction (default: 0.8)
            max_features: Maximum number of features (default: 1000)

        Returns:
            Tuple of (doc_ids, vectorizer, tfidf_matrix)

        Example:
            >>> doc_ids, vectorizer, matrix = kb._prepare_topic_model_corpus()
            >>> print(f"Prepared {len(doc_ids)} documents")
        """
        from sklearn.feature_extraction.text import TfidfVectorizer

        self.logger.info(f"Preparing corpus for topic modeling ({len(self.documents)} documents)")

        # Get all document texts
        docs = []
        doc_ids = []

        for doc_id, doc in self.documents.items():
            chunks = self._get_chunks_db(doc_id)
            if chunks:
                full_text = " ".join(chunk.content for chunk in chunks)
                docs.append(full_text)
                doc_ids.append(doc_id)

        if not docs:
            self.logger.warning("No documents found for topic modeling")
            return [], None, None

        self.logger.info(f"Creating TF-IDF vectorizer (min_df={min_df}, max_df={max_df}, max_features={max_features})")

        # Create TF-IDF vectorizer
        vectorizer = TfidfVectorizer(
            max_df=max_df,
            min_df=min_df,
            stop_words='english',
            max_features=max_features,
            ngram_range=(1, 2),
            lowercase=True,
            strip_accents='unicode'
        )

        try:
            tfidf_matrix = vectorizer.fit_transform(docs)
            self.logger.info(f"TF-IDF matrix: {tfidf_matrix.shape[0]} documents × {tfidf_matrix.shape[1]} features")
            return doc_ids, vectorizer, tfidf_matrix
        except Exception as e:
            self.logger.error(f"Failed to create TF-IDF matrix: {e}")
            raise

    def train_lda_model(self, num_topics: int = 10,
                       max_iter: int = 100,
                       random_state: int = 42,
                       min_doc_prob: float = 0.05) -> Dict[str, Any]:
        """
        Train Latent Dirichlet Allocation (LDA) topic model.

        LDA discovers latent topics in document collection by modeling
        documents as mixtures of topics and topics as mixtures of words.

        Args:
            num_topics: Number of topics to discover (default: 10)
            max_iter: Maximum iterations (default: 100)
            random_state: Random seed for reproducibility (default: 42)
            min_doc_prob: Minimum topic probability to assign document (default: 0.05)

        Returns:
            Dict with model statistics:
                - model_type: 'lda'
                - num_topics: Number of topics
                - topics: List of topic dicts with words and weights
                - perplexity: Model perplexity (lower = better)
                - num_documents: Number of documents processed

        Example:
            >>> results = kb.train_lda_model(num_topics=5)
            >>> for topic in results['topics']:
            >>>     print(f"Topic {topic['topic_number']}: {', '.join(topic['words'][:5])}")

        Raises:
            ImportError: If scikit-learn is not installed
        """
        try:
            from sklearn.decomposition import LatentDirichletAllocation
        except ImportError:
            self.logger.error("scikit-learn not installed. Run: pip install scikit-learn")
            raise ImportError("scikit-learn required for LDA. Install with: pip install scikit-learn")

        import hashlib
        from datetime import datetime

        self.logger.info(f"Training LDA model with {num_topics} topics")

        # Prepare corpus
        doc_ids, vectorizer, tfidf_matrix = self._prepare_topic_model_corpus()

        if not doc_ids:
            return {'error': 'No documents available for topic modeling'}

        # Train LDA
        self.logger.info("Fitting LDA model...")
        lda = LatentDirichletAllocation(
            n_components=num_topics,
            max_iter=max_iter,
            learning_method='online',
            random_state=random_state,
            n_jobs=-1,
            verbose=0
        )

        doc_topic_dist = lda.fit_transform(tfidf_matrix)

        self.logger.info(f"LDA training complete")

        # Extract topics
        feature_names = vectorizer.get_feature_names_out()
        topics = []

        for topic_idx, topic_weights in enumerate(lda.components_):
            # Get top 10 words for this topic
            top_indices = topic_weights.argsort()[-10:][::-1]
            top_words = [feature_names[i] for i in top_indices]
            word_weights = {feature_names[i]: float(topic_weights[i])
                           for i in top_indices}

            # Store topic to database
            topic_id = self._store_topic(
                model_type='lda',
                topic_number=topic_idx,
                top_words=top_words,
                word_weights=word_weights
            )

            topics.append({
                'topic_id': topic_id,
                'topic_number': topic_idx,
                'words': top_words,
                'weights': word_weights
            })

            self.logger.debug(f"Topic {topic_idx}: {', '.join(top_words[:5])}")

        # Assign documents to topics
        assignments = 0
        for doc_idx, doc_id in enumerate(doc_ids):
            topic_probs = doc_topic_dist[doc_idx]

            # Get top 3 topics for this document
            top_topic_indices = topic_probs.argsort()[-3:][::-1]

            for topic_idx in top_topic_indices:
                prob = topic_probs[topic_idx]
                if prob > min_doc_prob:
                    self._assign_document_to_topic(
                        doc_id=doc_id,
                        topic_id=topics[topic_idx]['topic_id'],
                        probability=float(prob),
                        model_type='lda'
                    )
                    assignments += 1

        self.logger.info(f"Created {assignments} document-topic assignments")

        # Calculate perplexity (lower is better)
        perplexity = lda.perplexity(tfidf_matrix)

        self.logger.info(f"LDA model perplexity: {perplexity:.2f}")

        return {
            'model_type': 'lda',
            'num_topics': num_topics,
            'topics': topics,
            'perplexity': perplexity,
            'num_documents': len(doc_ids),
            'num_assignments': assignments
        }

    def _store_topic(self, model_type: str, topic_number: int,
                    top_words: List[str], word_weights: Dict[str, float],
                    coherence_score: Optional[float] = None) -> str:
        """
        Store topic to database.

        Args:
            model_type: Model type ('lda', 'nmf', 'bertopic')
            topic_number: Topic index number
            top_words: List of top words for topic
            word_weights: Dict mapping words to weights
            coherence_score: Optional coherence score

        Returns:
            topic_id: Unique topic identifier
        """
        import hashlib
        import json
        from datetime import datetime

        # Generate topic ID
        topic_str = f"{model_type}_{topic_number}_{datetime.now().isoformat()}"
        topic_id = hashlib.sha256(topic_str.encode()).hexdigest()[:16]

        cursor = self.db_conn.cursor()

        # Store topic
        cursor.execute("""
            INSERT OR REPLACE INTO topics
            (topic_id, model_type, topic_number, top_words, word_weights,
             coherence_score, created_date, num_documents)
            VALUES (?, ?, ?, ?, ?, ?, ?, 0)
        """, (
            topic_id,
            model_type,
            topic_number,
            json.dumps(top_words),
            json.dumps(word_weights),
            coherence_score,
            datetime.now().isoformat()
        ))

        self.db_conn.commit()

        return topic_id

    def _assign_document_to_topic(self, doc_id: str, topic_id: str,
                                  probability: float, model_type: str):
        """
        Assign document to topic with probability.

        Args:
            doc_id: Document ID
            topic_id: Topic ID
            probability: Topic probability for document (0.0-1.0)
            model_type: Model type ('lda', 'nmf', 'bertopic')
        """
        import hashlib
        from datetime import datetime

        # Generate assignment ID
        assignment_str = f"{doc_id}_{topic_id}_{datetime.now().isoformat()}"
        assignment_id = hashlib.sha256(assignment_str.encode()).hexdigest()[:16]

        cursor = self.db_conn.cursor()

        # Store assignment
        cursor.execute("""
            INSERT OR REPLACE INTO document_topics
            (assignment_id, doc_id, topic_id, probability, model_type, assigned_date)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            assignment_id,
            doc_id,
            topic_id,
            probability,
            model_type,
            datetime.now().isoformat()
        ))

        # Update topic document count
        cursor.execute("""
            UPDATE topics
            SET num_documents = (
                SELECT COUNT(DISTINCT doc_id)
                FROM document_topics
                WHERE topic_id = ?
            )
            WHERE topic_id = ?
        """, (topic_id, topic_id))

        self.db_conn.commit()

    def train_nmf_model(self, num_topics: int = 10,
                       max_iter: int = 200,
                       random_state: int = 42,
                       min_doc_prob: float = 0.05) -> Dict[str, Any]:
        """
        Train Non-negative Matrix Factorization (NMF) topic model.

        NMF produces topics by factorizing the document-term matrix into
        two non-negative matrices. Often produces more coherent and
        interpretable topics than LDA.

        Args:
            num_topics: Number of topics to discover (default: 10)
            max_iter: Maximum iterations (default: 200)
            random_state: Random seed for reproducibility (default: 42)
            min_doc_prob: Minimum topic weight to assign document (default: 0.05)

        Returns:
            Dict with model statistics:
                - model_type: 'nmf'
                - num_topics: Number of topics
                - topics: List of topic dicts with words and weights
                - reconstruction_error: Model reconstruction error (lower = better)
                - num_documents: Number of documents processed

        Example:
            >>> results = kb.train_nmf_model(num_topics=5)
            >>> for topic in results['topics']:
            >>>     print(f"Topic {topic['topic_number']}: {', '.join(topic['words'][:5])}")

        Raises:
            ImportError: If scikit-learn is not installed
        """
        try:
            from sklearn.decomposition import NMF
        except ImportError:
            self.logger.error("scikit-learn not installed. Run: pip install scikit-learn")
            raise ImportError("scikit-learn required for NMF. Install with: pip install scikit-learn")

        import hashlib
        from datetime import datetime

        self.logger.info(f"Training NMF model with {num_topics} topics")

        # Prepare corpus
        doc_ids, vectorizer, tfidf_matrix = self._prepare_topic_model_corpus()

        if not doc_ids:
            return {'error': 'No documents available for topic modeling'}

        # Train NMF
        self.logger.info("Fitting NMF model...")
        nmf = NMF(
            n_components=num_topics,
            max_iter=max_iter,
            random_state=random_state,
            init='nndsvd',   # Non-negative Double SVD initialization
            verbose=0
        )

        # W = document-topic matrix, H = topic-word matrix
        doc_topic_matrix = nmf.fit_transform(tfidf_matrix)

        self.logger.info(f"NMF training complete")

        # Extract topics from H matrix (topic-word matrix)
        feature_names = vectorizer.get_feature_names_out()
        topics = []

        for topic_idx, topic_weights in enumerate(nmf.components_):
            # Get top 10 words for this topic
            top_indices = topic_weights.argsort()[-10:][::-1]
            top_words = [feature_names[i] for i in top_indices]
            word_weights = {feature_names[i]: float(topic_weights[i])
                           for i in top_indices}

            # Store topic to database
            topic_id = self._store_topic(
                model_type='nmf',
                topic_number=topic_idx,
                top_words=top_words,
                word_weights=word_weights
            )

            topics.append({
                'topic_id': topic_id,
                'topic_number': topic_idx,
                'words': top_words,
                'weights': word_weights
            })

            self.logger.debug(f"Topic {topic_idx}: {', '.join(top_words[:5])}")

        # Assign documents to topics
        assignments = 0
        for doc_idx, doc_id in enumerate(doc_ids):
            topic_weights = doc_topic_matrix[doc_idx]

            # Normalize to get probabilities
            topic_probs = topic_weights / topic_weights.sum() if topic_weights.sum() > 0 else topic_weights

            # Get top 3 topics for this document
            top_topic_indices = topic_probs.argsort()[-3:][::-1]

            for topic_idx in top_topic_indices:
                prob = topic_probs[topic_idx]
                if prob > min_doc_prob:
                    self._assign_document_to_topic(
                        doc_id=doc_id,
                        topic_id=topics[topic_idx]['topic_id'],
                        probability=float(prob),
                        model_type='nmf'
                    )
                    assignments += 1

        self.logger.info(f"Created {assignments} document-topic assignments")

        # Calculate reconstruction error (lower is better)
        reconstruction_error = nmf.reconstruction_err_

        self.logger.info(f"NMF model reconstruction error: {reconstruction_error:.2f}")

        return {
            'model_type': 'nmf',
            'num_topics': num_topics,
            'topics': topics,
            'reconstruction_error': reconstruction_error,
            'num_documents': len(doc_ids),
            'num_assignments': assignments
        }

    def train_bertopic_model(self, num_topics: int = 10,
                            min_cluster_size: int = 5,
                            random_state: int = 42) -> Dict[str, Any]:
        """
        Train BERTopic model using document embeddings.

        BERTopic is a state-of-the-art topic modeling technique that uses
        embeddings, UMAP dimensionality reduction, and HDBSCAN clustering.
        Produces highly coherent topics.

        Args:
            num_topics: Target number of topics (default: 10)
            min_cluster_size: Minimum documents per topic (default: 5)
            random_state: Random seed for reproducibility (default: 42)

        Returns:
            Dict with model statistics:
                - model_type: 'bertopic'
                - num_topics: Number of topics discovered
                - topics: List of topic dicts with words and scores
                - num_documents: Number of documents processed

        Example:
            >>> results = kb.train_bertopic_model(num_topics=5)
            >>> for topic in results['topics']:
            >>>     print(f"Topic {topic['topic_number']}: {', '.join(topic['words'][:5])}")

        Raises:
            ImportError: If bertopic, umap-learn, or hdbscan not installed
        """
        try:
            from bertopic import BERTopic
            from umap import UMAP
            from hdbscan import HDBSCAN
        except ImportError as e:
            missing = str(e).split("'")[1]
            self.logger.error(f"{missing} not installed. Run: pip install bertopic umap-learn hdbscan")
            raise ImportError(f"BERTopic dependencies required. Install with: pip install bertopic umap-learn hdbscan")

        import numpy as np
        from datetime import datetime

        self.logger.info(f"Training BERTopic model (target topics: {num_topics})")

        # Ensure embeddings model is loaded
        if self.embeddings_model is None:
            self._ensure_embeddings_loaded()

        # Get documents and generate/retrieve embeddings
        doc_ids = []
        documents = []

        for doc_id, doc in self.documents.items():
            chunks = self._get_chunks_db(doc_id)
            if chunks:
                doc_ids.append(doc_id)
                # Use first chunk as document representation
                documents.append(chunks[0].content[:500])  # Limit length for performance

        if not doc_ids:
            self.logger.warning("No documents found")
            return {'error': 'No documents available for topic modeling'}

        # Generate embeddings for all documents
        self.logger.info(f"Generating embeddings for {len(documents)} documents...")
        embeddings = self.embeddings_model.encode(documents, show_progress_bar=False)

        embeddings = np.array(embeddings)
        self.logger.info(f"Loaded {len(doc_ids)} documents with embeddings (shape: {embeddings.shape})")

        # Configure UMAP for dimensionality reduction
        umap_model = UMAP(
            n_neighbors=15,
            n_components=5,
            min_dist=0.0,
            metric='cosine',
            random_state=random_state
        )

        # Configure HDBSCAN for clustering
        hdbscan_model = HDBSCAN(
            min_cluster_size=min_cluster_size,
            metric='euclidean',
            cluster_selection_method='eom',
            prediction_data=True
        )

        # Create BERTopic model
        self.logger.info("Fitting BERTopic model...")
        topic_model = BERTopic(
            umap_model=umap_model,
            hdbscan_model=hdbscan_model,
            nr_topics=num_topics,
            verbose=False,
            calculate_probabilities=True
        )

        # Train model
        topic_assignments, probs = topic_model.fit_transform(documents, embeddings)

        # Get topic info
        topic_info = topic_model.get_topic_info()
        num_discovered_topics = len(topic_info) - 1  # -1 for outlier topic (-1)

        self.logger.info(f"BERTopic discovered {num_discovered_topics} topics")

        # Extract and store topics
        topics = []
        topic_id_map = {}  # Map BERTopic topic numbers to our topic_ids

        for idx, row in topic_info.iterrows():
            topic_num = row['Topic']
            if topic_num == -1:  # Skip outlier topic
                continue

            # Get top words for this topic
            topic_words = topic_model.get_topic(topic_num)
            if not topic_words:
                continue

            top_words = [word for word, _ in topic_words[:10]]
            word_weights = {word: float(score) for word, score in topic_words[:10]}

            # Store topic to database
            topic_id = self._store_topic(
                model_type='bertopic',
                topic_number=int(topic_num),
                top_words=top_words,
                word_weights=word_weights
            )

            topic_id_map[topic_num] = topic_id

            topics.append({
                'topic_id': topic_id,
                'topic_number': int(topic_num),
                'words': top_words,
                'weights': word_weights,
                'document_count': int(row['Count'])
            })

            self.logger.debug(f"Topic {topic_num}: {', '.join(top_words[:5])}")

        # Assign documents to topics
        assignments = 0
        for doc_idx, (doc_id, topic_num, prob) in enumerate(zip(doc_ids, topic_assignments, probs)):
            if topic_num == -1:  # Skip outlier documents
                continue

            if topic_num in topic_id_map:
                # For BERTopic, prob might be an array, take max
                if isinstance(prob, (list, np.ndarray)):
                    probability = float(np.max(prob))
                else:
                    probability = float(prob)

                self._assign_document_to_topic(
                    doc_id=doc_id,
                    topic_id=topic_id_map[topic_num],
                    probability=probability,
                    model_type='bertopic'
                )
                assignments += 1

        self.logger.info(f"Created {assignments} document-topic assignments")

        # Count outliers (topic -1)
        outlier_count = int(np.sum(np.array(topic_assignments) == -1))

        return {
            'model_type': 'bertopic',
            'num_topics': num_discovered_topics,
            'topics': topics,
            'num_documents': len(doc_ids),
            'num_assignments': assignments,
            'outliers': outlier_count
        }

    # ========================================
    # Phase 2: Document Clustering
    # ========================================

    def _store_cluster(self, algorithm: str, cluster_number: int,
                      centroid_vector: Optional[np.ndarray] = None,
                      silhouette_score: Optional[float] = None) -> str:
        """
        Store cluster to database.

        Args:
            algorithm: Clustering algorithm ('kmeans', 'dbscan', 'hdbscan')
            cluster_number: Cluster index number
            centroid_vector: Optional cluster centroid (for kmeans)
            silhouette_score: Optional silhouette score

        Returns:
            cluster_id: Unique cluster identifier
        """
        import hashlib
        import json
        from datetime import datetime

        # Generate cluster ID
        cluster_str = f"{algorithm}_{cluster_number}_{datetime.now().isoformat()}"
        cluster_id = hashlib.sha256(cluster_str.encode()).hexdigest()[:16]

        cursor = self.db_conn.cursor()

        # Convert centroid to JSON if provided
        centroid_json = json.dumps(centroid_vector.tolist()) if centroid_vector is not None else None

        # Store cluster
        cursor.execute("""
            INSERT OR REPLACE INTO clusters
            (cluster_id, algorithm, cluster_number, centroid_vector,
             silhouette_score, created_date, num_documents)
            VALUES (?, ?, ?, ?, ?, ?, 0)
        """, (
            cluster_id,
            algorithm,
            cluster_number,
            centroid_json,
            silhouette_score,
            datetime.now().isoformat()
        ))

        self.db_conn.commit()

        return cluster_id

    def _assign_document_to_cluster(self, doc_id: str, cluster_id: str,
                                    distance: float, algorithm: str):
        """
        Assign document to cluster with distance metric.

        Args:
            doc_id: Document ID
            cluster_id: Cluster ID
            distance: Distance from cluster centroid
            algorithm: Clustering algorithm
        """
        import hashlib
        from datetime import datetime

        cursor = self.db_conn.cursor()

        # Generate assignment ID
        assignment_str = f"{doc_id}_{cluster_id}_{datetime.now().isoformat()}"
        assignment_id = hashlib.sha256(assignment_str.encode()).hexdigest()[:16]

        cursor.execute("""
            INSERT OR REPLACE INTO document_clusters
            (assignment_id, doc_id, cluster_id, distance, algorithm, assigned_date)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (assignment_id, doc_id, cluster_id, distance, algorithm, datetime.now().isoformat()))

        self.db_conn.commit()

    def cluster_documents_kmeans(self, num_clusters: int = 10,
                                 random_state: int = 42) -> Dict[str, Any]:
        """
        Cluster documents using K-Means on embeddings.

        K-Means is a centroid-based clustering algorithm that partitions
        documents into K clusters based on embedding similarity.

        Args:
            num_clusters: Number of clusters (K) (default: 10)
            random_state: Random seed for reproducibility (default: 42)

        Returns:
            Dict with clustering statistics:
                - algorithm: 'kmeans'
                - num_clusters: Number of clusters
                - silhouette_score: Clustering quality metric (-1 to 1)
                - num_documents: Number of documents clustered

        Example:
            >>> results = kb.cluster_documents_kmeans(num_clusters=5)
            >>> print(f"Clustered {results['num_documents']} docs into 5 clusters")
            >>> print(f"Silhouette score: {results['silhouette_score']:.3f}")

        Raises:
            ImportError: If scikit-learn is not installed
        """
        try:
            from sklearn.cluster import KMeans
            from sklearn.metrics import silhouette_score
        except ImportError:
            self.logger.error("scikit-learn not installed. Run: pip install scikit-learn")
            raise ImportError("scikit-learn required for K-Means. Install with: pip install scikit-learn")

        import numpy as np

        self.logger.info(f"K-Means clustering with {num_clusters} clusters")

        # Ensure embeddings model is loaded
        if self.embeddings_model is None:
            self._ensure_embeddings_loaded()

        # Get documents and generate embeddings
        doc_ids = []
        documents = []

        for doc_id, doc in self.documents.items():
            chunks = self._get_chunks_db(doc_id)
            if chunks:
                doc_ids.append(doc_id)
                documents.append(chunks[0].content[:500])

        if not doc_ids:
            return {'error': 'No documents available for clustering'}

        # Generate embeddings
        self.logger.info(f"Generating embeddings for {len(documents)} documents...")
        embeddings = self.embeddings_model.encode(documents, show_progress_bar=False)
        embeddings = np.array(embeddings)

        # Train K-Means
        self.logger.info("Fitting K-Means model...")
        kmeans = KMeans(n_clusters=num_clusters, random_state=random_state, n_init=10)
        labels = kmeans.fit_predict(embeddings)

        # Calculate silhouette score
        silhouette = silhouette_score(embeddings, labels)
        self.logger.info(f"K-Means silhouette score: {silhouette:.3f}")

        # Store clusters and assign documents
        assignments = 0
        for cluster_num in range(num_clusters):
            # Find documents in this cluster
            cluster_doc_indices = [i for i, label in enumerate(labels) if label == cluster_num]

            # Store cluster
            cluster_id = self._store_cluster(
                algorithm='kmeans',
                cluster_number=cluster_num,
                centroid_vector=kmeans.cluster_centers_[cluster_num],
                silhouette_score=silhouette
            )

            # Assign documents to cluster
            for doc_idx in cluster_doc_indices:
                doc_id = doc_ids[doc_idx]
                distance = np.linalg.norm(embeddings[doc_idx] - kmeans.cluster_centers_[cluster_num])
                self._assign_document_to_cluster(doc_id, cluster_id, float(distance), 'kmeans')
                assignments += 1

        self.logger.info(f"Created {assignments} document-cluster assignments")

        return {
            'algorithm': 'kmeans',
            'num_clusters': num_clusters,
            'silhouette_score': float(silhouette),
            'num_documents': len(doc_ids),
            'num_assignments': assignments
        }

    def cluster_documents_dbscan(self, eps: float = 0.5,
                                 min_samples: int = 5) -> Dict[str, Any]:
        """
        Cluster documents using DBSCAN on embeddings.

        DBSCAN is a density-based clustering algorithm that finds
        arbitrary-shaped clusters and identifies outliers. Does not
        require specifying the number of clusters.

        Args:
            eps: Maximum distance between samples (default: 0.5)
            min_samples: Minimum samples in neighborhood (default: 5)

        Returns:
            Dict with clustering statistics:
                - algorithm: 'dbscan'
                - num_clusters: Number of clusters found
                - silhouette_score: Clustering quality metric
                - num_documents: Number of documents clustered
                - num_outliers: Number of outlier documents

        Example:
            >>> results = kb.cluster_documents_dbscan(eps=0.5, min_samples=5)
            >>> print(f"Found {results['num_clusters']} clusters")
            >>> print(f"Outliers: {results['num_outliers']}")

        Raises:
            ImportError: If scikit-learn is not installed
        """
        try:
            from sklearn.cluster import DBSCAN
            from sklearn.metrics import silhouette_score
        except ImportError:
            self.logger.error("scikit-learn not installed. Run: pip install scikit-learn")
            raise ImportError("scikit-learn required for DBSCAN. Install with: pip install scikit-learn")

        import numpy as np

        self.logger.info(f"DBSCAN clustering (eps={eps}, min_samples={min_samples})")

        # Ensure embeddings model is loaded
        if self.embeddings_model is None:
            self._ensure_embeddings_loaded()

        # Get documents and generate embeddings
        doc_ids = []
        documents = []

        for doc_id, doc in self.documents.items():
            chunks = self._get_chunks_db(doc_id)
            if chunks:
                doc_ids.append(doc_id)
                documents.append(chunks[0].content[:500])

        if not doc_ids:
            return {'error': 'No documents available for clustering'}

        # Generate embeddings
        self.logger.info(f"Generating embeddings for {len(documents)} documents...")
        embeddings = self.embeddings_model.encode(documents, show_progress_bar=False)
        embeddings = np.array(embeddings)

        # Train DBSCAN
        self.logger.info("Fitting DBSCAN model...")
        dbscan = DBSCAN(eps=eps, min_samples=min_samples, metric='cosine')
        labels = dbscan.fit_predict(embeddings)

        # Count clusters and outliers
        unique_labels = set(labels)
        num_clusters = len(unique_labels) - (1 if -1 in unique_labels else 0)
        num_outliers = int((labels == -1).sum())

        self.logger.info(f"DBSCAN found {num_clusters} clusters and {num_outliers} outliers")

        # Calculate silhouette score (excluding outliers)
        if num_clusters > 1 and num_outliers < len(labels):
            # Filter out outliers for silhouette calculation
            non_outlier_mask = labels != -1
            if non_outlier_mask.sum() > 1:
                silhouette = silhouette_score(
                    embeddings[non_outlier_mask],
                    labels[non_outlier_mask]
                )
            else:
                silhouette = 0.0
        else:
            silhouette = 0.0

        self.logger.info(f"DBSCAN silhouette score: {silhouette:.3f}")

        # Store clusters and assign documents
        assignments = 0
        cluster_centroids = {}

        for cluster_num in unique_labels:
            if cluster_num == -1:
                continue  # Skip outliers

            # Find documents in this cluster
            cluster_doc_indices = [i for i, label in enumerate(labels) if label == cluster_num]

            # Calculate cluster centroid
            cluster_embeddings = embeddings[cluster_doc_indices]
            centroid = np.mean(cluster_embeddings, axis=0)
            cluster_centroids[cluster_num] = centroid

            # Store cluster
            cluster_id = self._store_cluster(
                algorithm='dbscan',
                cluster_number=cluster_num,
                centroid_vector=centroid,
                silhouette_score=silhouette
            )

            # Assign documents to cluster
            for doc_idx in cluster_doc_indices:
                doc_id = doc_ids[doc_idx]
                distance = np.linalg.norm(embeddings[doc_idx] - centroid)
                self._assign_document_to_cluster(doc_id, cluster_id, float(distance), 'dbscan')
                assignments += 1

        self.logger.info(f"Created {assignments} document-cluster assignments")

        return {
            'algorithm': 'dbscan',
            'num_clusters': num_clusters,
            'silhouette_score': float(silhouette),
            'num_documents': len(doc_ids),
            'num_assignments': assignments,
            'num_outliers': num_outliers
        }

    def cluster_documents_hdbscan(self, min_cluster_size: int = 5,
                                  min_samples: Optional[int] = None) -> Dict[str, Any]:
        """
        Cluster documents using HDBSCAN on embeddings.

        HDBSCAN (Hierarchical DBSCAN) is an advanced density-based
        clustering algorithm that automatically selects clusters and
        handles varying densities better than DBSCAN.

        Args:
            min_cluster_size: Minimum samples per cluster (default: 5)
            min_samples: Minimum samples in neighborhood (default: None = min_cluster_size)

        Returns:
            Dict with clustering statistics:
                - algorithm: 'hdbscan'
                - num_clusters: Number of clusters found
                - num_documents: Number of documents clustered
                - num_outliers: Number of outlier documents

        Example:
            >>> results = kb.cluster_documents_hdbscan(min_cluster_size=5)
            >>> print(f"Found {results['num_clusters']} clusters")
            >>> print(f"Outliers: {results['num_outliers']}")

        Raises:
            ImportError: If hdbscan is not installed
        """
        try:
            import hdbscan
        except ImportError:
            self.logger.error("hdbscan not installed. Run: pip install hdbscan")
            raise ImportError("hdbscan required for HDBSCAN. Install with: pip install hdbscan")

        import numpy as np

        self.logger.info(f"HDBSCAN clustering (min_cluster_size={min_cluster_size})")

        # Ensure embeddings model is loaded
        if self.embeddings_model is None:
            self._ensure_embeddings_loaded()

        # Get documents and generate embeddings
        doc_ids = []
        documents = []

        for doc_id, doc in self.documents.items():
            chunks = self._get_chunks_db(doc_id)
            if chunks:
                doc_ids.append(doc_id)
                documents.append(chunks[0].content[:500])

        if not doc_ids:
            return {'error': 'No documents available for clustering'}

        # Generate embeddings
        self.logger.info(f"Generating embeddings for {len(documents)} documents...")
        embeddings = self.embeddings_model.encode(documents, show_progress_bar=False)
        embeddings = np.array(embeddings)

        # Train HDBSCAN
        self.logger.info("Fitting HDBSCAN model...")
        clusterer = hdbscan.HDBSCAN(
            min_cluster_size=min_cluster_size,
            min_samples=min_samples,
            metric='euclidean',
            cluster_selection_method='eom',
            prediction_data=True
        )
        labels = clusterer.fit_predict(embeddings)

        # Count clusters and outliers
        unique_labels = set(labels)
        num_clusters = len(unique_labels) - (1 if -1 in unique_labels else 0)
        num_outliers = int((labels == -1).sum())

        self.logger.info(f"HDBSCAN found {num_clusters} clusters and {num_outliers} outliers")

        # Store clusters and assign documents
        assignments = 0
        cluster_centroids = {}

        for cluster_num in unique_labels:
            if cluster_num == -1:
                continue  # Skip outliers

            # Find documents in this cluster
            cluster_doc_indices = [i for i, label in enumerate(labels) if label == cluster_num]

            # Calculate cluster centroid
            cluster_embeddings = embeddings[cluster_doc_indices]
            centroid = np.mean(cluster_embeddings, axis=0)
            cluster_centroids[cluster_num] = centroid

            # Store cluster
            cluster_id = self._store_cluster(
                algorithm='hdbscan',
                cluster_number=cluster_num,
                centroid_vector=centroid,
                silhouette_score=None  # HDBSCAN uses different metrics
            )

            # Assign documents to cluster
            for doc_idx in cluster_doc_indices:
                doc_id = doc_ids[doc_idx]
                distance = np.linalg.norm(embeddings[doc_idx] - centroid)
                self._assign_document_to_cluster(doc_id, cluster_id, float(distance), 'hdbscan')
                assignments += 1

        self.logger.info(f"Created {assignments} document-cluster assignments")

        return {
            'algorithm': 'hdbscan',
            'num_clusters': num_clusters,
            'num_documents': len(doc_ids),
            'num_assignments': assignments,
            'num_outliers': num_outliers,
            'cluster_persistence': clusterer.cluster_persistence_.tolist() if hasattr(clusterer, 'cluster_persistence_') else []
        }

    def generate_topic_wordcloud(self, topic_id: str, output_path: str,
                                 width: int = 800, height: int = 400,
                                 background_color: str = 'white') -> Dict[str, Any]:
        """
        Generate a word cloud visualization for a topic.

        Args:
            topic_id: Topic ID to visualize
            output_path: Path to save the word cloud image
            width: Image width in pixels
            height: Image height in pixels
            background_color: Background color for word cloud

        Returns:
            {
                'topic_id': str,
                'topic_number': int,
                'model_type': str,
                'output_path': str,
                'num_words': int
            }

        Examples:
            # Generate word cloud for a topic
            result = kb.generate_topic_wordcloud(
                'topic-123',
                'topic_wordcloud.png',
                width=1200,
                height=600
            )
        """
        from wordcloud import WordCloud
        import matplotlib.pyplot as plt

        cursor = self.db_conn.cursor()

        # Get topic details
        topic_row = cursor.execute(
            "SELECT topic_number, model_type, top_words, word_weights FROM topics WHERE topic_id = ?",
            (topic_id,)
        ).fetchone()

        if not topic_row:
            return {'error': f'Topic {topic_id} not found'}

        topic_number, model_type, top_words_json, weights_json = topic_row

        # Parse word weights
        import json
        weights = json.loads(weights_json)

        # Create word cloud
        wordcloud = WordCloud(
            width=width,
            height=height,
            background_color=background_color,
            colormap='viridis',
            relative_scaling=0.5
        ).generate_from_frequencies(weights)

        # Save to file
        plt.figure(figsize=(width/100, height/100), dpi=100)
        plt.imshow(wordcloud, interpolation='bilinear')
        plt.axis('off')
        plt.title(f'Topic {topic_number} ({model_type.upper()})', fontsize=16)
        plt.tight_layout(pad=0)
        plt.savefig(output_path, dpi=100, bbox_inches='tight')
        plt.close()

        return {
            'topic_id': topic_id,
            'topic_number': topic_number,
            'model_type': model_type,
            'output_path': output_path,
            'num_words': len(weights)
        }

    def visualize_cluster_scatter(self, algorithm: str, output_path: str,
                                  width: int = 1200, height: int = 800,
                                  n_neighbors: int = 15,
                                  min_dist: float = 0.1) -> Dict[str, Any]:
        """
        Generate 2D scatter plot of document clusters using UMAP projection.

        Args:
            algorithm: Clustering algorithm ('kmeans', 'dbscan', 'hdbscan')
            output_path: Path to save the scatter plot
            width: Image width in pixels
            height: Image height in pixels
            n_neighbors: UMAP n_neighbors parameter (controls local vs global structure)
            min_dist: UMAP min_dist parameter (controls cluster tightness)

        Returns:
            {
                'algorithm': str,
                'num_clusters': int,
                'num_documents': int,
                'output_path': str
            }

        Examples:
            # Visualize DBSCAN clusters
            result = kb.visualize_cluster_scatter('dbscan', 'clusters.png')
        """
        import umap
        import matplotlib.pyplot as plt
        import numpy as np

        cursor = self.db_conn.cursor()

        # Get clusters for this algorithm
        clusters = cursor.execute(
            "SELECT cluster_id, cluster_number FROM clusters WHERE algorithm = ?",
            (algorithm,)
        ).fetchall()

        if not clusters:
            return {'error': f'No clusters found for algorithm {algorithm}'}

        cluster_map = {cid: cnum for cid, cnum in clusters}

        # Get document embeddings and cluster assignments
        doc_data = []
        for doc_id, doc in self.documents.items():
            # Get cluster assignment
            assignment = cursor.execute(
                """SELECT c.cluster_number FROM document_clusters dc
                   JOIN clusters c ON dc.cluster_id = c.cluster_id
                   WHERE dc.doc_id = ? AND c.algorithm = ?""",
                (doc_id, algorithm)
            ).fetchone()

            if assignment:
                cluster_num = assignment[0]
                doc_data.append((doc_id, doc.title, cluster_num))

        if not doc_data:
            return {'error': f'No document assignments found for algorithm {algorithm}'}

        # Extract embeddings for assigned documents
        doc_ids = [d[0] for d in doc_data]
        titles = [d[1] for d in doc_data]
        cluster_labels = np.array([d[2] for d in doc_data])

        # Generate embeddings
        documents = [self.documents[doc_id] for doc_id in doc_ids]
        doc_texts = []
        for doc in documents:
            chunks = self._get_chunks_db(doc.doc_id)[:3]
            chunk_texts = [chunk.content for chunk in chunks]
            doc_text = doc.title + " " + " ".join(chunk_texts)
            doc_texts.append(doc_text)

        # Ensure embeddings model is loaded
        self._ensure_embeddings_loaded()
        embeddings = self.embeddings_model.encode(doc_texts, show_progress_bar=False)

        # Apply UMAP dimensionality reduction
        reducer = umap.UMAP(
            n_neighbors=n_neighbors,
            min_dist=min_dist,
            n_components=2,
            random_state=42,
            metric='cosine'
        )
        embedding_2d = reducer.fit_transform(embeddings)

        # Create scatter plot
        plt.figure(figsize=(width/100, height/100), dpi=100)

        # Get unique clusters (including outliers as -1)
        unique_clusters = np.unique(cluster_labels)
        colors = plt.cm.tab20(np.linspace(0, 1, len(unique_clusters)))

        for i, cluster_num in enumerate(unique_clusters):
            mask = cluster_labels == cluster_num
            label = f'Outliers' if cluster_num == -1 else f'Cluster {cluster_num}'

            plt.scatter(
                embedding_2d[mask, 0],
                embedding_2d[mask, 1],
                c=[colors[i]],
                label=label,
                alpha=0.6,
                s=50
            )

        plt.title(f'Document Clusters ({algorithm.upper()})', fontsize=16)
        plt.xlabel('UMAP Dimension 1', fontsize=12)
        plt.ylabel('UMAP Dimension 2', fontsize=12)
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        plt.tight_layout()
        plt.savefig(output_path, dpi=100, bbox_inches='tight')
        plt.close()

        return {
            'algorithm': algorithm,
            'num_clusters': len(unique_clusters),
            'num_documents': len(doc_data),
            'output_path': output_path
        }

    def generate_topic_heatmap(self, model_type: str, output_path: str,
                               max_topics: int = 20,
                               max_documents: int = 50) -> Dict[str, Any]:
        """
        Generate heatmap showing document-topic probability matrix.

        Args:
            model_type: Topic model type ('lda', 'nmf', 'bertopic')
            output_path: Path to save the heatmap
            max_topics: Maximum number of topics to include
            max_documents: Maximum number of documents to include

        Returns:
            {
                'model_type': str,
                'num_topics': int,
                'num_documents': int,
                'output_path': str
            }

        Examples:
            # Generate heatmap for LDA model
            result = kb.generate_topic_heatmap('lda', 'topic_heatmap.png')
        """
        import matplotlib.pyplot as plt
        import numpy as np
        import pandas as pd

        cursor = self.db_conn.cursor()

        # Get topics for this model
        topics = cursor.execute(
            "SELECT topic_id, topic_number FROM topics WHERE model_type = ? ORDER BY topic_number LIMIT ?",
            (model_type, max_topics)
        ).fetchall()

        if not topics:
            return {'error': f'No topics found for model type {model_type}'}

        topic_ids = [t[0] for t in topics]
        topic_numbers = [t[1] for t in topics]

        # Get document-topic assignments
        # Get top documents by average probability across topics
        doc_assignments = cursor.execute(
            f"""SELECT dt.doc_id, d.title, AVG(dt.probability) as avg_prob
               FROM document_topics dt
               JOIN documents d ON dt.doc_id = d.doc_id
               JOIN topics t ON dt.topic_id = t.topic_id
               WHERE t.model_type = ?
               GROUP BY dt.doc_id
               ORDER BY avg_prob DESC
               LIMIT ?""",
            (model_type, max_documents)
        ).fetchall()

        if not doc_assignments:
            return {'error': f'No document-topic assignments found for {model_type}'}

        doc_ids = [d[0] for d in doc_assignments]
        doc_titles = [d[1][:40] for d in doc_assignments]  # Truncate titles

        # Build probability matrix
        matrix = np.zeros((len(doc_ids), len(topic_ids)))

        for i, doc_id in enumerate(doc_ids):
            for j, topic_id in enumerate(topic_ids):
                prob = cursor.execute(
                    "SELECT probability FROM document_topics WHERE doc_id = ? AND topic_id = ?",
                    (doc_id, topic_id)
                ).fetchone()
                if prob:
                    matrix[i, j] = prob[0]

        # Create heatmap
        plt.figure(figsize=(12, max(8, len(doc_ids) * 0.3)))

        plt.imshow(matrix, cmap='YlOrRd', aspect='auto', interpolation='nearest')
        plt.colorbar(label='Topic Probability')

        # Set ticks
        plt.xticks(range(len(topic_numbers)), [f'T{n}' for n in topic_numbers], rotation=0)
        plt.yticks(range(len(doc_titles)), doc_titles, fontsize=8)

        plt.xlabel('Topics', fontsize=12)
        plt.ylabel('Documents', fontsize=12)
        plt.title(f'Document-Topic Heatmap ({model_type.upper()})', fontsize=14)

        plt.tight_layout()
        plt.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close()

        return {
            'model_type': model_type,
            'num_topics': len(topic_ids),
            'num_documents': len(doc_ids),
            'output_path': output_path
        }

    def visualize_cluster_distribution(self, algorithm: str, output_path: str,
                                       width: int = 1000, height: int = 600) -> Dict[str, Any]:
        """
        Generate bar chart showing cluster size distribution.

        Args:
            algorithm: Clustering algorithm ('kmeans', 'dbscan', 'hdbscan')
            output_path: Path to save the bar chart
            width: Image width in pixels
            height: Image height in pixels

        Returns:
            {
                'algorithm': str,
                'num_clusters': int,
                'output_path': str,
                'cluster_sizes': dict
            }

        Examples:
            # Visualize cluster distribution
            result = kb.visualize_cluster_distribution('kmeans', 'distribution.png')
        """
        import matplotlib.pyplot as plt
        import numpy as np

        cursor = self.db_conn.cursor()

        # Get cluster sizes
        cluster_sizes = cursor.execute(
            """SELECT c.cluster_number, COUNT(*) as size
               FROM document_clusters dc
               JOIN clusters c ON dc.cluster_id = c.cluster_id
               WHERE c.algorithm = ?
               GROUP BY c.cluster_number
               ORDER BY c.cluster_number""",
            (algorithm,)
        ).fetchall()

        if not cluster_sizes:
            return {'error': f'No clusters found for algorithm {algorithm}'}

        cluster_numbers = [c[0] for c in cluster_sizes]
        sizes = [c[1] for c in cluster_sizes]

        # Create bar chart
        plt.figure(figsize=(width/100, height/100), dpi=100)

        colors = ['red' if cn == -1 else 'steelblue' for cn in cluster_numbers]
        labels = ['Outliers' if cn == -1 else f'Cluster {cn}' for cn in cluster_numbers]

        bars = plt.bar(range(len(cluster_numbers)), sizes, color=colors, alpha=0.7)

        plt.xticks(range(len(cluster_numbers)), labels, rotation=45 if len(cluster_numbers) > 10 else 0)
        plt.ylabel('Number of Documents', fontsize=12)
        plt.xlabel('Cluster', fontsize=12)
        plt.title(f'Cluster Size Distribution ({algorithm.upper()})', fontsize=14)

        # Add value labels on bars
        for i, (bar, size) in enumerate(zip(bars, sizes)):
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height,
                    f'{int(size)}',
                    ha='center', va='bottom', fontsize=9)

        plt.tight_layout()
        plt.savefig(output_path, dpi=100, bbox_inches='tight')
        plt.close()

        # Build cluster sizes dict
        sizes_dict = {f'cluster_{cn}' if cn != -1 else 'outliers': size
                      for cn, size in cluster_sizes}

        return {
            'algorithm': algorithm,
            'num_clusters': len(cluster_numbers),
            'output_path': output_path,
            'cluster_sizes': sizes_dict
        }

    def extract_relationships_bulk(self, min_confidence: float = 0.6,
                                   max_docs: Optional[int] = None,
                                   skip_existing: bool = False) -> dict:
        """
        Bulk extract entity relationships for multiple documents.

        Args:
            min_confidence: Minimum confidence threshold for entities
            max_docs: Maximum number of documents to process (None = all)
            skip_existing: If True, skip documents that already have relationships

        Returns:
            {
                'processed': int,
                'failed': int,
                'skipped': int,
                'total_relationships': int,
                'results': [list of individual results]
            }

        Examples:
            # Extract relationships for all documents
            results = kb.extract_relationships_bulk()

            # Extract for documents with entities only
            results = kb.extract_relationships_bulk(skip_existing=True)
        """
        results = {
            'processed': 0,
            'failed': 0,
            'skipped': 0,
            'total_relationships': 0,
            'results': []
        }

        # Get documents with entities
        docs_to_process = []
        cursor = self.db_conn.cursor()

        for doc_id in self.documents.keys():
            cursor.execute("SELECT COUNT(*) FROM document_entities WHERE doc_id = ?", (doc_id,))
            entity_count = cursor.fetchone()[0]

            if entity_count > 0:
                docs_to_process.append(doc_id)

        if max_docs:
            docs_to_process = docs_to_process[:max_docs]

        self.logger.info(f"Extracting entity relationships for {len(docs_to_process)} documents")

        # Process each document
        for i, doc_id in enumerate(docs_to_process, 1):
            doc_results = {
                'doc_id': doc_id,
                'title': self.documents[doc_id].title,
                'status': 'unknown',
                'relationship_count': 0,
                'error': None
            }

            try:
                self.logger.info(f"[{i}/{len(docs_to_process)}] Extracting relationships from {doc_id}")

                result = self.extract_entity_relationships(doc_id, min_confidence=min_confidence)

                doc_results['status'] = 'success'
                doc_results['relationship_count'] = result['relationship_count']

                results['processed'] += 1
                results['total_relationships'] += result['relationship_count']
                results['results'].append(doc_results)

            except Exception as e:
                self.logger.error(f"[{i}/{len(docs_to_process)}] Failed to extract relationships from {doc_id}: {e}")
                doc_results['status'] = 'failed'
                doc_results['error'] = str(e)
                results['failed'] += 1
                results['results'].append(doc_results)

        self.logger.info(f"Bulk relationship extraction complete: processed={results['processed']}, "
                        f"failed={results['failed']}, total_relationships={results['total_relationships']}")

        return results

    def export_documents_bulk(self, doc_ids: Optional[list[str]] = None,
                              tags: Optional[list[str]] = None,
                              format: str = 'json') -> str:
        """
        Export metadata for multiple documents.

        Args:
            doc_ids: List of document IDs to export (if None, uses tags or exports all)
            tags: Export documents with any of these tags
            format: Export format ('json', 'csv', or 'markdown')

        Returns:
            Exported data as a string

        Examples:
            # Export all documents as JSON
            data = kb.export_documents_bulk(format='json')

            # Export documents with 'reference' tag as CSV
            data = kb.export_documents_bulk(tags=['reference'], format='csv')

            # Export specific documents as Markdown
            data = kb.export_documents_bulk(doc_ids=['doc1', 'doc2'], format='markdown')
        """
        # Collect docs to export
        docs_to_export = []

        if doc_ids:
            # Export specific documents
            for doc_id in doc_ids:
                if doc_id in self.documents:
                    docs_to_export.append(self.documents[doc_id])
        elif tags:
            # Export documents with specified tags
            for doc in self.documents.values():
                if any(tag in doc.tags for tag in tags):
                    docs_to_export.append(doc)
        else:
            # Export all documents
            docs_to_export = list(self.documents.values())

        self.logger.info(f"Bulk export: exporting {len(docs_to_export)} documents as {format}")

        # Format the output
        if format == 'json':
            export_data = []
            for doc in docs_to_export:
                export_data.append({
                    'doc_id': doc.doc_id,
                    'filename': doc.filename,
                    'title': doc.title,
                    'filepath': doc.filepath,
                    'file_type': doc.file_type,
                    'total_pages': doc.total_pages,
                    'total_chunks': doc.total_chunks,
                    'indexed_at': doc.indexed_at,
                    'tags': doc.tags,
                    'author': doc.author,
                    'subject': doc.subject,
                    'creator': doc.creator,
                    'creation_date': doc.creation_date
                })
            return json.dumps(export_data, indent=2)

        elif format == 'csv':
            import csv
            from io import StringIO

            output = StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow(['doc_id', 'filename', 'title', 'filepath', 'file_type',
                           'total_pages', 'total_chunks', 'indexed_at', 'tags',
                           'author', 'subject', 'creator', 'creation_date'])

            # Write data
            for doc in docs_to_export:
                writer.writerow([
                    doc.doc_id,
                    doc.filename,
                    doc.title,
                    doc.filepath,
                    doc.file_type,
                    doc.total_pages,
                    doc.total_chunks,
                    doc.indexed_at,
                    ', '.join(doc.tags),
                    doc.author or '',
                    doc.subject or '',
                    doc.creator or '',
                    doc.creation_date or ''
                ])

            return output.getvalue()

        elif format == 'markdown':
            lines = []
            lines.append("# Document Export")
            lines.append(f"\n**Total Documents:** {len(docs_to_export)}")
            lines.append(f"**Exported:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            lines.append("---\n")

            for i, doc in enumerate(docs_to_export, 1):
                lines.append(f"## {i}. {doc.title}")
                lines.append(f"- **ID:** `{doc.doc_id}`")
                lines.append(f"- **Filename:** {doc.filename}")
                lines.append(f"- **Type:** {doc.file_type}")
                lines.append(f"- **Pages:** {doc.total_pages}")
                lines.append(f"- **Chunks:** {doc.total_chunks}")
                lines.append(f"- **Tags:** {', '.join(doc.tags) if doc.tags else 'None'}")
                if doc.author:
                    lines.append(f"- **Author:** {doc.author}")
                if doc.subject:
                    lines.append(f"- **Subject:** {doc.subject}")
                lines.append(f"- **Indexed:** {doc.indexed_at}")
                lines.append(f"- **Path:** `{doc.filepath}`")
                lines.append("")

            return '\n'.join(lines)

        else:
            raise ValueError(f"Unsupported format: {format}. Use 'json', 'csv', or 'markdown'")

    def add_relationship(self, from_doc_id: str, to_doc_id: str,
                        relationship_type: str = "related", note: str = "") -> dict:
        """
        Add a relationship between two documents.

        Args:
            from_doc_id: Source document ID
            to_doc_id: Target document ID
            relationship_type: Type of relationship (e.g., 'related', 'references', 'prerequisite', 'sequel')
            note: Optional note about the relationship

        Returns:
            Dictionary with relationship details

        Examples:
            # Mark document as related
            kb.add_relationship("doc1", "doc2", "related", "Both cover VIC-II graphics")

            # Mark as prerequisite
            kb.add_relationship("basic_guide", "advanced_guide", "prerequisite", "Read basic first")
        """
        # Validate documents exist
        if from_doc_id not in self.documents:
            raise ValueError(f"Source document not found: {from_doc_id}")
        if to_doc_id not in self.documents:
            raise ValueError(f"Target document not found: {to_doc_id}")

        # Create relationships table if it doesn't exist
        cursor = self.db_conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS document_relationships (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                from_doc_id TEXT NOT NULL,
                to_doc_id TEXT NOT NULL,
                relationship_type TEXT NOT NULL,
                note TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (from_doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE,
                FOREIGN KEY (to_doc_id) REFERENCES documents(doc_id) ON DELETE CASCADE,
                UNIQUE(from_doc_id, to_doc_id, relationship_type)
            )
        """)

        # Check if relationship already exists
        cursor.execute("""
            SELECT id FROM document_relationships
            WHERE from_doc_id = ? AND to_doc_id = ? AND relationship_type = ?
        """, (from_doc_id, to_doc_id, relationship_type))

        if cursor.fetchone():
            raise ValueError(f"Relationship already exists: {from_doc_id} -> {to_doc_id} ({relationship_type})")

        # Insert relationship
        created_at = datetime.now().isoformat()
        cursor.execute("""
            INSERT INTO document_relationships (from_doc_id, to_doc_id, relationship_type, note, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (from_doc_id, to_doc_id, relationship_type, note, created_at))

        self.db_conn.commit()

        self.logger.info(f"Added relationship: {from_doc_id} -> {to_doc_id} ({relationship_type})")

        return {
            'from_doc_id': from_doc_id,
            'to_doc_id': to_doc_id,
            'relationship_type': relationship_type,
            'note': note,
            'created_at': created_at
        }

    def remove_relationship(self, from_doc_id: str, to_doc_id: str,
                           relationship_type: Optional[str] = None) -> bool:
        """
        Remove a relationship between two documents.

        Args:
            from_doc_id: Source document ID
            to_doc_id: Target document ID
            relationship_type: Optional specific relationship type to remove

        Returns:
            True if relationship was removed, False if not found
        """
        cursor = self.db_conn.cursor()

        if relationship_type:
            cursor.execute("""
                DELETE FROM document_relationships
                WHERE from_doc_id = ? AND to_doc_id = ? AND relationship_type = ?
            """, (from_doc_id, to_doc_id, relationship_type))
        else:
            cursor.execute("""
                DELETE FROM document_relationships
                WHERE from_doc_id = ? AND to_doc_id = ?
            """, (from_doc_id, to_doc_id))

        self.db_conn.commit()
        removed = cursor.rowcount > 0

        if removed:
            self.logger.info(f"Removed relationship: {from_doc_id} -> {to_doc_id}")

        return removed

    def get_entity_analytics(self, time_range_days: int = 30) -> dict:
        """
        Get comprehensive entity analytics for dashboard visualization.

        Provides aggregate statistics and trends for entities and relationships
        across the knowledge base.

        Args:
            time_range_days: Number of days to include in timeline (default: 30)

        Returns:
            {
                'entity_distribution': {                    # Entities by type
                    'hardware': 45,
                    'memory_address': 120,
                    'instruction': 89,
                    ...
                },
                'top_entities': [                           # Most common entities
                    {
                        'entity_text': 'VIC-II',
                        'entity_type': 'hardware',
                        'doc_count': 15,
                        'total_occurrences': 127,
                        'avg_confidence': 0.92
                    },
                    ...
                ],
                'relationship_stats': {                     # Relationship statistics
                    'total': 234,
                    'avg_strength': 0.75,
                    'by_type': {'references': 120, 'related_to': 114}
                },
                'top_relationships': [                      # Strongest relationships
                    {
                        'entity1': 'VIC-II',
                        'entity2': 'sprite',
                        'strength': 0.95,
                        'doc_count': 12
                    },
                    ...
                ],
                'extraction_timeline': [                    # Entities extracted over time
                    {'date': '2024-12-01', 'count': 45},
                    {'date': '2024-12-02', 'count': 67},
                    ...
                ]
            }
        """
        cursor = self.db_conn.cursor()
        analytics = {}

        # 1. Entity distribution by type
        cursor.execute("""
            SELECT entity_type, COUNT(DISTINCT entity_text) as count
            FROM document_entities
            GROUP BY entity_type
            ORDER BY count DESC
        """)
        analytics['entity_distribution'] = {row[0]: row[1] for row in cursor.fetchall()}

        # 2. Top entities (most common across documents)
        cursor.execute("""
            SELECT
                entity_text,
                entity_type,
                COUNT(DISTINCT doc_id) as doc_count,
                SUM(occurrence_count) as total_occurrences,
                AVG(confidence) as avg_confidence
            FROM document_entities
            GROUP BY entity_text, entity_type
            HAVING doc_count >= 2
            ORDER BY doc_count DESC, total_occurrences DESC
            LIMIT 50
        """)

        analytics['top_entities'] = []
        for row in cursor.fetchall():
            analytics['top_entities'].append({
                'entity_text': row[0],
                'entity_type': row[1],
                'doc_count': row[2],
                'total_occurrences': row[3],
                'avg_confidence': round(row[4], 2)
            })

        # 3. Relationship statistics
        cursor.execute("""
            SELECT COUNT(*) FROM entity_relationships
        """)
        total_relationships = cursor.fetchone()[0]

        cursor.execute("""
            SELECT AVG(strength) FROM entity_relationships
        """)
        avg_strength = cursor.fetchone()[0] or 0.0

        cursor.execute("""
            SELECT relationship_type, COUNT(*) as count
            FROM entity_relationships
            GROUP BY relationship_type
            ORDER BY count DESC
        """)
        by_type = {row[0]: row[1] for row in cursor.fetchall()}

        analytics['relationship_stats'] = {
            'total': total_relationships,
            'avg_strength': round(avg_strength, 2),
            'by_type': by_type
        }

        # 4. Top relationships (strongest entity pairs)
        cursor.execute("""
            SELECT
                r.entity1_text,
                r.entity1_type,
                r.entity2_text,
                r.entity2_type,
                r.strength,
                r.doc_count
            FROM entity_relationships r
            WHERE r.doc_count >= 1
            ORDER BY r.strength DESC, r.doc_count DESC
            LIMIT 50
        """)

        analytics['top_relationships'] = []
        for row in cursor.fetchall():
            analytics['top_relationships'].append({
                'entity1': row[0],
                'entity1_type': row[1],
                'entity2': row[2],
                'entity2_type': row[3],
                'strength': round(row[4], 2) if row[4] else 0.0,
                'doc_count': row[5]
            })

        # 5. Extraction timeline (entities extracted per day)
        from datetime import datetime, timedelta
        start_date = datetime.now() - timedelta(days=time_range_days)
        start_date_str = start_date.strftime('%Y-%m-%d')

        cursor.execute("""
            SELECT
                DATE(generated_at) as extraction_date,
                COUNT(DISTINCT entity_text) as count
            FROM document_entities
            WHERE generated_at >= ?
            GROUP BY DATE(generated_at)
            ORDER BY extraction_date ASC
        """, (start_date_str,))

        analytics['extraction_timeline'] = []
        for row in cursor.fetchall():
            analytics['extraction_timeline'].append({
                'date': row[0],
                'count': row[1]
            })

        # 6. Overall statistics
        cursor.execute("""
            SELECT
                COUNT(DISTINCT entity_text) as unique_entities,
                COUNT(*) as total_entity_instances,
                COUNT(DISTINCT doc_id) as docs_with_entities
            FROM document_entities
        """)
        row = cursor.fetchone()

        analytics['overall'] = {
            'unique_entities': row[0],
            'total_instances': row[1],
            'docs_with_entities': row[2],
            'avg_entities_per_doc': round(row[1] / row[2], 1) if row[2] > 0 else 0
        }

        return analytics

    def get_relationships(self, doc_id: str, direction: str = "both") -> list[dict]:
        """
        Get all relationships for a document.

        Args:
            doc_id: Document ID
            direction: 'outgoing', 'incoming', or 'both' (default)

        Returns:
            List of relationship dictionaries
        """
        if doc_id not in self.documents:
            raise ValueError(f"Document not found: {doc_id}")

        cursor = self.db_conn.cursor()
        relationships = []

        # Get outgoing relationships (this doc -> others)
        if direction in ["outgoing", "both"]:
            cursor.execute("""
                SELECT from_doc_id, to_doc_id, relationship_type, note, created_at
                FROM document_relationships
                WHERE from_doc_id = ?
            """, (doc_id,))

            for row in cursor.fetchall():
                relationships.append({
                    'direction': 'outgoing',
                    'from_doc_id': row[0],
                    'to_doc_id': row[1],
                    'relationship_type': row[2],
                    'note': row[3],
                    'created_at': row[4],
                    'related_doc_id': row[1]  # For convenience
                })

        # Get incoming relationships (others -> this doc)
        if direction in ["incoming", "both"]:
            cursor.execute("""
                SELECT from_doc_id, to_doc_id, relationship_type, note, created_at
                FROM document_relationships
                WHERE to_doc_id = ?
            """, (doc_id,))

            for row in cursor.fetchall():
                relationships.append({
                    'direction': 'incoming',
                    'from_doc_id': row[0],
                    'to_doc_id': row[1],
                    'relationship_type': row[2],
                    'note': row[3],
                    'created_at': row[4],
                    'related_doc_id': row[0]  # For convenience
                })

        return relationships

    def get_related_documents(self, doc_id: str, relationship_type: Optional[str] = None) -> list[dict]:
        """
        Get all documents related to a given document with full metadata.

        Args:
            doc_id: Document ID
            relationship_type: Optional filter by relationship type

        Returns:
            List of related documents with relationship info
        """
        relationships = self.get_relationships(doc_id)

        if relationship_type:
            relationships = [r for r in relationships if r['relationship_type'] == relationship_type]

        related_docs = []
        for rel in relationships:
            related_doc_id = rel['related_doc_id']
            if related_doc_id in self.documents:
                doc_meta = self.documents[related_doc_id]
                related_docs.append({
                    'doc_id': doc_meta.doc_id,
                    'title': doc_meta.title,
                    'filename': doc_meta.filename,
                    'tags': doc_meta.tags,
                    'relationship_type': rel['relationship_type'],
                    'relationship_direction': rel['direction'],
                    'note': rel['note'],
                    'created_at': rel['created_at']
                })

        return related_docs

    def get_relationship_graph(self, tags: Optional[list[str]] = None,
                              relationship_types: Optional[list[str]] = None) -> dict:
        """
        Get relationship graph data for visualization.

        Args:
            tags: Optional list of tags to filter documents
            relationship_types: Optional list of relationship types to include

        Returns:
            Dictionary with 'nodes' and 'edges' for graph visualization
        """
        cursor = self.db_conn.cursor()

        # Build WHERE clauses for filtering
        where_clauses = []
        params = []

        if relationship_types:
            placeholders = ','.join('?' * len(relationship_types))
            where_clauses.append(f"relationship_type IN ({placeholders})")
            params.extend(relationship_types)

        where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

        # Get all relationships
        query = f"""
            SELECT from_doc_id, to_doc_id, relationship_type, note
            FROM document_relationships
            {where_sql}
        """

        cursor.execute(query, params)
        relationships = cursor.fetchall()

        # Build nodes and edges
        nodes = {}
        edges = []

        for from_id, to_id, rel_type, note in relationships:
            # Check if documents match tag filter
            if tags:
                from_doc = self.documents.get(from_id)
                to_doc = self.documents.get(to_id)
                if not from_doc or not to_doc:
                    continue
                if not any(tag in from_doc.tags for tag in tags) and \
                   not any(tag in to_doc.tags for tag in tags):
                    continue

            # Add nodes if not already present
            if from_id not in nodes:
                doc = self.documents.get(from_id)
                if doc:
                    nodes[from_id] = {
                        'id': from_id,
                        'label': doc.title,
                        'title': f"{doc.title}\n{doc.filename}\nTags: {', '.join(doc.tags)}",
                        'tags': doc.tags,
                        'chunks': doc.total_chunks
                    }

            if to_id not in nodes:
                doc = self.documents.get(to_id)
                if doc:
                    nodes[to_id] = {
                        'id': to_id,
                        'label': doc.title,
                        'title': f"{doc.title}\n{doc.filename}\nTags: {', '.join(doc.tags)}",
                        'tags': doc.tags,
                        'chunks': doc.total_chunks
                    }

            # Add edge
            if from_id in nodes and to_id in nodes:
                edges.append({
                    'from': from_id,
                    'to': to_id,
                    'type': rel_type,
                    'label': rel_type,
                    'title': note if note else rel_type
                })

        return {
            'nodes': list(nodes.values()),
            'edges': edges,
            'stats': {
                'total_nodes': len(nodes),
                'total_edges': len(edges),
                'relationship_types': list(set(e['type'] for e in edges))
            }
        }

    # NOTE: translate_nl_query is defined later in the file (~line 9100)

    def search(self, query: str, max_results: int = 5, tags: Optional[list[str]] = None) -> list[dict]:
        """Search the knowledge base using BM25 ranking or simple term frequency."""
        start_time = time.time()

        # Check cache first
        if self._search_cache is not None:
            cache_key = self._cache_key('search',
                                       query=query,
                                       max_results=max_results,
                                       tags=tuple(sorted(tags)) if tags else None)
            if cache_key in self._search_cache:
                results = self._search_cache[cache_key]
                elapsed_ms = (time.time() - start_time) * 1000
                self.logger.debug(f"Cache hit for query: '{query}' ({len(results)} results, {elapsed_ms:.2f}ms)")
                return results

        self.logger.info(f"Search query: '{query}' (max_results={max_results}, tags={tags})")

        # Extract phrase queries (text in quotes)
        phrase_pattern = r'"([^"]*)"'
        phrases = re.findall(phrase_pattern, query)
        # Remove phrases from query to get regular terms
        query_without_phrases = re.sub(phrase_pattern, '', query)

        # Preprocess query terms (tokenize, remove stopwords, stem)
        query_terms_list = self._preprocess_text(query_without_phrases)
        query_terms = set(query_terms_list)
        query_terms = {term for term in query_terms if term}  # Remove empty strings

        # Check search backend preference (in priority order)
        # FTS5 can be enabled with environment variable: USE_FTS5=1
        # BM25 can be disabled with environment variable: USE_BM25=0
        use_fts5 = os.environ.get('USE_FTS5', '0') == '1'
        use_bm25 = os.environ.get('USE_BM25', '1') == '1'

        if use_fts5 and self._fts5_available():
            # Use SQLite FTS5 full-text search
            results = self._search_fts5(query, query_terms, phrases, tags, max_results)
            # Fall back to BM25/simple if FTS5 returns no results
            if not results:
                if use_bm25 and BM25_SUPPORT:
                    if self.bm25 is None:
                        self._build_bm25_index()
                    if self.bm25 is not None:
                        results = self._search_bm25(query, query_terms, phrases, tags, max_results)
                    else:
                        results = self._search_simple(query_terms, phrases, tags, max_results)
                else:
                    results = self._search_simple(query_terms, phrases, tags, max_results)
        elif use_bm25 and BM25_SUPPORT:
            # Build BM25 index if not already built
            if self.bm25 is None:
                self._build_bm25_index()

            if self.bm25 is not None:
                results = self._search_bm25(query, query_terms, phrases, tags, max_results)
            else:
                results = self._search_simple(query_terms, phrases, tags, max_results)
        else:
            results = self._search_simple(query_terms, phrases, tags, max_results)

        # Store in cache
        if self._search_cache is not None:
            cache_key = self._cache_key('search',
                                       query=query,
                                       max_results=max_results,
                                       tags=tuple(sorted(tags)) if tags else None)
            self._search_cache[cache_key] = results

        elapsed_ms = (time.time() - start_time) * 1000
        self.logger.info(f"Search completed: {len(results)} results in {elapsed_ms:.2f}ms")

        # Log search for analytics
        search_mode = 'fts5' if (use_fts5 and self._fts5_available() and results) else ('bm25' if use_bm25 else 'simple')
        self._log_search(query, search_mode, len(results), elapsed_ms, tags)

        return results

    def _search_bm25(self, query: str, query_terms: set, phrases: list, tags: Optional[list[str]], max_results: int) -> list[dict]:
        """Search using BM25 algorithm."""
        # Preprocess query for BM25
        tokenized_query = self._preprocess_text(query)

        # Get BM25 scores for all chunks
        bm25_scores = self.bm25.get_scores(tokenized_query)

        # Build results with scores
        results = []
        for idx, chunk in enumerate(self.chunks):
            # Filter by tags if specified
            if tags:
                doc = self.documents.get(chunk.doc_id)
                if doc and not any(t in doc.tags for t in tags):
                    continue

            score = bm25_scores[idx]

            # Boost score for phrase matches
            if phrases:
                content_lower = chunk.content.lower()
                for phrase in phrases:
                    if phrase.lower() in content_lower:
                        score *= 2  # 2x boost for phrase match

            # Combine query_terms and phrases for snippet extraction
            all_terms = query_terms | {p.lower() for p in phrases}
            snippet = self._extract_snippet(chunk.content, all_terms)
            results.append({
                'doc_id': chunk.doc_id,
                'filename': chunk.filename,
                'title': chunk.title,
                'chunk_id': chunk.chunk_id,
                'score': float(score),
                'snippet': snippet,
                'word_count': chunk.word_count
            })

        # Sort by score and return top results
        # BM25 scores can be negative for very small documents
        # Accept any non-zero score (positive or moderately negative)
        # Filter only exact zeros (true non-matches)
        results = [r for r in results if abs(r['score']) > 0.0001]
        results.sort(key=lambda x: x['score'], reverse=True)
        return results[:max_results]

    def _search_simple(self, query_terms: set, phrases: list, tags: Optional[list[str]], max_results: int) -> list[dict]:
        """Simple term frequency search with fuzzy matching support (fallback when BM25 not available)."""
        results = []

        for chunk in self.chunks:
            # Filter by tags if specified
            if tags:
                doc = self.documents.get(chunk.doc_id)
                if doc and not any(t in doc.tags for t in tags):
                    continue

            content_lower = chunk.content.lower()

            # Score based on term frequency (exact matches)
            score = 0
            for term in query_terms:
                # Exact word match (higher score)
                score += len(re.findall(r'\b' + re.escape(term) + r'\b', content_lower)) * 2
                # Partial match
                score += content_lower.count(term)

            # If fuzzy search is enabled and no exact matches, try fuzzy matching
            if self.use_fuzzy and score == 0:
                match_found, fuzzy_score = self._fuzzy_match_terms(list(query_terms), chunk.content)
                if match_found:
                    # Use fuzzy score (scaled down since it's less reliable than exact match)
                    score = fuzzy_score / 10.0

            # Boost score for phrase matches
            for phrase in phrases:
                if phrase.lower() in content_lower:
                    score += len(phrase.split()) * 10  # High boost for phrase match

            if score > 0:
                # Combine query_terms and phrases for snippet extraction
                all_terms = query_terms | {p.lower() for p in phrases}
                snippet = self._extract_snippet(chunk.content, all_terms)
                results.append({
                    'doc_id': chunk.doc_id,
                    'filename': chunk.filename,
                    'title': chunk.title,
                    'chunk_id': chunk.chunk_id,
                    'score': score,
                    'snippet': snippet,
                    'word_count': chunk.word_count
                })

        # Sort by score and return top results
        results.sort(key=lambda x: x['score'], reverse=True)
        return results[:max_results]

    def _search_fts5(self, query: str, query_terms: set, phrases: list,
                     tags: Optional[list[str]], max_results: int) -> list[dict]:
        """Search using SQLite FTS5 full-text search."""
        cursor = self.db_conn.cursor()

        # Build FTS5 query (FTS5 supports boolean operators and phrases natively)
        # Escape special characters that could cause syntax errors
        # Quote terms with hyphens or other special chars to treat them as phrases
        import re
        words = query.split()
        escaped_words = []
        for word in words:
            # If word contains hyphen or other special chars, quote it
            if re.search(r'[-:]', word):
                escaped_words.append(f'"{word}"')
            else:
                escaped_words.append(word)
        fts_query = ' '.join(escaped_words)

        try:
            # Execute FTS5 search with BM25 ranking
            cursor.execute("""
                SELECT
                    c.doc_id,
                    c.chunk_id,
                    c.content,
                    c.word_count,
                    c.page,
                    d.filename,
                    d.title,
                    fts.rank as score
                FROM chunks_fts5 fts
                JOIN chunks c ON c.rowid = fts.rowid
                JOIN documents d ON d.doc_id = c.doc_id
                WHERE chunks_fts5 MATCH ?
                ORDER BY rank
                LIMIT ?
            """, (fts_query, max_results * 2))  # Get 2x for tag filtering

            results = []
            for row in cursor.fetchall():
                doc_id, chunk_id, content, word_count, page, filename, title, score = row

                # Filter by tags if specified
                if tags:
                    doc = self.documents.get(doc_id)
                    if doc and not any(t in doc.tags for t in tags):
                        continue

                # Extract snippet with highlighting
                snippet = self._extract_snippet(content, query_terms | {p.lower() for p in phrases})

                results.append({
                    'doc_id': doc_id,
                    'filename': filename,
                    'title': title,
                    'chunk_id': chunk_id,
                    'score': abs(score),  # FTS5 returns negative scores (lower is better)
                    'snippet': snippet,
                    'word_count': word_count
                })

                if len(results) >= max_results:
                    break

            return results

        except Exception as e:
            self.logger.error(f"FTS5 search failed: {e}")
            # Fall back to simple search
            return []

    def _extract_snippet(self, content: str, query_terms: set, snippet_size: int = 300) -> str:
        """
        Extract a relevant snippet from content with highlighted search terms.
        Enhanced to extract complete sentences and find regions with high term density.
        """
        content_lower = content.lower()

        # Calculate term density across content in sliding windows
        window_size = snippet_size
        best_score = 0
        best_pos = 0

        # Slide through content and score each window by term matches
        for i in range(0, max(1, len(content) - window_size + 1), 50):  # Step by 50 chars
            window = content_lower[i:i + window_size]
            score = sum(window.count(term) for term in query_terms if term)
            if score > best_score:
                best_score = score
                best_pos = i

        # If no matches found, use beginning
        if best_score == 0:
            best_pos = 0

        # Expand to complete sentences
        # Find sentence boundaries (., !, ?, or newline followed by capital letter or code)
        sentence_pattern = r'[.!?\n][\s\n]+'

        # Find start of sentence
        start = best_pos
        # Look backwards for sentence start
        for match in re.finditer(sentence_pattern, content[:best_pos]):
            start = match.end()

        # If we're too far from best_pos, adjust
        if best_pos - start > snippet_size // 2:
            start = max(0, best_pos - snippet_size // 3)

        # Find end of sentence
        end = min(len(content), start + snippet_size)
        matches = list(re.finditer(sentence_pattern, content[start:]))
        for match in matches:
            potential_end = start + match.end()
            if potential_end - start >= snippet_size * 0.8:  # At least 80% of desired size
                end = potential_end
                break

        # If we couldn't find sentence end, use hard cutoff
        if end - start < snippet_size * 0.5:
            end = min(len(content), start + snippet_size)

        snippet = content[start:end].strip()

        # Preserve code blocks (lines starting with spaces/tabs)
        # Don't break in the middle of code
        lines = snippet.split('\n')
        if lines and (lines[0].startswith('    ') or lines[0].startswith('\t')):
            # Start of snippet is code, find complete code block
            code_end = 0
            for i, line in enumerate(lines):
                if line and not line[0].isspace():
                    code_end = i
                    break
            if code_end > 0:
                snippet = '\n'.join(lines[:code_end])

        # Highlight matching terms (case-insensitive, whole words)
        for term in query_terms:
            if len(term) >= 2:  # Only highlight terms with 2+ characters
                # Use word boundary for whole word matching when possible
                pattern = re.compile(f'\\b({re.escape(term)})\\b', re.IGNORECASE)
                snippet = pattern.sub(r'**\1**', snippet)

        # Add ellipsis if truncated
        if start > 0:
            snippet = "..." + snippet
        if end < len(content):
            snippet = snippet + "..."

        return snippet

    def _ensure_embeddings_loaded(self):
        """
        Lazy load embeddings model and index on first use.

        This significantly improves startup time by deferring the ~2.5 second
        model loading until semantic search is actually needed.
        """
        if not self.use_semantic or self._embeddings_loaded:
            return

        try:
            # Load the sentence transformer model (this is the slow part)
            model_name = os.getenv('SEMANTIC_MODEL', 'all-MiniLM-L6-v2')
            self.logger.info(f"Lazy loading embeddings model: {model_name} (first semantic search)")
            self.embeddings_model = SentenceTransformer(model_name)

            # Load the pre-computed embeddings index
            self._load_embeddings()

            self._embeddings_loaded = True
            self.logger.info("Embeddings model and index loaded successfully")
        except Exception as e:
            self.logger.error(f"Failed to lazy load embeddings: {e}")
            self.use_semantic = False
            self._embeddings_loaded = False

    def _load_embeddings(self):
        """Load FAISS embeddings index from disk."""
        if not self.use_semantic:
            return

        try:
            if self.embeddings_file.exists() and self.embeddings_map_file.exists():
                self.embeddings_index = faiss.read_index(str(self.embeddings_file))
                with open(self.embeddings_map_file, 'r') as f:
                    self.embeddings_doc_map = json.load(f)
                self.logger.info(f"Loaded embeddings index with {len(self.embeddings_doc_map)} vectors")
            else:
                self.logger.info("No existing embeddings found, will build on first use")
        except Exception as e:
            self.logger.error(f"Error loading embeddings: {e}")
            self.embeddings_index = None
            self.embeddings_doc_map = []

    def _save_embeddings(self):
        """Save FAISS embeddings index to disk."""
        if not self.use_semantic or self.embeddings_index is None:
            return

        try:
            faiss.write_index(self.embeddings_index, str(self.embeddings_file))
            with open(self.embeddings_map_file, 'w') as f:
                json.dump(self.embeddings_doc_map, f)
            self.logger.info(f"Saved embeddings index with {len(self.embeddings_doc_map)} vectors")
        except Exception as e:
            self.logger.error(f"Error saving embeddings: {e}")

    def _build_embeddings(self):
        """Build FAISS index from all document chunks."""
        if not self.use_semantic:
            return

        # Ensure model is loaded
        self._ensure_embeddings_loaded()

        if self.embeddings_model is None:
            return

        self.logger.info("Building embeddings index for all chunks...")
        start_time = time.time()

        # Load all chunks
        chunks = self._get_chunks_db()
        if not chunks:
            self.logger.warning("No chunks to embed")
            return

        # Generate embeddings
        texts = [chunk.content for chunk in chunks]
        embeddings = self.embeddings_model.encode(texts, show_progress_bar=True, convert_to_numpy=True)

        # Create FAISS index
        dimension = embeddings.shape[1]
        self.embeddings_index = faiss.IndexFlatIP(dimension)  # Inner product (cosine similarity)

        # Normalize embeddings for cosine similarity
        faiss.normalize_L2(embeddings)
        self.embeddings_index.add(embeddings)

        # Build doc map
        self.embeddings_doc_map = [(chunk.doc_id, chunk.chunk_id) for chunk in chunks]

        # Save to disk
        self._save_embeddings()

        # Invalidate similarity cache since embeddings changed
        if self._similar_cache is not None:
            self._similar_cache.clear()
            self.logger.info("Similarity cache invalidated after rebuilding embeddings")

        elapsed = time.time() - start_time
        self.logger.info(f"Built embeddings index in {elapsed:.2f}s")

    def _add_chunks_to_embeddings(self, chunks: list[DocumentChunk]):
        """
        Incrementally add new chunks to the FAISS embeddings index with batched processing.

        For large documents, processes chunks in batches to reduce memory usage and
        improve CPU cache utilization. Configurable via EMBEDDING_BATCH_SIZE env var.

        Args:
            chunks: List of DocumentChunk objects to add to embeddings
        """
        if not self.use_semantic or self.embeddings_model is None:
            return

        if not chunks:
            self.logger.debug("No chunks to add to embeddings")
            return

        self.logger.info(f"Adding {len(chunks)} chunks to embeddings index...")
        start_time = time.time()

        # Get batch size from environment (default: 32 for optimal memory/performance trade-off)
        batch_size = int(os.getenv('EMBEDDING_BATCH_SIZE', '32'))

        # If no existing index, create one (need to know dimension first)
        if self.embeddings_index is None or len(self.embeddings_doc_map) == 0:
            self.logger.info("Creating new embeddings index")
            # Generate first embedding to get dimension
            sample_text = chunks[0].content
            sample_embedding = self.embeddings_model.encode([sample_text], show_progress_bar=False, convert_to_numpy=True)
            dimension = sample_embedding.shape[1]
            self.embeddings_index = faiss.IndexFlatIP(dimension)  # Inner product (cosine similarity)
            self.embeddings_doc_map = []

        # Process chunks in batches for better memory usage
        total_chunks = len(chunks)
        for batch_start in range(0, total_chunks, batch_size):
            batch_end = min(batch_start + batch_size, total_chunks)
            batch_chunks = chunks[batch_start:batch_end]

            # Generate embeddings for this batch
            texts = [chunk.content for chunk in batch_chunks]
            batch_embeddings = self.embeddings_model.encode(texts, show_progress_bar=False, convert_to_numpy=True)

            # Normalize for cosine similarity
            faiss.normalize_L2(batch_embeddings)

            # Add to existing index
            self.embeddings_index.add(batch_embeddings)

            # Update doc map
            for chunk in batch_chunks:
                self.embeddings_doc_map.append((chunk.doc_id, chunk.chunk_id))

            if total_chunks > batch_size:
                self.logger.debug(f"Processed embedding batch {batch_start+1}-{batch_end}/{total_chunks}")

        # Save updated index to disk
        self._save_embeddings()

        # Invalidate similarity cache since embeddings changed
        if self._similar_cache is not None:
            self._similar_cache.clear()
            self.logger.debug("Similarity cache invalidated after adding chunks")

        elapsed = time.time() - start_time
        self.logger.info(f"Added {len(chunks)} chunks to embeddings index in {elapsed:.2f}s")

    def semantic_search(self, query: str, max_results: int = 5, tags: Optional[list[str]] = None) -> list[dict]:
        """
        Perform semantic search using embeddings and vector similarity.

        Args:
            query: Search query
            max_results: Maximum number of results to return
            tags: Optional list of tags to filter by

        Returns:
            List of search results with scores
        """
        if not self.use_semantic:
            raise RuntimeError("Semantic search not available. Enable with USE_SEMANTIC_SEARCH=1")

        # Lazy load embeddings model on first use (saves ~2.5s on startup)
        self._ensure_embeddings_loaded()

        if self.embeddings_model is None:
            raise RuntimeError("Failed to load embeddings model")

        # Build embeddings index if not yet built
        if self.embeddings_index is None or len(self.embeddings_doc_map) == 0:
            self._build_embeddings()
            if self.embeddings_index is None:
                return []

        # Check cache first
        start_time = time.time()
        if self._semantic_cache is not None:
            cache_key = self._cache_key('semantic_search',
                                       query=query,
                                       max_results=max_results,
                                       tags=tuple(sorted(tags)) if tags else None)
            if cache_key in self._semantic_cache:
                results = self._semantic_cache[cache_key]
                elapsed_ms = (time.time() - start_time) * 1000
                self.logger.debug(f"Semantic cache HIT for query: '{query}' ({len(results)} results, {elapsed_ms:.2f}ms)")
                return results

        self.logger.info(f"Semantic search query: '{query}' (max_results={max_results}, tags={tags})")

        # Check embedding cache first (expensive operation to avoid)
        cache_key = hashlib.md5(query.encode('utf-8')).hexdigest()
        query_embedding = None

        if self._embedding_cache is not None:
            query_embedding = self._embedding_cache.get(cache_key)
            if query_embedding is not None:
                self.logger.debug(f"Embedding cache HIT for query: '{query[:50]}'")

        # Encode query if not cached
        if query_embedding is None:
            self.logger.debug(f"Embedding cache MISS for query: '{query[:50]}'")
            query_embedding = self.embeddings_model.encode([query], convert_to_numpy=True)
            faiss.normalize_L2(query_embedding)

            # Cache the normalized embedding
            if self._embedding_cache is not None:
                self._embedding_cache[cache_key] = query_embedding
                self.logger.debug(f"Cached embedding for query: '{query[:50]}'")
        else:
            # Embedding was cached and already normalized
            pass

        # Search in FAISS index (get more results for filtering)
        k = min(max_results * 5, len(self.embeddings_doc_map))
        scores, indices = self.embeddings_index.search(query_embedding, k)

        # Build results
        results = []

        for score, idx in zip(scores[0], indices[0]):
            if idx < 0 or idx >= len(self.embeddings_doc_map):
                continue

            doc_id, chunk_id = self.embeddings_doc_map[idx]

            # Filter by tags if specified
            if tags:
                doc = self.documents.get(doc_id)
                if doc and not any(t in doc.tags for t in tags):
                    continue

            # Get chunk
            chunk = self.get_chunk(doc_id, chunk_id)
            if not chunk:
                continue

            # Extract snippet (highlight query terms)
            query_terms = set(query.lower().split())
            snippet = self._extract_snippet(chunk.content, query_terms)

            doc = self.documents.get(doc_id)
            results.append({
                'doc_id': doc_id,
                'filename': chunk.filename,
                'title': chunk.title,
                'chunk_id': chunk_id,
                'score': float(score),
                'snippet': snippet,
                'word_count': chunk.word_count,
                'similarity': float(score)  # Cosine similarity score
            })

            if len(results) >= max_results:
                break

        elapsed = (time.time() - start_time) * 1000
        self.logger.info(f"Semantic search completed: {len(results)} results in {elapsed:.2f}ms")

        # Log search for analytics
        self._log_search(query, 'semantic', len(results), elapsed, tags)

        # Store in cache
        if self._semantic_cache is not None:
            cache_key = self._cache_key('semantic_search',
                                       query=query,
                                       max_results=max_results,
                                       tags=tuple(sorted(tags)) if tags else None)
            self._semantic_cache[cache_key] = results

        return results

    def hybrid_search(self, query: str, max_results: int = 5, tags: Optional[list[str]] = None,
                     semantic_weight: float = 0.3) -> list[dict]:
        """
        Perform hybrid search combining FTS5 keyword search and semantic search.

        Args:
            query: Search query
            max_results: Maximum number of results to return
            tags: Optional list of tags to filter by
            semantic_weight: Weight for semantic score (0.0-1.0). Default 0.3 means 70% FTS5, 30% semantic.

        Returns:
            List of search results with combined scores
        """
        if not self.use_semantic or self.embeddings_model is None:
            # Fall back to regular search if semantic not available
            self.logger.warning("Semantic search not available, falling back to FTS5/BM25")
            return self.search(query, max_results, tags)

        # Check cache first
        start_time = time.time()
        if self._hybrid_cache is not None:
            cache_key = self._cache_key('hybrid_search',
                                       query=query,
                                       max_results=max_results,
                                       tags=tuple(sorted(tags)) if tags else None,
                                       semantic_weight=semantic_weight)
            if cache_key in self._hybrid_cache:
                results = self._hybrid_cache[cache_key]
                elapsed_ms = (time.time() - start_time) * 1000
                self.logger.debug(f"Hybrid cache HIT for query: '{query}' ({len(results)} results, {elapsed_ms:.2f}ms)")
                return results

        self.logger.info(f"Hybrid search query: '{query}' (max_results={max_results}, tags={tags}, semantic_weight={semantic_weight})")

        # Run both search methods in parallel for better performance
        # (FTS5 and semantic are independent operations)
        with ThreadPoolExecutor(max_workers=2) as executor:
            # Submit both searches concurrently
            fts_future = executor.submit(self.search, query, max_results * 2, tags)
            semantic_future = executor.submit(self.semantic_search, query, max_results * 2, tags)

            # Wait for both to complete
            fts_results = fts_future.result()
            semantic_results = semantic_future.result()

        self.logger.debug(f"Parallel search completed: FTS={len(fts_results)}, Semantic={len(semantic_results)}")

        # Normalize scores to 0-1 range
        # FTS5 scores: higher absolute value is better, normalize by max
        if fts_results:
            max_fts_score = max(r['score'] for r in fts_results)
            if max_fts_score > 0:
                for r in fts_results:
                    r['fts_score_normalized'] = r['score'] / max_fts_score
            else:
                for r in fts_results:
                    r['fts_score_normalized'] = 0.0

        # Semantic scores: already 0-1 cosine similarity
        for r in semantic_results:
            r['semantic_score_normalized'] = r.get('similarity', r['score'])

        # Merge results by (doc_id, chunk_id)
        merged = {}

        # Add FTS5 results
        for r in fts_results:
            key = (r['doc_id'], r['chunk_id'])
            merged[key] = {
                'doc_id': r['doc_id'],
                'filename': r['filename'],
                'title': r['title'],
                'chunk_id': r['chunk_id'],
                'snippet': r['snippet'],
                'word_count': r['word_count'],
                'fts_score': r['fts_score_normalized'],
                'semantic_score': 0.0  # Will be updated if found in semantic results
            }

        # Update with semantic results
        for r in semantic_results:
            key = (r['doc_id'], r['chunk_id'])
            if key in merged:
                merged[key]['semantic_score'] = r['semantic_score_normalized']
            else:
                # Not in FTS5 results, add it
                merged[key] = {
                    'doc_id': r['doc_id'],
                    'filename': r['filename'],
                    'title': r['title'],
                    'chunk_id': r['chunk_id'],
                    'snippet': r['snippet'],
                    'word_count': r['word_count'],
                    'fts_score': 0.0,
                    'semantic_score': r['semantic_score_normalized']
                }

        # Calculate hybrid scores
        results = []
        for item in merged.values():
            hybrid_score = (1.0 - semantic_weight) * item['fts_score'] + semantic_weight * item['semantic_score']
            results.append({
                'doc_id': item['doc_id'],
                'filename': item['filename'],
                'title': item['title'],
                'chunk_id': item['chunk_id'],
                'score': hybrid_score,
                'fts_score': item['fts_score'],
                'semantic_score': item['semantic_score'],
                'snippet': item['snippet'],
                'word_count': item['word_count']
            })

        # Sort by hybrid score (descending) and limit
        results.sort(key=lambda x: x['score'], reverse=True)
        results = results[:max_results]

        elapsed = (time.time() - start_time) * 1000
        self.logger.info(f"Hybrid search completed: {len(results)} results in {elapsed:.2f}ms")

        # Log search for analytics
        self._log_search(query, 'hybrid', len(results), elapsed, tags)

        # Store in cache
        if self._hybrid_cache is not None:
            cache_key = self._cache_key('hybrid_search',
                                       query=query,
                                       max_results=max_results,
                                       tags=tuple(sorted(tags)) if tags else None,
                                       semantic_weight=semantic_weight)
            self._hybrid_cache[cache_key] = results

        return results

    def fuzzy_search(self, query: str, max_results: int = 5, tags: Optional[list[str]] = None,
                     similarity_threshold: int = 80) -> list[dict]:
        """
        Search with typo tolerance using fuzzy string matching.

        Handles misspellings and variations:
        - "VIC2" → finds "VIC-II"
        - "asembly" → finds "assembly"
        - "6052" → finds "6502"

        Args:
            query: Search query (may contain typos)
            max_results: Maximum number of results to return
            tags: Optional list of tags to filter by
            similarity_threshold: Minimum similarity score (0-100, default: 80)

        Returns:
            List of search results, potentially with corrected query terms
        """
        start_time = time.time()

        # Try exact search first (fast path)
        exact_results = self.search(query, max_results, tags)
        if len(exact_results) >= max_results:
            elapsed_ms = (time.time() - start_time) * 1000
            self.logger.info(f"Fuzzy search (exact match): '{query}' ({len(exact_results)} results, {elapsed_ms:.2f}ms)")
            return exact_results

        # Build vocabulary from indexed terms if not already built
        if not hasattr(self, '_search_vocabulary') or self._search_vocabulary is None:
            self._build_search_vocabulary()

        # If vocabulary is still empty, fall back to exact search
        if not self._search_vocabulary:
            return exact_results

        # Attempt fuzzy matching on query terms
        if not FUZZY_SUPPORT:
            self.logger.warning("rapidfuzz not available, falling back to exact search")
            return exact_results

        from rapidfuzz import process

        # Split query into terms and correct each one
        query_terms = query.split()
        corrected_terms = []
        corrections_made = []

        for term in query_terms:
            # Find best match in vocabulary using fuzzy string matching
            best_match = process.extractOne(
                term,
                self._search_vocabulary,
                score_cutoff=similarity_threshold
            )

            if best_match:
                corrected_term = best_match[0]
                score = best_match[1]
                corrected_terms.append(corrected_term)

                if corrected_term.lower() != term.lower():
                    corrections_made.append({
                        'original': term,
                        'corrected': corrected_term,
                        'similarity': score
                    })
            else:
                # No match found, keep original term
                corrected_terms.append(term)

        # Build corrected query
        corrected_query = ' '.join(corrected_terms)

        # Log corrections if any were made
        if corrections_made:
            self.logger.info(f"Fuzzy search corrections: {corrections_made}")

        # Search with corrected query
        results = self.search(corrected_query, max_results, tags)

        # Add correction metadata to results
        if corrections_made and results:
            for result in results:
                result['fuzzy_corrections'] = corrections_made
                result['corrected_query'] = corrected_query

        elapsed_ms = (time.time() - start_time) * 1000
        self.logger.info(f"Fuzzy search: '{query}' → '{corrected_query}' ({len(results)} results, {elapsed_ms:.2f}ms)")

        return results

    def _build_search_vocabulary(self):
        """Extract all unique terms from indexed content for fuzzy matching."""
        vocabulary = set()

        # Add known C64 technical terms
        known_terms = [
            'VIC-II', 'VIC2', 'VIC', 'SID', 'CIA', '6502', '6581', '6567', '6569', '6526',
            'sprite', 'raster', 'screen', 'memory', 'register', 'address',
            'assembly', 'BASIC', 'machine', 'code', 'opcode', 'instruction',
            'bit', 'byte', 'word', 'pointer', 'variable', 'string', 'loop',
            'interrupt', 'timer', 'trigger', 'control', 'port', 'peripheral',
            'sound', 'music', 'voice', 'envelope', 'frequency', 'amplitude',
            'color', 'palette', 'graphics', 'pixel', 'bitmap', 'character',
            'kernel', 'ROM', 'RAM', 'disk', 'tape', 'storage',
            'program', 'subroutine', 'jump', 'branch', 'call', 'return',
            'accumulator', 'index', 'stack', 'status', 'flag', 'carry'
        ]
        vocabulary.update(known_terms)

        # Extract terms from chunks (limited to avoid overhead)
        try:
            chunks_sample = self.chunks[:min(1000, len(self.chunks))]  # Sample first 1000 chunks

            for chunk in chunks_sample:
                # Extract words (lowercase, alphanumeric + hyphen)
                words = re.findall(r'\b[a-z0-9-]{3,}\b', chunk.content.lower())
                vocabulary.update(words)
        except Exception as e:
            self.logger.warning(f"Failed to build search vocabulary: {e}")

        self._search_vocabulary = sorted(list(vocabulary))
        self.logger.debug(f"Built search vocabulary with {len(self._search_vocabulary)} terms")

    def search_within_results(self, previous_results: list[dict], refinement_query: str,
                             max_results: int = 5) -> list[dict]:
        """
        Search within a previous result set to refine results.

        Useful for progressive search refinement:
        1. results = search("VIC-II")  # 50 results
        2. refined = search_within_results(results, "sprite collision")  # 8 results

        Args:
            previous_results: Results from a previous search
            refinement_query: Query to refine within the previous results
            max_results: Maximum number of refined results

        Returns:
            Filtered and re-ranked results from the previous search set
        """
        start_time = time.time()

        if not previous_results:
            self.logger.warning("search_within_results called with empty previous results")
            return []

        # Extract unique doc_ids from previous results
        doc_ids = list(set([r['doc_id'] for r in previous_results]))
        self.logger.info(f"Searching within {len(doc_ids)} documents ({len(previous_results)} chunks) for: '{refinement_query}'")

        # Build refinement terms
        refinement_terms = self._preprocess_text(refinement_query)
        if not refinement_terms:
            self.logger.warning(f"No valid terms in refinement query: '{refinement_query}'")
            return previous_results[:max_results]

        # Score each previous result against refinement query
        scored_results = []

        for result in previous_results:
            # Get the full chunk content for better matching
            try:
                chunk = next((c for c in self.chunks
                            if c.doc_id == result['doc_id'] and c.chunk_id == result['chunk_id']),
                           None)

                if not chunk:
                    # Use snippet from search result if chunk not found
                    content = result.get('snippet', '')
                else:
                    content = chunk.content
            except Exception as e:
                self.logger.debug(f"Failed to get chunk content: {e}")
                content = result.get('snippet', '')

            # Calculate relevance score for refinement terms
            relevance_score = 0
            content_lower = content.lower()

            for term in refinement_terms:
                # Count occurrences and weight by position
                occurrences = content_lower.count(term)
                relevance_score += occurrences

                # Boost score if term appears near beginning (30% of content)
                if content_lower[:int(len(content) * 0.3)].count(term) > 0:
                    relevance_score += 2

            # Only include results with at least some match
            if relevance_score > 0:
                scored_results.append({
                    **result,
                    'refinement_score': relevance_score,
                    'original_score': result.get('score', 0)
                })

        # Sort by refinement score (descending), then by original score
        scored_results.sort(key=lambda x: (x['refinement_score'], x['original_score']), reverse=True)

        # Limit results
        refined_results = scored_results[:max_results]

        elapsed_ms = (time.time() - start_time) * 1000
        self.logger.info(f"Search within results: '{refinement_query}' ({len(refined_results)}/{len(previous_results)} results, {elapsed_ms:.2f}ms)")

        return refined_results

    def answer_question(self, question: str, max_context_chunks: int = 5,
                       force_search_mode: Optional[str] = None) -> dict:
        """
        Answer a question using RAG (Retrieval-Augmented Generation).

        Retrieves relevant documentation and uses LLM to synthesize an answer
        with citations to source material.

        Args:
            question: The question to answer about C64 documentation
            max_context_chunks: Maximum number of documentation chunks to use for context
            force_search_mode: Override search mode ('keyword', 'semantic', 'hybrid', or None for auto)

        Returns:
            Dictionary with answer, sources, confidence, and metadata:
            {
                'answer': str,              # Generated answer text
                'sources': list[dict],      # Citations with metadata
                'confidence': float,        # 0.0-1.0 confidence score
                'search_results': list,     # Top-N search results used
                'reasoning': str,           # Explanation of how answer was derived
                'model': str,               # LLM model used
                'error': str or None        # Error message if applicable
            }
        """
        start_time = time.time()

        if not question or len(question.strip()) < 3:
            return {
                'answer': None,
                'sources': [],
                'confidence': 0.0,
                'search_results': [],
                'reasoning': 'Invalid question: too short',
                'model': 'N/A',
                'error': 'Question must be at least 3 characters long'
            }

        self.logger.info(f"Answering question: '{question}'")

        try:
            # Step 1: Translate natural language query to understand intent
            translation = self.translate_nl_query(question, confidence_threshold=0.7)
            query_confidence = translation.get('confidence', 0.5)

            # Step 2: Choose search strategy
            if force_search_mode:
                search_mode = force_search_mode
            else:
                search_mode = translation.get('search_mode', 'hybrid')

            # Step 3: Retrieve relevant context
            if search_mode == 'semantic' and self.use_semantic:
                results = self.semantic_search(question, max_context_chunks * 2)
            elif search_mode == 'hybrid' and self.use_semantic:
                results = self.hybrid_search(question, max_context_chunks * 2)
            else:
                results = self.search(question, max_context_chunks * 2)

            if not results:
                self.logger.warning(f"No search results found for question: {question}")
                return {
                    'answer': f"I could not find relevant documentation to answer: {question}",
                    'sources': [],
                    'confidence': 0.1,
                    'search_results': [],
                    'reasoning': "Search found no results. Try searching directly with search_docs.",
                    'model': 'N/A',
                    'error': 'No relevant documents found'
                }

            # Step 4: Build context from top results
            context = self._build_rag_context(results[:max_context_chunks])

            # Step 5: Try to generate answer with LLM
            from llm_integration import get_llm_client

            llm_client = get_llm_client()

            if llm_client:
                # Generate answer using LLM with context
                answer_result = self._generate_answer_with_llm(
                    question=question,
                    context=context,
                    search_results=results[:max_context_chunks],
                    llm_client=llm_client
                )

                elapsed_ms = (time.time() - start_time) * 1000
                self.logger.info(f"Question answering completed in {elapsed_ms:.2f}ms")

                return {
                    'answer': answer_result['text'],
                    'sources': answer_result['citations'],
                    'confidence': answer_result.get('confidence', query_confidence),
                    'search_results': results[:max_context_chunks],
                    'reasoning': f"Question mode: {search_mode}. Retrieved {len(results[:max_context_chunks])} sources.",
                    'model': llm_client.provider.model,
                    'error': None
                }
            else:
                # Fallback: Return structured search summary without LLM
                self.logger.warning("LLM not configured, returning search-based fallback")
                return self._generate_summary_fallback(
                    question=question,
                    results=results[:max_context_chunks],
                    query_confidence=query_confidence,
                    search_mode=search_mode
                )

        except Exception as e:
            self.logger.error(f"Error answering question: {e}", exc_info=True)
            return {
                'answer': None,
                'sources': [],
                'confidence': 0.0,
                'search_results': [],
                'reasoning': 'An error occurred during answer generation',
                'model': 'N/A',
                'error': str(e)
            }

    def _build_rag_context(self, search_results: list[dict], max_tokens: int = 4000) -> str:
        """
        Build documentation context from search results for LLM prompt.

        Args:
            search_results: Search results with snippets and metadata
            max_tokens: Maximum tokens for context (~4 chars per token)

        Returns:
            Formatted documentation context string with source citations
        """
        context = ""
        token_count = 0
        CHARS_PER_TOKEN = 4
        SECTION_OVERHEAD = 200

        for i, result in enumerate(search_results, 1):
            # Build section header
            header = f"## Source {i}: {result['title']}\n"
            header += f"Document: {result['filename']} (ID: {result['doc_id']})\n"
            header += f"Chunk {result['chunk_id']}"

            if result.get('page'):
                header += f", Page {result['page']}"

            header += f" - Relevance: {result['score']:.2f}\n\n"

            # Get snippet (already highlighted by _extract_snippet)
            snippet = result.get('snippet', '') + "\n\n"

            # Build full section
            section = header + snippet
            section_tokens = len(section) // CHARS_PER_TOKEN

            # Check token budget
            if token_count + section_tokens + SECTION_OVERHEAD > max_tokens:
                remaining = len(search_results) - i
                if remaining > 0:
                    context += f"\n... ({remaining} more sources available, truncated for token limit)\n"
                break

            context += section
            token_count += section_tokens

        return context

    def _generate_answer_with_llm(self, question: str, context: str,
                                 search_results: list[dict],
                                 llm_client) -> dict:
        """
        Generate answer using LLM with retrieved context.

        Args:
            question: The user's question
            context: Formatted documentation context
            search_results: Original search results for citation tracking
            llm_client: LLMClient instance

        Returns:
            Dictionary with answer text and extracted citations
        """
        # Build prompt
        prompt = f"""You are a Commodore 64 documentation expert assistant.

Answer the following question based on the provided documentation excerpts. Be accurate and specific about technical details.

QUESTION: {question}

DOCUMENTATION:
{context}

INSTRUCTIONS:
1. Provide a clear, accurate answer based on the documentation
2. If the answer requires multiple sources, synthesize information from them
3. Be specific about memory addresses (like $D000), register names, and technical specifications
4. Reference which documentation sections you're using (e.g., "Source 1", "Source 2")
5. If the documentation is incomplete or you're unsure, state that explicitly
6. Use correct technical terminology for C64 concepts

Please provide your answer now:"""

        try:
            # Call LLM with low temperature for accuracy
            answer_text = llm_client.call(
                prompt,
                temperature=0.3,        # Deterministic, fact-based responses
                max_tokens=2048         # Enough for detailed answers
            )

            # Extract citations from answer
            citations = self._extract_citations(answer_text, search_results)

            # Calculate confidence based on citations
            if citations:
                confidence = 0.85  # Higher confidence if LLM cited sources
            else:
                confidence = 0.70  # Lower if no explicit citations found

            return {
                'text': answer_text,
                'citations': citations,
                'confidence': confidence
            }

        except Exception as e:
            self.logger.error(f"LLM answer generation failed: {e}")
            raise

    def _extract_citations(self, answer_text: str, search_results: list[dict]) -> list[dict]:
        """
        Extract citations from LLM-generated answer.

        Looks for "Source N" references and maps them to search results.

        Args:
            answer_text: The LLM-generated answer text
            search_results: List of search results to cite

        Returns:
            List of citation dictionaries with full metadata
        """
        citations = []
        seen = set()

        # Find all "Source X" references (case-insensitive)
        source_refs = re.findall(r'[Ss]ource\s+(\d+)', answer_text)

        for ref_str in source_refs:
            try:
                idx = int(ref_str) - 1  # Convert to 0-based index

                if 0 <= idx < len(search_results):
                    result = search_results[idx]
                    key = (result['doc_id'], result['chunk_id'])

                    # Avoid duplicates
                    if key not in seen:
                        citations.append({
                            'doc_id': result['doc_id'],
                            'filename': result['filename'],
                            'title': result['title'],
                            'chunk_id': result['chunk_id'],
                            'page': result.get('page'),
                            'score': result.get('score', 0.0)
                        })
                        seen.add(key)
            except (ValueError, IndexError, KeyError):
                # Skip invalid references
                pass

        return citations

    def _generate_summary_fallback(self, question: str, results: list[dict],
                                   query_confidence: float, search_mode: str) -> dict:
        """
        Fallback response when LLM is unavailable.

        Returns structured summary of search results without LLM-generated answer.

        Args:
            question: The user's question
            results: Search results found
            query_confidence: Confidence from query translation
            search_mode: The search mode that was used

        Returns:
            Response dictionary with search-based answer fallback
        """
        # Build summary from search results
        summary = f"# Search Results for: {question}\n\n"
        summary += f"Found {len(results)} relevant documentation section(s):\n\n"

        sources = []
        for i, result in enumerate(results, 1):
            summary += f"## {i}. {result['title']}\n"
            summary += f"**File**: {result['filename']} | **Chunk**: {result['chunk_id']}"
            if result.get('page'):
                summary += f" | **Page**: {result['page']}"
            summary += f" | **Relevance**: {result['score']:.2f}\n\n"
            summary += f"{result.get('snippet', '')}\n\n"

            sources.append({
                'doc_id': result['doc_id'],
                'filename': result['filename'],
                'title': result['title'],
                'chunk_id': result['chunk_id'],
                'page': result.get('page'),
                'score': result.get('score', 0.0)
            })

        summary += "\n*Note: LLM not configured. Showing search results instead of synthesized answer.*"
        summary += "\n*To enable LLM-based answers, set ANTHROPIC_API_KEY or OPENAI_API_KEY environment variables.*"

        return {
            'answer': summary,
            'sources': sources,
            'confidence': query_confidence * 0.7,  # Reduce confidence for fallback
            'search_results': results,
            'reasoning': f"LLM unavailable. Returned {search_mode} search results.",
            'model': 'fallback-search',
            'error': 'LLM not configured'
        }

    def translate_nl_query(self, query: str, confidence_threshold: float = 0.7) -> dict:
        """
        Translate natural language query to structured search parameters using LLM.

        Parses natural language queries and extracts:
        - Search terms (keywords to search for)
        - Entity mentions (hardware, addresses, people, etc.)
        - Facet filters (hardware types, memory ranges, etc.)
        - Recommended search mode (keyword, semantic, hybrid)

        Args:
            query: Natural language query (e.g., "find information about sprites on the VIC-II chip")
            confidence_threshold: Minimum confidence for entity extraction (0.0-1.0, default: 0.7)

        Returns:
            {
                'original_query': str,                    # Original query
                'search_terms': [str],                    # Extracted keywords
                'facet_filters': {                        # Mapped facet filters
                    'hardware': ['VIC-II'],
                    'registers': ['$D000-$D3FF'],
                    ...
                },
                'search_mode': 'hybrid',                  # Recommended mode: 'keyword', 'semantic', 'hybrid'
                'confidence': 0.85,                       # Overall confidence (0.0-1.0)
                'entities_found': [                       # Extracted entities
                    {'text': 'VIC-II', 'type': 'hardware', 'confidence': 0.95},
                    ...
                ],
                'reasoning': str                          # Explanation of translation
            }

        Example:
            >>> kb.translate_nl_query("find sprite information")
            {
                'search_terms': ['sprite', 'information'],
                'facet_filters': {'hardware': ['VIC-II']},
                'search_mode': 'hybrid',
                'entities_found': [{'text': 'sprite', 'type': 'graphics', 'confidence': 0.8}],
                ...
            }
        """
        self.logger.info(f"Translating natural language query: '{query}'")
        start_time = time.time()

        # Get LLM client
        try:
            from llm_integration import get_llm_client
        except ImportError:
            raise ImportError("llm_integration module not found")

        llm_client = get_llm_client()
        if not llm_client:
            # Fallback: Basic keyword extraction without LLM
            self.logger.warning("LLM not configured, using basic keyword extraction")
            words = query.lower().split()
            search_terms = [w for w in words if len(w) > 3]
            return {
                'original_query': query,
                'search_terms': search_terms,
                'facet_filters': {},
                'search_mode': 'keyword',
                'confidence': 0.5,
                'entities_found': [],
                'reasoning': 'LLM not available - basic keyword extraction used',
                'suggested_query': ' '.join(search_terms) or query
            }

        # Build structured prompt for LLM
        prompt = f"""You are a Commodore 64 technical documentation query parser.

User Query: "{query}"

Parse this query and extract structured information in JSON format:

{{
  "search_terms": ["keyword1", "keyword2"],        // Main keywords to search for
  "entities": [                                     // Detected technical entities
    {{"text": "VIC-II", "type": "hardware", "confidence": 0.95}},
    {{"text": "$D000", "type": "memory_address", "confidence": 0.90}}
  ],
  "search_mode": "hybrid",                         // "keyword", "semantic", or "hybrid"
  "confidence": 0.85,                              // Overall confidence (0.0-1.0)
  "reasoning": "Detected VIC-II chip mention..."  // Brief explanation
}}

Entity Types:
- hardware: Chip names (VIC-II, SID, CIA, 6502, 6526, 6581, etc.)
- memory_address: Memory addresses ($D000, $D020, 53280, 0xD020, etc.)
- instruction: Assembly instructions (LDA, STA, JMP, JSR, etc.)
- person: People mentioned (Bob Yannes, Jack Tramiel, etc.)
- register: Hardware registers (sprite registers, color registers, etc.)
- graphics: Graphics concepts (sprites, bitmap, character mode, etc.)
- audio: Sound concepts (voices, waveforms, ADSR, etc.)

Search Mode Guidelines:
- "keyword": For specific technical terms, exact matches
- "semantic": For conceptual questions, "how does X work", explanations
- "hybrid": For mixed queries with both specific terms and concepts

Return ONLY valid JSON, no additional text.
"""

        try:
            # Call LLM with low temperature for deterministic parsing
            response_text = llm_client.call(prompt, temperature=0.3, max_tokens=512)

            # Parse JSON response
            # Handle potential markdown code blocks
            if '```json' in response_text:
                # Extract JSON from code block
                json_start = response_text.find('```json') + 7
                json_end = response_text.find('```', json_start)
                response_text = response_text[json_start:json_end].strip()
            elif '```' in response_text:
                # Extract from generic code block
                json_start = response_text.find('```') + 3
                json_end = response_text.find('```', json_start)
                response_text = response_text[json_start:json_end].strip()

            parsed = json.loads(response_text)

            # Map entities to facet filters
            facet_filters = {}
            entities_found = parsed.get('entities', [])

            for entity in entities_found:
                if entity['confidence'] < confidence_threshold:
                    continue

                entity_type = entity['type']
                entity_text = entity['text']

                # Map entity types to facet categories
                if entity_type == 'hardware':
                    if 'hardware' not in facet_filters:
                        facet_filters['hardware'] = []
                    facet_filters['hardware'].append(entity_text)

                elif entity_type in ['memory_address', 'register']:
                    if 'registers' not in facet_filters:
                        facet_filters['registers'] = []
                    facet_filters['registers'].append(entity_text)

                elif entity_type == 'instruction':
                    if 'instructions' not in facet_filters:
                        facet_filters['instructions'] = []
                    facet_filters['instructions'].append(entity_text)

            # Build result
            result = {
                'original_query': query,
                'search_terms': parsed.get('search_terms', []),
                'facet_filters': facet_filters,
                'search_mode': parsed.get('search_mode', 'hybrid'),
                'confidence': parsed.get('confidence', 0.7),
                'entities_found': entities_found,
                'reasoning': parsed.get('reasoning', ''),
                'suggested_query': ' '.join(parsed.get('search_terms', [])) or query  # Reformulated query for searching
            }

            elapsed = (time.time() - start_time) * 1000
            self.logger.info(f"Query translation completed in {elapsed:.2f}ms - mode: {result['search_mode']}, "
                           f"entities: {len(entities_found)}, confidence: {result['confidence']:.2f}")

            return result

        except json.JSONDecodeError as e:
            self.logger.error(f"Failed to parse LLM JSON response: {e}")
            self.logger.error(f"Raw response: {response_text}")

            # Fallback to basic keyword extraction
            words = query.lower().split()
            search_terms = [w for w in words if len(w) > 3]
            return {
                'original_query': query,
                'search_terms': search_terms,
                'facet_filters': {},
                'search_mode': 'keyword',
                'confidence': 0.5,
                'entities_found': [],
                'reasoning': f'LLM response parsing failed: {str(e)}',
                'suggested_query': ' '.join(search_terms) or query
            }

        except Exception as e:
            self.logger.error(f"Query translation error: {e}")

            # Fallback to basic keyword extraction
            words = query.lower().split()
            search_terms = [w for w in words if len(w) > 3]
            return {
                'original_query': query,
                'search_terms': search_terms,
                'facet_filters': {},
                'search_mode': 'keyword',
                'confidence': 0.5,
                'entities_found': [],
                'reasoning': f'Translation error: {str(e)}',
                'suggested_query': ' '.join(search_terms) or query
            }

    def faceted_search(self, query: str, facet_filters: Optional[dict[str, list[str]]] = None,
                      max_results: int = 5, tags: Optional[list[str]] = None) -> list[dict]:
        """
        Perform search with faceted filtering.

        Args:
            query: Search query
            facet_filters: Dict of facet_type -> list of values to filter by.
                          Example: {'hardware': ['SID', 'VIC-II'], 'instruction': ['LDA', 'STA']}
            max_results: Maximum number of results to return
            tags: Optional list of tags to filter by

        Returns:
            List of search results filtered by facets, with facets included
        """
        # Check cache first
        start_time = time.time()
        if self._faceted_cache is not None:
            # Convert facet_filters dict to tuple for hashable cache key
            facet_filters_key = None
            if facet_filters:
                facet_filters_key = tuple(sorted((k, tuple(sorted(v))) for k, v in facet_filters.items()))

            cache_key = self._cache_key('faceted_search',
                                       query=query,
                                       facet_filters=facet_filters_key,
                                       max_results=max_results,
                                       tags=tuple(sorted(tags)) if tags else None)
            if cache_key in self._faceted_cache:
                results = self._faceted_cache[cache_key]
                elapsed_ms = (time.time() - start_time) * 1000
                self.logger.debug(f"Faceted cache HIT for query: '{query}' ({len(results)} results, {elapsed_ms:.2f}ms)")
                return results

        self.logger.info(f"Faceted search query: '{query}' (facets={facet_filters}, max_results={max_results}, tags={tags})")

        # First get regular search results (request more for filtering)
        results = self.search(query, max_results * 3, tags)

        # If no facet filters specified, just return regular results with facets added
        if not facet_filters:
            # Add facets to each result
            for result in results:
                result['facets'] = self._get_document_facets(result['doc_id'])
            final_results = results[:max_results]
            elapsed = (time.time() - start_time) * 1000
            self.logger.info(f"Faceted search (no filters) completed: {len(final_results)} results in {elapsed:.2f}ms")

            # Store in cache
            if self._faceted_cache is not None:
                facet_filters_key = None
                cache_key = self._cache_key('faceted_search',
                                           query=query,
                                           facet_filters=facet_filters_key,
                                           max_results=max_results,
                                           tags=tuple(sorted(tags)) if tags else None)
                self._faceted_cache[cache_key] = final_results

            return final_results

        # Filter results by facets
        filtered_results = []
        for result in results:
            # Get document facets
            doc_facets = self._get_document_facets(result['doc_id'])

            # Check if document matches all facet filters
            matches = True
            for facet_type, required_values in facet_filters.items():
                doc_values = doc_facets.get(facet_type, set())
                # Document must have at least one of the required values for this facet type
                if not any(val in doc_values for val in required_values):
                    matches = False
                    break

            if matches:
                result['facets'] = doc_facets
                filtered_results.append(result)

                if len(filtered_results) >= max_results:
                    break

        elapsed = (time.time() - start_time) * 1000
        self.logger.info(f"Faceted search completed: {len(filtered_results)} results in {elapsed:.2f}ms")

        # Log search for analytics
        self._log_search(query, 'faceted', len(filtered_results), elapsed, tags)

        # Store in cache
        if self._faceted_cache is not None:
            facet_filters_key = tuple(sorted((k, tuple(sorted(v))) for k, v in facet_filters.items()))
            cache_key = self._cache_key('faceted_search',
                                       query=query,
                                       facet_filters=facet_filters_key,
                                       max_results=max_results,
                                       tags=tuple(sorted(tags)) if tags else None)
            self._faceted_cache[cache_key] = filtered_results

        return filtered_results

    def _get_document_facets(self, doc_id: str) -> dict[str, set[str]]:
        """Get all facets for a document from the database."""
        cursor = self.db_conn.cursor()
        cursor.execute("""
            SELECT facet_type, facet_value
            FROM document_facets
            WHERE doc_id = ?
        """, (doc_id,))

        facets = {'hardware': set(), 'instruction': set(), 'register': set()}
        for row in cursor.fetchall():
            facet_type, facet_value = row
            if facet_type in facets:
                facets[facet_type].add(facet_value)

        return facets

    def find_by_reference(self, ref_type: str, ref_value: str, max_results: int = 10) -> list[dict]:
        """
        Find documents by cross-reference type and value.

        Args:
            ref_type: Type of reference ('memory_address', 'register_offset', 'page_reference')
            ref_value: The reference value to search for (e.g., '$D020', 'VIC+0', '156')
            max_results: Maximum number of results to return

        Returns:
            List of results with document info, chunk info, and context
        """
        self.logger.info(f"Finding documents by reference: {ref_type}={ref_value}")
        start_time = time.time()

        cursor = self.db_conn.cursor()

        # Query cross_references table
        cursor.execute("""
            SELECT
                xr.doc_id,
                xr.chunk_id,
                xr.ref_type,
                xr.ref_value,
                xr.context,
                d.filename,
                d.title,
                d.tags
            FROM cross_references xr
            JOIN documents d ON xr.doc_id = d.doc_id
            WHERE xr.ref_type = ? AND xr.ref_value = ?
            ORDER BY d.title, xr.chunk_id
            LIMIT ?
        """, (ref_type, ref_value, max_results))

        results = []
        for row in cursor.fetchall():
            doc_id, chunk_id, ref_type, ref_value, context, filename, title, tags_json = row
            results.append({
                'doc_id': doc_id,
                'chunk_id': chunk_id,
                'filename': filename,
                'title': title,
                'ref_type': ref_type,
                'ref_value': ref_value,
                'context': context,
                'tags': json.loads(tags_json) if tags_json else []
            })

        elapsed = (time.time() - start_time) * 1000
        self.logger.info(f"Found {len(results)} references in {elapsed:.2f}ms")

        return results

    def find_similar_documents(self, doc_id: str, chunk_id: Optional[int] = None,
                               max_results: int = 5, tags: Optional[list[str]] = None) -> list[dict]:
        """
        Find documents similar to the given document or chunk.

        Args:
            doc_id: Document ID to find similar documents for
            chunk_id: Optional chunk ID (if None, uses all chunks from document)
            max_results: Maximum number of results to return
            tags: Optional list of tags to filter by

        Returns:
            List of similar documents with similarity scores
        """
        # Check cache first
        if self._similar_cache is not None:
            cache_key = self._cache_key('find_similar',
                                       doc_id=doc_id,
                                       chunk_id=chunk_id,
                                       max_results=max_results,
                                       tags=tuple(sorted(tags)) if tags else None)
            if cache_key in self._similar_cache:
                results = self._similar_cache[cache_key]
                self.logger.debug(f"Cache hit for find_similar: doc_id={doc_id}, chunk_id={chunk_id}")
                return results

        # Prefer semantic search if available
        if self.use_semantic and self.embeddings_index is not None:
            results = self._find_similar_semantic(doc_id, chunk_id, max_results, tags)
        else:
            # Fall back to TF-IDF similarity
            results = self._find_similar_tfidf(doc_id, chunk_id, max_results, tags)

        # Store in cache
        if self._similar_cache is not None:
            cache_key = self._cache_key('find_similar',
                                       doc_id=doc_id,
                                       chunk_id=chunk_id,
                                       max_results=max_results,
                                       tags=tuple(sorted(tags)) if tags else None)
            self._similar_cache[cache_key] = results

        return results

    def _find_similar_semantic(self, doc_id: str, chunk_id: Optional[int],
                               max_results: int, tags: Optional[list[str]]) -> list[dict]:
        """Find similar documents using semantic embeddings."""
        if not self.use_semantic or self.embeddings_model is None:
            raise RuntimeError("Semantic search not available")

        # Build embeddings index if not yet built
        if self.embeddings_index is None or len(self.embeddings_doc_map) == 0:
            self._build_embeddings()
            if self.embeddings_index is None:
                return []

        # Get target embedding(s)
        if chunk_id is not None:
            # Find specific chunk's embedding
            try:
                target_idx = self.embeddings_doc_map.index((doc_id, chunk_id))
                target_embedding = self.embeddings_index.reconstruct(target_idx)
                target_embedding = target_embedding.reshape(1, -1)
            except ValueError:
                self.logger.error(f"Chunk not found in embeddings: {doc_id}, {chunk_id}")
                return []
        else:
            # Average all chunk embeddings for this document
            doc_indices = [i for i, (d, c) in enumerate(self.embeddings_doc_map) if d == doc_id]
            if not doc_indices:
                self.logger.error(f"Document not found in embeddings: {doc_id}")
                return []

            embeddings = np.array([self.embeddings_index.reconstruct(i) for i in doc_indices])
            target_embedding = np.mean(embeddings, axis=0).reshape(1, -1)

        # Normalize for cosine similarity
        faiss.normalize_L2(target_embedding)

        # Search for similar chunks (get more for filtering)
        k = min(max_results * 10, len(self.embeddings_doc_map))
        scores, indices = self.embeddings_index.search(target_embedding, k)

        # Build results, aggregating by document
        doc_scores = {}  # doc_id -> max similarity score
        doc_chunks = {}  # doc_id -> list of matching chunks

        for score, idx in zip(scores[0], indices[0]):
            if idx < 0 or idx >= len(self.embeddings_doc_map):
                continue

            found_doc_id, found_chunk_id = self.embeddings_doc_map[idx]

            # Skip the source document/chunk
            if found_doc_id == doc_id:
                if chunk_id is None or found_chunk_id == chunk_id:
                    continue

            # Filter by tags if specified
            if tags:
                doc = self.documents.get(found_doc_id)
                if doc and not any(t in doc.tags for t in tags):
                    continue

            # Track best score per document
            if found_doc_id not in doc_scores or score > doc_scores[found_doc_id]:
                doc_scores[found_doc_id] = float(score)
                doc_chunks[found_doc_id] = found_chunk_id

        # Sort documents by similarity score
        sorted_docs = sorted(doc_scores.items(), key=lambda x: x[1], reverse=True)[:max_results]

        # Build result list with document info
        results = []
        for found_doc_id, similarity in sorted_docs:
            doc = self.documents.get(found_doc_id)
            if not doc:
                continue

            best_chunk_id = doc_chunks[found_doc_id]
            chunk = self.get_chunk(found_doc_id, best_chunk_id)

            results.append({
                'doc_id': found_doc_id,
                'filename': doc.filename,
                'title': doc.title,
                'chunk_id': best_chunk_id,
                'similarity': similarity,
                'snippet': chunk.content[:300] + "..." if chunk else "",
                'total_chunks': doc.total_chunks,
                'tags': doc.tags
            })

        return results

    def _find_similar_tfidf(self, doc_id: str, chunk_id: Optional[int],
                            max_results: int, tags: Optional[list[str]]) -> list[dict]:
        """Find similar documents using TF-IDF (fallback when semantic search unavailable)."""
        from collections import Counter
        import math

        # Get target document chunks
        target_chunks = self._get_chunks_db(doc_id)
        if not target_chunks:
            return []

        # If specific chunk requested, use only that chunk
        if chunk_id is not None:
            target_chunks = [c for c in target_chunks if c.chunk_id == chunk_id]
            if not target_chunks:
                return []

        # Build target document term vector
        target_terms = []
        for chunk in target_chunks:
            words = chunk.content.lower().split()
            target_terms.extend(words)

        target_tf = Counter(target_terms)
        target_length = math.sqrt(sum(count**2 for count in target_tf.values()))

        # Calculate similarity with all other documents
        doc_similarities = []

        for other_doc_id, other_doc in self.documents.items():
            # Skip source document
            if other_doc_id == doc_id:
                continue

            # Filter by tags
            if tags and not any(t in other_doc.tags for t in tags):
                continue

            # Get chunks for comparison document
            other_chunks = self._get_chunks_db(other_doc_id)
            if not other_chunks:
                continue

            # Build term vector for other document
            other_terms = []
            for chunk in other_chunks:
                words = chunk.content.lower().split()
                other_terms.extend(words)

            other_tf = Counter(other_terms)
            other_length = math.sqrt(sum(count**2 for count in other_tf.values()))

            # Calculate cosine similarity
            dot_product = sum(target_tf[term] * other_tf[term] for term in target_tf if term in other_tf)

            if target_length > 0 and other_length > 0:
                similarity = dot_product / (target_length * other_length)
            else:
                similarity = 0.0

            if similarity > 0:
                # Find best matching chunk
                best_chunk = other_chunks[0] if other_chunks else None

                doc_similarities.append({
                    'doc_id': other_doc_id,
                    'filename': other_doc.filename,
                    'title': other_doc.title,
                    'chunk_id': best_chunk.chunk_id if best_chunk else 0,
                    'similarity': similarity,
                    'snippet': best_chunk.content[:300] + "..." if best_chunk else "",
                    'total_chunks': other_doc.total_chunks,
                    'tags': other_doc.tags
                })

        # Sort by similarity and return top results
        doc_similarities.sort(key=lambda x: x['similarity'], reverse=True)
        return doc_similarities[:max_results]

    def compare_documents(self, doc_id_1: str, doc_id_2: str,
                         comparison_type: str = 'full') -> dict:
        """
        Compare two documents side-by-side with similarity scoring and diff analysis.

        Args:
            doc_id_1: First document ID
            doc_id_2: Second document ID
            comparison_type: Type of comparison ('full', 'metadata', 'content')

        Returns:
            Dictionary with comparison results:
            {
                'similarity_score': float,         # Cosine similarity 0.0-1.0
                'metadata_diff': {                 # Metadata differences
                    'title': [title1, title2],
                    'filename': [filename1, filename2],
                    'tags': {
                        'common': [...],
                        'only_in_doc1': [...],
                        'only_in_doc2': [...]
                    },
                    'file_type': [type1, type2],
                    'total_pages': [pages1, pages2]
                },
                'chunk_count': [count1, count2],   # Number of chunks in each doc
                'content_diff': [                  # Unified diff lines (limited)
                    'diff line 1',
                    'diff line 2',
                    ...
                ],
                'entity_comparison': {             # Entity comparison
                    'common_entities': [           # Entities in both docs
                        {'text': str, 'type': str},
                        ...
                    ],
                    'unique_to_doc1': [...],       # Entities only in doc1
                    'unique_to_doc2': [...],       # Entities only in doc2
                    'total_doc1': int,
                    'total_doc2': int
                },
                'summary': str                     # Human-readable summary
            }

        Raises:
            ValueError: If either document not found

        Examples:
            >>> kb = KnowledgeBase()
            >>> result = kb.compare_documents('doc1', 'doc2')
            >>> print(f"Similarity: {result['similarity_score']:.1%}")
            >>> print(f"Common tags: {result['metadata_diff']['tags']['common']}")
        """
        from collections import Counter
        import math
        import difflib

        # Validate both documents exist
        if doc_id_1 not in self.documents:
            raise ValueError(f"Document not found: {doc_id_1}")
        if doc_id_2 not in self.documents:
            raise ValueError(f"Document not found: {doc_id_2}")

        doc1 = self.documents[doc_id_1]
        doc2 = self.documents[doc_id_2]

        # 1. Calculate Cosine Similarity using TF-IDF
        similarity_score = 0.0
        if comparison_type in ['full', 'content']:
            # Get chunks for both documents
            chunks1 = self._get_chunks_db(doc_id_1)
            chunks2 = self._get_chunks_db(doc_id_2)

            if chunks1 and chunks2:
                # Build term vectors
                terms1 = []
                for chunk in chunks1:
                    words = chunk.content.lower().split()
                    terms1.extend(words)

                terms2 = []
                for chunk in chunks2:
                    words = chunk.content.lower().split()
                    terms2.extend(words)

                tf1 = Counter(terms1)
                tf2 = Counter(terms2)

                length1 = math.sqrt(sum(count**2 for count in tf1.values()))
                length2 = math.sqrt(sum(count**2 for count in tf2.values()))

                # Calculate cosine similarity
                dot_product = sum(tf1[term] * tf2[term] for term in tf1 if term in tf2)

                if length1 > 0 and length2 > 0:
                    similarity_score = dot_product / (length1 * length2)

        # 2. Metadata Comparison
        metadata_diff = {
            'title': [doc1.title, doc2.title],
            'filename': [doc1.filename, doc2.filename],
            'file_type': [doc1.file_type, doc2.file_type],
            'total_pages': [doc1.total_pages, doc2.total_pages],
            'tags': {
                'common': list(set(doc1.tags) & set(doc2.tags)),
                'only_in_doc1': list(set(doc1.tags) - set(doc2.tags)),
                'only_in_doc2': list(set(doc2.tags) - set(doc1.tags))
            }
        }

        # 3. Chunk Count
        chunk_count = [doc1.total_chunks, doc2.total_chunks]

        # 4. Content Diff (limited to avoid huge output)
        content_diff = []
        if comparison_type in ['full', 'content']:
            # Get first few chunks for comparison
            chunks1 = self._get_chunks_db(doc_id_1)
            chunks2 = self._get_chunks_db(doc_id_2)

            if chunks1 and chunks2:
                # Concatenate first 5 chunks from each doc
                text1_sample = '\n\n'.join(c.content for c in chunks1[:5])
                text2_sample = '\n\n'.join(c.content for c in chunks2[:5])

                # Generate unified diff (limit to first 100 lines)
                diff_lines = list(difflib.unified_diff(
                    text1_sample.splitlines(keepends=True),
                    text2_sample.splitlines(keepends=True),
                    fromfile=doc1.filename,
                    tofile=doc2.filename,
                    lineterm=''
                ))
                content_diff = diff_lines[:100]  # Limit to 100 lines

        # 5. Entity Comparison
        entity_comparison = {
            'common_entities': [],
            'unique_to_doc1': [],
            'unique_to_doc2': [],
            'total_doc1': 0,
            'total_doc2': 0
        }

        if comparison_type in ['full', 'metadata']:
            cursor = self.db_conn.cursor()

            # Get entities for doc1
            cursor.execute("""
                SELECT entity_text, entity_type
                FROM document_entities
                WHERE doc_id = ?
            """, (doc_id_1,))
            entities1 = {(row[0], row[1]) for row in cursor.fetchall()}
            entity_comparison['total_doc1'] = len(entities1)

            # Get entities for doc2
            cursor.execute("""
                SELECT entity_text, entity_type
                FROM document_entities
                WHERE doc_id = ?
            """, (doc_id_2,))
            entities2 = {(row[0], row[1]) for row in cursor.fetchall()}
            entity_comparison['total_doc2'] = len(entities2)

            # Find common and unique entities
            common = entities1 & entities2
            unique1 = entities1 - entities2
            unique2 = entities2 - entities1

            entity_comparison['common_entities'] = [
                {'text': e[0], 'type': e[1]} for e in sorted(common)
            ]
            entity_comparison['unique_to_doc1'] = [
                {'text': e[0], 'type': e[1]} for e in sorted(unique1)
            ]
            entity_comparison['unique_to_doc2'] = [
                {'text': e[0], 'type': e[1]} for e in sorted(unique2)
            ]

        # 6. Generate Summary
        summary_parts = []
        summary_parts.append(f"Similarity: {similarity_score:.1%}")

        if metadata_diff['title'][0] == metadata_diff['title'][1]:
            summary_parts.append("Same title")
        else:
            summary_parts.append("Different titles")

        common_tags = len(metadata_diff['tags']['common'])
        if common_tags > 0:
            summary_parts.append(f"{common_tags} common tag(s)")

        if entity_comparison['total_doc1'] > 0 and entity_comparison['total_doc2'] > 0:
            common_ent = len(entity_comparison['common_entities'])
            summary_parts.append(f"{common_ent} common entit{'y' if common_ent == 1 else 'ies'}")

        summary = " | ".join(summary_parts)

        return {
            'similarity_score': round(similarity_score, 4),
            'metadata_diff': metadata_diff,
            'chunk_count': chunk_count,
            'content_diff': content_diff,
            'entity_comparison': entity_comparison,
            'summary': summary
        }

    def get_chunk(self, doc_id: str, chunk_id: int) -> Optional[DocumentChunk]:
        """Get a specific chunk from database."""
        cursor = self.db_conn.cursor()
        cursor.execute("""
            SELECT c.doc_id, d.filename, d.title, c.chunk_id, c.page, c.content, c.word_count
            FROM chunks c
            JOIN documents d ON c.doc_id = d.doc_id
            WHERE c.doc_id = ? AND c.chunk_id = ?
        """, (doc_id, chunk_id))

        row = cursor.fetchone()
        if not row:
            return None

        return DocumentChunk(
            doc_id=row[0],
            filename=row[1],
            title=row[2],
            chunk_id=row[3],
            page=row[4],
            content=row[5],
            word_count=row[6]
        )

    def get_document(self, doc_id: str) -> Optional[dict]:
        """Get document with all its chunks.

        Args:
            doc_id: Document ID

        Returns:
            Dictionary with document metadata and chunks, or None if not found
        """
        if doc_id not in self.documents:
            return None

        doc = self.documents[doc_id]
        chunks = self._get_chunks_db(doc_id)

        return {
            'doc_id': doc.doc_id,
            'title': doc.title,
            'filename': doc.filename,
            'chunks': [
                {
                    'chunk_id': chunk.chunk_id,
                    'content': chunk.content,
                    'page': chunk.page,
                    'word_count': chunk.word_count
                }
                for chunk in chunks
            ]
        }

    def get_document_content(self, doc_id: str) -> Optional[str]:
        """Get the full content of a document from database."""
        chunks = self._get_chunks_db(doc_id)
        if not chunks:
            return None
        return "\n\n".join(c.content for c in chunks)

    def list_documents(self) -> list[DocumentMeta]:
        """List all indexed documents."""
        return list(self.documents.values())
    
    def get_stats(self, use_cache: bool = True) -> dict:
        """
        Get knowledge base statistics from database.

        Args:
            use_cache: If True, use cached results if available (1 minute TTL)

        Returns:
            Dictionary with statistics
        """
        # Check cache first
        if use_cache and self._stats_cache is not None:
            cached_result = self._stats_cache.get('stats')
            if cached_result is not None:
                return cached_result

        cursor = self.db_conn.cursor()

        # Count total chunks and words
        cursor.execute("SELECT COUNT(*), SUM(word_count) FROM chunks")
        total_chunks, total_words = cursor.fetchone()
        total_chunks = total_chunks or 0
        total_words = total_words or 0

        # OPTIMIZED: Use sets comprehension more efficiently
        # Collect file_types and tags in a single pass
        file_types_set = set()
        all_tags_set = set()
        for doc in self.documents.values():
            file_types_set.add(doc.file_type)
            all_tags_set.update(doc.tags)

        stats = {
            'total_documents': len(self.documents),
            'total_chunks': total_chunks,
            'total_words': total_words,
            'file_types': sorted(list(file_types_set)),  # Sorted for consistent output
            'all_tags': sorted(list(all_tags_set))  # Sorted for consistent output
        }

        # Cache the result
        if use_cache and self._stats_cache is not None:
            self._stats_cache['stats'] = stats

        return stats

    def health_check(self, quick_check: bool = True, use_cache: bool = True) -> dict:
        """
        Perform health check on the knowledge base system.

        Args:
            quick_check: If True, skip expensive operations (integrity check, orphaned chunks)
            use_cache: If True, use cached results if available (5 minute TTL)

        Returns:
            Dictionary with health metrics and status information
        """
        # Check cache first
        cache_key = f"health_quick_{quick_check}"
        if use_cache and self._health_cache is not None:
            cached_result = self._health_cache.get(cache_key)
            if cached_result is not None:
                return cached_result

        health = {
            'status': 'healthy',
            'issues': [],
            'metrics': {},
            'features': {},
            'database': {},
            'performance': {}
        }

        try:
            # Database health
            cursor = self.db_conn.cursor()

            # Check database file size
            db_file = Path(self.data_dir) / "knowledge_base.db"
            if db_file.exists():
                db_size_mb = db_file.stat().st_size / (1024 * 1024)
                health['database']['size_mb'] = round(db_size_mb, 2)

                # Warn if database is very large
                if db_size_mb > 1000:  # 1GB
                    health['issues'].append(f"Database size is large: {db_size_mb:.2f} MB")

            # Check table integrity (EXPENSIVE - only on full check)
            if not quick_check:
                cursor.execute("PRAGMA integrity_check")
                integrity = cursor.fetchone()[0]
                health['database']['integrity'] = integrity
                if integrity != 'ok':
                    health['status'] = 'warning'
                    health['issues'].append(f"Database integrity check failed: {integrity}")

            # Document and chunk counts - OPTIMIZED: Single query instead of 3
            cursor.execute("""
                SELECT
                    (SELECT COUNT(*) FROM documents) as doc_count,
                    COUNT(*) as chunk_count,
                    SUM(word_count) as total_words
                FROM chunks
            """)
            doc_count, chunk_count, total_words = cursor.fetchone()
            health['metrics']['documents'] = doc_count or 0
            health['metrics']['chunks'] = chunk_count or 0
            health['metrics']['total_words'] = total_words or 0

            # Check for orphaned chunks (EXPENSIVE - only on full check)
            if not quick_check:
                cursor.execute("""
                    SELECT COUNT(*) FROM chunks c
                    LEFT JOIN documents d ON c.doc_id = d.doc_id
                    WHERE d.doc_id IS NULL
                """)
                orphaned = cursor.fetchone()[0]
                if orphaned > 0:
                    health['status'] = 'warning'
                    health['issues'].append(f"Found {orphaned} orphaned chunks")
                    health['database']['orphaned_chunks'] = orphaned

            # Feature availability
            health['features']['fts5_enabled'] = os.environ.get('USE_FTS5', '0') == '1'
            health['features']['fts5_available'] = self._fts5_available()
            health['features']['semantic_search_enabled'] = self.use_semantic
            # Check if embeddings are loaded OR if embeddings files exist (lazy loading)
            health['features']['semantic_search_available'] = (
                self.embeddings_index is not None or
                (self.embeddings_file.exists() and self.embeddings_map_file.exists())
            )
            health['features']['bm25_enabled'] = os.environ.get('USE_BM25', '1') == '1'
            health['features']['query_preprocessing'] = os.environ.get('USE_QUERY_PREPROCESSING', '1') == '1'

            # Check FTS5 index if enabled
            if health['features']['fts5_enabled']:
                if not health['features']['fts5_available']:
                    health['issues'].append("FTS5 is enabled but index not found")
                    health['status'] = 'warning'

            # Check semantic search if enabled
            if health['features']['semantic_search_enabled']:
                if not health['features']['semantic_search_available']:
                    health['issues'].append("Semantic search enabled but embeddings not built")
                    health['status'] = 'warning'
                else:
                    # Check embeddings file size (works for both loaded and lazy-loaded)
                    if self.embeddings_file.exists():
                        emb_size_mb = self.embeddings_file.stat().st_size / (1024 * 1024)
                        health['features']['embeddings_size_mb'] = round(emb_size_mb, 2)

                        # Get count from loaded index or from map file
                        if self.embeddings_index is not None:
                            health['features']['embeddings_count'] = len(self.embeddings_doc_map)
                        elif self.embeddings_map_file.exists():
                            try:
                                with open(self.embeddings_map_file, 'r') as f:
                                    embeddings_map = json.load(f)
                                    health['features']['embeddings_count'] = len(embeddings_map)
                            except Exception:
                                pass

            # Performance metrics
            health['performance']['cache_enabled'] = self._search_cache is not None
            if self._search_cache is not None:
                from cachetools import TTLCache
                if isinstance(self._search_cache, TTLCache):
                    health['performance']['cache_size'] = len(self._search_cache)
                    health['performance']['cache_capacity'] = self._search_cache.maxsize

            # BM25 index status
            if health['features']['bm25_enabled']:
                health['features']['bm25_index_built'] = self.bm25 is not None

            # Disk space check
            import shutil
            disk_usage = shutil.disk_usage(self.data_dir)
            free_gb = disk_usage.free / (1024 ** 3)
            health['database']['disk_free_gb'] = round(free_gb, 2)

            if free_gb < 1:  # Less than 1GB free
                health['status'] = 'warning'
                health['issues'].append(f"Low disk space: {free_gb:.2f} GB free")

            # Overall status
            if not health['issues']:
                health['status'] = 'healthy'
                health['message'] = 'All systems operational'
            else:
                health['message'] = f"System functional with {len(health['issues'])} issue(s)"

            # Cache the result
            if use_cache and self._health_cache is not None:
                self._health_cache[cache_key] = health

        except Exception as e:
            health['status'] = 'error'
            health['issues'].append(f"Health check error: {str(e)}")
            health['message'] = 'Health check failed'
            self.logger.error(f"Health check error: {e}", exc_info=True)

        return health

    def _log_search(self, query: str, search_mode: str, results_count: int, execution_time_ms: float,
                    tags: Optional[list[str]] = None, clicked_doc_id: Optional[str] = None):
        """Log a search query to the search_log table."""
        try:
            cursor = self.db_conn.cursor()
            cursor.execute("""
                INSERT INTO search_log (query, search_mode, results_count, execution_time_ms, tags, clicked_doc_id)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                query,
                search_mode,
                results_count,
                execution_time_ms,
                ','.join(tags) if tags else None,
                clicked_doc_id
            ))
            self.db_conn.commit()
        except Exception as e:
            self.logger.error(f"Error logging search: {e}")

    def get_search_analytics(self, days: int = 30, limit: int = 100) -> dict:
        """
        Get search analytics and insights.

        Args:
            days: Number of days to analyze (default: 30)
            limit: Maximum number of results for top queries (default: 100)

        Returns:
            Dictionary with analytics data including:
            - total_searches: Total number of searches
            - unique_queries: Number of unique queries
            - avg_results: Average number of results per search
            - avg_execution_time_ms: Average execution time
            - top_queries: Most frequent queries
            - failed_searches: Queries with zero results
            - search_modes: Breakdown by search mode
            - popular_tags: Most frequently used tags
        """
        cursor = self.db_conn.cursor()

        # Calculate cutoff date
        from datetime import datetime, timedelta
        cutoff_date = (datetime.now() - timedelta(days=days)).isoformat()

        analytics = {}

        try:
            # Total searches
            cursor.execute("""
                SELECT COUNT(*) FROM search_log
                WHERE timestamp >= ?
            """, (cutoff_date,))
            analytics['total_searches'] = cursor.fetchone()[0]

            # Unique queries
            cursor.execute("""
                SELECT COUNT(DISTINCT query) FROM search_log
                WHERE timestamp >= ?
            """, (cutoff_date,))
            analytics['unique_queries'] = cursor.fetchone()[0]

            # Average results count
            cursor.execute("""
                SELECT AVG(results_count) FROM search_log
                WHERE timestamp >= ?
            """, (cutoff_date,))
            avg_results = cursor.fetchone()[0]
            analytics['avg_results'] = round(avg_results, 2) if avg_results else 0

            # Average execution time
            cursor.execute("""
                SELECT AVG(execution_time_ms) FROM search_log
                WHERE timestamp >= ? AND execution_time_ms IS NOT NULL
            """, (cutoff_date,))
            avg_time = cursor.fetchone()[0]
            analytics['avg_execution_time_ms'] = round(avg_time, 2) if avg_time else 0

            # Top queries (most frequent)
            cursor.execute("""
                SELECT query, COUNT(*) as count, AVG(results_count) as avg_results
                FROM search_log
                WHERE timestamp >= ?
                GROUP BY query
                ORDER BY count DESC
                LIMIT ?
            """, (cutoff_date, limit))
            analytics['top_queries'] = [
                {
                    'query': row[0],
                    'count': row[1],
                    'avg_results': round(row[2], 1) if row[2] else 0
                }
                for row in cursor.fetchall()
            ]

            # Failed searches (zero results)
            cursor.execute("""
                SELECT query, COUNT(*) as count
                FROM search_log
                WHERE timestamp >= ? AND results_count = 0
                GROUP BY query
                ORDER BY count DESC
                LIMIT ?
            """, (cutoff_date, min(limit, 20)))
            analytics['failed_searches'] = [
                {'query': row[0], 'count': row[1]}
                for row in cursor.fetchall()
            ]

            # Search mode breakdown
            cursor.execute("""
                SELECT search_mode, COUNT(*) as count, AVG(results_count) as avg_results
                FROM search_log
                WHERE timestamp >= ?
                GROUP BY search_mode
                ORDER BY count DESC
            """, (cutoff_date,))
            analytics['search_modes'] = [
                {
                    'mode': row[0],
                    'count': row[1],
                    'avg_results': round(row[2], 1) if row[2] else 0
                }
                for row in cursor.fetchall()
            ]

            # Popular tags
            cursor.execute("""
                SELECT tags, COUNT(*) as count
                FROM search_log
                WHERE timestamp >= ? AND tags IS NOT NULL
                GROUP BY tags
                ORDER BY count DESC
                LIMIT ?
            """, (cutoff_date, 20))
            tag_counts = {}
            for row in cursor.fetchall():
                tags_str = row[0]
                count = row[1]
                # Split tags and count individually
                for tag in tags_str.split(','):
                    tag = tag.strip()
                    tag_counts[tag] = tag_counts.get(tag, 0) + count

            analytics['popular_tags'] = [
                {'tag': tag, 'count': count}
                for tag, count in sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)[:20]
            ]

        except Exception as e:
            self.logger.error(f"Error getting search analytics: {e}")
            analytics['error'] = str(e)

        return analytics

    def search_tables(self, query: str, max_results: int = 5, tags: Optional[list[str]] = None) -> list[dict]:
        """Search for tables in documents using FTS5.

        Returns a list of table dictionaries with structure:
        {
            'doc_id': str,
            'doc_title': str,
            'table_id': int,
            'page': int,
            'markdown': str,
            'row_count': int,
            'col_count': int,
            'score': float
        }
        """
        cursor = self.db_conn.cursor()

        # Check if tables_fts exists
        cursor.execute("""
            SELECT name FROM sqlite_master
            WHERE type='table' AND name='tables_fts'
        """)
        if not cursor.fetchone():
            self.logger.warning("tables_fts index not found, no tables to search")
            return []

        # Build FTS5 query
        # Escape special characters in query
        fts_query = query.replace('"', '""')
        fts_query = f'"{fts_query}"' if ' ' in fts_query else fts_query

        # Search tables_fts
        sql = """
            SELECT t.doc_id, d.title, t.table_id, t.page, t.markdown, t.row_count, t.col_count,
                   tables_fts.rank as score
            FROM tables_fts
            JOIN document_tables t ON tables_fts.doc_id = t.doc_id AND tables_fts.table_id = t.table_id
            JOIN documents d ON t.doc_id = d.doc_id
            WHERE tables_fts MATCH ?
        """

        # Add tag filtering if specified
        if tags:
            tag_conditions = " OR ".join(["d.tags LIKE ?" for _ in tags])
            sql += f" AND ({tag_conditions})"

        sql += " ORDER BY score DESC LIMIT ?"

        # Execute query
        params = [fts_query]
        if tags:
            params.extend([f'%"{tag}"%' for tag in tags])
        params.append(max_results)

        cursor.execute(sql, params)
        results = []

        for row in cursor.fetchall():
            results.append({
                'doc_id': row[0],
                'doc_title': row[1],
                'table_id': row[2],
                'page': row[3],
                'markdown': row[4],
                'row_count': row[5],
                'col_count': row[6],
                'score': abs(row[7])  # FTS5 rank is negative, take absolute value
            })

        self.logger.info(f"Table search for '{query}' returned {len(results)} results")
        return results

    def search_code(self, query: str, max_results: int = 5, block_type: Optional[str] = None,
                    tags: Optional[list[str]] = None) -> list[dict]:
        """Search for code blocks in documents using FTS5.

        Args:
            query: Search query
            max_results: Maximum number of results
            block_type: Filter by code type ('basic', 'assembly', 'hex')
            tags: Filter by document tags

        Returns a list of code block dictionaries with structure:
        {
            'doc_id': str,
            'doc_title': str,
            'block_id': int,
            'page': int,
            'block_type': str,
            'code': str,
            'line_count': int,
            'score': float
        }
        """
        cursor = self.db_conn.cursor()

        # Check if code_fts exists
        cursor.execute("""
            SELECT name FROM sqlite_master
            WHERE type='table' AND name='code_fts'
        """)
        if not cursor.fetchone():
            self.logger.warning("code_fts index not found, no code blocks to search")
            return []

        # Build FTS5 query
        # Escape special characters in query
        fts_query = query.replace('"', '""')
        fts_query = f'"{fts_query}"' if ' ' in fts_query else fts_query

        # Search code_fts
        sql = """
            SELECT c.doc_id, d.title, c.block_id, c.page, c.block_type, c.code, c.line_count,
                   code_fts.rank as score
            FROM code_fts
            JOIN document_code_blocks c ON code_fts.doc_id = c.doc_id AND code_fts.block_id = c.block_id
            JOIN documents d ON c.doc_id = d.doc_id
            WHERE code_fts MATCH ?
        """

        # Add block type filtering if specified
        if block_type:
            sql += " AND c.block_type = ?"

        # Add tag filtering if specified
        if tags:
            tag_conditions = " OR ".join(["d.tags LIKE ?" for _ in tags])
            sql += f" AND ({tag_conditions})"

        sql += " ORDER BY score DESC LIMIT ?"

        # Execute query
        params = [fts_query]
        if block_type:
            params.append(block_type)
        if tags:
            params.extend([f'%"{tag}"%' for tag in tags])
        params.append(max_results)

        cursor.execute(sql, params)
        results = []

        for row in cursor.fetchall():
            results.append({
                'doc_id': row[0],
                'doc_title': row[1],
                'block_id': row[2],
                'page': row[3],
                'block_type': row[4],
                'code': row[5],
                'line_count': row[6],
                'score': abs(row[7])  # FTS5 rank is negative, take absolute value
            })

        self.logger.info(f"Code search for '{query}' returned {len(results)} results (type={block_type or 'all'})")
        return results

    def create_backup(self, dest_dir: str, compress: bool = True) -> str:
        """
        Create full backup of knowledge base.

        Args:
            dest_dir: Destination directory for backup
            compress: Whether to compress backup to zip file (default: True)

        Returns:
            Path to backup (directory or zip file)
        """
        import shutil

        self.logger.info(f"Creating backup to {dest_dir}")
        start_time = time.time()

        # Create backup directory
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"kb_backup_{timestamp}"
        backup_path = Path(dest_dir) / backup_name

        try:
            backup_path.mkdir(parents=True, exist_ok=True)

            # Backup database file
            db_source = Path(self.data_dir) / "knowledge_base.db"
            db_dest = backup_path / "knowledge_base.db"
            if db_source.exists():
                shutil.copy2(db_source, db_dest)
                self.logger.info(f"Backed up database: {db_source.stat().st_size} bytes")

            # Backup embeddings if they exist
            embeddings_path = Path(self.data_dir) / "embeddings.faiss"
            embeddings_map_path = Path(self.data_dir) / "embeddings_map.json"

            if embeddings_path.exists():
                shutil.copy2(embeddings_path, backup_path / "embeddings.faiss")
                self.logger.info(f"Backed up embeddings index: {embeddings_path.stat().st_size} bytes")

            if embeddings_map_path.exists():
                shutil.copy2(embeddings_map_path, backup_path / "embeddings_map.json")
                self.logger.info("Backed up embeddings map")

            # Create metadata file
            metadata = {
                'timestamp': timestamp,
                'created_at': datetime.now().isoformat(),
                'document_count': len(self.documents),
                'total_chunks': sum(doc.total_chunks for doc in self.documents.values()),
                'database_size_bytes': db_source.stat().st_size if db_source.exists() else 0,
                'has_embeddings': embeddings_path.exists(),
                'version': '2.5.0'
            }

            with open(backup_path / "metadata.json", 'w') as f:
                json.dump(metadata, f, indent=2)

            self.logger.info(f"Created backup metadata: {metadata}")

            # Compress if requested
            if compress:
                self.logger.info("Compressing backup...")
                zip_path = shutil.make_archive(str(backup_path), 'zip', backup_path)
                shutil.rmtree(backup_path)  # Remove uncompressed directory

                elapsed = time.time() - start_time
                self.logger.info(f"Backup completed in {elapsed:.2f}s: {zip_path}")
                return zip_path
            else:
                elapsed = time.time() - start_time
                self.logger.info(f"Backup completed in {elapsed:.2f}s: {backup_path}")
                return str(backup_path)

        except Exception as e:
            self.logger.error(f"Backup failed: {str(e)}")
            # Cleanup partial backup
            if backup_path.exists():
                shutil.rmtree(backup_path, ignore_errors=True)
            raise

    def restore_from_backup(self, backup_path: str, verify: bool = True) -> dict:
        """
        Restore knowledge base from backup.

        Args:
            backup_path: Path to backup (directory or zip file)
            verify: Whether to verify backup integrity before restoring (default: True)

        Returns:
            Restoration metadata dict
        """
        import shutil
        import zipfile

        self.logger.info(f"Restoring from backup: {backup_path}")
        start_time = time.time()

        backup_path_obj = Path(backup_path)
        temp_dir = None

        try:
            # Extract if compressed
            if backup_path_obj.suffix == '.zip':
                self.logger.info("Extracting compressed backup...")
                temp_dir = Path(self.data_dir) / f"temp_restore_{int(time.time())}"
                temp_dir.mkdir(parents=True, exist_ok=True)

                with zipfile.ZipFile(backup_path, 'r') as zip_ref:
                    zip_ref.extractall(temp_dir)

                # Find the backup directory inside temp_dir
                extracted_dirs = list(temp_dir.iterdir())
                if len(extracted_dirs) == 1 and extracted_dirs[0].is_dir():
                    restore_from = extracted_dirs[0]
                else:
                    restore_from = temp_dir
            else:
                restore_from = backup_path_obj

            # Verify backup if requested
            if verify:
                self.logger.info("Verifying backup integrity...")

                # Check for required files
                db_file = restore_from / "knowledge_base.db"
                metadata_file = restore_from / "metadata.json"

                if not db_file.exists():
                    raise ValueError("Backup is missing database file")

                if not metadata_file.exists():
                    raise ValueError("Backup is missing metadata file")

                # Load and validate metadata
                with open(metadata_file, 'r') as f:
                    metadata = json.load(f)

                self.logger.info(f"Backup metadata: {metadata}")

            # Close current database connection
            self.close()

            # Backup current database before overwriting (safety measure)
            current_db = Path(self.data_dir) / "knowledge_base.db"
            if current_db.exists():
                safety_backup = Path(self.data_dir) / f"knowledge_base_pre_restore_{int(time.time())}.db"
                shutil.copy2(current_db, safety_backup)
                self.logger.info(f"Created safety backup: {safety_backup}")

            # Restore database
            db_source = restore_from / "knowledge_base.db"
            db_dest = Path(self.data_dir) / "knowledge_base.db"
            shutil.copy2(db_source, db_dest)
            self.logger.info(f"Restored database: {db_source.stat().st_size} bytes")

            # Restore embeddings if they exist in backup
            embeddings_source = restore_from / "embeddings.faiss"
            embeddings_map_source = restore_from / "embeddings_map.json"

            if embeddings_source.exists():
                embeddings_dest = Path(self.data_dir) / "embeddings.faiss"
                shutil.copy2(embeddings_source, embeddings_dest)
                self.logger.info("Restored embeddings index")

            if embeddings_map_source.exists():
                embeddings_map_dest = Path(self.data_dir) / "embeddings_map.json"
                shutil.copy2(embeddings_map_source, embeddings_map_dest)
                self.logger.info("Restored embeddings map")

            # Reload knowledge base
            self.logger.info("Reloading knowledge base...")
            self._init_database()
            self._load_documents()

            # Reload embeddings if they exist
            if self.use_semantic and embeddings_source.exists():
                self._load_embeddings()

            elapsed = time.time() - start_time

            result = {
                'success': True,
                'backup_metadata': metadata,
                'restored_documents': len(self.documents),
                'elapsed_seconds': elapsed
            }

            self.logger.info(f"Restore completed in {elapsed:.2f}s: {len(self.documents)} documents")
            return result

        except Exception as e:
            self.logger.error(f"Restore failed: {str(e)}")
            raise
        finally:
            # Cleanup temporary extraction directory
            if temp_dir and temp_dir.exists():
                shutil.rmtree(temp_dir, ignore_errors=True)

    def close(self):
        """Close the database connection and shutdown background workers."""
        # Signal worker threads to shutdown
        if hasattr(self, '_extraction_shutdown'):
            self._extraction_shutdown.set()
            self.logger.info("Signaled entity extraction worker to shutdown")

            # Wait for worker thread to finish (with timeout)
            if hasattr(self, '_extraction_worker') and self._extraction_worker.is_alive():
                self._extraction_worker.join(timeout=10.0)
                if self._extraction_worker.is_alive():
                    self.logger.warning("Entity extraction worker did not shutdown cleanly")
                else:
                    self.logger.info("Entity extraction worker shutdown complete")

        # Close database connection
        if self.db_conn:
            self.db_conn.close()
            self.logger.info("Database connection closed")


# Initialize the MCP server
server = Server("tdz-c64-knowledge")

# Get data directory from environment or use default
DATA_DIR = os.environ.get("TDZ_DATA_DIR", os.path.expanduser("~/.tdz-c64-knowledge"))
kb = KnowledgeBase(DATA_DIR)


@server.list_tools()
async def list_tools() -> list[Tool]:
    """List available tools."""
    return [
        Tool(
            name="search_docs",
            description="Search the C64 knowledge base for information. Use this to find documentation about memory maps, opcodes, BASIC commands, SID, VIC-II, CIA chips, etc.",
            inputSchema={
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Search query (keywords or phrases)"
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of results (default: 5)",
                        "default": 5
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter by document tags (optional)"
                    }
                },
                "required": ["query"]
            }
        ),
        Tool(
            name="translate_query",
            description="Translate natural language query to structured search parameters. Parses queries like 'find sprite information' or 'how does the SID chip work' into search terms, entity mentions, and facet filters. Returns recommended search mode (keyword/semantic/hybrid) and confidence score.",
            inputSchema={
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Natural language query to translate (e.g., 'find information about VIC-II sprites')"
                    },
                    "confidence_threshold": {
                        "type": "number",
                        "description": "Minimum confidence for entity extraction (0.0-1.0, default: 0.7)",
                        "default": 0.7
                    }
                },
                "required": ["query"]
            }
        ),
        Tool(
            name="get_chunk",
            description="Get the full content of a specific document chunk. Use after search_docs to read more context.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID from search results"
                    },
                    "chunk_id": {
                        "type": "integer",
                        "description": "Chunk ID from search results"
                    }
                },
                "required": ["doc_id", "chunk_id"]
            }
        ),
        Tool(
            name="get_document",
            description="Get the full content of a document by ID.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID"
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="list_docs",
            description="List all documents in the C64 knowledge base.",
            inputSchema={
                "type": "object",
                "properties": {}
            }
        ),
        Tool(
            name="add_document",
            description="Add a PDF or text file to the knowledge base.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Full path to the PDF or text file"
                    },
                    "title": {
                        "type": "string",
                        "description": "Document title (optional, defaults to filename)"
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Tags for categorization (e.g., 'memory-map', 'sid', 'basic', 'assembly')"
                    }
                },
                "required": ["filepath"]
            }
        ),
        Tool(
            name="scrape_url",
            description="Scrape a documentation website and add all pages to the knowledge base. Supports recursive scraping of entire sites by following links. Great for ingesting online documentation like http://www.sidmusic.org/sid/. Converts HTML to searchable markdown.",
            inputSchema={
                "type": "object",
                "properties": {
                    "url": {
                        "type": "string",
                        "description": "Starting URL to scrape (e.g., http://www.sidmusic.org/sid/)"
                    },
                    "title": {
                        "type": "string",
                        "description": "Base title for scraped documents (optional, defaults to page titles)"
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Tags for scraped documents (domain name auto-added)"
                    },
                    "follow_links": {
                        "type": "boolean",
                        "description": "Follow links to scrape sub-pages (default: true). Set to false to scrape only the single page.",
                        "default": True
                    },
                    "same_domain_only": {
                        "type": "boolean",
                        "description": "Only follow links on the same domain (default: true). Prevents scraping external sites.",
                        "default": True
                    },
                    "max_pages": {
                        "type": "integer",
                        "description": "Maximum number of pages to scrape (default: 50)",
                        "default": 50,
                        "minimum": 1,
                        "maximum": 500
                    },
                    "depth": {
                        "type": "integer",
                        "description": "Maximum link depth to follow (default: 3). Depth of 1=single page, 2=linked pages, 3=two levels deep.",
                        "default": 3,
                        "minimum": 1,
                        "maximum": 10
                    },
                    "limit": {
                        "type": "string",
                        "description": "Advanced: Only scrape URLs with this prefix (overrides same_domain_only)"
                    },
                    "threads": {
                        "type": "integer",
                        "description": "Number of concurrent download threads (default: 10)",
                        "default": 10,
                        "minimum": 1,
                        "maximum": 20
                    },
                    "delay": {
                        "type": "integer",
                        "description": "Delay between requests in milliseconds (default: 100)",
                        "default": 100,
                        "minimum": 0,
                        "maximum": 5000
                    },
                    "selector": {
                        "type": "string",
                        "description": "CSS selector for main content (optional, auto-detected)"
                    }
                },
                "required": ["url"]
            }
        ),
        Tool(
            name="rescrape_document",
            description="Re-scrape a URL-sourced document to check for updates. Removes the old version and re-scrapes the original URL with the same configuration.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID to re-scrape (must be a URL-sourced document)"
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="check_url_updates",
            description="Check all URL-sourced documents for updates by comparing Last-Modified headers. Detects when source URLs have been modified since last scrape.",
            inputSchema={
                "type": "object",
                "properties": {
                    "auto_rescrape": {
                        "type": "boolean",
                        "description": "Automatically re-scrape changed URLs (default: false)",
                        "default": False
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="remove_document",
            description="Remove a document from the knowledge base.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID to remove"
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="semantic_search",
            description="Search the knowledge base using semantic/conceptual similarity (requires USE_SEMANTIC_SEARCH=1). Finds documents based on meaning, not just keywords. Example: searching for 'movable objects' can find 'sprites'.",
            inputSchema={
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Search query (natural language)"
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of results (default: 5)",
                        "default": 5
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter by document tags (optional)"
                    }
                },
                "required": ["query"]
            }
        ),
        Tool(
            name="find_similar",
            description="Find documents similar to a given document. Uses semantic embeddings if available, falls back to TF-IDF. Great for discovering related content.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID to find similar documents for"
                    },
                    "chunk_id": {
                        "type": "integer",
                        "description": "Optional chunk ID (if omitted, uses entire document)"
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of results (default: 5)",
                        "default": 5
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter by document tags (optional)"
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="hybrid_search",
            description="Perform hybrid search combining FTS5 keyword search and semantic search. Best of both worlds - finds exact keyword matches AND conceptually related content. Returns results ranked by weighted combination of both scores.",
            inputSchema={
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Search query"
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of results (default: 5)",
                        "default": 5
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter by document tags (optional)"
                    },
                    "semantic_weight": {
                        "type": "number",
                        "description": "Weight for semantic score, 0.0-1.0 (default: 0.3). Higher values favor conceptual matches, lower values favor exact keyword matches.",
                        "default": 0.3
                    }
                },
                "required": ["query"]
            }
        ),
        Tool(
            name="fuzzy_search",
            description="Search with typo tolerance using fuzzy string matching. Handles misspellings and variations like 'VIC2' → 'VIC-II', 'asembly' → 'assembly', '6052' → '6502'. Returns exact matches first, then fuzzy matches if needed.",
            inputSchema={
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Search query (may contain typos)"
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of results (default: 5)",
                        "default": 5
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter by document tags (optional)"
                    },
                    "similarity_threshold": {
                        "type": "integer",
                        "description": "Minimum similarity score 0-100 (default: 80). Lower values are more forgiving of typos.",
                        "default": 80,
                        "minimum": 0,
                        "maximum": 100
                    }
                },
                "required": ["query"]
            }
        ),
        Tool(
            name="search_within_results",
            description="Refine previous search results with an additional query. Useful for progressive search refinement: first search broadly (e.g., 'VIC-II'), then refine (e.g., 'sprite collision'). Returns filtered and re-ranked results from previous search set.",
            inputSchema={
                "type": "object",
                "properties": {
                    "previous_results": {
                        "type": "array",
                        "items": {"type": "object"},
                        "description": "Results from a previous search (pass the full result objects)"
                    },
                    "refinement_query": {
                        "type": "string",
                        "description": "Query to refine the previous results"
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of refined results (default: 5)",
                        "default": 5
                    }
                },
                "required": ["previous_results", "refinement_query"]
            }
        ),
        Tool(
            name="suggest_tags",
            description="Get tag suggestions for a document based on content analysis. Detects hardware components (SID, VIC-II, CIA), programming topics (assembly, BASIC, graphics), document types (reference, tutorial), and difficulty levels. Useful for organizing documents.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID to analyze"
                    },
                    "confidence_threshold": {
                        "type": "number",
                        "description": "Minimum confidence for suggestions 0.0-1.0 (default: 0.6)",
                        "default": 0.6,
                        "minimum": 0.0,
                        "maximum": 1.0
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="get_tags_by_category",
            description="Browse all tags organized by category (hardware, programming, document-type, difficulty, custom). Shows tag usage count and sample documents. Useful for discovering and organizing content.",
            inputSchema={
                "type": "object",
                "properties": {}
            }
        ),
        Tool(
            name="answer_question",
            description="Answer questions about C64 documentation using RAG (Retrieval-Augmented Generation). Synthesizes information from multiple sources with citations. Returns answer text with source references and confidence score.",
            inputSchema={
                "type": "object",
                "properties": {
                    "question": {
                        "type": "string",
                        "description": "The question to answer about C64 documentation"
                    },
                    "max_sources": {
                        "type": "integer",
                        "description": "Maximum number of documentation sources to use for context (default: 5)",
                        "default": 5,
                        "minimum": 1,
                        "maximum": 20
                    },
                    "search_mode": {
                        "type": "string",
                        "enum": ["auto", "keyword", "semantic", "hybrid"],
                        "description": "Search strategy to use (default: auto for intelligent selection)",
                        "default": "auto"
                    }
                },
                "required": ["question"]
            }
        ),
        Tool(
            name="kb_stats",
            description="Get statistics about the knowledge base.",
            inputSchema={
                "type": "object",
                "properties": {}
            }
        ),
        Tool(
            name="health_check",
            description="Perform health check on the knowledge base system. Returns status, metrics, feature availability, and any issues detected.",
            inputSchema={
                "type": "object",
                "properties": {}
            }
        ),
        Tool(
            name="detect_anomalies",
            description="Detect anomalies in URL monitoring history. Analyzes patterns to identify unusual update frequencies, performance degradation, or unexpected content changes. Returns anomalies with severity scores (normal/minor/moderate/critical) based on learned baselines.",
            inputSchema={
                "type": "object",
                "properties": {
                    "min_severity": {
                        "type": "string",
                        "description": "Minimum severity level to include ('normal', 'minor', 'moderate', 'critical', default: 'moderate')",
                        "enum": ["normal", "minor", "moderate", "critical"],
                        "default": "moderate"
                    },
                    "days": {
                        "type": "integer",
                        "description": "Number of days of history to analyze (default: 7)",
                        "default": 7,
                        "minimum": 1,
                        "maximum": 90
                    }
                }
            }
        ),
        Tool(
            name="check_updates",
            description="Check all indexed documents for updates. Detects files that have been modified since indexing and optionally re-indexes them automatically.",
            inputSchema={
                "type": "object",
                "properties": {
                    "auto_update": {
                        "type": "boolean",
                        "description": "Automatically re-index changed documents (default: false)",
                        "default": False
                    }
                }
            }
        ),
        Tool(
            name="add_documents_bulk",
            description="Add multiple documents from a directory at once. Supports glob patterns for file matching.",
            inputSchema={
                "type": "object",
                "properties": {
                    "directory": {
                        "type": "string",
                        "description": "Directory to search for documents"
                    },
                    "pattern": {
                        "type": "string",
                        "description": "Glob pattern (default: **/*.{pdf,txt})",
                        "default": "**/*.{pdf,txt}"
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Tags to apply to all documents (optional)"
                    },
                    "recursive": {
                        "type": "boolean",
                        "description": "Search subdirectories (default: true)",
                        "default": True
                    },
                    "skip_duplicates": {
                        "type": "boolean",
                        "description": "Skip files with duplicate content (default: true)",
                        "default": True
                    }
                },
                "required": ["directory"]
            }
        ),
        Tool(
            name="remove_documents_bulk",
            description="Remove multiple documents by IDs or tags. Useful for cleaning up the knowledge base.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_ids": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "List of document IDs to remove (optional)"
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Remove all documents with these tags (optional)"
                    }
                }
            }
        ),
        Tool(
            name="update_tags_bulk",
            description="Update tags for multiple documents in bulk. Add, remove, or replace tags for documents selected by ID or existing tags. Useful for reorganizing and categorizing the knowledge base.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_ids": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "List of document IDs to update (optional, use existing_tags to find documents)"
                    },
                    "existing_tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Find documents with any of these tags (alternative to doc_ids)"
                    },
                    "add_tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Tags to add to the documents"
                    },
                    "remove_tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Tags to remove from the documents"
                    },
                    "replace_tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Replace all tags with these tags"
                    }
                }
            }
        ),
        Tool(
            name="export_documents_bulk",
            description="Export metadata for multiple documents in JSON, CSV, or Markdown format. Useful for creating reports, backups, or sharing document lists.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_ids": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "List of document IDs to export (optional, defaults to all or filtered by tags)"
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Export documents with any of these tags (optional)"
                    },
                    "format": {
                        "type": "string",
                        "description": "Export format (default: json)",
                        "enum": ["json", "csv", "markdown"],
                        "default": "json"
                    }
                }
            }
        ),
        Tool(
            name="search_tables",
            description="Search for tables in PDF documents. Tables contain structured data like memory maps, register definitions, and command references. Returns tables in markdown format with page numbers.",
            inputSchema={
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Search query for table content"
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of results (default: 5)",
                        "default": 5
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter by document tags (optional)"
                    }
                },
                "required": ["query"]
            }
        ),
        Tool(
            name="search_code",
            description="Search for code blocks in documents (BASIC, Assembly, Hex dumps). Finds programming examples and code snippets. Returns code with type (basic/assembly/hex), line count, and page numbers.",
            inputSchema={
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Search query for code content"
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of results (default: 5)",
                        "default": 5
                    },
                    "block_type": {
                        "type": "string",
                        "description": "Filter by code type: 'basic', 'assembly', or 'hex' (optional)",
                        "enum": ["basic", "assembly", "hex"]
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter by document tags (optional)"
                    }
                },
                "required": ["query"]
            }
        ),
        Tool(
            name="faceted_search",
            description="Search with faceted filtering. Filter results by hardware components (SID, VIC-II, CIA, etc.), assembly instructions (LDA, STA, etc.), or memory registers ($D000, etc.). Great for narrowing down search results to specific technical domains.",
            inputSchema={
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Search query"
                    },
                    "facet_filters": {
                        "type": "object",
                        "description": "Facet filters as dict of facet_type -> list of values. Example: {'hardware': ['SID', 'VIC-II'], 'instruction': ['LDA', 'STA']}",
                        "additionalProperties": {
                            "type": "array",
                            "items": {"type": "string"}
                        }
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of results (default: 5)",
                        "default": 5
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter by document tags (optional)"
                    }
                },
                "required": ["query"]
            }
        ),
        Tool(
            name="search_analytics",
            description="Get search analytics and insights. Shows popular queries, failed searches, search mode usage, and performance metrics.",
            inputSchema={
                "type": "object",
                "properties": {
                    "days": {
                        "type": "integer",
                        "description": "Number of days to analyze (default: 30)",
                        "default": 30
                    },
                    "limit": {
                        "type": "integer",
                        "description": "Maximum number of results for top queries (default: 100)",
                        "default": 100
                    }
                }
            }
        ),
        Tool(
            name="find_by_reference",
            description="Find documents by cross-reference. Search for documents containing specific memory addresses ($D020), register offsets (VIC+0, SID+4), or page references (page 156). Great for tracking how specific registers or memory locations are documented.",
            inputSchema={
                "type": "object",
                "properties": {
                    "ref_type": {
                        "type": "string",
                        "description": "Type of reference to search for",
                        "enum": ["memory_address", "register_offset", "page_reference"]
                    },
                    "ref_value": {
                        "type": "string",
                        "description": "The reference value (e.g., '$D020', 'VIC+0', '156')"
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of results (default: 10)",
                        "default": 10
                    }
                },
                "required": ["ref_type", "ref_value"]
            }
        ),
        Tool(
            name="suggest_queries",
            description="Get autocomplete suggestions for partial queries. Suggests technical terms, memory addresses, instructions, and concepts based on indexed content. Great for discovering searchable content and learning proper terminology.",
            inputSchema={
                "type": "object",
                "properties": {
                    "partial": {
                        "type": "string",
                        "description": "Partial query string (e.g., 'VIC', 'SID', '$D0')"
                    },
                    "max_suggestions": {
                        "type": "integer",
                        "description": "Maximum number of suggestions (default: 5)",
                        "default": 5
                    },
                    "category": {
                        "type": "string",
                        "description": "Optional category filter",
                        "enum": ["hardware", "register", "instruction", "concept"]
                    }
                },
                "required": ["partial"]
            }
        ),
        Tool(
            name="export_results",
            description="Export search results to various formats (markdown, json, html). Use this to save search results for offline use, sharing, or creating custom reference guides.",
            inputSchema={
                "type": "object",
                "properties": {
                    "results": {
                        "type": "array",
                        "description": "Search results array (from any search method)"
                    },
                    "format": {
                        "type": "string",
                        "description": "Export format",
                        "enum": ["markdown", "json", "html"],
                        "default": "markdown"
                    },
                    "query": {
                        "type": "string",
                        "description": "Optional query string to include in export"
                    }
                },
                "required": ["results"]
            }
        ),
        Tool(
            name="create_backup",
            description="Create a backup of the knowledge base. Backs up database and embeddings to a zip file. Use this regularly for data safety and before making major changes.",
            inputSchema={
                "type": "object",
                "properties": {
                    "dest_dir": {
                        "type": "string",
                        "description": "Destination directory for backup"
                    },
                    "compress": {
                        "type": "boolean",
                        "description": "Whether to compress backup to zip file (default: true)",
                        "default": True
                    }
                },
                "required": ["dest_dir"]
            }
        ),
        Tool(
            name="restore_backup",
            description="Restore knowledge base from a backup. WARNING: This will replace the current database. A safety backup is created automatically before restoration.",
            inputSchema={
                "type": "object",
                "properties": {
                    "backup_path": {
                        "type": "string",
                        "description": "Path to backup file or directory"
                    },
                    "verify": {
                        "type": "boolean",
                        "description": "Whether to verify backup integrity before restoring (default: true)",
                        "default": True
                    }
                },
                "required": ["backup_path"]
            }
        ),
        Tool(
            name="auto_tag_document",
            description="Automatically generate tags for a document using AI analysis. Analyzes document content and suggests relevant tags across categories: hardware (sid, vic-ii), programming (assembly, basic), document type (tutorial, reference), and difficulty level (beginner, advanced). Requires LLM configuration (set LLM_PROVIDER and API key).",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID to tag"
                    },
                    "confidence_threshold": {
                        "type": "number",
                        "description": "Minimum confidence to accept tag (0.0-1.0, default: 0.7)",
                        "default": 0.7,
                        "minimum": 0.0,
                        "maximum": 1.0
                    },
                    "max_tags": {
                        "type": "integer",
                        "description": "Maximum number of tags to suggest (default: 10)",
                        "default": 10
                    },
                    "append": {
                        "type": "boolean",
                        "description": "If true, append to existing tags; if false, replace (default: true)",
                        "default": True
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="auto_tag_all",
            description="Bulk auto-tag multiple documents using AI. Analyzes content and suggests relevant tags for all documents (or subset). Useful for initial organization or re-tagging collections. Can skip already-tagged documents and limit processing count. Requires LLM configuration.",
            inputSchema={
                "type": "object",
                "properties": {
                    "confidence_threshold": {
                        "type": "number",
                        "description": "Minimum confidence to accept tag (0.0-1.0, default: 0.7)",
                        "default": 0.7,
                        "minimum": 0.0,
                        "maximum": 1.0
                    },
                    "max_tags": {
                        "type": "integer",
                        "description": "Maximum tags per document (default: 10)",
                        "default": 10
                    },
                    "append": {
                        "type": "boolean",
                        "description": "If true, append to existing tags; if false, replace (default: true)",
                        "default": True
                    },
                    "skip_tagged": {
                        "type": "boolean",
                        "description": "If true, skip documents that already have tags (default: true)",
                        "default": True
                    },
                    "max_docs": {
                        "type": "integer",
                        "description": "Maximum number of documents to process (optional, for testing or rate limiting)"
                    }
                }
            }
        ),
        Tool(
            name="summarize_document",
            description="Generate an AI-powered summary of a document. Supports brief (200-300 words), detailed (500-800 words), or bullet-point summaries.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID to summarize"
                    },
                    "summary_type": {
                        "type": "string",
                        "enum": ["brief", "detailed", "bullet"],
                        "description": "Type of summary: 'brief' (default), 'detailed', or 'bullet'",
                        "default": "brief"
                    },
                    "force_regenerate": {
                        "type": "boolean",
                        "description": "If true, regenerate summary even if cached version exists (default: false)",
                        "default": False
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="get_summary",
            description="Retrieve a cached summary of a document without regenerating it.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID"
                    },
                    "summary_type": {
                        "type": "string",
                        "enum": ["brief", "detailed", "bullet"],
                        "description": "Type of summary (default: 'brief')",
                        "default": "brief"
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="summarize_all",
            description="Bulk generate summaries for all documents in the knowledge base.",
            inputSchema={
                "type": "object",
                "properties": {
                    "summary_types": {
                        "type": "array",
                        "items": {"type": "string", "enum": ["brief", "detailed", "bullet"]},
                        "description": "Types of summaries to generate (default: ['brief'])"
                    },
                    "force_regenerate": {
                        "type": "boolean",
                        "description": "If true, regenerate all summaries (default: false)",
                        "default": False
                    },
                    "max_docs": {
                        "type": "integer",
                        "description": "Maximum number of documents to process (optional, for testing)"
                    }
                }
            }
        ),
        Tool(
            name="extract_entities",
            description="Extract named entities from a C64 document using AI. Identifies hardware (SID, VIC-II, CIA, 6502), memory addresses ($D000), assembly instructions (LDA, STA), people, companies, products, and technical concepts. Returns entities with type, confidence score, and context. Requires LLM configuration.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID to extract entities from"
                    },
                    "confidence_threshold": {
                        "type": "number",
                        "description": "Minimum confidence to include entity (0.0-1.0, default: 0.6)",
                        "default": 0.6,
                        "minimum": 0.0,
                        "maximum": 1.0
                    },
                    "force_regenerate": {
                        "type": "boolean",
                        "description": "Force re-extraction even if entities already exist (default: false)",
                        "default": False
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="list_entities",
            description="List all entities extracted from a document, grouped by type (hardware, memory_address, instruction, person, company, product, concept). Great for getting an overview of what a document covers.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID"
                    },
                    "entity_types": {
                        "type": "array",
                        "items": {
                            "type": "string",
                            "enum": ["hardware", "memory_address", "instruction", "person", "company", "product", "concept"]
                        },
                        "description": "Filter by entity types (optional, returns all types if omitted)"
                    },
                    "min_confidence": {
                        "type": "number",
                        "description": "Minimum confidence threshold (0.0-1.0, default: 0.0)",
                        "default": 0.0,
                        "minimum": 0.0,
                        "maximum": 1.0
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="search_entities",
            description="Search for entities across all documents using full-text search. Find all documents mentioning specific hardware, addresses, instructions, people, companies, products, or concepts.",
            inputSchema={
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Search query (e.g., 'VIC-II', 'sprite', '$D000')"
                    },
                    "entity_types": {
                        "type": "array",
                        "items": {
                            "type": "string",
                            "enum": ["hardware", "memory_address", "instruction", "person", "company", "product", "concept"]
                        },
                        "description": "Filter by entity types (optional)"
                    },
                    "min_confidence": {
                        "type": "number",
                        "description": "Minimum confidence threshold (0.0-1.0, default: 0.0)",
                        "default": 0.0,
                        "minimum": 0.0,
                        "maximum": 1.0
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of documents to return (default: 20)",
                        "default": 20,
                        "minimum": 1,
                        "maximum": 100
                    }
                },
                "required": ["query"]
            }
        ),
        Tool(
            name="entity_stats",
            description="Get statistics about extracted entities in the knowledge base. Shows breakdown by type, top entities, and documents with most entities. Useful for understanding the knowledge base content.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity_type": {
                        "type": "string",
                        "enum": ["hardware", "memory_address", "instruction", "person", "company", "product", "concept"],
                        "description": "Filter statistics by entity type (optional, shows all types if omitted)"
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="get_entity_analytics",
            description="Get comprehensive entity analytics for dashboard visualization. Provides entity distribution by type, top entities by document count, relationship statistics, top entity relationships, and extraction timeline trends over time.",
            inputSchema={
                "type": "object",
                "properties": {
                    "time_range_days": {
                        "type": "integer",
                        "description": "Number of days to include in timeline analysis (default: 30)",
                        "default": 30,
                        "minimum": 1,
                        "maximum": 365
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="extract_entities_bulk",
            description="Bulk extract entities from multiple documents in the knowledge base. Processes documents in batch, skips documents that already have entities (unless force_regenerate). Returns statistics about processed documents and extracted entities.",
            inputSchema={
                "type": "object",
                "properties": {
                    "confidence_threshold": {
                        "type": "number",
                        "description": "Minimum confidence to include entity (0.0-1.0, default: 0.6)",
                        "default": 0.6,
                        "minimum": 0.0,
                        "maximum": 1.0
                    },
                    "force_regenerate": {
                        "type": "boolean",
                        "description": "Force re-extraction even if entities already exist (default: false)",
                        "default": False
                    },
                    "max_docs": {
                        "type": "integer",
                        "description": "Maximum number of documents to process (optional, for testing)",
                        "minimum": 1
                    },
                    "skip_existing": {
                        "type": "boolean",
                        "description": "Skip documents that already have entities (default: true)",
                        "default": True
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="extract_entity_relationships",
            description="Extract co-occurrence relationships between entities in a document. Analyzes how entities appear together (e.g., VIC-II + raster interrupt, SID + sound programming). Returns entity pairs with relationship strength and context snippets.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID to extract relationships from"
                    },
                    "min_confidence": {
                        "type": "number",
                        "description": "Minimum confidence threshold for entities (0.0-1.0, default: 0.6)",
                        "default": 0.6,
                        "minimum": 0.0,
                        "maximum": 1.0
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="get_entity_relationships",
            description="Get all entities related to a specific entity. Shows which other entities frequently co-occur with the target entity, sorted by relationship strength. Great for discovering related concepts, hardware, and techniques.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity_text": {
                        "type": "string",
                        "description": "The entity to find relationships for (e.g., 'VIC-II', 'SID', 'LDA')"
                    },
                    "min_strength": {
                        "type": "number",
                        "description": "Minimum relationship strength (0.0-1.0, default: 0.0)",
                        "default": 0.0,
                        "minimum": 0.0,
                        "maximum": 1.0
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of related entities to return (default: 20)",
                        "default": 20,
                        "minimum": 1,
                        "maximum": 100
                    }
                },
                "required": ["entity_text"]
            }
        ),
        Tool(
            name="find_related_entities",
            description="Discover entities related to a given entity (simplified version of get_entity_relationships). Returns top related entities for quick exploration and discovery.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity_text": {
                        "type": "string",
                        "description": "The entity to find related entities for"
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of related entities (default: 10)",
                        "default": 10,
                        "minimum": 1,
                        "maximum": 50
                    }
                },
                "required": ["entity_text"]
            }
        ),
        Tool(
            name="search_entity_pair",
            description="Find documents that contain both entities. Useful for finding documentation about specific combinations (e.g., 'VIC-II' AND 'raster interrupt'). Returns documents with both entity counts and context snippets.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity1": {
                        "type": "string",
                        "description": "First entity to search for"
                    },
                    "entity2": {
                        "type": "string",
                        "description": "Second entity to search for"
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum number of documents to return (default: 10)",
                        "default": 10,
                        "minimum": 1,
                        "maximum": 50
                    }
                },
                "required": ["entity1", "entity2"]
            }
        ),
        Tool(
            name="translate_query",
            description="Translate a natural language query into structured search parameters. Uses AI to extract entities, keywords, and determine optimal search strategy. Perfect for conversational queries like 'find info about sprites on VIC-II' or 'how does sound work?'. Returns structured parameters that can be used directly with other search tools.",
            inputSchema={
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Natural language query to translate (e.g., 'find sprite information on the VIC-II chip')"
                    },
                    "confidence_threshold": {
                        "type": "number",
                        "description": "Minimum confidence score for entity extraction (0.0-1.0, default: 0.7)",
                        "default": 0.7,
                        "minimum": 0.0,
                        "maximum": 1.0
                    }
                },
                "required": ["query"]
            }
        ),
        Tool(
            name="compare_documents",
            description="Compare two documents side-by-side with similarity scoring, metadata diff, content diff, and entity comparison. Perfect for finding differences between document versions, comparing related documents, or analyzing document similarity. Returns comprehensive comparison with cosine similarity score (0.0-1.0).",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id_1": {
                        "type": "string",
                        "description": "First document ID to compare"
                    },
                    "doc_id_2": {
                        "type": "string",
                        "description": "Second document ID to compare"
                    },
                    "comparison_type": {
                        "type": "string",
                        "description": "Type of comparison: 'full' (all), 'metadata' (metadata + entities), 'content' (metadata + similarity + diff)",
                        "enum": ["full", "metadata", "content"],
                        "default": "full"
                    }
                },
                "required": ["doc_id_1", "doc_id_2"]
            }
        ),
        Tool(
            name="export_entities",
            description="Export all extracted entities to CSV or JSON format. Includes entity text, type, confidence scores, document counts, and occurrence counts. Useful for data analysis, reporting, or importing into other tools.",
            inputSchema={
                "type": "object",
                "properties": {
                    "format": {
                        "type": "string",
                        "description": "Export format: 'csv' or 'json'",
                        "enum": ["csv", "json"],
                        "default": "csv"
                    },
                    "entity_types": {
                        "type": "array",
                        "description": "Filter by entity types (e.g., ['hardware', 'instruction']). Leave empty for all types.",
                        "items": {
                            "type": "string",
                            "enum": ["hardware", "memory_address", "instruction", "person", "company", "product", "concept"]
                        }
                    },
                    "min_confidence": {
                        "type": "number",
                        "description": "Minimum confidence threshold (0.0-1.0)",
                        "default": 0.0,
                        "minimum": 0.0,
                        "maximum": 1.0
                    },
                    "output_path": {
                        "type": "string",
                        "description": "Optional file path to save export (if not provided, returns as string)"
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="export_relationships",
            description="Export all entity relationships to CSV or JSON format. Includes entity pairs, types, relationship strength scores (0.0-1.0), and document counts. Perfect for network analysis, visualization, or data export.",
            inputSchema={
                "type": "object",
                "properties": {
                    "format": {
                        "type": "string",
                        "description": "Export format: 'csv' or 'json'",
                        "enum": ["csv", "json"],
                        "default": "csv"
                    },
                    "min_strength": {
                        "type": "number",
                        "description": "Minimum relationship strength (0.0-1.0)",
                        "default": 0.0,
                        "minimum": 0.0,
                        "maximum": 1.0
                    },
                    "entity_types": {
                        "type": "array",
                        "description": "Filter by entity types (applies to either entity in pair). Leave empty for all types.",
                        "items": {
                            "type": "string",
                            "enum": ["hardware", "memory_address", "instruction", "person", "company", "product", "concept"]
                        }
                    },
                    "output_path": {
                        "type": "string",
                        "description": "Optional file path to save export (if not provided, returns as string)"
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="queue_entity_extraction",
            description="Queue a document for background entity extraction. Extraction happens asynchronously without blocking. Use this to extract entities from documents without waiting for LLM processing. Returns job ID for tracking progress.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID to extract entities from"
                    },
                    "confidence_threshold": {
                        "type": "number",
                        "description": "Minimum confidence threshold (0.0-1.0, default: 0.6)",
                        "default": 0.6,
                        "minimum": 0.0,
                        "maximum": 1.0
                    },
                    "skip_if_exists": {
                        "type": "boolean",
                        "description": "Skip if entities already exist or job is queued (default: true)",
                        "default": True
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="get_extraction_status",
            description="Get the entity extraction status for a document. Shows whether entities exist, extraction job status (queued/running/completed/failed), timestamps, and error messages if any. Use this to check if extraction is complete before querying entities.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID to check extraction status for"
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="get_extraction_jobs",
            description="Get all entity extraction jobs with optional status filtering. Shows job queue, running jobs, completed extractions, and failed jobs. Useful for monitoring background extraction progress across all documents.",
            inputSchema={
                "type": "object",
                "properties": {
                    "status_filter": {
                        "type": "string",
                        "description": "Filter by status: 'queued', 'running', 'completed', or 'failed'. Leave empty for all jobs.",
                        "enum": ["queued", "running", "completed", "failed"]
                    },
                    "limit": {
                        "type": "integer",
                        "description": "Maximum number of jobs to return (default: 100)",
                        "default": 100,
                        "minimum": 1,
                        "maximum": 1000
                    }
                },
                "required": []
            }
        ),
        # ============================================================
        # Knowledge Graph Tools (v2.24.0 - Phase 1, Task 1.4)
        # ============================================================
        Tool(
            name="build_knowledge_graph",
            description="Build a knowledge graph from extracted entities and relationships. Returns graph statistics including node count, edge count, density, and connected components. The graph can be filtered by entity types, minimum occurrence counts, and relationship strength thresholds.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity_types": {
                        "type": "array",
                        "description": "Filter to specific entity types (e.g., ['hardware', 'instruction']). Leave empty for all types.",
                        "items": {
                            "type": "string",
                            "enum": ["hardware", "memory_address", "instruction", "person", "company", "product", "concept"]
                        }
                    },
                    "min_occurrences": {
                        "type": "integer",
                        "description": "Minimum entity occurrences to include in graph",
                        "default": 2,
                        "minimum": 1
                    },
                    "min_relationship_strength": {
                        "type": "number",
                        "description": "Minimum relationship strength threshold (0.0-1.0)",
                        "default": 0.3,
                        "minimum": 0.0,
                        "maximum": 1.0
                    },
                    "use_cache": {
                        "type": "boolean",
                        "description": "Use cached graph if available",
                        "default": True
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="compute_graph_metrics",
            description="Compute comprehensive graph analysis metrics including PageRank (importance), betweenness centrality (bridge nodes), degree centrality (connections), and community detection. Returns detailed metrics for all entities and graph-level statistics. Metrics are stored in the database for later retrieval.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity_types": {
                        "type": "array",
                        "description": "Filter graph to specific entity types",
                        "items": {
                            "type": "string",
                            "enum": ["hardware", "memory_address", "instruction", "person", "company", "product", "concept"]
                        }
                    },
                    "min_occurrences": {
                        "type": "integer",
                        "description": "Minimum entity occurrences for graph building",
                        "default": 2,
                        "minimum": 1
                    },
                    "min_relationship_strength": {
                        "type": "number",
                        "description": "Minimum relationship strength",
                        "default": 0.3,
                        "minimum": 0.0,
                        "maximum": 1.0
                    },
                    "store_results": {
                        "type": "boolean",
                        "description": "Store computed metrics to database",
                        "default": True
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="get_entity_metrics",
            description="Retrieve stored graph metrics for a specific entity. Returns PageRank score, betweenness centrality, degree centrality, community ID, entity type, and computation timestamp. Useful for understanding an entity's importance and role in the knowledge graph.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity_text": {
                        "type": "string",
                        "description": "Entity to retrieve metrics for (e.g., 'VIC-II', 'SID', 'Commodore 64')"
                    },
                    "metric_types": {
                        "type": "array",
                        "description": "Specific metrics to retrieve. Leave empty for all metrics.",
                        "items": {
                            "type": "string",
                            "enum": ["pagerank", "betweenness", "degree", "community"]
                        }
                    }
                },
                "required": ["entity_text"]
            }
        ),
        Tool(
            name="find_entity_path",
            description="Find the shortest path between two entities in the knowledge graph. Returns the complete path, path length, and relationship details for each connection. Useful for discovering how concepts are related and understanding knowledge connections.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity1": {
                        "type": "string",
                        "description": "Source entity (e.g., 'STA')"
                    },
                    "entity2": {
                        "type": "string",
                        "description": "Target entity (e.g., 'VIC-II')"
                    },
                    "max_path_length": {
                        "type": "integer",
                        "description": "Maximum path length to search",
                        "default": 6,
                        "minimum": 1,
                        "maximum": 10
                    },
                    "store_result": {
                        "type": "boolean",
                        "description": "Store computed path to database",
                        "default": True
                    }
                },
                "required": ["entity1", "entity2"]
            }
        ),
        Tool(
            name="get_entity_community",
            description="Get all entities in the same community as the specified entity. Communities are groups of closely related entities detected through graph analysis. Returns community ID, member count, and list of community members with their types.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity_text": {
                        "type": "string",
                        "description": "Entity to find community for (e.g., 'SID')"
                    },
                    "max_members": {
                        "type": "integer",
                        "description": "Maximum number of community members to return",
                        "default": 50,
                        "minimum": 1
                    }
                },
                "required": ["entity_text"]
            }
        ),
        Tool(
            name="get_top_entities",
            description="Get top-ranked entities by a specific metric (PageRank, betweenness, or degree centrality). Returns ranked list of entities with their scores, types, and other metrics. Useful for discovering the most important, central, or well-connected entities in the knowledge graph.",
            inputSchema={
                "type": "object",
                "properties": {
                    "metric": {
                        "type": "string",
                        "description": "Metric to rank by",
                        "enum": ["pagerank", "betweenness", "degree"],
                        "default": "pagerank"
                    },
                    "limit": {
                        "type": "integer",
                        "description": "Number of top entities to return",
                        "default": 10,
                        "minimum": 1,
                        "maximum": 100
                    },
                    "entity_types": {
                        "type": "array",
                        "description": "Filter to specific entity types",
                        "items": {
                            "type": "string",
                            "enum": ["hardware", "memory_address", "instruction", "person", "company", "product", "concept"]
                        }
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="visualize_graph",
            description="Generate interactive HTML visualization of the knowledge graph using PyVis. Creates a beautiful network diagram with customizable node colors (by entity type or community), node sizes (by PageRank/betweenness/degree), and interactive physics simulation. Perfect for exploring entity relationships and discovering patterns in the knowledge base.",
            inputSchema={
                "type": "object",
                "properties": {
                    "output_path": {
                        "type": "string",
                        "description": "Output file path for HTML visualization (relative to data directory or absolute path)",
                        "default": "knowledge_graph.html"
                    },
                    "entity_types": {
                        "type": "array",
                        "description": "Filter to specific entity types",
                        "items": {
                            "type": "string",
                            "enum": ["hardware", "memory_address", "instruction", "person", "company", "product", "concept"]
                        }
                    },
                    "min_occurrences": {
                        "type": "integer",
                        "description": "Minimum entity occurrences to include",
                        "default": 2,
                        "minimum": 1
                    },
                    "min_relationship_strength": {
                        "type": "number",
                        "description": "Minimum relationship strength",
                        "default": 0.3,
                        "minimum": 0.0,
                        "maximum": 1.0
                    },
                    "color_by": {
                        "type": "string",
                        "description": "Node coloring scheme",
                        "enum": ["entity_type", "community", "uniform"],
                        "default": "entity_type"
                    },
                    "size_by": {
                        "type": "string",
                        "description": "Node sizing metric",
                        "enum": ["pagerank", "degree", "betweenness", "uniform"],
                        "default": "pagerank"
                    },
                    "physics_enabled": {
                        "type": "boolean",
                        "description": "Enable physics simulation for dynamic layout",
                        "default": True
                    },
                    "height": {
                        "type": "string",
                        "description": "Visualization height (CSS format)",
                        "default": "800px"
                    },
                    "width": {
                        "type": "string",
                        "description": "Visualization width (CSS format)",
                        "default": "100%"
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="train_lda_topics",
            description="Train LDA topic model on all documents to discover latent topics. Uses Latent Dirichlet Allocation to find topics based on word co-occurrence patterns.",
            inputSchema={
                "type": "object",
                "properties": {
                    "num_topics": {
                        "type": "integer",
                        "description": "Number of topics to discover",
                        "default": 10,
                        "minimum": 2
                    },
                    "max_iter": {
                        "type": "integer",
                        "description": "Maximum training iterations",
                        "default": 100
                    },
                    "max_features": {
                        "type": "integer",
                        "description": "Maximum vocabulary size",
                        "default": 1000
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="train_nmf_topics",
            description="Train NMF topic model on all documents. Non-negative Matrix Factorization often produces more coherent topics than LDA and is faster.",
            inputSchema={
                "type": "object",
                "properties": {
                    "num_topics": {
                        "type": "integer",
                        "description": "Number of topics to discover",
                        "default": 10,
                        "minimum": 2
                    },
                    "max_iter": {
                        "type": "integer",
                        "description": "Maximum training iterations",
                        "default": 200
                    },
                    "max_features": {
                        "type": "integer",
                        "description": "Maximum vocabulary size",
                        "default": 1000
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="train_bertopic",
            description="Train BERTopic model using embeddings + UMAP + HDBSCAN. State-of-the-art transformer-based topic modeling that automatically discovers topics.",
            inputSchema={
                "type": "object",
                "properties": {
                    "num_topics": {
                        "type": "integer",
                        "description": "Target number of topics (actual may vary)",
                        "default": 10,
                        "minimum": 2
                    },
                    "min_topic_size": {
                        "type": "integer",
                        "description": "Minimum documents per topic",
                        "default": 5,
                        "minimum": 2
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="get_document_topics",
            description="Get topic assignments for a specific document or all documents. Shows which topics each document belongs to with probability scores.",
            inputSchema={
                "type": "object",
                "properties": {
                    "model_type": {
                        "type": "string",
                        "description": "Topic model type",
                        "enum": ["lda", "nmf", "bertopic"]
                    },
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID (optional, omit to get all documents)"
                    },
                    "min_probability": {
                        "type": "number",
                        "description": "Minimum topic probability to include",
                        "default": 0.1,
                        "minimum": 0.0,
                        "maximum": 1.0
                    }
                },
                "required": ["model_type"]
            }
        ),
        Tool(
            name="cluster_documents_kmeans",
            description="Cluster documents using K-Means algorithm on embeddings. Partitions documents into K clusters based on similarity.",
            inputSchema={
                "type": "object",
                "properties": {
                    "num_clusters": {
                        "type": "integer",
                        "description": "Number of clusters (K)",
                        "default": 10,
                        "minimum": 2
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="cluster_documents_dbscan",
            description="Cluster documents using DBSCAN (density-based clustering). Automatically discovers clusters and identifies outliers without needing to specify number of clusters.",
            inputSchema={
                "type": "object",
                "properties": {
                    "eps": {
                        "type": "number",
                        "description": "Maximum distance between neighbors",
                        "default": 0.5,
                        "minimum": 0.0
                    },
                    "min_samples": {
                        "type": "integer",
                        "description": "Minimum samples in neighborhood",
                        "default": 3,
                        "minimum": 2
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="cluster_documents_hdbscan",
            description="Cluster documents using HDBSCAN (hierarchical density-based clustering). Improved version of DBSCAN with better handling of varying density clusters.",
            inputSchema={
                "type": "object",
                "properties": {
                    "min_cluster_size": {
                        "type": "integer",
                        "description": "Minimum cluster size",
                        "default": 5,
                        "minimum": 2
                    },
                    "min_samples": {
                        "type": "integer",
                        "description": "Minimum samples in neighborhood",
                        "default": 3,
                        "minimum": 1
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="get_cluster_documents",
            description="Get documents in a specific cluster or all clusters. Shows cluster membership, distances, and cluster characteristics.",
            inputSchema={
                "type": "object",
                "properties": {
                    "algorithm": {
                        "type": "string",
                        "description": "Clustering algorithm",
                        "enum": ["kmeans", "dbscan", "hdbscan"]
                    },
                    "cluster_number": {
                        "type": "integer",
                        "description": "Cluster number (optional, omit to get all clusters)"
                    }
                },
                "required": ["algorithm"]
            }
        ),
        # Phase 2: Visualization Tools
        Tool(
            name="generate_topic_wordcloud",
            description="Generate word cloud visualization for a topic. Creates an image file showing the most important words in a topic with size proportional to their weights.",
            inputSchema={
                "type": "object",
                "properties": {
                    "topic_id": {
                        "type": "string",
                        "description": "Topic ID to visualize"
                    },
                    "output_path": {
                        "type": "string",
                        "description": "Path to save the word cloud image (PNG)"
                    },
                    "width": {
                        "type": "integer",
                        "description": "Image width in pixels (default: 800)",
                        "default": 800
                    },
                    "height": {
                        "type": "integer",
                        "description": "Image height in pixels (default: 400)",
                        "default": 400
                    },
                    "background_color": {
                        "type": "string",
                        "description": "Background color (default: 'white')",
                        "default": "white"
                    }
                },
                "required": ["topic_id", "output_path"]
            }
        ),
        Tool(
            name="visualize_cluster_scatter",
            description="Generate 2D scatter plot of document clusters using UMAP projection. Shows how documents are distributed across clusters in 2D space.",
            inputSchema={
                "type": "object",
                "properties": {
                    "algorithm": {
                        "type": "string",
                        "description": "Clustering algorithm",
                        "enum": ["kmeans", "dbscan", "hdbscan"]
                    },
                    "output_path": {
                        "type": "string",
                        "description": "Path to save the scatter plot image (PNG)"
                    },
                    "width": {
                        "type": "integer",
                        "description": "Image width in pixels (default: 1200)",
                        "default": 1200
                    },
                    "height": {
                        "type": "integer",
                        "description": "Image height in pixels (default: 800)",
                        "default": 800
                    },
                    "n_neighbors": {
                        "type": "integer",
                        "description": "UMAP n_neighbors parameter (default: 15)",
                        "default": 15
                    },
                    "min_dist": {
                        "type": "number",
                        "description": "UMAP min_dist parameter (default: 0.1)",
                        "default": 0.1
                    }
                },
                "required": ["algorithm", "output_path"]
            }
        ),
        Tool(
            name="generate_topic_heatmap",
            description="Generate heatmap showing document-topic probability matrix. Visualizes which topics are most prominent in which documents.",
            inputSchema={
                "type": "object",
                "properties": {
                    "model_type": {
                        "type": "string",
                        "description": "Topic model type",
                        "enum": ["lda", "nmf", "bertopic"]
                    },
                    "output_path": {
                        "type": "string",
                        "description": "Path to save the heatmap image (PNG)"
                    },
                    "max_topics": {
                        "type": "integer",
                        "description": "Maximum number of topics to include (default: 20)",
                        "default": 20
                    },
                    "max_documents": {
                        "type": "integer",
                        "description": "Maximum number of documents to include (default: 50)",
                        "default": 50
                    }
                },
                "required": ["model_type", "output_path"]
            }
        ),
        Tool(
            name="visualize_cluster_distribution",
            description="Generate bar chart showing cluster size distribution. Shows how many documents are in each cluster.",
            inputSchema={
                "type": "object",
                "properties": {
                    "algorithm": {
                        "type": "string",
                        "description": "Clustering algorithm",
                        "enum": ["kmeans", "dbscan", "hdbscan"]
                    },
                    "output_path": {
                        "type": "string",
                        "description": "Path to save the bar chart image (PNG)"
                    },
                    "width": {
                        "type": "integer",
                        "description": "Image width in pixels (default: 1000)",
                        "default": 1000
                    },
                    "height": {
                        "type": "integer",
                        "description": "Image height in pixels (default: 600)",
                        "default": 600
                    }
                },
                "required": ["algorithm", "output_path"]
            }
        ),
        # Phase 3: Temporal Analysis Tools
        Tool(
            name="extract_document_events",
            description="Extract temporal events from a document (product releases, company milestones, technical innovations, cultural events). Detects event patterns and dates, then stores to database.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID to extract events from"
                    },
                    "min_confidence": {
                        "type": "number",
                        "description": "Minimum confidence threshold (0.0-1.0)",
                        "default": 0.5,
                        "minimum": 0.0,
                        "maximum": 1.0
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="get_timeline",
            description="Get chronological timeline of events with optional filtering. Returns timeline entries sorted by date with event details.",
            inputSchema={
                "type": "object",
                "properties": {
                    "start_year": {
                        "type": "integer",
                        "description": "Filter events from this year onwards (inclusive)"
                    },
                    "end_year": {
                        "type": "integer",
                        "description": "Filter events up to this year (inclusive)"
                    },
                    "category": {
                        "type": "string",
                        "description": "Filter by category (e.g., '1980s-release', '1970s-innovation')"
                    },
                    "min_importance": {
                        "type": "integer",
                        "description": "Minimum importance level (1-5, where 5 is highest)",
                        "default": 1,
                        "minimum": 1,
                        "maximum": 5
                    },
                    "limit": {
                        "type": "integer",
                        "description": "Maximum number of entries to return"
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="search_events_by_date",
            description="Search for events within a date range. Filter by year range, event type (release, milestone, innovation, cultural, update), and confidence.",
            inputSchema={
                "type": "object",
                "properties": {
                    "start_year": {
                        "type": "integer",
                        "description": "Start year (inclusive)"
                    },
                    "end_year": {
                        "type": "integer",
                        "description": "End year (inclusive)"
                    },
                    "event_type": {
                        "type": "string",
                        "description": "Event type filter",
                        "enum": ["release", "milestone", "innovation", "cultural", "update"]
                    },
                    "min_confidence": {
                        "type": "number",
                        "description": "Minimum confidence threshold",
                        "default": 0.5,
                        "minimum": 0.0,
                        "maximum": 1.0
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="get_historical_context",
            description="Get historical context for a specific year. Returns events from the target year plus surrounding years to provide temporal context.",
            inputSchema={
                "type": "object",
                "properties": {
                    "year": {
                        "type": "integer",
                        "description": "Target year to get context for"
                    },
                    "context_years": {
                        "type": "integer",
                        "description": "Number of years before/after to include",
                        "default": 2,
                        "minimum": 0,
                        "maximum": 10
                    }
                },
                "required": ["year"]
            }
        ),
        # Phase 1: Knowledge Graph Tools (v2.24.0)
        Tool(
            name="build_knowledge_graph",
            description="Build a knowledge graph from entities and relationships in the C64 knowledge base. The graph represents entities as nodes and their relationships as weighted edges. Use this to understand the structure of knowledge and find connections between concepts.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity_types": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter to specific entity types (e.g., ['person', 'product', 'hardware']). Omit for all types."
                    },
                    "min_occurrences": {
                        "type": "integer",
                        "description": "Minimum entity occurrences to include (default: 2)",
                        "default": 2,
                        "minimum": 1
                    },
                    "min_relationship_strength": {
                        "type": "number",
                        "description": "Minimum relationship strength 0.0-1.0 (default: 0.3)",
                        "default": 0.3,
                        "minimum": 0.0,
                        "maximum": 1.0
                    }
                }
            }
        ),
        Tool(
            name="analyze_graph_pagerank",
            description="Calculate PageRank scores for entities in the knowledge graph. PageRank identifies the most 'important' or 'central' entities based on their connections. Higher scores indicate entities that are more connected and influential in the knowledge network.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity_types": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter graph to specific entity types (optional)"
                    },
                    "min_occurrences": {
                        "type": "integer",
                        "description": "Minimum entity occurrences (default: 2)",
                        "default": 2
                    },
                    "top_n": {
                        "type": "integer",
                        "description": "Number of top entities to return (default: 20)",
                        "default": 20,
                        "minimum": 1,
                        "maximum": 100
                    },
                    "alpha": {
                        "type": "number",
                        "description": "Damping parameter for PageRank (default: 0.85)",
                        "default": 0.85,
                        "minimum": 0.0,
                        "maximum": 1.0
                    }
                }
            }
        ),
        Tool(
            name="detect_graph_communities",
            description="Detect communities (clusters) in the knowledge graph. Communities are groups of entities that are more densely connected to each other than to the rest of the graph. This helps identify topic clusters and thematic groupings.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity_types": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter graph to specific entity types (optional)"
                    },
                    "min_occurrences": {
                        "type": "integer",
                        "description": "Minimum entity occurrences (default: 2)",
                        "default": 2
                    },
                    "algorithm": {
                        "type": "string",
                        "description": "Detection algorithm: 'louvain' (best for large graphs), 'label_propagation' (fast), or 'greedy_modularity'",
                        "default": "louvain",
                        "enum": ["louvain", "label_propagation", "greedy_modularity"]
                    }
                }
            }
        ),
        Tool(
            name="calculate_graph_centrality",
            description="Calculate centrality measures for entities in the knowledge graph. Returns betweenness, closeness, and degree centrality. These measures identify entities that bridge different parts of the graph, are close to all others, or have many connections.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity_types": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter graph to specific entity types (optional)"
                    },
                    "min_occurrences": {
                        "type": "integer",
                        "description": "Minimum entity occurrences (default: 2)",
                        "default": 2
                    },
                    "top_n": {
                        "type": "integer",
                        "description": "Number of top entities per measure (default: 10)",
                        "default": 10,
                        "minimum": 1,
                        "maximum": 50
                    }
                }
            }
        ),
        Tool(
            name="find_entity_path",
            description="Find the shortest path between two entities in the knowledge graph. Shows how entities are connected through intermediate relationships. Useful for understanding conceptual connections and knowledge pathways.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity1": {
                        "type": "string",
                        "description": "Source entity name"
                    },
                    "entity2": {
                        "type": "string",
                        "description": "Target entity name"
                    },
                    "min_occurrences": {
                        "type": "integer",
                        "description": "Minimum entity occurrences for graph (default: 2)",
                        "default": 2
                    }
                },
                "required": ["entity1", "entity2"]
            }
        ),
        Tool(
            name="get_graph_statistics",
            description="Get statistical overview of the knowledge graph including node count, edge count, density, connected components, and degree distribution. Provides insight into the overall structure and complexity of the knowledge network.",
            inputSchema={
                "type": "object",
                "properties": {
                    "entity_types": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter graph to specific entity types (optional)"
                    },
                    "min_occurrences": {
                        "type": "integer",
                        "description": "Minimum entity occurrences (default: 2)",
                        "default": 2
                    }
                }
            }
        ),
        # ========================================
        # Phase 2: Topic Modeling & Clustering
        # ========================================
        Tool(
            name="train_lda_topics",
            description="Train LDA (Latent Dirichlet Allocation) topic model on documents. Discovers latent topics using probabilistic modeling. Returns topics with top words and document assignments.",
            inputSchema={
                "type": "object",
                "properties": {
                    "num_topics": {
                        "type": "integer",
                        "description": "Number of topics to discover (default: 10)",
                        "default": 10
                    },
                    "max_iter": {
                        "type": "integer",
                        "description": "Maximum iterations (default: 100)",
                        "default": 100
                    },
                    "random_state": {
                        "type": "integer",
                        "description": "Random seed for reproducibility (default: 42)",
                        "default": 42
                    }
                }
            }
        ),
        Tool(
            name="train_nmf_topics",
            description="Train NMF (Non-negative Matrix Factorization) topic model on documents. Often produces more coherent topics than LDA using matrix factorization. Returns topics with top words and document assignments.",
            inputSchema={
                "type": "object",
                "properties": {
                    "num_topics": {
                        "type": "integer",
                        "description": "Number of topics to discover (default: 10)",
                        "default": 10
                    },
                    "max_iter": {
                        "type": "integer",
                        "description": "Maximum iterations (default: 200)",
                        "default": 200
                    },
                    "random_state": {
                        "type": "integer",
                        "description": "Random seed for reproducibility (default: 42)",
                        "default": 42
                    }
                }
            }
        ),
        Tool(
            name="train_bertopic",
            description="Train BERTopic model using document embeddings. State-of-the-art topic modeling with UMAP + HDBSCAN clustering. Automatically discovers topics from semantic embeddings.",
            inputSchema={
                "type": "object",
                "properties": {
                    "num_topics": {
                        "type": "integer",
                        "description": "Target number of topics (default: 10)",
                        "default": 10
                    },
                    "min_cluster_size": {
                        "type": "integer",
                        "description": "Minimum documents per topic (default: 5)",
                        "default": 5
                    }
                }
            }
        ),
        Tool(
            name="get_document_topics",
            description="Get topics assigned to a specific document, including probabilities and top words for each topic. Shows which topics the document belongs to.",
            inputSchema={
                "type": "object",
                "properties": {
                    "doc_id": {
                        "type": "string",
                        "description": "Document ID"
                    },
                    "model_type": {
                        "type": "string",
                        "description": "Topic model type: 'lda', 'nmf', or 'bertopic' (optional)",
                        "enum": ["lda", "nmf", "bertopic"]
                    }
                },
                "required": ["doc_id"]
            }
        ),
        Tool(
            name="cluster_documents_kmeans",
            description="Cluster documents using K-Means algorithm on embeddings. Partitions documents into K clusters. Returns cluster assignments and silhouette score.",
            inputSchema={
                "type": "object",
                "properties": {
                    "num_clusters": {
                        "type": "integer",
                        "description": "Number of clusters (default: 10)",
                        "default": 10
                    },
                    "random_state": {
                        "type": "integer",
                        "description": "Random seed (default: 42)",
                        "default": 42
                    }
                }
            }
        ),
        Tool(
            name="cluster_documents_dbscan",
            description="Cluster documents using DBSCAN (density-based) algorithm. Automatically discovers clusters and identifies outliers. Does not require specifying number of clusters.",
            inputSchema={
                "type": "object",
                "properties": {
                    "eps": {
                        "type": "number",
                        "description": "Maximum distance between samples (default: 0.5)",
                        "default": 0.5
                    },
                    "min_samples": {
                        "type": "integer",
                        "description": "Minimum samples in neighborhood (default: 5)",
                        "default": 5
                    }
                }
            }
        ),
        Tool(
            name="cluster_documents_hdbscan",
            description="Cluster documents using HDBSCAN (hierarchical density-based) algorithm. Advanced clustering that handles varying densities. Automatically discovers clusters and outliers.",
            inputSchema={
                "type": "object",
                "properties": {
                    "min_cluster_size": {
                        "type": "integer",
                        "description": "Minimum samples per cluster (default: 5)",
                        "default": 5
                    },
                    "min_samples": {
                        "type": "integer",
                        "description": "Minimum samples in neighborhood (optional)"
                    }
                }
            }
        ),
        Tool(
            name="get_cluster_documents",
            description="Get all documents in a specific cluster, including distances from centroid. Shows which documents are grouped together.",
            inputSchema={
                "type": "object",
                "properties": {
                    "cluster_id": {
                        "type": "string",
                        "description": "Cluster ID"
                    },
                    "algorithm": {
                        "type": "string",
                        "description": "Clustering algorithm: 'kmeans', 'dbscan', or 'hdbscan' (optional)",
                        "enum": ["kmeans", "dbscan", "hdbscan"]
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Maximum documents to return (default: 50)",
                        "default": 50
                    }
                },
                "required": ["cluster_id"]
            }
        )
    ]


@server.call_tool()
async def call_tool(name: str, arguments: dict) -> list[TextContent]:
    """Handle tool calls."""
    
    if name == "search_docs":
        query = arguments.get("query", "")
        max_results = arguments.get("max_results", 5)
        tags = arguments.get("tags")
        
        results = kb.search(query, max_results, tags)
        
        if not results:
            return [TextContent(type="text", text=f"No results found for: {query}")]
        
        output = f"Found {len(results)} results for '{query}':\n\n"
        for i, r in enumerate(results, 1):
            output += f"--- Result {i} ---\n"
            output += f"Document: {r['title']} ({r['filename']})\n"
            output += f"Doc ID: {r['doc_id']}, Chunk: {r['chunk_id']}\n"
            output += f"Score: {r['score']}\n"
            output += f"Snippet:\n{r['snippet']}\n\n"
        
        return [TextContent(type="text", text=output)]

    elif name == "translate_query":
        query = arguments.get("query", "")
        confidence_threshold = arguments.get("confidence_threshold", 0.7)

        try:
            result = kb.translate_nl_query(query, confidence_threshold)
        except Exception as e:
            return [TextContent(type="text", text=f"Query translation error: {str(e)}")]

        # Format the translation result
        output = "Natural Language Query Translation\n"
        output += f"{'='*60}\n\n"
        output += f"Original Query: \"{result['original_query']}\"\n\n"

        output += f"Search Mode: {result['search_mode']} (confidence: {result['confidence']:.2f})\n"
        output += f"Reasoning: {result['reasoning']}\n\n"

        if result['search_terms']:
            output += "Search Terms:\n"
            for term in result['search_terms']:
                output += f"  - {term}\n"
            output += "\n"

        if result['facet_filters']:
            output += "Facet Filters:\n"
            for facet_type, values in result['facet_filters'].items():
                output += f"  {facet_type}: {', '.join(values)}\n"
            output += "\n"

        if result['entities_found']:
            output += f"Entities Detected ({len(result['entities_found'])}):\n"
            for entity in result['entities_found']:
                output += f"  - {entity['text']} ({entity['type']}, confidence: {entity['confidence']:.2f})\n"
            output += "\n"

        # Add suggested next steps
        output += "Suggested Action:\n"
        if result['search_mode'] == 'keyword':
            output += f"  Use search_docs with terms: {', '.join(result['search_terms'][:3])}\n"
        elif result['search_mode'] == 'semantic':
            output += f"  Use semantic_search with query: \"{result['original_query']}\"\n"
        else:  # hybrid
            output += "  Use hybrid_search for best results\n"

        if result['facet_filters']:
            output += f"  Apply facet filters: {result['facet_filters']}\n"

        return [TextContent(type="text", text=output)]

    elif name == "semantic_search":
        if not kb.use_semantic:
            return [TextContent(
                type="text",
                text="Semantic search is not enabled. Set USE_SEMANTIC_SEARCH=1 and install sentence-transformers and faiss-cpu."
            )]

        query = arguments.get("query", "")
        max_results = arguments.get("max_results", 5)
        tags = arguments.get("tags")

        try:
            results = kb.semantic_search(query, max_results, tags)
        except Exception as e:
            return [TextContent(type="text", text=f"Semantic search error: {str(e)}")]

        if not results:
            return [TextContent(type="text", text=f"No results found for: {query}")]

        output = f"Found {len(results)} semantic results for '{query}':\n\n"
        for i, r in enumerate(results, 1):
            output += f"--- Result {i} ---\n"
            output += f"Document: {r['title']} ({r['filename']})\n"
            output += f"Doc ID: {r['doc_id']}, Chunk: {r['chunk_id']}\n"
            output += f"Similarity Score: {r['similarity']:.4f}\n"
            output += f"Snippet:\n{r['snippet']}\n\n"

        return [TextContent(type="text", text=output)]

    elif name == "hybrid_search":
        if not kb.use_semantic:
            return [TextContent(
                type="text",
                text="Hybrid search requires semantic search. Set USE_SEMANTIC_SEARCH=1 and install sentence-transformers and faiss-cpu."
            )]

        query = arguments.get("query", "")
        max_results = arguments.get("max_results", 5)
        tags = arguments.get("tags")
        semantic_weight = arguments.get("semantic_weight", 0.3)

        try:
            results = kb.hybrid_search(query, max_results, tags, semantic_weight)
        except Exception as e:
            return [TextContent(type="text", text=f"Hybrid search error: {str(e)}")]

        if not results:
            return [TextContent(type="text", text=f"No results found for: {query}")]

        output = f"Found {len(results)} hybrid results for '{query}' (semantic_weight={semantic_weight}):\n\n"
        for i, r in enumerate(results, 1):
            output += f"--- Result {i} ---\n"
            output += f"Document: {r['title']} ({r['filename']})\n"
            output += f"Doc ID: {r['doc_id']}, Chunk: {r['chunk_id']}\n"
            output += f"Hybrid Score: {r['score']:.4f} (FTS: {r['fts_score']:.4f}, Semantic: {r['semantic_score']:.4f})\n"
            output += f"Snippet:\n{r['snippet']}\n\n"

        return [TextContent(type="text", text=output)]

    elif name == "fuzzy_search":
        query = arguments.get("query", "")
        max_results = arguments.get("max_results", 5)
        tags = arguments.get("tags")
        similarity_threshold = arguments.get("similarity_threshold", 80)

        try:
            results = kb.fuzzy_search(query, max_results, tags, similarity_threshold)
        except Exception as e:
            return [TextContent(type="text", text=f"Fuzzy search error: {str(e)}")]

        if not results:
            return [TextContent(type="text", text=f"No results found for: {query}")]

        output = f"Found {len(results)} fuzzy search results for '{query}':\n\n"

        # Show corrections if any
        if results and 'fuzzy_corrections' in results[0]:
            corrections = results[0]['fuzzy_corrections']
            output += "Corrections applied:\n"
            for corr in corrections:
                output += f"  '{corr['original']}' → '{corr['corrected']}' (similarity: {corr['similarity']}%)\n"
            output += f"Corrected query: '{results[0]['corrected_query']}'\n\n"

        for i, r in enumerate(results, 1):
            output += f"--- Result {i} ---\n"
            output += f"Document: {r['title']} ({r['filename']})\n"
            output += f"Doc ID: {r['doc_id']}, Chunk: {r['chunk_id']}\n"
            output += f"Score: {r['score']}\n"
            output += f"Snippet:\n{r['snippet']}\n\n"

        return [TextContent(type="text", text=output)]

    elif name == "search_within_results":
        previous_results = arguments.get("previous_results", [])
        refinement_query = arguments.get("refinement_query", "")
        max_results = arguments.get("max_results", 5)

        try:
            results = kb.search_within_results(previous_results, refinement_query, max_results)
        except Exception as e:
            return [TextContent(type="text", text=f"Search within results error: {str(e)}")]

        if not results:
            return [TextContent(type="text", text=f"No matching results in previous set for refinement query: {refinement_query}")]

        output = f"Refined search: found {len(results)} results for '{refinement_query}' within previous results:\n\n"

        for i, r in enumerate(results, 1):
            output += f"--- Result {i} ---\n"
            output += f"Document: {r['title']} ({r['filename']})\n"
            output += f"Doc ID: {r['doc_id']}, Chunk: {r['chunk_id']}\n"
            output += f"Original Score: {r.get('score', 'N/A')}\n"
            output += f"Refinement Score: {r.get('refinement_score', 'N/A')}\n"
            output += f"Snippet:\n{r['snippet']}\n\n"

        return [TextContent(type="text", text=output)]

    elif name == "suggest_tags":
        doc_id = arguments.get("doc_id", "")
        confidence_threshold = arguments.get("confidence_threshold", 0.6)

        try:
            suggestions = kb.suggest_tags(doc_id, confidence_threshold)
        except Exception as e:
            return [TextContent(type="text", text=f"Tag suggestion error: {str(e)}")]

        if not suggestions:
            return [TextContent(type="text", text=f"No tag suggestions found for document {doc_id}. Content may not match known categories.")]

        # Organize by category
        by_category = {}
        for suggestion in suggestions:
            category = suggestion['category']
            if category not in by_category:
                by_category[category] = []
            by_category[category].append(suggestion)

        output = f"Tag suggestions for document {doc_id}:\n\n"

        for category in sorted(by_category.keys()):
            tags_in_cat = by_category[category]
            output += f"**{category.replace('-', ' ').title()}:**\n"
            for tag in tags_in_cat:
                output += f"  - {tag['tag']} (confidence: {tag['confidence']:.0%})\n"
            output += "\n"

        output += "\n**To apply these tags:**\n"
        output += "Use the 'add_document' tool to update the document with tags.\n"
        output += "Or use update_document_tags tool to modify tags directly.\n"

        return [TextContent(type="text", text=output)]

    elif name == "get_tags_by_category":
        try:
            tag_categories = kb.get_tags_by_category()
        except Exception as e:
            return [TextContent(type="text", text=f"Tag browsing error: {str(e)}")]

        output = "Knowledge Base Tags by Category\n"
        output += "=" * 60 + "\n\n"

        total_tags = 0

        for category in sorted(tag_categories.keys()):
            tags = tag_categories[category]
            if not tags:
                continue

            output += f"**{category.replace('-', ' ').title()}** ({len(tags)} tags):\n"

            for tag_info in tags:
                tag = tag_info['tag']
                count = tag_info['count']
                output += f"  - {tag}: {count} document{'s' if count != 1 else ''}\n"
                total_tags += 1

            output += "\n"

        output += f"\nTotal: {total_tags} tags across {len(tag_categories)} categories\n"

        return [TextContent(type="text", text=output)]

    elif name == "faceted_search":
        query = arguments.get("query", "")
        facet_filters = arguments.get("facet_filters")
        max_results = arguments.get("max_results", 5)
        tags = arguments.get("tags")

        try:
            results = kb.faceted_search(query, facet_filters, max_results, tags)
        except Exception as e:
            return [TextContent(type="text", text=f"Faceted search error: {str(e)}")]

        if not results:
            if facet_filters:
                return [TextContent(type="text", text=f"No results found for '{query}' with facet filters: {facet_filters}")]
            else:
                return [TextContent(type="text", text=f"No results found for: {query}")]

        # Format output
        filter_desc = f" with facets: {facet_filters}" if facet_filters else ""
        output = f"Found {len(results)} results for '{query}'{filter_desc}:\n\n"

        for i, r in enumerate(results, 1):
            output += f"--- Result {i} ---\n"
            output += f"Document: {r['title']} ({r['filename']})\n"
            output += f"Doc ID: {r['doc_id']}, Chunk: {r['chunk_id']}\n"
            output += f"Score: {r['score']:.4f}\n"

            # Show document facets
            if 'facets' in r:
                facets = r['facets']
                facet_parts = []
                if facets.get('hardware'):
                    facet_parts.append(f"Hardware: {', '.join(sorted(facets['hardware']))}")
                if facets.get('instruction'):
                    facet_parts.append(f"Instructions: {', '.join(sorted(facets['instruction']))}")
                if facets.get('register'):
                    # Show only first 5 registers if many
                    regs = sorted(facets['register'])
                    if len(regs) > 5:
                        facet_parts.append(f"Registers: {', '.join(regs[:5])} (+{len(regs)-5} more)")
                    else:
                        facet_parts.append(f"Registers: {', '.join(regs)}")
                if facet_parts:
                    output += f"Facets: {' | '.join(facet_parts)}\n"

            output += f"Snippet:\n{r['snippet']}\n\n"

        return [TextContent(type="text", text=output)]

    elif name == "get_chunk":
        doc_id = arguments.get("doc_id")
        chunk_id = arguments.get("chunk_id")
        
        chunk = kb.get_chunk(doc_id, chunk_id)
        if not chunk:
            return [TextContent(type="text", text=f"Chunk not found: {doc_id}/{chunk_id}")]
        
        output = f"Document: {chunk.title}\n"
        output += f"Chunk {chunk.chunk_id} ({chunk.word_count} words):\n\n"
        output += chunk.content
        
        return [TextContent(type="text", text=output)]
    
    elif name == "get_document":
        doc_id = arguments.get("doc_id")
        
        content = kb.get_document_content(doc_id)
        if not content:
            return [TextContent(type="text", text=f"Document not found: {doc_id}")]
        
        doc = kb.documents.get(doc_id)
        output = f"Document: {doc.title}\n"
        output += f"File: {doc.filename}\n"
        output += f"{'='*50}\n\n"
        output += content
        
        return [TextContent(type="text", text=output)]
    
    elif name == "list_docs":
        docs = kb.list_documents()
        
        if not docs:
            return [TextContent(type="text", text="No documents in knowledge base. Use add_document to add PDFs or text files.")]
        
        output = f"Documents in knowledge base ({len(docs)}):\n\n"
        for doc in docs:
            output += f"- {doc.title}\n"
            output += f"  ID: {doc.doc_id}\n"
            output += f"  File: {doc.filename} ({doc.file_type})\n"
            if doc.total_pages:
                output += f"  Pages: {doc.total_pages}\n"
            output += f"  Chunks: {doc.total_chunks}\n"
            if doc.tags:
                output += f"  Tags: {', '.join(doc.tags)}\n"
            output += f"  Indexed: {doc.indexed_at}\n\n"
        
        return [TextContent(type="text", text=output)]
    
    elif name == "add_document":
        filepath = arguments.get("filepath")
        title = arguments.get("title")
        tags = arguments.get("tags", [])
        
        try:
            doc = kb.add_document(filepath, title, tags)
            output = "Successfully added document:\n"
            output += f"  Title: {doc.title}\n"
            output += f"  ID: {doc.doc_id}\n"
            output += f"  Type: {doc.file_type}\n"
            output += f"  Chunks: {doc.total_chunks}\n"
            if doc.total_pages:
                output += f"  Pages: {doc.total_pages}\n"
            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error adding document: {str(e)}")]

    elif name == "scrape_url":
        url = arguments.get("url")
        title = arguments.get("title")
        tags = arguments.get("tags", [])
        follow_links = arguments.get("follow_links", True)
        same_domain_only = arguments.get("same_domain_only", True)
        max_pages = arguments.get("max_pages", 50)
        depth = arguments.get("depth", 3)
        limit = arguments.get("limit")
        threads = arguments.get("threads", 10)
        delay = arguments.get("delay", 100)
        selector = arguments.get("selector")

        try:
            result = kb.scrape_url(
                url=url,
                title=title,
                tags=tags,
                follow_links=follow_links,
                same_domain_only=same_domain_only,
                max_pages=max_pages,
                depth=depth,
                limit=limit,
                threads=threads,
                delay=delay,
                selector=selector
            )

            output = "Scraping Result:\n\n"
            output += f"Status: {result['status']}\n"
            output += f"URL: {result['url']}\n"
            output += f"Files scraped: {result['files_scraped']}\n"
            output += f"Documents added: {result['docs_added']}\n"

            if result['docs_failed'] > 0:
                output += f"Documents failed: {result['docs_failed']}\n"

            if result.get('error'):
                output += f"\nError: {result['error']}\n"

            if result.get('doc_ids'):
                output += "\nAdded document IDs:\n"
                for doc_id in result['doc_ids'][:10]:  # Show first 10
                    doc = kb.documents.get(doc_id)
                    if doc:
                        output += f"  - {doc.title} ({doc_id})\n"
                if len(result['doc_ids']) > 10:
                    output += f"  ... and {len(result['doc_ids']) - 10} more\n"

            if result['status'] == 'success':
                output += f"\nOutput directory: {result['output_dir']}\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error scraping URL: {str(e)}")]

    elif name == "rescrape_document":
        doc_id = arguments.get("doc_id")

        try:
            result = kb.rescrape_document(doc_id)

            output = "Re-scrape Result:\n\n"
            output += f"Status: {result['status']}\n"
            output += f"Original doc ID: {result['old_doc_id']}\n"
            output += f"Documents added: {result['docs_added']}\n"

            if result.get('error'):
                output += f"Error: {result['error']}\n"

            if result.get('doc_ids'):
                output += "\nNew document IDs:\n"
                for doc_id in result['doc_ids'][:5]:
                    doc = kb.documents.get(doc_id)
                    if doc:
                        output += f"  - {doc.title} ({doc_id})\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error re-scraping: {str(e)}")]

    elif name == "check_url_updates":
        auto_rescrape = arguments.get("auto_rescrape", False)

        try:
            results = kb.check_url_updates(auto_rescrape)

            output = "URL Update Check:\n\n"

            if results['unchanged']:
                output += f"✓ {len(results['unchanged'])} documents unchanged\n"

            if results['changed']:
                output += f"⚠ {len(results['changed'])} documents have updates:\n"
                for doc in results['changed'][:10]:  # Show first 10
                    output += f"  - {doc['title']}\n"
                    output += f"    URL: {doc['url']}\n"
                    output += f"    Last modified: {doc['last_modified']}\n"
                if len(results['changed']) > 10:
                    output += f"  ... and {len(results['changed']) - 10} more\n"
                output += "\n"

            if results['failed']:
                output += f"✗ {len(results['failed'])} checks failed:\n"
                for doc in results['failed'][:5]:
                    output += f"  - {doc['title']}: {doc['error']}\n"
                output += "\n"

            if auto_rescrape and results['rescraped']:
                output += f"✓ {len(results['rescraped'])} documents re-scraped\n"
            elif not auto_rescrape and results['changed']:
                output += "\nTip: Use auto_rescrape=true to automatically update changed documents.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error checking updates: {str(e)}")]

    elif name == "remove_document":
        doc_id = arguments.get("doc_id")
        
        if kb.remove_document(doc_id):
            return [TextContent(type="text", text=f"Successfully removed document: {doc_id}")]
        else:
            return [TextContent(type="text", text=f"Document not found: {doc_id}")]
    
    elif name == "find_similar":
        doc_id = arguments.get("doc_id")
        chunk_id = arguments.get("chunk_id")
        max_results = arguments.get("max_results", 5)
        tags = arguments.get("tags")

        # Check if document exists
        if doc_id not in kb.documents:
            return [TextContent(type="text", text=f"Document not found: {doc_id}")]

        try:
            results = kb.find_similar_documents(doc_id, chunk_id, max_results, tags)
        except Exception as e:
            return [TextContent(type="text", text=f"Error finding similar documents: {str(e)}")]

        if not results:
            return [TextContent(type="text", text=f"No similar documents found for: {doc_id}")]

        source_doc = kb.documents[doc_id]
        if chunk_id is not None:
            output = f"Documents similar to '{source_doc.title}' (chunk {chunk_id}):\n\n"
        else:
            output = f"Documents similar to '{source_doc.title}':\n\n"

        for i, r in enumerate(results, 1):
            output += f"--- {i}. {r['title']} ({r['filename']}) ---\n"
            output += f"Doc ID: {r['doc_id']}, Similarity: {r['similarity']:.4f}\n"
            output += f"Tags: {', '.join(r['tags']) if r['tags'] else 'none'}\n"
            output += f"Snippet:\n{r['snippet']}\n\n"

        return [TextContent(type="text", text=output)]

    elif name == "answer_question":
        question = arguments.get("question", "")
        max_sources = arguments.get("max_sources", 5)
        search_mode = arguments.get("search_mode", "auto")

        # Validate input
        if not question or len(question.strip()) < 3:
            return [TextContent(type="text", text="Error: Question must be at least 3 characters long")]

        # Call RAG implementation
        try:
            result = kb.answer_question(
                question,
                max_context_chunks=max_sources,
                force_search_mode=None if search_mode == "auto" else search_mode
            )
        except Exception as e:
            return [TextContent(type="text", text=f"Error answering question: {str(e)}")]

        # Format response
        output = f"# Answer\n\n{result['answer']}\n\n"

        # Add sources section if available
        if result['sources']:
            output += "## Sources\n\n"
            for i, source in enumerate(result['sources'], 1):
                doc = kb.documents.get(source['doc_id'])
                if doc:
                    output += f"**{i}. {doc.title}** ({doc.filename})\n"
                    output += f"   - Chunk ID: {source['chunk_id']}\n"
                    output += f"   - Relevance: {source['score']:.2%}\n"
                    if 'page' in source and source['page']:
                        output += f"   - Page: {source['page']}\n"
                    output += "\n"

        # Add metadata section
        output += "## Answer Metadata\n\n"
        output += f"- **Confidence**: {result['confidence']:.1%}\n"
        output += f"- **Search Mode**: {result.get('reasoning', 'Unknown')}\n"
        output += f"- **LLM Model**: {result.get('model', 'Fallback')}\n"
        output += f"- **Sources Used**: {len(result['sources'])} documents\n"

        # Add error note if applicable
        if result.get('error'):
            output += f"\n⚠️ **Note**: {result['error']}\n"

        return [TextContent(type="text", text=output)]

    elif name == "kb_stats":
        stats = kb.get_stats()
        output = "Knowledge Base Statistics:\n"
        output += f"  Documents: {stats['total_documents']}\n"
        output += f"  Chunks: {stats['total_chunks']}\n"
        output += f"  Total Words: {stats['total_words']:,}\n"
        output += f"  File Types: {', '.join(stats['file_types']) or 'none'}\n"
        output += f"  Tags: {', '.join(stats['all_tags']) or 'none'}\n"
        return [TextContent(type="text", text=output)]

    elif name == "health_check":
        health = kb.health_check()

        # Format output
        output = f"System Health Check\n{'='*50}\n\n"
        output += f"Status: {health['status'].upper()}\n"
        output += f"Message: {health['message']}\n\n"

        # Metrics
        if health['metrics']:
            output += "Metrics:\n"
            for key, value in health['metrics'].items():
                if isinstance(value, int):
                    output += f"  {key}: {value:,}\n"
                else:
                    output += f"  {key}: {value}\n"
            output += "\n"

        # Database
        if health['database']:
            output += "Database:\n"
            for key, value in health['database'].items():
                output += f"  {key}: {value}\n"
            output += "\n"

        # Features
        if health['features']:
            output += "Features:\n"
            for key, value in health['features'].items():
                status = "✓" if value else "✗"
                output += f"  {status} {key}: {value}\n"
            output += "\n"

        # Performance
        if health['performance']:
            output += "Performance:\n"
            for key, value in health['performance'].items():
                output += f"  {key}: {value}\n"
            output += "\n"

        # Issues
        if health['issues']:
            output += f"Issues ({len(health['issues'])}):\n"
            for i, issue in enumerate(health['issues'], 1):
                output += f"  {i}. {issue}\n"
        else:
            output += "✓ No issues detected\n"

        return [TextContent(type="text", text=output)]

    elif name == "detect_anomalies":
        min_severity = arguments.get("min_severity", "moderate")
        days = arguments.get("days", 7)

        results = kb.detect_anomalies(min_severity=min_severity, days=days)

        if 'error' in results:
            return [TextContent(type="text", text=f"Error: {results['error']}")]

        # Format output
        output = f"Anomaly Detection Results\n{'='*60}\n\n"
        output += f"Time Range: Last {results['time_range_days']} days\n"
        output += f"Minimum Severity: {min_severity}\n"
        output += f"Total Anomalies: {results['total_anomalies']}\n"

        if results['by_severity']:
            output += f"\nBreakdown by Severity:\n"
            for severity, count in sorted(results['by_severity'].items()):
                output += f"  {severity.title()}: {count}\n"

        if results['total_anomalies'] > 0:
            output += f"\nAverage Score: {results['avg_score']}/100\n\n"
            output += f"Anomalies Detected:\n"
            output += "-" * 60 + "\n\n"

            for anomaly in results['anomalies']:
                output += f"Document: {anomaly.get('doc_title', 'Unknown')}\n"
                output += f"  Severity: {anomaly.get('severity', 'unknown').upper()} (Score: {anomaly.get('score', 0)}/100)\n"
                output += f"  Check Date: {anomaly.get('check_date', 'N/A')}\n"
                output += f"  Status: {anomaly.get('status', 'N/A')}\n"

                if 'components' in anomaly:
                    output += f"  Score Components:\n"
                    for component, score in anomaly['components'].items():
                        output += f"    - {component}: {score:.1f}\n"

                if 'change_type' in anomaly and anomaly['change_type']:
                    output += f"  Change Type: {anomaly['change_type']}\n"

                if 'http_status' in anomaly and anomaly['http_status']:
                    output += f"  HTTP Status: {anomaly['http_status']}\n"

                if 'error_message' in anomaly and anomaly['error_message']:
                    output += f"  Error: {anomaly['error_message']}\n"

                output += "\n"
        else:
            output += "\n✓ No anomalies detected in the specified time range.\n"

        return [TextContent(type="text", text=output)]

    elif name == "search_analytics":
        days = arguments.get("days", 30)
        limit = arguments.get("limit", 100)

        analytics = kb.get_search_analytics(days, limit)

        if 'error' in analytics:
            return [TextContent(type="text", text=f"Error getting analytics: {analytics['error']}")]

        # Format output
        output = f"Search Analytics (Last {days} days)\n{'='*50}\n\n"

        # Overall stats
        output += "Overview:\n"
        output += f"  Total Searches: {analytics.get('total_searches', 0):,}\n"
        output += f"  Unique Queries: {analytics.get('unique_queries', 0):,}\n"
        output += f"  Avg Results per Search: {analytics.get('avg_results', 0)}\n"
        output += f"  Avg Execution Time: {analytics.get('avg_execution_time_ms', 0):.2f}ms\n\n"

        # Search modes
        if analytics.get('search_modes'):
            output += "Search Mode Usage:\n"
            for mode in analytics['search_modes']:
                output += f"  {mode['mode']}: {mode['count']:,} searches (avg {mode['avg_results']} results)\n"
            output += "\n"

        # Top queries
        if analytics.get('top_queries'):
            output += f"Top {min(10, len(analytics['top_queries']))} Most Popular Queries:\n"
            for i, query in enumerate(analytics['top_queries'][:10], 1):
                output += f"  {i}. \"{query['query']}\" - {query['count']} times (avg {query['avg_results']} results)\n"
            output += "\n"

        # Failed searches
        if analytics.get('failed_searches'):
            output += f"Top {min(10, len(analytics['failed_searches']))} Failed Searches (0 results):\n"
            for i, failed in enumerate(analytics['failed_searches'][:10], 1):
                output += f"  {i}. \"{failed['query']}\" - {failed['count']} times\n"
            output += "\n"

        # Popular tags
        if analytics.get('popular_tags'):
            output += f"Top {min(10, len(analytics['popular_tags']))} Most Used Tags:\n"
            for i, tag in enumerate(analytics['popular_tags'][:10], 1):
                output += f"  {i}. {tag['tag']}: {tag['count']} searches\n"

        return [TextContent(type="text", text=output)]

    elif name == "find_by_reference":
        ref_type = arguments.get("ref_type")
        ref_value = arguments.get("ref_value")
        max_results = arguments.get("max_results", 10)

        if not ref_type or not ref_value:
            return [TextContent(type="text", text="Error: ref_type and ref_value are required")]

        results = kb.find_by_reference(ref_type, ref_value, max_results)

        if not results:
            return [TextContent(type="text", text=f"No references found for {ref_type}={ref_value}")]

        output = f"Found {len(results)} references for {ref_type}={ref_value}:\n\n"
        for i, r in enumerate(results, 1):
            output += f"--- Reference {i} ---\n"
            output += f"Document: {r['title']} ({r['filename']})\n"
            output += f"Doc ID: {r['doc_id']}, Chunk: {r['chunk_id']}\n"
            output += f"Type: {r['ref_type']}, Value: {r['ref_value']}\n"
            output += f"Context:\n{r['context']}\n\n"

        return [TextContent(type="text", text=output)]

    elif name == "check_updates":
        auto_update = arguments.get("auto_update", False)
        results = kb.check_all_updates(auto_update)

        output = "Document Update Check:\n\n"

        if results['unchanged']:
            output += f"✓ {len(results['unchanged'])} documents unchanged\n"

        if results['changed']:
            output += f"⚠ {len(results['changed'])} documents changed:\n"
            for doc in results['changed']:
                output += f"  - {doc['title']} ({doc['filepath']})\n"
            output += "\n"

        if results['missing']:
            output += f"✗ {len(results['missing'])} documents missing (files not found):\n"
            for doc in results['missing']:
                output += f"  - {doc['title']} ({doc['filepath']})\n"
            output += "\n"

        if auto_update and results['updated']:
            output += f"✓ {len(results['updated'])} documents re-indexed:\n"
            for doc in results['updated']:
                output += f"  - {doc['title']} ({doc['filepath']})\n"

        if not auto_update and results['changed']:
            output += "\nRun with auto_update=true to automatically re-index changed documents.\n"

        return [TextContent(type="text", text=output)]

    elif name == "add_documents_bulk":
        directory = arguments.get("directory")
        pattern = arguments.get("pattern", "**/*.{pdf,txt}")
        tags = arguments.get("tags")
        recursive = arguments.get("recursive", True)
        skip_duplicates = arguments.get("skip_duplicates", True)

        try:
            results = kb.add_documents_bulk(directory, pattern, tags, recursive, skip_duplicates)
        except Exception as e:
            return [TextContent(type="text", text=f"Error in bulk add: {str(e)}")]

        output = "Bulk Document Add Results:\n\n"

        if results['added']:
            output += f"✓ {len(results['added'])} documents added:\n"
            for doc in results['added']:
                output += f"  - {doc['title']} ({doc['filename']})\n"
                output += f"    ID: {doc['doc_id']}, Chunks: {doc['chunks']}\n"
            output += "\n"

        if results['skipped']:
            output += f"⊘ {len(results['skipped'])} documents skipped (duplicates):\n"
            for doc in results['skipped']:
                output += f"  - {doc['filepath']}\n"
            output += "\n"

        if results['failed']:
            output += f"✗ {len(results['failed'])} documents failed:\n"
            for failure in results['failed']:
                output += f"  - {failure['filepath']}: {failure['error']}\n"
            output += "\n"

        output += f"Total: {len(results['added'])} added, {len(results['skipped'])} skipped, {len(results['failed'])} failed"

        return [TextContent(type="text", text=output)]

    elif name == "remove_documents_bulk":
        doc_ids = arguments.get("doc_ids")
        tags = arguments.get("tags")

        try:
            results = kb.remove_documents_bulk(doc_ids, tags)
        except Exception as e:
            return [TextContent(type="text", text=f"Error in bulk remove: {str(e)}")]

        output = "Bulk Document Remove Results:\n\n"

        if results['removed']:
            output += f"✓ {len(results['removed'])} documents removed:\n"
            for doc_id in results['removed']:
                output += f"  - {doc_id}\n"
            output += "\n"

        if results['failed']:
            output += f"✗ {len(results['failed'])} documents failed to remove:\n"
            for failure in results['failed']:
                output += f"  - {failure['doc_id']}: {failure['error']}\n"
            output += "\n"

        output += f"Total: {len(results['removed'])} removed, {len(results['failed'])} failed"

        return [TextContent(type="text", text=output)]

    elif name == "update_tags_bulk":
        doc_ids = arguments.get("doc_ids")
        existing_tags = arguments.get("existing_tags")
        add_tags = arguments.get("add_tags")
        remove_tags = arguments.get("remove_tags")
        replace_tags = arguments.get("replace_tags")

        try:
            results = kb.update_tags_bulk(
                doc_ids=doc_ids,
                existing_tags=existing_tags,
                add_tags=add_tags,
                remove_tags=remove_tags,
                replace_tags=replace_tags
            )
        except Exception as e:
            return [TextContent(type="text", text=f"Error in bulk tag update: {str(e)}")]

        output = "Bulk Tag Update Results:\n\n"

        if results['updated']:
            output += f"✓ {len(results['updated'])} documents updated:\n"
            for update in results['updated']:
                output += f"  - {update['doc_id']}\n"
                output += f"    Old tags: {', '.join(update['old_tags']) if update['old_tags'] else 'None'}\n"
                output += f"    New tags: {', '.join(update['new_tags']) if update['new_tags'] else 'None'}\n"
            output += "\n"

        if results['failed']:
            output += f"✗ {len(results['failed'])} documents failed to update:\n"
            for failure in results['failed']:
                output += f"  - {failure['doc_id']}: {failure['error']}\n"
            output += "\n"

        output += f"Total: {len(results['updated'])} updated, {len(results['failed'])} failed"

        return [TextContent(type="text", text=output)]

    elif name == "export_documents_bulk":
        doc_ids = arguments.get("doc_ids")
        tags = arguments.get("tags")
        format = arguments.get("format", "json")

        try:
            export_data = kb.export_documents_bulk(doc_ids=doc_ids, tags=tags, format=format)
        except Exception as e:
            return [TextContent(type="text", text=f"Error in bulk export: {str(e)}")]

        # Determine document count from the export
        if format == "json":
            import json as json_lib
            doc_count = len(json_lib.loads(export_data))
        else:
            doc_count = export_data.count('\n') if format == 'csv' else export_data.count('## ') - 1

        output = f"Document Export ({format.upper()}):\n\n"
        output += f"Exported {doc_count} document(s)\n\n"
        output += "=" * 80 + "\n\n"
        output += export_data

        return [TextContent(type="text", text=output)]

    elif name == "search_tables":
        query = arguments.get("query", "")
        max_results = arguments.get("max_results", 5)
        tags = arguments.get("tags")

        try:
            results = kb.search_tables(query, max_results, tags)
        except Exception as e:
            return [TextContent(type="text", text=f"Error searching tables: {str(e)}")]

        if not results:
            return [TextContent(type="text", text=f"No tables found for query: '{query}'")]

        output = f"Found {len(results)} table(s) for '{query}':\n\n"

        for i, result in enumerate(results, 1):
            output += f"Result {i}:\n"
            output += f"  Document: {result['doc_title']}\n"
            output += f"  Page: {result['page']}\n"
            output += f"  Size: {result['row_count']} rows × {result['col_count']} columns\n"
            output += f"  Score: {result['score']:.2f}\n\n"
            output += f"Table content:\n{result['markdown']}\n\n"
            output += "-" * 80 + "\n\n"

        return [TextContent(type="text", text=output)]

    elif name == "search_code":
        query = arguments.get("query", "")
        max_results = arguments.get("max_results", 5)
        block_type = arguments.get("block_type")
        tags = arguments.get("tags")

        try:
            results = kb.search_code(query, max_results, block_type, tags)
        except Exception as e:
            return [TextContent(type="text", text=f"Error searching code: {str(e)}")]

        if not results:
            type_filter = f" (type: {block_type})" if block_type else ""
            return [TextContent(type="text", text=f"No code blocks found for query: '{query}'{type_filter}")]

        output = f"Found {len(results)} code block(s) for '{query}':\n\n"

        for i, result in enumerate(results, 1):
            output += f"Result {i}:\n"
            output += f"  Document: {result['doc_title']}\n"
            output += f"  Page: {result['page'] or 'N/A'}\n"
            output += f"  Type: {result['block_type']}\n"
            output += f"  Lines: {result['line_count']}\n"
            output += f"  Score: {result['score']:.2f}\n\n"
            output += f"Code:\n```{result['block_type']}\n{result['code']}\n```\n\n"
            output += "-" * 80 + "\n\n"

        return [TextContent(type="text", text=output)]

    elif name == "suggest_queries":
        partial = arguments.get("partial", "")
        max_suggestions = arguments.get("max_suggestions", 5)
        category = arguments.get("category")

        if not partial:
            return [TextContent(type="text", text="Error: partial query string is required")]

        try:
            suggestions = kb.get_query_suggestions(partial, max_suggestions, category)
        except Exception as e:
            return [TextContent(type="text", text=f"Error getting suggestions: {str(e)}")]

        if not suggestions:
            return [TextContent(type="text", text=f"No suggestions found for: '{partial}'")]

        output = f"Query suggestions for '{partial}':\n\n"
        for i, sug in enumerate(suggestions, 1):
            output += f"{i}. {sug['term']} ({sug['category']}) - used {sug['frequency']} times\n"

        return [TextContent(type="text", text=output)]

    elif name == "export_results":
        results = arguments.get("results", [])
        format = arguments.get("format", "markdown")
        query = arguments.get("query")

        if not results:
            return [TextContent(type="text", text="Error: results array is required")]

        try:
            exported = kb.export_search_results(results, format, query)

            # Return the exported content
            output = f"Search results exported to {format} format:\n\n"
            output += "=" * 80 + "\n\n"
            output += exported

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error exporting results: {str(e)}")]

    elif name == "create_backup":
        dest_dir = arguments.get("dest_dir")
        compress = arguments.get("compress", True)

        if not dest_dir:
            return [TextContent(type="text", text="Error: dest_dir is required")]

        try:
            backup_path = kb.create_backup(dest_dir, compress)

            output = "✓ Backup created successfully!\n\n"
            output += f"Location: {backup_path}\n"
            output += f"Format: {'Compressed (ZIP)' if compress else 'Uncompressed directory'}\n\n"
            output += "The backup includes:\n"
            output += f"- Database ({len(kb.documents)} documents)\n"
            output += "- Embeddings (if available)\n"
            output += "- Metadata file with timestamp and version info\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error creating backup: {str(e)}")]

    elif name == "restore_backup":
        backup_path = arguments.get("backup_path")
        verify = arguments.get("verify", True)

        if not backup_path:
            return [TextContent(type="text", text="Error: backup_path is required")]

        try:
            result = kb.restore_from_backup(backup_path, verify)

            output = "✓ Restore completed successfully!\n\n"
            output += f"Backup source: {backup_path}\n"
            output += f"Documents restored: {result['restored_documents']}\n"
            output += f"Time elapsed: {result['elapsed_seconds']:.2f}s\n\n"

            if 'backup_metadata' in result:
                metadata = result['backup_metadata']
                output += "Backup info:\n"
                output += f"- Created: {metadata.get('created_at', 'Unknown')}\n"
                output += f"- Version: {metadata.get('version', 'Unknown')}\n"
                output += f"- Original document count: {metadata.get('document_count', 'Unknown')}\n"

            output += "\nNote: A safety backup was created before restoration."

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error restoring backup: {str(e)}")]

    elif name == "auto_tag_document":
        doc_id = arguments.get("doc_id")
        confidence_threshold = arguments.get("confidence_threshold", 0.7)
        max_tags = arguments.get("max_tags", 10)
        append = arguments.get("append", True)

        if not doc_id:
            return [TextContent(type="text", text="Error: doc_id is required")]

        try:
            result = kb.auto_tag_document(
                doc_id,
                confidence_threshold=confidence_threshold,
                max_tags=max_tags,
                append=append
            )

            output = "✓ Auto-tagged document successfully!\n\n"
            output += f"**Document:** {result['doc_title']}\n"
            output += f"**Document ID:** {doc_id}\n\n"

            output += f"**Applied Tags ({len(result['applied_tags'])}):**\n"
            for tag_info in result['suggested_tags']:
                if tag_info['tag'] in result['applied_tags']:
                    output += f"  - {tag_info['tag']} (confidence: {tag_info['confidence']:.2f})\n"
                    output += f"    Reason: {tag_info.get('reason', 'N/A')}\n"

            if result['skipped_tags']:
                output += f"\n**Skipped Tags (below {confidence_threshold} threshold):**\n"
                for tag_info in result['suggested_tags']:
                    if tag_info['tag'] in result['skipped_tags']:
                        output += f"  - {tag_info['tag']} (confidence: {tag_info['confidence']:.2f})\n"

            output += "\n**Tag Summary:**\n"
            output += f"  - Existing tags: {', '.join(result['existing_tags']) if result['existing_tags'] else 'None'}\n"
            output += f"  - New tags: {', '.join(result['new_tags'])}\n"
            output += f"  - Total tags added: {len(set(result['new_tags']) - set(result['existing_tags']))}\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error auto-tagging document: {str(e)}\n\nNote: Auto-tagging requires LLM configuration. Set LLM_PROVIDER and appropriate API key (ANTHROPIC_API_KEY or OPENAI_API_KEY).")]

    elif name == "auto_tag_all":
        confidence_threshold = arguments.get("confidence_threshold", 0.7)
        max_tags = arguments.get("max_tags", 10)
        append = arguments.get("append", True)
        skip_tagged = arguments.get("skip_tagged", True)
        max_docs = arguments.get("max_docs")

        try:
            results = kb.auto_tag_all_documents(
                confidence_threshold=confidence_threshold,
                max_tags=max_tags,
                append=append,
                skip_tagged=skip_tagged,
                max_docs=max_docs
            )

            output = "✓ Bulk auto-tagging complete!\n\n"
            output += "**Statistics:**\n"
            output += f"  - Documents processed: {results['processed']}\n"
            output += f"  - Documents skipped: {results['skipped']}\n"
            output += f"  - Documents failed: {results['failed']}\n"
            output += f"  - Total tags added: {results['total_tags_added']}\n\n"

            if results['processed'] > 0:
                output += "**Sample Results (first 5):**\n"
                for i, result in enumerate(results['results'][:5], 1):
                    if 'error' in result:
                        output += f"\n{i}. {result['doc_id']} - ERROR: {result['error']}\n"
                    else:
                        output += f"\n{i}. {result.get('doc_title', 'Unknown')}\n"
                        output += f"   - Applied: {', '.join(result['applied_tags']) if result['applied_tags'] else 'None'}\n"
                        output += f"   - Total tags: {len(result['new_tags'])}\n"

                if results['processed'] > 5:
                    output += f"\n... and {results['processed'] - 5} more documents\n"

            if results['failed'] > 0:
                output += f"\n**Warning:** {results['failed']} documents failed to process. Check logs for details.\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error in bulk auto-tagging: {str(e)}\n\nNote: Auto-tagging requires LLM configuration. Set LLM_PROVIDER and appropriate API key (ANTHROPIC_API_KEY or OPENAI_API_KEY).")]

    elif name == "summarize_document":
        doc_id = arguments.get("doc_id")
        summary_type = arguments.get("summary_type", "brief")
        force_regenerate = arguments.get("force_regenerate", False)

        if not doc_id:
            return [TextContent(type="text", text="Error: doc_id is required")]

        try:
            summary = kb.generate_summary(
                doc_id,
                summary_type=summary_type,
                force_regenerate=force_regenerate
            )

            output = f"✓ Summary generated ({summary_type})\n\n"
            output += f"**Document:** {kb.documents[doc_id].title}\n\n"
            output += f"**Summary ({summary_type}):**\n\n"
            output += summary

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error generating summary: {str(e)}\n\nNote: Summarization requires LLM configuration. Set LLM_PROVIDER and appropriate API key (ANTHROPIC_API_KEY or OPENAI_API_KEY).")]

    elif name == "get_summary":
        doc_id = arguments.get("doc_id")
        summary_type = arguments.get("summary_type", "brief")

        if not doc_id:
            return [TextContent(type="text", text="Error: doc_id is required")]

        try:
            summary = kb.get_summary(doc_id, summary_type)

            if not summary:
                return [TextContent(type="text", text=f"No cached summary found for {doc_id} ({summary_type}). Use 'summarize_document' to generate one.")]

            output = f"✓ Cached summary retrieved ({summary_type})\n\n"
            output += f"**Document:** {kb.documents[doc_id].title}\n\n"
            output += f"**Summary ({summary_type}):**\n\n"
            output += summary

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error retrieving summary: {str(e)}")]

    elif name == "summarize_all":
        summary_types = arguments.get("summary_types", ["brief"])
        force_regenerate = arguments.get("force_regenerate", False)
        max_docs = arguments.get("max_docs")

        try:
            results = kb.generate_summary_all(
                summary_types=summary_types,
                force_regenerate=force_regenerate,
                max_docs=max_docs
            )

            output = "✓ Bulk summarization complete!\n\n"
            output += "**Statistics:**\n"
            output += f"  - Documents processed: {results['processed']}\n"
            output += f"  - Documents failed: {results['failed']}\n"
            output += f"  - Total summaries generated: {results['total_summaries']}\n"
            output += "  - By type:\n"
            for summary_type, count in results['by_type'].items():
                output += f"    - {summary_type}: {count}\n"

            output += "\n**Sample Results (first 3):**\n"
            for i, result in enumerate(results['results'][:3], 1):
                output += f"\n{i}. {result['title']}\n"
                for summary_type, summary_result in result['summaries'].items():
                    if summary_result['success']:
                        output += f"   - {summary_type}: {summary_result['word_count']} words\n"
                    else:
                        output += f"   - {summary_type}: ERROR - {summary_result['error']}\n"

            if results['processed'] > 3:
                output += f"\n... and {results['processed'] - 3} more documents\n"

            if results['failed'] > 0:
                output += f"\n**Warning:** {results['failed']} documents failed to process. Check logs for details.\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error in bulk summarization: {str(e)}\n\nNote: Summarization requires LLM configuration. Set LLM_PROVIDER and appropriate API key (ANTHROPIC_API_KEY or OPENAI_API_KEY).")]

    elif name == "extract_entities":
        doc_id = arguments.get("doc_id")
        confidence_threshold = arguments.get("confidence_threshold", 0.6)
        force_regenerate = arguments.get("force_regenerate", False)

        if not doc_id:
            return [TextContent(type="text", text="Error: doc_id is required")]

        try:
            result = kb.extract_entities(
                doc_id,
                confidence_threshold=confidence_threshold,
                force_regenerate=force_regenerate
            )

            output = "✓ Entity extraction complete!\n\n"
            output += f"**Document:** {result['doc_title']}\n"
            output += f"**Document ID:** {doc_id}\n"
            output += f"**Entities Found:** {result['entity_count']}\n\n"

            # Group by entity type
            if result['entities']:
                for entity_type in sorted(result['types'].keys()):
                    entities_of_type = [e for e in result['entities'] if e['entity_type'] == entity_type]
                    output += f"**{entity_type.upper().replace('_', ' ')}** ({len(entities_of_type)}):\n"

                    # Show first 5 of each type
                    for entity in entities_of_type[:5]:
                        output += f"  - **{entity['entity_text']}** (confidence: {entity['confidence']:.2f}"
                        if entity.get('occurrence_count', 1) > 1:
                            output += f", occurs {entity['occurrence_count']}x"
                        output += ")\n"
                        if entity.get('context'):
                            context = entity['context'][:80] + "..." if len(entity['context']) > 80 else entity['context']
                            output += f"    *{context}*\n"

                    if len(entities_of_type) > 5:
                        output += f"  ... and {len(entities_of_type) - 5} more\n"
                    output += "\n"

                output += "Use `list_entities` tool to see all entities with filtering options.\n"
            else:
                output += "No entities found with the current confidence threshold.\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error extracting entities: {str(e)}\n\nNote: Entity extraction requires LLM configuration. Set LLM_PROVIDER and appropriate API key (ANTHROPIC_API_KEY or OPENAI_API_KEY).")]

    elif name == "list_entities":
        doc_id = arguments.get("doc_id")
        entity_types = arguments.get("entity_types")
        min_confidence = arguments.get("min_confidence", 0.0)

        if not doc_id:
            return [TextContent(type="text", text="Error: doc_id is required")]

        try:
            result = kb.get_entities(
                doc_id,
                entity_types=entity_types,
                min_confidence=min_confidence
            )

            output = f"**Entities for Document:** {result['doc_title']}\n"
            output += f"**Document ID:** {doc_id}\n"

            # Show filters if applied
            if entity_types or min_confidence > 0:
                output += "**Filters:** "
                filters = []
                if entity_types:
                    filters.append(f"types={', '.join(entity_types)}")
                if min_confidence > 0:
                    filters.append(f"min_confidence={min_confidence}")
                output += ', '.join(filters) + "\n"

            output += f"**Total Entities:** {result['entity_count']}\n\n"

            if result['entities']:
                # Group by type
                for entity_type in sorted(result['types'].keys()):
                    entities_of_type = [e for e in result['entities'] if e['entity_type'] == entity_type]
                    output += f"**{entity_type.upper().replace('_', ' ')}** ({len(entities_of_type)}):\n"

                    for entity in entities_of_type:
                        output += f"  - **{entity['entity_text']}** (conf: {entity['confidence']:.2f}"
                        if entity.get('occurrence_count', 1) > 1:
                            output += f", {entity['occurrence_count']}x"
                        output += ")\n"
                        if entity.get('context'):
                            context = entity['context'][:100] + "..." if len(entity['context']) > 100 else entity['context']
                            output += f"    *{context}*\n"

                    output += "\n"
            else:
                output += "No entities found. Use `extract_entities` tool to extract entities first.\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error listing entities: {str(e)}")]

    elif name == "search_entities":
        query = arguments.get("query")
        entity_types = arguments.get("entity_types")
        min_confidence = arguments.get("min_confidence", 0.0)
        max_results = arguments.get("max_results", 20)

        if not query:
            return [TextContent(type="text", text="Error: query is required")]

        try:
            result = kb.search_entities(
                query,
                entity_types=entity_types,
                min_confidence=min_confidence,
                max_results=max_results
            )

            output = f"**Entity Search Results for:** {result['query']}\n"
            output += f"**Total Matches:** {result['total_matches']}\n"

            # Show filters if applied
            if entity_types or min_confidence > 0:
                output += "**Filters:** "
                filters = []
                if entity_types:
                    filters.append(f"types={', '.join(entity_types)}")
                if min_confidence > 0:
                    filters.append(f"min_confidence={min_confidence}")
                output += ', '.join(filters) + "\n"

            output += f"**Documents Found:** {len(result['documents'])}\n\n"

            if result['documents']:
                for doc in result['documents']:
                    output += f"**{doc['doc_title']}** ({doc['doc_id']})\n"
                    output += f"  Matches: {doc['match_count']}\n"

                    # Show first 3 matches per document
                    for match in doc['matches'][:3]:
                        output += f"  - **{match['entity_text']}** ({match['entity_type']}, conf: {match['confidence']:.2f}"
                        if match.get('occurrence_count', 1) > 1:
                            output += f", {match['occurrence_count']}x"
                        output += ")\n"
                        if match.get('context'):
                            context = match['context'][:80] + "..." if len(match['context']) > 80 else match['context']
                            output += f"    *{context}*\n"

                    if doc['match_count'] > 3:
                        output += f"  ... and {doc['match_count'] - 3} more matches\n"
                    output += "\n"
            else:
                output += "No entities found matching your query.\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error searching entities: {str(e)}")]

    elif name == "entity_stats":
        entity_type = arguments.get("entity_type")

        try:
            result = kb.get_entity_stats(entity_type=entity_type)

            output = "**Entity Statistics**\n"
            if entity_type:
                output += f"**Type Filter:** {entity_type}\n"
            output += "\n"

            output += f"**Total Entities:** {result['total_entities']}\n"
            output += f"**Documents with Entities:** {result['total_documents_with_entities']}\n\n"

            # Breakdown by type
            if result['by_type']:
                output += "**Entities by Type:**\n"
                for ent_type, count in sorted(result['by_type'].items(), key=lambda x: x[1], reverse=True):
                    output += f"  - {ent_type.replace('_', ' ')}: {count}\n"
                output += "\n"

            # Top entities
            if result['top_entities']:
                output += "**Top Entities (by document count):**\n"
                for i, entity in enumerate(result['top_entities'][:10], 1):
                    output += f"{i}. **{entity['entity_text']}** ({entity['entity_type']})\n"
                    output += f"   - Found in {entity['document_count']} document(s)\n"
                    output += f"   - Total occurrences: {entity['total_occurrences']}\n"
                    output += f"   - Avg confidence: {entity['avg_confidence']:.2f}\n"
                output += "\n"

            # Documents with most entities
            if result['documents_with_most_entities']:
                output += "**Documents with Most Entities:**\n"
                for i, doc in enumerate(result['documents_with_most_entities'], 1):
                    output += f"{i}. **{doc['doc_title']}**: {doc['entity_count']} entities\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error getting entity stats: {str(e)}")]

    elif name == "get_entity_analytics":
        time_range_days = arguments.get("time_range_days", 30)

        try:
            result = kb.get_entity_analytics(time_range_days=time_range_days)

            output = "**Entity Analytics Dashboard**\n"
            output += f"**Time Range:** Last {time_range_days} days\n\n"

            # Summary Metrics
            if 'summary_metrics' in result:
                metrics = result['summary_metrics']
                output += "**Summary Metrics:**\n"
                output += f"  - Total Entities: {metrics.get('total_entities', 0)}\n"
                output += f"  - Unique Entity Texts: {metrics.get('unique_entity_texts', 0)}\n"
                output += f"  - Total Relationships: {metrics.get('total_relationships', 0)}\n"
                output += f"  - Documents with Entities: {metrics.get('documents_with_entities', 0)}\n"
                output += f"  - Avg Entities per Document: {metrics.get('avg_entities_per_doc', 0):.2f}\n\n"

            # Entity Distribution by Type
            if 'entity_distribution' in result and result['entity_distribution']:
                output += "**Entity Distribution by Type:**\n"
                for entity_type, count in sorted(result['entity_distribution'].items(), key=lambda x: x[1], reverse=True):
                    output += f"  - {entity_type.replace('_', ' ').title()}: {count}\n"
                output += "\n"

            # Top Entities
            if 'top_entities' in result and result['top_entities']:
                output += "**Top 10 Entities (by document count):**\n"
                for i, entity in enumerate(result['top_entities'][:10], 1):
                    output += f"{i}. **{entity['entity_text']}** ({entity['entity_type']})\n"
                    output += f"   - Documents: {entity['doc_count']}\n"
                    output += f"   - Avg Confidence: {entity['avg_confidence']:.2f}\n"
                output += "\n"

            # Relationship Statistics
            if 'relationship_stats' in result and result['relationship_stats']:
                stats = result['relationship_stats']
                output += "**Relationship Statistics:**\n"
                output += f"  - Total Relationships: {stats.get('total', 0)}\n"
                output += f"  - Avg Strength: {stats.get('avg_strength', 0):.3f}\n"
                output += f"  - Max Strength: {stats.get('max_strength', 0):.3f}\n"
                if 'by_type' in stats and stats['by_type']:
                    output += "  - By Type:\n"
                    for rel_type, count in sorted(stats['by_type'].items(), key=lambda x: x[1], reverse=True)[:5]:
                        output += f"    - {rel_type}: {count}\n"
                output += "\n"

            # Top Relationships
            if 'top_relationships' in result and result['top_relationships']:
                output += "**Top 10 Entity Relationships:**\n"
                for i, rel in enumerate(result['top_relationships'][:10], 1):
                    output += f"{i}. **{rel['entity1']}** ({rel['entity1_type']}) <-> **{rel['entity2']}** ({rel['entity2_type']})\n"
                    output += f"   - Strength: {rel['strength']:.3f}\n"
                    output += f"   - Documents: {rel['doc_count']}\n"
                output += "\n"

            # Extraction Timeline
            if 'extraction_timeline' in result and result['extraction_timeline']:
                output += "**Extraction Timeline (Recent Activity):**\n"
                for entry in result['extraction_timeline'][:7]:  # Last 7 days
                    output += f"  - {entry['date']}: {entry['count']} entities extracted\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error getting entity analytics: {str(e)}")]

    elif name == "extract_entities_bulk":
        confidence_threshold = arguments.get("confidence_threshold", 0.6)
        force_regenerate = arguments.get("force_regenerate", False)
        max_docs = arguments.get("max_docs")
        skip_existing = arguments.get("skip_existing", True)

        try:
            result = kb.extract_entities_bulk(
                confidence_threshold=confidence_threshold,
                force_regenerate=force_regenerate,
                max_docs=max_docs,
                skip_existing=skip_existing
            )

            output = "**Bulk Entity Extraction Complete**\n\n"
            output += f"**Processed:** {result['processed']} documents\n"
            output += f"**Skipped:** {result['skipped']} documents (already have entities)\n"
            output += f"**Failed:** {result['failed']} documents\n"
            output += f"**Total Entities Extracted:** {result['total_entities']}\n\n"

            if result['by_type']:
                output += "**Entities by Type:**\n"
                for ent_type, count in sorted(result['by_type'].items(), key=lambda x: x[1], reverse=True):
                    output += f"  - {ent_type.replace('_', ' ')}: {count}\n"
                output += "\n"

            # Show sample results
            if result['results']:
                output += "**Sample Results (first 10):**\n"
                for i, doc_result in enumerate(result['results'][:10], 1):
                    status_emoji = "✓" if doc_result['status'] == 'success' else "⊗" if doc_result['status'] == 'failed' else "⊘"
                    output += f"{i}. {status_emoji} **{doc_result['title']}**"
                    if doc_result['status'] == 'success':
                        output += f" - {doc_result['entity_count']} entities"
                    elif doc_result['status'] == 'skipped':
                        output += f" - skipped ({doc_result['entity_count']} entities)"
                    elif doc_result['status'] == 'failed':
                        output += f" - ERROR: {doc_result.get('error', 'unknown error')}"
                    output += "\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error in bulk entity extraction: {str(e)}\n\nNote: Entity extraction requires LLM configuration. Set LLM_PROVIDER and appropriate API key (ANTHROPIC_API_KEY or OPENAI_API_KEY).")]

    elif name == "extract_entity_relationships":
        doc_id = arguments.get("doc_id")
        min_confidence = arguments.get("min_confidence", 0.6)

        if not doc_id:
            return [TextContent(type="text", text="Error: doc_id is required")]

        try:
            result = kb.extract_entity_relationships(doc_id, min_confidence=min_confidence)

            output = "**Entity Relationship Extraction Complete**\n\n"
            output += f"**Document:** {kb.documents[doc_id].title}\n"
            output += f"**Relationships Found:** {result['relationship_count']}\n\n"

            if result['relationships']:
                output += "**Top Relationships (by strength):**\n\n"
                # Show top 10 relationships
                for i, rel in enumerate(result['relationships'][:10], 1):
                    output += f"{i}. **{rel['entity1']}** ({rel['entity1_type']}) ↔ **{rel['entity2']}** ({rel['entity2_type']})\n"
                    output += f"   Strength: {rel['strength']:.2f}\n"
                    if rel.get('context'):
                        context = rel['context'][:100] + "..." if len(rel['context']) > 100 else rel['context']
                        output += f"   *{context}*\n"
                    output += "\n"

                if len(result['relationships']) > 10:
                    output += f"... and {len(result['relationships']) - 10} more relationships\n\n"

                output += "Use `get_entity_relationships` to explore specific entities.\n"
            else:
                output += "No relationships found. The document may not have enough entities extracted.\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error extracting relationships: {str(e)}")]

    elif name == "get_entity_relationships":
        entity_text = arguments.get("entity_text")
        min_strength = arguments.get("min_strength", 0.0)
        max_results = arguments.get("max_results", 20)

        if not entity_text:
            return [TextContent(type="text", text="Error: entity_text is required")]

        try:
            relationships = kb.get_entity_relationships(
                entity_text=entity_text,
                min_strength=min_strength,
                max_results=max_results
            )

            if not relationships:
                return [TextContent(type="text", text=f"No relationships found for entity '{entity_text}'.\n\nThis entity may not have been extracted yet, or it doesn't co-occur with other entities.")]

            output = f"**Entities Related to '{entity_text}'**\n\n"
            output += f"Found {len(relationships)} related entities:\n\n"

            for i, rel in enumerate(relationships, 1):
                output += f"{i}. **{rel['related_entity']}** ({rel['related_type']})\n"
                output += f"   Strength: {rel['strength']:.2f} | Found in {rel['doc_count']} document(s)\n"
                if rel.get('context_sample'):
                    context = rel['context_sample'][:100] + "..." if len(rel['context_sample']) > 100 else rel['context_sample']
                    output += f"   *{context}*\n"
                output += "\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error getting relationships: {str(e)}")]

    elif name == "find_related_entities":
        entity_text = arguments.get("entity_text")
        max_results = arguments.get("max_results", 10)

        if not entity_text:
            return [TextContent(type="text", text="Error: entity_text is required")]

        try:
            related = kb.find_related_entities(entity_text=entity_text, max_results=max_results)

            if not related:
                return [TextContent(type="text", text=f"No related entities found for '{entity_text}'.")]

            output = f"**Entities Related to '{entity_text}'** (Top {len(related)})\n\n"

            for i, rel in enumerate(related, 1):
                output += f"{i}. **{rel['related_entity']}** ({rel['related_type']}) - strength: {rel['strength']:.2f}\n"

            output += "\nUse `get_entity_relationships` for more details and context.\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error finding related entities: {str(e)}")]

    elif name == "search_entity_pair":
        entity1 = arguments.get("entity1")
        entity2 = arguments.get("entity2")
        max_results = arguments.get("max_results", 10)

        if not entity1 or not entity2:
            return [TextContent(type="text", text="Error: Both entity1 and entity2 are required")]

        try:
            results = kb.search_by_entity_pair(entity1=entity1, entity2=entity2, max_results=max_results)

            if not results:
                return [TextContent(type="text", text=f"No documents found containing both '{entity1}' and '{entity2}'.")]

            output = f"**Documents Containing Both '{entity1}' AND '{entity2}'**\n\n"
            output += f"Found {len(results)} document(s):\n\n"

            for i, doc in enumerate(results, 1):
                output += f"**{i}. {doc['title']}**\n"
                output += f"   '{entity1}': {doc['entity1_count']} mention(s) | '{entity2}': {doc['entity2_count']} mention(s)\n"
                output += f"   Doc ID: `{doc['doc_id']}`\n"

                if doc.get('contexts'):
                    output += "   **Context snippets:**\n"
                    for j, context in enumerate(doc['contexts'][:2], 1):
                        ctx_short = context[:150] + "..." if len(context) > 150 else context
                        output += f"   {j}. *{ctx_short}*\n"
                output += "\n"

            return [TextContent(type="text", text=output)]
        except Exception as e:
            return [TextContent(type="text", text=f"Error searching entity pair: {str(e)}")]

    elif name == "translate_query":
        query = arguments.get("query")
        confidence_threshold = arguments.get("confidence_threshold", 0.7)

        if not query:
            return [TextContent(type="text", text="Error: query is required")]

        try:
            result = kb.translate_nl_query(query=query, confidence_threshold=confidence_threshold)

            # Build formatted output
            output = "**Natural Language Query Translation**\n\n"
            output += f"**Original Query:** {result['original_query']}\n"
            output += f"**Suggested Query:** {result['suggested_query']}\n"
            output += f"**Confidence:** {result['confidence']:.2f}\n"
            output += f"**Search Mode:** {result['search_mode']}\n\n"

            # Show extracted search terms
            if result.get('search_terms'):
                output += f"**Search Terms:** {', '.join(result['search_terms'])}\n\n"

            # Show entities found
            if result.get('entities_found'):
                output += f"**Entities Detected:** ({len(result['entities_found'])} found)\n"
                for entity in result['entities_found'][:10]:  # Show top 10
                    source_badge = "🔍" if entity['source'] == 'regex' else "🤖"
                    output += f"  {source_badge} **{entity['text']}** ({entity['type']}) - confidence: {entity['confidence']:.2f}\n"
                if len(result['entities_found']) > 10:
                    output += f"  ... and {len(result['entities_found']) - 10} more\n"
                output += "\n"

            # Show facet filters
            if result.get('facet_filters'):
                output += "**Facet Filters:**\n"
                for facet_type, values in result['facet_filters'].items():
                    output += f"  - {facet_type}: {', '.join(values)}\n"
                output += "\n"

            # Show intent if available
            if result.get('intent'):
                output += f"**Intent:** {result['intent']}\n\n"

            # Warn if fallback mode
            if result.get('fallback'):
                output += "⚠️ **Fallback Mode:** LLM unavailable, using simple keyword extraction\n\n"

            # Add helpful next steps
            output += "---\n**Next Steps:**\n"
            if result['search_mode'] == 'keyword':
                output += f"• Use `search_docs` with query: \"{result['suggested_query']}\"\n"
            elif result['search_mode'] == 'semantic':
                output += f"• Use `semantic_search` with query: \"{result['suggested_query']}\"\n"
            elif result['search_mode'] == 'hybrid':
                output += f"• Use `hybrid_search` with query: \"{result['suggested_query']}\"\n"

            if result.get('facet_filters'):
                output += f"• Use `faceted_search` with filters: {result['facet_filters']}\n"

            return [TextContent(type="text", text=output)]
        except ValueError as e:
            return [TextContent(type="text", text=f"Translation error: {str(e)}\n\nMake sure LLM_PROVIDER and API key are configured.")]
        except Exception as e:
            return [TextContent(type="text", text=f"Error translating query: {str(e)}")]

    elif name == "compare_documents":
        doc_id_1 = arguments.get("doc_id_1")
        doc_id_2 = arguments.get("doc_id_2")
        comparison_type = arguments.get("comparison_type", "full")

        if not doc_id_1 or not doc_id_2:
            return [TextContent(type="text", text="Error: Both doc_id_1 and doc_id_2 are required")]

        try:
            result = kb.compare_documents(doc_id_1, doc_id_2, comparison_type)

            # Build formatted output
            output = "**Document Comparison**\n\n"
            output += f"**Similarity Score:** {result['similarity_score']:.1%}\n"
            output += f"**Summary:** {result['summary']}\n\n"

            # Metadata differences
            output += "**Metadata Comparison:**\n"
            md = result['metadata_diff']
            output += "- **Titles:** \n"
            output += f"  - Doc 1: {md['title'][0]}\n"
            output += f"  - Doc 2: {md['title'][1]}\n"
            output += f"- **Files:** {md['filename'][0]} vs {md['filename'][1]}\n"
            output += f"- **Types:** {md['file_type'][0]} vs {md['file_type'][1]}\n"
            output += f"- **Chunks:** {result['chunk_count'][0]} vs {result['chunk_count'][1]}\n\n"

            # Tags comparison
            tags = md['tags']
            if tags['common']:
                output += f"**Common Tags ({len(tags['common'])}):** {', '.join(tags['common'])}\n"
            if tags['only_in_doc1']:
                output += f"**Only in Doc 1 ({len(tags['only_in_doc1'])}):** {', '.join(tags['only_in_doc1'])}\n"
            if tags['only_in_doc2']:
                output += f"**Only in Doc 2 ({len(tags['only_in_doc2'])}):** {', '.join(tags['only_in_doc2'])}\n"
            output += "\n"

            # Entity comparison
            ec = result['entity_comparison']
            if ec['total_doc1'] > 0 or ec['total_doc2'] > 0:
                output += "**Entity Comparison:**\n"
                output += f"- Total in Doc 1: {ec['total_doc1']}\n"
                output += f"- Total in Doc 2: {ec['total_doc2']}\n"
                output += f"- Common Entities: {len(ec['common_entities'])}\n"

                if ec['common_entities'][:5]:  # Show first 5
                    output += "\n**Top Common Entities:**\n"
                    for ent in ec['common_entities'][:5]:
                        output += f"  - **{ent['text']}** ({ent['type']})\n"

                if ec['unique_to_doc1'][:3]:
                    output += "\n**Sample Unique to Doc 1:**\n"
                    for ent in ec['unique_to_doc1'][:3]:
                        output += f"  - **{ent['text']}** ({ent['type']})\n"

                if ec['unique_to_doc2'][:3]:
                    output += "\n**Sample Unique to Doc 2:**\n"
                    for ent in ec['unique_to_doc2'][:3]:
                        output += f"  - **{ent['text']}** ({ent['type']})\n"

            # Content diff preview
            if result['content_diff']:
                output += f"\n**Content Diff Preview:** (First {min(len(result['content_diff']), 20)} lines)\n"
                output += "```diff\n"
                for line in result['content_diff'][:20]:
                    output += line + "\n"
                output += "```\n"
                if len(result['content_diff']) > 20:
                    output += f"... {len(result['content_diff']) - 20} more lines\n"

            return [TextContent(type="text", text=output)]

        except ValueError as e:
            return [TextContent(type="text", text=f"Error: {str(e)}")]
        except Exception as e:
            return [TextContent(type="text", text=f"Error comparing documents: {str(e)}")]

    elif name == "export_entities":
        format = arguments.get("format", "csv")
        entity_types = arguments.get("entity_types")
        min_confidence = arguments.get("min_confidence", 0.0)
        output_path = arguments.get("output_path")

        try:
            result = kb.export_entities(
                format=format,
                entity_types=entity_types,
                min_confidence=min_confidence,
                output_path=output_path
            )

            # Count entities
            if format.lower() == 'csv':
                entity_count = result.count('\n') - 1  # Subtract header
            else:  # json
                import json
                entity_count = len(json.loads(result))

            output = "**Entity Export Complete**\n\n"
            output += f"**Format:** {format.upper()}\n"
            output += f"**Entities Exported:** {entity_count}\n"
            output += f"**Min Confidence:** {min_confidence:.2f}\n"

            if entity_types:
                output += f"**Filtered Types:** {', '.join(entity_types)}\n"

            if output_path:
                output += f"**Saved to:** `{output_path}`\n\n"
            else:
                output += "\n**Preview (first 500 chars):**\n```\n"
                output += result[:500]
                if len(result) > 500:
                    output += "\n... (truncated)"
                output += "\n```\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error exporting entities: {str(e)}")]

    elif name == "export_relationships":
        format = arguments.get("format", "csv")
        min_strength = arguments.get("min_strength", 0.0)
        entity_types = arguments.get("entity_types")
        output_path = arguments.get("output_path")

        try:
            result = kb.export_relationships(
                format=format,
                min_strength=min_strength,
                entity_types=entity_types,
                output_path=output_path
            )

            # Count relationships
            if format.lower() == 'csv':
                rel_count = result.count('\n') - 1  # Subtract header
            else:  # json
                import json
                rel_count = len(json.loads(result))

            output = "**Relationship Export Complete**\n\n"
            output += f"**Format:** {format.upper()}\n"
            output += f"**Relationships Exported:** {rel_count}\n"
            output += f"**Min Strength:** {min_strength:.2f}\n"

            if entity_types:
                output += f"**Filtered Types:** {', '.join(entity_types)}\n"

            if output_path:
                output += f"**Saved to:** `{output_path}`\n\n"
            else:
                output += "\n**Preview (first 500 chars):**\n```\n"
                output += result[:500]
                if len(result) > 500:
                    output += "\n... (truncated)"
                output += "\n```\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error exporting relationships: {str(e)}")]

    elif name == "queue_entity_extraction":
        doc_id = arguments.get("doc_id")
        confidence_threshold = arguments.get("confidence_threshold", 0.6)
        skip_if_exists = arguments.get("skip_if_exists", True)

        if not doc_id:
            return [TextContent(type="text", text="Error: doc_id is required")]

        try:
            result = kb.queue_entity_extraction(
                doc_id=doc_id,
                confidence_threshold=confidence_threshold,
                skip_if_exists=skip_if_exists
            )

            if result.get('queued'):
                output = "**Entity Extraction Queued**\n\n"
                output += f"**Job ID:** {result['job_id']}\n"
                output += f"**Document ID:** {doc_id}\n"
                output += f"**Confidence Threshold:** {confidence_threshold:.1%}\n\n"
                output += "✅ Extraction job has been queued and will run in the background.\n"
                output += "Use `get_extraction_status` to check progress.\n"
            else:
                output = "**Entity Extraction Not Queued**\n\n"
                output += f"**Reason:** {result.get('reason', 'Unknown')}\n"
                if 'existing_job_id' in result:
                    output += f"**Existing Job ID:** {result['existing_job_id']}\n"
                if 'existing_entities' in result:
                    output += f"**Existing Entities:** {result['existing_entities']}\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error queuing extraction: {str(e)}")]

    elif name == "get_extraction_status":
        doc_id = arguments.get("doc_id")

        if not doc_id:
            return [TextContent(type="text", text="Error: doc_id is required")]

        try:
            status = kb.get_extraction_status(doc_id)

            output = f"**Entity Extraction Status for Document: {doc_id}**\n\n"
            output += f"**Has Entities:** {'✅ Yes' if status['has_entities'] else '❌ No'}\n"
            output += f"**Entity Count:** {status['entity_count']}\n\n"

            if status['jobs']:
                output += "**Extraction Jobs:**\n\n"
                for job in status['jobs']:
                    output += f"- **Job {job['job_id']}**\n"
                    output += f"  - Status: {job['status'].upper()}\n"
                    output += f"  - Confidence: {job['confidence_threshold']:.1%}\n"
                    output += f"  - Queued: {job['queued_at']}\n"
                    if job['started_at']:
                        output += f"  - Started: {job['started_at']}\n"
                    if job['completed_at']:
                        output += f"  - Completed: {job['completed_at']}\n"
                    if job['entities_extracted']:
                        output += f"  - Entities Extracted: {job['entities_extracted']}\n"
                    if job['error_message']:
                        output += f"  - Error: {job['error_message']}\n"
                    output += "\n"
            else:
                output += "**No extraction jobs found for this document.**\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error getting extraction status: {str(e)}")]

    elif name == "get_extraction_jobs":
        status_filter = arguments.get("status_filter")
        limit = arguments.get("limit", 100)

        try:
            jobs = kb.get_all_extraction_jobs(
                status_filter=status_filter,
                limit=limit
            )

            output = "**Entity Extraction Jobs**\n\n"
            if status_filter:
                output += f"**Filter:** {status_filter.upper()}\n"
            output += f"**Total Jobs:** {len(jobs)}\n\n"

            if jobs:
                # Group by status
                by_status = {}
                for job in jobs:
                    status = job['status']
                    if status not in by_status:
                        by_status[status] = []
                    by_status[status].append(job)

                # Show summary
                output += "**Summary:**\n"
                for status, job_list in sorted(by_status.items()):
                    output += f"- {status.upper()}: {len(job_list)}\n"
                output += "\n"

                # Show recent jobs (limit to 10 for readability)
                output += f"**Recent Jobs (showing {min(len(jobs), 10)} of {len(jobs)}):**\n\n"
                for job in jobs[:10]:
                    output += f"- **Job {job['job_id']}** ({job['status'].upper()})\n"
                    output += f"  - Document: {job['doc_title'][:50]}...\n"
                    output += f"  - Doc ID: {job['doc_id']}\n"
                    output += f"  - Queued: {job['queued_at']}\n"
                    if job['entities_extracted']:
                        output += f"  - Entities: {job['entities_extracted']}\n"
                    if job['error_message']:
                        output += f"  - Error: {job['error_message'][:100]}...\n"
                    output += "\n"
            else:
                output += "**No jobs found.**\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error getting extraction jobs: {str(e)}")]

    # ============================================================
    # Knowledge Graph Tool Handlers (v2.24.0 - Phase 1, Task 1.4)
    # ============================================================

    elif name == "build_knowledge_graph":
        entity_types = arguments.get("entity_types")
        min_occurrences = arguments.get("min_occurrences", 2)
        min_relationship_strength = arguments.get("min_relationship_strength", 0.3)
        use_cache = arguments.get("use_cache", True)

        try:
            import networkx as nx

            G = kb.build_knowledge_graph(
                entity_types=entity_types,
                min_occurrences=min_occurrences,
                min_relationship_strength=min_relationship_strength,
                use_cache=use_cache
            )

            # Build comprehensive output
            output = "# Knowledge Graph Built\n\n"
            output += f"**Graph Statistics:**\n"
            output += f"- Nodes (entities): {G.number_of_nodes()}\n"
            output += f"- Edges (relationships): {G.number_of_edges()}\n"

            if G.number_of_nodes() > 0:
                density = nx.density(G)
                output += f"- Density: {density:.4f}\n"
                output += f"- Connected components: {nx.number_connected_components(G)}\n"

                # Show sample nodes
                sample_nodes = list(G.nodes())[:10]
                output += f"\n**Sample Entities (first 10):**\n"
                for node in sample_nodes:
                    node_data = G.nodes[node]
                    output += f"- {node} (type: {node_data.get('type', 'unknown')}, "
                    output += f"occurrences: {node_data.get('occurrences', 0)})\n"

                # Show sample edges
                if G.number_of_edges() > 0:
                    sample_edges = list(G.edges(data=True))[:10]
                    output += f"\n**Sample Relationships (first 10):**\n"
                    for e1, e2, data in sample_edges:
                        output += f"- {e1} ↔ {e2} (strength: {data.get('weight', 0):.3f})\n"
            else:
                output += "\n**No entities found matching the criteria.**\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error building knowledge graph: {str(e)}")]

    elif name == "compute_graph_metrics":
        entity_types = arguments.get("entity_types")
        min_occurrences = arguments.get("min_occurrences", 2)
        min_relationship_strength = arguments.get("min_relationship_strength", 0.3)
        store_results = arguments.get("store_results", True)

        try:
            metrics = kb.compute_graph_metrics(
                G=None,  # Will build new graph
                entity_types=entity_types,
                min_occurrences=min_occurrences,
                min_relationship_strength=min_relationship_strength,
                store_results=store_results
            )

            # Build output
            output = "# Graph Metrics Computed\n\n"

            # Graph statistics
            output += f"**Graph Statistics:**\n"
            stats = metrics['graph_stats']
            output += f"- Nodes: {stats['nodes']}\n"
            output += f"- Edges: {stats['edges']}\n"
            output += f"- Density: {stats['density']:.4f}\n"
            output += f"- Connected components: {stats['connected_components']}\n"
            output += f"- Communities detected: {metrics['num_communities']}\n"
            output += f"- Computed: {stats['computed_at']}\n"

            # Top entities by PageRank
            if metrics['pagerank']:
                output += f"\n**Top 15 Entities by PageRank:**\n"
                top_pr = sorted(metrics['pagerank'].items(), key=lambda x: x[1], reverse=True)[:15]
                for rank, (entity, score) in enumerate(top_pr, 1):
                    community = metrics['communities'].get(entity, 'N/A')
                    output += f"{rank:2d}. {entity:30s} (PR: {score:.6f}, Community: {community})\n"

            # Top entities by Betweenness
            if metrics['betweenness']:
                output += f"\n**Top 10 Bridge Entities (Betweenness):**\n"
                top_bt = sorted(metrics['betweenness'].items(), key=lambda x: x[1], reverse=True)[:10]
                for rank, (entity, score) in enumerate(top_bt, 1):
                    output += f"{rank:2d}. {entity:30s} (Betweenness: {score:.6f})\n"

            if store_results:
                output += f"\n**Metrics stored to database for {len(metrics['pagerank'])} entities.**\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error computing graph metrics: {str(e)}")]

    elif name == "get_entity_metrics":
        entity_text = arguments.get("entity_text")
        metric_types = arguments.get("metric_types")

        if not entity_text:
            return [TextContent(type="text", text="Error: entity_text is required")]

        try:
            result = kb.get_entity_metrics(entity_text, metric_types)

            if not result['found']:
                return [TextContent(type="text", text=f"No metrics found for entity: '{entity_text}'\n\nThis entity may not be in the knowledge graph, or metrics have not been computed yet. Run 'compute_graph_metrics' first.")]

            # Build output
            output = f"# Entity Metrics: {entity_text}\n\n"
            output += f"**Entity Type:** {result['entity_type']}\n"
            output += f"**Computed:** {result['computed_date']}\n\n"

            output += f"**Metrics:**\n"
            metrics = result['metrics']

            if 'pagerank' in metrics and metrics['pagerank'] is not None:
                output += f"- PageRank: {metrics['pagerank']:.6f} (importance score)\n"

            if 'betweenness' in metrics and metrics['betweenness'] is not None:
                output += f"- Betweenness: {metrics['betweenness']:.6f} (bridge score)\n"

            if 'degree' in metrics and metrics['degree'] is not None:
                output += f"- Degree: {metrics['degree']:.6f} (connection score)\n"

            if 'community' in metrics and metrics['community'] is not None:
                output += f"- Community ID: {metrics['community']}\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error retrieving entity metrics: {str(e)}")]

    elif name == "find_entity_path":
        entity1 = arguments.get("entity1")
        entity2 = arguments.get("entity2")
        max_path_length = arguments.get("max_path_length", 6)
        store_result = arguments.get("store_result", True)

        if not entity1 or not entity2:
            return [TextContent(type="text", text="Error: both entity1 and entity2 are required")]

        try:
            result = kb.find_shortest_path(
                entity1=entity1,
                entity2=entity2,
                G=None,  # Will build graph
                max_path_length=max_path_length,
                store_result=store_result
            )

            if result is None:
                return [TextContent(type="text", text=f"Error: Could not compute path. One or both entities may not exist in the graph.")]

            # Build output
            output = f"# Path: {entity1} → {entity2}\n\n"

            if result['exists']:
                output += f"**Path Found:** {result['length']} edges\n\n"
                output += f"**Path:**\n"
                output += " → ".join(result['path']) + "\n\n"

                if result['relationships']:
                    output += f"**Relationship Details:**\n"
                    for i, rel in enumerate(result['relationships'], 1):
                        output += f"{i}. {rel['from']} → {rel['to']}\n"
                        output += f"   - Strength: {rel['weight']:.3f}\n"
                        output += f"   - Co-occurrences: {rel['doc_count']} documents\n"

                if store_result:
                    output += f"\n**Path stored to database.**\n"
            else:
                output += f"**No path found.**\n\n"
                output += f"The entities '{entity1}' and '{entity2}' are in different connected components "
                output += f"of the knowledge graph (no relationship path exists between them).\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error finding path: {str(e)}")]

    elif name == "get_entity_community":
        entity_text = arguments.get("entity_text")
        max_members = arguments.get("max_members", 50)

        if not entity_text:
            return [TextContent(type="text", text="Error: entity_text is required")]

        try:
            # First, get the entity's community ID
            entity_metrics = kb.get_entity_metrics(entity_text)

            if not entity_metrics['found']:
                return [TextContent(type="text", text=f"Entity '{entity_text}' not found in graph metrics.\n\nRun 'compute_graph_metrics' first to detect communities.")]

            if 'community' not in entity_metrics['metrics'] or entity_metrics['metrics']['community'] is None:
                return [TextContent(type="text", text=f"No community information for '{entity_text}'.\n\nRun 'compute_graph_metrics' first.")]

            community_id = entity_metrics['metrics']['community']

            # Get all entities in the same community
            cursor = kb.db_conn.cursor()
            rows = cursor.execute("""
                SELECT entity_text, entity_type, pagerank, degree_centrality
                FROM graph_metrics
                WHERE community_id = ?
                ORDER BY pagerank DESC
                LIMIT ?
            """, (community_id, max_members)).fetchall()

            # Build output
            output = f"# Community {community_id}\n\n"
            output += f"**Anchor Entity:** {entity_text} ({entity_metrics['entity_type']})\n"
            output += f"**Total Members:** {len(rows)}\n\n"

            output += f"**Community Members (ranked by PageRank):**\n"
            for i, (ent_text, ent_type, pr, degree) in enumerate(rows, 1):
                marker = " ← anchor" if ent_text == entity_text else ""
                output += f"{i:2d}. {ent_text:30s} (type: {ent_type}, PR: {pr:.6f}){marker}\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error getting community: {str(e)}")]

    elif name == "get_top_entities":
        metric = arguments.get("metric", "pagerank")
        limit = arguments.get("limit", 10)
        entity_types = arguments.get("entity_types")

        try:
            # Build query based on metric
            metric_column_map = {
                'pagerank': 'pagerank',
                'betweenness': 'betweenness_centrality',
                'degree': 'degree_centrality'
            }

            metric_column = metric_column_map.get(metric, 'pagerank')

            cursor = kb.db_conn.cursor()

            # Build query
            query = f"""
                SELECT entity_text, entity_type, pagerank, betweenness_centrality,
                       degree_centrality, community_id
                FROM graph_metrics
                WHERE {metric_column} IS NOT NULL
            """
            params = []

            if entity_types:
                placeholders = ','.join('?' * len(entity_types))
                query += f" AND entity_type IN ({placeholders})"
                params.extend(entity_types)

            query += f" ORDER BY {metric_column} DESC LIMIT ?"
            params.append(limit)

            rows = cursor.execute(query, params).fetchall()

            if not rows:
                return [TextContent(type="text", text=f"No entities found.\n\nRun 'compute_graph_metrics' first to generate metrics.")]

            # Build output
            metric_display = metric.replace('_', ' ').title()
            output = f"# Top {len(rows)} Entities by {metric_display}\n\n"

            if entity_types:
                output += f"**Filtered by types:** {', '.join(entity_types)}\n\n"

            output += f"**Ranking:**\n"
            for rank, (ent_text, ent_type, pr, betw, deg, comm) in enumerate(rows, 1):
                output += f"{rank:2d}. {ent_text:30s}\n"
                output += f"    Type: {ent_type}, Community: {comm}\n"
                output += f"    PageRank: {pr:.6f}, Betweenness: {betw:.6f}, Degree: {deg:.6f}\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error getting top entities: {str(e)}")]

    elif name == "visualize_graph":
        output_path = arguments.get("output_path", "knowledge_graph.html")
        entity_types = arguments.get("entity_types")
        min_occurrences = arguments.get("min_occurrences", 2)
        min_relationship_strength = arguments.get("min_relationship_strength", 0.3)
        color_by = arguments.get("color_by", "entity_type")
        size_by = arguments.get("size_by", "pagerank")
        physics_enabled = arguments.get("physics_enabled", True)
        height = arguments.get("height", "800px")
        width = arguments.get("width", "100%")

        try:
            # Generate visualization
            saved_path = kb.visualize_knowledge_graph(
                G=None,  # Will build graph
                output_path=output_path,
                entity_types=entity_types,
                min_occurrences=min_occurrences,
                min_relationship_strength=min_relationship_strength,
                color_by=color_by,
                size_by=size_by,
                physics_enabled=physics_enabled,
                height=height,
                width=width
            )

            if not saved_path:
                return [TextContent(type="text", text="Error: Could not generate visualization (empty graph or error occurred).")]

            # Build output message
            output = "# Knowledge Graph Visualization Generated\n\n"
            output += f"**File saved to:** `{saved_path}`\n\n"

            output += f"**Visualization Settings:**\n"
            output += f"- Nodes colored by: {color_by}\n"
            output += f"- Node size based on: {size_by}\n"
            output += f"- Physics simulation: {'Enabled' if physics_enabled else 'Disabled'}\n"

            if entity_types:
                output += f"- Filtered to types: {', '.join(entity_types)}\n"

            output += f"- Min occurrences: {min_occurrences}\n"
            output += f"- Min relationship strength: {min_relationship_strength}\n\n"

            output += f"**Interactive Features:**\n"
            output += f"- Hover over nodes to see entity details and metrics\n"
            output += f"- Hover over edges to see relationship strength\n"
            output += f"- Click and drag nodes to rearrange\n"
            output += f"- Zoom with mouse wheel\n"
            output += f"- Pan by dragging the background\n\n"

            output += f"Open the HTML file in your web browser to explore the interactive visualization!\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error generating visualization: {str(e)}")]

    elif name == "train_lda_topics":
        num_topics = arguments.get("num_topics", 10)
        max_iter = arguments.get("max_iter", 100)
        max_features = arguments.get("max_features", 1000)

        try:
            results = kb.train_lda_model(
                num_topics=num_topics,
                max_iter=max_iter,
                max_features=max_features,
                store_results=True
            )

            output = f"# LDA Topic Model Trained\n\n"
            output += f"**Model Statistics:**\n"
            output += f"- Number of topics: {results['num_topics']}\n"
            output += f"- Documents processed: {results['num_documents']}\n"
            output += f"- Vocabulary size: {results['vocabulary_size']}\n"
            output += f"- Perplexity: {results['perplexity']:.2f}\n"
            output += f"- Training time: {results['training_time']:.2f}s\n\n"

            output += f"**Top Topics:**\n\n"
            for i, topic in enumerate(results['topics'][:5], 1):
                output += f"{i}. **Topic {topic['topic_number']}:** {', '.join(topic['top_words'][:5])}\n"

            output += f"\nResults stored to database. Use `get_document_topics` to see document assignments.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error training LDA model: {str(e)}")]

    elif name == "train_nmf_topics":
        num_topics = arguments.get("num_topics", 10)
        max_iter = arguments.get("max_iter", 200)
        max_features = arguments.get("max_features", 1000)

        try:
            results = kb.train_nmf_model(
                num_topics=num_topics,
                max_iter=max_iter,
                max_features=max_features,
                store_results=True
            )

            output = f"# NMF Topic Model Trained\n\n"
            output += f"**Model Statistics:**\n"
            output += f"- Number of topics: {results['num_topics']}\n"
            output += f"- Documents processed: {results['num_documents']}\n"
            output += f"- Vocabulary size: {results['vocabulary_size']}\n"
            output += f"- Reconstruction error: {results['reconstruction_error']:.2f}\n"
            output += f"- Training time: {results['training_time']:.2f}s\n\n"

            output += f"**Top Topics:**\n\n"
            for i, topic in enumerate(results['topics'][:5], 1):
                output += f"{i}. **Topic {topic['topic_number']}:** {', '.join(topic['top_words'][:5])}\n"

            output += f"\nResults stored to database. Use `get_document_topics` to see document assignments.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error training NMF model: {str(e)}")]

    elif name == "train_bertopic":
        num_topics = arguments.get("num_topics", 10)
        min_topic_size = arguments.get("min_topic_size", 5)

        try:
            results = kb.train_bertopic_model(
                num_topics=num_topics,
                min_topic_size=min_topic_size,
                store_results=True
            )

            output = f"# BERTopic Model Trained\n\n"
            output += f"**Model Statistics:**\n"
            output += f"- Number of topics: {results['num_topics']}\n"
            output += f"- Documents processed: {results['num_documents']}\n"
            output += f"- Outliers: {results['outliers']}\n"
            output += f"- Training time: {results['training_time']:.2f}s\n\n"

            output += f"**Top Topics:**\n\n"
            for i, topic in enumerate(results['topics'][:5], 1):
                output += f"{i}. **Topic {topic['topic_number']}:** {', '.join(topic['top_words'][:5])}\n"

            output += f"\nResults stored to database. Use `get_document_topics` to see document assignments.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error training BERTopic model: {str(e)}")]

    elif name == "get_document_topics":
        model_type = arguments.get("model_type")
        doc_id = arguments.get("doc_id")
        min_probability = arguments.get("min_probability", 0.1)

        try:
            cursor = kb.db_conn.cursor()

            if doc_id:
                # Get topics for specific document
                assignments = cursor.execute("""
                    SELECT dt.topic_id, t.topic_number, t.top_words, dt.probability
                    FROM document_topics dt
                    JOIN topics t ON dt.topic_id = t.topic_id
                    WHERE dt.doc_id = ? AND dt.model_type = ? AND dt.probability >= ?
                    ORDER BY dt.probability DESC
                """, (doc_id, model_type, min_probability)).fetchall()

                if not assignments:
                    return [TextContent(type="text", text=f"No topic assignments found for document {doc_id} with model {model_type}")]

                doc = kb.documents.get(doc_id)
                doc_title = doc.title if doc else "Unknown"

                output = f"# Topic Assignments for: {doc_title}\n\n"
                output += f"**Model:** {model_type}\n\n"

                for topic_id, topic_num, top_words_json, probability in assignments:
                    import json
                    top_words = json.loads(top_words_json)
                    output += f"- **Topic {topic_num}** (prob: {probability:.3f}): {', '.join(top_words[:5])}\n"

            else:
                # Get summary of all documents
                stats = cursor.execute("""
                    SELECT COUNT(DISTINCT doc_id), COUNT(*)
                    FROM document_topics
                    WHERE model_type = ?
                """, (model_type,)).fetchone()

                if not stats or stats[0] == 0:
                    return [TextContent(type="text", text=f"No topic assignments found for model {model_type}")]

                num_docs, num_assignments = stats

                output = f"# Topic Assignments Summary\n\n"
                output += f"**Model:** {model_type}\n"
                output += f"**Documents:** {num_docs}\n"
                output += f"**Total assignments:** {num_assignments}\n\n"

                # Get topic distribution
                topic_dist = cursor.execute("""
                    SELECT t.topic_number, t.top_words, COUNT(DISTINCT dt.doc_id) as doc_count
                    FROM topics t
                    LEFT JOIN document_topics dt ON t.topic_id = dt.topic_id
                    WHERE t.model_type = ?
                    GROUP BY t.topic_number
                    ORDER BY doc_count DESC
                """, (model_type,)).fetchall()

                output += f"**Topic Distribution:**\n\n"
                for topic_num, top_words_json, doc_count in topic_dist:
                    import json
                    top_words = json.loads(top_words_json)
                    output += f"- **Topic {topic_num}** ({doc_count} docs): {', '.join(top_words[:5])}\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error getting document topics: {str(e)}")]

    elif name == "cluster_documents_kmeans":
        num_clusters = arguments.get("num_clusters", 10)

        try:
            results = kb.cluster_documents_kmeans(
                num_clusters=num_clusters,
                store_results=True
            )

            output = f"# K-Means Clustering Complete\n\n"
            output += f"**Clustering Statistics:**\n"
            output += f"- Number of clusters: {results['num_clusters']}\n"
            output += f"- Documents clustered: {results['num_documents']}\n"
            output += f"- Silhouette score: {results['silhouette_score']:.3f}\n"
            output += f"- Training time: {results['training_time']:.2f}s\n\n"

            output += f"**Cluster Sizes:**\n\n"
            for i, cluster in enumerate(results['clusters'][:10], 1):
                output += f"{i}. **Cluster {cluster['cluster_number']}** ({cluster['num_documents']} docs): {', '.join(cluster['top_terms'][:5])}\n"

            output += f"\nResults stored to database. Use `get_cluster_documents` to see cluster contents.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error clustering documents: {str(e)}")]

    elif name == "cluster_documents_dbscan":
        eps = arguments.get("eps", 0.5)
        min_samples = arguments.get("min_samples", 3)

        try:
            results = kb.cluster_documents_dbscan(
                eps=eps,
                min_samples=min_samples,
                store_results=True
            )

            output = f"# DBSCAN Clustering Complete\n\n"
            output += f"**Clustering Statistics:**\n"
            output += f"- Number of clusters: {results['num_clusters']}\n"
            output += f"- Documents clustered: {results['num_documents']}\n"
            output += f"- Outliers: {results['num_outliers']}\n"
            output += f"- Training time: {results['training_time']:.2f}s\n\n"

            if results['clusters']:
                output += f"**Cluster Sizes:**\n\n"
                for i, cluster in enumerate(results['clusters'][:10], 1):
                    output += f"{i}. **Cluster {cluster['cluster_number']}** ({cluster['num_documents']} docs): {', '.join(cluster['top_terms'][:5])}\n"

            output += f"\nResults stored to database. Use `get_cluster_documents` to see cluster contents.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error clustering documents: {str(e)}")]

    elif name == "cluster_documents_hdbscan":
        min_cluster_size = arguments.get("min_cluster_size", 5)
        min_samples = arguments.get("min_samples", 3)

        try:
            results = kb.cluster_documents_hdbscan(
                min_cluster_size=min_cluster_size,
                min_samples=min_samples,
                store_results=True
            )

            output = f"# HDBSCAN Clustering Complete\n\n"
            output += f"**Clustering Statistics:**\n"
            output += f"- Number of clusters: {results['num_clusters']}\n"
            output += f"- Documents clustered: {results['num_documents']}\n"
            output += f"- Outliers: {results['num_outliers']}\n"
            output += f"- Training time: {results['training_time']:.2f}s\n\n"

            if results['clusters']:
                output += f"**Cluster Sizes:**\n\n"
                for i, cluster in enumerate(results['clusters'][:10], 1):
                    output += f"{i}. **Cluster {cluster['cluster_number']}** ({cluster['num_documents']} docs): {', '.join(cluster['top_terms'][:5])}\n"

            output += f"\nResults stored to database. Use `get_cluster_documents` to see cluster contents.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error clustering documents: {str(e)}")]

    elif name == "get_cluster_documents":
        algorithm = arguments.get("algorithm")
        cluster_number = arguments.get("cluster_number")

        try:
            cursor = kb.db_conn.cursor()

            if cluster_number is not None:
                # Get documents in specific cluster
                cluster_id = f"{algorithm}_cluster_{cluster_number}"

                # Get cluster info
                cluster_info = cursor.execute("""
                    SELECT num_documents, top_terms, silhouette_score
                    FROM clusters
                    WHERE cluster_id = ?
                """, (cluster_id,)).fetchone()

                if not cluster_info:
                    return [TextContent(type="text", text=f"Cluster {cluster_number} not found for algorithm {algorithm}")]

                num_docs, top_terms_json, silhouette = cluster_info
                import json
                top_terms = json.loads(top_terms_json)

                # Get documents
                docs = cursor.execute("""
                    SELECT dc.doc_id, dc.distance
                    FROM document_clusters dc
                    WHERE dc.cluster_id = ?
                    ORDER BY dc.distance ASC
                    LIMIT 20
                """, (cluster_id,)).fetchall()

                output = f"# Cluster {cluster_number} ({algorithm})\n\n"
                output += f"**Cluster Info:**\n"
                output += f"- Documents: {num_docs}\n"
                output += f"- Top terms: {', '.join(top_terms[:10])}\n"
                if silhouette:
                    output += f"- Silhouette score: {silhouette:.3f}\n"
                output += f"\n**Documents (closest to centroid):**\n\n"

                for doc_id, distance in docs:
                    doc = kb.documents.get(doc_id)
                    if doc:
                        output += f"- {doc.title[:60]} (distance: {distance:.3f})\n"

            else:
                # Get summary of all clusters
                clusters = cursor.execute("""
                    SELECT cluster_number, num_documents, top_terms
                    FROM clusters
                    WHERE algorithm = ?
                    ORDER BY num_documents DESC
                """, (algorithm,)).fetchall()

                if not clusters:
                    return [TextContent(type="text", text=f"No clusters found for algorithm {algorithm}")]

                output = f"# Clusters ({algorithm})\n\n"
                output += f"**Total clusters:** {len(clusters)}\n\n"

                for cluster_num, num_docs, top_terms_json in clusters[:20]:
                    import json
                    top_terms = json.loads(top_terms_json)
                    output += f"- **Cluster {cluster_num}** ({num_docs} docs): {', '.join(top_terms[:5])}\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error getting cluster documents: {str(e)}")]

    # Phase 3: Temporal Analysis Tools
    elif name == "extract_document_events":
        doc_id = arguments.get("doc_id")
        min_confidence = arguments.get("min_confidence", 0.5)

        try:
            result = kb.extract_document_events(doc_id, min_confidence=min_confidence)

            output = f"# Event Extraction Complete\n\n"
            output += f"**Document:** {result['title']}\n"
            output += f"**Statistics:**\n"
            output += f"- Total events detected: {result['event_count']}\n"
            output += f"- Events with confidence >= {min_confidence}: {result['filtered_count']}\n"
            output += f"- Events stored to database: {result['stored_count']}\n\n"

            if result['events']:
                output += f"**Extracted Events:**\n\n"
                for i, event in enumerate(result['events'][:10], 1):
                    date_str = event['date_info']['text'] if event['date_info'] else 'No date'
                    output += f"{i}. **[{event['type']}]** {event['title'][:80]}...\n"
                    output += f"   - Date: {date_str}\n"
                    output += f"   - Confidence: {event['confidence']:.2f}\n"
                    if event['entities']:
                        output += f"   - Entities: {', '.join(event['entities'][:3])}\n"
                    output += "\n"

                if len(result['events']) > 10:
                    output += f"... and {len(result['events']) - 10} more events\n\n"

            output += f"Events stored to database. Use `get_timeline` or `search_events_by_date` to query.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error extracting events: {str(e)}")]

    elif name == "get_timeline":
        start_year = arguments.get("start_year")
        end_year = arguments.get("end_year")
        category = arguments.get("category")
        min_importance = arguments.get("min_importance", 1)
        limit = arguments.get("limit")

        try:
            timeline = kb.get_timeline(
                start_year=start_year,
                end_year=end_year,
                category=category,
                min_importance=min_importance,
                limit=limit
            )

            if not timeline:
                # Try to build timeline if empty
                kb.build_timeline()
                timeline = kb.get_timeline(
                    start_year=start_year,
                    end_year=end_year,
                    category=category,
                    min_importance=min_importance,
                    limit=limit
                )

            if not timeline:
                return [TextContent(type="text", text="No timeline entries found. Extract events from documents first using `extract_document_events`.")]

            output = f"# Timeline\n\n"

            # Add filter info
            filters = []
            if start_year:
                filters.append(f"from {start_year}")
            if end_year:
                filters.append(f"to {end_year}")
            if category:
                filters.append(f"category: {category}")
            if min_importance > 1:
                filters.append(f"importance >= {min_importance}")

            if filters:
                output += f"**Filters:** {', '.join(filters)}\n"

            output += f"**Total entries:** {len(timeline)}\n\n"
            output += f"**Timeline Entries:**\n\n"

            for entry in timeline[:50]:  # Limit to 50 entries
                importance_stars = "⭐" * entry['importance']
                output += f"**[{entry['display_date']}]** {entry['title'][:70]}...\n"
                output += f"  - Type: {entry['event_type']} | Importance: {importance_stars} ({entry['importance']}/5)\n"
                output += f"  - Confidence: {entry['confidence']:.2f}\n\n"

            if len(timeline) > 50:
                output += f"... and {len(timeline) - 50} more entries (use `limit` parameter to see more)\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error getting timeline: {str(e)}")]

    elif name == "search_events_by_date":
        start_year = arguments.get("start_year")
        end_year = arguments.get("end_year")
        event_type = arguments.get("event_type")
        min_confidence = arguments.get("min_confidence", 0.5)

        try:
            events = kb.search_events_by_date(
                start_year=start_year,
                end_year=end_year,
                event_type=event_type,
                min_confidence=min_confidence
            )

            if not events:
                return [TextContent(type="text", text="No events found matching the criteria.")]

            output = f"# Event Search Results\n\n"

            # Add search criteria
            criteria = []
            if start_year and end_year:
                criteria.append(f"{start_year}-{end_year}")
            elif start_year:
                criteria.append(f"from {start_year}")
            elif end_year:
                criteria.append(f"to {end_year}")
            if event_type:
                criteria.append(f"type: {event_type}")
            criteria.append(f"confidence >= {min_confidence}")

            output += f"**Search criteria:** {', '.join(criteria)}\n"
            output += f"**Results:** {len(events)} events\n\n"

            # Group events by year
            events_by_year = {}
            for event in events:
                year = event['year']
                if year not in events_by_year:
                    events_by_year[year] = []
                events_by_year[year].append(event)

            for year in sorted(events_by_year.keys()):
                year_events = events_by_year[year]
                output += f"## {year} ({len(year_events)} events)\n\n"

                for event in year_events[:20]:  # Limit per year
                    output += f"- **[{event['date_normalized']}]** {event['title'][:70]}...\n"
                    output += f"  Type: {event['event_type']}, Confidence: {event['confidence']:.2f}\n"

                if len(year_events) > 20:
                    output += f"  ... and {len(year_events) - 20} more events from {year}\n"

                output += "\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error searching events: {str(e)}")]

    elif name == "get_historical_context":
        year = arguments.get("year")
        context_years = arguments.get("context_years", 2)

        try:
            context = kb.get_historical_context(year, context_years=context_years)

            if context['total_events'] == 0:
                return [TextContent(type="text", text=f"No events found in the period {context['year_range'][0]}-{context['year_range'][1]}.")]

            output = f"# Historical Context for {year}\n\n"
            output += f"**Context range:** {context['year_range'][0]} - {context['year_range'][1]}\n"
            output += f"**Total events:** {context['total_events']}\n\n"

            # Show events by year
            for event_year in sorted(context['events_by_year'].keys()):
                year_events = context['events_by_year'][event_year]
                marker = "📍 **TARGET YEAR**" if event_year == year else ""

                output += f"## {event_year} ({len(year_events)} events) {marker}\n\n"

                for event in year_events[:10]:
                    output += f"- **[{event['event_type']}]** {event['title'][:70]}...\n"
                    output += f"  Date: {event['date_normalized']}, Confidence: {event['confidence']:.2f}\n"

                if len(year_events) > 10:
                    output += f"  ... and {len(year_events) - 10} more events\n"

                output += "\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error getting historical context: {str(e)}")]

    # Phase 1: Knowledge Graph Tools (v2.24.0)
    elif name == "build_knowledge_graph":
        try:
            import networkx as nx
        except ImportError:
            return [TextContent(type="text", text="NetworkX not installed. Run: pip install networkx>=3.0")]

        entity_types = arguments.get("entity_types")
        min_occurrences = arguments.get("min_occurrences", 2)
        min_relationship_strength = arguments.get("min_relationship_strength", 0.3)

        try:
            G = kb.build_knowledge_graph(
                entity_types=entity_types,
                min_occurrences=min_occurrences,
                min_relationship_strength=min_relationship_strength
            )

            output = f"# Knowledge Graph Built\n\n"
            output += f"**Nodes (entities):** {G.number_of_nodes()}\n"
            output += f"**Edges (relationships):** {G.number_of_edges()}\n"
            output += f"**Density:** {nx.density(G):.4f}\n"
            output += f"**Connected components:** {nx.number_connected_components(G)}\n\n"

            # Show largest connected component
            if G.number_of_nodes() > 0:
                largest_cc = max(nx.connected_components(G), key=len)
                output += f"**Largest component:** {len(largest_cc)} nodes\n\n"

            # Show sample nodes with highest degree
            degrees = dict(G.degree())
            if degrees:
                top_nodes = sorted(degrees.items(), key=lambda x: x[1], reverse=True)[:10]
                output += "## Top Connected Entities\n\n"
                for entity, degree in top_nodes:
                    node_type = G.nodes[entity].get('type', 'unknown')
                    occurrences = G.nodes[entity].get('occurrences', 0)
                    output += f"- **{entity}** (type: {node_type}, degree: {degree}, occurrences: {occurrences})\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error building knowledge graph: {str(e)}")]

    elif name == "analyze_graph_pagerank":
        try:
            import networkx as nx
        except ImportError:
            return [TextContent(type="text", text="NetworkX not installed. Run: pip install networkx>=3.0")]

        entity_types = arguments.get("entity_types")
        min_occurrences = arguments.get("min_occurrences", 2)
        top_n = arguments.get("top_n", 20)
        alpha = arguments.get("alpha", 0.85)

        try:
            # Build graph
            G = kb.build_knowledge_graph(
                entity_types=entity_types,
                min_occurrences=min_occurrences
            )

            if G.number_of_nodes() == 0:
                return [TextContent(type="text", text="No entities found for graph construction.")]

            # Calculate PageRank
            pagerank = kb.analyze_pagerank(G, alpha=alpha)

            # Format results
            output = f"# PageRank Analysis\n\n"
            output += f"**Total entities:** {len(pagerank)}\n"
            output += f"**Damping factor (alpha):** {alpha}\n\n"
            output += f"## Top {top_n} Entities by PageRank\n\n"

            for i, (entity, score) in enumerate(list(pagerank.items())[:top_n], 1):
                node_data = G.nodes[entity]
                output += f"{i}. **{entity}**\n"
                output += f"   - PageRank: {score:.6f}\n"
                output += f"   - Type: {node_data['type']}\n"
                output += f"   - Occurrences: {node_data['occurrences']}\n\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error analyzing PageRank: {str(e)}")]

    elif name == "detect_graph_communities":
        try:
            import networkx as nx
        except ImportError:
            return [TextContent(type="text", text="NetworkX not installed. Run: pip install networkx>=3.0")]

        entity_types = arguments.get("entity_types")
        min_occurrences = arguments.get("min_occurrences", 2)
        algorithm = arguments.get("algorithm", "louvain")

        try:
            # Build graph
            G = kb.build_knowledge_graph(
                entity_types=entity_types,
                min_occurrences=min_occurrences
            )

            if G.number_of_nodes() == 0:
                return [TextContent(type="text", text="No entities found for graph construction.")]

            # Detect communities
            communities = kb.detect_communities(G, algorithm=algorithm)
            num_communities = len(set(communities.values()))

            # Format results
            output = f"# Community Detection\n\n"
            output += f"**Algorithm:** {algorithm}\n"
            output += f"**Total entities:** {len(communities)}\n"
            output += f"**Communities detected:** {num_communities}\n\n"

            # Group entities by community
            community_groups = {}
            for entity, comm_id in communities.items():
                if comm_id not in community_groups:
                    community_groups[comm_id] = []
                community_groups[comm_id].append(entity)

            # Show communities sorted by size
            sorted_communities = sorted(community_groups.items(), key=lambda x: len(x[1]), reverse=True)

            output += "## Communities (by size)\n\n"
            for i, (comm_id, members) in enumerate(sorted_communities[:10], 1):
                output += f"### Community {comm_id} ({len(members)} members)\n\n"
                # Show first 10 members
                for entity in members[:10]:
                    node_type = G.nodes[entity].get('type', 'unknown')
                    output += f"- {entity} (type: {node_type})\n"
                if len(members) > 10:
                    output += f"- ... and {len(members) - 10} more\n"
                output += "\n"

            if len(sorted_communities) > 10:
                output += f"\n... and {len(sorted_communities) - 10} more communities\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error detecting communities: {str(e)}")]

    elif name == "calculate_graph_centrality":
        try:
            import networkx as nx
        except ImportError:
            return [TextContent(type="text", text="NetworkX not installed. Run: pip install networkx>=3.0")]

        entity_types = arguments.get("entity_types")
        min_occurrences = arguments.get("min_occurrences", 2)
        top_n = arguments.get("top_n", 10)

        try:
            # Build graph
            G = kb.build_knowledge_graph(
                entity_types=entity_types,
                min_occurrences=min_occurrences
            )

            if G.number_of_nodes() == 0:
                return [TextContent(type="text", text="No entities found for graph construction.")]

            # Calculate centrality
            centrality = kb.calculate_centrality(G)

            # Format results
            output = f"# Centrality Analysis\n\n"
            output += f"**Total entities:** {G.number_of_nodes()}\n\n"

            # Show top entities for each centrality measure
            for measure_name, measure_values in centrality.items():
                sorted_values = sorted(measure_values.items(), key=lambda x: x[1], reverse=True)

                output += f"## Top {top_n} by {measure_name.capitalize()} Centrality\n\n"
                output += f"*{measure_name.capitalize()} measures "
                if measure_name == 'betweenness':
                    output += "entities that bridge different parts of the graph*\n\n"
                elif measure_name == 'closeness':
                    output += "entities that are close to all other entities*\n\n"
                elif measure_name == 'degree':
                    output += "entities with many direct connections*\n\n"

                for i, (entity, score) in enumerate(sorted_values[:top_n], 1):
                    node_data = G.nodes[entity]
                    output += f"{i}. **{entity}**\n"
                    output += f"   - Score: {score:.6f}\n"
                    output += f"   - Type: {node_data['type']}\n\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error calculating centrality: {str(e)}")]

    elif name == "find_entity_path":
        try:
            import networkx as nx
        except ImportError:
            return [TextContent(type="text", text="NetworkX not installed. Run: pip install networkx>=3.0")]

        entity1 = arguments.get("entity1")
        entity2 = arguments.get("entity2")
        min_occurrences = arguments.get("min_occurrences", 2)

        try:
            # Build graph
            G = kb.build_knowledge_graph(min_occurrences=min_occurrences)

            if G.number_of_nodes() == 0:
                return [TextContent(type="text", text="No entities found for graph construction.")]

            # Find path
            path = kb.find_shortest_path(G, entity1, entity2)

            if path is None:
                # Check if entities exist in graph
                if entity1 not in G.nodes():
                    return [TextContent(type="text", text=f"Entity '{entity1}' not found in graph. Try lowering min_occurrences or check spelling.")]
                if entity2 not in G.nodes():
                    return [TextContent(type="text", text=f"Entity '{entity2}' not found in graph. Try lowering min_occurrences or check spelling.")]

                return [TextContent(type="text", text=f"No path found between '{entity1}' and '{entity2}'. They are in different connected components.")]

            # Format results
            output = f"# Entity Path: {entity1} → {entity2}\n\n"
            output += f"**Path length:** {len(path)} steps\n"
            output += f"**Path:** {' → '.join(path)}\n\n"

            # Show detailed path with relationship info
            output += "## Detailed Path\n\n"
            for i in range(len(path) - 1):
                curr_entity = path[i]
                next_entity = path[i + 1]

                curr_type = G.nodes[curr_entity].get('type', 'unknown')
                next_type = G.nodes[next_entity].get('type', 'unknown')

                if G.has_edge(curr_entity, next_entity):
                    edge_data = G.edges[curr_entity, next_entity]
                    weight = edge_data.get('weight', 0.0)
                    co_occurrences = edge_data.get('co_occurrences', 0)

                    output += f"{i + 1}. **{curr_entity}** (type: {curr_type})\n"
                    output += f"   ↓ *relationship strength: {weight:.2f}, co-occurrences: {co_occurrences}*\n"

            # Add final entity
            final_entity = path[-1]
            final_type = G.nodes[final_entity].get('type', 'unknown')
            output += f"{len(path)}. **{final_entity}** (type: {final_type})\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error finding entity path: {str(e)}")]

    elif name == "get_graph_statistics":
        try:
            import networkx as nx
        except ImportError:
            return [TextContent(type="text", text="NetworkX not installed. Run: pip install networkx>=3.0")]

        entity_types = arguments.get("entity_types")
        min_occurrences = arguments.get("min_occurrences", 2)

        try:
            # Build graph
            G = kb.build_knowledge_graph(
                entity_types=entity_types,
                min_occurrences=min_occurrences
            )

            if G.number_of_nodes() == 0:
                return [TextContent(type="text", text="No entities found for graph construction.")]

            # Calculate statistics
            density = nx.density(G)
            num_components = nx.number_connected_components(G)

            # Get degree distribution
            degrees = [d for _, d in G.degree()]
            avg_degree = sum(degrees) / len(degrees) if degrees else 0
            max_degree = max(degrees) if degrees else 0
            min_degree = min(degrees) if degrees else 0

            # Format results
            output = f"# Knowledge Graph Statistics\n\n"
            output += f"## Basic Metrics\n\n"
            output += f"- **Nodes (entities):** {G.number_of_nodes()}\n"
            output += f"- **Edges (relationships):** {G.number_of_edges()}\n"
            output += f"- **Density:** {density:.4f} (0 = sparse, 1 = complete)\n"
            output += f"- **Connected components:** {num_components}\n\n"

            # Show largest connected component
            if num_components > 0:
                components = sorted(nx.connected_components(G), key=len, reverse=True)
                output += f"## Connected Components\n\n"
                output += f"- Largest: {len(components[0])} nodes ({len(components[0])/G.number_of_nodes()*100:.1f}%)\n"
                if len(components) > 1:
                    output += f"- Second largest: {len(components[1])} nodes\n"
                output += f"- Isolated nodes: {sum(1 for c in components if len(c) == 1)}\n\n"

            # Degree statistics
            output += f"## Degree Distribution\n\n"
            output += f"- Average degree: {avg_degree:.2f}\n"
            output += f"- Maximum degree: {max_degree}\n"
            output += f"- Minimum degree: {min_degree}\n\n"

            # Entity type distribution
            type_counts = {}
            for node in G.nodes():
                node_type = G.nodes[node].get('type', 'unknown')
                type_counts[node_type] = type_counts.get(node_type, 0) + 1

            output += f"## Entity Types\n\n"
            for etype, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
                output += f"- {etype}: {count} ({count/G.number_of_nodes()*100:.1f}%)\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error getting graph statistics: {str(e)}")]

    # ========================================
    # Phase 2: Topic Modeling & Clustering
    # ========================================

    elif name == "train_lda_topics":
        num_topics = arguments.get("num_topics", 10)
        max_iter = arguments.get("max_iter", 100)
        random_state = arguments.get("random_state", 42)

        try:
            result = kb.train_lda_model(num_topics, max_iter, random_state)

            if 'error' in result:
                return [TextContent(type="text", text=f"Error: {result['error']}")]

            output = f"LDA Topic Modeling Complete\n{'='*60}\n\n"
            output += f"Model: {result['model_type']}\n"
            output += f"Topics: {result['num_topics']}\n"
            output += f"Documents: {result['num_documents']}\n"
            output += f"Assignments: {result['num_assignments']}\n"
            output += f"Perplexity: {result['perplexity']:.2f}\n\n"

            output += "Discovered Topics:\n\n"
            for topic in result['topics'][:5]:  # Show first 5 topics
                top_words = topic['words'][:5]
                output += f"Topic {topic['topic_number']}: {', '.join(top_words)}\n"

            if len(result['topics']) > 5:
                output += f"\n... and {len(result['topics']) - 5} more topics\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error training LDA model: {str(e)}")]

    elif name == "train_nmf_topics":
        num_topics = arguments.get("num_topics", 10)
        max_iter = arguments.get("max_iter", 200)
        random_state = arguments.get("random_state", 42)

        try:
            result = kb.train_nmf_model(num_topics, max_iter, random_state)

            if 'error' in result:
                return [TextContent(type="text", text=f"Error: {result['error']}")]

            output = f"NMF Topic Modeling Complete\n{'='*60}\n\n"
            output += f"Model: {result['model_type']}\n"
            output += f"Topics: {result['num_topics']}\n"
            output += f"Documents: {result['num_documents']}\n"
            output += f"Assignments: {result['num_assignments']}\n"
            output += f"Reconstruction Error: {result['reconstruction_error']:.2f}\n\n"

            output += "Discovered Topics:\n\n"
            for topic in result['topics'][:5]:
                top_words = topic['words'][:5]
                output += f"Topic {topic['topic_number']}: {', '.join(top_words)}\n"

            if len(result['topics']) > 5:
                output += f"\n... and {len(result['topics']) - 5} more topics\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error training NMF model: {str(e)}")]

    elif name == "train_bertopic":
        num_topics = arguments.get("num_topics", 10)
        min_cluster_size = arguments.get("min_cluster_size", 5)

        try:
            result = kb.train_bertopic_model(num_topics, min_cluster_size)

            if 'error' in result:
                return [TextContent(type="text", text=f"Error: {result['error']}")]

            output = f"BERTopic Modeling Complete\n{'='*60}\n\n"
            output += f"Model: {result['model_type']}\n"
            output += f"Topics Discovered: {result['num_topics']}\n"
            output += f"Documents: {result['num_documents']}\n"
            output += f"Assignments: {result['num_assignments']}\n"
            output += f"Outliers: {result['outliers']}\n\n"

            output += "Discovered Topics:\n\n"
            for topic in result['topics'][:5]:
                top_words = topic['words'][:5]
                output += f"Topic {topic['topic_number']}: {', '.join(top_words)}\n"

            if len(result['topics']) > 5:
                output += f"\n... and {len(result['topics']) - 5} more topics\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error training BERTopic model: {str(e)}")]

    elif name == "get_document_topics":
        doc_id = arguments.get("doc_id")
        model_type = arguments.get("model_type")

        try:
            cursor = kb.db_conn.cursor()

            # Build query
            if model_type:
                query = """
                    SELECT t.model_type, t.topic_number, t.top_words, dt.probability
                    FROM document_topics dt
                    JOIN topics t ON dt.topic_id = t.topic_id
                    WHERE dt.doc_id = ? AND t.model_type = ?
                    ORDER BY dt.probability DESC
                """
                results = cursor.execute(query, (doc_id, model_type)).fetchall()
            else:
                query = """
                    SELECT t.model_type, t.topic_number, t.top_words, dt.probability
                    FROM document_topics dt
                    JOIN topics t ON dt.topic_id = t.topic_id
                    WHERE dt.doc_id = ?
                    ORDER BY t.model_type, dt.probability DESC
                """
                results = cursor.execute(query, (doc_id,)).fetchall()

            if not results:
                return [TextContent(type="text", text=f"No topics found for document: {doc_id}")]

            # Get document info
            doc = kb.documents.get(doc_id)
            doc_title = doc.title if doc else "Unknown"

            output = f"Topics for Document: {doc_title}\n{'='*60}\n\n"
            output += f"Document ID: {doc_id}\n"
            output += f"Total Topics: {len(results)}\n\n"

            current_model = None
            for model, topic_num, top_words_json, prob in results:
                if model != current_model:
                    output += f"\n{model.upper()} Model Topics:\n"
                    output += "-" * 40 + "\n"
                    current_model = model

                import json
                top_words = json.loads(top_words_json)
                words_str = ', '.join(top_words[:5])
                output += f"Topic {topic_num} (prob={prob:.3f}): {words_str}\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error getting document topics: {str(e)}")]

    elif name == "cluster_documents_kmeans":
        num_clusters = arguments.get("num_clusters", 10)
        random_state = arguments.get("random_state", 42)

        try:
            result = kb.cluster_documents_kmeans(num_clusters, random_state)

            if 'error' in result:
                return [TextContent(type="text", text=f"Error: {result['error']}")]

            output = f"K-Means Clustering Complete\n{'='*60}\n\n"
            output += f"Algorithm: {result['algorithm']}\n"
            output += f"Clusters: {result['num_clusters']}\n"
            output += f"Documents: {result['num_documents']}\n"
            output += f"Assignments: {result['num_assignments']}\n"
            output += f"Silhouette Score: {result['silhouette_score']:.3f}\n\n"

            output += "Clustering quality:\n"
            if result['silhouette_score'] > 0.5:
                output += "  Excellent - Strong cluster separation\n"
            elif result['silhouette_score'] > 0.3:
                output += "  Good - Reasonable cluster structure\n"
            elif result['silhouette_score'] > 0.0:
                output += "  Fair - Weak but present structure\n"
            else:
                output += "  Poor - Overlapping clusters\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error clustering with K-Means: {str(e)}")]

    elif name == "cluster_documents_dbscan":
        eps = arguments.get("eps", 0.5)
        min_samples = arguments.get("min_samples", 5)

        try:
            result = kb.cluster_documents_dbscan(eps, min_samples)

            if 'error' in result:
                return [TextContent(type="text", text=f"Error: {result['error']}")]

            output = f"DBSCAN Clustering Complete\n{'='*60}\n\n"
            output += f"Algorithm: {result['algorithm']}\n"
            output += f"Clusters Found: {result['num_clusters']}\n"
            output += f"Documents: {result['num_documents']}\n"
            output += f"Assignments: {result['num_assignments']}\n"
            output += f"Outliers: {result['num_outliers']} ({result['num_outliers']/result['num_documents']*100:.1f}%)\n"
            output += f"Silhouette Score: {result['silhouette_score']:.3f}\n\n"

            output += "Note: DBSCAN automatically discovers clusters without\n"
            output += "specifying K. Outliers are documents that don't fit\n"
            output += "well into any cluster.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error clustering with DBSCAN: {str(e)}")]

    elif name == "cluster_documents_hdbscan":
        min_cluster_size = arguments.get("min_cluster_size", 5)
        min_samples = arguments.get("min_samples")

        try:
            result = kb.cluster_documents_hdbscan(min_cluster_size, min_samples)

            if 'error' in result:
                return [TextContent(type="text", text=f"Error: {result['error']}")]

            output = f"HDBSCAN Clustering Complete\n{'='*60}\n\n"
            output += f"Algorithm: {result['algorithm']}\n"
            output += f"Clusters Found: {result['num_clusters']}\n"
            output += f"Documents: {result['num_documents']}\n"
            output += f"Assignments: {result['num_assignments']}\n"
            output += f"Outliers: {result['num_outliers']} ({result['num_outliers']/result['num_documents']*100:.1f}%)\n\n"

            output += "Note: HDBSCAN is hierarchical and handles varying\n"
            output += "densities better than DBSCAN. More conservative with\n"
            output += "cluster assignment, resulting in higher quality clusters.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error clustering with HDBSCAN: {str(e)}")]

    elif name == "get_cluster_documents":
        cluster_id = arguments.get("cluster_id")
        algorithm = arguments.get("algorithm")
        max_results = arguments.get("max_results", 50)

        try:
            cursor = kb.db_conn.cursor()

            # Build query
            if algorithm:
                query = """
                    SELECT d.doc_id, d.title, d.filename, dc.distance, c.algorithm
                    FROM document_clusters dc
                    JOIN documents d ON dc.doc_id = d.doc_id
                    JOIN clusters c ON dc.cluster_id = c.cluster_id
                    WHERE dc.cluster_id = ? AND c.algorithm = ?
                    ORDER BY dc.distance ASC
                    LIMIT ?
                """
                results = cursor.execute(query, (cluster_id, algorithm, max_results)).fetchall()
            else:
                query = """
                    SELECT d.doc_id, d.title, d.filename, dc.distance, c.algorithm
                    FROM document_clusters dc
                    JOIN documents d ON dc.doc_id = d.doc_id
                    JOIN clusters c ON dc.cluster_id = c.cluster_id
                    WHERE dc.cluster_id = ?
                    ORDER BY dc.distance ASC
                    LIMIT ?
                """
                results = cursor.execute(query, (cluster_id, max_results)).fetchall()

            if not results:
                return [TextContent(type="text", text=f"No documents found in cluster: {cluster_id}")]

            # Get cluster info
            cluster_info = cursor.execute("""
                SELECT algorithm, cluster_number
                FROM clusters
                WHERE cluster_id = ?
            """, (cluster_id,)).fetchone()

            if not cluster_info:
                return [TextContent(type="text", text=f"Cluster not found: {cluster_id}")]

            algo, cluster_num = cluster_info

            output = f"Documents in Cluster {cluster_num} ({algo.upper()})\n{'='*60}\n\n"
            output += f"Cluster ID: {cluster_id}\n"
            output += f"Total Documents: {len(results)}\n\n"

            for doc_id, title, filename, distance, _ in results[:20]:  # Show first 20
                output += f"- {title}\n"
                output += f"  File: {filename}\n"
                output += f"  Distance: {distance:.4f}\n\n"

            if len(results) > 20:
                output += f"... and {len(results) - 20} more documents\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error getting cluster documents: {str(e)}")]

    # Phase 2: Visualization Tools
    elif name == "generate_topic_wordcloud":
        topic_id = arguments.get("topic_id")
        output_path = arguments.get("output_path")
        width = arguments.get("width", 800)
        height = arguments.get("height", 400)
        background_color = arguments.get("background_color", "white")

        try:
            result = kb.generate_topic_wordcloud(
                topic_id=topic_id,
                output_path=output_path,
                width=width,
                height=height,
                background_color=background_color
            )

            if 'error' in result:
                return [TextContent(type="text", text=f"Error: {result['error']}")]

            output = f"Topic Word Cloud Generated\n{'='*60}\n\n"
            output += f"Topic: {result['topic_number']} ({result['model_type'].upper()})\n"
            output += f"Output File: {result['output_path']}\n"
            output += f"Image Size: {width}x{height}px\n"
            output += f"Words Visualized: {result['num_words']}\n\n"
            output += "The word cloud has been saved. Larger words indicate higher importance in the topic.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error generating word cloud: {str(e)}")]

    elif name == "visualize_cluster_scatter":
        algorithm = arguments.get("algorithm")
        output_path = arguments.get("output_path")
        width = arguments.get("width", 1200)
        height = arguments.get("height", 800)
        n_neighbors = arguments.get("n_neighbors", 15)
        min_dist = arguments.get("min_dist", 0.1)

        try:
            result = kb.visualize_cluster_scatter(
                algorithm=algorithm,
                output_path=output_path,
                width=width,
                height=height,
                n_neighbors=n_neighbors,
                min_dist=min_dist
            )

            if 'error' in result:
                return [TextContent(type="text", text=f"Error: {result['error']}")]

            output = f"Cluster Scatter Plot Generated\n{'='*60}\n\n"
            output += f"Algorithm: {result['algorithm'].upper()}\n"
            output += f"Clusters Visualized: {result['num_clusters']}\n"
            output += f"Documents Plotted: {result['num_documents']}\n"
            output += f"Output File: {result['output_path']}\n"
            output += f"Image Size: {width}x{height}px\n\n"
            output += "The scatter plot shows document clusters in 2D space using UMAP projection.\n"
            output += "Documents in the same cluster appear closer together in the visualization.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error generating scatter plot: {str(e)}")]

    elif name == "generate_topic_heatmap":
        model_type = arguments.get("model_type")
        output_path = arguments.get("output_path")
        max_topics = arguments.get("max_topics", 20)
        max_documents = arguments.get("max_documents", 50)

        try:
            result = kb.generate_topic_heatmap(
                model_type=model_type,
                output_path=output_path,
                max_topics=max_topics,
                max_documents=max_documents
            )

            if 'error' in result:
                return [TextContent(type="text", text=f"Error: {result['error']}")]

            output = f"Topic Heatmap Generated\n{'='*60}\n\n"
            output += f"Model Type: {result['model_type'].upper()}\n"
            output += f"Topics: {result['num_topics']}\n"
            output += f"Documents: {result['num_documents']}\n"
            output += f"Output File: {result['output_path']}\n\n"
            output += "The heatmap shows topic-document probability matrix.\n"
            output += "Brighter colors indicate higher topic probability for a document.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error generating heatmap: {str(e)}")]

    elif name == "visualize_cluster_distribution":
        algorithm = arguments.get("algorithm")
        output_path = arguments.get("output_path")
        width = arguments.get("width", 1000)
        height = arguments.get("height", 600)

        try:
            result = kb.visualize_cluster_distribution(
                algorithm=algorithm,
                output_path=output_path,
                width=width,
                height=height
            )

            if 'error' in result:
                return [TextContent(type="text", text=f"Error: {result['error']}")]

            output = f"Cluster Distribution Chart Generated\n{'='*60}\n\n"
            output += f"Algorithm: {result['algorithm'].upper()}\n"
            output += f"Clusters: {result['num_clusters']}\n"
            output += f"Output File: {result['output_path']}\n\n"
            output += "Cluster Sizes:\n"

            # Show cluster sizes
            for cluster_name, size in sorted(result['cluster_sizes'].items()):
                if cluster_name == 'outliers':
                    output += f"  Outliers: {size} documents\n"
                else:
                    cluster_num = cluster_name.replace('cluster_', '')
                    output += f"  Cluster {cluster_num}: {size} documents\n"

            output += "\nThe bar chart shows the distribution of documents across clusters.\n"

            return [TextContent(type="text", text=output)]

        except Exception as e:
            return [TextContent(type="text", text=f"Error generating distribution chart: {str(e)}")]

    return [TextContent(type="text", text=f"Unknown tool: {name}")]


@server.list_resources()
async def list_resources() -> list[Resource]:
    """List available resources."""
    resources = []
    for doc in kb.list_documents():
        resources.append(Resource(
            uri=f"c64kb://{doc.doc_id}",
            name=doc.title,
            description=f"{doc.file_type.upper()} document with {doc.total_chunks} chunks",
            mimeType="text/plain"
        ))
    return resources


@server.read_resource()
async def read_resource(uri: str) -> str:
    """Read a resource."""
    if uri.startswith("c64kb://"):
        doc_id = uri[8:]
        content = kb.get_document_content(doc_id)
        if content:
            return content
    return f"Resource not found: {uri}"


async def main():
    """Run the MCP server."""
    # Log version information
    logger = logging.getLogger(__name__)
    logger.info("=" * 60)
    logger.info(f"Starting {get_full_version_string()}")
    logger.info(f"Build Date: {__build_date__}")
    logger.info("=" * 60)

    async with stdio_server() as (read_stream, write_stream):
        await server.run(
            read_stream,
            write_stream,
            server.create_initialization_options()
        )


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
