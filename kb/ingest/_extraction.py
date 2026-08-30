"""File-type extraction and content-parsing helpers for IngestMixin.

Split out of kb/ingest.py by R12 step 4 (module-size reduction). Move, not a
rewrite: every method body below is unchanged from the original.
"""
from features import MARKITDOWN_SUPPORT
from features import OCR_SUPPORT
from features import PDFPLUMBER_SUPPORT
from features import PDF_SUPPORT
from features import PdfReader
from features import convert_from_path
from features import pdfplumber
from features import pytesseract
from models import DocumentChunk
from models import DocumentNotFoundError
from models import KnowledgeBaseError
from models import UnsupportedFileTypeError
from typing import Optional
from util import http_get_polite
import json
import os
import re

# Extensions handed to markitdown. Deliberately an EXPLICIT allowlist and not
# a catch-all: a catch-all would mean UnsupportedFileTypeError never fires
# again and any stray binary gets ingested as text. Every extension a native
# handler already claims is absent here on purpose - see _extract_via_markitdown.
# .msg is NOT listed: it needs markitdown's 'outlook' extra (olefile), which
# this project does not declare.
_MARKITDOWN_EXTENSIONS = frozenset({
    '.docx', '.pptx', '.epub', '.csv', '.json', '.xml',
})


def _markitdown_enabled() -> bool:
    """Whether the markitdown fallback may run: installed AND not switched off."""
    return MARKITDOWN_SUPPORT and os.environ.get('TDZ_MARKITDOWN', '1') != '0'


class _ExtractionMixin:

    def _chunk_text(self, text: str, chunk_size: int = 1500, overlap: int = 200) -> list[str]:
        """
        Split text into overlapping chunks with optimized algorithm.

        Uses single-pass processing to minimize string operations.
        For large documents (>10K words), this is ~15% faster than naive split/join.

        Args:
            text: Text to chunk
            chunk_size: Number of words per chunk (default: 1500)
            overlap: Number of words to overlap between chunks (default: 200)

        Returns:
            List of text chunks
        """
        words = text.split()

        if len(words) <= chunk_size:
            return [text]

        chunks = []
        start = 0

        while start < len(words):
            end = start + chunk_size
            chunk_words = words[start:end]
            chunks.append(' '.join(chunk_words))
            start = end - overlap

        return chunks

    def _extract_pdf_with_ocr(self, filepath: str) -> tuple[str, int]:
        """Extract text from scanned PDF using OCR, returns (text, page_count)."""
        if not self.use_ocr:
            raise RuntimeError("OCR not enabled. Set USE_OCR=1 and install Tesseract.")

        if not self.poppler_available:
            raise RuntimeError(
                "Poppler not found! OCR requires poppler-utils.\n"
                "Install instructions:\n"
                "  Windows: Download from https://github.com/oschwartz10612/poppler-windows/releases/\n"
                "  Set POPPLER_PATH environment variable to the bin directory\n"
                "  Example: POPPLER_PATH=C:\\path\\to\\poppler-24.08.0\\Library\\bin"
            )

        try:
            self.logger.info(f"Using OCR to extract text from scanned PDF: {filepath}")
            # Convert PDF pages to images
            if self.poppler_path:
                images = convert_from_path(filepath, poppler_path=self.poppler_path)
            else:
                images = convert_from_path(filepath)

            pages = []
            for i, image in enumerate(images):
                try:
                    # Extract text using Tesseract OCR
                    text = pytesseract.image_to_string(image)
                    pages.append(text)
                    self.logger.debug(f"OCR processed page {i + 1}/{len(images)}")
                except Exception as e:
                    self.logger.error(f"OCR failed for page {i + 1}: {e}")
                    pages.append("")  # Empty text for failed page

            full_text = "\n\n--- PAGE BREAK ---\n\n".join(pages)
            self.logger.info(f"OCR extraction complete: {len(full_text)} characters from {len(images)} pages")
            return full_text, len(images)

        except Exception as e:
            self.logger.error(f"OCR extraction failed: {e}")
            raise RuntimeError(f"OCR extraction failed: {e}")

    def _check_poppler_available(self) -> bool:
        """Check if poppler is available for pdf2image.

        Returns True if poppler can be found, False otherwise.
        """
        if not OCR_SUPPORT:
            return False

        try:
            # Try to import pdf2image

            # Create a minimal test - just try to access pdfinfo
            import subprocess

            # Determine the command to check based on poppler_path
            if self.poppler_path:
                # Check if pdfinfo exists in the specified path
                pdfinfo_cmd = os.path.join(self.poppler_path, 'pdfinfo')
                if os.name == 'nt':  # Windows
                    pdfinfo_cmd += '.exe'

                if not os.path.exists(pdfinfo_cmd):
                    return False

                # Try to run pdfinfo -v (poppler uses -v not --version)
                result = subprocess.run(
                    [pdfinfo_cmd, '-v'],
                    capture_output=True,
                    timeout=5
                )
                return result.returncode == 0
            else:
                # Check if pdfinfo is in PATH
                if os.name == 'nt':  # Windows
                    result = subprocess.run(
                        ['where', 'pdfinfo'],
                        capture_output=True,
                        timeout=5
                    )
                else:  # Unix-like
                    result = subprocess.run(
                        ['which', 'pdfinfo'],
                        capture_output=True,
                        timeout=5
                    )
                return result.returncode == 0

        except Exception as e:
            self.logger.debug(f"Poppler check failed: {e}")
            return False

    def _extract_pdf_text(self, filepath: str) -> tuple[str, int, dict]:
        """Extract text from PDF with automatic OCR fallback for scanned PDFs.

        Returns (text, page_count, metadata).
        """
        if not PDF_SUPPORT:
            raise RuntimeError("PDF support not available. Install pypdf: pip install pypdf")

        reader = PdfReader(filepath)
        pages = []
        total_text_length = 0

        for page in reader.pages:
            text = page.extract_text() or ""
            pages.append(text)
            total_text_length += len(text.strip())

        # Check if PDF appears to be scanned (very little or no text extracted)
        # Threshold: if we get less than 100 characters total, likely scanned
        is_scanned = total_text_length < 100 and len(reader.pages) > 0

        if is_scanned and self.use_ocr:
            self.logger.info(f"PDF appears to be scanned ({total_text_length} chars extracted), falling back to OCR")
            try:
                ocr_text, page_count = self._extract_pdf_with_ocr(filepath)
                # Use OCR text instead of extracted pages
                pages = [ocr_text]
                # Still extract metadata from PDF
            except Exception as e:
                self.logger.warning(f"OCR fallback failed: {e}, using extracted text anyway")

        # Extract metadata
        metadata = {}
        if reader.metadata:
            # Convert metadata values to strings to handle IndirectObject references
            author = reader.metadata.get('/Author')
            metadata['author'] = str(author) if author else None

            subject = reader.metadata.get('/Subject')
            metadata['subject'] = str(subject) if subject else None

            creator = reader.metadata.get('/Creator')
            metadata['creator'] = str(creator) if creator else None

            creation_date = reader.metadata.get('/CreationDate')
            if creation_date:
                # Try to parse PDF date format (D:YYYYMMDDHHmmSS)
                try:
                    creation_date_str = str(creation_date)
                    if creation_date_str.startswith('D:'):
                        date_str = creation_date_str[2:16]  # Extract YYYYMMDDHHmmSS
                        metadata['creation_date'] = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
                    else:
                        metadata['creation_date'] = creation_date_str
                except (ValueError, TypeError, AttributeError, IndexError):
                    metadata['creation_date'] = str(creation_date)

        return "\n\n--- PAGE BREAK ---\n\n".join(pages), len(reader.pages), metadata

    def _extract_text_file(self, filepath: str) -> str:
        """Extract text from a text file."""
        encodings = ['utf-8', 'latin-1', 'cp1252']
        for enc in encodings:
            try:
                with open(filepath, 'r', encoding=enc) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
        raise RuntimeError(f"Could not decode {filepath}")

    def _extract_excel_file(self, filepath: str) -> tuple[str, int]:
        """
        Extract text from Excel file (.xlsx, .xls).

        Returns:
            Tuple of (text_content, sheet_count)
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("openpyxl not installed. Install with: pip install openpyxl")

        try:
            workbook = load_workbook(filepath, data_only=True)
            sheets_text = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Add sheet header
                sheets_text.append(f"\n{'='*60}\nSheet: {sheet_name}\n{'='*60}\n")

                # Extract all cell values
                rows_text = []
                for row in sheet.iter_rows(values_only=True):
                    # Filter out None values and convert to strings
                    row_values = [str(cell) if cell is not None else '' for cell in row]
                    # Skip completely empty rows
                    if any(val.strip() for val in row_values):
                        rows_text.append('\t'.join(row_values))

                sheets_text.append('\n'.join(rows_text))

            text_content = '\n\n'.join(sheets_text)
            sheet_count = len(workbook.sheetnames)

            self.logger.info(f"Extracted {sheet_count} sheets from Excel file")
            return text_content, sheet_count

        except Exception as e:
            raise RuntimeError(f"Error reading Excel file: {str(e)}")

    def _extract_html_file(self, filepath: str) -> str:
        """
        Extract text from HTML file (.html, .htm).

        Returns:
            Extracted text content
        """
        try:
            from bs4 import BeautifulSoup
        except ImportError:
            raise RuntimeError("beautifulsoup4 not installed. Install with: pip install beautifulsoup4")

        try:
            # Read the HTML file with encoding detection
            with open(filepath, 'rb') as f:
                raw_data = f.read()

            # Detect encoding
            import chardet
            detected = chardet.detect(raw_data)
            encoding = detected.get('encoding', 'utf-8')

            # Decode with detected encoding
            html_content = raw_data.decode(encoding, errors='replace')

            # Parse with BeautifulSoup
            soup = BeautifulSoup(html_content, 'html.parser')

            # Remove script and style elements
            for script in soup(['script', 'style', 'nav', 'footer', 'header']):
                script.decompose()

            # Extract text content
            text_parts = []

            # Get title if available
            if soup.title and soup.title.string:
                text_parts.append(f"Title: {soup.title.string.strip()}\n")

            # Process body or entire document
            body = soup.body if soup.body else soup

            # Handle code blocks specially to preserve formatting
            for pre in body.find_all(['pre', 'code']):
                # Mark code blocks so they're preserved
                pre_text = pre.get_text()
                if pre_text.strip():
                    text_parts.append(f"\n--- CODE BLOCK ---\n{pre_text}\n--- END CODE BLOCK ---\n")
                pre.decompose()  # Remove so we don't process again

            # Get remaining text
            main_text = body.get_text(separator='\n', strip=True)

            # Clean up excessive whitespace
            lines = [line.strip() for line in main_text.split('\n')]
            lines = [line for line in lines if line]  # Remove empty lines
            text_parts.append('\n'.join(lines))

            text_content = '\n\n'.join(text_parts)

            self.logger.info(f"Extracted HTML file ({len(text_content)} characters)")
            return text_content

        except Exception as e:
            raise RuntimeError(f"Error reading HTML file: {str(e)}")

    def _extract_tables(self, filepath: str) -> list[dict]:
        """Extract tables from PDF using pdfplumber.

        Returns a list of table dictionaries with structure:
        {
            'table_id': int,
            'page': int,
            'markdown': str,
            'searchable_text': str,
            'row_count': int,
            'col_count': int
        }
        """
        if not PDFPLUMBER_SUPPORT:
            self.logger.debug("pdfplumber not available, skipping table extraction")
            return []

        tables = []
        table_id = 0

        try:
            with pdfplumber.open(filepath) as pdf:
                for page_num, page in enumerate(pdf.pages, start=1):
                    # Extract tables from this page
                    page_tables = page.extract_tables()

                    if page_tables:
                        for table_data in page_tables:
                            if not table_data or len(table_data) == 0:
                                continue

                            # Convert table to markdown
                            markdown = self._table_to_markdown(table_data)

                            # Create searchable text (all cells joined with spaces)
                            searchable_text = " ".join(
                                str(cell).strip()
                                for row in table_data
                                for cell in row
                                if cell and str(cell).strip()
                            )

                            tables.append({
                                'table_id': table_id,
                                'page': page_num,
                                'markdown': markdown,
                                'searchable_text': searchable_text,
                                'row_count': len(table_data),
                                'col_count': len(table_data[0]) if table_data else 0
                            })
                            table_id += 1

        except Exception as e:
            self.logger.warning(f"Error extracting tables from {filepath}: {e}")
            return []

        self.logger.info(f"Extracted {len(tables)} tables from PDF")
        return tables

    def _table_to_markdown(self, table_data: list[list]) -> str:
        """Convert a table (list of lists) to markdown format."""
        if not table_data or len(table_data) == 0:
            return ""

        lines = []

        # Header row
        header = table_data[0]
        lines.append("| " + " | ".join(str(cell or "").strip() for cell in header) + " |")

        # Separator row
        lines.append("| " + " | ".join("---" for _ in header) + " |")

        # Data rows
        for row in table_data[1:]:
            lines.append("| " + " | ".join(str(cell or "").strip() for cell in row) + " |")

        return "\n".join(lines)

    def _detect_code_blocks(self, text: str) -> list[dict]:
        """Detect code blocks in text (BASIC, Assembly, Hex dumps).

        Returns a list of code block dictionaries with structure:
        {
            'block_id': int,
            'page': None,  # Page detection happens in add_document
            'block_type': str,  # 'basic', 'assembly', or 'hex'
            'code': str,
            'searchable_text': str,
            'line_count': int
        }
        """
        code_blocks = []
        block_id = 0

        # Pattern 1: BASIC code (lines starting with line numbers)
        # Example: "10 PRINT "HELLO"", "20 GOTO 10"
        basic_pattern = r'(?:^|\n)((?:\d+\s+[A-Z]+[^\n]*\n?){3,})'

        for match in re.finditer(basic_pattern, text, re.MULTILINE):
            code = match.group(1).strip()
            lines = code.split('\n')

            code_blocks.append({
                'block_id': block_id,
                'page': None,
                'block_type': 'basic',
                'code': code,
                'searchable_text': code,
                'line_count': len(lines)
            })
            block_id += 1

        # Pattern 2: Assembly code (mnemonics: LDA, STA, JMP, etc.)
        # Example: "    LDA #$00", "    STA $D020"
        assembly_pattern = r'(?:^|\n)((?:\s*(?:LDA|STA|LDX|STX|LDY|STY|JMP|JSR|RTS|BEQ|BNE|BCC|BCS|ADC|SBC|AND|ORA|EOR|INC|DEC|CMP|CPX|CPY|ASL|LSR|ROL|ROR|BIT|NOP|CLC|SEC|CLI|SEI|CLD|SED|CLV|PHA|PLA|PHP|PLP|TAX|TAY|TXA|TYA|TSX|TXS|INX|INY|DEX|DEY|BMI|BPL|BVC|BVS)[^\n]*\n?){3,})'

        for match in re.finditer(assembly_pattern, text, re.MULTILINE | re.IGNORECASE):
            code = match.group(1).strip()
            lines = code.split('\n')

            # Avoid duplicates (check if this code is already captured)
            if not any(block['code'] == code for block in code_blocks):
                code_blocks.append({
                    'block_id': block_id,
                    'page': None,
                    'block_type': 'assembly',
                    'code': code,
                    'searchable_text': code,
                    'line_count': len(lines)
                })
                block_id += 1

        # Pattern 3: Hex dumps (lines with hex values)
        # Example: "D000: 00 01 02 03 04 05 06 07"
        hex_pattern = r'(?:^|\n)((?:[0-9A-F]{4}:\s*(?:[0-9A-F]{2}\s*){8,}\n?){3,})'

        for match in re.finditer(hex_pattern, text, re.MULTILINE | re.IGNORECASE):
            code = match.group(1).strip()
            lines = code.split('\n')

            code_blocks.append({
                'block_id': block_id,
                'page': None,
                'block_type': 'hex',
                'code': code,
                'searchable_text': code,
                'line_count': len(lines)
            })
            block_id += 1

        basic_count = sum(1 for b in code_blocks if b['block_type'] == 'basic')
        assembly_count = sum(1 for b in code_blocks if b['block_type'] == 'assembly')
        hex_count = sum(1 for b in code_blocks if b['block_type'] == 'hex')
        self.logger.info(f"Detected {len(code_blocks)} code blocks ({basic_count} BASIC, {assembly_count} Assembly, {hex_count} Hex)")
        return code_blocks

    def _extract_facets(self, text: str) -> dict[str, set[str]]:
        """Extract categorizable terms for faceted search.

        Returns a dictionary of facet types to sets of values:
        {
            'hardware': {'SID', 'VIC-II', 'CIA'},
            'instruction': {'LDA', 'STA', 'JMP'},
            'register': {'$D000', '$D400'}
        }
        """
        facets = {
            'hardware': set(),
            'instruction': set(),
            'register': set()
        }

        # Extract hardware components
        facets['hardware'] = self._extract_hardware_refs(text)

        # Extract 6502 instructions
        facets['instruction'] = self._extract_instructions(text)

        # Extract register addresses
        facets['register'] = self._extract_registers(text)

        return facets

    def _extract_hardware_refs(self, text: str) -> set[str]:
        """Extract hardware component mentions from text."""
        hardware = set()

        # Hardware patterns (case-insensitive)
        patterns = {
            'SID': r'\b(?:SID|6581|8580|Sound\s+Interface\s+Device)\b',
            'VIC-II': r'\b(?:VIC-?II|VIC\s*2|6569|6567|Video\s+Interface\s+Chip)\b',
            'CIA': r'\b(?:CIA|6526|Complex\s+Interface\s+Adapter)\b',
            '6502': r'\b6502\b',
            'PLA': r'\b(?:PLA|82S100|Programmable\s+Logic\s+Array)\b',
            'Datasette': r'\b(?:Datasette|1530|C2N)\b',
            'Disk Drive': r'\b(?:1541|1571|1581|Disk\s+Drive)\b',
        }

        for component, pattern in patterns.items():
            if re.search(pattern, text, re.IGNORECASE):
                hardware.add(component)

        return hardware

    def _extract_instructions(self, text: str) -> set[str]:
        """Extract 6502 assembly instructions from text."""
        instructions = set()

        # Common 6502 mnemonics
        mnemonics = [
            'LDA', 'STA', 'LDX', 'STX', 'LDY', 'STY',
            'JMP', 'JSR', 'RTS', 'RTI',
            'BEQ', 'BNE', 'BCC', 'BCS', 'BMI', 'BPL', 'BVC', 'BVS',
            'ADC', 'SBC', 'AND', 'ORA', 'EOR',
            'INC', 'DEC', 'INX', 'INY', 'DEX', 'DEY',
            'CMP', 'CPX', 'CPY',
            'ASL', 'LSR', 'ROL', 'ROR',
            'BIT', 'NOP',
            'CLC', 'SEC', 'CLI', 'SEI', 'CLD', 'SED', 'CLV',
            'PHA', 'PLA', 'PHP', 'PLP',
            'TAX', 'TAY', 'TXA', 'TYA', 'TSX', 'TXS'
        ]

        for mnemonic in mnemonics:
            # Look for mnemonic as whole word (not part of another word)
            pattern = r'\b' + mnemonic + r'\b'
            if re.search(pattern, text, re.IGNORECASE):
                instructions.add(mnemonic)

        return instructions

    def _extract_registers(self, text: str) -> set[str]:
        """Extract memory register addresses from text."""
        registers = set()

        # Find all 4-digit hex addresses with $ prefix
        # Common C64 ranges: $D000-$DFFF (I/O), $A000-$BFFF (BASIC ROM), $E000-$FFFF (KERNAL ROM)
        register_pattern = r'\$[0-9A-Fa-f]{4}'
        matches = re.findall(register_pattern, text)

        # Normalize to uppercase and add to set
        for match in matches:
            registers.add(match.upper())

        return registers

    def _extract_cross_references(self, chunks: list[DocumentChunk], doc_id: str) -> list[dict]:
        """
        Extract cross-references from document chunks.

        Returns list of cross-reference dictionaries with keys:
        - doc_id, chunk_id, ref_type, ref_value, context
        """
        cross_refs = []

        for chunk in chunks:
            text = chunk.content
            chunk_id = chunk.chunk_id

            # Extract memory addresses ($D000-$FFFF)
            addresses = self._extract_memory_addresses(text)
            for addr in addresses:
                # Get context (sentence containing the address)
                context = self._get_reference_context(text, addr)
                cross_refs.append({
                    'doc_id': doc_id,
                    'chunk_id': chunk_id,
                    'ref_type': 'memory_address',
                    'ref_value': addr,
                    'context': context
                })

            # Extract register offsets (VIC+0, SID+4, etc.)
            offsets = self._extract_register_offsets(text)
            for offset in offsets:
                context = self._get_reference_context(text, offset)
                cross_refs.append({
                    'doc_id': doc_id,
                    'chunk_id': chunk_id,
                    'ref_type': 'register_offset',
                    'ref_value': offset,
                    'context': context
                })

            # Extract page references ("see page 156")
            page_refs = self._extract_page_references(text)
            for page_ref in page_refs:
                context = self._get_reference_context(text, str(page_ref))
                cross_refs.append({
                    'doc_id': doc_id,
                    'chunk_id': chunk_id,
                    'ref_type': 'page_reference',
                    'ref_value': str(page_ref),
                    'context': context
                })

        return cross_refs

    def _extract_memory_addresses(self, text: str) -> set[str]:
        """Extract memory addresses like $D000, $D020, etc."""
        addresses = set()
        # Match $xxxx format (4-digit hex)
        pattern = r'\$[0-9A-Fa-f]{4}\b'
        matches = re.findall(pattern, text)
        for match in matches:
            addresses.add(match.upper())
        return addresses

    def _extract_register_offsets(self, text: str) -> set[str]:
        """Extract register offset references like VIC+0, SID+4, CIA1+0."""
        offsets = set()
        # Match patterns like: VIC+0, SID+4, CIA1+12, etc.
        pattern = r'\b(VIC|SID|CIA[12]?|PLA)\s*\+\s*(\d+)\b'
        matches = re.findall(pattern, text, re.IGNORECASE)
        for chip, offset in matches:
            offsets.add(f"{chip.upper()}+{offset}")
        return offsets

    def _extract_page_references(self, text: str) -> set[int]:
        """Extract page number references like 'see page 156', 'page 42'."""
        page_nums = set()
        # Match patterns like: "page 123", "see page 456", "on page 789"
        pattern = r'\b(?:see\s+)?page\s+(\d+)\b'
        matches = re.findall(pattern, text, re.IGNORECASE)
        for page_num in matches:
            page_nums.add(int(page_num))
        return page_nums

    def _get_reference_context(self, text: str, reference: str, context_chars: int = 100) -> str:
        """Get surrounding context for a reference."""
        # Find the reference in text
        pos = text.find(reference)
        if pos == -1:
            # Try case-insensitive
            pos = text.lower().find(reference.lower())

        if pos == -1:
            return ""

        # Get surrounding context
        start = max(0, pos - context_chars)
        end = min(len(text), pos + len(reference) + context_chars)

        context = text[start:end].strip()

        # Add ellipsis if truncated
        if start > 0:
            context = "..." + context
        if end < len(text):
            context = context + "..."

        return context

    def _extract_text_for_file(self, filepath: str) -> tuple[str, str, Optional[int], dict]:
        """Extract raw text and file-type metadata for a document file.

        Shared by add_document (full ingest) and update_document (which peeks
        a new file's declared card id before deciding what to supersede).

        Returns:
            (text, file_type, total_pages, pdf_metadata)
        """
        filename = os.path.basename(filepath)
        file_ext = os.path.splitext(filename)[1].lower()

        total_pages = None
        pdf_metadata = {}
        try:
            if file_ext == '.pdf':
                text, total_pages, pdf_metadata = self._extract_pdf_text(filepath)
                file_type = 'pdf'
                self.logger.info(f"Extracted {total_pages} pages from PDF")
            elif file_ext in ['.xlsx', '.xls']:
                text, sheet_count = self._extract_excel_file(filepath)
                file_type = 'excel'
                total_pages = sheet_count  # Treat sheets as "pages"
                self.logger.info(f"Extracted Excel file with {sheet_count} sheets ({len(text)} characters)")
            elif file_ext in ['.html', '.htm']:
                text = self._extract_html_file(filepath)
                file_type = 'html'
                self.logger.info(f"Extracted HTML file ({len(text)} characters)")
            elif file_ext in ['.txt', '.md', '.asm', '.bas', '.inc', '.s']:
                text = self._extract_text_file(filepath)
                file_type = 'text'
                self.logger.info(f"Extracted text file ({len(text)} characters)")
            elif file_ext in ['.sid', '.psid', '.rsid']:
                text = self._extract_sid_file(filepath)
                file_type = 'sid'
                self.logger.info(f"Extracted SID header ({len(text)} characters)")
            elif file_ext == '.zip':
                text, tune_count = self._extract_sid_archive(filepath)
                file_type = 'sid-archive'
                total_pages = tune_count  # tunes, the way excel counts sheets
                self.logger.info(f"Extracted {tune_count} SID tunes from archive")
            elif file_ext in _MARKITDOWN_EXTENSIONS:
                if not _markitdown_enabled():
                    raise UnsupportedFileTypeError(
                        f"Unsupported file type: {file_ext} - this format needs the "
                        f"optional markitdown extra (pip install -e \".[markitdown]\"); "
                        f"set TDZ_MARKITDOWN=0 to disable it deliberately"
                    )
                text = self._extract_via_markitdown(filepath)
                file_type = file_ext.lstrip('.')
                self.logger.info(f"Extracted {file_type} via markitdown ({len(text)} characters)")
            else:
                raise UnsupportedFileTypeError(f"Unsupported file type: {file_ext}")
        except (UnsupportedFileTypeError, DocumentNotFoundError):
            raise
        except Exception as e:
            self.logger.error(f"Error extracting {filepath}: {e}")
            raise KnowledgeBaseError(f"Error extracting document: {e}")

        return text, file_type, total_pages, pdf_metadata

    def _extract_via_markitdown(self, filepath: str) -> str:
        """Convert one document markitdown handles and this pipeline does not.

        A FALLBACK, never a replacement: the dispatcher reaches this only for
        extensions no native handler claims. PDF in particular stays on
        _extract_pdf_text, which reports page boundaries and falls back to OCR;
        markitdown returns one flat markdown string with neither, and figure
        extraction, page references and the PDF viewer all depend on exactly
        what it would throw away. .zip likewise stays on the HVSC SID path.

        convert_local, NOT convert: convert() routes a string starting http:,
        https:, file: or data: to convert_uri() and fetches it over the network,
        so a filename shaped like a URI would turn a local ingest into an
        outbound request.

        The MarkItDown instance is built once and cached: its __init__ constructs
        a magika.Magika(), which loads ONNX Runtime. That is far too expensive to
        pay per document - and the import stays in here, not at module level, so
        it is never paid during the 30s MCP initialize handshake at all.
        """
        from markitdown import MarkItDown

        converter = getattr(self, '_markitdown_converter', None)
        if converter is None:
            converter = MarkItDown()
            self._markitdown_converter = converter

        return converter.convert_local(filepath).text_content

    def _extract_sid_file(self, filepath: str) -> str:
        """Read one PSID/RSID tune into a searchable text card.

        A malformed SID raises SidHeaderError, which the caller wraps as a
        KnowledgeBaseError - deliberately not UnsupportedFileTypeError, since
        the type IS supported and this particular file is broken.
        """
        from ._sid import MAX_MEMBER_BYTES, parse_sid_header, sid_card_text

        with open(filepath, 'rb') as fh:
            data = fh.read(MAX_MEMBER_BYTES)
        meta = parse_sid_header(data)
        return sid_card_text(meta, os.path.basename(filepath))

    def _extract_sid_archive(self, filepath: str) -> tuple[str, int]:
        """Read every SID tune in a zip (HVSC's distribution format).

        Nothing is written to disk: members are read in memory, so hostile
        member names have no path to act on. See kb/ingest/_sid.py.

        One archive becomes ONE document, because add_document's contract is
        one file in, one DocumentMeta out. The text is the concatenation of
        every tune's card, so each tune's title/author/released is searchable.
        """
        from ._sid import parse_sid_zip, sid_card_text

        parsed, failures = parse_sid_zip(filepath)
        if not parsed:
            raise UnsupportedFileTypeError(
                f"zip holds no readable SID tunes: {os.path.basename(filepath)}"
            )

        name = os.path.basename(filepath)
        parts = [
            f"# SID archive: {name}",
            "",
            f"{len(parsed)} tune(s) parsed from `{name}`.",
            "",
        ]
        for meta in parsed:
            parts.append(sid_card_text(meta, meta.get('member', name)))
            parts.append("")
        if failures:
            parts.append("## Members that could not be parsed")
            parts.append("")
            for member, reason in failures:
                parts.append(f"- `{member}`: {reason}")
            parts.append("")
        return "\n".join(parts), len(parsed)

    def _extract_deepsid_metadata(self, fullname: str, base_url: Optional[str] = None) -> str:
        """Fetch one tune's metadata from DeepSID's JSON API as a text card.

        Not reachable from _extract_text_for_file's extension dispatch: this
        takes a collection path, not a file on disk, so there is no extension
        to dispatch on. See kb/ingest/_sid.py for the endpoint's contract.

        The request goes through http_get_polite for the identifying
        User-Agent and 429/5xx backoff the rest of this server's fetches use -
        DeepSID is one volunteer's site, and the politeness defaults exist for
        exactly this kind of target. Note http_get_polite does
        kwargs.setdefault('headers', ...), so passing a bare headers dict
        would REPLACE the User-Agent rather than add to it; http_headers()
        merges instead, which is why the XHR header is threaded through it.
        """
        from ._sid import (
            DEEPSID_BASE_URL,
            DeepSidError,
            deepsid_card_text,
            deepsid_info_url,
            parse_deepsid_payload,
        )
        from util import http_headers, robots_allows

        base = base_url or os.environ.get('TDZ_DEEPSID_BASE_URL', DEEPSID_BASE_URL)
        url = deepsid_info_url(fullname, base)

        if not robots_allows(url):
            raise DeepSidError(f"robots.txt disallows fetching {url}")

        try:
            response = http_get_polite(
                url,
                headers=http_headers({'X-Requested-With': 'XMLHttpRequest'}),
            )
        except Exception as e:
            raise DeepSidError(f"DeepSID request failed: {e}")

        # Checked before the body: a 404 page would otherwise be reported as
        # "non-JSON body", which names the symptom instead of the cause.
        if response.status_code != 200:
            raise DeepSidError(
                f"DeepSID returned HTTP {response.status_code} for {fullname!r}"
            )

        info = parse_deepsid_payload(response.text)
        self.logger.info(
            f"Fetched DeepSID metadata for {fullname!r} "
            f"({info.get('name') or 'untitled'} by {info.get('author') or 'unknown'})"
        )
        return deepsid_card_text(info, fullname, base)

    def _extract_card_id(self, text: str) -> Optional[str]:
        """Parse the logical `id` out of a knowledge card's fenced ```json block.

        Cards are markdown documents whose identity is the `id` field of their
        first fenced json block, not their (content-hash) doc_id. Anything
        without a well-formed json block with a string `id` is not a card -
        PDFs, scrapes, ezines, etc. return None here and are left untouched
        by the upsert/refuse logic in add_document/update_document.
        """
        if not text:
            return None
        match = re.search(r'```json\s*\n(.*?)```', text, re.DOTALL)
        if not match:
            return None
        try:
            data = json.loads(match.group(1))
        except (json.JSONDecodeError, ValueError):
            return None
        if isinstance(data, dict):
            card_id = data.get('id')
            if isinstance(card_id, str) and card_id.strip():
                return card_id.strip()
        return None

    def _detect_and_extract_frames(self, url: str) -> list[str]:
        """Detect HTML frames/iframes and extract their source URLs.

        Args:
            url: The URL to check for frames

        Returns:
            List of frame source URLs (relative URLs converted to absolute)
        """
        import re
        import requests  # for the exception types in the handlers below

        try:
            # Fetch the HTML content with longer timeout
            response = http_get_polite(url, timeout=30)
            response.raise_for_status()
            html_content = response.text

            # Check if this is a frameset page
            if '<FRAMESET' not in html_content.upper() and '<FRAME' not in html_content.upper() and '<IFRAME' not in html_content.upper():
                return []

            self.logger.info(f"Detected frameset page at {url}")

            # Extract frame/iframe src attributes using regex
            # Match <frame src="..."> and <iframe src="...">
            frame_pattern = r'<(?:frame|iframe)[^>]+src=["\']([\w\-\.\/]+)["\']'
            matches = re.findall(frame_pattern, html_content, re.IGNORECASE)

            if not matches:
                return []

            # Convert relative URLs to absolute
            from urllib.parse import urljoin
            frame_urls = []
            for src in matches:
                # Skip external URLs and special targets
                if src.startswith('http://') or src.startswith('https://') or src.startswith('javascript:'):
                    continue

                absolute_url = urljoin(url, src)
                frame_urls.append(absolute_url)
                self.logger.info(f"Found frame source: {absolute_url}")

            return frame_urls

        except requests.exceptions.Timeout:
            self.logger.debug(f"Timeout detecting frames at {url} (will use mdscrape instead)")
            return []
        except Exception as e:
            self.logger.debug(f"Could not detect frames at {url}: {e} (will use mdscrape instead)")
            return []

    def _fetch_deepsid_folder(
        self, folder: str, base_url: Optional[str] = None
    ) -> tuple[list[dict], list[str], list[str]]:
        """List one DeepSID folder in ONE request: (rows, paths, subfolders).

        THE ROWS ARE RETURNED, not just the paths, and that is the whole
        point. A music.php file row carries 36 keys - author, released,
        sidmodel, lengths, player, the full stil text - so a folder listing
        already holds everything needed to build a card for every tune in it.
        Returning only paths would force a follow-up info.php call per tune:
        96 requests where 1 suffices, against one volunteer's server.

        `paths` are collection_paths rebuilt from folder + filename, in the
        shape info.php's `fullname` wants, so a caller that DOES need the one
        field a row lacks (`name`) can fetch just that tune. Recursing into
        `subfolders` is how a whole collection gets walked; this lists one
        directory and does not walk the tree itself.

        Was a module-level function until this change, deliberately, to avoid
        editing _EXPECTED_METHODS out of scope - see kb/ingest/__init__.py.
        It is a method now, and that frozenset gained both new names in the
        same commit, which is the constraint that comment anticipated.
        """
        from ._sid import (
            DEEPSID_BASE_URL,
            DeepSidError,
            deepsid_folder_collection_paths,
            deepsid_folder_url,
            parse_deepsid_folder_payload,
        )
        from util import http_headers, robots_allows

        base = base_url or os.environ.get('TDZ_DEEPSID_BASE_URL', DEEPSID_BASE_URL)
        url = deepsid_folder_url(folder, base)

        if not robots_allows(url):
            raise DeepSidError(f"robots.txt disallows fetching {url}")

        try:
            response = http_get_polite(
                url,
                headers=http_headers({'X-Requested-With': 'XMLHttpRequest'}),
            )
        except Exception as e:
            raise DeepSidError(f"DeepSID request failed: {e}")

        # Checked before the body: php/music.php answers HTTP 500 with an EMPTY
        # body (an uncaught PHP TypeError from scandir() on a folder value it
        # cannot resolve), not a JSON error - see kb/ingest/_sid.py. Reported as
        # "non-JSON body" that would name the symptom instead of the cause.
        if response.status_code != 200:
            raise DeepSidError(
                f"DeepSID returned HTTP {response.status_code} for folder {folder!r}"
            )

        files, folders = parse_deepsid_folder_payload(response.text)
        paths = deepsid_folder_collection_paths(folder, files)
        subfolders = [str(f.get("foldername") or "") for f in folders]
        self.logger.info(
            f"DeepSID folder {folder!r}: {len(files)} tune(s), "
            f"{len(subfolders)} subfolder(s) in one request"
        )
        return files, paths, subfolders
